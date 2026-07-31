"""
Preenche o MODELO.xlsx da Cemig para cada UC do dados_cemig.xlsx.
Extrai endereço dos PDFs nas CARIMBOS DIGITADOS.
Salva um arquivo por UC em ProtocoloDemandaCemig/.
"""

import re
import os
import glob
import shutil
import pdfplumber
import openpyxl
import pandas as pd

# ── Caminhos ────────────────────────────────────────────────────────────────
BASE_DIR   = r"\\10.10.250.21\Energia\ARQUIVOS ENZO\ProtocoloDemandaCemig"
MODELO     = os.path.join(BASE_DIR, "MODELO.xlsx")
DADOS      = os.path.join(BASE_DIR, "dados_cemig.xlsx")
CARIMBOS   = r"\\10.10.250.21\Energia\CONTROLE BB\DIGITADOS\CARIMBOS DIGITADOS"

# ── Mapeamento UC → arquivo PDF (já localizado) ──────────────────────────────
PDF_MAP = {
    "7.014.706.018-61": os.path.join(CARIMBOS, "08042026", "BB_2002715.pdf"),
    "7.015.705.018-30": os.path.join(CARIMBOS, "08042026", "BB_2002735.pdf"),
    "7.021.829.018-37": os.path.join(CARIMBOS, "08042026", "BB_2002728.pdf"),
    "7.021.751.018-80": os.path.join(CARIMBOS, "08042026", "BB_2002698.pdf"),
    "7.015.172.018-00": os.path.join(CARIMBOS, "08042026", "BB_2002720.pdf"),
    "7.014.835.018-90": os.path.join(CARIMBOS, "08042026", "BB_2002699.pdf"),
    "7.015.696.018-94": os.path.join(CARIMBOS, "08042026", "BB_2002719.pdf"),
    "7.015.699.018-60": os.path.join(CARIMBOS, "08042026", "BB_2002736.pdf"),
    "7.022.192.018-29": os.path.join(CARIMBOS, "08042026", "BB_2002708.pdf"),
    "7.014.310.018-08": os.path.join(CARIMBOS, "08042026", "BB_2002691.pdf"),
    "7.015.149.018-90": os.path.join(CARIMBOS, "08042026", "BB_2002732.pdf"),
    "7.014.274.018-92": os.path.join(CARIMBOS, "08042026", "BB_2002710.pdf"),
    "7.014.196.018-69": os.path.join(CARIMBOS, "08042026", "BB_2002706.pdf"),
    "7.016.313.018-03": os.path.join(CARIMBOS, "08042026", "BB_2002726.pdf"),
    "7.015.729.018-66": os.path.join(CARIMBOS, "08042026", "BB_2002675.pdf"),
    "7.016.343.018-59": os.path.join(CARIMBOS, "08042026", "BB_2002696.pdf"),
    "7.016.424.018-59": os.path.join(CARIMBOS, "08042026", "BB_2002747.pdf"),
    "7.021.750.018-95": os.path.join(CARIMBOS, "08042026", "BB_2002690.pdf"),
    "7.021.432.018-98": os.path.join(CARIMBOS, "08042026", "BB_2002712.pdf"),
    "7.026.308.018-30": os.path.join(CARIMBOS, "08042026", "BB_2002692.pdf"),
    "7.019.867.018-39": os.path.join(CARIMBOS, "08042026", "BB_2002686.pdf"),
    # 3 sem PDF no CARIMBOS – endereço será parcial (CEP + cidade)
    "7.016.582.018-37": None,  # Guanh\u00e3es, MG – CEP 39740-000
    "7.014.848.018-48": None,  # Varginha, MG   – CEP 37002-030
    "7.016.408.018-57": None,  # Nanuque, MG    – CEP 39860-000
}

# Endereços parciais para as 3 UCs sem PDF
ENDERECO_FALLBACK = {
    "7.016.582.018-37": {"cep": "39740-000", "cidade": "GUANHAES",       "estado": "MG", "rua": "", "numero": "", "complemento": "", "bairro": ""},
    "7.014.848.018-48": {"cep": "37002-030", "cidade": "VARGINHA",        "estado": "MG", "rua": "", "numero": "", "complemento": "", "bairro": ""},
    "7.016.408.018-57": {"cep": "39860-000", "cidade": "NANUQUE",         "estado": "MG", "rua": "", "numero": "", "complemento": "", "bairro": ""},
}

# ── Campos fixos (iguais para todas as UCs) ───────────────────────────────────
FIXED = {
    # Seção 1
    "O34":  "Mercado Livre",
    "AD34": "Aumento de Demanda",
    # Seção 2
    "S39":  "BANCO DO BRASIL",
    "H40":  "00.000.000/0001-91",   # CNPJ sede BB – ajuste se necessário
    "O41":  "bbenergia@acaoengenharia.com.br",
    "AL41": 1138836080,
    "O42":  "sophia.chuba@acaoengenharia.com.br",
    "AL42": 11943087270,
    # Seção 3 – Correspondência
    "AD46": "Não",
    "S47":  "E-mail",
    "AK47": "cesup.consumoenergia@bb.com.br",
    # Seção 4 – Dados da UC
    "U58":  "Outros Serviços e Outras Atividades",
    "AQ58": "64.22-1-00 - Bancos múltiplos, com carteira comercial",
    "AA59": "Urbana",
    # APP / Reserva Legal
    "BG72": "Não",
    "BG73": "Não",
    # Seção 5 – Técnico
    "S80":  "Sim",
    "S85":  13.8,
    "R94":  0.02,
    "BH134": "Não",   # troca tipo subestação
    "AP136": "Subestação N° 2",
    "AE138": "Não",   # faturamento monômio
    "AA143": "Verde",
    "AR143": "Não",   # demanda escalonada
    "AL158": "Não",   # paralelismo momentâneo
    "AL160": "Não",   # GRID ZERO
    "AL162": "Não",   # BT na mesma propriedade
    "BC169": 2,       # tipo de ramal
    # Data
    "C190":  "São Paulo, 16 de Abril de 2026",
}


# ── Funções ──────────────────────────────────────────────────────────────────

def extrair_endereco(pdf_path: str) -> dict:
    """
    Lê a fatura Cemig e devolve dicionário com os campos de endereço.
    Formato esperado na fatura:
      linha i  : BANCO DO BRASIL SA   Referente a  ...
      linha i+1: <RUA> <NUMERO> <COMPL>  <MES/ANO> <DATA> <VALOR>
      linha i+2: <BAIRRO>
      linha i+3: <CEP> <MUNICIPIO>, <UF>  NOTA FISCAL ...
    """
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""

    linhas = text.split("\n")

    rua = numero = complemento = bairro = cep = cidade = estado = ""

    for i, linha in enumerate(linhas):
        if "BANCO DO BRASIL" in linha and "Referente" in linha:
            addr_raw  = linhas[i + 1] if i + 1 < len(linhas) else ""
            bairro    = linhas[i + 2].strip() if i + 2 < len(linhas) else ""
            cep_linha = linhas[i + 3] if i + 3 < len(linhas) else ""

            # Remove a parte " MES/ANO DATA VALOR" do final da linha de endereço
            addr_part = re.split(r"\s+[A-Z]{3}/\d{4}", addr_raw)[0].strip()

            # Tenta separar: LOGRADOURO NUMERO COMPLEMENTO
            m = re.match(r"^(.+?)\s+(\d+)\s*(.*)$", addr_part)
            if m:
                rua         = m.group(1).strip()
                numero      = m.group(2).strip()
                complemento = m.group(3).strip()
            else:
                rua = addr_part

            # CEP + Cidade, Estado
            m2 = re.match(r"(\d{5}-\d{3})\s+(.+?),\s*([A-Z]{2})", cep_linha)
            if m2:
                cep    = m2.group(1)
                cidade = m2.group(2).strip()
                estado = m2.group(3)
            break

    return {
        "rua": rua, "numero": numero, "complemento": complemento,
        "bairro": bairro, "cep": cep, "cidade": cidade, "estado": estado,
    }


def preencher_modelo(uc: str, row: pd.Series, endereco: dict, sheet_name: str):
    """Copia o MODELO, preenche e salva como <uc>.xlsx."""
    saida = os.path.join(BASE_DIR, f"{uc}.xlsx")
    shutil.copy2(MODELO, saida)

    wb = openpyxl.load_workbook(saida)
    ws = wb[sheet_name]

    # Campos fixos
    for cel, val in FIXED.items():
        ws[cel] = val

    # Seção 1 – número da instalação
    ws["BB34"] = uc

    # Seção 3 – endereço de correspondência
    ws["M50"] = endereco["rua"]
    ws["AO50"] = endereco["numero"]
    ws["AZ50"] = endereco["complemento"]
    ws["M51"] = endereco["bairro"]
    ws["AO51"] = endereco["cep"]
    ws["M52"] = endereco["cidade"]
    ws["AO52"] = endereco["estado"]

    # Seção 4 – endereço da UC
    ws["F59"]  = endereco["cep"]
    ws["AQ59"] = endereco["cidade"]
    ws["BF59"] = endereco["estado"]
    ws["S66"]  = endereco["rua"]
    ws["AK66"] = endereco["numero"] if endereco["numero"] else "SN"
    ws["AR66"] = endereco["bairro"]
    ws["BE66"] = endereco["complemento"]

    # Transformadores
    kva  = int(row["Potência dos transformafores"])
    qtd  = int(row["Quant "])
    ws["H94"]  = kva
    ws["M94"]  = qtd
    ws["C101"] = "∑"
    ws["H101"] = kva
    ws["M101"] = qtd

    # Alteração de demanda
    ws["AE134"] = kva
    ws["AO134"] = qtd
    ws["AE135"] = kva   # mesma potência (sem troca de trafo)
    ws["AO135"] = qtd

    # Demandas
    ws["AF145"] = int(row["kW Cont"])
    ws["AT145"] = int(row["Solicitado"])

    wb.save(saida)
    print(f"  OK  {uc}  ->  {os.path.basename(saida)}")
    return saida


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    df = pd.read_excel(DADOS)
    print(f"UCs carregadas: {len(df)}\n")

    # Detecta nome real da aba do formulário
    wb_modelo = openpyxl.load_workbook(MODELO)
    sheet_name = next((s for s in wb_modelo.sheetnames if "ormul" in s), wb_modelo.sheetnames[0])
    wb_modelo.close()
    print(f"Aba do formulário: '{sheet_name}'\n")

    erros = []

    for _, row in df.iterrows():
        uc = str(row["Instalação"]).strip()
        print(f"Processando {uc} …")

        pdf_path = PDF_MAP.get(uc)

        if pdf_path and os.path.exists(pdf_path):
            try:
                endereco = extrair_endereco(pdf_path)
            except Exception as e:
                print(f"  AVISO - Erro ao ler PDF: {e}")
                endereco = ENDERECO_FALLBACK.get(uc, {
                    "rua": "", "numero": "", "complemento": "",
                    "bairro": "", "cep": "", "cidade": "", "estado": "MG"
                })
        else:
            endereco = ENDERECO_FALLBACK.get(uc, {
                "rua": "", "numero": "", "complemento": "",
                "bairro": "", "cep": "", "cidade": "", "estado": "MG"
            })
            if not pdf_path:
                print("  AVISO - PDF nao encontrado - endereco parcial")

        try:
            preencher_modelo(uc, row, endereco, sheet_name)
        except Exception as e:
            print(f"  ERRO ao preencher: {e}")
            erros.append((uc, str(e)))

    print(f"\n{'='*50}")
    print(f"Concluído. {len(df) - len(erros)} arquivos gerados em:")
    print(f"  {BASE_DIR}")
    if erros:
        print(f"\nErros ({len(erros)}):")
        for uc, msg in erros:
            print(f"  {uc}: {msg}")


if __name__ == "__main__":
    main()
