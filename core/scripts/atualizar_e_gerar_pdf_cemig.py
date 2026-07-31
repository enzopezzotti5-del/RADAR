"""
1. Busca enderecos das 3 UCs sem PDF na Relacao dos Prefixos
2. Corrige CNPJ para 00.000.000/5435-60 em todos os 24 arquivos
3. Valida campos obrigatorios
4. Gera PDF da aba "Formulario MT" de cada arquivo
"""

import os
import re
import io
import shutil
import subprocess
import pandas as pd
import openpyxl
from xlsx2html import xlsx2html

BASE_DIR = r"\\10.10.250.21\Energia\ARQUIVOS ENZO\ProtocoloDemandaCemig"
RELACAO  = os.path.join(BASE_DIR,
    "Relação dos Prefixos BB - Consumo (kWh) - Custo (R$) - Data da Leitura - Data Vencimento.xlsx")

CNPJ_CORRETO = "00.000.000/5435-60"

# UCs sem PDF (endereco vem da Relacao)
UCS_SEM_PDF = {"7.016.582.018-37", "7.014.848.018-48", "7.016.408.018-57"}

# Campos obrigatorios a validar
CAMPOS_OBRIG = {
    "BB34": "Instalacao",
    "O34":  "Opcao Atendimento",
    "AD34": "Finalidade",
    "S39":  "Nome",
    "H40":  "CNPJ",
    "M50":  "Rua Correspondencia",
    "AO51": "CEP",
    "M52":  "Municipio",
    "F59":  "CEP UC",
    "AQ59": "Municipio UC",
    "S66":  "Rua UC",
    "H94":  "Trafo kVA",
    "M94":  "Trafo Qtd",
    "AF145":"Demanda Atual",
    "AT145":"Demanda Futura",
}


def parse_endereco(endereco_str):
    """Transforma 'AV GETULIO VARGAS,300' em (rua, numero, complemento)."""
    s = str(endereco_str).strip()
    parts = [p.strip() for p in s.split(",")]
    rua = parts[0] if parts else s
    numero = ""
    complemento = ""
    if len(parts) >= 2:
        # primeira parte apos virgula pode conter numero e complemento
        resto = parts[1]
        m = re.match(r"^(S/N|\d+)\s*(.*)", resto, re.IGNORECASE)
        if m:
            numero = m.group(1)
            complemento = m.group(2).strip()
        else:
            complemento = resto
        if len(parts) > 2:
            complemento = (complemento + " " + " ".join(parts[2:])).strip()
    return rua, numero, complemento


def carregar_relacao():
    # Lê apenas as 6 colunas necessárias (muito mais rápido)
    df = pd.read_excel(RELACAO, sheet_name="Export",
                       usecols=["Instalacao", "CNPJ", "Endereco", "Cidade", "UF", "CEP"])
    return df.copy()


def atualizar_arquivo(xlsx_path, uc, rel_row, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]

    # Corrige CNPJ em todos os arquivos
    ws["H40"] = CNPJ_CORRETO

    # Se UC sem PDF, preenche endereco com dados da Relacao
    if uc in UCS_SEM_PDF:
        rua, numero, complemento = parse_endereco(rel_row["Endereco"])
        cidade = str(rel_row["Cidade"]).strip().upper()
        cep_raw = str(rel_row["CEP"]).strip()
        # formata CEP se vier sem hifen
        cep = cep_raw if "-" in cep_raw else cep_raw[:5] + "-" + cep_raw[5:]
        estado = str(rel_row["UF"]).strip().upper()

        # Correspondencia (secao 3)
        ws["M50"]  = rua
        ws["AO50"] = numero if numero else "S/N"
        ws["AZ50"] = complemento
        ws["M51"]  = ""          # bairro nao consta na Relacao
        ws["AO51"] = cep
        ws["M52"]  = cidade
        ws["AO52"] = estado

        # Endereco UC (secao 4)
        ws["F59"]  = cep
        ws["AA59"] = "Urbana"
        ws["AQ59"] = cidade
        ws["BF59"] = estado
        ws["S66"]  = rua
        ws["AK66"] = numero if numero else "S/N"
        ws["AR66"] = ""          # bairro
        ws["BE66"] = complemento

    wb.save(xlsx_path)


def validar(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    faltando = []
    for cel, nome in CAMPOS_OBRIG.items():
        val = ws[cel].value
        if val is None or str(val).strip() == "":
            faltando.append(nome)
    return faltando


CHROME = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
TMP    = r"C:\Users\Revit\AppData\Local\Temp"


def gerar_pdfs_em_lote(arquivos_xlsx, sheet_name, base_dir):
    """Converte cada xlsx -> HTML -> PDF via Chrome headless."""
    gerados = []
    erros   = []

    for arq in arquivos_xlsx:
        xlsx_net  = os.path.join(base_dir, arq)           # rede
        xlsx_tmp  = os.path.join(TMP, arq)                # local temp
        html_tmp  = xlsx_tmp.replace(".xlsx", ".html")
        pdf_net   = xlsx_net.replace(".xlsx", ".pdf")
        pdf_tmp   = xlsx_tmp.replace(".xlsx", ".pdf")

        try:
            # 1. Copia xlsx para local (xlsx2html nao suporta UNC)
            shutil.copy2(xlsx_net, xlsx_tmp)

            # 2. Converte xlsx -> HTML (somente a aba do formulario)
            out = io.StringIO()
            xlsx2html(xlsx_tmp, out, sheet=sheet_name)
            with open(html_tmp, "w", encoding="utf-8") as f:
                f.write(out.getvalue())

            # 3. Chrome headless: HTML -> PDF
            subprocess.run(
                [CHROME, "--headless=new", "--disable-gpu",
                 f"--print-to-pdf={pdf_tmp}", "--no-margins",
                 html_tmp],
                capture_output=True, timeout=30
            )

            # 4. Copia PDF de volta para a rede
            if os.path.exists(pdf_tmp):
                shutil.copy2(pdf_tmp, pdf_net)
                gerados.append(pdf_net)
                print(f"        PDF: {os.path.basename(pdf_net)}")
            else:
                raise RuntimeError("Chrome nao gerou o PDF")

        except Exception as e:
            erros.append((arq, str(e)))
            print(f"        ERRO PDF {arq}: {e}")
        finally:
            for f in [xlsx_tmp, html_tmp, pdf_tmp]:
                try:
                    os.remove(f)
                except Exception:
                    pass

    return gerados, erros


def main():
    # Carrega Relacao
    rel = carregar_relacao()
    rel_idx = rel.set_index("Instalacao")

    # Lista todos os xlsx gerados (exclui MODELO e planilhas de apoio)
    arquivos = sorted([
        f for f in os.listdir(BASE_DIR)
        if f.endswith(".xlsx") and f[0].isdigit()
    ])
    print(f"Arquivos encontrados: {len(arquivos)}\n")

    # Detecta nome real da aba
    wb_modelo = openpyxl.load_workbook(
        os.path.join(BASE_DIR, arquivos[0]))
    sheet_name = next(
        (s for s in wb_modelo.sheetnames if "ormul" in s),
        wb_modelo.sheetnames[0])
    wb_modelo.close()
    print(f"Aba: '{sheet_name}'\n")

    erros_valid = []
    arquivos_ok = []

    # Passo 1: atualizar e validar todos os xlsx
    print("--- Passo 1: Atualizando e validando xlsx ---")
    for arq in arquivos:
        uc = arq.replace(".xlsx", "")
        xlsx_path = os.path.join(BASE_DIR, arq)
        rel_row = rel_idx.loc[uc] if uc in rel_idx.index else None

        try:
            atualizar_arquivo(xlsx_path, uc, rel_row, sheet_name)
        except Exception as e:
            print(f"  ERRO ao atualizar {uc}: {e}")
            continue

        faltando = validar(xlsx_path, sheet_name)
        if faltando:
            print(f"  AVISO {uc}: campos vazios -> {faltando}")
            erros_valid.append((uc, faltando))
        else:
            print(f"  OK    {uc}")
        arquivos_ok.append(arq)

    # Passo 2: gerar todos os PDFs de uma vez (Excel abre uma unica vez)
    print(f"\n--- Passo 2: Gerando {len(arquivos_ok)} PDFs (Excel unico) ---")
    pdfs_gerados, erros_pdf = gerar_pdfs_em_lote(arquivos_ok, sheet_name, BASE_DIR)

    print(f"\n{'='*55}")
    print(f"PDFs gerados : {len(pdfs_gerados)}")
    print(f"Erros valid  : {len(erros_valid)}")
    print(f"Erros PDF    : {len(erros_pdf)}")
    if erros_valid:
        print("\nCampos faltando:")
        for uc, campos in erros_valid:
            print(f"  {uc}: {campos}")
    if erros_pdf:
        print("\nErros ao gerar PDF:")
        for uc, msg in erros_pdf:
            print(f"  {uc}: {msg}")


if __name__ == "__main__":
    main()
