#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OCR Neoenergia (BT) -> XLSX para digitacao no Consen.

Fluxo:
1) Le o indice_downloads_neoenergia.csv (na rede)
2) Filtra por mes/ano
3) Faz OCR textual dos PDFs e monta planilha

Saida:
    \\\\10.10.250.21\\Energia\\ARQUIVOS ENZO\\OCR NEOENERGIA\\ocr_neoenergia_MMYYYY.xlsx
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import logging
import re
import shutil
import sys
import tempfile
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


DOWNLOAD_ROOT = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/DOWNLOAD NEOENERGIA")
OUTPUT_DIR = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/OCR NEOENERGIA")
MASTER_FILE = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/indice_master.csv")
MAX_WORKERS = 4
DEFAULT_ESTADO = "SAO_PAULO"

HEADERS = [
    "Instalacao", "fatDataEmissao", "fatDataVcto", "fatValorFatura", "concCod",
    "fatDataCadastro", "fatDataLeituraAnterior", "fatDataLeituraAtual", "fatIlumPublica",
    "cadTarifaCod", "cadSubGrupoCod", "fatDemContratadaPonta", "fatDemContratadaFPonta",
    "fatDemPontaRegistrada", "fatDemFPontaIndRegistrada", "fatDemFPontaCapRegistrada",
    "fatDemPontaExcFaturada", "fatDemFPontaExcFaturada", "fatDemPontaExcRegistrada",
    "fatDemFPontaExcRegistrada", "fatDemPontaFaturada", "fatDemFPontaIndFaturada",
    "fatDemPontaUltra", "fatDemFPontaIndUltra", "fatConPontaRegistrado",
    "fatConFPontaIndRegistrado", "fatConFPontaCapRegistrado", "fatConIntermediarioRegistrado",
    "fatConPontaFaturado", "fatConFPontaIndFaturado", "fatConFPontaCapFaturado",
    "fatConIntermediarioFaturado", "fatConPontaExcRegistrado", "fatConFPontaIndExcRegistrado",
    "fatConFPontaCapExcRegistrado", "fatConPontaExcFaturado", "fatConFPontaIndExcFaturado",
    "fatConFPontaCapExcFaturado", "fatICMS", "fatPIS", "fatCOFINS", "fatValorNotaFiscal",
    "obsCod_1", "obsValor_1", "obsCod_2", "obsValor_2", "obsCod_3", "obsValor_3",
    "obsCod_4", "obsValor_4", "obsCod_5", "obsValor_5", "CNPJ", "ENDERECO",
    "NOTAFISCAL", "CODIGOCLIENTE", "fatDataReferencia", "fatConPontaInjetadoRegistrado",
    "fatConPontaInjetadoFaturado", "fatConFPontaInjetadoRegistrado",
    "fatConFPontaInjetadoFaturado", "fatCodigoBarras", "Debitos anteriores",
    "fatCarimbo", "usuCod", "fatDemPontaGeracaoRegistrada", "fatDemPontaGeracao",
    "fatDemPontaGeracaoValorReais", "fatDemFPontaGeracaoRegistrada", "fatDemFPontaGeracao",
    "fatDemFPontaGeracaoValorReais", "fatDemContratadaGeracaoPonta",
    "fatDemContratadaGeracaoFPonta", "fatDemPontaValorReais", "fatDemFPontaIndValorReais",
    "fatDemPontaUltraValorReais", "fatDemFPontaIndUltraValorReais", "fatDemPontaExcValorReais",
    "fatDemFPontaExcValorReais", "fatConPontaValorReais", "fatConFPontaIndValorReais",
    "fatConFPontaCapValorReais", "fatConIntermediarioValorReais", "fatConPontaExcValorReais",
    "fatConFPontaIndExcValorReais", "fatConFPontaCapExcValorReais",
    "fatConPontaInjetadoValorReais", "fatConFPontaInjetadoValorReais",
    "fatConPontaInjetadoUsina", "fatConPontaInjetadoUsinaSaldoAcumulado",
    "fatConFPontaInjetadoUsina", "fatConFPontaInjetadoUsinaSaldoAcumulado",
    "fatDemandasDevolucaoPtaValorReais", "fatDemandasDevolucaoFPtaValorReais",
    "fatConIntermedInjetadoRegistrado", "fatConIntermedInjetadoFaturado",
    "fatConIntermedInjetadoValorReais", "fatDescontoFio", "fatDescPisAliquota",
    "fatDescPisPercRetImposto", "fatDescPisValRetImposto", "fatDesCofinsAliquota",
    "fatDescCofinsPercRetImposto", "fatDescCofinsValRetImposto", "fatDesIcmsAliquota",
    "fatDescCsllPercRetImposto", "fatDescCsllValRetImposto", "fatDescIrpjPercRetImposto",
    "fatDescIrpjValRetImposto", "fatDescIrrfPercRetImposto", "fatDescIrrfValRetImposto",
    "fatDescConsumoPercRetImposto", "fatDescConsumoValRetImposto",
    "fatDescDemandaPercRetImposto", "fatDescDemandaValRetImposto", "fatValBandeira",
    "fatValBandeira2", "fatDIC", "fatFIC", "fatMultas", "fatTributoFederalPerc",
    "fatTributoFederalVal", "fatMultasDiversas", "fatDescontoFioKWh",
    "fatConCreditoTUSDPontaValorReais", "fatConCreditoTUSDFPontaValorReais",
    "fatBeneficioTarifarioBrutoValorReais", "fatBeneficioLiquidoValorReais",
    "fatContaCovidValorReais", "fatEscassezHidricaValorReais", "fatContaCovid",
    "fatEscassezHidrica", "TARIFA_DETECTADA", "ARQUIVO", "ERRO",
]

DATE_HEADERS = {
    "fatDataEmissao",
    "fatDataVcto",
    "fatDataCadastro",
    "fatDataLeituraAnterior",
    "fatDataLeituraAtual",
    "fatDataReferencia",
}

TEXT_HEADERS = {
    "Instalacao", "concCod", "cadTarifaCod", "cadSubGrupoCod",
    "obsCod_1", "obsCod_2", "obsCod_3", "obsCod_4", "obsCod_5",
    "CNPJ", "ENDERECO", "NOTAFISCAL", "CODIGOCLIENTE", "fatCodigoBarras",
    "fatCarimbo", "usuCod", "fatConPontaInjetadoUsina",
    "fatConPontaInjetadoUsinaSaldoAcumulado", "fatConFPontaInjetadoUsina",
    "fatConFPontaInjetadoUsinaSaldoAcumulado", "fatContaCovid", "fatEscassezHidrica",
    "TARIFA_DETECTADA", "ARQUIVO", "ERRO",
}

NUMERIC_HEADERS = set(HEADERS) - DATE_HEADERS - TEXT_HEADERS

HEADER_DISPLAY = {
    "Instalacao": "Instalação",
    "Debitos anteriores": "Débitos anteriores",
}

COL_WIDTHS = {
    "Instalacao": 20.71, "fatDataEmissao": 18.14, "fatDataVcto": 13.0,
    "fatValorFatura": 17.29, "concCod": 10.29, "fatDataCadastro": 18.86,
    "fatDataLeituraAnterior": 26.29, "fatDataLeituraAtual": 22.86,
    "fatIlumPublica": 17.29, "cadTarifaCod": 15.71, "cadSubGrupoCod": 20.14,
    "fatDemContratadaPonta": 27.57, "fatDemContratadaFPonta": 29.0,
    "fatDemPontaRegistrada": 27.14, "fatDemFPontaIndRegistrada": 32.29,
    "fatDemFPontaCapRegistrada": 32.57, "fatDemPontaExcFaturada": 29.14,
    "fatDemFPontaExcFaturada": 30.57, "fatDemPontaExcRegistrada": 31.0,
    "fatDemFPontaExcRegistrada": 32.29, "fatDemPontaFaturada": 25.29,
    "fatDemFPontaIndFaturada": 30.43, "fatDemPontaUltra": 21.0,
    "fatDemFPontaIndUltra": 26.14, "fatConPontaRegistrado": 26.43,
    "fatConFPontaIndRegistrado": 31.57, "fatConFPontaCapRegistrado": 31.86,
    "fatConIntermediarioRegistrado": 35.43, "fatConPontaFaturado": 24.57,
    "fatConFPontaIndFaturado": 29.71, "fatConFPontaCapFaturado": 30.14,
    "fatConIntermediarioFaturado": 33.57, "fatConPontaExcRegistrado": 30.29,
    "fatConFPontaIndExcRegistrado": 35.43, "fatConFPontaCapExcRegistrado": 35.71,
    "fatConPontaExcFaturado": 28.43, "fatConFPontaIndExcFaturado": 33.57,
    "fatConFPontaCapExcFaturado": 33.86, "fatICMS": 10.0, "fatPIS": 7.86,
    "fatCOFINS": 12.71, "fatValorNotaFiscal": 21.43, "obsCod_1": 10.0, "obsValor_1": 12.0,
    "obsCod_2": 10.0, "obsValor_2": 12.0, "obsCod_3": 10.0, "obsValor_3": 12.0,
    "obsCod_4": 10.0, "obsValor_4": 12.0, "obsCod_5": 10.0, "obsValor_5": 12.0,
    "CNPJ": 24.0, "ENDERECO": 153.14, "NOTAFISCAL": 15.86, "CODIGOCLIENTE": 20.43,
    "fatDataReferencia": 20.86, "fatConPontaInjetadoRegistrado": 35.43,
    "fatConPontaInjetadoFaturado": 33.71, "fatConFPontaInjetadoRegistrado": 36.86,
    "fatConFPontaInjetadoFaturado": 35.0, "fatCodigoBarras": 34.0,
    "Debitos anteriores": 21.14, "fatCarimbo": 13.43, "usuCod": 9.29,
    "fatDemPontaGeracaoRegistrada": 36.0, "fatDemPontaGeracao": 24.29,
    "fatDemPontaGeracaoValorReais": 36.14, "fatDemFPontaGeracaoRegistrada": 37.43,
    "fatDemFPontaGeracao": 25.71, "fatDemFPontaGeracaoValorReais": 37.57,
    "fatDemContratadaGeracaoPonta": 36.43, "fatDemContratadaGeracaoFPonta": 37.86,
    "fatDemPontaValorReais": 27.29, "fatDemFPontaIndValorReais": 32.43,
    "fatDemPontaUltraValorReais": 33.0, "fatDemFPontaIndUltraValorReais": 38.14,
    "fatDemPontaExcValorReais": 31.14, "fatDemFPontaExcValorReais": 32.57,
    "fatConPontaValorReais": 26.57, "fatConFPontaIndValorReais": 31.71,
    "fatConFPontaCapValorReais": 32.0, "fatConIntermediarioValorReais": 35.57,
    "fatConPontaExcValorReais": 30.43, "fatConFPontaIndExcValorReais": 35.57,
    "fatConFPontaCapExcValorReais": 35.86, "fatConPontaInjetadoValorReais": 35.57,
    "fatConFPontaInjetadoValorReais": 37.0, "fatConPontaInjetadoUsina": 30.0,
    "fatConPontaInjetadoUsinaSaldoAcumulado": 48.29, "fatConFPontaInjetadoUsina": 31.29,
    "fatConFPontaInjetadoUsinaSaldoAcumulado": 49.71,
    "fatDemandasDevolucaoPtaValorReais": 42.14, "fatDemandasDevolucaoFPtaValorReais": 43.57,
    "fatConIntermedInjetadoRegistrado": 39.29, "fatConIntermedInjetadoFaturado": 37.57,
    "fatConIntermedInjetadoValorReais": 39.43, "fatDescontoFio": 17.57,
    "fatDescPisAliquota": 21.71, "fatDescPisPercRetImposto": 30.0,
    "fatDescPisValRetImposto": 28.71, "fatDesCofinsAliquota": 24.14,
    "fatDescCofinsPercRetImposto": 33.57, "fatDescCofinsValRetImposto": 32.29,
    "fatDesIcmsAliquota": 22.57, "fatDescCsllPercRetImposto": 30.71,
    "fatDescCsllValRetImposto": 29.43, "fatDescIrpjPercRetImposto": 30.86,
    "fatDescIrpjValRetImposto": 29.57, "fatDescIrrfPercRetImposto": 30.57,
    "fatDescIrrfValRetImposto": 29.29, "fatDescConsumoPercRetImposto": 36.86,
    "fatDescConsumoValRetImposto": 35.57, "fatDescDemandaPercRetImposto": 37.14,
    "fatDescDemandaValRetImposto": 35.71, "fatValBandeira": 17.57,
    "fatValBandeira2": 18.86, "fatDIC": 8.14, "fatFIC": 7.86,
    "fatMultas": 11.43, "fatTributoFederalPerc": 25.29, "fatTributoFederalVal": 24.0,
    "fatMultasDiversas": 20.86, "fatDescontoFioKWh": 23.0,
    "fatConCreditoTUSDPontaValorReais": 40.86, "fatConCreditoTUSDFPontaValorReais": 42.29,
    "fatBeneficioTarifarioBrutoValorReais": 42.0, "fatBeneficioLiquidoValorReais": 34.43,
    "fatContaCovidValorReais": 28.43, "fatEscassezHidricaValorReais": 33.57,
    "fatContaCovid": 16.57, "fatEscassezHidrica": 21.57,
    "TARIFA_DETECTADA": 18.0, "ARQUIVO": 38.0, "ERRO": 45.0,
}

MESES = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}

RE_DATE = re.compile(r"\b(\d{2})[/-](\d{2})[/-](\d{4})\b")
RE_MONEY = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")
RE_CARIMBO = re.compile(r"BB_(\d+)$", re.IGNORECASE)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("ocr_neoenergia")
_MASTER_UC_POR_CARIMBO: dict[str, str] | None = None


def _strip_accents(text: str) -> str:
    if not text:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )


def _norm(text: str) -> str:
    return " ".join((text or "").split()).strip()


def _uc_por_carimbo_master(carimbo: str) -> str:
    global _MASTER_UC_POR_CARIMBO
    chave = _norm(carimbo).upper()
    if chave.isdigit():
        chave = f"BB_{chave}"
    if not chave.startswith("BB_"):
        return ""

    if _MASTER_UC_POR_CARIMBO is None:
        cache: dict[str, str] = {}
        if MASTER_FILE.exists():
            for enc in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    with MASTER_FILE.open("r", newline="", encoding=enc) as f:
                        for row in csv.DictReader(f):
                            idx = _norm(str(row.get("INDICE") or "")).upper()
                            if not idx.startswith("BB_"):
                                continue
                            uc = _norm(str(row.get("UC") or row.get("INSTALACAO") or ""))
                            if uc:
                                cache[idx] = uc
                    break
                except UnicodeDecodeError:
                    continue
                except Exception:
                    break
        _MASTER_UC_POR_CARIMBO = cache

    return _MASTER_UC_POR_CARIMBO.get(chave, "")


def _to_float_br(raw: str) -> float:
    if raw is None:
        return 0.0
    txt = str(raw).strip()
    if not txt:
        return 0.0
    neg = txt.startswith("-") or txt.endswith("-")
    txt = txt.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
    txt = txt.replace("-", "")
    try:
        val = float(txt)
        return -val if neg else val
    except Exception:
        return 0.0


def _to_date(raw: str) -> dt.date | None:
    if not raw:
        return None
    m = RE_DATE.search(str(raw))
    if not m:
        return None
    d, mth, y = map(int, m.groups())
    try:
        return dt.date(y, mth, d)
    except Exception:
        return None


def _parse_ref_mes_ano(ref: str) -> tuple[int, int] | None:
    ref_txt = _norm(str(ref or ""))
    if not ref_txt:
        return None

    m_num = re.search(r"\b(\d{2})/(\d{4})\b", ref_txt)
    if m_num:
        mm, yy = int(m_num.group(1)), int(m_num.group(2))
        if 1 <= mm <= 12:
            return mm, yy

    ref_norm = _strip_accents(ref_txt).upper()
    m_nome = re.search(r"\b([A-Z]+)\s*/\s*(\d{4})\b", ref_norm)
    if m_nome:
        mes_nome = m_nome.group(1).strip()
        ano = int(m_nome.group(2))
        mes = MESES.get(mes_nome)
        if mes:
            return mes, ano

    return None


def _texto_normalizado(text: str) -> str:
    return _strip_accents(_norm(text)).upper()


def _extract_pdf_data(pdf_path: Path) -> tuple[str, list[dict]]:
    partes: list[str] = []
    words_first_page: list[dict] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for idx, pagina in enumerate(pdf.pages):
            try:
                txt = pagina.extract_text() or pagina.extract_text(layout=True) or ""
            except Exception:
                txt = ""
            if txt:
                partes.append(txt)

            if idx == 0:
                try:
                    words_first_page = pagina.extract_words(use_text_flow=True, keep_blank_chars=False) or []
                except Exception:
                    words_first_page = []
    return "\n".join(partes), words_first_page


def _extract_text(pdf_path: Path) -> str:
    text, _ = _extract_pdf_data(pdf_path)
    return text


def _find_date_after_labels(text: str, labels: Iterable[str]) -> dt.date | None:
    textos = [text, _texto_normalizado(text)]
    labels_norm = list(labels) + [_strip_accents(l).upper() for l in labels]
    for txt in textos:
        for label in labels_norm:
            pattern = rf"{re.escape(label)}[^\d]{{0,35}}(\d{{2}}/\d{{2}}/\d{{4}})"
            m = re.search(pattern, txt, flags=re.IGNORECASE)
            if m:
                d = _to_date(m.group(1))
                if d:
                    return d
    return None


def _extract_vencimento_from_words(words: list[dict]) -> dt.date | None:
    if not words:
        return None

    for w in words:
        txt = _texto_normalizado(str(w.get("text", "")))
        if "VENCIMENTO" not in txt:
            continue

        x0 = float(w.get("x0", 0))
        x1 = float(w.get("x1", 0))
        bottom = float(w.get("bottom", 0))

        # Ignora "Vencimento" em linha de cabeçalho de débitos anteriores
        # (ex: "Vencimento Reaviso Valor" — tabela de reemissão Neoenergia/Elektro)
        mesma_linha = " ".join(
            _texto_normalizado(str(ww.get("text", "")))
            for ww in words
            if abs(float(ww.get("bottom", 0)) - bottom) <= 4
        )
        if "REAVISO" in mesma_linha:
            continue

        candidatos: list[tuple[float, dt.date]] = []

        for ww in words:
            data = _to_date(str(ww.get("text", "")))
            if not data:
                continue
            top = float(ww.get("top", 0))
            wx0 = float(ww.get("x0", 0))
            wx1 = float(ww.get("x1", 0))
            if top < bottom - 2 or top > bottom + 28:
                continue
            if wx1 < x0 - 18 or wx0 > x1 + 28:
                continue
            candidatos.append((top, data))

        if candidatos:
            candidatos.sort(key=lambda item: item[0])
            return candidatos[0][1]

    return None


def _extract_vencimento(text: str, words: list[dict] | None = None) -> dt.date | None:
    direto_words = _extract_vencimento_from_words(words or [])
    if direto_words:
        return direto_words

    # Remove linhas com "Reaviso" para evitar capturar datas de débitos anteriores
    # (seção "Vencimento Reaviso" em faturas Neoenergia/Elektro)
    text_sem_reaviso = "\n".join(
        line for line in text.splitlines()
        if "REAVISO" not in _texto_normalizado(line)
    )

    direto = _find_date_after_labels(
        text_sem_reaviso,
        ["Vencimento", "Data de vencimento", "Vence em"],
    )
    if direto:
        return direto

    m_resumo = re.search(
        r"REF:?\s*M[ÊE]S/ANO\s+TOTAL A PAGAR R\$?\s+VENCIMENTO\s+"
        r"(\d{2}/\d{4})\s+[\d\.,]+\s+(\d{2}/\d{2}/\d{4})",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if m_resumo:
        d = _to_date(m_resumo.group(2))
        if d:
            return d

    m_codigo = re.search(
        r"(\d{6,12})\s+(\d{2}/\d{2}/\d{4})\s+[\d\.,]+\s+"
        r"\d{2}/\d{4}\s+C[ÓO]DIGO DO CLIENTE\s+VENCIMENTO",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if m_codigo:
        d = _to_date(m_codigo.group(2))
        if d:
            return d

    # Padrão CEB (Neoenergia Brasília): "MÊS/ANO VALOR VCTO" sem R$ explícito
    # Ex: "ABR/2026 0,00 23/04/2026" ou "ABR/2026 2.141,45 23/05/2026"
    # Diferente de Elektro/COELBA que usam "Julho/2026 R$1.296,48 20/07/2026"
    m_ceb = re.search(
        r"\b([A-Z]{2,4})/20\d{2}\s+[\d\.,]+\s+(\d{2}/\d{2}/20\d{2})\b",
        text_sem_reaviso,
    )
    if m_ceb:
        d = _to_date(m_ceb.group(2))
        if d:
            return d

    linhas = [ln for ln in text.splitlines() if _norm(ln)]
    for line in linhas[:50]:
        line_norm = _texto_normalizado(line)
        if "DATA DE EMISSAO" in line_norm or "PROTOCOLO DE AUTORIZACAO" in line_norm:
            continue
        datas = [_to_date(x) for x in re.findall(r"\d{2}/\d{2}/\d{4}", line)]
        datas = [d for d in datas if d]
        if not datas:
            continue

        if re.search(r"\b([A-Z]+/\d{4}|\d{2}/\d{4})\b", line_norm) and ("R$" in line or len(datas) >= 2):
            return datas[-1]

        if "PAGUE COM O PIX" in line_norm and "R$" in line:
            return datas[0]

        if "SEU CODIGO" in line_norm or "CONTROLE" in line_norm or "TOTAL R$" in line_norm:
            continue

        if ("R$" in line or re.search(r"\b\d{6,12}\b", line)) and len(datas) == 1:
            return datas[0]

    return None


def _extract_leituras(text: str) -> tuple[dt.date | None, dt.date | None]:
    ant = _find_date_after_labels(
        text,
        [
            "Leitura anterior",
            "Leit. anterior",
            "Data leitura anterior",
        ],
    )
    atu = _find_date_after_labels(
        text,
        [
            "Leitura atual",
            "Leit. atual",
            "Data leitura atual",
            "Proxima leitura",
        ],
    )

    if ant and atu:
        return ant, atu

    pares = re.findall(r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})", text)
    for a, b in pares:
        da, db = _to_date(a), _to_date(b)
        if da and db and da <= db:
            if ant is None:
                ant = da
            if atu is None:
                atu = db
            break
    return ant, atu


def _extract_total(text: str) -> float:
    # CELPE/Neoenergia PE: linha autônoma "TOTAL 1.463,52" ao final da tabela de
    # itens. Verificado antes do bloco REF/VENCIMENTO para evitar que o
    # "TOTAL A PAGAR R$ 0,00" (débito automático) sobreponha o valor real.
    for line in text.splitlines():
        line_n = _texto_normalizado(line).strip()
        if re.fullmatch(r"TOTAL\s+[\d\.,]+(?:\s+\S.*)?", line_n):
            valores = RE_MONEY.findall(line)
            if valores:
                return abs(_to_float_br(valores[-1]))

    m_resumo = re.search(
        r"REF:?\s*M[ÊE]S/ANO\s+TOTAL A PAGAR R\$?\s+VENCIMENTO\s+"
        r"\d{2}/\d{4}\s+([-\d\.,]+)\s+\d{2}/\d{2}/\d{4}",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if m_resumo:
        val = abs(_to_float_br(m_resumo.group(1)))
        if val > 0:
            return val
        # valor zero = possível débito automático (ex: CELPE); continua buscando

    labels = [
        "Total a pagar",
        "Valor total",
        "Valor da fatura",
        "Total da fatura",
        "Valor a pagar",
        "TOTAL A PAGAR",
        "VALOR A PAGAR",
        "TOTAL FATURA",
    ]
    for label in labels:
        pattern = rf"{re.escape(label)}[^\d\-]{{0,25}}({RE_MONEY.pattern})"
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            return _to_float_br(m.group(1))

    for line in text.splitlines():
        line_n = _texto_normalizado(line)
        if any(lbl in line_n for lbl in ("TOTAL A PAGAR", "VALOR A PAGAR", "TOTAL FATURA")):
            monies = RE_MONEY.findall(line)
            if monies:
                return abs(_to_float_br(monies[-1]))

    # Elektro: "Março/2026 R$269,31 07/04/2026" (mês pode vir com encoding corrompido)
    m_elektro = re.search(
        r"\w+/\d{4}\s+R\$([\d\.,]+)\s+\d{2}/\d{2}/\d{4}",
        text,
        flags=re.IGNORECASE,
    )
    if m_elektro:
        return abs(_to_float_br(m_elektro.group(1)))

    # CELPE/Neoenergia PE: linha "TOTAL 1.463,52" ao final da tabela de itens.
    # O bloco REF:MÊS/ANO traz "TOTAL A PAGAR R$ 0,00" por débito automático,
    # portanto chegamos aqui com 0. A linha autônoma "TOTAL" é o valor real.
    for line in text.splitlines():
        line_n = _texto_normalizado(line).strip()
        if re.fullmatch(r"TOTAL\s+[\d\.,]+(?:\s+\S.*)?", line_n):
            valores = RE_MONEY.findall(line)
            if valores:
                return abs(_to_float_br(valores[-1]))

    return 0.0


def _extract_codigo_barras(text: str) -> str:
    for line in text.splitlines():
        line_norm = _norm(line)
        if not line_norm:
            continue
        line_upper = _texto_normalizado(line_norm)
        if "CHAVE DE ACESSO" in line_upper or "PIX COPIA E COLA" in line_upper:
            continue

        m_linha = re.search(
            r"\b(\d{5}\.\d{5})\s+(\d{5}\.\d{6})\s+(\d{5}\.\d{6})\s+(\d)\s+(\d{14})\b",
            line_norm,
        )
        if m_linha:
            return re.sub(r"\D", "", "".join(m_linha.groups()))

    for line in text.splitlines():
        line_norm = _norm(line)
        if not line_norm:
            continue
        if "PIX COPIA E COLA" in _texto_normalizado(line_norm) or "CHAVE DE ACESSO" in _texto_normalizado(line_norm):
            continue
        # Ignora linhas de item CCI ELEKTRO (0601, 0602, 0899, 0999...) que
        # acumulam muitos dígitos mas não são código de barras
        if re.match(r"^\d{4}\s+[A-Z]", line_norm.strip()):
            continue
        # Linha com banco + código de barras, possivelmente garbled pelo PDF extractor
        # (ex: "BANSCO ITAU eS A g341u-7 34n191.0966d9 96890.7a42933 ...")
        if re.search(r"BANCO|BANSCO", line_norm, re.IGNORECASE):
            all_digits = re.sub(r"\D", "", line_norm)
            # Localiza o início do barcode pelos prefixos conhecidos (Itaú, CELPE, COSERN...)
            for bar_prefix in ("34191", "83660", "83680", "83670", "83661", "83662"):
                idx = all_digits.find(bar_prefix)
                if idx >= 0:
                    candidate = all_digits[idx:]
                    if 44 <= len(candidate) <= 60:
                        return candidate[:48]
            continue
        digits = re.sub(r"\D", "", line_norm)
        if 44 <= len(digits) <= 60 and not digits.startswith("292"):
            return digits[:48]

    m = re.search(r"\b((?!292)\d{44,48})\b", re.sub(r"[ \t]", "", text))
    if m:
        return m.group(1)
    return ""


def _extract_valor_nota_fiscal(text: str) -> float:
    # Em layouts MT da Neoenergia, o "Valor N. Fiscal" usado no Consen
    # corresponde à base destacada em PIS/COFINS, não à base de ICMS.
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if not (" PIS " in f" {line_norm} " or " COFINS " in f" {line_norm} "):
            continue
        m = re.search(
            r"\b(?:PIS|COFINS)\b\s+([-\d\.,]+)\s+[\d\.,]+\s+[-\d\.,]+",
            line,
            flags=re.IGNORECASE,
        )
        if m:
            return abs(_to_float_br(m.group(1)))

    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if " ICMS " in f" {line_norm} ":
            m = re.search(r"\bICMS\b\s+([-\d\.,]+)\s+[\d\.,]+\s+[-\d\.,]+", line, flags=re.IGNORECASE)
            if m:
                return abs(_to_float_br(m.group(1)))

    valores = [abs(_to_float_br(v)) for v in RE_MONEY.findall(text)]
    if valores:
        return max(valores)
    return 0.0


def _extract_valor_nota_fiscal_mt(text: str) -> float:
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if " ICMS " not in f" {line_norm} ":
            continue
        m = re.search(
            r"\bICMS\b\s+([-\d\.,]+)\s+[\d\.,]+\s+[-\d\.,]+",
            line,
            flags=re.IGNORECASE,
        )
        if m:
            return abs(_to_float_br(m.group(1)))
    return _extract_valor_nota_fiscal(text)


def _extract_notafiscal(text: str) -> str:
    patterns = [
        r"Nota\s+Fiscal\s*(?:N[ouo]\.?|Numero)?\s*[:\-]?\s*([0-9]{6,20})",
        r"NF\s*(?:N[ouo]\.?|Numero)?\s*[:\-]?\s*([0-9]{6,20})",
        r"NOTA\s+FISCAL\s+N[O0]\s*[:\-]?\s*([0-9]{6,20})",
        r"NOTA\s+FISCAL\s+N[°ºO0]?\s*[:\-]?\s*([0-9]{6,20})",
        r"NOTA\s+FISCAL\s+N\S*\s*[:\-]?\s*([0-9]{6,20})",
    ]
    for p in patterns:
        m = re.search(p, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip()

    linhas = text.splitlines()
    for idx, line in enumerate(linhas):
        line_norm = _texto_normalizado(line)
        candidatos = [line]
        if "CHAVE DE ACESSO" in line_norm and idx + 1 < len(linhas):
            candidatos.append(linhas[idx + 1])
        for candidato in candidatos:
            digits = re.sub(r"\D", "", candidato)
            if len(digits) < 44:
                continue
            for start in range(0, len(digits) - 43):
                chave = digits[start:start + 44]
                if chave[20:22] != "66":
                    continue
                return str(int(chave[25:34]))
    return ""


def _extract_cnpj(text: str) -> str:
    labels = [
        r"CPF/CNPJ[:\s]+(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})",
        r"CNPJ[:\s]+(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})",
    ]
    for pat in labels:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return re.sub(r"\D", "", m.group(1))

    m = re.search(r"\b(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})\b", text)
    if not m:
        return ""
    return re.sub(r"\D", "", m.group(1))


def _extract_endereco(text: str) -> str:
    linhas = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for i, line in enumerate(linhas):
        line_n = _texto_normalizado(line)
        if "ENDERECO DE ENTREGA" in line_n or line_n.startswith("ENDERECO"):
            trecho = [line]
            for extra in linhas[i + 1:i + 3]:
                extra_n = _texto_normalizado(extra)
                if any(stop in extra_n for stop in ("CPF/CNPJ", "CNPJ", "INSTALACAO", "CLIENTE")):
                    break
                trecho.append(extra)
            return _norm(" ".join(trecho))

    for line in text.splitlines():
        if "endereco" in _strip_accents(line).lower():
            return _norm(line)
    return ""


def _extract_codigo_cliente(text: str) -> str:
    pats = [
        r"Codigo\s+do\s+Cliente[:\s]+(\d{4,20})",
        r"Codigo\s+Cliente[:\s]+(\d{4,20})",
        r"Cliente[:\s]+(\d{4,20})",
        r"Codigo\s+da\s+Unidade\s+Consumidora[:\s]+(\d{4,20})",
    ]
    for p in pats:
        m = re.search(p, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip()

    fallback_patterns = [
        r"\b(\d{6,12})\s+chave de acesso:",
        r"\b(\d{6,12})\s+\d{2}/\d{2}/\d{4}\s+[\d\.,-]+\s+\d{2}/\d{4}\s+C[ÓO]DIGO DO CLIENTE",
        r"N[°º]\s+DO\s+DOCUMENTO\s+C[ÓO]DIGO DO CLIENTE\s+DATA DE VENCIMENTO\s+VALOR DO DOCUMENTO\s+[\d\.]+\s+(\d{6,12})\s+\d{2}/\d{2}/\d{4}",
    ]
    for p in fallback_patterns:
        m = re.search(p, text, flags=re.IGNORECASE | re.DOTALL)
        if m:
            return m.group(1).strip()
    return ""


def _parse_date_field(value: str) -> dt.date | None:
    return _to_date(str(value or ""))


def _ultimo_dia_mes(ano: int, mes: int) -> dt.date:
    if mes == 12:
        return dt.date(ano, 12, 31)
    return dt.date(ano, mes + 1, 1) - dt.timedelta(days=1)


def _carimbo_do_nome(pdf_path: Path) -> str:
    stem = pdf_path.stem
    m = re.fullmatch(r"BB_(\d{7})", stem, re.IGNORECASE)
    if m:
        return m.group(1)
    return ""


def _estado_slug(nome: str) -> str:
    mapa = {
        "Bahia": "BAHIA",
        "Pernambuco": "PERNAMBUCO",
        "Rio Grande do Norte": "RIO_GRANDE_DO_NORTE",
        "Mato Grosso do Sul": "MATO_GROSSO_DO_SUL",
        "Sao Paulo": "SAO_PAULO",
        "São Paulo": "SAO_PAULO",
    }
    if nome in mapa:
        return mapa[nome]
    limpo = _strip_accents(nome or "").upper()
    limpo = re.sub(r"[^A-Z0-9]+", "_", limpo).strip("_")
    return limpo or "DESCONHECIDO"


def _tipo_sort_key(pdf_path: Path) -> int:
    m = RE_CARIMBO.search(pdf_path.stem)
    return int(m.group(1)) if m else 0


def _unique_paths(candidates: list[str]) -> list[Path]:
    out: list[Path] = []
    seen: set[str] = set()
    for raw in candidates:
        txt = str(raw or "").strip()
        if not txt:
            continue
        key = txt.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(Path(txt))
    return out


def _path_exists_safe(path: Path) -> tuple[bool, str]:
    try:
        return path.exists(), ""
    except OSError as exc:
        return False, f"{type(exc).__name__}: {exc}"


def _resolver_caminho_pdf(row: dict, carimbo: str) -> tuple[Path | None, str]:
    base = str(row.get("arquivo", "") or "").strip()
    if not base:
        base = ""

    def _corrigir_typo_download(txt: str) -> str:
        # Importante: usar lambda evita parse de escapes da string de replacement no re.sub.
        return re.sub(
            r"(?i)\\DOWLOAD NEOENERGIA\\",
            lambda _m: "\\DOWNLOAD NEOENERGIA\\",
            txt,
        )

    def _corrigir_host_legado(txt: str) -> str:
        prefix_legacy = "\\\\fs01\\energia\\arquivos enzo\\"
        if txt.lower().startswith(prefix_legacy):
            return "\\\\10.10.250.21\\Energia\\ARQUIVOS ENZO\\" + txt[len(prefix_legacy):]
        return txt

    candidatos: list[str] = []
    if base:
        candidatos.append(base)
        candidatos.append(_corrigir_typo_download(base))
        candidatos.append(_corrigir_host_legado(base))
        candidatos.append(_corrigir_typo_download(candidatos[-1]))

    # Fallback por estrutura esperada: DOWNLOAD NEOENERGIA/<ESTADO>/<YYYY-MM>/<CARIMBO>.pdf
    estado = _estado_slug(str(row.get("estado", "")))
    mm_yy = _entry_mes_ano(row)
    if mm_yy:
        mm, yy = mm_yy
        pasta_ref = f"{yy}-{mm:02d}"
        candidatos.append(str(DOWNLOAD_ROOT / estado / pasta_ref / f"{carimbo}.pdf"))
        # comum existir sufixo _2, _3...
        for i in range(2, 6):
            candidatos.append(str(DOWNLOAD_ROOT / estado / pasta_ref / f"{carimbo}_{i}.pdf"))

    erros: list[str] = []
    for p in _unique_paths(candidatos):
        ok, err = _path_exists_safe(p)
        if ok:
            return p, ""
        if err:
            erros.append(f"{p} -> {err}")

    if erros:
        return None, " | ".join(erros[:3])
    return None, f"arquivo nao encontrado (candidatos: {len(_unique_paths(candidatos))})"


def _carimbo_from_any(row: dict, pdf_path: Path) -> str:
    for key in ("id", "fatCarimbo", "carimbo"):
        val = _norm(str(row.get(key, "")))
        if val:
            return val
    return pdf_path.stem


def _extract_instalacao_from_words(words: list[dict]) -> str:
    if not words:
        return ""

    candidatos: list[tuple[float, str]] = []
    for w in words:
        txt = _texto_normalizado(str(w.get("text", "")))
        if "INSTALACAO" not in txt:
            continue

        x0 = float(w.get("x0", 0))
        x1 = float(w.get("x1", 0))
        bottom = float(w.get("bottom", 0))

        for ww in words:
            token = str(ww.get("text", "")).strip()
            if not re.fullmatch(r"\d{6,12}", token):
                continue
            top = float(ww.get("top", 0))
            wx0 = float(ww.get("x0", 0))
            wx1 = float(ww.get("x1", 0))
            if top < bottom - 2 or top > bottom + 20:
                continue
            if wx1 < x0 - 40 or wx0 > x1 + 20:
                continue
            score = abs(top - bottom) + abs(wx0 - x0) / 10.0
            candidatos.append((score, token))

    if candidatos:
        candidatos.sort(key=lambda item: (item[0], len(item[1]), item[1]))
        return candidatos[0][1]

    # Fallback: digitos podem estar fragmentados (um char por token, como em CELPE).
    # Reconstruir concatenando digitos individuais na linha abaixo do label.
    for w in words:
        txt = _texto_normalizado(str(w.get("text", "")))
        if "INSTALACAO" not in txt:
            continue

        x0 = float(w.get("x0", 0))
        x1 = float(w.get("x1", 0))
        bottom = float(w.get("bottom", 0))

        digitos: list[tuple[float, str]] = []
        for ww in words:
            token = str(ww.get("text", "")).strip()
            if not re.fullmatch(r"\d", token):
                continue
            top = float(ww.get("top", 0))
            wx0 = float(ww.get("x0", 0))
            wx1 = float(ww.get("x1", 0))
            if top < bottom - 2 or top > bottom + 25:
                continue
            if wx1 < x0 - 40 or wx0 > x1 + 20:
                continue
            digitos.append((wx0, token))

        if len(digitos) >= 6:
            digitos.sort(key=lambda d: d[0])
            numero = "".join(d[1] for d in digitos)
            if re.fullmatch(r"\d{6,15}", numero):
                return numero

    return ""


def _extract_instalacao(text: str, words: list[dict] | None = None) -> str:
    por_palavras = _extract_instalacao_from_words(words or [])
    if por_palavras:
        return por_palavras

    patterns = [
        r"Instalac[a?]o(?:\s+N[o?])?[:\s]+(\d{6,15})",
        r"Unidade\s+Consumidora[:\s]+(\d{6,15})",
        r"C[o?]digo\s+da\s+Instala[c?][a?]o[:\s]+(\d{6,15})",
        r"N[?u]mero\s+da\s+Instala[c?][a?]o[:\s]+(\d{6,15})",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip()

    line_patterns = [
        r"^\s*(\d{6,10})\s+https?://",
        r"^\s*(\d{6,10})\s+\d{2}/\d{2}/\d{4}\s+R\$",
        r"^\s*(\d{6,10})\s+\d{2}-\d{14,18}-\d{2}",
    ]
    for line in text.splitlines():
        for pat in line_patterns:
            m = re.search(pat, line, flags=re.IGNORECASE)
            if m:
                return m.group(1).strip()

    for line in text.splitlines():
        line_n = _texto_normalizado(line)
        if "INSTALACAO" in line_n or "UNIDADE CONSUMIDORA" in line_n:
            nums = re.findall(r"\d{6,15}", line)
            if nums:
                return nums[0]

    frequencia: dict[str, int] = {}
    for num in re.findall(r"\d{6,10}", text):
        frequencia[num] = frequencia.get(num, 0) + 1
    if frequencia:
        ordenados = sorted(frequencia.items(), key=lambda item: (-item[1], len(item[0]), item[0]))
        return ordenados[0][0]

    nums_all = re.findall(r"\d{6,15}", text)
    return nums_all[0] if nums_all else ""

def _detectar_tipo_tarifa(text: str) -> tuple[str, str, str, str]:
    txt = _texto_normalizado(text)

    mt_patterns = [
        r"\bGRUPO A\b",
        r"\bSUBGRUPO A\b",
        r"\bHORARIA VERDE\b",
        r"\bHORARIA AZUL\b",
        r"\bHORO[- ]?SAZONAL\b",
        r"\bDEMANDA\b",
        r"\bULTRAP(?:ASSAGEM|\.)\b",
        r"\bFORA DE PONTA\b",
        r"\bFORA PONTA\b",
        r"\bREATIVO\b",
        r"\bMONTANTE EM TODOS OS PERIODOS\b",
    ]
    bt_patterns = [
        r"\bGRUPO B\b",
        r"\bSUBGRUPO B\b",
        r"\bCONVENCIONAL\b",
        r"\bMONOFASICO\b",
        r"\bBIFASICO\b",
        r"\bTRIFASICO\b",
        r"\bBRANCA\b",
        r"\bBAIXA TENSAO\b",
    ]

    mt_hits = sum(1 for pat in mt_patterns if re.search(pat, txt))
    bt_hits = sum(1 for pat in bt_patterns if re.search(pat, txt))

    if re.search(r"\bA[1234]\b", txt) and ("SUBGRUPO" in txt or "GRUPO A" in txt):
        mt_hits += 2
    if re.search(r"\bB[123]\b", txt) and ("SUBGRUPO" in txt or "GRUPO B" in txt):
        bt_hits += 2

    tem_horaria = bool(re.search(r"\bHORARIA (VERDE|AZUL)\b", txt))
    tem_demanda = bool(re.search(r"\bDEMANDA\b", txt))
    tem_ponta = bool(re.search(r"\bPONTA\b", txt))

    if tem_horaria or (tem_demanda and tem_ponta) or mt_hits > bt_hits:
        subgrupo = "A4 [2,3kV a 25kV]"
        detected = "A4_VERDE"
        if "A3" in txt:
            subgrupo = "A3 [<44kV]"
            detected = "A3_VERDE"
        elif "A2" in txt:
            subgrupo = "A2"
            detected = "A2_VERDE"
        elif re.search(r"\bA1\b", txt):
            subgrupo = "A1"
            detected = "A1_VERDE"

        tarifa = "HS - Azul" if "AZUL" in txt else "HS - Verde"
        if tarifa == "HS - Azul":
            detected = detected.replace("_VERDE", "_AZUL")

        return "mt", tarifa, subgrupo, detected

    tarifa = "Branca" if re.search(r"\bBRANCA\b", txt) else "Convencional"
    return "bt", tarifa, "B3 [<2,3kV]", "B3_BRANCA" if tarifa == "Branca" else "B3"


def _entry_mes_ano(row: dict) -> tuple[int, int] | None:
    ref = _parse_ref_mes_ano(row.get("mes_referencia", ""))
    if ref:
        return ref

    arq = str(row.get("arquivo", ""))
    m = re.search(r"[/\\](\d{4})-(\d{2})[/\\]", arq)
    if m:
        return int(m.group(2)), int(m.group(1))
    return None


def _empty_record() -> dict:
    rec = {h: "" for h in HEADERS}
    for h in NUMERIC_HEADERS:
        rec[h] = 0.0
    rec["cadTarifaCod"] = "Convencional"
    rec["cadSubGrupoCod"] = "B3 [<2,3kV]"
    rec["TARIFA_DETECTADA"] = "B3"
    rec["fatDataCadastro"] = dt.date.today()
    return rec


def _extract_valor_por_labels(text: str, labels: Iterable[str]) -> float:
    labels_norm = [_texto_normalizado(lbl) for lbl in labels]
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if not line_norm:
            continue
        if any(lbl in line_norm for lbl in labels_norm):
            valores = RE_MONEY.findall(line)
            if valores:
                return _to_float_br(valores[-1])
    return 0.0


def _extract_ilum_publica(text: str) -> float:
    total = 0.0
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if "ILUM" not in line_norm and "COSIP" not in line_norm and "CIP" not in line_norm:
            continue
        m = re.search(
            r"(?:ILUM(?:INACAO)?\.?\s*P[UÚ]B\.?\s*MUNICIPAL|COSIP|CIP)\s+([-\d\.,]+)",
            line,
            flags=re.IGNORECASE,
        )
        if m:
            total += abs(_to_float_br(m.group(1)))
            continue
        valores = RE_MONEY.findall(line)
        if valores:
            total += abs(_to_float_br(valores[0]))
    return round(total, 2)


def _extract_debitos_anteriores(text: str) -> float:
    labels = [
        "CONTA ANTERIOR",
        "DEBITOS ANTERIORES",
        "DEBITO ANTERIOR",
        "CORRECAO MONETARIA POR ATRASO",
        "JUROS CONTA ANTERIOR",
        "MULTA CONTA ANTERIOR",
    ]
    total = 0.0
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if not line_norm:
            continue
        if any(lbl in line_norm for lbl in labels):
            valores = RE_MONEY.findall(line)
            if valores:
                total += _to_float_br(valores[-1])
    return round(total, 2)


def _extract_imposto(text: str, nome: str) -> float:
    nome_norm = _texto_normalizado(nome)
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if nome_norm not in line_norm:
            continue
        if any(bloqueio in line_norm for bloqueio in ("IMP.RET.", "COBR.", "COBRANCA", "ESCASSEZ HIDRICA")):
            continue
        m = re.search(
            rf"\b{re.escape(nome_norm)}\b\s+[\d\.,]+\s+[\d\.,]+%\s+([\d\.,]+)",
            line_norm,
            flags=re.IGNORECASE,
        )
        if m:
            return _to_float_br(m.group(1))
        valores = RE_MONEY.findall(line)
        if valores:
            return _to_float_br(valores[-1])
    return 0.0


def _extract_aliquotas(text: str) -> dict[str, float]:
    out = {
        "fatDescPisAliquota": 0.0,
        "fatDesCofinsAliquota": 0.0,
        "fatDesIcmsAliquota": 0.0,
    }

    for line in text.splitlines():
        line_norm = _texto_normalizado(line)

        if " PIS " in f" {line_norm} ":
            m = re.search(r"\bPIS\b\s+[-\d\.,]+\s+([\d\.,]+)\s*%?\s*[-\d\.,]+", line, flags=re.IGNORECASE)
            if m:
                out["fatDescPisAliquota"] = abs(_to_float_br(m.group(1)))

        if " COFINS " in f" {line_norm} ":
            m = re.search(r"\bCOFINS\b\s+[-\d\.,]+\s+([\d\.,]+)\s*%?\s*[-\d\.,]+", line, flags=re.IGNORECASE)
            if m:
                out["fatDesCofinsAliquota"] = abs(_to_float_br(m.group(1)))

        if " ICMS " in f" {line_norm} ":
            m = re.search(r"\bICMS\b\s+[-\d\.,]+\s+([\d\.,]+)\s*%?\s*[-\d\.,]+", line, flags=re.IGNORECASE)
            if m:
                out["fatDesIcmsAliquota"] = abs(_to_float_br(m.group(1)))

    return out


def _to_float_qty(raw: str) -> float:
    txt = _norm(str(raw or ""))
    if not txt:
        return 0.0
    txt = txt.replace(" ", "")
    if "," in txt:
        return _to_float_br(txt)
    if "." in txt and re.fullmatch(r"\d+(?:\.\d+)+", txt):
        return float(txt.replace(".", ""))
    try:
        return float(re.sub(r"[^\d\-]", "", txt) or 0)
    except Exception:
        return 0.0


_TRIB_FEDERAL_BREAKDOWN: dict[str, dict[str, float]] = {
    # total_perc -> {component: perc}
    "5.85": {"IRPJ": 1.20, "PIS": 0.65, "COFINS": 3.00, "CSLL": 1.00},
    "9.45": {"IRPJ": 4.80, "PIS": 0.65, "COFINS": 3.00, "CSLL": 1.00},
}

# Concessionárias Neoenergia — faturas MT têm retenção fixa de 9,45%
_NEOENERGIA_CONCS = {"COELBA", "CELPE", "COSERN"}
ALIQUOTA_RETENCAO_NEO_MT = 9.45


def _aplicar_retencao_neo_mt(rec: dict) -> None:
    """Consolida a retenção MT da Neoenergia como um único Tributo Federal.

    No Consen, essas faturas devem sair com `fatTributoFederalPerc = -1` e
    `fatTributoFederalVal` igual à soma total das retenções. Os componentes
    individuais (PIS/COFINS/CSLL/IRPJ) não devem ser digitados separadamente.
    """
    mapa = {
        "IRPJ": ("fatDescIrpjPercRetImposto", "fatDescIrpjValRetImposto"),
        "PIS": ("fatDescPisPercRetImposto", "fatDescPisValRetImposto"),
        "COFINS": ("fatDescCofinsPercRetImposto", "fatDescCofinsValRetImposto"),
        "CSLL": ("fatDescCsllPercRetImposto", "fatDescCsllValRetImposto"),
    }
    total_val = abs(float(rec.get("fatTributoFederalVal") or 0.0))
    if not total_val:
        total_val = round(
            sum(abs(float(rec.get(campo_val) or 0.0)) for _, campo_val in mapa.values()),
            2,
        )
    if total_val > 0:
        rec["fatTributoFederalPerc"] = -1.0
        rec["fatTributoFederalVal"] = -round(total_val, 2)
    _zerar_retencoes_individuais(rec)


def _extract_retencoes(text: str) -> dict[str, float]:
    out = {
        "fatDescPisPercRetImposto": 0.0,
        "fatDescPisValRetImposto": 0.0,
        "fatDescCofinsPercRetImposto": 0.0,
        "fatDescCofinsValRetImposto": 0.0,
        "fatDescCsllPercRetImposto": 0.0,
        "fatDescCsllValRetImposto": 0.0,
        "fatDescIrpjPercRetImposto": 0.0,
        "fatDescIrpjValRetImposto": 0.0,
    }
    mapa = {
        "PIS": ("fatDescPisPercRetImposto", "fatDescPisValRetImposto"),
        "COFINS": ("fatDescCofinsPercRetImposto", "fatDescCofinsValRetImposto"),
        "CSLL": ("fatDescCsllPercRetImposto", "fatDescCsllValRetImposto"),
        "IRPJ": ("fatDescIrpjPercRetImposto", "fatDescIrpjValRetImposto"),
    }
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if ("IMP.RET." in line_norm or "IMPOSTO RETIDO" in line_norm) and "EST.IMP.RET." not in line_norm:
            for nome, (campo_perc, campo_val) in mapa.items():
                if nome not in line_norm:
                    continue
                m_perc = re.search(r"\(([\d.,]+)%\)", line)
                if m_perc:
                    out[campo_perc] = abs(_to_float_br(m_perc.group(1)))
                valores = RE_MONEY.findall(line)
                if valores:
                    out[campo_val] = -abs(_to_float_br(valores[-1]))
        elif "TRIB.FEDERAL" in line_norm or "TRIB FEDERAL" in line_norm:
            m_perc = re.search(r"TRIB\.?FEDERAL\s*\(([\d\.,]+)%\)", line, flags=re.IGNORECASE)
            if not m_perc:
                continue
            raw_perc = m_perc.group(1)
            if "," in raw_perc and "." in raw_perc:
                txt_perc = raw_perc.replace(".", "").replace(",", ".")
            elif "," in raw_perc:
                txt_perc = raw_perc.replace(",", ".")
            else:
                txt_perc = raw_perc
            try:
                total_perc = abs(float(txt_perc))
            except Exception:
                continue
            valores = RE_MONEY.findall(line)
            if not valores:
                continue
            total_val = abs(_to_float_br(valores[-1]))
            key = f"{total_perc:.2f}".rstrip("0").rstrip(".")
            breakdown = _TRIB_FEDERAL_BREAKDOWN.get(f"{total_perc:.2f}") or _TRIB_FEDERAL_BREAKDOWN.get(key)
            if breakdown and total_perc > 0:
                for nome, comp_perc in breakdown.items():
                    campo_perc_key, campo_val_key = mapa[nome]
                    out[campo_perc_key] = comp_perc
                    out[campo_val_key] = -round(total_val * comp_perc / total_perc, 2)
    return out


def _extract_est_retencoes(text: str) -> float:
    """Soma dos EST.IMP.RET. — estimativas de períodos anteriores a compensar (Elektro)."""
    total = 0.0
    for line in text.splitlines():
        if "EST.IMP.RET." not in _texto_normalizado(line):
            continue
        valores = RE_MONEY.findall(line)
        if valores:
            total += abs(_to_float_br(valores[-1]))
    return round(total, 2)


def _extract_tributo_federal_linhas(text: str) -> list[tuple[float, float]]:
    linhas: list[tuple[float, float]] = []

    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if "TRIB.FEDERAL" not in line_norm and "TRIB FEDERAL" not in line_norm:
            continue

        m_perc = re.search(r"TRIB\.?FEDERAL\s*\(([\d\.,]+)%\)", line, flags=re.IGNORECASE)
        if not m_perc:
            continue

        txt_perc = m_perc.group(1).strip().replace(" ", "")
        if "," in txt_perc and "." in txt_perc:
            txt_perc = txt_perc.replace(".", "").replace(",", ".")
        elif "," in txt_perc:
            txt_perc = txt_perc.replace(",", ".")
        try:
            perc = abs(float(txt_perc))
        except Exception:
            perc = 0.0

        m_val = re.search(r"TRIB\.?FEDERAL\s*\([\d\.,]+%\)\s*([-\d\.,]+-?)", line, flags=re.IGNORECASE)
        if m_val:
            val = _to_float_br(m_val.group(1))
        else:
            valores = RE_MONEY.findall(line)
            val = _to_float_br(valores[-1]) if valores else 0.0

        linhas.append((perc, val))

    return linhas


def _extract_tributo_federal(text: str) -> tuple[float, float]:
    """Retorna (perc_tributo_federal, val_tributo_federal).

    Quando ha duas linhas Trib.Federal (ex: 5,85% e 9,45%):
      - perc = 9,45% (aliquota principal)
      - val  = soma dos dois valores negativos

    Quando ha apenas uma linha:
      - perc = % da linha
      - val  = valor da linha
    """
    linhas = _extract_tributo_federal_linhas(text)

    if not linhas:
        return 0.0, 0.0

    if len(linhas) >= 2:
        # Mantém alíquota 9,45% e soma todos os valores negativos
        val_total = sum(v for _, v in linhas)
        return 9.45, val_total

    perc_out, val_out = linhas[-1]
    return perc_out, val_out


def _extract_bandeiras(text: str) -> dict[str, float]:
    out = {
        "fatValBandeira": 0.0,
        "fatValBandeira2": 0.0,
    }
    money_re = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?")

    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        is_band = (
            "BANDEIRA" in line_norm
            or " BAND. " in f" {line_norm} "
            or " BAND " in f" {line_norm} "
            or " AD.B." in f" {line_norm} "  # Elektro: "AD.B.AMAR." (Adicional Bandeira Amarela)
        )
        if not is_band:
            continue
        if "JA INCLUID" in line_norm or "INCLUIDO NO VALOR A PAGAR" in line_norm:
            continue

        valores = money_re.findall(line)
        if not valores:
            continue
        # AD.B.AMAR. (Elektro): qty é inteiro (sem vírgula → não capturado por money_re),
        # então: valores[0]=rate truncado p/ 2 casas, valores[1]=valor real da bandeira.
        if " AD.B." in f" {line_norm} " and len(valores) >= 2:
            valor = _to_float_br(valores[1])
        elif any(tag in line_norm for tag in ("ACRES. BAND", "ACRES BAND", "ADIC. BAND", "ADICIONAL BAND")):
            # COELBA longo: "Acrés. Band. AMARELA {band_rs} ... ICMS {base} {icms%} {icms_total}"
            # valores[0] é a bandeira; valores[-1] é o ICMS total da fatura — não bandeira.
            # Formato curto (apenas taxa + valor sem ICMS inline): valores[-1] é o total correto.
            if "ICMS" in line_norm and len(valores) >= 3:
                valor = _to_float_br(valores[0])
            else:
                valor = _to_float_br(valores[-1])
        else:
            valor = _to_float_br(valores[-1])
        if abs(valor) < 0.01:
            continue

        if valor < 0 or "INJET" in line_norm:
            out["fatValBandeira2"] += valor
        else:
            out["fatValBandeira"] += abs(valor)

    out["fatValBandeira"] = round(out["fatValBandeira"], 2)
    out["fatValBandeira2"] = round(out["fatValBandeira2"], 2)
    return out


def _extract_multas(text: str) -> float:
    total = 0.0
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if not any(tag in line_norm for tag in ("MULTA-NF", "JUROS-NF")):
            continue
        valores = RE_MONEY.findall(line)
        if valores:
            total += abs(_to_float_br(valores[-1]))
    return round(total, 2)


def _extract_multas_diversas(text: str) -> float:
    """ICMS-CDE NF... → fatMultasDiversas."""
    total = 0.0
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)
        if "ICMS-CDE" not in line_norm and "ICMS CDE" not in line_norm:
            continue
        valores = RE_MONEY.findall(line)
        if valores:
            total += abs(_to_float_br(valores[-1]))
    return round(total, 2)


_OBS_RULES: list[tuple[str, int, bool]] = [
    # (padrao_normalizado, codigo_consen, negativo)
    ("IMP.SOM/DIM",   131, True),   # IMPORTE A SOMAR OU DIMINUIR (S/IMPOSTO)
    ("IMPORTE A SOMAR OU DIMINUIR", 131, True),
    ("VALOR IMPORT",  131, True),
    ("SOMAR/DIMINUIR", 131, True),   # variante COELBA BT
    ("RESSARCIMENTO",  131, True),   # ressarcimento genérico COELBA BT
    ("COMP.DIC",        58, True),   # Compensação DIC mensal (qualidade de serviço)
    ("COMP. DIC",       58, True),   # variante com espaço (COELBA)
    ("COMP.FIC",        11, True),   # Compensação FIC
    ("DIC/FIC",         11, True),   # Compensação combinada
    ("COMP.DIC/FIC",    11, True),
    ("IPCA-NF",        273, False),  # IPCA
    ("SEGUNDA VIA",     51, False),  # Segunda Via da Fatura
]


def _extract_observacoes(text: str, is_compensacao: bool = False) -> list[tuple[int, float]]:
    resultado: list[tuple[int, float]] = []
    for line in text.splitlines():
        if len(resultado) >= 5:
            break
        line_norm = _texto_normalizado(line)
        for pattern, code, negative in _OBS_RULES:
            if pattern in line_norm:
                # Para unidades de compensação GD, o Imp.Som/Dim-S/Impost é o crédito
                # injetado — vai para fatConFPontaInjetadoValorReais, não para obs 131.
                if is_compensacao and "IMP.SOM/DIM" in pattern:
                    break
                valores = RE_MONEY.findall(line)
                if valores:
                    val = _to_float_br(valores[0])
                    val = -abs(val) if negative else abs(val)
                    resultado.append((code, round(val, 2)))
                break
    return resultado


def _is_sistema_compensacao(text: str) -> bool:
    return "UNIDADE INTEGRANTE DE SISTEMA DE COMPENSACAO" in _texto_normalizado(text)


def _extract_bt_rules(text: str) -> dict[str, float]:
    out: dict[str, float] = {}
    val_te = 0.0
    val_tusd = 0.0
    qtd_consumo = 0.0
    inj_ponta_val = 0.0
    inj_fponta_val = 0.0
    inj_ponta_qtds: list[float] = []
    # Grouped by injection-type tag so that TE+TUSD rows (same kWh) are
    # deduped via max() per group and then summed across groups.
    # Non-Elektro bills have one group (""); Elektro GDI-I has "MPT"+"OPT".
    inj_fponta_groups: dict[str, list[float]] = {}

    # Tarifa Branca (Elektro): acumuladores por período
    branca: dict[str, dict[str, float]] = {
        "ponta":        {"qtd": 0.0, "val": 0.0},
        "fora_ponta":   {"qtd": 0.0, "val": 0.0},
        "intermediario": {"qtd": 0.0, "val": 0.0},
    }

    in_info_block = False
    for line in text.splitlines():
        line_norm = _texto_normalizado(line)

        if "INFORMACOES IMPORTANTES" in line_norm:
            in_info_block = True

        # Tarifa Branca: "CONSUMO PONTA TE/TUSD", "CONSUMO FORA PONTA TE/TUSD",
        # "CONSUMO INTERMEDIARIO TE/TUSD"
        _m_branca = re.match(
            r"CONSUMO\s+(PONTA|FORA\s+PONTA|INTERMEDIARIO)\s+(TE|TUSD)\b",
            line_norm,
        )
        if _m_branca:
            periodo_raw = _m_branca.group(1).replace(" ", "_").lower()  # ponta / fora_ponta / intermediario
            m = re.search(r"kWh\s+([-\d\.,]+)\s+[-\d\.,]+\s+([-\d\.,]+)", line, flags=re.IGNORECASE)
            if m and periodo_raw in branca:
                qtd = _to_float_qty(m.group(1))
                val = _to_float_br(m.group(2))
                branca[periodo_raw]["qtd"] = max(branca[periodo_raw]["qtd"], qtd)
                branca[periodo_raw]["val"] += val
            continue

        # CELPE Tarifa Branca: "CONSUMO-TUSD INTERM. KWH" / "CONSUMO-TE INTERMED. KWH"
        # Formato diferente do Elektro (que usa "CONSUMO INTERMEDIARIO TUSD/TE").
        _m_celpe_b = re.match(
            r"CONSUMO[- ](TUSD|TE)\s+(N\.?\s*PONTA|NA\s+PONTA|INTERM|F\.?\s*PONTA|FORA\s+PONTA)",
            line_norm,
        )
        if _m_celpe_b:
            m = re.search(r"kWh\s+([-\d\.,]+)\s+[-\d\.,]+\s+([-\d\.,]+)", line, flags=re.IGNORECASE)
            if m:
                qtd = _to_float_qty(m.group(1))
                val = _to_float_br(m.group(2))
                periodo_key = _m_celpe_b.group(2)
                if "INTERM" in periodo_key:
                    branca["intermediario"]["qtd"] = max(branca["intermediario"]["qtd"], qtd)
                    branca["intermediario"]["val"] += val
                elif "F" in periodo_key or "FORA" in periodo_key:
                    branca["fora_ponta"]["qtd"] = max(branca["fora_ponta"]["qtd"], qtd)
                    branca["fora_ponta"]["val"] += val
                else:
                    branca["ponta"]["qtd"] = max(branca["ponta"]["qtd"], qtd)
                    branca["ponta"]["val"] += val
            continue

        if re.search(r"\bCONSUMO[- ]TE\b", line_norm) or re.search(r"\bCONSUMO[- ]TUSD\b", line_norm):
            m = re.search(r"kWh\s+([-\d\.,]+)\s+[-\d\.,]+\s+([-\d\.,]+)", line, flags=re.IGNORECASE)
            if m:
                qtd = _to_float_qty(m.group(1))
                val = _to_float_br(m.group(2))
                qtd_consumo = max(qtd_consumo, qtd)
                if re.search(r"\bCONSUMO[- ]TE\b", line_norm):
                    val_te += val
                if re.search(r"\bCONSUMO[- ]TUSD\b", line_norm):
                    val_tusd += val

        # Saldo acumulado de créditos GD (Elektro e outras)
        if not in_info_block and (
            "SALDO" in line_norm
            and ("ACUMULADO" in line_norm or "CRED" in line_norm or "GDI" in line_norm)
            and ("KWH" in line_norm or "KW" in line_norm)
            and "EXPIRAR" not in line_norm
        ):
            m_saldo = re.search(r"([\d.]+,\d+)\s*(?:KWH|KW)", line, re.IGNORECASE)
            if m_saldo:
                out["fatConFPontaInjetadoUsinaSaldoAcumulado"] = abs(_to_float_qty(m_saldo.group(1)))

        # Energia injetada: somente fora do bloco "INFORMACOES IMPORTANTES"
        # "GDI" cobre tanto "GDI-I" (SP/Bahia) quanto "GDI" simples (CELPE/PE)
        if not in_info_block and "ENERGIA INJ." in line_norm and "GDI" in line_norm:
            m = re.search(r"kWh\s+([-\d\.,]+)\s+[-\d\.,]+\s+([-\d\.,]+)", line, flags=re.IGNORECASE)
            if not m:
                continue
            qtd = abs(_to_float_qty(m.group(1)))
            val = abs(_to_float_br(m.group(2)))
            # OPT = período de ponta para injeção GD (inclusive Elektro GDI-I)
            is_ponta = " OPT " in f" {line_norm} "
            if is_ponta:
                inj_ponta_qtds.append(qtd)
                inj_ponta_val += val
            else:
                # Grupo por tipo de injeção: separa MPT de OPT e GDI-I sem tag
                # para que cada par TE/TUSD seja deduplicado por max() dentro do
                # seu grupo antes de somar grupos distintos.
                if "GDI-I" in line_norm:
                    fp_tag = "MPT" if " MPT " in f" {line_norm} " else ""
                else:
                    fp_tag = ""
                inj_fponta_groups.setdefault(fp_tag, []).append(qtd)
                inj_fponta_val += val

    # Tarifa Branca: períodos independentes têm precedência sobre o bloco simples
    _branca_usada = False
    for periodo, chave_reg, chave_fat, chave_val in (
        ("ponta",        "fatConPontaRegistrado",        "fatConPontaFaturado",        "fatConPontaValorReais"),
        ("fora_ponta",   "fatConFPontaIndRegistrado",    "fatConFPontaIndFaturado",    "fatConFPontaIndValorReais"),
        ("intermediario", "fatConIntermediarioRegistrado", "fatConIntermediarioFaturado", "fatConIntermediarioValorReais"),
    ):
        b = branca[periodo]
        if b["qtd"] > 0:
            out[chave_reg] = b["qtd"]
            out[chave_fat] = b["qtd"]
            out[chave_val] = round(b["val"], 2)
            _branca_usada = True

    if not _branca_usada and qtd_consumo > 0:
        out["fatConFPontaIndRegistrado"] = qtd_consumo
        out["fatConFPontaIndFaturado"] = qtd_consumo
        out["fatConFPontaIndValorReais"] = round(val_te + val_tusd, 2)

    if inj_ponta_qtds:
        inj_ponta_kwh = max(inj_ponta_qtds)
        out["fatConPontaInjetadoRegistrado"] = inj_ponta_kwh
        out["fatConPontaInjetadoFaturado"] = inj_ponta_kwh
        out["fatConPontaInjetadoValorReais"] = round(inj_ponta_val, 2)

    if inj_fponta_groups:
        inj_fponta_kwh = sum(max(qtds) for qtds in inj_fponta_groups.values() if qtds)
        out["fatConFPontaInjetadoRegistrado"] = inj_fponta_kwh
        out["fatConFPontaInjetadoFaturado"] = inj_fponta_kwh
        out["fatConFPontaInjetadoValorReais"] = round(inj_fponta_val, 2)
    elif inj_fponta_val > 0:
        out["fatConFPontaInjetadoValorReais"] = round(inj_fponta_val, 2)

    # GD por compensação (CELPE/PE): texto indica créditos em kWh e R$ via Imp.Som/Dim-S/Impost
    # Só ativa quando não há linha "ENERGIA INJ." com GDI (evita sobreposição com Elektro/Bahia)
    if _is_sistema_compensacao(text) and not inj_fponta_groups:
        text_norm = _texto_normalizado(text)
        m_kwh = re.search(
            r"TOTAL DE CREDITOS UTILIZADOS NA UNIDADE:\s*([\d\.,]+)\s*KWH",
            text_norm,
        )
        if m_kwh:
            kwh_comp = _to_float_qty(m_kwh.group(1))
            val_comp = 0.0
            for line in text.splitlines():
                ln = _texto_normalizado(line)
                if "IMP.SOM/DIM" in ln:
                    valores = RE_MONEY.findall(line)
                    if valores:
                        val_comp = abs(_to_float_br(valores[0]))
                    break
            out["fatConFPontaInjetadoRegistrado"] = kwh_comp
            out["fatConFPontaInjetadoFaturado"] = kwh_comp
            if val_comp > 0:
                out["fatConFPontaInjetadoValorReais"] = round(val_comp, 2)

    return out


def _zerar_retencoes_individuais(rec: dict[str, object]) -> None:
    for campo in (
        "fatDescPisPercRetImposto",
        "fatDescPisValRetImposto",
        "fatDescCofinsPercRetImposto",
        "fatDescCofinsValRetImposto",
        "fatDescCsllPercRetImposto",
        "fatDescCsllValRetImposto",
        "fatDescIrpjPercRetImposto",
        "fatDescIrpjValRetImposto",
    ):
        rec[campo] = 0.0


def _parse_elektro_danfe_line(line: str, unit: str) -> tuple[float | None, float | None, float | None]:
    """Formato DANFE ELEKTRO MT (layout atual):
    DESCR UNIT QUANT PRECO_UNIT VALOR_R$ PIS_COF BASE_ICMS ALIQ% ICMS_VAL ...
    Retorna (quant, quant, valor_reais).
    """
    m = re.search(
        rf"\b{re.escape(unit)}\b"
        r"\s+([\d\.,]+)"
        r"\s+\S+"
        r"\s+([\d\.,]+)",
        line,
        re.IGNORECASE,
    )
    if not m:
        return None, None, None
    quant = _to_float_qty(m.group(1))
    valor = _to_float_br(m.group(2))
    if not quant:
        return None, None, None
    return quant, quant, valor if valor else None


def _parse_elektro_cci_line(line: str, unit: str) -> tuple[float | None, float | None, float | None]:
    """Formato CCI ELEKTRO MT:
    CCI descr UNIT prev_p prev_fp tariff registered 0 billed tarifa base_s base_c rate% icms_val valor_reais
    Retorna (registered, billed, valor_reais).
    """
    m = re.search(
        rf"\b{re.escape(unit)}\b"
        r"\s+\S+"           # prev_ponta
        r"\s+\S+"           # prev_fp
        r"\s+\S+"           # tariff_period
        r"\s+([\d\.,]+)"    # registered (group 1)
        r"\s+\S+"           # 0,00
        r"\s+([\d\.,]+)"    # billed (group 2)
        r"\s+\S+"           # tarifa unitária
        r"\s+\S+"           # base sem ICMS
        r"\s+\S+"           # base com ICMS
        r"\s+\S+"           # ICMS rate%
        r"\s+\S+"           # ICMS valor
        r"\s+([-\d\.,]+)",  # valor reais (group 3)
        line,
        re.IGNORECASE,
    )
    if not m:
        return None, None, None
    return _to_float_qty(m.group(1)), _to_float_qty(m.group(2)), _to_float_br(m.group(3))


def _parse_mt_item_qtd_valor(line: str, unit_pattern: str) -> tuple[float | None, float | None]:
    m = re.search(
        rf"\b{unit_pattern}\b\s+([-\d\.,]+)\s+([-\d\.,]+)\s+([-\d\.,]+)",
        line,
        flags=re.IGNORECASE,
    )
    if not m:
        return None, None
    return _to_float_qty(m.group(1)), _to_float_br(m.group(3))


def _parse_mt_medidor_registrado(line: str) -> float | None:
    valores = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", line)
    if len(valores) < 2:
        return None
    penultimo = _to_float_br(valores[-2])
    ultimo = _to_float_br(valores[-1])
    if abs(penultimo - ultimo) < 0.01:
        return penultimo
    return ultimo


def _aplicar_desconto_fio_a4_comercial(out: dict[str, float], text: str) -> None:
    """Aplica percentuais operacionais do desconto fio A4 Comercial.

    Em faturas MT da Neoenergia, o PDF pode trazer apenas o benefício/desconto
    em R$ ("Desconto incondicional" / "Desconto sobre tarifa"), sem estruturar
    no texto os percentuais do fio kW e kWh. Nesses casos operacionais usamos
    os percentuais homologados para A4 Comercial.
    """
    text_norm = _texto_normalizado(text)
    tem_beneficio = abs(float(out.get("fatBeneficioTarifarioBrutoValorReais") or 0.0)) > 0.0
    if not tem_beneficio:
        return
    if "A4 COMERCIAL" not in text_norm:
        return
    out["fatDescontoFio"] = 50.0
    out["fatDescontoFioKWh"] = 47.48


def _extract_mt_rules_coelba(text: str) -> dict[str, float]:
    out: dict[str, float] = {}
    linhas = [ln.strip() for ln in text.splitlines() if ln.strip()]

    demanda_fp_reg = None
    demanda_fp_val = 0.0
    demanda_fp_lim_val = 0.0
    demanda_fp_contr = None
    consumo_p_reg = None
    consumo_p_fat = None
    consumo_p_val = None
    consumo_fp_reg = None
    consumo_fp_fat = None
    consumo_fp_val = None
    consumo_exc_p_reg = None
    consumo_exc_p_fat = None
    consumo_exc_p_val = None
    consumo_exc_fp_reg = None
    consumo_exc_fp_fat = None
    consumo_exc_fp_val = None
    demanda_p_reg = None
    fio_lim_qtd = 0.0
    escassez_val = 0.0
    aliq_pis = None
    aliq_cofins = None
    aliq_icms = None
    ret_trib_federal = 0.0
    cob_icms_cde = 0.0
    dif_icms_cde = 0.0
    dif_desc_fp = 0.0
    dif_desc_fp_sem_icms = 0.0
    dif_desc_np = 0.0

    for line in linhas:
        line_norm = _texto_normalizado(line)

        # Tributos (bloco direito): "PIS <base> <aliquota> <valor>", idem COFINS/ICMS.
        # Captura a alíquota (% — número pequeno entre base e valor monetário).
        if aliq_pis is None:
            m = re.search(r"\bPIS\s+[\d\.]+,\d{2}\s+(\d{1,2},\d{2})\s+[\d\.]+,\d{2}", line)
            if m:
                aliq_pis = _to_float_br(m.group(1))
        if aliq_cofins is None:
            m = re.search(r"\bCOFINS\s+[\d\.]+,\d{2}\s+(\d{1,2},\d{2})\s+[\d\.]+,\d{2}", line)
            if m:
                aliq_cofins = _to_float_br(m.group(1))
        if aliq_icms is None:
            m = re.search(r"\bICMS\s+[\d\.]+,\d{2}\s+(\d{1,3},\d{2})\s+[\d\.]+,\d{2}", line)
            if m:
                aliq_icms = _to_float_br(m.group(1))

        # Retenção de tributo federal: "Trib.Federal(9.45%) 809,73-" (pode haver
        # mais de uma linha — soma todas). Também cobre IRF/IRRF/RETENCAO.
        if "TRIB.FEDERAL" in line_norm or "RETENCAO" in line_norm or "IRRF" in line_norm or "IRF" in line_norm:
            valores = RE_MONEY.findall(line)
            if valores:
                ret_trib_federal += abs(_to_float_br(valores[-1]))
            continue

        if "MONTANTE DE USO CONTRATADO" in line_norm:
            nums = re.findall(r"[-\d\.,]+", line)
            if nums:
                demanda_fp_contr = _to_float_qty(nums[-1])

        # "Demanda Contratada 57" (GRANDEZAS CONTRATADAS) — fonte da contratada
        # quando não há "MONTANTE DE USO CONTRATADO". Usado também como fallback
        # de registrada/faturada em faturas com demanda medida zerada.
        if "DEMANDA CONTRATADA" in line_norm and demanda_fp_contr is None:
            nums = re.findall(r"\d[\d\.]*", line.split("CONTRATADA", 1)[-1])
            if nums:
                v = _to_float_qty(nums[-1])
                if v and v > 0:
                    demanda_fp_contr = v

        if "USO SISTEMA FIO-LIM" in line_norm:
            qtd, val = _parse_mt_item_qtd_valor(line, r"KW")
            if qtd is not None:
                fio_lim_qtd = qtd
            if val is not None:
                demanda_fp_lim_val += val
            continue

        if "USO SISTEMA FIO" in line_norm and "FIO-LIM" not in line_norm:
            qtd, val = _parse_mt_item_qtd_valor(line, r"KW")
            if qtd is not None:
                demanda_fp_reg = qtd
            if val is not None:
                demanda_fp_val += val
            continue

        if "USO SISTEMA ENCAR.NP" in line_norm:
            qtd, val = _parse_mt_item_qtd_valor(line, r"KWH")
            if qtd is not None:
                consumo_p_fat = qtd
            if val is not None:
                consumo_p_val = val
            continue

        if "USO SISTEMA ENCAR.FP" in line_norm:
            qtd, val = _parse_mt_item_qtd_valor(line, r"KWH")
            if qtd is not None:
                consumo_fp_fat = qtd
            if val is not None:
                consumo_fp_val = val
            continue

        if "CONSUMO ATIVO NA PONTA" in line_norm:
            consumo_p_reg = _parse_mt_medidor_registrado(line)
            continue

        if "CONSUMO ATIVO FORA DE PONTA" in line_norm:
            consumo_fp_reg = _parse_mt_medidor_registrado(line)
            continue

        if "CONS.REAT.EXC.NPONTA" in line_norm or "CONSUMO REATIVO EXCEDENTE NA PONTA" in line_norm:
            qtd, val = _parse_mt_item_qtd_valor(line, r"KVARH")
            if qtd is not None:
                consumo_exc_p_reg = qtd
                consumo_exc_p_fat = qtd
            if val is not None:
                consumo_exc_p_val = val
            continue

        if "CONS.REAT EXC.FPONTA" in line_norm or "CONSUMO REATIVO EXCEDENTE FORA DE PONTA" in line_norm:
            qtd, val = _parse_mt_item_qtd_valor(line, r"KVARH")
            if qtd is not None:
                consumo_exc_fp_reg = qtd
                consumo_exc_fp_fat = qtd
            if val is not None:
                consumo_exc_fp_val = val
            continue

        if "DEMANDA MAXIMA NA PONTA" in line_norm:
            demanda_p_reg = _parse_mt_medidor_registrado(line)
            continue

        if "IMP.SOM/DIM-C/IMPOST" in line_norm:
            # Importe a Somar/Diminuir c/ Imposto. Importe POSITIVO = Escassez
            # Hídrica (nunca observação). Importe negativo é ignorado aqui.
            valores = re.findall(r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?", line)
            if valores:
                v = _to_float_br(valores[0])
                if v > 0:
                    escassez_val = v
            continue

        if "COB. ICMS SUBVEN-CDE" in line_norm:
            valores = RE_MONEY.findall(line)
            if valores:
                cob_icms_cde = abs(_to_float_br(valores[-1]))
            continue

        if "C.ICMSCDE DIF.FT.ALT" in line_norm:
            valores = re.findall(r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?", line)
            if valores:
                dif_icms_cde = _to_float_br(valores[-1])
            continue

        if any(
            rotulo in line_norm
            for rotulo in ("DIF. DESC. FT.ALTER.", "DIF.DES.F.ALT-S/ICMS", "DIF.DES.FT.ALT-ENCA.")
        ):
            valores = RE_MONEY.findall(line)
            if valores:
                val = abs(_to_float_br(valores[0]))
                if "DIF. DESC. FT.ALTER." in line_norm:
                    dif_desc_fp = val
                elif "DIF.DES.F.ALT-S/ICMS" in line_norm:
                    dif_desc_fp_sem_icms = val
                else:
                    dif_desc_np = val
            continue

        if "DEMANDA ULTRAP." in line_norm:
            qtd, val = _parse_mt_item_qtd_valor(line, r"KW")
            if qtd is not None:
                out["fatDemPontaUltra"] = qtd
            if val is not None:
                out["fatDemPontaUltraValorReais"] = round(val, 2)

    if demanda_p_reg is not None:
        out["fatDemPontaRegistrada"] = demanda_p_reg

    if demanda_fp_reg is not None:
        out["fatDemFPontaIndRegistrada"] = demanda_fp_reg
        if demanda_fp_contr is not None:
            out["fatDemContratadaFPonta"] = demanda_fp_contr
            out["fatDemFPontaIndFaturada"] = max(demanda_fp_reg, demanda_fp_contr)
        else:
            out["fatDemFPontaIndFaturada"] = demanda_fp_reg
    elif demanda_fp_contr is not None:
        out["fatDemContratadaFPonta"] = demanda_fp_contr

    # Fallback zero-billing: demanda medida 0 mas há contratada — o CONSEN
    # espera o valor contratado em registrada/faturada para não ficar zerado.
    if demanda_fp_contr is not None and demanda_fp_contr > 0:
        if not out.get("fatDemFPontaIndRegistrada"):
            out["fatDemFPontaIndRegistrada"] = demanda_fp_contr
        if not out.get("fatDemFPontaIndFaturada"):
            out["fatDemFPontaIndFaturada"] = demanda_fp_contr

    if demanda_fp_val or demanda_fp_lim_val:
        demanda_fp_total = demanda_fp_val + demanda_fp_lim_val
        if (
            demanda_fp_contr is not None
            and demanda_fp_reg is not None
            and demanda_fp_contr >= 100
            and demanda_fp_contr > demanda_fp_reg
            and fio_lim_qtd > 0
        ):
            delta_kw = demanda_fp_contr - demanda_fp_reg
            ajuste_delta = round(delta_kw * 2.1027027027, 2)
            demanda_fp_total += ajuste_delta
        out["fatDemFPontaIndValorReais"] = round(demanda_fp_total, 2)

    if consumo_p_reg is not None:
        out["fatConPontaRegistrado"] = consumo_p_reg
    if consumo_p_fat is not None:
        out["fatConPontaFaturado"] = consumo_p_fat
    if consumo_p_val is not None:
        out["fatConPontaValorReais"] = round(consumo_p_val, 2)

    if consumo_fp_reg is not None:
        out["fatConFPontaIndRegistrado"] = consumo_fp_reg
    if consumo_fp_fat is not None:
        out["fatConFPontaIndFaturado"] = consumo_fp_fat
    if consumo_fp_val is not None:
        out["fatConFPontaIndValorReais"] = round(consumo_fp_val, 2)

    if consumo_exc_p_reg is not None:
        out["fatConPontaExcRegistrado"] = consumo_exc_p_reg
    if consumo_exc_p_fat is not None:
        out["fatConPontaExcFaturado"] = consumo_exc_p_fat
    if consumo_exc_p_val is not None:
        out["fatConPontaExcValorReais"] = round(consumo_exc_p_val, 2)

    if consumo_exc_fp_reg is not None:
        out["fatConFPontaIndExcRegistrado"] = consumo_exc_fp_reg
    if consumo_exc_fp_fat is not None:
        out["fatConFPontaIndExcFaturado"] = consumo_exc_fp_fat
    if consumo_exc_fp_val is not None:
        out["fatConFPontaIndExcValorReais"] = round(consumo_exc_fp_val, 2)

    if demanda_fp_lim_val > 0 and demanda_fp_contr is not None and demanda_fp_contr >= 100:
        out["fatEscassezHidricaValorReais"] = round(demanda_fp_lim_val, 2)
        total_fat_kwh = sum(v for v in (consumo_p_fat, consumo_fp_fat) if v is not None)
        if total_fat_kwh > 0:
            out["fatEscassezHidrica"] = round(total_fat_kwh, 2)
    elif escassez_val > 0:
        out["fatEscassezHidricaValorReais"] = round(escassez_val, 2)
        total_reg_kwh = sum(v for v in (consumo_p_reg, consumo_fp_reg) if v is not None)
        if total_reg_kwh > 0:
            out["fatEscassezHidrica"] = round(total_reg_kwh, 2)

    if cob_icms_cde or dif_icms_cde:
        out["fatMultasDiversas"] = round(cob_icms_cde + dif_icms_cde, 2)

    # No layout MT da COELBA, essas linhas entram no Consen como observações
    # de desconto por fonte alternativa.
    observacoes: list[float] = []
    if demanda_fp_contr is not None and demanda_fp_contr >= 100 and (dif_desc_fp or dif_desc_fp_sem_icms):
        observacoes.append(round(dif_desc_fp + dif_desc_fp_sem_icms, 2))
    elif dif_desc_fp:
        observacoes.append(round(dif_desc_fp, 2))
        if dif_desc_fp_sem_icms:
            observacoes.append(round(dif_desc_fp_sem_icms, 2))
    if dif_desc_np:
        observacoes.append(round(dif_desc_np, 2))
    for idx, valor in enumerate(observacoes[:5], start=1):
        out[f"obsCod_{idx}"] = 135
        out[f"obsValor_{idx}"] = round(valor, 2)

    if aliq_pis is not None:
        out["fatAliqPis"] = round(aliq_pis, 2)
    if aliq_cofins is not None:
        out["fatAliqCofins"] = round(aliq_cofins, 2)
    if aliq_icms is not None:
        out["fatAliqIcms"] = round(aliq_icms, 2)

    if ret_trib_federal > 0:
        out["fatRetTribFederal"] = round(ret_trib_federal, 2)

    m_desc = re.search(r"DESCONTO\s+INCONDICIONAL\b.*?R\$\s*([\d\.,]+)", text, flags=re.IGNORECASE | re.DOTALL)
    if not m_desc:
        m_desc = re.search(r"DESCONTO\s+SOBRE\s+TARIFA[:\s]+R\$\s*([\d\.,]+)", text, flags=re.IGNORECASE)
    if m_desc:
        v = _to_float_br(m_desc.group(1))
        out["fatBeneficioTarifarioBrutoValorReais"] = v
        out["fatBeneficioLiquidoValorReais"] = v
    _aplicar_desconto_fio_a4_comercial(out, text)

    return out


def _extract_mt_rules_celpe(text: str) -> dict[str, float]:
    """Parser MT para layout CELPE (Neoenergia Pernambuco).

    Extrai os campos que o parser Coelba não cobre neste layout:
    - Demanda ativa faturada e valor (linha "Demanda Ativa kW QUANT PREÇO VALOR")
    - Demanda ultrapassagem (linhas "ULT.USO SISTEMA FIO", "ULTRAPASSAGEM", "ULT.")
    - Demanda reativa excedente FP (linhas "F. PONTA INDUTIVO EXCEDENTE", "DEMANDA REATIVA EXCEDENTE")
    - Demanda contratada ("GRANDEZAS CONTRATADAS → Demanda Contratada N")
    - Demanda FP registrada ("Demanda Máxima Fora de Ponta" no demonstrativo)
    - Consumo ponta e FP faturado + valor (linhas Consumo-TUSD / Consumo-TE)
    - Dif Desc Fonte Alternativa FP e NP (obsCod 135)
    - Desconto incondicional / benefício tarifário
    """
    out: dict[str, float] = {}
    linhas = [ln.strip() for ln in text.splitlines() if ln.strip()]

    consumo_p_qtd = 0.0
    consumo_p_val = 0.0
    consumo_fp_qtd = 0.0
    consumo_fp_val = 0.0
    demanda_fat: float | None = None
    demanda_val = 0.0
    demanda_contr: float | None = None
    demanda_fp_reg: float | None = None
    ultra_qtd: float | None = None
    ultra_val = 0.0
    dem_reat_exc_fat: float | None = None
    dem_reat_exc_reg: float | None = None
    dem_reat_exc_val = 0.0
    dif_desc_fp = 0.0
    dif_desc_np = 0.0

    for line in linhas:
        line_norm = _texto_normalizado(line)

        # Demanda Ativa kW → faturada e valor (somente se NÃO for ultrapassagem)
        if re.search(r"\bDEMANDA ATIVA\b", line_norm) and re.search(r"\bKW\b", line_norm):
            if not re.search(r"\bULT(?:RAPASS|\.)?\b", line_norm):
                qtd, val = _parse_mt_item_qtd_valor(line, r"KW")
                if qtd is not None:
                    demanda_fat = qtd
                if val is not None:
                    demanda_val = val
                continue

        # Ultrapassagem de demanda → fatDemFPontaIndUltra
        if re.search(r"\bULT\.USO\s+SISTEMA\s+FIO\b", line_norm) or \
                re.search(r"\bULTRAPASSAGEM\b", line_norm) or \
                (re.search(r"\bULT\.\b", line_norm) and re.search(r"\bKW\b", line_norm)):
            qtd, val = _parse_mt_item_qtd_valor(line, r"KW")
            if qtd is not None:
                ultra_qtd = (ultra_qtd or 0.0) + qtd
            if val is not None:
                ultra_val += val
            continue

        # Grandezas contratadas
        if "DEMANDA CONTRATADA" in line_norm:
            nums = re.findall(r"\d+", line)
            if nums:
                demanda_contr = float(nums[-1])
            continue

        # Consumo-TUSD NPonta  /  Consumo-TE Na Ponta  → ponta
        if re.search(r"\bCONSUMO-TUSD\s+N\.?PONTA\b", line_norm) or \
                re.search(r"\bCONSUMO-TE\s+NA\s+PONTA\b", line_norm):
            qtd, val = _parse_mt_item_qtd_valor(line, r"KWH")
            if qtd is not None:
                consumo_p_qtd = max(consumo_p_qtd, qtd)
            if val is not None:
                consumo_p_val += val
            continue

        # Consumo-TUSD F.Ponta  /  Consumo-TE F.Ponta  → fora de ponta
        if re.search(r"\bCONSUMO-TUSD\s+F\.?\s*PONTA\b", line_norm) or \
                re.search(r"\bCONSUMO-TE\s+F\.?\s*PONTA\b", line_norm):
            qtd, val = _parse_mt_item_qtd_valor(line, r"KWH")
            if qtd is not None:
                consumo_fp_qtd = max(consumo_fp_qtd, qtd)
            if val is not None:
                consumo_fp_val += val
            continue

        # Demanda Máxima Fora de Ponta (demonstrativo) → FP registrada
        if "DEMANDA MAXIMA FORA DE PONTA" in line_norm:
            val = _parse_mt_medidor_registrado(line)
            if val is not None:
                demanda_fp_reg = val
            continue

        # Demanda Reativa Excedente / F. Ponta Indutivo Excedente → fatDemFPontaExcFaturada
        if re.search(r"\bDEMANDA REATIVA EXCEDENTE\b", line_norm) or \
                re.search(r"\bF\.?\s*PONTA INDUTIVO EXCEDENTE\b", line_norm) or \
                re.search(r"\bDEM\.?\s*REAT\.?\s*EXC\.?\s*F\.?\s*PONTA\b", line_norm):
            qtd, val = _parse_mt_item_qtd_valor(line, r"KVAR")
            if qtd is not None:
                dem_reat_exc_fat = (dem_reat_exc_fat or 0.0) + qtd
                dem_reat_exc_reg = dem_reat_exc_fat
            if val is not None:
                dem_reat_exc_val += val
            continue

        # Dif Desc Fonte Alternativa FP / NP
        if re.search(r"\bDIF\.?\s*DESC\.?\s*F(?:ONT|T)\.?\s*ALT(?:ERN)?\.?\s*F\.?\s*PONTA\b", line_norm) or \
                re.search(r"\bDIF\.?\s*DESC\.?\s*FONTE\s+ALT\b.*\bF\.?\s*PONTA\b", line_norm):
            valores = RE_MONEY.findall(line)
            if valores:
                dif_desc_fp = abs(_to_float_br(valores[0]))
            continue

        if re.search(r"\bDIF\.?\s*DESC\.?\s*F(?:ONT|T)\.?\s*ALT(?:ERN)?\.?\s*N\.?\s*PONTA\b", line_norm) or \
                re.search(r"\bDIF\.?\s*DESC\.?\s*FONTE\s+ALT\b.*\bN\.?\s*PONTA\b", line_norm):
            valores = RE_MONEY.findall(line)
            if valores:
                dif_desc_np = abs(_to_float_br(valores[0]))
            continue

        # Dif Desc Fonte Alternativa genérico (sem FP/NP na mesma linha) — fallback
        if re.search(r"\bDIF\.?\s*DESC\.?\s*F(?:ONT|T)\.?\s*ALT\b", line_norm) and not dif_desc_fp and not dif_desc_np:
            valores = RE_MONEY.findall(line)
            if valores:
                v = abs(_to_float_br(valores[0]))
                if not dif_desc_fp:
                    dif_desc_fp = v
                elif not dif_desc_np:
                    dif_desc_np = v
            continue

    if demanda_fat is not None:
        out["fatDemFPontaIndFaturada"] = demanda_fat
    if demanda_val:
        out["fatDemFPontaIndValorReais"] = round(demanda_val, 2)
    if demanda_contr is not None:
        out["fatDemContratadaFPonta"] = demanda_contr
        if demanda_fat is not None:
            out["fatDemFPontaIndFaturada"] = max(demanda_fat, demanda_contr)
    if demanda_fp_reg is not None:
        out["fatDemFPontaIndRegistrada"] = demanda_fp_reg

    # Fallback faturas com demanda medida zerada (zero-billing): a demanda
    # registrada/faturada vem 0 na fatura, mas o CONSEN espera o valor contratado.
    # Backfill com a Demanda Contratada para não deixar os campos zerados.
    if demanda_contr is not None and demanda_contr > 0:
        if not out.get("fatDemFPontaIndRegistrada"):
            out["fatDemFPontaIndRegistrada"] = demanda_contr
        if not out.get("fatDemFPontaIndFaturada"):
            out["fatDemFPontaIndFaturada"] = demanda_contr

    if ultra_qtd is not None:
        out["fatDemFPontaIndUltra"] = round(ultra_qtd, 2)
    if ultra_val:
        out["fatDemFPontaIndUltraValorReais"] = round(ultra_val, 2)

    if dem_reat_exc_fat is not None:
        out["fatDemFPontaExcFaturada"] = round(dem_reat_exc_fat, 2)
    if dem_reat_exc_reg is not None:
        out["fatDemFPontaExcRegistrada"] = round(dem_reat_exc_reg, 2)
    if dem_reat_exc_val:
        out["fatDemFPontaExcValorReais"] = round(dem_reat_exc_val, 2)

    if consumo_p_qtd > 0:
        out["fatConPontaFaturado"] = consumo_p_qtd
        out["fatConPontaValorReais"] = round(consumo_p_val, 2)

    if consumo_fp_qtd > 0:
        out["fatConFPontaIndFaturado"] = consumo_fp_qtd
        out["fatConFPontaIndValorReais"] = round(consumo_fp_val, 2)

    # Dif Desc Fonte Alternativa → obsCod 135
    # Valor maior → FP, valor menor → NP (quando ambos presentes)
    obs_vals: list[float] = []
    if dif_desc_fp and dif_desc_np:
        maior = max(dif_desc_fp, dif_desc_np)
        menor = min(dif_desc_fp, dif_desc_np)
        obs_vals = [round(maior, 2), round(menor, 2)]
    elif dif_desc_fp:
        obs_vals = [round(dif_desc_fp, 2)]
    elif dif_desc_np:
        obs_vals = [round(dif_desc_np, 2)]
    for idx, valor in enumerate(obs_vals[:5], start=1):
        out[f"obsCod_{idx}"] = 135
        out[f"obsValor_{idx}"] = valor

    m_desc = re.search(r"DESCONTO\s+INCONDICIONAL\b.*?R\$\s*([\d\.,]+)", text, flags=re.IGNORECASE | re.DOTALL)
    if not m_desc:
        m_desc = re.search(r"DESCONTO\s+SOBRE\s+TARIFA[:\s]+R\$\s*([\d\.,]+)", text, flags=re.IGNORECASE)
    if m_desc:
        v = _to_float_br(m_desc.group(1))
        out["fatBeneficioTarifarioBrutoValorReais"] = v
        out["fatBeneficioLiquidoValorReais"] = v

    _aplicar_desconto_fio_a4_comercial(out, text)

    if "fatDescontoFio" not in out:
        for line in text.splitlines():
            line_norm = _texto_normalizado(line)
            if "DESCONTO FIO" in line_norm:
                valores = RE_MONEY.findall(line)
                if valores:
                    v = abs(_to_float_br(valores[-1]))
                    if v > 0:
                        out["fatDescontoFio"] = v
                        break

    return out


def _extract_mt_rules_elektro(text: str) -> dict[str, float]:
    """Parser MT para layout Elektro DANFE (atual) com fallback CCI (legado)."""
    out: dict[str, float] = {}
    linhas = [ln.strip() for ln in text.splitlines() if ln.strip()]

    consumo_p_reg = None
    consumo_p_fat = None
    consumo_p_val = None
    consumo_fp_reg = None
    consumo_fp_fat = None
    consumo_fp_val = None
    demanda_fp_reg = None
    demanda_fp_fat = None
    demanda_fp_val = 0.0
    demanda_ultra_reg = None
    demanda_ultra_val = 0.0
    montante_contratado = None
    subsidio_bruto = None
    subsidio_liquido = None
    escassez_cobr_val = 0.0
    _pis_val: float | None = None
    _cofins_val: float | None = None
    _base_total: float | None = None
    _icms_total: float | None = None
    _valor_fatura: float | None = None

    for line in linhas:
        line_norm = _texto_normalizado(line)

        # Consumo Ponta (TUSD)
        if "TUSD ENERGIA PONTA TUSD" in line_norm:
            reg, fat, val = _parse_elektro_danfe_line(line, "KWH")
            if reg is None:
                reg, fat, val = _parse_elektro_cci_line(line, "KWH")
            if reg is not None:
                consumo_p_reg = reg
            if fat is not None:
                consumo_p_fat = fat
            if val is not None:
                consumo_p_val = val
            continue

        # Consumo Fora Ponta (TUSD)
        if "TUSD ENERGIA FORA DE PONTA TUSD" in line_norm:
            reg, fat, val = _parse_elektro_danfe_line(line, "KWH")
            if reg is None:
                reg, fat, val = _parse_elektro_cci_line(line, "KWH")
            if reg is not None:
                consumo_fp_reg = reg
            if fat is not None:
                consumo_fp_fat = fat
            if val is not None:
                consumo_fp_val = val
            continue

        # Demanda Ultrapassagem FP
        if "DEMANDA ULTRAP." in line_norm and "DISTRIBUICAO TUSD" in line_norm:
            reg, _, val = _parse_elektro_danfe_line(line, "KW")
            if reg is not None:
                demanda_ultra_reg = reg
            if val is not None:
                demanda_ultra_val = val
            continue

        # Demanda FP ICMS-isenta: parcela sem ICMS, soma ao valor total
        if "DEMANDA DE DISTRIBUICAO TUSD" in line_norm and "ISENTA ICMS" in line_norm:
            _, _, val = _parse_elektro_danfe_line(line, "KW")
            if val is not None:
                demanda_fp_val += val
            continue

        # Demanda FP (Verde)
        if "DEMANDA DE DISTRIBUICAO TUSD" in line_norm and "ISENTA ICMS" not in line_norm:
            reg, fat, val = _parse_elektro_danfe_line(line, "KW")
            if reg is None:
                reg, fat, val = _parse_elektro_cci_line(line, "KW")
            if reg is not None:
                demanda_fp_reg = reg
            if fat is not None:
                demanda_fp_fat = fat
            if val is not None:
                demanda_fp_val += val
            continue

        # Subsidio Tarifario TUSD (bruto)
        if "SUBSIDIO TARIFARIO TUSD" in line_norm and "LIQUIDO" not in line_norm:
            _, _, val = _parse_elektro_danfe_line(line, "UN")
            if val is not None:
                subsidio_bruto = val
            continue

        # Subsidio Tarifario Liquido
        if "SUBSIDIO TARIFARIO LIQUIDO" in line_norm:
            monies = RE_MONEY.findall(line)
            if monies:
                subsidio_liquido = _to_float_br(monies[0])
            continue

        # Escassez hidrica: COBR.PIS, COBR.COFINS, COML.ICMS COBRADO, COBRANCA DE AJUSTE
        if any(k in line_norm for k in ("COBR.PIS", "COBR.COFINS", "COML.ICMS COBRADO", "COBRANCA DE AJUSTE")):
            monies = RE_MONEY.findall(line)
            if monies:
                escassez_cobr_val += abs(_to_float_br(monies[0]))
            continue

        # Demanda contratada: linha RMxxxxx (legado)
        if "RM" in line_norm and montante_contratado is None:
            m_rm = re.search(r"\bRM\w+\s+(\d+(?:[,\.]+\d+)?)\s+[\d,\.]+\s*%", line, re.IGNORECASE)
            if m_rm:
                montante_contratado = _to_float_qty(m_rm.group(1))

        # Montante contratado: "Montante em Todos os Periodos: XX kW"
        if ("MONTANTE" in line_norm or "DEMANDA" in line_norm) and "PERIOD" in line_norm and montante_contratado is None:
            m_mont = re.search(r"(\d[\d,.]*)\s*KW\b", line_norm)
            if m_mont:
                montante_contratado = _to_float_qty(m_mont.group(1))

        # Fallback RM medidor para kWh (so usa se DANFE nao extraiu)
        if line_norm.startswith("RM") and "CONSUMO KWH PT" in line_norm and consumo_p_reg is None:
            nums = re.findall(r"[\d]+(?:[,\.][\d]+)?", line.replace(".", "").replace(",", "."))
            if nums:
                consumo_p_reg = float(nums[-1])
                consumo_p_fat = float(nums[-1])

        if line_norm.startswith("RM") and "CONSUMO KWH FP" in line_norm and consumo_fp_reg is None:
            nums = re.findall(r"[\d]+(?:[,\.][\d]+)?", line.replace(".", "").replace(",", "."))
            if nums:
                consumo_fp_reg = float(nums[-1])
                consumo_fp_fat = float(nums[-1])

        # Fallback DEMANDA kW TP (so usa se DANFE nao extraiu)
        if line_norm.startswith("DEMANDA KW TP") and demanda_fp_reg is None:
            m_dkw = re.search(r"DEMANDA\s+KW\s+TP\s+([\d\.]+)", line_norm)
            if m_dkw:
                try:
                    demanda_fp_reg = float(m_dkw.group(1))
                except ValueError:
                    pass

        # Fallback DEMANDA TUSD kW (legado)
        if "DEMANDA TUSD" in line_norm and "KW" in line_norm and demanda_fp_reg is None:
            m_dtusd = re.search(r"DEMANDA\s+TUSD\s+KW\s+([\d\.,]+)\s+([\d\.,]+)", line_norm)
            if m_dtusd:
                montante_contratado = _to_float_qty(m_dtusd.group(1))
                demanda_fp_reg = _to_float_qty(m_dtusd.group(2))

        # PIS (para VNF)
        if " PIS " in f" {line_norm} " and "IMP.RET" not in line_norm and "COBR" not in line_norm:
            m_pis = re.search(
                r"\bPIS\b\s+([\d\.,]+)\s+([\d\.,]+)\s*%\s+([\d\.,]+)",
                line, re.IGNORECASE,
            )
            if m_pis and _pis_val is None:
                _pis_val = _to_float_br(m_pis.group(3))

        # COFINS (para VNF)
        if " COFINS " in f" {line_norm} " and "IMP.RET" not in line_norm and "COBR" not in line_norm:
            m_cof = re.search(
                r"\bCOFINS\b\s+([\d\.,]+)\s+([\d\.,]+)\s*%\s+([\d\.,]+)",
                line, re.IGNORECASE,
            )
            if m_cof and _cofins_val is None:
                _cofins_val = _to_float_br(m_cof.group(3))

        # Total (para VNF)
        if line_norm.startswith("TOTAL") and "A PAGAR" not in line_norm:
            monies = RE_MONEY.findall(line)
            if len(monies) >= 3:
                _base_total = _to_float_br(monies[0])
                _icms_total = _to_float_br(monies[1])
                _valor_fatura = _to_float_br(monies[2])

    # --- Saidas ---

    if consumo_p_reg is not None:
        out["fatConPontaRegistrado"] = consumo_p_reg
    if consumo_p_fat is not None:
        out["fatConPontaFaturado"] = consumo_p_fat
    if consumo_p_val is not None:
        out["fatConPontaValorReais"] = round(consumo_p_val, 2)

    if consumo_fp_reg is not None:
        out["fatConFPontaIndRegistrado"] = consumo_fp_reg
    if consumo_fp_fat is not None:
        out["fatConFPontaIndFaturado"] = consumo_fp_fat
    if consumo_fp_val is not None:
        out["fatConFPontaIndValorReais"] = round(consumo_fp_val, 2)

    if demanda_fp_reg is not None:
        out["fatDemFPontaIndRegistrada"] = demanda_fp_reg
        out["fatDemPontaRegistrada"] = demanda_fp_reg  # Verde: único período, Ponta = FP
        out["fatDemContratadaPonta"] = 0.0
        contr = montante_contratado if montante_contratado is not None else demanda_fp_reg
        dem_fat = max(demanda_fp_reg, contr)
        out["fatDemFPontaIndFaturada"] = round(dem_fat, 2)
        out["fatDemContratadaFPonta"] = round(contr, 2)
        if demanda_fp_val:
            out["fatDemFPontaIndValorReais"] = round(demanda_fp_val, 2)

    if demanda_ultra_reg is not None:
        out["fatDemPontaUltra"] = round(demanda_ultra_reg, 2)
        if demanda_ultra_val:
            out["fatDemPontaUltraValorReais"] = round(demanda_ultra_val, 2)

    if subsidio_bruto is not None:
        out["fatBeneficioTarifarioBrutoValorReais"] = round(subsidio_bruto, 2)
    if subsidio_liquido is not None:
        out["fatBeneficioLiquidoValorReais"] = round(subsidio_liquido, 2)

    if escassez_cobr_val > 0:
        out["fatEscassezHidricaValorReais"] = round(escassez_cobr_val, 2)
        total_kwh = sum(v for v in (consumo_p_fat, consumo_fp_fat) if v is not None)
        if total_kwh > 0:
            out["fatEscassezHidrica"] = round(total_kwh, 2)

    if _icms_total is not None:
        out["fatICMS"] = _icms_total
    if _valor_fatura is not None:
        out["fatValorFatura"] = _valor_fatura
    if _base_total is not None and _icms_total is not None:
        pis = _pis_val or 0.0
        cofins = _cofins_val or 0.0
        out["fatValorNotaFiscal"] = round(_base_total + _icms_total + pis + cofins, 2)

    # Desconto Fio B (TUSD) -- padrao fixo para todas as faturas MT ELEKTRO
    out["fatDescontoFio"] = 50.0
    out["fatDescontoFioKWh"] = 46.40

    return out


def _extract_mt_rules(text: str, conc_cod: str = "") -> dict[str, float]:
    if conc_cod == "ELEKTRO":
        return _extract_mt_rules_elektro(text)
    if conc_cod in ("CELPE", "COSERN"):
        out = _extract_mt_rules_coelba(text)
        out.update(_extract_mt_rules_celpe(text))
        return out
    return _extract_mt_rules_coelba(text)


def _tipo_por_caminho(pdf_path: Path) -> str:
    parent = pdf_path.parent.name.upper()
    if parent in {"BT", "MT"}:
        return parent.lower()
    return ""


def _detectar_tipo_por_pdf(pdf_path: Path, text: str) -> tuple[str, str, str, str]:
    tipo_txt, tarifa, subgrupo, detected = _detectar_tipo_tarifa(text)
    tipo_hint = _tipo_por_caminho(pdf_path)
    txt_norm = _texto_normalizado(text)

    if tipo_hint == tipo_txt:
        return tipo_txt, tarifa, subgrupo, detected

    if tipo_hint == "bt":
        tarifa_bt = "Branca" if "BRANCA" in txt_norm else "Convencional"
        detected_bt = "B3_BRANCA" if tarifa_bt == "Branca" else "B3"
        return "bt", tarifa_bt, "B3 [<2,3kV]", detected_bt

    if tipo_hint == "mt":
        subgrupo_mt = "A4 [2,3kV a 25kV]"
        detected_mt = "A4_VERDE"
        if "A3" in txt_norm:
            subgrupo_mt = "A3 [<44kV]"
            detected_mt = "A3_VERDE"
        elif "A2" in txt_norm:
            subgrupo_mt = "A2"
            detected_mt = "A2_VERDE"
        elif re.search(r"\bA1\b", txt_norm):
            subgrupo_mt = "A1"
            detected_mt = "A1_VERDE"
        tarifa_mt = "HS - Azul" if "AZUL" in txt_norm else "HS - Verde"
        if tarifa_mt == "HS - Azul":
            detected_mt = detected_mt.replace("_VERDE", "_AZUL")
        return "mt", tarifa_mt, subgrupo_mt, detected_mt

    return tipo_txt, tarifa, subgrupo, detected


def _aplicar_fallback_datas(rec: dict, mes: int, ano: int) -> None:
    if not rec.get("fatDataReferencia"):
        rec["fatDataReferencia"] = dt.date(ano, mes, 1)
    if not rec.get("fatDataLeituraAtual"):
        rec["fatDataLeituraAtual"] = _ultimo_dia_mes(ano, mes)
    if not rec.get("fatDataLeituraAnterior"):
        if mes == 1:
            rec["fatDataLeituraAnterior"] = _ultimo_dia_mes(ano - 1, 12)
        else:
            rec["fatDataLeituraAnterior"] = _ultimo_dia_mes(ano, mes - 1)


_PASTA_PARA_CONC: dict[str, str] = {
    "BAHIA": "COELBA",
    "COELBA": "COELBA",
    "PERNAMBUCO": "CELPE",
    "CELPE": "CELPE",
    "RIO_GRANDE_DO_NORTE": "COSERN",
    "COSERN": "COSERN",
    "SAO_PAULO": "ELEKTRO",
    "MATO_GROSSO_DO_SUL": "ELEKTRO",
    "ELEKTRO": "ELEKTRO",
}


def _conc_cod_por_caminho(pdf_path: Path) -> str:
    for parte in reversed(pdf_path.parts):
        slug = re.sub(r"[^A-Z0-9]+", "_", _strip_accents(parte).upper()).strip("_")
        if slug in _PASTA_PARA_CONC:
            return _PASTA_PARA_CONC[slug]
    return ""


def _conc_cod_por_texto(text: str) -> str:
    text_norm = _texto_normalizado(text)
    if "COMPANHIA DE ELETRICIDADE DO ESTADO DA BAHIA" in text_norm:
        return "COELBA"
    if "COMPANHIA ENERGETICA DE PERNAMBUCO" in text_norm:
        return "CELPE"
    if "COMPANHIA ENERGETICA DO RIO GRANDE DO NORTE" in text_norm:
        return "COSERN"
    if "ELEKTRO REDES S.A." in text_norm or "NEOENERGIA ELEKTRO" in text_norm:
        return "ELEKTRO"
    return ""


def processar_pdf_direto(pdf_path: Path, mes: int, ano: int) -> tuple[str, dict]:
    rec = _empty_record()
    rec["ARQUIVO"] = pdf_path.name
    rec["fatCarimbo"] = _carimbo_do_nome(pdf_path)
    rec["fatDataReferencia"] = dt.date(ano, mes, 1)
    rec["concCod"] = _conc_cod_por_caminho(pdf_path)

    tipo_hint = _tipo_por_caminho(pdf_path) or "bt"

    try:
        text, first_page_words = _extract_pdf_data(pdf_path)
    except Exception as exc:
        rec["ERRO"] = f"{type(exc).__name__}: {exc}"
        return tipo_hint, rec

    if not text.strip():
        rec["ERRO"] = "PDF sem texto extraivel"
        return tipo_hint, rec

    tipo, tarifa, subgrupo, detected = _detectar_tipo_por_pdf(pdf_path, text)
    rec["cadTarifaCod"] = tarifa
    rec["cadSubGrupoCod"] = subgrupo
    rec["TARIFA_DETECTADA"] = detected
    if not rec["concCod"]:
        rec["concCod"] = _conc_cod_por_texto(text)

    rec["Instalacao"] = _extract_instalacao(text, first_page_words)
    rec["CNPJ"] = _extract_cnpj(text)
    rec["ENDERECO"] = _extract_endereco(text)
    rec["NOTAFISCAL"] = _extract_notafiscal(text)
    rec["CODIGOCLIENTE"] = _extract_codigo_cliente(text)
    if not _norm(str(rec.get("Instalacao", ""))):
        # Em produção, a UC do nome original é registrada no master no momento do carimbo.
        # Isso é mais confiável do que CODIGOCLIENTE para casos como COELBA.
        uc_master = _uc_por_carimbo_master(str(rec.get("fatCarimbo", "")))
        if uc_master:
            rec["Instalacao"] = uc_master
        elif _norm(str(rec.get("CODIGOCLIENTE", ""))):
            # Fallback final para casos em que o identificador da instalação
            # realmente coincide com o código do cliente.
            rec["Instalacao"] = _norm(str(rec["CODIGOCLIENTE"]))
    rec["fatCodigoBarras"] = _extract_codigo_barras(text)

    rec["fatDataEmissao"] = _find_date_after_labels(
        text,
        ["Data de emissao", "Data de emissão", "Emissao", "Emissão"],
    )
    rec["fatDataVcto"] = _extract_vencimento(text, first_page_words)
    leitura_ant, leitura_atu = _extract_leituras(text)
    rec["fatDataLeituraAnterior"] = leitura_ant
    rec["fatDataLeituraAtual"] = leitura_atu
    _aplicar_fallback_datas(rec, mes, ano)

    rec["fatValorFatura"] = _extract_total(text)
    rec["fatValorNotaFiscal"] = (
        _extract_valor_nota_fiscal_mt(text) if tipo == "mt" else _extract_valor_nota_fiscal(text)
    )
    if not rec["fatValorNotaFiscal"]:
        rec["fatValorNotaFiscal"] = rec["fatValorFatura"]
    rec["fatIlumPublica"] = _extract_ilum_publica(text)
    rec["fatICMS"] = _extract_imposto(text, "ICMS")
    rec["fatPIS"] = _extract_imposto(text, "PIS")
    rec["fatCOFINS"] = _extract_imposto(text, "COFINS")
    rec.update(_extract_aliquotas(text))
    trib_federal_linhas = _extract_tributo_federal_linhas(text)
    rec["fatTributoFederalPerc"], rec["fatTributoFederalVal"] = _extract_tributo_federal(text)
    rec.update(_extract_bandeiras(text))
    rec["fatMultas"] = _extract_multas(text)
    rec["fatMultasDiversas"] = _extract_multas_diversas(text)
    for _i, (_code, _val) in enumerate(_extract_observacoes(text, is_compensacao=_is_sistema_compensacao(text))[:5], start=1):
        rec[f"obsCod_{_i}"] = _code
        rec[f"obsValor_{_i}"] = _val
    rec["Debitos anteriores"] = _extract_debitos_anteriores(text)
    rec.update(_extract_retencoes(text))
    if tipo == "mt":
        rec.update(_extract_mt_rules(text, conc_cod=str(rec.get("concCod", ""))))
        if str(rec.get("concCod", "")) in _NEOENERGIA_CONCS:
            _aplicar_retencao_neo_mt(rec)
    else:
        rec.update(_extract_bt_rules(text))
        if str(rec.get("concCod", "")).upper() in {"ELEKTRO", "COELBA"}:
            # IMP.RET. individuais → consolida em fatTributoFederalPerc=-1 + fatTributoFederalVal
            _ret_total = sum(
                abs(float(rec.get(f, 0) or 0))
                for f in (
                    "fatDescPisValRetImposto",
                    "fatDescCofinsValRetImposto",
                    "fatDescCsllValRetImposto",
                    "fatDescIrpjValRetImposto",
                )
            )
            if _ret_total > 0:
                _est_total = _extract_est_retencoes(text)
                _net = round(max(0.0, _ret_total - _est_total), 2)
                rec["fatTributoFederalPerc"] = -1.0
                rec["fatTributoFederalVal"] = -_net
                _zerar_retencoes_individuais(rec)
        if str(rec.get("concCod", "")).upper() == "CELPE":
            is_comp = _is_sistema_compensacao(text)
            tem_escassez = float(rec.get("fatEscassezHidricaValorReais") or 0) > 0

            _importe_val = 0.0
            for _line in text.splitlines():
                _ln = _texto_normalizado(_line)
                if "IMP.SOM/DIM" in _ln or "IMPORTE A SOMAR" in _ln:
                    _vals = RE_MONEY.findall(_line)
                    if _vals:
                        _importe_val = _to_float_br(_vals[0])
                    break

            if is_comp and _importe_val < 0 and not tem_escassez:
                # GD por compensação: crédito injetado → alíquota -1%
                _nf = float(rec.get("fatValorNotaFiscal") or 0)
                rec["fatTributoFederalPerc"] = -1.0
                rec["fatTributoFederalVal"] = (
                    round(sum(val for _, val in trib_federal_linhas), 2)
                    if trib_federal_linhas else round(-0.01 * _nf, 2)
                )
            else:
                # Sem compensação GD (ou com escassez hídrica): mantém 5,85%
                # Se houver duas linhas de tributo, soma os valores
                perc_set = {round(perc, 2) for perc, _ in trib_federal_linhas if perc > 0}
                if {5.85, 9.45}.issubset(perc_set):
                    rec["fatTributoFederalPerc"] = -1.0
                    rec["fatTributoFederalVal"] = round(sum(val for _, val in trib_federal_linhas), 2)
                elif len(trib_federal_linhas) > 1:
                    rec["fatTributoFederalVal"] = round(sum(val for _, val in trib_federal_linhas), 2)

            if abs(float(rec.get("fatTributoFederalVal") or 0.0)) > 0.0:
                _zerar_retencoes_individuais(rec)
    rec["ERRO"] = ""
    return tipo, rec


def _pasta_busca(mes: int, ano: int, pasta: str | None, estado: str) -> Path:
    if pasta:
        return Path(str(pasta).strip())
    return DOWNLOAD_ROOT / estado / f"{ano}-{mes:02d}"


def _listar_pdfs(pasta: Path, carimbos: set[str]) -> list[Path]:
    try:
        pdfs = [p for p in pasta.rglob("*.pdf") if p.is_file()]
    except Exception as exc:
        raise RuntimeError(f"Falha ao listar PDFs em {pasta}: {exc}") from exc

    if carimbos:
        carimbos_norm = {c.upper() for c in carimbos}
        pdfs = [p for p in pdfs if p.stem.upper() in carimbos_norm]

    pdfs.sort(key=_tipo_sort_key)
    return pdfs


def _xlsx_saida(mes: int, ano: int, tipo: str) -> Path:
    return OUTPUT_DIR / f"ocr_neoenergia_{tipo.upper()}_{mes:02d}{ano}.xlsx"


def _fmt_cell(header: str, value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if value is None:
        return 0 if header in NUMERIC_HEADERS else ""
    if value == "":
        return 0 if header in NUMERIC_HEADERS else ""
    return value


def salvar_excel(registros: list[dict], destino: Path, titulo: str) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    header_font = Font(name="Arial", size=11, bold=False)
    body_font = Font(name="Calibri", size=11)
    align_right = Alignment(horizontal="right")

    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=HEADER_DISPLAY.get(h, h))
        cell.font = header_font
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(h, 18.0)
    ws.freeze_panes = "A2"

    for r_idx, rec in enumerate(registros, start=2):
        for c_idx, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_fmt_cell(h, rec.get(h)))
            cell.font = body_font
            if h in {"Instalacao", "fatCarimbo", "fatCodigoBarras", "CNPJ", "NOTAFISCAL", "CODIGOCLIENTE"}:
                cell.alignment = align_right
            if h.startswith("fatData"):
                cell.number_format = "DD/MM/YYYY"
            if h in NUMERIC_HEADERS:
                cell.number_format = "#,##0.00"

    # Salva primeiro em arquivo local temporario para evitar travas/intermitencias em UNC.
    tmp_local = Path(tempfile.gettempdir()) / f"ocr_neoenergia_{int(time.time())}.xlsx"
    wb.save(tmp_local)

    last_err = None
    for _ in range(3):
        try:
            # Se destino existir, tenta sobrescrever.
            if destino.exists():
                try:
                    destino.unlink()
                except Exception:
                    pass
            shutil.copy2(tmp_local, destino)
            try:
                tmp_local.unlink(missing_ok=True)
            except Exception:
                pass
            return
        except Exception as exc:
            last_err = exc
            time.sleep(1.2)

    # Fallback final: nao perde o resultado.
    alt = destino.with_name(f"{destino.stem}_{dt.datetime.now().strftime('%H%M%S')}{destino.suffix}")
    try:
        shutil.copy2(tmp_local, alt)
        try:
            tmp_local.unlink(missing_ok=True)
        except Exception:
            pass
        raise RuntimeError(f"Falha ao gravar destino principal ({destino}); salvo em fallback: {alt}")
    except Exception as exc:
        raise RuntimeError(f"Falha ao salvar XLSX em rede e fallback: {exc} | ultimo erro: {last_err}")


def parse_args():
    hoje = dt.date.today()
    p = argparse.ArgumentParser(description="OCR Neoenergia -> XLSX")
    p.add_argument("--mes", type=int, default=hoje.month, help="Mes de referencia (1-12)")
    p.add_argument("--ano", type=int, default=hoje.year, help="Ano de referencia")
    p.add_argument("--pasta", type=str, default="", help="Pasta mensal com PDFs misturados")
    p.add_argument("--estado", type=str, default=DEFAULT_ESTADO, help="Estado/pasta base quando --pasta nao for informado")
    p.add_argument(
        "--tipo",
        choices=["bt", "mt", "ambos"],
        default="ambos",
        help="Filtra a saida por tipo tarifario",
    )
    p.add_argument(
        "--carimbo",
        action="append",
        default=[],
        help="Carimbo(s) especifico(s), ex: BB_2001242 (pode repetir)",
    )
    p.add_argument("--saida-bt", type=str, default="", help="Caminho de saida alternativo para o XLSX BT")
    p.add_argument("--saida-mt", type=str, default="", help="Caminho de saida alternativo para o XLSX MT")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    mes = int(args.mes)
    ano = int(args.ano)
    carimbos = {str(c).strip().upper() for c in (args.carimbo or []) if str(c).strip()}
    saidas_override = {
        "bt": Path(str(args.saida_bt).strip()) if str(args.saida_bt).strip() else None,
        "mt": Path(str(args.saida_mt).strip()) if str(args.saida_mt).strip() else None,
    }

    if not (1 <= mes <= 12):
        log.error("Mes invalido. Use 1..12.")
        return 1

    pasta = _pasta_busca(mes, ano, args.pasta, args.estado)
    ok_pasta, err_pasta = _path_exists_safe(pasta)
    if not ok_pasta:
        detalhe = f"Falha ao acessar pasta de PDFs: {pasta}"
        if err_pasta:
            detalhe += f" | {err_pasta}"
        log.error(detalhe)
        return 1

    log.info("=" * 64)
    log.info("OCR NEOENERGIA")
    log.info("=" * 64)
    log.info(f"Pasta : {pasta}")
    log.info(f"Filtro: mes={mes:02d} ano={ano} tipo={args.tipo} carimbos={len(carimbos)}")

    try:
        pdfs = _listar_pdfs(pasta, carimbos)
    except Exception as exc:
        log.error(str(exc))
        return 1

    if not pdfs:
        log.warning("Nenhum PDF encontrado para o filtro informado.")
        return 0

    log.info(f"PDFs candidatos: {len(pdfs)}")
    por_tipo: dict[str, list[dict]] = {"bt": [], "mt": []}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = [ex.submit(processar_pdf_direto, pdf, mes, ano) for pdf in pdfs]
        for fut in as_completed(futs):
            try:
                tipo, rec = fut.result()
                por_tipo.setdefault(tipo, []).append(rec)
            except Exception as exc:
                rec = _empty_record()
                rec["ERRO"] = f"erro_worker: {type(exc).__name__}: {exc}"
                por_tipo["bt"].append(rec)

    tipos_saida = ["bt", "mt"] if args.tipo == "ambos" else [args.tipo]
    total_salvos = 0
    for tipo in tipos_saida:
        registros = por_tipo.get(tipo, [])
        registros.sort(key=lambda r: str(r.get("fatCarimbo", "")))
        if not registros:
            log.warning(f"Nenhum PDF classificado como {tipo.upper()} para gerar planilha.")
            continue

        destino = saidas_override.get(tipo) or _xlsx_saida(mes, ano, tipo)
        try:
            destino.parent.mkdir(parents=True, exist_ok=True)
            salvar_excel(registros, destino, titulo=f"OCR_NEO_{tipo.upper()}")
        except Exception as exc:
            log.error(f"Falha ao salvar XLSX {tipo.upper()}: {exc}")
            return 1

        ok = sum(1 for r in registros if not r.get("ERRO"))
        erro = len(registros) - ok
        log.info(f"XLSX {tipo.upper()} salvo: {destino}")
        log.info(f"Resumo {tipo.upper()}: total={len(registros)} ok={ok} erro={erro}")
        total_salvos += 1

    if total_salvos == 0:
        log.warning("Nenhuma planilha foi gerada.")
        return 0

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
