#!/usr/bin/env python3
"""
ocr_cemig.py  —  OCR autônomo de faturas CEMIG
================================================
Mesma arquitetura do ocr_enel_ENEL.py.
Usa os 4 extratores originais sem alterar nada neles:
  cemig_b3.py | cemig_b1.py | cemig_tusd_a4_verde.py | cemig_ths_a4.py

Estrutura de pastas esperada:
    DOWNLOAD CEMIG / 03.2026 / BT / BB_2000001.pdf
    DOWNLOAD CEMIG / 03.2026 / MT / BB_2000050.pdf

Saida:
    OCR CEMIG / ocr_cemig_BT_032026.xlsx   <- SEMPRE recriado do zero
    OCR CEMIG / ocr_cemig_MT_032026.xlsx   <- SEMPRE recriado do zero
    OCR CEMIG / ocr_cemig.log

Uso:
    python ocr_cemig.py                        # todos os meses (padrao)
    python ocr_cemig.py --mes 03 --ano 2026    # mes especifico
    python ocr_cemig.py --pasta "03.2026"      # nome exato da subpasta
    python ocr_cemig.py --todos                # forca todos os meses
    python ocr_cemig.py --tipo bt              # so BT
    python ocr_cemig.py --tipo mt              # so MT

MUDANÇA vs versão anterior:
    - salvar_excel() SEMPRE recria o arquivo do zero (sem append).
    - _carimbos_no_xlsx() e filtro de duplicatas REMOVIDOS.
    - Todos os PDFs da pasta são processados a cada execução.
"""

from __future__ import annotations

import argparse
import datetime as dt
import importlib.util
import logging
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURACAO
# =============================================================================

PASTA_DOWNLOAD = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/DOWNLOAD CEMIG")
PASTA_SAIDA    = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/OCR CEMIG")
PASTA_LOGS     = Path(__file__).resolve().parent / "logs"
MAX_WORKERS    = 4

NOMES_BT = {"bt", "b3", "b1", "baixa tensao", "baixa_tensao"}
NOMES_MT = {"mt", "a4", "a3", "a2", "a1", "media tensao", "media_tensao",
            "mt_a4", "ths", "tusd"}

# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# =============================================================================
# HEADERS — identicos aos campos usados pelos 4 extratores originais
# =============================================================================

HEADERS_REF: List[str] = [
    # ── Identificação (mesma ordem da ENEL) ───────────────────────────────────
    "Instalacao",          # col A - sem acento (extratores gravam 'Instalacao', exibido como 'Instalação'
    "fatDataEmissao",
    "fatDataVcto",
    "fatValorFatura",
    "concCod",
    "fatDataCadastro",
    "fatDataLeituraAnterior",
    "fatDataLeituraAtual",
    "fatIlumPublica",
    "cadTarifaCod",        # texto real do select (ex: 'Convencional', 'HS - Verde')
    "cadSubGrupoCod",      # texto real do select (ex: 'B3', 'A4')
    # ── Demandas ──────────────────────────────────────────────────────────────
    "fatDemContratadaPonta",
    "fatDemContratadaFPonta",
    "fatDemPontaRegistrada",
    "fatDemFPontaIndRegistrada",
    "fatDemFPontaCapRegistrada",
    "fatDemPontaExcFaturada",
    "fatDemFPontaExcFaturada",
    "fatDemPontaExcRegistrada",
    "fatDemFPontaExcRegistrada",
    "fatDemPontaFaturada",
    "fatDemFPontaIndFaturada",
    "fatDemPontaUltra",
    "fatDemFPontaIndUltra",
    # ── Consumo ───────────────────────────────────────────────────────────────
    "fatConPontaRegistrado",
    "fatConFPontaIndRegistrado",
    "fatConFPontaCapRegistrado",
    "fatConIntermediarioRegistrado",
    "fatConPontaFaturado",
    "fatConFPontaIndFaturado",
    "fatConFPontaCapFaturado",
    "fatConIntermediarioFaturado",
    "fatConPontaExcRegistrado",
    "fatConFPontaIndExcRegistrado",
    "fatConPontaReativoExcedente",
    "fatConFPontaIndReativoExcedente",
    "fatConFPontaCapExcRegistrado",
    "fatConPontaExcFaturado",
    "fatConFPontaIndExcFaturado",
    "fatConPontaReativoFaturado",
    "fatConFPontaIndReativoFaturado",
    "fatConFPontaCapExcFaturado",
    # ── Impostos / financeiros ────────────────────────────────────────────────
    "fatICMS",
    "fatICMSBase",
    "fatPIS",
    "fatCOFINS",
    "fatValorNotaFiscal",
    # ── Observações (até 5 por fatura) ──────────────────────────────────────
    "obsCod_1",    "obsValor_1",
    "obsCod_2",    "obsValor_2",
    "obsCod_3",    "obsValor_3",
    "obsCod_4",    "obsValor_4",
    "obsCod_5",    "obsValor_5",
    "CNPJ",
    "ENDERECO",
    "NOTAFISCAL",
    "CODIGOCLIENTE",
    "fatDataReferencia",
    # ── Injetado / GD ────────────────────────────────────────────────────────
    "fatConPontaInjetadoRegistrado",
    "fatConPontaInjetadoFaturado",
    "fatConFPontaInjetadoRegistrado",
    "fatConFPontaInjetadoFaturado",
    "fatCodigoBarras",
    "Debitos anteriores",
    "fatCarimbo",
    "usuCod",
    # ── Demandas geração ──────────────────────────────────────────────────────
    "fatDemPontaGeracaoRegistrada",
    "fatDemPontaGeracao",
    "fatDemPontaGeracaoValorReais",
    "fatDemFPontaGeracaoRegistrada",
    "fatDemFPontaGeracao",
    "fatDemFPontaGeracaoValorReais",
    "fatDemContratadaGeracaoPonta",
    "fatDemContratadaGeracaoFPonta",
    # ── Demanda Reativa HFP (UFDR kW) ────────────────────────────────────────
    "fatDemFPontaIndReativoExcedente",
    "fatDemFPontaIndReativoFaturado",
    "fatDemFPontaIndReativoValorReais",
    # ── Valores R$ ───────────────────────────────────────────────────────────
    "fatDemPontaValorReais",
    "fatDemFPontaIndValorReais",
    "fatDemPontaUltraValorReais",
    "fatDemFPontaIndUltraValorReais",
    "fatDemPontaExcValorReais",
    "fatDemFPontaExcValorReais",
    "fatConPontaValorReais",
    "fatConFPontaIndValorReais",
    "fatConFPontaCapValorReais",
    "fatConIntermediarioValorReais",
    "fatConPontaExcValorReais",
    "fatConFPontaIndExcValorReais",
    "fatConFPontaCapExcValorReais",
    "fatConPontaInjetadoValorReais",
    "fatConFPontaInjetadoValorReais",
    "fatConPontaInjetadoUsina",
    "fatConPontaInjetadoUsinaSaldoAcumulado",
    "fatConFPontaInjetadoUsina",
    "fatConFPontaInjetadoUsinaSaldoAcumulado",
    "fatDemandasDevolucaoPtaValorReais",
    "fatDemandasDevolucaoFPtaValorReais",
    "fatConIntermedInjetadoRegistrado",
    "fatConIntermedInjetadoFaturado",
    "fatConIntermedInjetadoValorReais",
    # ── TUSD / Fio ───────────────────────────────────────────────────────────
    "fatDescontoFio",
    "fatDescPisAliquota",
    "fatDescPisPercRetImposto",
    "fatDescPisValRetImposto",
    "fatDesCofinsAliquota",
    "fatDescCofinsPercRetImposto",
    "fatDescCofinsValRetImposto",
    "fatDesIcmsAliquota",
    "fatDescCsllPercRetImposto",
    "fatDescCsllValRetImposto",
    "fatDescIrpjPercRetImposto",
    "fatDescIrpjValRetImposto",
    "fatDescIrrfPercRetImposto",
    "fatDescIrrfValRetImposto",
    "fatDescConsumoPercRetImposto",
    "fatDescConsumoValRetImposto",
    "fatDescDemandaPercRetImposto",
    "fatDescDemandaValRetImposto",
    "fatValBandeira",
    "fatValBandeira2",
    "fatDIC",
    "fatFIC",
    "fatMultas",
    "fatTributoFederalPerc",
    "fatTributoFederalVal",
    "fatMultasDiversas",
    "fatDescontoFioKWh",
    "fatConCreditoTUSDPontaValorReais",
    "fatConCreditoTUSDFPontaValorReais",
    "fatBeneficioTarifarioBrutoValorReais",
    "fatBeneficioLiquidoValorReais",
    "fatContaCovidValorReais",
    "fatEscassezHidricaValorReais",
    "fatContaCovid",
    "fatEscassezHidrica",
    # ── Controle ─────────────────────────────────────────────────────────────
    "TARIFA_DETECTADA",
    "ARQUIVO",
    "ERRO",
]

_HEADER_DISPLAY = {
    "Instalacao": 'Instalação',   # exibe com acento no Excel, chave sem
}

_COL_WIDTH_DEFAULT = 20.0
_COL_WIDTHS = {
    "Instalação": 22.0, "fatCarimbo": 14.0, "fatDataCadastro": 18.0,
    "concCod": 10.0, "cadTarifaCod": 18.0, "cadSubGrupoCod": 22.0,
    "fatDataEmissao": 18.0, "fatDataVcto": 14.0, "fatDataReferencia": 20.0,
    "fatDataLeituraAnterior": 26.0, "fatDataLeituraAtual": 22.0,
    "fatValorFatura": 18.0, "fatValorNotaFiscal": 22.0,
    "CNPJ": 24.0, "ENDERECO": 60.0, "NOTAFISCAL": 18.0,
    "TARIFA_DETECTADA": 22.0, "ARQUIVO": 40.0, "ERRO": 45.0,
}

# =============================================================================
# DETECCAO DE TARIFA
# Ordem de prioridade: TUSD Livre > THS A4 > B3 > B1
# =============================================================================

def _texto_rapido(pdf_path: str, max_paginas: int = 3) -> str:
    partes = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for p in pdf.pages[:max_paginas]:
                try:
                    txt = p.extract_text(layout=True) or p.extract_text() or ""
                except Exception:
                    txt = ""
                partes.append(txt)
    except Exception:
        pass
    return "\n".join(partes).upper()


def detectar_tarifa(pdf_path: str) -> str:
    """
    Detecta a tarifa CEMIG pelo texto do PDF.
    Retorna: 'B3' | 'B1' | 'TUSD_A4_VERDE' | 'THS_A4' | 'DESCONHECIDA'
    """
    txt = _texto_rapido(pdf_path)

    # MT — TUSD Livre A4 Verde (mercado livre)
    # Marcadores exclusivos: COMPONENTE FIO, COMPONENTE ENCARGO, TUSD LIVRE, desconto fio
    _is_tusd = (
        "COMPONENTE FIO" in txt
        or "COMPONENTE ENCARGO" in txt
        or "TUSD LIVRE" in txt
        or "DESCONTO COMP. FIO" in txt
        or "ULTRAPASSAGEM C. FIO" in txt
        or ("TUSD" in txt and "LIVRE" in txt)
        or ("TUSD" in txt and "A4" in txt and "VERDE" in txt)
    )
    if _is_tusd:
        return "TUSD_A4_VERDE"

    # MT — THS Verde A4 (cativo horosazonal)
    if "THS" in txt or "DEMANDA ATIVA HFP" in txt or "HOROSAZON" in txt:
        return "THS_A4"

    # BT — B3 Convencional
    if re.search(r"CONVENCIONAL\s+B3", txt):
        return "B3"

    # BT — B1 Residencial
    if "B1" in txt or "RESIDENCIAL" in txt:
        return "B1"

    return "DESCONHECIDA"


# =============================================================================
# IMPORTACAO DOS EXTRATORES (dinamica, com cache)
# =============================================================================

_CACHE_EXTRATORES: dict = {}


def _carregar_extrator(tarifa: str):
    if tarifa in _CACHE_EXTRATORES:
        return _CACHE_EXTRATORES[tarifa]

    MAPA = {
        "B3":            "cemig_b3",
        "B1":            "cemig_b1",
        "TUSD_A4_VERDE": "cemig_tusd_a4_verde",
        "THS_A4":        "cemig_ths_a4",
    }
    nome = MAPA.get(tarifa)
    if not nome:
        return None

    base = Path(__file__).resolve().parent
    candidatos = [
        base / f"{nome}.py",
        base / "tarifas" / f"{nome}.py",
        base / "tarifas" / f"{nome.replace('cemig_', '')}.py",
    ]

    for caminho in candidatos:
        if caminho.exists():
            spec = importlib.util.spec_from_file_location(nome, str(caminho))
            mod  = importlib.util.module_from_spec(spec)
            sys.modules[nome] = mod
            spec.loader.exec_module(mod)
            _CACHE_EXTRATORES[tarifa] = mod
            return mod

    return None


# =============================================================================
# CARIMBO DO NOME DO ARQUIVO
# =============================================================================

def _carimbo_do_nome(filename: str) -> str:
    stem = Path(filename).stem
    m = re.search(r"[Bb][Bb]_(\d+)", stem)
    if m:
        return m.group(1)
    m = re.search(r"(\d{4,})", stem)
    return m.group(1) if m else stem


# =============================================================================
# MAPA TARIFA → TEXTOS REAIS DOS SELECTS DO CONSEN
# =============================================================================

_TARIFA_TEXTOS_CEMIG = {
    "B3":            ("Convencional",      "B3"),
    "B1":            ("Convencional",      "B1"),
    "THS_A4":        ("HS - Verde",         "A4"),
    "TUSD_A4_VERDE": ("TUSD Livre Verde",  "A4"),
}


def _textos_tarifa_cemig(tarifa: str):
    return _TARIFA_TEXTOS_CEMIG.get(tarifa, (None, None))


def _numero_cemig(valor) -> float:
    if valor in (None, ""):
        return 0.0
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    texto = str(valor).strip()
    if not texto:
        return 0.0
    if re.match(r"^-?[\d\.]+,\d{2}$", texto):
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return round(float(texto), 2)
    except Exception:
        return 0.0


def _normalizar_observacoes_cemig(dados: dict) -> None:
    """
    Garante que a família DIC/FIC da CEMIG ocupe no máximo uma observação.
    Preferência de código: 58 > 149 > 11.
    """
    pares: list[tuple[str, float]] = []
    for i in range(1, 6):
        cod = str(dados.get(f"obsCod_{i}") or "").strip()
        if not cod or cod == "0":
            continue
        valor = _numero_cemig(dados.get(f"obsValor_{i}"))
        if abs(valor) <= 0.004:
            continue
        pares.append((cod, valor))

    dic_codigos = {"11", "58", "149"}
    dic_pares = [(cod, val) for cod, val in pares if cod in dic_codigos]
    outros = [(cod, val) for cod, val in pares if cod not in dic_codigos]
    fat_dic = _numero_cemig(dados.get("fatDIC"))
    fat_fic = _numero_cemig(dados.get("fatFIC"))

    if not dic_pares and (abs(fat_dic) > 0.004 or abs(fat_fic) > 0.004):
        if abs(fat_dic) > 0.004:
            dic_pares = [("58", -abs(fat_dic))]
        else:
            dic_pares = [("11", -abs(fat_fic))]

    if len(dic_pares) > 1:
        preferencia = ("58", "149", "11")
        cod_escolhido = next((cod for cod in preferencia if any(c == cod for c, _ in dic_pares)), dic_pares[0][0])
        valor_escolhido = next(val for cod, val in dic_pares if cod == cod_escolhido)
        pares = outros + [(cod_escolhido, valor_escolhido)]
    else:
        pares = outros + dic_pares

    for i in range(1, 6):
        dados[f"obsCod_{i}"] = ""
        dados[f"obsValor_{i}"] = 0
    for i, (cod, valor) in enumerate(pares[:5], start=1):
        dados[f"obsCod_{i}"] = cod
        dados[f"obsValor_{i}"] = _numero_cemig(valor)


# =============================================================================
# PROCESSAMENTO DE UM PDF
# =============================================================================

def processar_pdf(pdf_path: str, tipo_pasta: str) -> dict:
    filename = Path(pdf_path).name
    carimbo  = _carimbo_do_nome(filename)

    try:
        tarifa = detectar_tarifa(pdf_path)

        if tarifa == "DESCONHECIDA":
            tarifa = "B3" if tipo_pasta == "bt" else "THS_A4"
            log.warning(f"    {filename}: tarifa nao detectada, usando fallback '{tarifa}'")

        extrator = _carregar_extrator(tarifa)
        if extrator is None:
            raise ImportError(
                f"Extrator para '{tarifa}' nao encontrado. "
                f"Certifique-se que cemig_b3.py, cemig_b1.py, "
                f"cemig_tusd_a4_verde.py e cemig_ths_a4.py estao "
                f"no mesmo diretorio de ocr_cemig.py."
            )

        linha = extrator.extrair_linha(pdf_path, HEADERS_REF, carimbo=carimbo)
        dados = dict(zip(HEADERS_REF, linha))

        obs_list = dados.pop("_obs_list", None)
        if obs_list:
            for _i, (_cod, _val) in enumerate(obs_list[:5], start=1):
                dados[f"obsCod_{_i}"]   = _cod
                dados[f"obsValor_{_i}"] = _numero_cemig(_val)
        _normalizar_observacoes_cemig(dados)

        dados["TARIFA_DETECTADA"] = tarifa
        dados["ARQUIVO"]          = filename
        dados["ERRO"]             = ""

        if not dados.get("fatCarimbo"):
            dados["fatCarimbo"] = carimbo

        tname, sname = _textos_tarifa_cemig(tarifa)
        if tname:
            dados["cadTarifaCod"]   = tname
        if sname:
            subgrupo_extraido = str(dados.get("cadSubGrupoCod") or "").strip()
            manter_subgrupo_mt = (
                tarifa in {"THS_A4", "TUSD_A4_VERDE"}
                and (
                    subgrupo_extraido.upper().startswith("A")
                    or subgrupo_extraido.upper().startswith("AS")
                )
            )
            if not manter_subgrupo_mt:
                dados["cadSubGrupoCod"] = sname

        if "Instalação" in dados and "Instalacao" not in dados:
            dados["Instalacao"] = dados["Instalação"]
        if "Instalação" in dados and "Instalacao" not in dados:
            dados["Instalacao"] = dados["Instalação"]

        inst = dados.get("Instalacao") or dados.get("Instalação") or "?"
        log.info(f"    OK  {filename}  ->  {tarifa}  (inst={inst}  carimbo={carimbo})")
        return dados

    except Exception as exc:
        log.error(f"    ERRO  {filename}: {exc}")
        return {
            "fatCarimbo":       carimbo,
            "TARIFA_DETECTADA": "ERRO",
            "ARQUIVO":          filename,
            "ERRO":             str(exc),
        }


# =============================================================================
# EXCEL — sempre recria do zero
# =============================================================================

_FONT_HEADER = Font(name="Arial",   size=11, bold=False)
_FONT_DADOS  = Font(name="Calibri", size=11)
_ALIGN_DIR   = Alignment(horizontal="right")


def _fmt_cell(header: str, value):
    CAMPOS_TEXTO = {
        "Instalacao", "Instalação", "CNPJ", "ENDERECO", "NOTAFISCAL",
        "fatCodigoBarras", "ARQUIVO", "ERRO", "TARIFA_DETECTADA",
        "cadTarifaCod", "cadSubGrupoCod", "CODIGOCLIENTE",
        "obsCod_1", "obsCod_2", "obsCod_3", "obsCod_4", "obsCod_5",
    }

    if value is None or value == "":
        if header.startswith("fat") or header in ("concCod", "usuCod"):
            return 0
        if header.startswith("obsValor"):
            return 0
        if header.startswith("obsCod"):
            return ""
        return ""

    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (float, int)):
        if header in CAMPOS_TEXTO:
            return str(value)
        return value

    if isinstance(value, str):
        v = value.strip()
        if v.startswith("'"):
            v = v[1:]

        if header in CAMPOS_TEXTO:
            return v

        if re.match(r"^\d{2}/\d{2}/\d{4}$", v):
            try:
                return dt.datetime.strptime(v, "%d/%m/%Y").date()
            except Exception:
                pass
        if re.match(r"^-?[\d\.]+,\d{2}$", v):
            try:
                return float(v.replace(".", "").replace(",", "."))
            except Exception:
                pass
        if re.match(r"^-?\d+$", v):
            try:
                return int(v)
            except Exception:
                pass
        return v

    return value


def _criar_header_excel(ws, headers: list):
    for col, h in enumerate(headers, 1):
        display = _HEADER_DISPLAY.get(h, h)
        cell = ws.cell(row=1, column=col, value=display)
        cell.font = _FONT_HEADER
        ws.column_dimensions[get_column_letter(col)].width = (
            _COL_WIDTHS.get(h, _COL_WIDTH_DEFAULT)
        )
    ws.row_dimensions[1].height = 21.0
    ws.freeze_panes = "A2"


def salvar_excel(registros: list, caminho: Path):
    """
    SEMPRE recria o arquivo do zero.
    Se já existir, apaga e recria — não há append.
    """
    if caminho.exists():
        caminho.unlink()
        log.info(f"  Removido xlsx anterior: {caminho.name}")

    wb = Workbook()
    ws = wb.active
    ws.title = caminho.stem
    _criar_header_excel(ws, HEADERS_REF)

    for linha_idx, reg in enumerate(registros, start=2):
        for col, h in enumerate(HEADERS_REF, 1):
            cell = ws.cell(
                row=linha_idx, column=col,
                value=_fmt_cell(h, reg.get(h, ""))
            )
            cell.font      = _FONT_DADOS
            cell.alignment = _ALIGN_DIR

    wb.save(caminho)
    ok   = sum(1 for r in registros if not r.get("ERRO"))
    erro = len(registros) - ok
    log.info(f"  Criado: {caminho.name}  OK={ok}  ERRO={erro}  TOTAL={len(registros)}")


# =============================================================================
# NAVEGACAO DE PASTAS
# =============================================================================

def _pasta_label(pasta: Path) -> str:
    m = re.search(r"(\d{2})[_\-\.\s]?(\d{4})", pasta.name)
    return f"{m.group(1)}{m.group(2)}" if m else pasta.name


def _listar_pastas_mes(base: Path) -> list:
    padrao = re.compile(r"^(\d{2})[_\-\.\s]?(\d{4})$")
    try:
        filhos = list(base.iterdir())
    except OSError as e:
        log.warning(f"  iterdir falhou em {base}: {e} — sem pastas de mes")
        return []
    return sorted(
        [p for p in filhos
         if _is_dir_unc(p) and padrao.match(p.name.strip())],
        key=lambda p: p.name,
    )


def _subpasta(pasta_mes: Path, nomes_aceitos: set) -> Optional[Path]:
    """Localiza subpasta por caminho direto, sem iterdir (evita WinError UNC)."""
    # Tenta maiusculo primeiro (ex: BT, MT), depois os outros nomes do set
    candidatos = []
    for nome in nomes_aceitos:
        candidatos.append(pasta_mes / nome.upper())
        candidatos.append(pasta_mes / nome.lower())
        candidatos.append(pasta_mes / nome.capitalize())
    for candidato in candidatos:
        try:
            if candidato.is_dir():
                log.info(f"  Subpasta encontrada: {candidato.name}")
                return candidato
        except OSError:
            continue
    return None


# =============================================================================
# PROCESSAMENTO DE UM MES — sem filtro de duplicatas
# =============================================================================

def _processar_subpasta(pasta_sub: Path, tipo: str, xlsx_saida: Path):
    pdfs = sorted(pasta_sub.glob("*.pdf"))
    if not pdfs:
        log.warning(f"  Sem PDFs em: {pasta_sub}")
        return

    label = "BT" if tipo == "bt" else "MT"
    log.info(f"  {label}  ->  {pasta_sub.name}  ({len(pdfs)} PDFs)")

    registros = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(processar_pdf, str(p), tipo): p for p in pdfs}
        for f in as_completed(futures):
            registros.append(f.result())

    def _sort_key(r):
        try:
            return int(str(r.get("fatCarimbo", 0) or 0))
        except (ValueError, TypeError):
            return 0

    registros.sort(key=_sort_key)
    salvar_excel(registros, xlsx_saida)


def processar_mes(pasta_mes: Path, fazer_bt: bool = True, fazer_mt: bool = True):
    label = _pasta_label(pasta_mes)
    log.info(f"\n{'='*60}")
    log.info(f"  {pasta_mes.name}  ->  {label}")
    log.info(f"{'='*60}")

    if fazer_bt:
        sub = _subpasta(pasta_mes, NOMES_BT)
        if sub:
            _processar_subpasta(sub, "bt", PASTA_SAIDA / f"ocr_cemig_BT_{label}.xlsx")
        else:
            log.warning(f"  Subpasta BT nao encontrada em: {pasta_mes.name}")

    if fazer_mt:
        sub = _subpasta(pasta_mes, NOMES_MT)
        if sub:
            _processar_subpasta(sub, "mt", PASTA_SAIDA / f"ocr_cemig_MT_{label}.xlsx")
        else:
            log.warning(f"  Subpasta MT nao encontrada em: {pasta_mes.name}")


# =============================================================================
# CLI
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(description="OCR CEMIG  ->  planilhas BT e MT (recria do zero)")
    p.add_argument("--mes",          type=str)
    p.add_argument("--ano",          type=str)
    p.add_argument("--pasta",        type=str)
    p.add_argument("--todos",        action="store_true")
    p.add_argument("--tipo",         choices=["bt", "mt", "ambos"], default="ambos")
    p.add_argument("--base-dir",     type=str, default=None,
                   help="Pasta raiz alternativa (substitui PASTA_DOWNLOAD).")
    p.add_argument("--pasta-direta", type=str, default=None,
                   help="Pasta com PDFs para processar diretamente (sem estrutura BT/MT). "
                        "Requer --saida.")
    p.add_argument("--saida",        type=str, default=None,
                   help="Caminho do xlsx de saida (usado com --pasta-direta).")
    return p.parse_args()


def _resolver_pasta(args) -> Path:
    base = Path(args.base_dir) if getattr(args, "base_dir", None) else PASTA_DOWNLOAD

    if args.pasta:
        for sep in [".", "-", "_", " ", ""]:
            for tentativa in [
                args.pasta,
                args.pasta.replace(".", sep),
                args.pasta.replace("-", sep),
            ]:
                p = base / tentativa
                if _is_dir_unc(p):
                    return p
        log.error(f"Pasta '{args.pasta}' nao encontrada em {base}")
        sys.exit(1)

    hoje = dt.date.today()
    mes  = args.mes or f"{hoje.month:02d}"
    ano  = args.ano or str(hoje.year)

    for sep in [".", "-", "_", " ", ""]:
        p = base / f"{mes}{sep}{ano}"
        if _is_dir_unc(p):
            return p

    log.error(f"Pasta {mes}/{ano} nao encontrada em {base}")
    sys.exit(1)



def _is_dir_unc(p: Path) -> bool:
    """is_dir() tolerante ao WinError 1398 em caminhos UNC."""
    try:
        return p.is_dir()
    except OSError:
        return True  # assume que existe — o acesso real confirmara

def _mkdir_seguro(pasta):
    """mkdir tolerante ao WinError 1398 (diferenca de relogio com servidor UNC)."""
    try:
        pasta.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass  # pasta ja existe no servidor — WinError 1398 e falso positivo

def main():
    args = parse_args()
    _mkdir_seguro(PASTA_SAIDA)

    _mkdir_seguro(PASTA_LOGS)
    fh = logging.FileHandler(
        PASTA_LOGS / "ocr_cemig.log", encoding="utf-8", errors="replace"
    )
    fh.setFormatter(logging.Formatter(
        "%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    ))
    log.addHandler(fh)

    fazer_bt = args.tipo in ("bt", "ambos")
    fazer_mt = args.tipo in ("mt", "ambos")

    log.info("=" * 60)
    log.info("  OCR CEMIG  -  BT + MT  [MODO: recria do zero]".center(60))
    log.info("=" * 60)

    # Modo pasta-direta: processa PDFs de uma pasta arbitraria sem estrutura BT/MT
    if getattr(args, "pasta_direta", None):
        pasta_dir = Path(args.pasta_direta)
        if not args.saida:
            log.error("--pasta-direta requer --saida")
            sys.exit(1)
        xlsx_saida = Path(args.saida)
        _mkdir_seguro(xlsx_saida.parent)
        pdfs = sorted(pasta_dir.glob("*.pdf"))
        if not pdfs:
            log.warning(f"Sem PDFs em: {pasta_dir}")
            sys.exit(0)
        tipo = "bt" if fazer_bt and not fazer_mt else ("mt" if fazer_mt and not fazer_bt else "ambos")
        log.info(f"  Pasta      : {pasta_dir}")
        log.info(f"  PDFs       : {len(pdfs)}")
        log.info(f"  Saida      : {xlsx_saida}")
        registros = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futures = {ex.submit(processar_pdf, str(p), tipo): p for p in pdfs}
            for f in as_completed(futures):
                registros.append(f.result())
        registros.sort(key=lambda r: int(str(r.get("fatCarimbo", 0) or 0)) if str(r.get("fatCarimbo", 0) or 0).isdigit() else 0)
        salvar_excel(registros, xlsx_saida)
        log.info(f"\nConcluido. total={len(registros)} ok={sum(1 for r in registros if not r.get('ERRO'))} erro={sum(1 for r in registros if r.get('ERRO'))}")
        return

    base_download = Path(args.base_dir) if getattr(args, "base_dir", None) else PASTA_DOWNLOAD
    log.info(f"  Tipo       : {args.tipo.upper()}")
    log.info(f"  Download   : {base_download}")
    log.info(f"  Saida      : {PASTA_SAIDA}")
    log.info(f"  Workers    : {MAX_WORKERS}")
    log.info("  Extratores : cemig_b3 | cemig_b1 | cemig_tusd_a4_verde | cemig_ths_a4")

    modo_todos = args.todos or (not args.mes and not args.ano and not args.pasta)

    if modo_todos:
        pastas = _listar_pastas_mes(base_download)
        if not pastas:
            log.error(f"Nenhuma pasta de mes encontrada em: {base_download}")
            sys.exit(1)
        log.info(f"  Meses      : {len(pastas)}")
        for pasta in pastas:
            processar_mes(pasta, fazer_bt, fazer_mt)
    else:
        pasta = _resolver_pasta(args)
        processar_mes(pasta, fazer_bt, fazer_mt)

    log.info("\nConcluido.")


if __name__ == "__main__":
    main()
