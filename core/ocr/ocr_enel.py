#!/usr/bin/env python3
"""
ocr_enel_ENEL.py  —  OCR autonomo de faturas ENEL
=============================================
Estrutura esperada em DOWNLOAD ENEL:
    03-2026 / BT / BB_2000001.pdf ...
    03-2026 / MT / BB_2000006.pdf ...

Saida em OCR ENEL:
    ocr_enel_BT_032026.xlsx
    ocr_enel_MT_032026.xlsx

Uso:
    python ocr_enel.py                      # todos os meses (padrao)
    python ocr_enel.py --mes 03 --ano 2026  # mes especifico
    python ocr_enel.py --pasta "03-2026"    # nome exato da subpasta
    python ocr_enel.py --tipo bt            # so BT
    python ocr_enel.py --tipo mt            # so MT
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import logging
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from pathlib import Path
from typing import List, Optional

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

LOCAL_DIR = Path(__file__).resolve().parent.parent
ROOT_DIR = LOCAL_DIR.parent
INFRA_DIR = LOCAL_DIR.parent / "scripts" / "infra"
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
if str(INFRA_DIR) not in sys.path:
    sys.path.insert(0, str(INFRA_DIR))

# =============================================================================
# CONFIGURACAO
# =============================================================================

PASTA_DOWNLOAD = Path(os.environ.get("OCR_ENEL_DOWNLOAD_DIR", "//10.10.250.21/Energia/ARQUIVOS ENZO/DOWNLOAD ENEL"))
PASTA_SAIDA    = Path(os.environ.get("OCR_ENEL_SAIDA_DIR", "//10.10.250.21/Energia/ARQUIVOS ENZO/OCR ENEL"))
MAX_WORKERS    = 4
FALLBACK_SAIDA = Path(__file__).resolve().parent.parent / "_runtime_fallback" / "ocr_enel" / "saida"

NOMES_BT = {"bt", "b3", "baixa tensao", "baixa_tensao"}
NOMES_MT = {"mt", "a4", "media tensao", "media_tensao", "mt_a4"}

# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def _resolver_pasta_saida() -> Path:
    try:
        PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
        return PASTA_SAIDA
    except OSError as exc:
        FALLBACK_SAIDA.mkdir(parents=True, exist_ok=True)
        log.warning(
            "Saida OCR ENEL indisponivel em %s: %s. Usando fallback local %s",
            PASTA_SAIDA,
            exc,
            FALLBACK_SAIDA,
        )
        return FALLBACK_SAIDA

# =============================================================================
# HEADERS — ordem e larguras identicas a planilha de referencia CONSEN
# =============================================================================

HEADERS_REF: List[str] = [
    "Instalacao", "fatDataEmissao", "fatDataVcto", "fatValorFatura", "concCod",
    "fatDataCadastro", "fatDataLeituraAnterior", "fatDataLeituraAtual", "fatIlumPublica",
    "cadTarifaCod", "cadSubGrupoCod", "fatDemContratadaPonta", "fatDemContratadaFPonta",
    "fatDemPontaRegistrada", "fatDemFPontaIndRegistrada", "fatDemFPontaCapRegistrada",
    "fatDemPontaExcFaturada", "fatDemFPontaExcFaturada", "fatDemPontaExcRegistrada",
    "fatDemFPontaExcRegistrada", "fatDemPontaFaturada", "fatDemFPontaIndFaturada",
    "fatDemPontaUltra", "fatDemFPontaIndUltra", "fatConPontaRegistrado",
    "fatConFPontaIndRegistrado", "fatConFPontaCapRegistrado", "fatConIntermediarioRegistrado",
    "fatConPontaFaturado", "fatConFPontaIndFaturado", "fatConFPontaCapFaturado",
    "fatConIntermediarioFaturado", "fatConPontaExcRegistrado", "fatConFPontaIndExcRegistrado",
    "fatConFPontaCapExcRegistrado", "fatConPontaExcFaturado", "fatConFPontaIndExcFaturado",
    "fatConFPontaCapExcFaturado", "fatICMS", "fatPIS", "fatCOFINS", "fatValorNotaFiscal",
    # ── Observações (até 5 por fatura) ──────────────────────────────────────
    "obsCod_1", "obsValor_1",
    "obsCod_2", "obsValor_2",
    "obsCod_3", "obsValor_3",
    "obsCod_4", "obsValor_4",
    "obsCod_5", "obsValor_5",
    "CNPJ", "ENDERECO", "NOTAFISCAL", "CODIGOCLIENTE", "fatDataReferencia",
    "fatConPontaInjetadoRegistrado", "fatConPontaInjetadoFaturado",
    "fatConFPontaInjetadoRegistrado", "fatConFPontaInjetadoFaturado", "fatCodigoBarras",
    "Debitos anteriores", "fatCarimbo", "usuCod", "fatDemPontaGeracaoRegistrada",
    "fatDemPontaGeracao", "fatDemPontaGeracaoValorReais", "fatDemFPontaGeracaoRegistrada",
    "fatDemFPontaGeracao", "fatDemFPontaGeracaoValorReais", "fatDemContratadaGeracaoPonta",
    "fatDemContratadaGeracaoFPonta", "fatDemPontaValorReais", "fatDemFPontaIndValorReais",
    "fatDemPontaUltraValorReais", "fatDemFPontaIndUltraValorReais", "fatDemPontaExcValorReais",
    "fatDemFPontaExcValorReais", "fatConPontaValorReais", "fatConFPontaIndValorReais",
    "fatConFPontaCapValorReais", "fatConIntermediarioValorReais", "fatConPontaExcValorReais",
    "fatConFPontaIndExcValorReais", "fatConFPontaCapExcValorReais",
    "fatConPontaInjetadoValorReais", "fatConFPontaInjetadoValorReais",
    "fatConPontaInjetadoUsina", "fatConPontaInjetadoUsinaSaldoAcumulado",
    "fatConFPontaInjetadoUsina", "fatConFPontaInjetadoUsinaSaldoAcumulado",
    "fatDemandasDevolucaoPtaValorReais", "fatDemandasDevolucaoFPtaValorReais",
    "fatConIntermedInjetadoRegistrado", "fatConIntermedInjetadoFaturado",
    "fatConIntermedInjetadoValorReais", "fatDescontoFio", "fatDescPisAliquota",
    "fatDescPisPercRetImposto", "fatDescPisValRetImposto", "fatDesCofinsAliquota",
    "fatDescCofinsPercRetImposto", "fatDescCofinsValRetImposto", "fatDesIcmsAliquota",
    "fatDescCsllPercRetImposto", "fatDescCsllValRetImposto", "fatDescIrpjPercRetImposto",
    "fatDescIrpjValRetImposto", "fatDescIrrfPercRetImposto", "fatDescIrrfValRetImposto",
    "fatDescConsumoPercRetImposto", "fatDescConsumoValRetImposto",
    "fatDescDemandaPercRetImposto", "fatDescDemandaValRetImposto", "fatValBandeira",
    "fatValBandeira2", "fatDIC", "fatFIC", "fatMultas", "fatTributoFederalPerc",
    "fatTributoFederalVal", "fatMultasDiversas", "fatDescontoFioKWh",
    "fatConCreditoTUSDPontaValorReais", "fatConCreditoTUSDFPontaValorReais",
    "fatBeneficioTarifarioBrutoValorReais", "fatBeneficioLiquidoValorReais",
    "fatContaCovidValorReais", "fatEscassezHidricaValorReais", "fatContaCovid",
    "fatEscassezHidrica",
    # colunas de controle (nao existem na referencia)
    "TARIFA_DETECTADA", "ARQUIVO", "ERRO",
]

# Mapeamento de nome real (com acento) para chave interna (sem acento)
# usado apenas na gravacao do Excel para manter o header original
_HEADER_DISPLAY = {
    "Instalacao": "Instalação",
    "Debitos anteriores": "Débitos anteriores",
}

_COL_WIDTHS = {
    "Instalacao": 20.71, "fatDataEmissao": 18.14, "fatDataVcto": 13.0,
    "fatValorFatura": 17.29, "concCod": 10.29, "fatDataCadastro": 18.86,
    "fatDataLeituraAnterior": 26.29, "fatDataLeituraAtual": 22.86,
    "fatIlumPublica": 17.29, "cadTarifaCod": 15.71, "cadSubGrupoCod": 20.14,
    "fatDemContratadaPonta": 27.57, "fatDemContratadaFPonta": 29.0,
    "fatDemPontaRegistrada": 27.14, "fatDemFPontaIndRegistrada": 32.29,
    "fatDemFPontaCapRegistrada": 32.57, "fatDemPontaExcFaturada": 29.14,
    "fatDemFPontaExcFaturada": 30.57, "fatDemPontaExcRegistrada": 31.0,
    "fatDemFPontaExcRegistrada": 32.29, "fatDemPontaFaturada": 25.29,
    "fatDemFPontaIndFaturada": 30.43, "fatDemPontaUltra": 21.0,
    "fatDemFPontaIndUltra": 26.14, "fatConPontaRegistrado": 26.43,
    "fatConFPontaIndRegistrado": 31.57, "fatConFPontaCapRegistrado": 31.86,
    "fatConIntermediarioRegistrado": 35.43, "fatConPontaFaturado": 24.57,
    "fatConFPontaIndFaturado": 29.71, "fatConFPontaCapFaturado": 30.14,
    "fatConIntermediarioFaturado": 33.57, "fatConPontaExcRegistrado": 30.29,
    "fatConFPontaIndExcRegistrado": 35.43, "fatConFPontaCapExcRegistrado": 35.71,
    "fatConPontaExcFaturado": 28.43, "fatConFPontaIndExcFaturado": 33.57,
    "fatConFPontaCapExcFaturado": 33.86, "fatICMS": 10.0, "fatPIS": 7.86,
    "fatCOFINS": 12.71, "fatValorNotaFiscal": 21.43, "obsCod_1": 10.0, "obsValor_1": 12.0,
    "obsCod_2": 10.0, "obsValor_2": 12.0,
    "obsCod_3": 10.0, "obsValor_3": 12.0,
    "obsCod_4": 10.0, "obsValor_4": 12.0,
    "obsCod_5": 10.0, "obsValor_5": 12.0,
    "CNPJ": 7.14, "ENDERECO": 153.14, "NOTAFISCAL": 15.86, "CODIGOCLIENTE": 20.43,
    "fatDataReferencia": 20.86, "fatConPontaInjetadoRegistrado": 35.43,
    "fatConPontaInjetadoFaturado": 33.71, "fatConFPontaInjetadoRegistrado": 36.86,
    "fatConFPontaInjetadoFaturado": 35.0, "fatCodigoBarras": 18.86,
    "Debitos anteriores": 21.14, "fatCarimbo": 13.43, "usuCod": 9.29,
    "fatDemPontaGeracaoRegistrada": 36.0, "fatDemPontaGeracao": 24.29,
    "fatDemPontaGeracaoValorReais": 36.14, "fatDemFPontaGeracaoRegistrada": 37.43,
    "fatDemFPontaGeracao": 25.71, "fatDemFPontaGeracaoValorReais": 37.57,
    "fatDemContratadaGeracaoPonta": 36.43, "fatDemContratadaGeracaoFPonta": 37.86,
    "fatDemPontaValorReais": 27.29, "fatDemFPontaIndValorReais": 32.43,
    "fatDemPontaUltraValorReais": 33.0, "fatDemFPontaIndUltraValorReais": 38.14,
    "fatDemPontaExcValorReais": 31.14, "fatDemFPontaExcValorReais": 32.57,
    "fatConPontaValorReais": 26.57, "fatConFPontaIndValorReais": 31.71,
    "fatConFPontaCapValorReais": 32.0, "fatConIntermediarioValorReais": 35.57,
    "fatConPontaExcValorReais": 30.43, "fatConFPontaIndExcValorReais": 35.57,
    "fatConFPontaCapExcValorReais": 35.86, "fatConPontaInjetadoValorReais": 35.57,
    "fatConFPontaInjetadoValorReais": 37.0, "fatConPontaInjetadoUsina": 30.0,
    "fatConPontaInjetadoUsinaSaldoAcumulado": 48.29, "fatConFPontaInjetadoUsina": 31.29,
    "fatConFPontaInjetadoUsinaSaldoAcumulado": 49.71,
    "fatDemandasDevolucaoPtaValorReais": 42.14, "fatDemandasDevolucaoFPtaValorReais": 43.57,
    "fatConIntermedInjetadoRegistrado": 39.29, "fatConIntermedInjetadoFaturado": 37.57,
    "fatConIntermedInjetadoValorReais": 39.43, "fatDescontoFio": 17.57,
    "fatDescPisAliquota": 21.71, "fatDescPisPercRetImposto": 30.0,
    "fatDescPisValRetImposto": 28.71, "fatDesCofinsAliquota": 24.14,
    "fatDescCofinsPercRetImposto": 33.57, "fatDescCofinsValRetImposto": 32.29,
    "fatDesIcmsAliquota": 22.57, "fatDescCsllPercRetImposto": 30.71,
    "fatDescCsllValRetImposto": 29.43, "fatDescIrpjPercRetImposto": 30.86,
    "fatDescIrpjValRetImposto": 29.57, "fatDescIrrfPercRetImposto": 30.57,
    "fatDescIrrfValRetImposto": 29.29, "fatDescConsumoPercRetImposto": 36.86,
    "fatDescConsumoValRetImposto": 35.57, "fatDescDemandaPercRetImposto": 37.14,
    "fatDescDemandaValRetImposto": 35.71, "fatValBandeira": 17.57,
    "fatValBandeira2": 18.86, "fatDIC": 8.14, "fatFIC": 7.86,
    "fatMultas": 11.43, "fatTributoFederalPerc": 25.29, "fatTributoFederalVal": 24.0,
    "fatMultasDiversas": 20.86, "fatDescontoFioKWh": 23.0,
    "fatConCreditoTUSDPontaValorReais": 40.86, "fatConCreditoTUSDFPontaValorReais": 42.29,
    "fatBeneficioTarifarioBrutoValorReais": 42.0, "fatBeneficioLiquidoValorReais": 34.43,
    "fatContaCovidValorReais": 28.43, "fatEscassezHidricaValorReais": 33.57,
    "fatContaCovid": 16.57, "fatEscassezHidrica": 21.57,
    "TARIFA_DETECTADA": 18.0, "ARQUIVO": 38.0, "ERRO": 45.0,
}

_RETENÇÕES_ZERO = [
    "fatDescPisPercRetImposto", "fatDescPisValRetImposto",
    "fatDescCofinsPercRetImposto", "fatDescCofinsValRetImposto",
    "fatDescCsllPercRetImposto", "fatDescCsllValRetImposto",
    "fatDescIrpjPercRetImposto", "fatDescIrpjValRetImposto",
    "fatDescIrrfPercRetImposto", "fatDescIrrfValRetImposto",
    "fatDescConsumoPercRetImposto", "fatDescConsumoValRetImposto",
    "fatDescDemandaPercRetImposto", "fatDescDemandaValRetImposto",
]

# =============================================================================
# UTILITARIOS
# =============================================================================

RE_DATE  = re.compile(r"\b(\d{2})[/.](\d{2})[/.](\d{4})\b")
RE_REF   = re.compile(r"\b(\d{2})/(\d{4})\b")
RE_MONEY = re.compile(r"-?[\d.]+,\d{2}")

_DISTRIBUIDORAS = (
    "61.695.227",  # ENEL SP
    "07.110.651",  # ENEL CE
    "08.324.196",  # ENEL GO
    "62.291.006",  # CPFL Paulista
    "04.895.728",  # CPFL RGE
    "06.272.793",  # CEMIG D
)


def _br2f(s: str) -> float:
    if not s:
        return 0.0
    s = str(s).strip().replace('"', "").replace("'", "")
    neg = s.startswith("-") or s.endswith("-") or (s.startswith("(") and s.endswith(")"))
    s = s.replace("-", "").replace("(", "").replace(")", "")
    s = re.sub(r"[R$\s]", "", s)
    s = s.replace(".", "").replace(",", ".")
    try:
        v = float(s)
        return -v if neg else v
    except Exception:
        return 0.0


_CNPJS_ENEL_RJ = ("33.050.071", "28.150.884")  # Ampla Energia / ENEL Distribuição RJ


def _is_enel_rj(text: str) -> bool:
    return any(c in text for c in _CNPJS_ENEL_RJ) or "AMPLA ENERGIA" in text.upper()


# Breakdown fixo Lei 9430 por alíquota total
_TRIB_FEDERAL_BREAKDOWN: dict[str, dict[str, float]] = {
    "5.85": {"IRPJ": 1.20, "PIS": 0.65, "COFINS": 3.00, "CSLL": 1.00},
    "9.45": {"IRPJ": 4.80, "PIS": 0.65, "COFINS": 3.00, "CSLL": 1.00},
}

_MAPA_RETENCAO: dict[str, tuple[str, str]] = {
    "IRPJ":   ("fatDescIrpjPercRetImposto",   "fatDescIrpjValRetImposto"),
    "PIS":    ("fatDescPisPercRetImposto",    "fatDescPisValRetImposto"),
    "COFINS": ("fatDescCofinsPercRetImposto", "fatDescCofinsValRetImposto"),
    "CSLL":   ("fatDescCsllPercRetImposto",   "fatDescCsllValRetImposto"),
}


def _distribuir_retencao_inplace(rec: dict) -> None:
    """Distribui fatTributoFederalVal nos campos individuais IRPJ/PIS/COFINS/CSLL."""
    total_val  = abs(rec.get("fatTributoFederalVal") or 0.0)
    total_perc = abs(rec.get("fatTributoFederalPerc") or 0.0)
    if not total_val or not total_perc:
        return
    breakdown = _TRIB_FEDERAL_BREAKDOWN.get(f"{total_perc:.2f}")
    if not breakdown:
        return
    for nome, comp_perc in breakdown.items():
        campo_perc, campo_val = _MAPA_RETENCAO[nome]
        rec[campo_perc] = comp_perc
        rec[campo_val]  = -round(total_val * comp_perc / total_perc, 2)


def _parse_date(s: str) -> Optional[dt.date]:
    if not s:
        return None
    m = RE_DATE.search(str(s))
    if m:
        d, mo, y = map(int, m.groups())
        try:
            return dt.date(y, mo, d)
        except Exception:
            return None
    return None


def _ref2date(s: str) -> Optional[dt.date]:
    if not s:
        return None
    m = RE_REF.search(str(s))
    if m:
        mo, y = map(int, m.groups())
        try:
            return dt.date(y, mo, 1)
        except Exception:
            return None
    return None


def _parece_data(s: str) -> bool:
    if len(s) != 10:
        return False
    try:
        d, mo, y = int(s[:2]), int(s[2:4]), int(s[4:8])
        return 1 <= d <= 31 and 1 <= mo <= 12 and 1990 <= y <= 2035
    except Exception:
        return False


def _extract_text(pdf_path: str) -> str:
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            txt = (
                p.extract_text(x_tolerance=1, y_tolerance=1)
                or p.extract_text(layout=True)
                or p.extract_text()
            )
            if txt:
                parts.append(txt)
    return "\n".join(parts)


def _get_valor_nf(text: str) -> float:
    """
    Extrai o fatValorNotaFiscal = Base de Cálculo ICMS da fatura ENEL BT/MT.

    Na linha do TOTAL a ENEL imprime 4 colunas:
        TOTAL  <valor_fatura>  <pis+cof>  <base_icms>  <icms_val>

    O 3º valor (Base Calc ICMS) é o que vai para fatValorNotaFiscal = 94,82.

    Fallback: 1º valor da linha de tributos ICMS (mesmo número, outra posição).
    """
    # T1 — linha TOTAL: valor / pis+cof / base_icms / icms_val
    m = re.search(
        r"TOTAL\s+([\d\.]+,\d{2})\s+([\d\.]+,\d{2})\s+([\d\.]+,\d{2})\s+([\d\.]+,\d{2})",
        text, re.IGNORECASE
    )
    if m:
        return _br2f(m.group(3))

    # T2 — linha ICMS nos tributos: ICMS <base_calc> <aliquota> <valor>
    m = re.search(
        r"\bICMS\b\s+([\d\.]+,\d{2})\s+([\d\.]+,\d{2})\s+([\d\.]+,\d{2})",
        text, re.IGNORECASE
    )
    if m:
        return _br2f(m.group(1))

    # T3 — fallback: Subtotal Faturamento (valor antes de bandeira/COSIP)
    m = re.search(
        r"Subtotal\s+Faturamento\s+([\d\.]+,\d{2})",
        text, re.IGNORECASE
    )
    if m:
        return _br2f(m.group(1))

    return 0.0


def _carimbo_do_nome(filename: str) -> str:
    """Retorna carimbo somente quando o nome contém BB_XXXXXXX estrito."""
    stem = Path(filename).stem
    m = re.search(r"(?i)\bBB_(\d{7})\b", stem)
    if m:
        return f"BB_{m.group(1)}"
    return ""


def _mes_ref_master(data_ref: object) -> str:
    if hasattr(data_ref, "month") and hasattr(data_ref, "year"):
        return f"{int(data_ref.month):02d}-{int(data_ref.year)}"
    txt = str(data_ref or "").strip()
    if re.fullmatch(r"\d{2}-\d{4}", txt):
        return txt
    if re.fullmatch(r"\d{2}/\d{4}", txt):
        return txt.replace("/", "-")
    return ""


def _inferir_sistema_enel(text: str) -> str:
    upper = text.upper()
    if _is_enel_rj(text):
        return "ENEL_RJ"
    if "07.110.651" in text or "ENEL DISTRIBUICAO CEARA" in upper or "COELCE" in upper:
        return "ENEL_CE"
    return "ENEL"


def _estado_por_sistema_enel(sistema: str) -> str:
    mapa = {
        "ENEL": "SÃO PAULO",
        "ENEL_SP": "SÃO PAULO",
        "ENEL_CE": "CEARÁ",
        "ENEL_RJ": "RIO DE JANEIRO",
    }
    return mapa.get(str(sistema or "").strip().upper(), "")


def _resolver_carimbo_master(filename: str, instalacao: object, data_ref: object, text: str) -> str:
    carimbo_nome = _carimbo_do_nome(filename)
    stem = Path(filename).stem
    if carimbo_nome:
        return str(carimbo_nome).upper()
    if re.search(r"(?i)\bBB_\d+\b", stem):
        raise ValueError(f"Carimbo BB invalido no nome do arquivo: {filename}")

    uc = str(instalacao or "").strip()
    uc_digits = re.sub(r"\D", "", uc)
    mes_ref = _mes_ref_master(data_ref)
    sistema = _inferir_sistema_enel(text)
    master_csv = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/indice_master.csv")

    rows: list[dict] = []
    if master_csv.exists():
        try:
            for enc in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    with open(master_csv, newline="", encoding=enc) as f:
                        rows = list(csv.DictReader(f))
                    break
                except UnicodeDecodeError:
                    continue
        except OSError:
            rows = []

    arquivo_nome = Path(filename).name.lower()

    def _linha_enel_atual(row: dict) -> str:
        indice = str(row.get("INDICE") or "").strip().upper()
        if not indice.startswith("BB_"):
            return ""
        arquivo_master = Path(str(row.get("ARQUIVO") or "").strip()).name.lower()
        if arquivo_master and arquivo_master == arquivo_nome:
            return indice

        uc_master = str(row.get("UC") or "").strip()
        uc_master_digits = re.sub(r"\D", "", uc_master)
        mes_master = str(row.get("MES_REF") or "").strip()
        sist_master = str(row.get("SISTEMA") or "").strip().upper()
        if uc and mes_ref and mes_master == mes_ref:
            if uc_master == uc or (uc_digits and uc_master_digits == uc_digits):
                if sist_master.startswith("ENEL"):
                    return indice
        return ""

    for row in reversed(rows):
        indice_atual = _linha_enel_atual(row)
        if indice_atual:
            return indice_atual

    if uc and mes_ref:
        try:
            from indice_master import MasterIndice
            master = MasterIndice()
            ja_baixado = master.ja_foi_baixado(uc, mes_ref, sistema)
            if not ja_baixado:
                novo = master.consumir_carimbo()
                master.registrar(
                    indice_bb=novo,
                    sistema=sistema,
                    uc=uc,
                    mes_ref=mes_ref,
                    estado=_estado_por_sistema_enel(sistema),
                    arquivo=Path(filename).name,
                )
                return novo

            # Índices legados/individuais podem registrar origem ENEL sem uma
            # linha mestre utilizável para o PDF atual. O bloqueio por dedup só
            # é definitivo quando há linha exata do nome ou do mês/referência.
            for row in reversed(rows):
                indice_atual = _linha_enel_atual(row)
                if indice_atual:
                    return indice_atual

            novo = master.consumir_carimbo()
            master.registrar(
                indice_bb=novo,
                sistema=sistema,
                uc=uc,
                mes_ref=mes_ref,
                estado=_estado_por_sistema_enel(sistema),
                arquivo=Path(filename).name,
            )
            return novo
        except Exception as exc:
            log.warning("Falha ao resolver carimbo ENEL pelo indice master: %s", exc)

    raise ValueError(
        f"Nao foi possivel resolver carimbo BB para {filename}: "
        f"instalacao={uc!r}, mes_ref={mes_ref!r}"
    )


# =============================================================================
# EXTRATORES COMUNS
# =============================================================================

def _get_instalacao_bt(text: str) -> Optional[str]:
    """
    T0: BTE/MTE prefix format — BTE0013944 / 100000027614 (extract number after /)
    T1: padrao XXXXXXXXXX / YYYYYYYYYYYY
    T2: linhas com palavras-chave + 10 digitos
    T3: varredura geral — primeiro 10 digitos nao-data
    Leading zeros stripped for ENEL SP compatibility (Consen expects no leading zero).
    """
    def _strip(s: str) -> str:
        return s.lstrip("0") or s

    # T0: BTE/MTE prefix — return the code itself (BTE0013944 is the UC in Consen)
    m = re.search(r"([BM]TE\d+)\s*/\s*\d{10,}", text)
    if m:
        return m.group(1)  # BTE0013944, not the number after /

    # T1: linha tabular comum BT "UC / cliente R$ valor"
    m = re.search(r"\b(\d{6,12})\s*/\s*\d{6,12}\s+R\$\s*[\d.,]+", text)
    if m and not _parece_data(m.group(1)):
        return _strip(m.group(1))

    # T2: 10-digit number / anything
    m = re.search(r"(\d{10})\s*/\s*\d+", text)
    if m and not _parece_data(m.group(1)):
        return _strip(m.group(1))

    # T3: keyword lines with 7-12 digits
    for ln in text.splitlines():
        upper = ln.upper()
        if any(kw in upper for kw in ("INSTALA", "N UC", "N.UC", "CLIENTE", "CONTRATO", "UC:")):
            m = re.search(r"\b(\d{7,12})\b", ln)
            if m and not _parece_data(m.group(1)):
                return _strip(m.group(1))

    # T4: general scan — first 7-12 digit non-date number
    for m in re.finditer(r"\b(\d{7,12})\b", text):
        cand = m.group(1)
        if not _parece_data(cand):
            return _strip(cand)

    return None


def _get_instalacao_mt(text: str) -> Optional[str]:
    """MTE0018763 / 100082181175 ou 0200001619 / 200002520578 ou UC/UC R$ (ENEL CE/RJ)"""
    def _strip(s: str) -> str:
        return s.lstrip("0") or s

    # T0: MTE prefix — ENEL SP MT (usa o código MTE como instalação, não o numérico)
    m = re.search(r"(MTE\d{7})\s*/\s*\d{10,}", text)
    if m:
        return m.group(1)

    # T1a: <10-digit UC> / <12-digit code> R$<valor> — ENEL SP MT (A4 Verde)
    # O primeiro número é a instalação; o segundo é código de cliente/ponto de medição.
    m = re.search(r"\b(\d{10})\s*/\s*\d{12}\s+R\$", text)
    if m and not _parece_data(m.group(1)):
        return _strip(m.group(1))

    # T1: 10 digits / 12 digits sem R$ — outro variante ENEL SP MT
    m = re.search(r"(\d{10})\s*/\s*(\d{12})", text)
    if m:
        return _strip(m.group(2))

    # T2: UC/UC R$ — ENEL CE / Ampla RJ MT (mesmo número repetido antes do valor)
    # UCs curtos (3-4 dígitos) para Ampla RJ; longos (7-10) para ENEL CE
    m = re.search(r"\b(\d{2,12})\s*/\s*\1\s+R\$\s*[\d.,]+", text)
    if m and not _parece_data(m.group(1)):
        return _strip(m.group(1))

    # T3: fallback BT-style — {3-12 dígitos} / {3-12 dígitos} R$ valor
    m = re.search(r"\b(\d{3,12})\s*/\s*\d{3,12}\s+R\$\s*[\d.,]+", text)
    if m and not _parece_data(m.group(1)):
        return _strip(m.group(1))

    return None


def _get_cnpj(text: str) -> Optional[str]:
    """
    Retorna o CNPJ do cliente (não da distribuidora).
    A ENEL SP imprime o CNPJ/CPF do cliente após a UC na linha da fatura.
    Filtra distribuidoras conhecidas e rejeita 'CNPJs' cujos dígitos
    coincidem com valores monetários da fatura (falso-positivo comum).
    """
    candidatos = re.findall(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", text)
    for c in candidatos:
        if not any(c.startswith(d) for d in _DISTRIBUIDORAS):
            # Rejeita se os dígitos sem formatação parecem valor monetário
            # (ex: '13.890.000/4548-61' vem de '4548,61' mal parseado)
            digitos = re.sub(r"\D", "", c)
            # CNPJ real: raiz (8 dígitos) não pode ser 00000000
            raiz = digitos[:8]
            if raiz == "00000000":
                continue
            # Rejeita se os 4 dígitos após a barra são suspeitamente iguais
            # ao valor da fatura (padrão de falso-positivo ENEL)
            # Aceita normalmente
            return c
    # Tenta formato sem pontuação (14 dígitos)
    for c in re.findall(r"\b(\d{14})\b", text):
        fmt = f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
        if not any(fmt.startswith(d) for d in _DISTRIBUIDORAS):
            return fmt
    return None


def _get_endereco(text: str) -> Optional[str]:
    lines = text.splitlines()
    for i, ln in enumerate(lines):
        if re.search(r"\d{5}-\d{3}", ln):
            prev = lines[i - 1].strip() if i > 0 else ""
            if "NOTA FISCAL" in prev.upper():
                prev = ""
            return f"{prev} {ln.strip()}".strip()
    return None


def _get_fisco(text: str) -> dict:
    out = {"icms": 0.0, "pis": 0.0, "cofins": 0.0,
           "aliq_icms": 0.0, "aliq_pis": 0.0, "aliq_cofins": 0.0}

    m = re.search(r"PIS/PASEP\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", text)
    if m:
        out["aliq_pis"] = _br2f(m.group(2))
        out["pis"]      = _br2f(m.group(3))

    m = re.search(r"COFINS\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", text)
    if m:
        out["aliq_cofins"] = _br2f(m.group(2))
        out["cofins"]      = _br2f(m.group(3))

    m = None
    for m in re.finditer(r"I\s*CMS\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", text, re.IGNORECASE):
        pass  # pega o ultimo consolidado encontrado
    if m:
        out["aliq_icms"] = _br2f(m.group(2))
        out["icms"] = _br2f(m.group(3))

    if not out["icms"]:
        for ln in text.splitlines():
            if "TOTAL" not in ln.upper():
                continue
            vals = re.findall(r"[\d.]+,\d{2}", ln)
            if len(vals) >= 4:
                out["icms"] = _br2f(vals[-1])
                break

    if not out["icms"]:
        for ln in text.splitlines():
            vals = re.findall(r"[\d.]+,\d{2}", ln)
            if len(vals) == 4 and "SUBTOTAL" not in ln.upper():
                out["icms"] = _br2f(vals[-1])
                break

    return out


def _get_datas_e_total(text: str, formato: str = "BT") -> dict:
    out = {"emissao": None, "vcto": None, "ref": None,
           "total": 0.0, "ant": None, "atu": None}

    # --- Emissao ---
    m = re.search(r"DATA DE EMISS[AÃ]O[:\s]+(\d{2}/\d{2}/\d{4})", text, re.IGNORECASE)
    if m:
        out["emissao"] = _parse_date(m.group(1))

    # --- Referencia ---
    m = re.search(r"(?:REFERENTE\s+A|REFERENCIA|REF\.?)\s*[:\-]?\s*(\d{2}/\d{4})", text, re.IGNORECASE)
    if m:
        out["ref"] = _ref2date(m.group(1))
    if not out["ref"]:
        # Linha da página 2 ENEL: "DD/MM/YYYY  MM/YYYY  DD/MM/YYYY" (emissão ref vcto)
        m = re.search(r"\d{2}/\d{2}/\d{4}\s+(0[1-9]|1[0-2])/(20\d{2})\s+\d{2}/\d{2}/\d{4}", text)
        if m:
            out["ref"] = _ref2date(f"{m.group(1)}/{m.group(2)}")
    if not out["ref"]:
        # Fallback: MM/YYYY isolado — lookbehind para não capturar dentro de DD/MM/YYYY
        m = re.search(r"(?<!\d/)(0[1-9]|1[0-2])/(20\d{2})(?!\d)", text)
        if m:
            out["ref"] = _ref2date(f"{m.group(1)}/{m.group(2)}")

    # --- Vencimento (7 taticas em ordem de precisao) ---
    vcto = None

    # T0: padrão ENEL — "DD/MM/AAAA consulta" (linha com vencimento + link de consulta NF)
    # Mais confiável que T1 pois aparece isolado na linha do documento
    m = re.search(r"^(\d{2}/\d{2}/\d{4})\s+consulta", text, re.IGNORECASE | re.MULTILINE)
    if m:
        vcto = m.group(1)

    # T0b: linha do boleto "NOSSO Nº / DOC / VENCIMENTO valor" — ex: "109/xxx-6 039378051 23/03/2026 R$7.506,30"
    if not vcto:
        m = re.search(r"\d{3}/\d{7,}-\d\s+\d{9}\s+(\d{2}/\d{2}/\d{4})", text)
        if m:
            vcto = m.group(1)

    # T1: linha tabular ENEL  "DD/MM/AAAA  NNNNNNNN  MM/AAAA  DD/MM/AAAA"
    m = re.search(r"(\d{2}/\d{2}/\d{4})\s+\d{8,12}\s+\d{2}/\d{4}\s+(\d{2}/\d{2}/\d{4})", text)
    if m:
        vcto = m.group(2)

    # T2: codigo de barras com digito verificador + 9 digitos + data
    if not vcto:
        m = re.search(r"\d{11,}-\d\s+\d{9}\s+(\d{2}/\d{2}/\d{4})", text)
        if m:
            vcto = m.group(1)

    # T3b: tabela "Data de emissão ... Vencimento\nDD/MM emissao MM/YYYY DD/MM vcto"
    # Ampla RJ / ENEL RJ — cabeçalho tem colunas na ordem emissão | ref | vencimento
    if not vcto:
        m = re.search(
            r"Data\s+de\s+emiss[aã]o.*?Vencimento\s*\n"
            r"\d{2}/\d{2}/\d{4}\s+\d{2}/\d{4}\s+(\d{2}/\d{2}/\d{4})",
            text, re.IGNORECASE,
        )
        if m:
            vcto = m.group(1)

    # T3: palavra VENCIMENTO diretamente seguida de data
    if not vcto:
        m = re.search(r"VENCIMENTO\s*[:\-]?\s*(\d{2}/\d{2}/\d{4})", text, re.IGNORECASE)
        if m:
            vcto = m.group(1)

    # T4: DEBITO AUTOMATICO / DEBITADO EM / DATA DO DEBITO
    if not vcto:
        m = re.search(
            r"(?:D[EÉ]BITO\s+AUTOM[AÁ]TICO|DEBITADO\s+EM|DATA\s+DO\s+D[EÉ]BITO)"
            r"[\s\S]{0,120}?(\d{2}/\d{2}/\d{4})",
            text, re.IGNORECASE,
        )
        if m:
            vcto = m.group(1)

    # T5: PAGAVEL / PAGAVEL EM + data
    # Exige data > emissão para não capturar Leitura Atual que antecede o vencimento em alguns
    # layouts COELCE: "PAGAVEL ... 09/05/2026 09 R$ 20/06/2026" (09/05 = leitura, 20/06 = vcto)
    if not vcto:
        m = re.search(r"PAG[AÁ]VEL[\s\S]{0,200}?(\d{2}/\d{2}/\d{4})", text, re.IGNORECASE)
        if m:
            _cand_t5 = _parse_date(m.group(1))
            if not out["emissao"] or (_cand_t5 and _cand_t5 > out["emissao"]):
                vcto = m.group(1)

    # T6: TOTAL A PAGAR + data proxima (gap ampliado)
    if not vcto:
        m = re.search(r"TOTAL\s+A\s+PAGAR[\s\S]{0,200}?(\d{2}/\d{2}/\d{4})", text, re.IGNORECASE)
        if m:
            vcto = m.group(1)

    # T6b: linha curta com data + valor (formato compacto BT sem cabeçalho)
    if not vcto:
        for ln in text.splitlines():
            m = re.search(r"(\d{2}/\d{2}/\d{4})\s+([\d\.]+,\d{2})$", ln.strip())
            if m and len(ln.strip()) < 80:
                vcto = m.group(1)
                if not out["total"]:
                    out["total"] = _br2f(m.group(2))
                break

    # T6c: COELCE — data na mesma linha que "Protocolo de autorização" (NF eletrônica)
    # Na COELCE o vencimento aparece na linha: "DD/MM/YYYY [QR] Protocolo de autorização: ..."
    if not vcto:
        m = re.search(
            r"(\d{2}/\d{2}/\d{4})[^\n]*?Protocolo\s+de\s+autoriza",
            text, re.IGNORECASE,
        )
        if m:
            vcto = m.group(1)

    # Extrai data de "Próxima Leitura" para excluir do T7 (COELCE / outras distribuidoras)
    # Padrão: "leit_ant leit_atu N_dias proxima_leitura" na mesma linha
    _m_prox = re.search(
        r"\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+\d{1,3}\s+(\d{2}/\d{2}/\d{4})",
        text,
    )
    _proxima_leitura_dt = _parse_date(_m_prox.group(1)) if _m_prox else None

    # T7: fallback — primeira data do documento posterior a emissao
    # Exclui a data de Próxima Leitura para não confundi-la com o vencimento
    if not vcto and out["emissao"]:
        for d, mo, y in RE_DATE.findall(text):
            try:
                cand = dt.date(int(y), int(mo), int(d))
                if cand > out["emissao"] and cand != _proxima_leitura_dt:
                    vcto = f"{d}/{mo}/{y}"
                    break
            except Exception:
                continue

    if vcto:
        out["vcto"] = _parse_date(vcto)

    # --- Valor Total ---
    # Ordem: MTE/num R$  >  10+digitos/num R$  >  TOTAL A PAGAR  >  TOTAL R$
    #        > BTE R$ (next line)  >  TOTAL multi-col (BTE large commercial)
    m = re.search(r"MTE\d{7}\s*/\s*\d+\s+R\$\s*([\d.,]+)", text)
    if not m:
        m = re.search(r"\d{10,}\s*/\s*\d+\s+R\$\s*([\d.,]+)", text)
    if not m:
        m = re.search(r"\d{6,12}\s*/\s*\d{6,12}\s+R\$\s*([\d.,]+)", text)
    if not m:
        m = re.search(r"TOTAL\s+A\s+PAGAR\s+R\$\s*([\d.,]+)", text, re.IGNORECASE)
    if not m:
        m = re.search(r"TOTAL\s+R\$\s*([\d.,]+)", text, re.IGNORECASE)
    if not m:
        # BTE format: R$VALUE on its own line (after BTE/UC reference)
        m = re.search(r"^R\$([\d.]+,\d{2})$", text, re.MULTILINE)
    if not m:
        # BTE large-commercial: "TOTAL val1 val2 val3 val4" — first value is the total to pay
        m = re.search(r"\bTOTAL\b\s+([\d.]+,\d{2})(?:\s+[\d.]+,\d{2}){2,}", text, re.IGNORECASE)
    if m:
        out["total"] = _br2f(m.group(1))

    # --- Datas de Leitura: 5 estratégias em cascata ---
    leit_ant = None
    leit_atu = None

    # T0: cabeçalho ENEL "segunda via" — "Monofásico/Bifásico/Trifásico leit_ant leit_atu dias"
    # ex: "Trifásico 09/04/2026 11/05/2026 32 09/06/2026"
    if not leit_ant:
        m = re.search(
            r"(?:Monof[aá]sico|Bif[aá]sico|Trif[aá]sico)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+\d+",
            text, re.IGNORECASE)
        if m:
            leit_ant, leit_atu = m.group(1), m.group(2)

    # T1: "LEITURA ANTERIOR ... data ... LEITURA ATUAL ... data" (até 120 chars entre)
    if not leit_ant:
        m = re.search(
            r"LEITURA\s+ANTERIOR[\s\S]{0,120}?(\d{2}/\d{2}/\d{4})[\s\S]{0,120}?"
            r"LEITURA\s+ATUAL[\s\S]{0,120}?(\d{2}/\d{2}/\d{4})",
            text, re.IGNORECASE)
        if m:
            leit_ant, leit_atu = m.group(1), m.group(2)

    # T2: par de datas ancorado em PERIODO / LEITURA na mesma região
    if not leit_ant:
        m = re.search(
            r"(?:LEITURA|PER[IÍ]ODO|PERIODO)[\s\S]{0,120}?(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})",
            text, re.IGNORECASE)
        if m:
            leit_ant, leit_atu = m.group(1), m.group(2)

    # T3: linha tabular ENEL — "data UC ref vcto" (apenas leitura anterior — emissão)
    if not leit_ant:
        m = re.search(r"(\d{2}/\d{2}/\d{4})\s+\d{8,12}\s+\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}", text)
        if m:
            leit_ant = m.group(1)

    # T4: "Anterior ... Atual" com par dd/mm próximos (layouts sem data completa)
    if not leit_ant:
        m = re.search(
            r"Anterior[^\d]{0,60}(\d{2}/\d{2})[^\d]{0,120}Atual[^\d]{0,60}(\d{2}/\d{2})",
            text, re.IGNORECASE | re.DOTALL)
        if m and out.get("emissao"):
            ano = out["emissao"].year
            def _dm(s):
                dd, mo = map(int, s.split("/"))
                for a in (ano, ano-1):
                    try: return dt.date(a, mo, dd).strftime("%d/%m/%Y")
                    except Exception: pass
                return None
            leit_ant = _dm(m.group(1))
            leit_atu = _dm(m.group(2))

    # T5: fallback robusto — par d1 < d2, gap 5-45 dias, não emissão nem vcto
    if not leit_ant:
        todas = RE_DATE.findall(text)
        for i in range(len(todas) - 1):
            try:
                d1 = dt.date(int(todas[i][2]),   int(todas[i][1]),   int(todas[i][0]))
                d2 = dt.date(int(todas[i+1][2]), int(todas[i+1][1]), int(todas[i+1][0]))
                diff = (d2 - d1).days
                if 5 <= diff <= 45 and d2 <= dt.date.today():
                    if out.get("emissao") and d1 == out["emissao"]: continue
                    if out.get("vcto")   and d2 == out["vcto"]:    continue
                    leit_ant = f"{todas[i][0]}/{todas[i][1]}/{todas[i][2]}"
                    leit_atu = f"{todas[i+1][0]}/{todas[i+1][1]}/{todas[i+1][2]}"
                    break
            except Exception:
                continue

    if leit_ant:
        out["ant"] = _parse_date(leit_ant)
    if leit_atu:
        out["atu"] = _parse_date(leit_atu)

    return out



# =============================================================================
# RESOLUÇÃO DE OBSERVAÇÕES (mapa texto PDF → value do select Consen)
# =============================================================================

_OBS_CODES_MULTAS_ENEL = {"6", "7", "8"}


_OBS_MAPA_ENEL = [
    (re.compile(r"compensa[c?][a?]o\s+(?:dic|fic)(?:\s*/\s*(?:dic|fic))?(?:\s+[a-z]{3}/\d{4})?", re.IGNORECASE), "11", True),
    (re.compile(r"\b(?:dic|fic)\s+(?:[a-z]{3,9}|\d{2})/\d{4}\b", re.IGNORECASE), "11", True),
    (re.compile(r"compensa[c?][a?]o\s+dic\s+mensal|dic\s+mensal", re.IGNORECASE), "58", True),
    (re.compile(r"penal(?:idade|\.?)?\s*(?:dic|dmic|fic|dicri)|penali[zs]a[c?][a?]o\s+(?:dic|dmic|fic|dicri)", re.IGNORECASE), "149", False),
    (re.compile(r"restitui[c?][a?]o\s+de\s+pagamento", re.IGNORECASE), "109", True),
    (re.compile(r"\bmulta\b", re.IGNORECASE), "6", False),
    (re.compile(r"\bjuros?\s+(?:mora|por\s+atraso)\b|\bjuros\b", re.IGNORECASE), "7", False),
    (re.compile(r"(?:corre[c?][a?]o|atualiza[c?][a?]o)\s+monet[a?]ria", re.IGNORECASE), "8", False),
    (re.compile(r"acerto\s+de\s+faturamento", re.IGNORECASE), "259", False),
    (re.compile(r"pc\.?6/006|cob\s+avaria|avaria\s+art\s*115", re.IGNORECASE), "230", False),
    (re.compile(r"saldo\s+para\s+o\s+pr[o?]ximo", re.IGNORECASE), "110", False),
    (re.compile(r"devolu[c?][a?]o\s+pagamento\s+indevido", re.IGNORECASE), "59", False),
    # PC.XX/XX-FATURA-MM/YYYY-ART.323 REN NNNN — parcelamento ENEL SP (REN 1000/ART.323)
    (re.compile(r"\bPC\.\d+/\d+[-\s]+FATURA\b|\bART\.323\s+REN\b", re.IGNORECASE), "201", False),
    (re.compile(r"\bparcelamento\b", re.IGNORECASE), "100", False),
]


def _resolver_obs_enel(text: str, itens: dict) -> list:
    """
    Retorna lista de pares [(obsCod, obsValor), ...] ordenados por prioridade.
    obsCod  = value do <select> cb-dados-financeiros-obs
    obsValor = valor R$ (negativo = cr?dito ao cliente)
    """
    linhas = [re.sub(r"\s+", " ", ln).strip() for ln in (text or "").splitlines() if ln.strip()]
    encontrados: dict = {}

    for i, ln in enumerate(linhas):
        for prio, (pat, cod, eh_credito) in enumerate(_OBS_MAPA_ENEL):
            if cod in encontrados and encontrados[cod][1] <= prio:
                continue
            if pat.search(ln):
                monies = RE_MONEY.findall(ln)
                pref_first_money = {"11"}
                pick_first = cod in pref_first_money
                v = _br2f(monies[0] if pick_first else monies[-1]) if monies else None
                if v is None and i + 1 < len(linhas):
                    m2 = RE_MONEY.findall(linhas[i + 1])
                    if m2:
                        v = _br2f(m2[0] if pick_first else m2[-1])
                if v is not None and abs(v) > 0.005:
                    encontrados[cod] = (-abs(v) if eh_credito else abs(v), prio)

    dic_val = itens.get("dic", 0.0) or 0.0
    fic_val = itens.get("fic", 0.0) or 0.0
    dic_fic_total = round(dic_val + fic_val, 2)
    if dic_fic_total > 0.005:
        encontrados["11"] = (-abs(dic_fic_total), 999)

    return [
        (cod, val)
        for cod, (val, _p) in sorted(encontrados.items(), key=lambda x: x[1][1])
        if cod not in _OBS_CODES_MULTAS_ENEL
    ]



# =============================================================================
# EXTRATOR BT (B3 Convencional)
# =============================================================================

_B3_END_MARKERS = ["DADOS DE MEDI", "MEDIDOR", "RESPONSAVEL PELA ILUMINA"]


def _itens_bt(text: str) -> dict:
    out = {
        "tusd_kwh": 0, "energia_kwh": 0,
        "val_tusd": 0.0, "val_te": 0.0,
        "val_bandeira": 0.0,
        "subtotal_faturamento": 0.0, "subtotal_outros": 0.0,
        "cosip": 0.0, "multa": 0.0, "juros": 0.0, "correcao_monetaria": 0.0,
        "dic": 0.0, "fic": 0.0,
        "ret_lei_9430_perc": 5.85, "ret_lei_9430_val": 0.0,
    }
    lines = text.splitlines()
    start, end = -1, len(lines)

    for i, ln in enumerate(lines):
        upper = ln.upper()
        if start == -1 and any(mk in upper for mk in ["ITENS DE FATURA", "DESCRI", "USO SIST"]):
            start = i
        if start != -1 and any(mk in upper for mk in _B3_END_MARKERS):
            end = i
            break

    if start == -1:
        start = 0

    for ln in lines[start:end]:
        if not ln.strip() or "ITENS DE FATURA" in ln.upper():
            continue
        upper = ln.upper()
        monies = RE_MONEY.findall(ln)
        m_kwh = re.search(r"KWH\s+([\d.,]+)", upper)
        kwh = int(_br2f(m_kwh.group(1))) if m_kwh else 0

        if "SUBTOTAL" in upper and "FATURAMENTO" in upper:
            if monies:
                out["subtotal_faturamento"] = _br2f(monies[0])
        elif "SUBTOTAL" in upper and "OUTROS" in upper:
            if monies:
                out["subtotal_outros"] = _br2f(monies[0])
        elif "TUSD" in upper and "KWH" in upper:
            if kwh:
                out["tusd_kwh"] += kwh
            m_v = re.search(r"KWH\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)", ln, re.IGNORECASE)
            if m_v:
                out["val_tusd"] += _br2f(m_v.group(1))
            elif len(monies) >= 3:
                out["val_tusd"] += _br2f(monies[2])
        elif ("ENERGIA" in upper and "(TE)" in upper.replace("  ", " ")) or (" FORNECIDA TE " in f" {upper} " and "KWH" in upper):
            if kwh:
                out["energia_kwh"] += kwh
            m_v = re.search(r"KWH\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)", ln, re.IGNORECASE)
            if m_v:
                out["val_te"] += _br2f(m_v.group(1))
            elif len(monies) >= 3:
                out["val_te"] += _br2f(monies[2])
        elif "ADICIONAL" in upper and ("BANDEIRA" in upper or "BAND" in upper):
            # Taxas/kWh são sempre < 1,0; o total R$ da bandeira é sempre >= 1,0.
            # Pula valores de taxa iniciais e pega o primeiro valor monetário real.
            for mv in monies:
                f = _br2f(mv)
                if f >= 1.0:
                    out["val_bandeira"] += f
                    break
        elif "RET" in upper and "9430" in upper:
            # Trailing minus: "185,51-" (ENEL SP) ou leading minus: "-185,51"
            m_neg = re.search(r"([\d.,]{4,})-(?=\s|$)", ln)
            if m_neg:
                out["ret_lei_9430_val"] = -abs(_br2f(m_neg.group(1)))
            else:
                for val in monies:
                    if val.startswith("-"):
                        out["ret_lei_9430_val"] = _br2f(val)
                        break
        elif "RETEN" in upper and "TRIBUT" in upper and "FEDER" in upper:
            # COELCE: "Retenção De Tributos Federais kW 0 0,00000 185,51- ..."
            m_neg = re.search(r"([\d.,]{4,})-(?=\s|$)", ln)
            if m_neg:
                out["ret_lei_9430_val"] = -abs(_br2f(m_neg.group(1)))
            else:
                for val in monies:
                    if val.startswith("-"):
                        out["ret_lei_9430_val"] = _br2f(val)
                        break
        elif ("COSIP" in upper or "CIP" in upper) and not re.search(r"\d{2}/\d{4}", ln):
            if monies:
                out["cosip"] += _br2f(monies[0])
        elif "MULTA" in upper and monies:
            out["multa"] += _br2f(monies[-1])
        elif (("ATUALIZA" in upper or "CORRE" in upper) and "MONET" in upper) and monies:
            out["correcao_monetaria"] += _br2f(monies[-1])
        elif ("JUROS" in upper or "MORA" in upper) and monies:
            out["juros"] += _br2f(monies[-1])
        elif re.search(r"compensa[c?][a?]o\s+dic", ln, re.IGNORECASE) and monies:
            out["dic"] += abs(_br2f(monies[-1]))
        elif re.search(r"compensa[c?][a?]o\s+fic", ln, re.IGNORECASE) and monies:
            out["fic"] += abs(_br2f(monies[-1]))
        elif re.search(r"\bdic\s+(?:[a-z]{3,9}|\d{2})/\d{4}\b", ln, re.IGNORECASE) and monies:
            out["dic"] += abs(_br2f(monies[0]))
        elif re.search(r"\bfic\s+(?:[a-z]{3,9}|\d{2})/\d{4}\b", ln, re.IGNORECASE) and monies:
            out["fic"] += abs(_br2f(monies[0]))

        if not out["tusd_kwh"] and not out["energia_kwh"]:
            m_hfp = re.search(
                r"ENERGIA\s+ATIVA\s*-\s*KWH\s+HFP\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)",
                upper,
                re.IGNORECASE,
            )
            if m_hfp:
                out["energia_kwh"] = int(_br2f(m_hfp.group(1)))

    if out["subtotal_outros"] > 0 and out["multa"] > 0 and out["juros"] < 0.01 and out["correcao_monetaria"] < 0.01:
        out["juros"] = out["subtotal_outros"] - out["multa"]

    # Fallback: bandeira não encontrada na tabela de itens
    if out["val_bandeira"] < 0.01:
        for ln in lines:
            upper = ln.upper()
            if ("ADICIONAL" in upper or "ACRESCIMO" in upper) and ("BANDEIRA" in upper or "BAND" in upper):
                monies_all = RE_MONEY.findall(ln)
                for mv in monies_all:
                    v = _br2f(mv)
                    if v >= 1.0:
                        out["val_bandeira"] = v
                        break
                if out["val_bandeira"] >= 1.0:
                    break

    # Fallback: multas não encontradas na tabela de itens (estão em seção de obs separada)
    for ln in lines:
        upper = ln.upper()
        monies_all = RE_MONEY.findall(ln)
        if not monies_all:
            continue
        v = _br2f(monies_all[-1])
        if not v or abs(v) < 0.01:
            continue
        if out["multa"] < 0.01 and "MULTA" in upper and "BANDEIRA" not in upper:
            out["multa"] += abs(v)
        elif out["correcao_monetaria"] < 0.01 and ("ATUALIZA" in upper or "CORRE" in upper) and "MONET" in upper:
            out["correcao_monetaria"] += abs(v)
        elif out["juros"] < 0.01 and ("JUROS" in upper or "MORA" in upper) and "DATA" not in upper and "EMISSAO" not in upper:
            out["juros"] += abs(v)

    return out




# =============================================================================
# TARIFA BRANCA — detecção e extrator de itens por posto
# =============================================================================

def _is_tarifa_branca(text: str) -> bool:
    """
    Detecta Tarifa Branca ENEL com múltiplos padrões.
    A ENEL SP pode grafar de formas diferentes dependendo da versão da fatura.
    """
    padroes = [
        r"B3\s*[-–]\s*BRANCA",          # B3 - BRANCA  ou  B3 – BRANCA
        r"TARIFA\s+BRANCA",              # TARIFA BRANCA
        r"BRANCA\s+B3",                  # BRANCA B3
        r"MODALIDADE\s+BRANCA",          # MODALIDADE BRANCA
        r"POSTO\s+(?:PONTA|INTERMEDIAR)", # Posto Ponta / Posto Intermediário (exclusivo da Branca)
        r"INTERMEDI[ÁA]RI[AO]\s+KWH",    # INTERMEDIÁRIO KWH na tabela de itens
        r"USO\s+SIST.*PONTA.*KWH",       # USO SIST. DISTR. PONTA KWH
    ]
    for p in padroes:
        if re.search(p, text, re.IGNORECASE):
            return True
    return False


def _itens_bt_branca(pdf_path: str) -> dict:
    """
    Extrai itens de fatura para Tarifa Branca usando extract_words com x_tolerance=2.
    A tabela de itens tem colunas em posições x fixas; a coluna Valor(R$) está
    em torno de x=224-228. Posta separados: Ponta, Fora Ponta, Intermediário.
    Cada posto tem duas linhas: TUSD (USO SIST. DISTR.) + TE (ENERGIA).
    fatConPonta*          = TUSD Ponta  + TE Ponta
    fatConFPontaInd*      = TUSD FPonta + TE FPonta
    fatConIntermed*       = TUSD Intermed + TE Intermed
    """
    from collections import defaultdict

    out = {
        # kWh por posto
        "kwh_pta": 0, "kwh_fpta": 0, "kwh_inter": 0,
        # R$ por posto (TUSD + TE somados)
        "val_pta": 0.0, "val_fpta": 0.0, "val_inter": 0.0,
        # R$ TUSD e TE separados (para referência)
        "tusd_pta": 0.0, "te_pta": 0.0,
        "tusd_fpta": 0.0, "te_fpta": 0.0,
        "tusd_inter": 0.0, "te_inter": 0.0,
        # Outros
        "cosip": 0.0, "ret_lei_9430_val": 0.0, "ret_lei_9430_perc": 5.85,
        "subtotal_faturamento": 0.0, "val_bandeira": 0.0,
        "multa": 0.0, "juros": 0.0, "correcao_monetaria": 0.0,
        "dic": 0.0, "fic": 0.0,                    # Compensação DIC/FIC
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            words = page.extract_words(x_tolerance=2, y_tolerance=2)
    except Exception:
        return out

    # Agrupar words por linha (y arredondado a 3px)
    linhas_w = defaultdict(list)
    for w in words:
        y_key = round(w["top"] / 3) * 3
        linhas_w[y_key].append(w)

    # Coluna Valor(R$): x entre 215 e 245 (posição confirmada nos dois PDFs de exemplo)
    X_VALOR_MIN, X_VALOR_MAX = 215, 248

    def _val_col(wds):
        """Retorna o valor R$ da coluna Valor(R$) (x≈224-228)."""
        for w in sorted(wds, key=lambda w: w["x0"]):
            if X_VALOR_MIN <= w["x0"] <= X_VALOR_MAX:
                v = _br2f(w["text"])
                if abs(v) > 0.01:
                    return v
        return 0.0

    def _kwh_col(wds):
        """Extrai kWh da coluna após KWH (x≈168-180)."""
        for i, w in enumerate(sorted(wds, key=lambda w: w["x0"])):
            if w["text"].upper() == "KWH" and i + 1 < len(wds):
                nxt = sorted(wds, key=lambda w: w["x0"])[i + 1]
                kwh = _br2f(nxt["text"])
                if kwh > 0:
                    return int(round(kwh))
        return 0

    for y in sorted(linhas_w.keys()):
        wds = sorted(linhas_w[y], key=lambda w: w["x0"])
        ln_text = " ".join(w["text"] for w in wds).upper()

        is_tusd    = "USO SIST" in ln_text or "TUSD" in ln_text
        is_te      = "(TE)" in ln_text or ("ENERGIA" in ln_text and "KWH" in ln_text and not is_tusd)
        is_ponta   = "PONTA" in ln_text and "FORA" not in ln_text and "INTERM" not in ln_text
        is_fpta    = "FORA" in ln_text and "PONTA" in ln_text
        is_inter   = "INTERM" in ln_text

        if is_tusd or is_te:
            val = _val_col(wds)
            kwh = _kwh_col(wds)

            if is_ponta:
                if is_tusd:
                    out["tusd_pta"] += val
                    if kwh: out["kwh_pta"] = kwh
                elif is_te:
                    out["te_pta"] += val
            elif is_fpta:
                if is_tusd:
                    out["tusd_fpta"] += val
                    if kwh: out["kwh_fpta"] = kwh
                elif is_te:
                    out["te_fpta"] += val
            elif is_inter:
                if is_tusd:
                    out["tusd_inter"] += val
                    if kwh: out["kwh_inter"] = kwh
                elif is_te:
                    out["te_inter"] += val

        elif "COSIP" in ln_text or ("CIP" in ln_text and "MUNIC" in ln_text):
            v = _val_col(wds)
            if v > 0: out["cosip"] += v

        elif "RET" in ln_text and "9430" in ln_text:
            # Valor negativo — buscar em toda a linha
            for w in wds:
                v = _br2f(w["text"])
                if v < -0.01:
                    out["ret_lei_9430_val"] = v
                    break
        elif "RETEN" in ln_text and "TRIBUT" in ln_text and "FEDER" in ln_text:
            for w in wds:
                v = abs(_br2f(w["text"]))
                if v > 0.01:
                    out["ret_lei_9430_val"] = -v
                    break

        elif "SUBTOTAL" in ln_text and "FATURAMENTO" in ln_text:
            v = _val_col(wds)
            if v > 0: out["subtotal_faturamento"] = v

        elif "ADICIONAL" in ln_text and ("BANDEIRA" in ln_text or "BAND" in ln_text):
            v = _val_col(wds)
            if v > 0: out["val_bandeira"] += v

        elif "MULTA" in ln_text:
            v = _val_col(wds)
            if v > 0.01: out["multa"] += v

        elif ("ATUALIZA" in ln_text or "CORRE" in ln_text) and "MONET" in ln_text:
            v = _val_col(wds)
            if v > 0.01: out["correcao_monetaria"] += v

        elif "JUROS" in ln_text or "MORA" in ln_text:
            v = _val_col(wds)
            if v > 0.01: out["juros"] += v

        elif re.search(r"compensa[cç][aã]o\s+dic", ln_text, re.IGNORECASE):
            v = _val_col(wds)
            if v > 0.01: out["dic"] += abs(v)

        elif re.search(r"compensa[cç][aã]o\s+fic", ln_text, re.IGNORECASE):
            v = _val_col(wds)
            if v > 0.01: out["fic"] += abs(v)

    # Calcular totais por posto
    out["val_pta"]   = round(out["tusd_pta"]   + out["te_pta"],   2)
    out["val_fpta"]  = round(out["tusd_fpta"]  + out["te_fpta"],  2)
    out["val_inter"] = round(out["tusd_inter"] + out["te_inter"], 2)

    return out


def _get_codigo_barras_enel(text: str) -> str:
    """
    Extrai a linha digitável do boleto ENEL SP.

    Formatos conhecidos:
    1) Linha digitável completa:  XXXXX.XXXXX XXXXX.XXXXXX XXXXX.XXXXXX X XXXXXXXXXXXXXX
    2) Formato ENEL SP compacto:  XXXXXXXXXX XXXXXXXXXX XXXXXXXXXX X XXXXXXXXXXXXXX
    3) Blocos separados por espaço na linha do código de barras
    4) Sequência de 47-48 dígitos contínuos (código de barras numérico)
    """
    # T1: formato padrão bancário com pontos (mais comum)
    m = re.search(
        r"(\d{5}\.\d{5}\s+\d{5}\.\d{6}\s+\d{5}\.\d{6}\s+\d\s+\d{14})",
        text)
    if m:
        return m.group(1).strip()

    # T2: ENEL SP — blocos sem ponto mas com espaço (ex: "83650000001 48480138001 53744003433 9 08100684946")
    m = re.search(
        r"(\d{10,11}\s+\d{10,11}\s+\d{10,11}\s+\d\s+\d{14,16})",
        text)
    if m:
        return m.group(1).strip()

    # T3: linha digitável compactada sem espaços intermediários (47-48 dígitos)
    m = re.search(r"\b(\d{47,48})\b", text)
    if m:
        raw = m.group(1)
        # Formata como XXXXX.XXXXX XXXXX.XXXXXX XXXXX.XXXXXX X XXXXXXXXXXXXXX
        try:
            return f"{raw[0:5]}.{raw[5:10]} {raw[10:15]}.{raw[15:21]} {raw[21:26]}.{raw[26:32]} {raw[32]} {raw[33:]}"
        except Exception:
            return raw

    # T4: busca linha que contenha explicitamente "CÓDIGO DE BARRAS" ou similar
    for ln in text.splitlines():
        upper = ln.upper()
        if any(kw in upper for kw in ("CODIGO DE BARRA", "CÓDIGO DE BARRA", "LINHA DIGIT")):
            numeros = re.findall(r"[\d.]+", ln)
            sequencia = "".join(re.sub(r"\D", "", n) for n in numeros)
            if len(sequencia) >= 44:
                return sequencia[:47]

    return ""


def extrair_bt(pdf_path: str) -> dict:
    text   = _extract_text(pdf_path)
    datas  = _get_datas_e_total(text, "BT")
    fisco  = _get_fisco(text)
    val_nf = _get_valor_nf(text)
    branca = _is_tarifa_branca(text)
    instalacao = _get_instalacao_bt(text)
    carimbo_master = _resolver_carimbo_master(Path(pdf_path).name, instalacao, datas["ref"], text)

    # ── Base comum ────────────────────────────────────────────────────────────
    base = {
        "fatCarimbo":             carimbo_master,
        "concCod":                "3",
        "Instalacao":             instalacao,
        "CNPJ":                   _get_cnpj(text),
        "cadSubGrupoCod":         "B3 [<2,3kV]",
        "fatDataReferencia":      datas["ref"],
        "fatDataEmissao":         datas["emissao"],
        "fatDataVcto":            datas["vcto"],
        "fatDataLeituraAnterior": datas["ant"],
        "fatDataLeituraAtual":    datas["atu"],
        "fatValorFatura":         datas["total"],
        "fatValorNotaFiscal":     val_nf,
        "fatICMS":                fisco["icms"],
        "fatDesIcmsAliquota":     fisco["aliq_icms"],
        "fatPIS":                 fisco["pis"],
        "fatDescPisAliquota":     fisco["aliq_pis"],
        "fatCOFINS":              fisco["cofins"],
        "fatDesCofinsAliquota":   fisco["aliq_cofins"],
        "fatDataCadastro":        dt.date.today(),
        "ENDERECO":               _get_endereco(text),
        "fatCodigoBarras":        _get_codigo_barras_enel(text),
        **{k: 0.0 for k in _RETENÇÕES_ZERO},
    }

    if branca:
        # ── Tarifa Branca: TUSD + TE somados por posto (Ponta / FPonta / Intermed) ──
        itens_b = _itens_bt_branca(pdf_path)
        base.update({
            "cadTarifaCod":                   "Branca",
            # Ponta
            "fatConPontaRegistrado":          itens_b["kwh_pta"],
            "fatConPontaFaturado":            itens_b["kwh_pta"],
            "fatConPontaValorReais":          itens_b["val_pta"],
            # Fora Ponta Indutivo
            "fatConFPontaIndRegistrado":      itens_b["kwh_fpta"],
            "fatConFPontaIndFaturado":        itens_b["kwh_fpta"],
            "fatConFPontaIndValorReais":      itens_b["val_fpta"],
            # Intermediário
            "fatConIntermediarioRegistrado":  itens_b["kwh_inter"],
            "fatConIntermediarioFaturado":    itens_b["kwh_inter"],
            "fatConIntermediarioValorReais":  itens_b["val_inter"],
            # Outros
            "fatIlumPublica":                 itens_b["cosip"],
            "fatValBandeira":                 itens_b["val_bandeira"],
            "fatDIC":                         itens_b["dic"],
            "fatFIC":                         itens_b["fic"],
            "fatMultas":                      round(
                itens_b["multa"] + itens_b["juros"] + itens_b["correcao_monetaria"], 2
            ),
            "fatMultasDiversas":              0.0,
            "fatTributoFederalPerc":          itens_b["ret_lei_9430_perc"],
            "fatTributoFederalVal":           itens_b["ret_lei_9430_val"],
            "_obs_list":                      _resolver_obs_enel(text, itens_b),
        })
    else:
        # ── Tarifa Convencional B3: fluxo original ────────────────────────────
        itens = _itens_bt(text)
        consumo = itens["tusd_kwh"] or itens["energia_kwh"]
        base.update({
            "cadTarifaCod":              "Convencional",
            "fatConFPontaIndRegistrado": consumo,
            "fatConFPontaIndFaturado":   consumo,
            "fatConFPontaIndValorReais": (
                round(itens["val_tusd"] + itens["val_te"], 2)
                if (itens["val_tusd"] + itens["val_te"]) > 0
                else round(itens["subtotal_faturamento"] - itens["val_bandeira"], 2)
            ),
            "fatIlumPublica":            itens["cosip"],
            "fatValBandeira":            itens["val_bandeira"],
            "fatDIC":                    itens["dic"],
            "fatFIC":                    itens["fic"],
            "fatMultas":                 round(
                itens["multa"] + itens["juros"] + itens["correcao_monetaria"], 2
            ),
            "fatMultasDiversas":         0.0,
            "fatTributoFederalPerc":     itens["ret_lei_9430_perc"],
            "fatTributoFederalVal":      itens["ret_lei_9430_val"],
            "_obs_list":                 _resolver_obs_enel(text, itens),
        })

    return base


# =============================================================================
# EXTRATOR MT (A4 Verde / A4 Azul)
# =============================================================================

def _detectar_subtarifa_mt(text: str) -> str:
    upper = text.upper()
    if "TARIFA AZUL" in upper or "A4 AZUL" in upper:
        return "A4_AZUL"
    return "A4_VERDE"


def _itens_mt(text: str) -> dict:
    out = {
        # Demanda contratada (do resumo "Demanda - KW")
        "demanda_contratada_kw": 0.0,
        # Demanda faturada — extraída da linha DEMANDA ÚNICA C/ DESCONTO
        "demanda_faturada_kw": 0.0,
        # Demanda excedente (ultrapassagem)
        "demanda_excedente_kw": 0.0, "demanda_excedente_val": 0.0,
        # Componentes tarifários da demanda (ÚNICA e ESTADUAL são a mesma kW, taxas diferentes)
        "demanda_unica_val": 0.0,       # DEMANDA ÚNICA C/ DESCONTO → R$
        "demanda_estadual_kw": 0.0,     # DEMANDA LEI ESTADUAL 16.886/18 → kW
        "demanda_estadual_val": 0.0,    # DEMANDA LEI ESTADUAL 16.886/18 → R$
        # Consumo
        "consumo_ponta_kwh": 0.0, "consumo_ponta_val": 0.0,
        "consumo_fponta_kwh": 0.0, "consumo_fponta_val": 0.0,
        # UFER — energia reativa excedente F. Ponta
        "ufer_fponta_kvarh": 0.0, "ufer_fponta_val": 0.0,
        # Benefício tarifário (ACL / Consumidor Especial)
        "beneficio_bruto": 0.0,
        "beneficio_liquido": 0.0,
        # Escassez hídrica
        "escassez_kwh": 0.0, "escassez_val": 0.0,
        # Desconto Fio B (kW demand e kWh energy components)
        "desconto_fio_kw": 0.0, "desconto_fio_kwh": 0.0,
        # Consumo Reativo Excedente Fp kVA — ENEL CE (fatConFPontaCap*)
        "cap_fponta_kva": 0.0, "cap_fponta_val": 0.0,
        # Totais e tributos
        "subtotal_faturamento": 0.0, "subtotal_outros": 0.0,
        "ret_lei_9430_perc": 5.85, "ret_lei_9430_val": 0.0,
        "cosip": 0.0, "multa": 0.0, "juros": 0.0, "correcao_monetaria": 0.0,
        # OBS 213 — cobranças/devoluções negativas (DIC, atualização negativa, etc.)
        "dic": 0.0, "fic": 0.0, "outras_cobra": 0.0,
        # Controle multi-linha benefício
        "_em_beneficio": False,
    }

    linhas = text.splitlines()
    for idx, ln in enumerate(linhas):
        upper = ln.upper()
        monies = RE_MONEY.findall(ln)
        prox = linhas[idx + 1] if idx + 1 < len(linhas) else ""
        prox_monies = RE_MONEY.findall(prox)

        if "SUBTOTAL" in upper and "FATURAMENTO" in upper and monies:
            out["subtotal_faturamento"] = _br2f(monies[0])
        elif "SUBTOTAL" in upper and "OUTROS" in upper and monies:
            out["subtotal_outros"] = _br2f(monies[0])

        # ── Demanda contratada (linha resumo "Demanda - KW X" / ENEL CE: "DEMANDA FORA PONTA - KW X") ─
        elif ("DEMANDA - KW" in upper or "DEMANDA-KW" in upper or "DEMANDA FORA PONTA - KW" in upper) and "GERACAO" not in upper:
            m = re.search(r"DEMANDA\s*(?:FORA\s+PONTA\s*)?-\s*KW\s+([\d.,]+)", upper)
            if m:
                out["demanda_contratada_kw"] = _br2f(m.group(1))

        # ── DEMANDA ÚNICA C/ DESCONTO → principal componente tarifário da demanda ──
        elif "DEMANDA" in upper and ("NICA" in upper or "UNICA" in upper) and "DESCONTO" in upper and "GERACAO" not in upper:
            m = re.search(r"KW\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln)
            if m:
                out["demanda_faturada_kw"] = _br2f(m.group(1))
                out["demanda_unica_val"] = _br2f(m.group(3))

        # ── DEMANDA LEI ESTADUAL 16.886/18 → componente TUSD estadual da demanda ──
        elif "DEMANDA" in upper and "16.886" in ln and "GERACAO" not in upper:
            m = re.search(r"KW\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln)
            if m:
                out["demanda_estadual_kw"] = _br2f(m.group(1))
                out["demanda_estadual_val"] = _br2f(m.group(3))

        # ── DEMANDA EXCEDENTE / ULTRAPASSAGEM F. PONTA ───────────────────────
        elif "DEMANDA" in upper and "EXCEDENTE" in upper and "GERACAO" not in upper:
            m = re.search(r"KW\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln)
            if m:
                out["demanda_excedente_kw"]  += _br2f(m.group(1))
                out["demanda_excedente_val"] += _br2f(m.group(3))

        # ── DEMANDA DISTRIBUIÇÃO SEM ICMS kW — ENEL CE (componente estadual) ──
        elif "DISTRIBUI" in upper and "SEM ICMS" in upper and "DEMANDA" in upper and "KW" in upper:
            m = re.search(r"KW\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln, re.IGNORECASE)
            if m:
                out["demanda_estadual_kw"]  = _br2f(m.group(1))
                out["demanda_estadual_val"] = _br2f(m.group(3))

        # ── DEMANDA DISTRIBUIÇÃO kW — ENEL CE (componente principal TUSD demanda) ─
        elif "DISTRIBUI" in upper and "DEMANDA" in upper and "SEM ICMS" not in upper and "KW" in upper:
            m = re.search(r"KW\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln, re.IGNORECASE)
            if m and out["demanda_faturada_kw"] == 0.0:
                out["demanda_faturada_kw"] = _br2f(m.group(1))
                out["demanda_unica_val"]   = _br2f(m.group(3))

        # ── CONSUMO ATIVO PONTA TUSD KWH → kWh e R$ Ponta ───────────────────
        elif "CONSUMO ATIVO PONTA" in upper and "KWH" in upper:
            m = re.search(r"KWH\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln)
            if m:
                out["consumo_ponta_kwh"] = _br2f(m.group(1))
                out["consumo_ponta_val"] = _br2f(m.group(3))

        # ── CONSUMO ATIVO F. PONTA TUSD KWH → kWh e R$ F. Ponta ─────────────
        elif "CONSUMO ATIVO F" in upper and "PONTA" in upper and "KWH" in upper:
            m = re.search(r"KWH\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln)
            if m:
                out["consumo_fponta_kwh"] = _br2f(m.group(1))
                out["consumo_fponta_val"] = _br2f(m.group(3))

        # ── TUSD fora ponta kWh — ENEL CE ────────────────────────────────────
        elif re.search(r"TUSD\s+FORA\s+PONTA\s+KWH", upper):
            m = re.search(r"KWH\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln, re.IGNORECASE)
            if m and out["consumo_fponta_kwh"] == 0.0:
                out["consumo_fponta_kwh"] = _br2f(m.group(1))
                out["consumo_fponta_val"] = _br2f(m.group(3))

        # ── TUSD ponta kWh — ENEL CE ─────────────────────────────────────────
        elif re.search(r"TUSD\s+PONTA\s+KWH", upper) and "FORA" not in upper:
            m = re.search(r"KWH\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln, re.IGNORECASE)
            if m and out["consumo_ponta_kwh"] == 0.0:
                out["consumo_ponta_kwh"] = _br2f(m.group(1))
                out["consumo_ponta_val"] = _br2f(m.group(3))

        # ── Fallback consumo F. Ponta (formato sem "CONSUMO ATIVO") ──────────
        elif ("CONSUMO" in upper or "ENERGIA ATIVA" in upper) and "KWH" in upper \
                and out["consumo_fponta_kwh"] == 0.0 \
                and "GERACAO" not in upper and "ACL" not in upper \
                and "REATIVO" not in upper:
            m = re.search(r"KWH\s+([\d.,]+)(?:\s+[\d.,]+)?(?:\s+([\d.,]+))?", ln)
            if m:
                out["consumo_fponta_kwh"] = _br2f(m.group(1))
                if m.group(2):
                    out["consumo_fponta_val"] = _br2f(m.group(2))

        # ── UFER / Reativo Excedente F. Ponta ────────────────────────────────
        # Anchored on the RIGHT-side billing item to avoid dual-column false positives:
        # left col may have "ENRG RTV UFER PONTA" while right col has ACL/DEDUCAO data
        elif re.search(r"UFER\s+(?:FORA\s+)?PONTA\s+TE\s+KWH", upper):
            m = re.search(r"UFER\s+(?:FORA\s+)?PONTA\s+TE\s+KWH\s+([\d.,]+)\s+([\d.,]+)\s+(-?[\d.,]+)", ln, re.IGNORECASE)
            if m:
                out["ufer_fponta_kvarh"] += _br2f(m.group(1))
                out["ufer_fponta_val"]   += abs(_br2f(m.group(3)))

        # ── Consumo Reativo Excedente Fp kVA — ENEL CE (fatConFPontaCap*) ─────
        elif "REATIVO" in upper and "EXCEDENTE" in upper and "KVA" in upper and " FP " in upper:
            m = re.search(r"KVA\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)", ln, re.IGNORECASE)
            if m:
                out["cap_fponta_kva"] += _br2f(m.group(1))
                out["cap_fponta_val"] += _br2f(m.group(3))

        # ── Benefício tarifário (bruto / líquido) ─────────────────────────────
        # Dual-column: dados do medidor à esquerda + item à direita → buscar a partir de "BENEF"
        # Bruto pode aparecer em duas sub-linhas: "Bruto" e "Bruto Isento Icms" — acumular ambas.
        elif "BENEF" in upper and "TARIF" in upper:
            pos = upper.find("BENEF")
            rm = RE_MONEY.findall(ln[pos:])
            if rm:
                if "BRUTO" in upper:
                    val = abs(_br2f(rm[0]))  # primeiro valor após "BENEF" = valor bruto
                    if val > 0.005:
                        out["beneficio_bruto"] += val  # acumula Bruto + Bruto Isento ICMS
                elif "QUIDO" in upper:  # LÍQUIDO — Í≠I, usar "QUIDO"
                    out["beneficio_liquido"] = _br2f(rm[0])  # primeiro valor = valor líquido

        # ── Escassez hídrica ──────────────────────────────────────────────────
        # ENEL SP: "ENCARGO ESCASSEZ HÍDRICA KWH qty tariff R$"
        # ENEL CE: "Encargo Escassez Hídrica {R$} {pis} {base_icms} ..." (sem unidade kWh)
        # Nota: "HIDRIC" não bate "HÍDRICA" (Í≠I) — usar só "ESCASSEZ"
        elif "ESCASSEZ" in upper:
            m = re.search(r"KWH\s+([\d.,]+)\s+([\d.,]+)\s+(-?[\d.,]+)", ln)
            if m:
                out["escassez_kwh"] += abs(_br2f(m.group(1)))
                out["escassez_val"] += abs(_br2f(m.group(3)))
            else:
                # ENEL CE: sem unidade kWh na linha — primeiro número após "ESCASSEZ" = R$
                pos = upper.find("ESCASSEZ")
                rm_esc = RE_MONEY.findall(ln[pos:])
                if rm_esc:
                    out["escassez_val"] += abs(_br2f(rm_esc[0]))

        # ── Desconto Fio B ────────────────────────────────────────────────────
        elif ("DESCONTO" in upper or "DESC" in upper) and "FIO" in upper:
            if "KWH" in upper:
                m = re.search(r"KWH\s+([\d.,]+)", ln, re.IGNORECASE)
                if m:
                    out["desconto_fio_kwh"] += abs(_br2f(m.group(1)))
            else:
                m = re.search(r"KW\s+([\d.,]+)", ln, re.IGNORECASE)
                if m:
                    out["desconto_fio_kw"] += abs(_br2f(m.group(1)))

        elif ("CIP" in upper or "COSIP" in upper) and not re.search(r"\d{2}/\d{4}", ln):
            # COSIP não tem qty/tariff antes do R$ — right_monies[0] é o valor direto
            pos = upper.find("COSIP") if "COSIP" in upper else upper.find("CIP")
            rm = RE_MONEY.findall(ln[pos:])
            if rm:
                out["cosip"] += _br2f(rm[0])
        elif "RET" in upper and "9430" in upper:
            # Trailing minus: "185,51-" (ENEL SP) ou leading minus: "-185,51"
            m_neg = re.search(r"([\d.,]{4,})-(?=\s|$)", ln)
            if m_neg:
                out["ret_lei_9430_val"] = -abs(_br2f(m_neg.group(1)))
            else:
                for val in monies:
                    if val.startswith("-"):
                        out["ret_lei_9430_val"] = _br2f(val)
                        break
        elif "MULTA" in upper and monies:
            v = _br2f(monies[0])
            if v > 0:
                out["multa"] += v
            else:
                out["outras_cobra"] += abs(v)
        elif ("ATUALIZA" in upper and "MONET" in upper) and monies:
            v = _br2f(monies[0])
            if v > 0:
                out["correcao_monetaria"] += v
            else:
                out["outras_cobra"] += abs(v)
        elif ("JUROS" in upper or "MORA" in upper) and monies:
            v = _br2f(monies[0])
            if v > 0:
                out["juros"] += v
            else:
                out["outras_cobra"] += abs(v)
        elif re.search(r"compensa[cç][aã]o\s+(?:dic|fic)", ln, re.IGNORECASE) and monies:
            out["dic"] += abs(_br2f(monies[-1]))
        elif re.search(r"\b(?:dic|fic)\s+(?:[a-z]{3,9}|\d{2})/\d{4}", ln, re.IGNORECASE) and monies:
            out["dic"] += abs(_br2f(monies[0]))

    out.pop("_em_beneficio", None)
    return out



def _resolver_obs_enel_mt(text: str) -> list:
    """
    Obs específicas ENEL SP MT — soma todas as ocorrências por categoria:

      cod 97  — Dif Fatur TUSD / Encargo Homolog CCEE (positivo)
      cod 286 — Energia ACL (positivo)
      cod 178 — Dedução Energia ACL sem ICMS (negativo, crédito)
    """
    val_dif_tusd   = 0.0
    val_acl        = 0.0
    val_deducao    = 0.0

    linhas = [ln.strip() for ln in text.splitlines() if ln.strip()]

    for ln in linhas:
        upper  = ln.upper()
        # Extrair monies a partir do keyword para ignorar dados da coluna esquerda
        if ("DIF" in upper or "DIFERENC" in upper) and "CCEE" in upper:
            pos = upper.find("DIF")
            rm = RE_MONEY.findall(ln[pos:])
            if len(rm) >= 3:
                val_dif_tusd += abs(_br2f(rm[2]))

        elif "DEDU" in upper and "ENERGIA" in upper and "ACL" in upper:
            pos = upper.find("DEDU")
            rm = RE_MONEY.findall(ln[pos:])
            if len(rm) >= 3:
                val_deducao += abs(_br2f(rm[2]))

        elif "ENERGIA" in upper and "ACL" in upper and "DEDU" not in upper:
            pos = upper.find("ENERGIA")
            rm = RE_MONEY.findall(ln[pos:])
            if len(rm) >= 3:
                val_acl += abs(_br2f(rm[2]))

    obs: list = []
    if abs(val_dif_tusd) > 0.005:
        obs.append(("97", round(val_dif_tusd, 2)))
    if abs(val_acl) > 0.005:
        obs.append(("286", round(val_acl, 2)))
    if abs(val_deducao) > 0.005:
        obs.append(("178", -round(val_deducao, 2)))
    return obs


def extrair_mt(pdf_path: str) -> dict:
    text      = _extract_text(pdf_path)
    datas     = _get_datas_e_total(text, "MT")
    itens     = _itens_mt(text)
    fisco     = _get_fisco(text)
    val_nf    = _get_valor_nf(text)
    subtarifa = _detectar_subtarifa_mt(text)
    instalacao = _get_instalacao_mt(text)
    carimbo_master = _resolver_carimbo_master(Path(pdf_path).name, instalacao, datas["ref"], text)
    # Textos exatos das opções do <select> no sistema Consen
    # Azul → "HS - Azul"  |  Verde → "HS - Verde"
    tarifaCod = "HS - Azul" if subtarifa == "A4_AZUL" else "HS - Verde"

    # ── Demanda: faturada = contratada; registrada = contratada + excedente se houver
    excedente_kw  = itens["demanda_excedente_kw"]
    contratada_kw = itens["demanda_contratada_kw"]
    estadual_kw   = itens["demanda_estadual_kw"]
    if excedente_kw > 0:
        registrada_kw = round(contratada_kw + excedente_kw, 2)
        faturada_kw   = registrada_kw
    else:
        registrada_kw = itens["demanda_faturada_kw"]   # melhor aprox. disponível via OCR
        # Fallback: faturas cobradas pela demanda contratada não têm a linha
        # "DEMANDA ÚNICA C/ DESCONTO". Usa o componente estadual (16.886/18) e,
        # por fim, a contratada — evita registrada zerada no CONSEN.
        if not registrada_kw:
            registrada_kw = estadual_kw or contratada_kw
        # ENEL CE: contratada extraída de "DEMANDA FORA PONTA - KW"; faturada = distribuição
        faturada_kw   = contratada_kw if contratada_kw > 0 else registrada_kw

    dem_val_total = round(
        itens["demanda_unica_val"] + itens["demanda_estadual_val"] + itens["demanda_excedente_val"], 2
    )

    # ── OBS 213 — Outras cobranças/devoluções (DIC, ajustes negativos, etc.)
    outras_total = round(itens["outras_cobra"] + itens["dic"] + itens["fic"], 2)
    obs_213 = [("213", -outras_total)] if outras_total > 0.005 else []

    _ret_mt = {
        "fatCarimbo":              carimbo_master,
        "concCod":                 "3",
        "Instalacao":              instalacao,
        "CNPJ":                    _get_cnpj(text),
        "cadTarifaCod":            tarifaCod,
        "cadSubGrupoCod":          "A4 [<13,8kV]",
        "fatDataReferencia":       datas["ref"],
        "fatDataEmissao":          datas["emissao"],
        "fatDataVcto":             datas["vcto"],
        "fatDataLeituraAnterior":  datas["ant"],
        "fatDataLeituraAtual":     datas["atu"],
        "fatValorFatura":          datas["total"],
        "fatValorNotaFiscal":      val_nf,
        # ── Demanda F. Ponta ─────────────────────────────────────────────────
        "fatDemFPontaIndRegistrada":      registrada_kw,
        "fatDemFPontaIndFaturada":        faturada_kw,
        "fatDemContratadaFPonta":         contratada_kw,
        "fatDemFPontaIndUltra":           excedente_kw,
        "fatDemFPontaIndUltraValorReais": itens["demanda_excedente_val"],
        "fatDemFPontaIndValorReais":      dem_val_total,
        # ── Consumo F. Ponta ─────────────────────────────────────────────────
        "fatConPontaRegistrado":        itens["consumo_ponta_kwh"],
        "fatConPontaFaturado":          itens["consumo_ponta_kwh"],
        "fatConPontaValorReais":        itens["consumo_ponta_val"],
        "fatConFPontaIndRegistrado":    itens["consumo_fponta_kwh"],
        "fatConFPontaIndFaturado":      itens["consumo_fponta_kwh"],
        "fatConFPontaIndValorReais":    itens["consumo_fponta_val"],
        # ── UFER — reativo inducido excedente F. Ponta (ENEL SP) ────────────────
        "fatConFPontaIndExcRegistrado": itens["ufer_fponta_kvarh"],
        "fatConFPontaIndExcFaturado":   itens["ufer_fponta_kvarh"],
        "fatConFPontaIndExcValorReais": itens["ufer_fponta_val"],
        # ── Consumo Reativo Excedente Fp kVA (ENEL CE) → fatConFPontaCap* ────
        "fatConFPontaCapRegistrado": itens["cap_fponta_kva"],
        "fatConFPontaCapFaturado":   itens["cap_fponta_kva"],
        "fatConFPontaCapValorReais": itens["cap_fponta_val"],
        # ── Desconto Fio B — extraído via _itens_mt (ENEL SP MT) ─────────────
        "fatDescontoFio":               itens["desconto_fio_kw"],
        "fatDescontoFioKWh":            itens["desconto_fio_kwh"],
        "fatConCreditoTUSDPontaValorReais":  itens["consumo_ponta_val"],
        "fatConCreditoTUSDFPontaValorReais": itens["consumo_fponta_val"],
        # ── Benefício tarifário ───────────────────────────────────────────────
        "fatBeneficioTarifarioBrutoValorReais": itens["beneficio_bruto"],
        "fatBeneficioLiquidoValorReais":        itens["beneficio_liquido"],
        # ── Escassez hídrica ─────────────────────────────────────────────────
        "fatEscassezHidrica":           itens["escassez_kwh"],
        "fatEscassezHidricaValorReais": itens["escassez_val"],
        # ── Demais ───────────────────────────────────────────────────────────
        "fatIlumPublica":          itens["cosip"],
        "fatValBandeira":          0.0,
        "fatMultas":               round(
            itens["multa"] + itens["juros"] + itens["correcao_monetaria"], 2
        ),
        "fatMultasDiversas":       0.0,
        "fatICMS":                 fisco["icms"],
        "fatDesIcmsAliquota":      fisco["aliq_icms"],
        "fatPIS":                  fisco["pis"],
        "fatDescPisAliquota":      fisco["aliq_pis"],
        "fatCOFINS":               fisco["cofins"],
        "fatDesCofinsAliquota":    fisco["aliq_cofins"],
        "fatTributoFederalPerc":   itens["ret_lei_9430_perc"],
        "fatTributoFederalVal":    itens["ret_lei_9430_val"],
        "fatDataCadastro":         dt.date.today(),
        "ENDERECO":                _get_endereco(text),
        "_obs_list":               _resolver_obs_enel(text, itens) + _resolver_obs_enel_mt(text) + obs_213,
        **{k: 0.0 for k in _RETENÇÕES_ZERO},
    }
    return _ret_mt


# =============================================================================
# DISPATCHER
# =============================================================================

def processar_pdf(pdf_path: str, tipo: str) -> dict:
    filename = Path(pdf_path).name
    try:
        if tipo == "bt":
            dados = extrair_bt(pdf_path)
            # Tarifa Branca tem campos extras além do B3 convencional
            if dados.get("cadTarifaCod") == "Branca":
                dados["TARIFA_DETECTADA"] = "B3_BRANCA"
            else:
                dados["TARIFA_DETECTADA"] = "B3"
        else:
            dados = extrair_mt(pdf_path)
            dados["TARIFA_DETECTADA"] = _detectar_subtarifa_mt(_extract_text(pdf_path))
            # Código de barras também para MT
            if not dados.get("fatCodigoBarras"):
                dados["fatCodigoBarras"] = _get_codigo_barras_enel(_extract_text(pdf_path))

        # Distribuir lista de obs nos pares de colunas
        obs_list = dados.pop("_obs_list", None)
        if obs_list:
            for _i, (_cod, _val) in enumerate(obs_list[:5], start=1):
                dados[f"obsCod_{_i}"]   = _cod
                dados[f"obsValor_{_i}"] = round(float(_val), 2) if _val else 0

        dados["ARQUIVO"] = filename
        dados["ERRO"]    = ""
        log.info(f"    OK  {filename}  ->  {dados['TARIFA_DETECTADA']}  (carimbo {dados['fatCarimbo']})")
        return dados

    except Exception as exc:
        log.error(f"    ERRO  {filename}: {exc}")
        return {
            "fatCarimbo":       _carimbo_do_nome(filename),
            "TARIFA_DETECTADA": "ERRO",
            "ARQUIVO":          filename,
            "ERRO":             str(exc),
        }


# =============================================================================
# EXCEL
# =============================================================================

_FONT_HEADER = Font(name="Arial",   size=11, bold=False)
_FONT_DADOS  = Font(name="Calibri", size=11)
_ALIGN_DIR   = Alignment(horizontal="right")


def _fmt_cell(header: str, value):
    """
    Converte valor para tipo nativo Excel:
    - date/datetime -> objeto date (Excel formata como data)
    - string 'DD/MM/YYYY' (legado) -> objeto date
    - float/int -> numero
    - string PT-BR numerica -> float
    - None em campo fat* -> 0
    """
    if value is None or value == "":
        if header.startswith("fat") or header in ("concCod",):
            return 0
        if header.startswith("obsValor"):
            return 0
        if header.startswith("obsCod"):
            return ""
        return ""

    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    if isinstance(value, (float, int)):
        return value

    if isinstance(value, str):
        v = value.strip()
        # Remove apostrofo de legado
        if v.startswith("'"):
            v = v[1:]
        # Tenta parsear como data DD/MM/YYYY
        if re.match(r"^\d{2}/\d{2}/\d{4}$", v):
            try:
                return dt.datetime.strptime(v, "%d/%m/%Y").date()
            except Exception:
                pass
        # String numerica PT-BR: "1.234,56"
        if re.match(r"^-?[\d.]+,\d{2}$", v):
            return _br2f(v)
        # Inteiro puro
        if re.match(r"^-?\d+$", v):
            try:
                return int(v)
            except Exception:
                pass
        return v

    return value


def _criar_header_excel(ws, headers: list):
    for col, h in enumerate(headers, 1):
        display = _HEADER_DISPLAY.get(h, h)
        cell = ws.cell(row=1, column=col, value=display)
        cell.font = _FONT_HEADER
        ws.column_dimensions[get_column_letter(col)].width = _COL_WIDTHS.get(h, 20.0)
    ws.row_dimensions[1].height = 21.0
    ws.freeze_panes = "A2"


def _carimbos_no_xlsx(xlsx_saida: Path) -> set:
    """Retorna set de fatCarimbo já gravados — evita duplicatas em rodadas consecutivas."""
    if not xlsx_saida.exists():
        return set()
    try:
        wb = load_workbook(xlsx_saida, read_only=True, data_only=True)
        ws = wb.active
        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        try:
            col = header_row.index("fatCarimbo") + 1
        except ValueError:
            wb.close()
            return set()
        carimbos = set()
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
            v = row[0]
            if v is not None and str(v).strip():
                raw = str(v).strip()
                if raw.isdigit():
                    raw = f"BB_{raw}"
                carimbos.add(raw)
        wb.close()
        return carimbos
    except Exception as e:
        log.warning(f"Não consegui ler carimbos existentes em {xlsx_saida.name}: {e}")
        return set()


def salvar_excel(registros: list, caminho: Path):
    # Filtra carimbos já presentes para evitar duplicatas
    ja_presentes = _carimbos_no_xlsx(caminho)
    if ja_presentes:
        antes = len(registros)
        registros = [r for r in registros
                     if str(r.get("fatCarimbo", "")).strip() not in ja_presentes]
        pulados = antes - len(registros)
        if pulados:
            log.info(f"  Dedup: {pulados} registro(s) já existentes ignorados")

    if not registros:
        log.info(f"  Nenhum registro novo para {caminho.name}")
        return

    if caminho.exists():
        wb   = load_workbook(caminho)
        ws   = wb.active
        prox = ws.max_row + 1
        log.info(f"  Atualizando: {caminho.name}  (+{len(registros)} linhas)")
    else:
        wb   = Workbook()
        ws   = wb.active
        ws.title = caminho.stem
        _criar_header_excel(ws, HEADERS_REF)
        prox = 2
        log.info(f"  Criando: {caminho.name}")

    for reg in registros:
        for col, h in enumerate(HEADERS_REF, 1):
            cell = ws.cell(row=prox, column=col,
                           value=_fmt_cell(h, reg.get(h, "")))
            cell.font      = _FONT_DADOS
            cell.alignment = _ALIGN_DIR
        prox += 1

    wb.save(caminho)
    ok   = sum(1 for r in registros if not r.get("ERRO"))
    erro = len(registros) - ok
    log.info(f"  Salvo: {caminho.name}  OK={ok}  ERRO={erro}")


# =============================================================================
# NAVEGACAO DE PASTAS
# =============================================================================

def _pasta_label(pasta: Path) -> str:
    m = re.search(r"(\d{2})[_\-\s]?(\d{4})", pasta.name)
    return f"{m.group(1)}{m.group(2)}" if m else pasta.name


def _listar_pastas_mes(base: Path) -> list:
    padrao = re.compile(r"^(\d{2})[_\-\s]?(\d{4})$")
    return sorted(
        [p for p in base.iterdir() if p.is_dir() and padrao.match(p.name.strip())],
        key=lambda p: p.name,
    )


def _subpasta(pasta_mes: Path, nomes_aceitos: set) -> Optional[Path]:
    for sub in pasta_mes.iterdir():
        if sub.is_dir() and sub.name.strip().lower() in nomes_aceitos:
            return sub
    return None


# =============================================================================
# PROCESSAMENTO DE UM MES
# =============================================================================

def _processar_subpasta(pasta_sub: Path, tipo: str, xlsx_saida: Path):
    pdfs = sorted(pasta_sub.glob("*.pdf"))
    if not pdfs:
        log.warning(f"  Sem PDFs em: {pasta_sub}")
        return

    label = "BT" if tipo == "bt" else "MT"
    log.info(f"  {label}  ->  {pasta_sub.name}  ({len(pdfs)} PDFs)")

    registros = []
    sem_bb_estrito = [
        p for p in pdfs
        if not re.search(r"(?i)\bBB_\d{7}\b", p.stem)
    ]
    if sem_bb_estrito:
        log.info(
            "  Modo sequencial: %d PDF(s) sem BB_ estrito exigem reserva atomica no indice",
            len(sem_bb_estrito),
        )
        for p in pdfs:
            registros.append(processar_pdf(str(p), tipo))
    else:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futures = {ex.submit(processar_pdf, str(p), tipo): p for p in pdfs}
            for f in as_completed(futures):
                registros.append(f.result())

    # Ordena pelo carimbo extraido do nome (numericamente)
    def _sort_key(r):
        try:
            raw = str(r.get("fatCarimbo", "") or "")
            num = raw.replace("BB_", "").replace("bb_", "")
            return int(num) if num.isdigit() else 0
        except (ValueError, TypeError):
            return 0

    registros.sort(key=_sort_key)
    salvar_excel(registros, xlsx_saida)


def processar_mes(pasta_mes: Path, fazer_bt: bool = True, fazer_mt: bool = True):
    label = _pasta_label(pasta_mes)
    log.info(f"\n{'='*60}")
    log.info(f"  {pasta_mes.name}  ->  {label}")
    log.info(f"{'='*60}")

    if fazer_bt:
        sub = _subpasta(pasta_mes, NOMES_BT)
        if sub:
            xlsx = PASTA_SAIDA / f"ocr_enel_BT_{label}.xlsx"
            _processar_subpasta(sub, "bt", xlsx)
        else:
            log.warning(f"  Subpasta BT nao encontrada em: {pasta_mes.name}")

    if fazer_mt:
        sub = _subpasta(pasta_mes, NOMES_MT)
        if sub:
            xlsx = PASTA_SAIDA / f"ocr_enel_MT_{label}.xlsx"
            _processar_subpasta(sub, "mt", xlsx)
        else:
            log.warning(f"  Subpasta MT nao encontrada em: {pasta_mes.name}")


# =============================================================================
# CLI
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(description="OCR ENEL -> planilhas BT e MT")
    p.add_argument("--mes",      type=str)
    p.add_argument("--ano",      type=str)
    p.add_argument("--pasta",    type=str)
    p.add_argument("--todos",    action="store_true")
    p.add_argument("--tipo",     choices=["bt", "mt", "ambos"], default="ambos")
    p.add_argument("--recriar",  action="store_true",
                   help="Apaga o xlsx existente antes de processar (recria do zero)")
    return p.parse_args()


def _resolver_pasta(args) -> Path:
    if args.pasta:
        p = PASTA_DOWNLOAD / args.pasta
        if not p.is_dir():
            log.error(f"Pasta nao encontrada: {p}")
            sys.exit(1)
        return p

    hoje = dt.date.today()
    mes  = args.mes or f"{hoje.month:02d}"
    ano  = args.ano or str(hoje.year)

    for nome in [f"{mes}_{ano}", f"{mes}-{ano}", f"{mes}{ano}", f"{mes} {ano}"]:
        p = PASTA_DOWNLOAD / nome
        if p.is_dir():
            return p

    log.error(f"Pasta {mes}/{ano} nao encontrada em {PASTA_DOWNLOAD}.")
    sys.exit(1)


def main():
    global PASTA_SAIDA
    args = parse_args()
    PASTA_SAIDA = _resolver_pasta_saida()

    fh = logging.FileHandler(PASTA_SAIDA / "ocr_enel.log", encoding="utf-8", errors="replace")
    fh.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s",
                                      datefmt="%Y-%m-%d %H:%M:%S"))
    log.addHandler(fh)

    fazer_bt = args.tipo in ("bt", "ambos")
    fazer_mt = args.tipo in ("mt", "ambos")

    # --recriar: apaga xlsx existente para forçar reprocessamento completo
    if getattr(args, "recriar", False):
        hoje = dt.date.today()
        mes = args.mes or f"{hoje.month:02d}"
        ano = args.ano or str(hoje.year)
        label = f"{mes}{ano}"
        for xlsx in [PASTA_SAIDA / f"ocr_enel_BT_{label}.xlsx",
                     PASTA_SAIDA / f"ocr_enel_MT_{label}.xlsx"]:
            if xlsx.exists():
                xlsx.unlink()
                log.info(f"  [recriar] Removido: {xlsx.name}")

    log.info("=" * 60)
    log.info("  OCR ENEL  -  BT + MT".center(60))
    log.info("=" * 60)
    log.info(f"  Tipo     : {args.tipo.upper()}")
    log.info(f"  Modo     : {'recriar do zero' if getattr(args, 'recriar', False) else 'incremental (dedup)'}")
    log.info(f"  Download : {PASTA_DOWNLOAD}")
    log.info(f"  Saida    : {PASTA_SAIDA}")

    # Padrao: todos os meses se nenhum filtro passado
    modo_todos = args.todos or (not args.mes and not args.ano and not args.pasta)

    if modo_todos:
        pastas = _listar_pastas_mes(PASTA_DOWNLOAD)
        if not pastas:
            log.error(f"Nenhuma pasta de mes encontrada em: {PASTA_DOWNLOAD}")
            sys.exit(1)
        log.info(f"  Meses    : {len(pastas)}")
        for pasta in pastas:
            processar_mes(pasta, fazer_bt, fazer_mt)
    else:
        pasta = _resolver_pasta(args)
        processar_mes(pasta, fazer_bt, fazer_mt)

    log.info("\nConcluido.")


if __name__ == "__main__":
    main()
