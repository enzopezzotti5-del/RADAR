#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline_enel_mt_lote.py
------------------------
Pipeline completo para lote avulso ENEL MT:
  0. (Opcional) Staging — carimba e copia PDFs para DOWNLOAD ENEL
  1. OCR          — extrai dados dos PDFs do lote
  2. Digitação    — digita no Consen via Selenium
  3. Filtro       — move PDFs digitados para Digitadas

O lote fica em:
    DOWNLOAD ENEL / {lote} / MT / BB_XXXXXXX.pdf

O xlsx de OCR fica em:
    OCR ENEL / ocr_enel_MT_{mes}{ano}.xlsx

Uso:
    python pipeline_enel_mt_lote.py --mes 06 --ano 2026
    python pipeline_enel_mt_lote.py --mes 06 --ano 2026 --so-ocr
    python pipeline_enel_mt_lote.py --mes 06 --ano 2026 --so-digitacao
    python pipeline_enel_mt_lote.py --mes 06 --ano 2026 --so-filtro
    python pipeline_enel_mt_lote.py --mes 06 --ano 2026 --pular-staging
"""
from __future__ import annotations

import argparse
import os
import subprocess
import sys
from pathlib import Path

_RAIZ = Path(__file__).resolve().parent.parent.parent
_CORE = _RAIZ / "core"
sys.path.insert(0, str(_RAIZ))
sys.path.insert(0, str(_CORE))

try:
    from pipelines._visual_safe import _p, _info, _ok, _fail, _warn, _sep, _banner, _rodar as _rodar_visual
except ModuleNotFoundError:
    from _visual_safe import _p, _info, _ok, _fail, _warn, _sep, _banner, _rodar as _rodar_visual

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

SERVIDOR       = Path("//10.10.250.21/Energia")
DOWNLOAD_ENEL  = SERVIDOR / "ARQUIVOS ENZO" / "DOWNLOAD ENEL"
OCR_SAIDA_DIR  = SERVIDOR / "ARQUIVOS ENZO" / "OCR ENEL"
PIPELINE_SAIDA = SERVIDOR / "ARQUIVOS ENZO" / "ENEL_pipeline_saida"
DIGITADAS_DIR  = SERVIDOR / "CONTASDEENERGIAELETRICA" / "BB" / "ENZO" / "Digitadas"

LOTE_NOME_PADRAO = "lote_enel_mt_062026"

STAGING_SCRIPT   = _CORE / "pipelines" / "staging_enel_mt_lote.py"
OCR_SCRIPT       = _CORE / "ocr"       / "ocr_enel.py"
DIGITACAO_SCRIPT = _CORE / "digitacao_consen" / "digitacao_consen_enel.py"
FILTRO_SCRIPT    = _CORE / "digitacao_consen" / "enel_filtro.py"

PYTHON_EXE = str(Path(sys.executable).parent / "python.exe")

FALLBACK_ROOT = _RAIZ / "_runtime_fallback" / "pipeline_enel_mt_lote"


# =============================================================================
# HELPERS
# =============================================================================

def _rodar(descricao: str, cmd: list[str], env_extra: dict | None = None) -> int:
    return _rodar_visual(descricao, cmd, env_extra=env_extra)


def _resolver_dir(preferido: Path, fallback: Path, rotulo: str) -> Path:
    try:
        preferido.mkdir(parents=True, exist_ok=True)
        return preferido
    except OSError as e:
        fallback.mkdir(parents=True, exist_ok=True)
        _warn(f"{rotulo} indisponível: {e}. Fallback: {fallback}")
        return fallback


def _resetar_auditoria(pasta_saida: Path) -> None:
    arq = pasta_saida / "auditoria_resultados.csv"
    if arq.exists():
        arq.unlink()
        _info("[reset] auditoria_resultados.csv removido antes da digitação.")


def _atualizar_master_pos_filtro(auditoria_csv: Path) -> None:
    try:
        from indice_master import MasterIndice, marcar_digitados_do_auditoria
        contadores = marcar_digitados_do_auditoria(auditoria_csv, MasterIndice())
        _info(f"[MASTER] Digitação atualizada: {contadores}")
    except Exception as e:
        _warn(f"[MASTER] Não foi possível atualizar o master: {e}")


# =============================================================================
# ETAPAS
# =============================================================================

def etapa_staging(lote: str, fonte: Path | None = None) -> int:
    cmd = [PYTHON_EXE, str(STAGING_SCRIPT), "--lote", lote]
    if fonte is not None:
        cmd += ["--fonte", str(fonte)]
    return _rodar(f"STAGING ENEL MT  ({lote})", cmd)


def etapa_ocr(mes: str, ano: str, lote: str, recriar: bool) -> int:
    cmd = [
        PYTHON_EXE, str(OCR_SCRIPT),
        "--pasta", lote,
        "--tipo",  "mt",
        "--mes",   mes,
        "--ano",   ano,
    ]
    if recriar:
        cmd.append("--recriar")
    env_extra = {
        "OCR_ENEL_DOWNLOAD_DIR": str(DOWNLOAD_ENEL),
        "OCR_ENEL_SAIDA_DIR":    str(OCR_SAIDA_DIR),
    }
    return _rodar(f"OCR ENEL MT  ({lote}  {mes}/{ano})", cmd, env_extra)


def etapa_digitacao(
    xlsx: Path,
    pasta_pdfs: Path,
    pipeline_saida: Path,
    preservar_auditoria: bool,
) -> int:
    if not xlsx.exists():
        _fail(f"Planilha não encontrada: {xlsx}")
        return 1

    if not preservar_auditoria:
        _resetar_auditoria(pipeline_saida)

    env_extra = {
        "ENEL_EXCEL_PATH":               str(xlsx),
        "ENEL_DIGITACAO_PASTA_PDFS":     str(pasta_pdfs),
        "CONSEN_PIPELINE_SAIDA":         str(pipeline_saida),
        "CONSEN_INTERATIVO_FECHAR":      "0",
        "CONSEN_INVESTIGAR_ZEROS":       "0",
        "DIGITACAO_FATOR_VELOCIDADE":    "0.25",
        "CONSEN_SENHA":                  "Acao2026",
        "ENEL_AJUSTAR_LEITURA_ULTIMO_DIA": "1",
    }
    return _rodar(f"DIGITAÇÃO ENEL MT  ({pasta_pdfs.parent.name})", [PYTHON_EXE, str(DIGITACAO_SCRIPT)], env_extra)


def etapa_validar_ocr(xlsx: Path, pasta_pdfs: Path) -> None:
    """Valida as linhas do XLSX com Pydantic antes de enviar ao CONSEN.

    Não bloqueia a digitação — registra avisos para campos suspeitos.
    Filtra apenas as linhas cujo PDF existe na pasta (mesmo critério da digitação).
    """
    try:
        import openpyxl
        from schemas.fatura import validar_linha_ocr
    except ImportError:
        _warn("[VALIDAÇÃO] openpyxl ou schemas não disponíveis — pulando")
        return

    if not xlsx.exists():
        return

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    pdfs_presentes = {p.stem for p in pasta_pdfs.glob("*.pdf")} if pasta_pdfs.exists() else set()

    arquivo_idx  = next((i for i, h in enumerate(headers) if h == "ARQUIVO"), None)
    carimbo_idx  = next((i for i, h in enumerate(headers) if h == "fatCarimbo"), None)

    total = validos = invalidos = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        arquivo = str(row[arquivo_idx] or "") if arquivo_idx is not None else ""
        stem    = Path(arquivo).stem if arquivo else ""
        if stem and stem not in pdfs_presentes:
            continue  # não está no lote atual

        total += 1
        d = dict(zip(headers, row))
        resultado = validar_linha_ocr(d, tipo="MT")

        if resultado.avisos:
            for av in resultado.avisos:
                _warn(f"[VALIDAÇÃO] {resultado.carimbo} ({resultado.instalacao}): {av}")

        if not resultado.valido:
            invalidos += 1
            for err in resultado.erros:
                _warn(f"[VALIDAÇÃO] {resultado.carimbo} ({resultado.instalacao}): {err}")
        else:
            validos += 1

    _info(f"[VALIDAÇÃO OCR] {total} linhas  |  {validos} OK  |  {invalidos} com problema(s)")


def etapa_filtro(pasta_pdfs: Path, pipeline_saida: Path) -> int:
    auditoria_csv = pipeline_saida / "auditoria_resultados.csv"
    if not auditoria_csv.exists():
        _warn(f"auditoria_resultados.csv não encontrado — filtro pulado: {auditoria_csv}")
        return 0
    investigar_dir = SERVIDOR / "CONTASDEENERGIAELETRICA" / "BB" / "ENZO" / "Investigar"
    env_extra = {
        "ENEL_FILTRO_CSV":        str(auditoria_csv),
        "ENEL_FILTRO_PDFS":       str(pasta_pdfs),
        "ENEL_FILTRO_DESTINO":    str(DIGITADAS_DIR),
        "ENEL_FILTRO_JA_EXISTIAM": str(investigar_dir),
    }
    codigo = _rodar(f"FILTRO ENEL MT  ({pasta_pdfs.parent.name})", [PYTHON_EXE, str(FILTRO_SCRIPT)], env_extra)
    if codigo == 0:
        _atualizar_master_pos_filtro(auditoria_csv)
    else:
        _warn("[MASTER] Atualização do master pulada — filtro falhou.")
    return codigo


# =============================================================================
# MAIN
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(description="Pipeline ENEL MT — lote avulso")
    p.add_argument("--mes",   default="06",   help="Mês (ex: 06)")
    p.add_argument("--ano",   default="2026", help="Ano (ex: 2026)")
    p.add_argument("--lote",  default="",     help="Nome do lote (subpasta em DOWNLOAD ENEL); padrão: lote_enel_mt_{mes}{ano}")
    p.add_argument("--pasta", default="",     help="Pasta fonte dos PDFs originais (watcher passa aqui)")
    p.add_argument("--pular-staging",  action="store_true", help="Não roda staging (PDFs já foram copiados)")
    p.add_argument("--so-staging",     action="store_true")
    p.add_argument("--so-ocr",         action="store_true")
    p.add_argument("--so-digitacao",   action="store_true")
    p.add_argument("--so-filtro",      action="store_true")
    p.add_argument("--recriar",        action="store_true", help="Apaga xlsx do OCR antes de reprocessar")
    p.add_argument("--preservar-auditoria", action="store_true",
                   help="Não apaga auditoria_resultados.csv antes da digitação (retomar lote interrompido)")
    return p.parse_args()


def main():
    args = parse_args()

    mes   = args.mes
    ano   = args.ano
    lote  = args.lote or f"lote_enel_mt_{mes}{ano}"
    pasta = Path(args.pasta) if args.pasta else None

    tudo          = not any([args.so_staging, args.so_ocr, args.so_digitacao, args.so_filtro])
    fazer_staging = (tudo or args.so_staging) and not args.pular_staging
    fazer_ocr     = tudo or args.so_ocr
    fazer_dig     = tudo or args.so_digitacao
    fazer_filtro  = tudo or args.so_filtro

    # Caminhos resolvidos
    ocr_saida    = _resolver_dir(OCR_SAIDA_DIR,  FALLBACK_ROOT / "ocr_saida",  "Saída OCR")
    pip_saida    = _resolver_dir(PIPELINE_SAIDA, FALLBACK_ROOT / "saida",      "Saída pipeline")
    dig_destino  = _resolver_dir(DIGITADAS_DIR,  FALLBACK_ROOT / "digitadas",  "Digitadas")

    xlsx      = ocr_saida / f"ocr_enel_MT_{mes}{ano}.xlsx"
    pasta_mt  = DOWNLOAD_ENEL / lote / "MT"

    _banner(f"PIPELINE ENEL MT LOTE  {lote}  {mes}/{ano}", [
        f"PDFs lote : {pasta_mt}",
        f"xlsx OCR  : {xlsx}",
        f"Saída     : {pip_saida}",
        f"Digitadas : {dig_destino}",
    ])

    falhou = False

    # 0. Staging
    if fazer_staging:
        cod = etapa_staging(lote, fonte=pasta)
        if cod != 0:
            _fail("Staging falhou — abortando.")
            sys.exit(1)

    # Verifica que a pasta de lote existe após staging
    if fazer_ocr or fazer_dig or fazer_filtro:
        if not pasta_mt.exists():
            _fail(f"Pasta do lote não encontrada: {pasta_mt}")
            _warn("Rode staging primeiro ou passe --pular-staging se a pasta já existe.")
            sys.exit(1)

    # 1. OCR
    if fazer_ocr:
        cod = etapa_ocr(mes, ano, lote, recriar=args.recriar)
        if cod != 0:
            _fail("OCR falhou — abortando.")
            sys.exit(1)

    # 1.5 Validação Pydantic (não bloqueia — registra avisos)
    if fazer_ocr or fazer_dig:
        etapa_validar_ocr(xlsx, pasta_mt)

    # 2. Digitação
    if fazer_dig:
        preservar = args.preservar_auditoria or os.environ.get(
            "PIPELINE_PRESERVAR_AUDITORIA", ""
        ).strip().lower() in {"1", "true", "sim", "s"}
        cod = etapa_digitacao(xlsx, pasta_mt, pip_saida, preservar)
        if cod != 0:
            _warn(f"Digitação terminou com exit {cod} — continuando para filtro.")

    # 3. Filtro
    if fazer_filtro:
        cod = etapa_filtro(pasta_mt, pip_saida)
        if cod != 0:
            _fail("Filtro falhou.")
            falhou = True

    _p()
    _sep("-")
    if falhou:
        _fail("PIPELINE ENEL MT LOTE FINALIZADO COM FALHAS")
    else:
        _ok("PIPELINE ENEL MT LOTE FINALIZADO COM SUCESSO")
    _sep("-")
    sys.exit(1 if falhou else 0)


if __name__ == "__main__":
    main()
