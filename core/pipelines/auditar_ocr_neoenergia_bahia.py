#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Auditoria automatica do OCR Neoenergia Bahia.

Uso:
    python auditar_ocr_neoenergia_bahia.py --mes 03 --ano 2026 --tipo bt
    python auditar_ocr_neoenergia_bahia.py --xlsx "\\\\...\\ocr_neoenergia_bahia_BT_032026.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import logging
import math
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SERVIDOR = Path("//10.10.250.21/Energia")
OCR_DIR = SERVIDOR / "ARQUIVOS ENZO" / "OCR NEOENERGIA" / "BAHIA"
AUDITORIA_DIR = OCR_DIR / "auditoria"

CRITICOS = [
    "Instalacao",
    "CODIGOCLIENTE",
    "fatDataEmissao",
    "fatDataVcto",
    "fatValorFatura",
    "CNPJ",
    "fatCodigoBarras",
    "fatCarimbo",
]

COLUNAS_CSV = [
    "ARQUIVO",
    "fatCarimbo",
    "Instalacao",
    "CODIGOCLIENTE",
    "fatDataEmissao",
    "fatDataVcto",
    "fatValorFatura",
    "CNPJ",
    "fatCodigoBarras",
    "fatICMS",
    "fatPIS",
    "fatCOFINS",
    "fatTributoFederalPerc",
    "fatTributoFederalVal",
    "ERRO",
    "PROBLEMAS",
]


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("auditar_ocr_neoenergia_bahia")


def _mkdir_seguro(pasta: Path) -> None:
    try:
        pasta.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass


def _xlsx_default(mes: str, ano: str, tipo: str) -> Path:
    return OCR_DIR / f"ocr_neoenergia_bahia_{tipo.upper()}_{mes}{ano}.xlsx"


def _csv_saida(mes: str, ano: str, tipo: str) -> Path:
    return AUDITORIA_DIR / f"auditoria_ocr_neoenergia_bahia_{tipo.upper()}_{mes}{ano}.csv"


def _txt_saida(mes: str, ano: str, tipo: str) -> Path:
    return AUDITORIA_DIR / f"auditoria_ocr_neoenergia_bahia_{tipo.upper()}_{mes}{ano}.txt"


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def _digits(value: Any) -> str:
    return "".join(ch for ch in _as_text(value) if ch.isdigit())


def _as_date(value: Any) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    txt = _as_text(value)
    if not txt:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def _as_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return 0.0
    txt = _as_text(value)
    if not txt:
        return 0.0
    txt = txt.replace("R$", "").replace(" ", "")
    neg = txt.startswith("-")
    txt = txt.replace("-", "").replace(".", "").replace(",", ".")
    try:
        val = float(txt)
        return -val if neg else val
    except Exception:
        return 0.0


def _ler_planilha(xlsx: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(v).strip() if v is not None else "" for v in next(rows)]
        saida: list[dict[str, Any]] = []
        for row in rows:
            rec = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
            if not any(v is not None and _as_text(v) for v in rec.values()):
                continue
            saida.append(rec)
        return saida
    finally:
        wb.close()


def _avaliar_registro(rec: dict[str, Any]) -> list[str]:
    probs: list[str] = []

    erro = _as_text(rec.get("ERRO"))
    instalacao = _digits(rec.get("Instalacao"))
    cliente = _digits(rec.get("CODIGOCLIENTE"))
    emissao = _as_date(rec.get("fatDataEmissao"))
    vcto = _as_date(rec.get("fatDataVcto"))
    valor = _as_float(rec.get("fatValorFatura"))
    cnpj = _digits(rec.get("CNPJ"))
    barras = _digits(rec.get("fatCodigoBarras"))
    carimbo = _digits(rec.get("fatCarimbo"))
    icms = _as_float(rec.get("fatICMS"))
    pis = _as_float(rec.get("fatPIS"))
    cofins = _as_float(rec.get("fatCOFINS"))
    trib_fed_perc = _as_float(rec.get("fatTributoFederalPerc"))
    trib_fed_val = _as_float(rec.get("fatTributoFederalVal"))

    if erro:
        probs.append("OCR_ERRO")

    if not instalacao:
        probs.append("SEM_INSTALACAO")
    elif not (6 <= len(instalacao) <= 12):
        probs.append("INSTALACAO_FORMATO")

    if not cliente:
        probs.append("SEM_CODIGOCLIENTE")
    elif not (6 <= len(cliente) <= 12):
        probs.append("CODIGOCLIENTE_FORMATO")

    if not emissao:
        probs.append("SEM_EMISSAO")

    if not vcto:
        probs.append("SEM_VENCIMENTO")
    elif emissao and vcto < emissao:
        probs.append("VENCIMENTO_ANTES_EMISSAO")

    if valor <= 0:
        probs.append("VALOR_FATURA_INVALIDO")

    if not cnpj:
        probs.append("SEM_CNPJ")
    elif len(cnpj) != 14:
        probs.append("CNPJ_INVALIDO")

    if not barras:
        probs.append("SEM_CODIGO_BARRAS")
    elif len(barras) < 44:
        probs.append("CODIGO_BARRAS_CURTO")

    if not carimbo:
        probs.append("SEM_CARIMBO")
    elif len(carimbo) != 7:
        probs.append("CARIMBO_FORMATO")

    if icms <= 0:
        probs.append("ICMS_ZERADO")
    if pis <= 0:
        probs.append("PIS_ZERADO")
    if cofins <= 0:
        probs.append("COFINS_ZERADO")
    if trib_fed_perc <= 0:
        probs.append("TRIB_FEDERAL_PERC_ZERADO")
    if trib_fed_val <= 0:
        probs.append("TRIB_FEDERAL_VAL_ZERADO")

    return probs


def _normalizar_csv(rec: dict[str, Any], problemas: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for col in COLUNAS_CSV:
        if col == "PROBLEMAS":
            out[col] = " | ".join(problemas)
        else:
            out[col] = _as_text(rec.get(col))
    return out


def _gravar_csv(path: Path, linhas: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLUNAS_CSV, delimiter=";")
        w.writeheader()
        w.writerows(linhas)


def _gravar_txt(
    path: Path,
    xlsx: Path,
    tipo: str,
    total: int,
    suspeitas: int,
    counters: Counter,
    faltas_criticas: Counter,
) -> None:
    linhas = [
        "=" * 72,
        "AUDITORIA OCR NEOENERGIA BAHIA",
        "=" * 72,
        f"Planilha : {xlsx}",
        f"Tipo     : {tipo.upper()}",
        f"Gerado   : {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        f"Total de linhas        : {total}",
        f"Linhas suspeitas       : {suspeitas}",
        f"Linhas sem suspeita    : {max(total - suspeitas, 0)}",
        "",
        "Faltas em campos criticos:",
    ]

    for campo in CRITICOS:
        linhas.append(f"  - {campo}: {faltas_criticas.get(campo, 0)}")

    linhas.append("")
    linhas.append("Ocorrencias por regra:")
    for regra, qtd in counters.most_common():
        linhas.append(f"  - {regra}: {qtd}")

    path.write_text("\n".join(linhas), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    hoje = dt.date.today()
    p = argparse.ArgumentParser(description="Auditoria do OCR Neoenergia Bahia")
    p.add_argument("--mes", type=str, default=f"{hoje.month:02d}")
    p.add_argument("--ano", type=str, default=str(hoje.year))
    p.add_argument("--tipo", choices=["bt", "mt"], default="bt")
    p.add_argument("--xlsx", type=str, default="", help="Caminho manual da planilha OCR")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    mes = f"{int(args.mes):02d}"
    ano = str(int(args.ano))
    tipo = args.tipo.lower()
    xlsx = Path(args.xlsx) if str(args.xlsx).strip() else _xlsx_default(mes, ano, tipo)

    if not xlsx.exists():
        log.error(f"Planilha nao encontrada: {xlsx}")
        return 1

    _mkdir_seguro(AUDITORIA_DIR)

    log.info("=" * 60)
    log.info("AUDITORIA OCR NEOENERGIA BAHIA")
    log.info("=" * 60)
    log.info(f"Planilha : {xlsx}")
    log.info(f"Tipo     : {tipo.upper()}")

    registros = _ler_planilha(xlsx)
    if not registros:
        log.warning("Nenhuma linha encontrada na planilha.")
        return 0

    suspeitas_csv: list[dict[str, str]] = []
    counters: Counter = Counter()
    faltas_criticas: Counter = Counter()

    for rec in registros:
        probs = _avaliar_registro(rec)
        if probs:
            suspeitas_csv.append(_normalizar_csv(rec, probs))
        for p in probs:
            counters[p] += 1

        for campo in CRITICOS:
            valor = rec.get(campo)
            if campo in {"fatDataEmissao", "fatDataVcto"}:
                vazio = _as_date(valor) is None
            elif campo == "fatValorFatura":
                vazio = _as_float(valor) <= 0
            else:
                vazio = not _as_text(valor)
            if vazio:
                faltas_criticas[campo] += 1

    csv_saida = _csv_saida(mes, ano, tipo)
    txt_saida = _txt_saida(mes, ano, tipo)
    _gravar_csv(csv_saida, suspeitas_csv)
    _gravar_txt(txt_saida, xlsx, tipo, len(registros), len(suspeitas_csv), counters, faltas_criticas)

    log.info(f"Total linhas      : {len(registros)}")
    log.info(f"Linhas suspeitas  : {len(suspeitas_csv)}")
    for campo in CRITICOS:
        log.info(f"Falta {campo:<16}: {faltas_criticas.get(campo, 0)}")
    if counters:
        log.info("Ocorrencias por regra:")
        for regra, qtd in counters.most_common():
            log.info(f"  - {regra}: {qtd}")

    log.info(f"CSV auditoria : {csv_saida}")
    log.info(f"Resumo TXT    : {txt_saida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
