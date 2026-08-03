#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
staging_celesc_mt.py
--------------------
Carimba PDFs CELESC MT vindos da pasta Faturas e os copia para
DOWNLOAD CELESC/MM.YYYY/MT/ prontos para OCR/digitação/filtro.

Uso standalone:
    python staging_celesc_mt.py --pasta "\\\\srv\\Faturas\\CELESC"
    python staging_celesc_mt.py --pasta "..." --dry-run
"""
from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
import tempfile
import re
from pathlib import Path

_RAIZ = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_RAIZ))
sys.path.insert(0, str(_RAIZ / "core"))
import _venv_check  # noqa

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

from indice_master import MasterIndice

DOWNLOAD_CELESC = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/DOWNLOAD CELESC")
OCR_SCRIPT = _RAIZ / "core" / "ocr" / "ocr_celesc_mt.py"
PYTHON_EXE = str(Path(sys.executable))


def _rodar_ocr(pasta: Path, mes: str, ano: str) -> dict[str, dict]:
    """Roda ocr_celesc.py na pasta e devolve {stem_arquivo: rec}."""
    if openpyxl is None:
        print("[ERRO] openpyxl não instalado")
        return {}

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_tmp = Path(tmp.name)

    cmd = [
        PYTHON_EXE, str(OCR_SCRIPT),
        "--pasta", str(pasta),
        "--mes", mes,
        "--ano", ano,
        "--saida", str(xlsx_tmp),
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        print(f"[ERRO] OCR falhou:\n{res.stderr[-2000:]}")
        xlsx_tmp.unlink(missing_ok=True)
        return {}

    wb = openpyxl.load_workbook(str(xlsx_tmp), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    recs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        arq = str(d.get("ARQUIVO", "") or "")
        stem = Path(arq).stem if arq else ""
        if stem:
            recs[stem] = d
    xlsx_tmp.unlink(missing_ok=True)
    return recs


def _texto_pdf(path: Path) -> str:
    try:
        import pdfplumber
    except ImportError:
        return ""
    try:
        partes: list[str] = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages[:2]:
                partes.append(page.extract_text() or "")
        return "\n".join(partes)
    except Exception:
        return ""


def _parece_celesc_bt(texto: str) -> bool:
    txt = (texto or "").upper()
    return bool(
        re.search(r"GRUPO/SUBGRUPO\s+TENS[AÃ]O\s*:\s*B\s*/\s*B[1-4]\b", txt)
        or re.search(r"\bGRUPO\s+B\b", txt)
        or re.search(r"\bSUBGRUPO\s+B[1-4]\b", txt)
        or re.search(r"\bBAIXA\s*TENS[AÃ]O\b", txt)
    )


def _mt_confirmado(texto: str, rec: dict) -> bool:
    txt = (texto or "").upper()
    subgrupo = str(rec.get("SUBGRUPO_DETECTADO") or rec.get("cadSubGrupoCod") or "").upper()
    tarifa = str(rec.get("TARIFA_DETECTADA") or rec.get("cadTarifaCod") or "").upper()
    return bool(
        subgrupo.startswith("A")
        or re.search(r"GRUPO/SUBGRUPO\s+TENS[AÃ]O\s*:\s*A\s*/\s*A", txt)
        or re.search(r"TENS[AÃ]O\s+FORNECIMENTO[^\n]{0,20}\b(13[,.]8|23|34[,.]5)\b", txt)
        or "VERDE" in tarifa
        or "AZUL" in tarifa
    )


def main() -> int:
    hoje_import = __import__("datetime").date.today()
    ap = argparse.ArgumentParser(description="Staging CELESC MT — carimbagem e cópia para DOWNLOAD CELESC")
    ap.add_argument("--pasta",   required=True, help="Pasta com os PDFs originais (ex: Faturas/CELESC/_processado)")
    ap.add_argument("--mes",     default=f"{hoje_import.month:02d}", help="Mês padrão para OCR (MM)")
    ap.add_argument("--ano",     default=str(hoje_import.year),      help="Ano padrão para OCR (YYYY)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    pasta = Path(args.pasta)
    if not pasta.exists():
        print(f"[ERRO] Pasta não encontrada: {pasta}")
        return 1

    # Processa só PDFs originais (não BB_*.pdf que já foram carimbados)
    pdfs = sorted(p for p in pasta.glob("*.pdf") if p.is_file() and not p.name.startswith("BB_"))
    if not pdfs:
        print(f"Nenhum PDF novo em: {pasta}")
        return 0

    print(f"\n{'[DRY-RUN] ' if args.dry_run else ''}STAGING CELESC MT")
    print(f"  Pasta   : {pasta}")
    print(f"  PDFs    : {len(pdfs)}")

    print(f"  Rodando OCR...")
    recs = _rodar_ocr(pasta, args.mes, args.ano)
    if not recs:
        print("[ERRO] OCR não retornou resultados")
        return 1

    master = MasterIndice()
    copiados = pulados = erros = 0

    for pdf in pdfs:
        rec = recs.get(pdf.stem)
        if not rec:
            print(f"  [SEM OCR] {pdf.name}")
            erros += 1
            continue

        texto = _texto_pdf(pdf)
        if _parece_celesc_bt(texto):
            print(f"  [BLOQUEADO BT] {pdf.name}: texto indica Grupo/Subgrupo B; nao enviar ao staging MT")
            erros += 1
            continue
        if not _mt_confirmado(texto, rec):
            print(f"  [MT NAO CONFIRMADO] {pdf.name}: sem subgrupo/tensao/modalidade MT explicita")
            erros += 1
            continue

        uc = str(rec.get("Instalacao") or "").strip()
        if not uc:
            print(f"  [SEM UC] {pdf.name}")
            erros += 1
            continue

        ref_raw = rec.get("fatDataReferencia")
        import datetime as dt
        if hasattr(ref_raw, "month"):
            mes_ref = f"{ref_raw.month:02d}-{ref_raw.year}"
        elif isinstance(ref_raw, str) and ref_raw:
            try:
                d = dt.date.fromisoformat(ref_raw[:10])
                mes_ref = f"{d.month:02d}-{d.year}"
            except Exception:
                mes_ref = f"{args.mes}-{args.ano}"
        else:
            mes_ref = f"{args.mes}-{args.ano}"

        cnpj = str(rec.get("CNPJ") or "").strip()

        if master.ja_foi_baixado(uc, mes_ref, "CELESC"):
            print(f"  [JÁ NO MASTER] UC={uc}  ref={mes_ref}  ({pdf.name})")
            pulados += 1
            continue

        if args.dry_run:
            print(f"  [DRY] {pdf.name:40s} UC={uc}  ref={mes_ref}")
            continue

        carimbo = master.consumir_carimbo()
        mes_n, ano_n = mes_ref.split("-")
        pasta_dest = DOWNLOAD_CELESC / f"{mes_n}.{ano_n}" / "MT"
        destino = pasta_dest / f"{carimbo}.pdf"

        try:
            pasta_dest.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(pdf), str(destino))
        except OSError as e:
            print(f"  [ERRO CÓPIA] {pdf.name}: {e}")
            erros += 1
            continue

        master.registrar(
            indice_bb=carimbo,
            sistema="CELESC",
            uc=uc,
            mes_ref=mes_ref,
            cnpj=cnpj,
            estado="SANTA CATARINA",
            arquivo=pdf.name,
        )
        print(f"  [OK] {pdf.name:40s} -> {destino.name}  UC={uc}  {mes_ref}")
        copiados += 1

    print(f"\n  Copiados : {copiados}")
    print(f"  Pulados  : {pulados}")
    print(f"  Erros    : {erros}")
    return 0 if erros == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
