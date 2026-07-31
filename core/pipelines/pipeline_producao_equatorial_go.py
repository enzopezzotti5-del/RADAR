#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Pipeline de producao Equatorial GO
=================================

Processa os PDFs brutos que chegam na pasta:
    \\10.10.250.21\Energia\CONTASDEENERGIAELETRICA\BB\ENZO

Fluxo:
    1. Varre apenas os PDFs da raiz da pasta de producao
    2. Identifica os PDFs da Equatorial GO e separa BT/MT
    3. Carimba todos os PDFs da Equatorial GO na propria origem e registra no master
    4. Copia para staging BT/MT
    5. Executa OCR separado:
         - BT -> XLSX BT
         - MT -> XLSX MT
    6. Digita somente BT no Consen
    7. Filtra somente BT, movendo para Digitadas apenas os carimbos aprovados
       pela auditoria; PDFs nao digitados ficam na pasta raiz

Uso:
    .venv\\Scripts\\python.exe pipelines\\pipeline_producao_equatorial_go.py
    .venv\\Scripts\\python.exe pipelines\\pipeline_producao_equatorial_go.py --so-carimbo
    .venv\\Scripts\\python.exe pipelines\\pipeline_producao_equatorial_go.py --so-ocr
    .venv\\Scripts\\python.exe pipelines\\pipeline_producao_equatorial_go.py --so-digitacao
    .venv\\Scripts\\python.exe pipelines\\pipeline_producao_equatorial_go.py --so-filtro
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import logging
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook

LOCAL_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(LOCAL_DIR))
sys.path.insert(0, str(LOCAL_DIR.parent))          # ENERGIA/ root
sys.path.insert(0, str(LOCAL_DIR.parent / "scripts"))  # ENERGIA/scripts — onde indice_master.py está

from indice_master import MASTER_FIELDS, MasterIndice, marcar_digitados_do_auditoria
from ocr import ocr_equatorial_go as ocr_eq


SERVIDOR = Path("//10.10.250.21/Energia")
PASTA_ENZO = SERVIDOR / "CONTASDEENERGIAELETRICA" / "BB" / "ENZO"
DIGITADAS_DIR = PASTA_ENZO / "Digitadas"

OCR_SAIDA_DIR = SERVIDOR / "ARQUIVOS ENZO" / "OCR PRODUCAO EQUATORIAL GO"
PIPELINE_SAIDA_DIR = SERVIDOR / "ARQUIVOS ENZO" / "EQUATORIAL_GO_producao_saida"

DIGITACAO_SCRIPT = LOCAL_DIR / "digitacao_consen" / "digitacao_consen_enel.py"
FILTRO_SCRIPT = LOCAL_DIR / "digitacao_consen" / "enel_filtro.py"

PYTHON_EXE = str(Path(sys.executable).parent / "python.exe")
STATUS_MOVER_ESTRITO = {"sucesso_auditoria", "auditoria_sem_valor"}
_SLUG_FILE = LOCAL_DIR / "pipelines" / ".producao_equatorial_go_slug"


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("pipeline_producao_equatorial_go")


def _mkdir_seguro(path: Path) -> None:
    try:
        path.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass


def _to_ascii_upper(text: str) -> str:
    return unicodedata.normalize("NFKD", text or "").encode("ASCII", "ignore").decode("ASCII").upper()


def _rodar(descricao: str, cmd: list[str], env_extra: dict | None = None) -> int:
    log.info("=" * 60)
    log.info("  %s", descricao)
    log.info("=" * 60)
    log.info("  Comando: %s", " ".join(cmd))

    env = os.environ.copy()
    env.update({"PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8", "PYTHONUNBUFFERED": "1"})
    if env_extra:
        env.update({k: str(v) for k, v in env_extra.items()})

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
        env=env,
        creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if os.name == "nt" else 0,
    )

    def _drenar(stream, prefixo: str) -> None:
        for linha in iter(stream.readline, ""):
            linha = linha.rstrip()
            if linha:
                log.info("  [%s] %s", prefixo, linha)

    t_out = threading.Thread(target=_drenar, args=(proc.stdout, "OUT"), daemon=True)
    t_err = threading.Thread(target=_drenar, args=(proc.stderr, "ERR"), daemon=True)
    t_out.start()
    t_err.start()
    t_out.join()
    t_err.join()

    proc.wait()
    code = int(proc.returncode or 0)
    log.info("%s  %s -> exit %d", "OK" if code == 0 else "FALHA", descricao, code)
    return code


def _normalizar_carimbo(valor: str) -> str:
    txt = str(valor or "").strip().upper()
    if not txt:
        return ""
    if txt.endswith(".0"):
        txt = txt[:-2]
    if txt.startswith("BB_"):
        return txt
    if txt.isdigit():
        return f"BB_{txt}"
    return txt


def _carimbos_da_pasta(pasta: Path) -> list[str]:
    if not pasta.exists():
        return []
    return sorted(_normalizar_carimbo(p.stem) for p in pasta.glob("BB_*.pdf"))


def _atualizar_arquivo_master(master: MasterIndice, carimbo: str, arquivo: Path) -> None:
    linhas: list[dict] = []
    encontrado = False

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(master.master_file, newline="", encoding=enc) as f:
                linhas = list(csv.DictReader(f))
            break
        except UnicodeDecodeError:
            continue
    if not linhas:
        return

    for row in linhas:
        if _normalizar_carimbo(row.get("INDICE", "")) == carimbo:
            row["ARQUIVO"] = str(arquivo)
            encontrado = True

    if not encontrado:
        return

    tmp = master.master_file.with_suffix(".tmp")
    with open(tmp, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=MASTER_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in linhas:
            writer.writerow(row)
    tmp.replace(master.master_file)


def _carimbar_pdf_na_origem(pdf: Path, carimbo: str, master: MasterIndice) -> Path:
    destino = pdf.with_name(f"{carimbo}.pdf")
    if pdf == destino:
        _atualizar_arquivo_master(master, carimbo, destino)
        return destino

    if destino.exists():
        log.warning("  Arquivo carimbado ja existe na origem: %s", destino.name)
        _atualizar_arquivo_master(master, carimbo, destino)
        return destino

    try:
        pdf.rename(destino)
        log.info("  Carimbado na origem: %s -> %s", pdf.name, destino.name)
        _atualizar_arquivo_master(master, carimbo, destino)
        return destino
    except PermissionError as exc:
        # Se o PDF estiver aberto por outro processo, nao derruba o lote todo.
        # Mantemos o caminho original e seguimos com staging/OCR/digitacao.
        log.warning("  Arquivo em uso, carimbo em nome adiado: %s (%s)", pdf.name, exc)
        return pdf
    except OSError as exc:
        log.warning("  Falha ao renomear %s -> %s: %s", pdf.name, destino.name, exc)
        return pdf


def _identificar_pdf_equatorial(pdf_path: Path) -> dict:
    info = {
        "sistema": "DESCONHECIDA",
        "instalacao": "",
        "mes_ref": "",
        "grupo": "",
        "tipo": "",
    }

    try:
        text = ocr_eq._extract_text(str(pdf_path))
    except Exception as exc:
        log.warning("  Falha ao ler PDF %s: %s", pdf_path.name, exc)
        return info

    if not ocr_eq._eh_layout_equatorial_go(text, pdf_path.name):
        return info

    tipo = ocr_eq._detectar_tipo_equatorial_go(text)
    instalacao = ocr_eq._resolver_instalacao_equatorial_go(pdf_path, text, tipo)
    mes_ref = ocr_eq._eq_mes_ref_texto(text)

    info.update(
        {
            "sistema": "EQUATORIAL",
            "instalacao": instalacao,
            "mes_ref": mes_ref,
            "grupo": "A" if tipo == "mt" else "B",
            "tipo": tipo,
        }
    )
    return info


def _buscar_carimbo_existente(master: MasterIndice, uc: str, mes_ref: str, sistema: str) -> str:
    try:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(master.master_file, newline="", encoding=enc) as f:
                    for row in csv.DictReader(f):
                        if (
                            row.get("SISTEMA", "").strip().upper() == sistema.upper()
                            and row.get("UC", "").strip().lstrip("0") == uc.strip().lstrip("0")
                            and row.get("MES_REF", "").strip() == mes_ref.strip()
                        ):
                            ind = row.get("INDICE", "").strip()
                            if ind.startswith("BB_"):
                                return ind
                break
            except UnicodeDecodeError:
                continue
    except Exception:
        pass
    return ""


def etapa_carimbos(pdfs: list[Path], master: MasterIndice, cache_info: dict[Path, dict]) -> dict[Path, str]:
    log.info("=" * 60)
    log.info("  ETAPA 1 - Carimbos Equatorial GO")
    log.info("=" * 60)

    mapa: dict[Path, str] = {}
    novos = existentes = ignorados = 0

    for pdf in pdfs:
        info = cache_info.get(pdf) or _identificar_pdf_equatorial(pdf)
        sistema = info.get("sistema", "")
        instalacao = info.get("instalacao", "")
        mes_ref = info.get("mes_ref", "")

        if sistema != "EQUATORIAL":
            ignorados += 1
            continue

        if not instalacao or not mes_ref:
            log.warning("  IGNORADO (sem instalacao ou mes_ref): %s | %s", pdf.name, info)
            ignorados += 1
            continue

        if master.ja_foi_baixado(instalacao, mes_ref, sistema):
            carimbo = _buscar_carimbo_existente(master, instalacao, mes_ref, sistema)
            if not carimbo:
                carimbo = master.consumir_carimbo()
                master.registrar(
                    indice_bb=carimbo,
                    sistema="EQUATORIAL",
                    uc=instalacao,
                    mes_ref=mes_ref,
                    estado="GO",
                    arquivo=str(pdf),
                )
                log.info("  NOVO (reregistrado) -> %s | UC=%s | %s", carimbo, instalacao, mes_ref)
            else:
                log.info("  JA REGISTRADO -> %s | UC=%s | %s", carimbo, instalacao, mes_ref)
            pdf_carimbado = _carimbar_pdf_na_origem(pdf, carimbo, master)
            mapa[pdf_carimbado] = carimbo
            cache_info[pdf_carimbado] = info
            existentes += 1
            continue

        carimbo = master.consumir_carimbo()
        master.registrar(
            indice_bb=carimbo,
            sistema="EQUATORIAL",
            uc=instalacao,
            mes_ref=mes_ref,
            estado="GO",
            arquivo=str(pdf),
        )
        pdf_carimbado = _carimbar_pdf_na_origem(pdf, carimbo, master)
        mapa[pdf_carimbado] = carimbo
        cache_info[pdf_carimbado] = info
        novos += 1
        log.info("  NOVO -> %s | %s | %s | %s", carimbo, info.get("tipo", "").upper(), instalacao, mes_ref)

    log.info("-" * 60)
    log.info("  Total PDFs origem : %d", len(pdfs))
    log.info("  Novos carimbados  : %d", novos)
    log.info("  Ja existentes     : %d", existentes)
    log.info("  Ignorados         : %d", ignorados)
    log.info("  Proximo carimbo   : %s", master.proximo_carimbo)

    return mapa


def etapa_staging(
    mapa: dict[Path, str],
    staging_root: Path,
    cache_info: dict[Path, dict],
) -> tuple[Path, Path, dict[str, Path], list[str], list[str]]:
    log.info("=" * 60)
    log.info("  ETAPA 2 - Staging BT/MT")
    log.info("=" * 60)

    pasta_bt = staging_root / "BT"
    pasta_mt = staging_root / "MT"
    _mkdir_seguro(pasta_bt)
    _mkdir_seguro(pasta_mt)

    mapa_reverso: dict[str, Path] = {}
    bt_carimbos: list[str] = []
    mt_carimbos: list[str] = []

    for pdf, carimbo in mapa.items():
        info = cache_info.get(pdf) or {}
        tipo = info.get("tipo", "")
        if tipo == "mt":
            destino = pasta_mt / f"{carimbo}.pdf"
            mt_carimbos.append(carimbo)
        else:
            destino = pasta_bt / f"{carimbo}.pdf"
            bt_carimbos.append(carimbo)

        if not destino.exists():
            shutil.copy2(pdf, destino)
            log.info("  Staging: %s -> %s", pdf.name, destino)

        mapa_reverso[carimbo] = pdf

    log.info("  BT em staging : %d", len(bt_carimbos))
    log.info("  MT em staging : %d", len(mt_carimbos))
    return pasta_bt, pasta_mt, mapa_reverso, sorted(bt_carimbos), sorted(mt_carimbos)


def _reconstruir_staging_da_raiz(
    pasta_pdfs: Path,
    staging_root: Path,
) -> tuple[Path, Path, dict[str, Path], list[str], list[str]]:
    log.info("=" * 60)
    log.info("  RECONSTRUCAO DE STAGING PELA PASTA RAIZ")
    log.info("=" * 60)

    if not pasta_pdfs.exists():
        log.warning("  Pasta de producao nao encontrada para reconstruir staging: %s", pasta_pdfs)
        return staging_root / "BT", staging_root / "MT", {}, [], []

    pdfs = sorted(p for p in pasta_pdfs.glob("*.pdf") if p.is_file())
    cache_info: dict[Path, dict] = {}
    mapa_carimbo: dict[Path, str] = {}

    for pdf in pdfs:
        info = _identificar_pdf_equatorial(pdf)
        if info.get("sistema") != "EQUATORIAL":
            continue
        carimbo = _normalizar_carimbo(pdf.stem)
        if not carimbo.startswith("BB_"):
            continue
        cache_info[pdf] = info
        mapa_carimbo[pdf] = carimbo

    if not mapa_carimbo:
        log.warning("  Nenhum PDF Equatorial GO carimbado encontrado na pasta raiz para reconstruir staging.")
        return staging_root / "BT", staging_root / "MT", {}, [], []

    log.info("  PDFs Equatorial GO reaproveitados da raiz: %d", len(mapa_carimbo))
    return etapa_staging(mapa_carimbo, staging_root, cache_info)


def etapa_ocr_tipo(pasta_staging: Path, tipo: str, xlsx_saida: Path, carimbos: list[str]) -> int:
    if not carimbos:
        log.info("  Nenhum PDF %s para OCR.", tipo.upper())
        return 0

    _mkdir_seguro(xlsx_saida.parent)
    if xlsx_saida.exists():
        try:
            xlsx_saida.unlink()
            log.info("  XLSX %s anterior removido para recriacao: %s", tipo.upper(), xlsx_saida)
        except Exception as exc:
            log.warning("  Falha ao remover XLSX %s anterior (%s): %s", tipo.upper(), xlsx_saida, exc)
    pdfs = [pasta_staging / f"{carimbo}.pdf" for carimbo in carimbos if (pasta_staging / f"{carimbo}.pdf").exists()]
    if not pdfs:
        log.warning("  Nenhum PDF fisico encontrado para OCR %s em %s", tipo.upper(), pasta_staging)
        return 1

    log.info("=" * 60)
    log.info("  ETAPA 3 - OCR %s", tipo.upper())
    log.info("=" * 60)
    log.info("  PDFs %s a processar: %d", tipo.upper(), len(pdfs))

    registros: list[dict] = []
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(ocr_eq.processar_pdf, str(pdf), tipo): pdf for pdf in pdfs}
        for future in as_completed(futures):
            registros.append(future.result())

    def _sort_key(row: dict) -> int:
        try:
            return int(row.get("fatCarimbo", 0) or 0)
        except (TypeError, ValueError):
            return 0

    registros.sort(key=_sort_key)
    ocr_eq.salvar_excel(registros, xlsx_saida)
    log.info("  XLSX %s salvo em: %s", tipo.upper(), xlsx_saida)
    return 0


def _xlsx_bt_parece_zerado(xlsx_bt: Path) -> tuple[bool, str]:
    try:
        wb = load_workbook(xlsx_bt, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(v).strip() if v is not None else "" for v in next(rows)]
        idx_raw = {nome: pos for pos, nome in enumerate(header)}
        idx_norm = {_to_ascii_upper(nome): pos for pos, nome in enumerate(header)}

        aliases = {
            "fatValorFatura": ["fatValorFatura"],
            "fatConFPontaIndRegistrado": ["fatConFPontaIndRegistrado"],
            "fatConFPontaIndValorReais": ["fatConFPontaIndValorReais"],
            "Instalacao": ["Instalacao", "Instalação"],
        }

        idx: dict[str, int] = {}
        faltando: list[str] = []
        for canonico, opcoes in aliases.items():
            pos = None
            for nome in opcoes:
                if nome in idx_raw:
                    pos = idx_raw[nome]
                    break
                pos = idx_norm.get(_to_ascii_upper(nome))
                if pos is not None:
                    break
            if pos is None:
                faltando.append(canonico)
            else:
                idx[canonico] = pos

        if faltando:
            wb.close()
            return True, f"colunas ausentes no XLSX BT: {', '.join(faltando)}"

        total = 0
        suspeitas = 0
        exemplos: list[str] = []
        for row in rows:
            if not row or all(v in (None, "") for v in row):
                continue
            total += 1
            instal = str(row[idx["Instalacao"]] or "").strip()
            valor = float(row[idx["fatValorFatura"]] or 0)
            consumo = float(row[idx["fatConFPontaIndRegistrado"]] or 0)
            valor_cons = float(row[idx["fatConFPontaIndValorReais"]] or 0)
            if valor <= 0 or (consumo <= 0 and valor_cons <= 0):
                suspeitas += 1
                if len(exemplos) < 5:
                    exemplos.append(f"{instal or 'sem_inst'}: valor={valor} consumo={consumo} valorCons={valor_cons}")
        wb.close()

        if total == 0:
            return True, "XLSX BT sem linhas de dados"
        if suspeitas == total:
            return True, "todas as linhas BT parecem zeradas: " + " | ".join(exemplos)
        if suspeitas / total >= 0.8:
            return True, f"{suspeitas}/{total} linhas BT parecem zeradas: " + " | ".join(exemplos)
        return False, f"sanidade BT ok ({total - suspeitas}/{total} linhas parecem consistentes)"
    except Exception as exc:
        return True, f"falha na validacao do XLSX BT: {exc}"


def etapa_digitacao_bt(xlsx_bt: Path, pasta_saida: Path) -> int:
    if not xlsx_bt.exists():
        log.warning("  XLSX BT nao encontrado, digitacao pulada: %s", xlsx_bt)
        return 0
    if not DIGITACAO_SCRIPT.exists():
        log.error("Script de digitacao nao encontrado: %s", DIGITACAO_SCRIPT)
        return 1

    suspeito, detalhe = _xlsx_bt_parece_zerado(xlsx_bt)
    if suspeito:
        log.error("  XLSX BT reprovado na sanidade - digitacao abortada: %s", detalhe)
        return 2
    log.info("  Sanidade XLSX BT: %s", detalhe)

    _mkdir_seguro(pasta_saida)
    for nome in ("auditoria_resultados.csv", "resultado_preenchimento_csv.csv"):
        arquivo = pasta_saida / nome
        if arquivo.exists():
            try:
                arquivo.unlink()
                log.info("  Arquivo de saida anterior removido para nova digitacao: %s", arquivo)
            except Exception as exc:
                log.warning("  Falha ao remover saida anterior %s: %s", arquivo, exc)

    for pendente in pasta_saida.glob("pendentes_*.csv"):
        try:
            pendente.unlink()
        except Exception:
            pass

    env_extra = {
        "ENEL_EXCEL_PATH": str(xlsx_bt),
        "CONSEN_PIPELINE_SAIDA": str(pasta_saida),
        "CONSEN_INTERATIVO_FECHAR": "0",
        "DIGITACAO_FATOR_VELOCIDADE":    "0.25",
        "CONSEN_PERMITIR_LOTE_COMPLETO": "1",
    }
    return _rodar("DIGITACAO EQUATORIAL GO BT", [PYTHON_EXE, str(DIGITACAO_SCRIPT)], env_extra=env_extra)


def _ler_carimbos_moveis(auditoria_csv: Path) -> set[str]:
    carimbos: set[str] = set()
    linhas: list[list[str]] = []

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(auditoria_csv, newline="", encoding=enc) as f:
                linhas = list(csv.reader(f, delimiter=";"))
            break
        except UnicodeDecodeError:
            continue
        except Exception:
            return carimbos

    if not linhas:
        return carimbos

    header = [c.strip().lower() for c in linhas[0]]
    try:
        idx_carimbo = header.index("carimbo")
        idx_status = header.index("status")
    except ValueError:
        return carimbos

    for cols in linhas[1:]:
        if len(cols) <= max(idx_carimbo, idx_status):
            continue
        carimbo = _normalizar_carimbo(cols[idx_carimbo].strip())
        status_col = cols[-1] if len(cols) > len(header) else cols[idx_status]
        status = status_col.strip().lower()
        if carimbo and status in STATUS_MOVER_ESTRITO:
            carimbos.add(carimbo)

    return carimbos


def _carregar_arquivo_original_por_carimbo(master: MasterIndice, carimbos: set[str]) -> dict[str, Path]:
    if not carimbos:
        return {}

    mapa: dict[str, Path] = {}
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(master.master_file, newline="", encoding=enc) as f:
                for row in csv.DictReader(f):
                    carimbo = _normalizar_carimbo(row.get("INDICE", ""))
                    if carimbo not in carimbos:
                        continue
                    arquivo = str(row.get("ARQUIVO", "")).strip()
                    if not arquivo:
                        continue
                    original = Path(arquivo)
                    if original.exists():
                        mapa[carimbo] = original
            return mapa
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            log.warning("  Falha ao ler master para reconstruir staging: %s", exc)
            return mapa
    return mapa


def _reconstruir_staging_para_filtro(auditoria_csv: Path, pasta_staging: Path, master: MasterIndice) -> dict[str, Path]:
    carimbos = _ler_carimbos_moveis(auditoria_csv)
    if not carimbos:
        log.info("  Nenhum carimbo elegivel para reconstruir staging.")
        return {}

    mapa_reverso = _carregar_arquivo_original_por_carimbo(master, carimbos)
    if not mapa_reverso:
        log.warning("  Nao foi possivel reconstruir staging via master.")
        return {}

    _mkdir_seguro(pasta_staging)
    for carimbo in sorted(carimbos):
        original = mapa_reverso.get(carimbo)
        if not original or not original.exists():
            continue
        destino = pasta_staging / f"{carimbo}.pdf"
        if not destino.exists():
            try:
                shutil.copy2(original, destino)
            except Exception as exc:
                log.warning("  Falha ao copiar %s -> %s: %s", original.name, destino.name, exc)
    return mapa_reverso


def etapa_filtro_bt(pasta_saida: Path, pasta_staging_bt: Path, mapa_reverso: dict[str, Path]) -> int:
    auditoria_csv = pasta_saida / "auditoria_resultados.csv"
    if not auditoria_csv.exists():
        log.warning("  auditoria_resultados.csv nao encontrado em %s - filtro pulado", pasta_saida)
        return 0
    if not FILTRO_SCRIPT.exists():
        log.error("Script de filtro nao encontrado: %s", FILTRO_SCRIPT)
        return 1

    _mkdir_seguro(DIGITADAS_DIR)
    env_extra = {
        "ENEL_FILTRO_CSV": str(auditoria_csv),
        "ENEL_FILTRO_PDFS": str(pasta_staging_bt),
        "ENEL_FILTRO_DESTINO": str(DIGITADAS_DIR),
    }
    code = _rodar("FILTRO EQUATORIAL GO BT", [PYTHON_EXE, str(FILTRO_SCRIPT)], env_extra=env_extra)

    carimbos_moveis = _ler_carimbos_moveis(auditoria_csv)
    removidos = nao_encontrados = 0
    for carimbo in sorted(carimbos_moveis):
        original = mapa_reverso.get(carimbo)
        if original and original.exists():
            try:
                original.unlink()
                removidos += 1
                log.info("  Original removido da raiz: %s", original.name)
            except Exception as exc:
                log.warning("  Falha ao remover original %s: %s", original.name, exc)
        else:
            nao_encontrados += 1

    log.info("  Originais BT removidos da raiz: %d | nao encontrados: %d", removidos, nao_encontrados)

    try:
        contadores = marcar_digitados_do_auditoria(auditoria_csv, MasterIndice())
        log.info("  [MASTER] %s", contadores)
    except Exception as exc:
        log.warning("  [MASTER] Nao foi possivel atualizar: %s", exc)

    return code


def _xlsx_mt_parece_zerado(xlsx_mt: Path) -> tuple[bool, str]:
    try:
        wb = load_workbook(xlsx_mt, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(v).strip() if v is not None else "" for v in next(rows)]
        idx_raw = {nome: pos for pos, nome in enumerate(header)}
        idx_norm = {_to_ascii_upper(nome): pos for pos, nome in enumerate(header)}

        aliases = {
            "fatValorFatura":         ["fatValorFatura"],
            "fatDemContratadaFPonta": ["fatDemContratadaFPonta"],
            "fatConFPontaIndRegistrado": ["fatConFPontaIndRegistrado"],
            "Instalacao":             ["Instalacao", "Instalação"],
        }
        idx: dict[str, int] = {}
        faltando: list[str] = []
        for canonico, opcoes in aliases.items():
            pos = None
            for nome in opcoes:
                if nome in idx_raw:
                    pos = idx_raw[nome]
                    break
                pos = idx_norm.get(_to_ascii_upper(nome))
                if pos is not None:
                    break
            if pos is None:
                faltando.append(canonico)
            else:
                idx[canonico] = pos

        if faltando:
            wb.close()
            return True, f"colunas ausentes no XLSX MT: {', '.join(faltando)}"

        total = suspeitas = 0
        exemplos: list[str] = []
        for row in rows:
            if not row or all(v in (None, "") for v in row):
                continue
            total += 1
            instal = str(row[idx["Instalacao"]] or "").strip()
            valor  = float(row[idx["fatValorFatura"]] or 0)
            dem    = float(row[idx["fatDemContratadaFPonta"]] or 0)
            consumo = float(row[idx["fatConFPontaIndRegistrado"]] or 0)
            if valor <= 0 or (dem <= 0 and consumo <= 0):
                suspeitas += 1
                if len(exemplos) < 5:
                    exemplos.append(
                        f"{instal or 'sem_inst'}: valor={valor} dem={dem} consumo={consumo}"
                    )
        wb.close()

        if total == 0:
            return True, "XLSX MT sem linhas de dados"
        if suspeitas == total:
            return True, "todas as linhas MT parecem zeradas: " + " | ".join(exemplos)
        if suspeitas / total >= 0.8:
            return True, f"{suspeitas}/{total} linhas MT parecem zeradas: " + " | ".join(exemplos)
        return False, f"sanidade MT ok ({total - suspeitas}/{total} linhas parecem consistentes)"
    except Exception as exc:
        return True, f"falha na validacao do XLSX MT: {exc}"


def etapa_digitacao_mt(xlsx_mt: Path, pasta_saida: Path) -> int:
    if not xlsx_mt.exists():
        log.warning("  XLSX MT nao encontrado, digitacao MT pulada: %s", xlsx_mt)
        return 0
    if not DIGITACAO_SCRIPT.exists():
        log.error("Script de digitacao nao encontrado: %s", DIGITACAO_SCRIPT)
        return 1

    suspeito, detalhe = _xlsx_mt_parece_zerado(xlsx_mt)
    if suspeito:
        log.error("  XLSX MT reprovado na sanidade - digitacao MT abortada: %s", detalhe)
        return 2
    log.info("  Sanidade XLSX MT: %s", detalhe)

    _mkdir_seguro(pasta_saida)
    for nome in ("auditoria_resultados.csv", "resultado_preenchimento_csv.csv"):
        arquivo = pasta_saida / nome
        if arquivo.exists():
            try:
                arquivo.unlink()
                log.info("  Arquivo de saida anterior removido: %s", arquivo)
            except Exception as exc:
                log.warning("  Falha ao remover saida anterior %s: %s", arquivo, exc)

    for pendente in pasta_saida.glob("pendentes_*.csv"):
        try:
            pendente.unlink()
        except Exception:
            pass

    env_extra = {
        "ENEL_EXCEL_PATH":           str(xlsx_mt),
        "CONSEN_PIPELINE_SAIDA":     str(pasta_saida),
        "CONSEN_INTERATIVO_FECHAR":  "0",
        "DIGITACAO_FATOR_VELOCIDADE":    "0.25",
        "CONSEN_PERMITIR_LOTE_COMPLETO": "1",
    }
    return _rodar("DIGITACAO EQUATORIAL GO MT", [PYTHON_EXE, str(DIGITACAO_SCRIPT)], env_extra=env_extra)


def etapa_filtro_mt(pasta_saida: Path, pasta_staging_mt: Path, mapa_reverso: dict[str, Path]) -> int:
    auditoria_csv = pasta_saida / "auditoria_resultados.csv"
    if not auditoria_csv.exists():
        log.warning("  auditoria_resultados.csv nao encontrado em %s - filtro MT pulado", pasta_saida)
        return 0
    if not FILTRO_SCRIPT.exists():
        log.error("Script de filtro nao encontrado: %s", FILTRO_SCRIPT)
        return 1

    _mkdir_seguro(DIGITADAS_DIR)
    env_extra = {
        "ENEL_FILTRO_CSV":      str(auditoria_csv),
        "ENEL_FILTRO_PDFS":     str(pasta_staging_mt),
        "ENEL_FILTRO_DESTINO":  str(DIGITADAS_DIR),
    }
    code = _rodar("FILTRO EQUATORIAL GO MT", [PYTHON_EXE, str(FILTRO_SCRIPT)], env_extra=env_extra)

    carimbos_moveis = _ler_carimbos_moveis(auditoria_csv)
    removidos = nao_encontrados = 0
    for carimbo in sorted(carimbos_moveis):
        original = mapa_reverso.get(carimbo)
        if original and original.exists():
            try:
                original.unlink()
                removidos += 1
                log.info("  Original MT removido da raiz: %s", original.name)
            except Exception as exc:
                log.warning("  Falha ao remover original MT %s: %s", original.name, exc)
        else:
            nao_encontrados += 1

    log.info("  Originais MT removidos da raiz: %d | nao encontrados: %d", removidos, nao_encontrados)

    try:
        contadores = marcar_digitados_do_auditoria(auditoria_csv, MasterIndice())
        log.info("  [MASTER] MT %s", contadores)
    except Exception as exc:
        log.warning("  [MASTER] Nao foi possivel atualizar MT: %s", exc)

    return code


def _carregar_ou_criar_slug() -> str:
    if _SLUG_FILE.exists():
        slug = _SLUG_FILE.read_text(encoding="utf-8").strip()
        if slug:
            return slug
    slug = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    _SLUG_FILE.write_text(slug, encoding="utf-8")
    return slug


def _novo_slug() -> str:
    slug = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    _SLUG_FILE.write_text(slug, encoding="utf-8")
    return slug


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Pipeline de producao Equatorial GO a partir da pasta BB/ENZO")
    parser.add_argument("--pasta", type=str, default=str(PASTA_ENZO), help="Pasta raiz com os PDFs da producao geral")
    parser.add_argument("--so-carimbo", action="store_true", help="So carimba e registra no master")
    parser.add_argument("--so-ocr", action="store_true", help="So OCR (usa staging ja preparado)")
    parser.add_argument("--so-digitacao", action="store_true", help="So digitacao BT (usa XLSX BT ja existente)")
    parser.add_argument("--so-filtro", action="store_true", help="So filtro BT (usa auditoria_resultados.csv ja existente)")
    parser.add_argument("--novo-slug", action="store_true", help="Forca novo slug para uma execucao do zero")
    parser.add_argument("--manter-staging", action="store_true", help="Nao apaga o staging ao final")
    parser.add_argument("--slug", type=str, default="", help="Slug de uma execucao anterior para retomar")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    pasta_pdfs = Path(args.pasta.strip())

    if args.slug:
        slug = args.slug.strip()
        _SLUG_FILE.write_text(slug, encoding="utf-8")
    else:
        slug = _novo_slug() if args.novo_slug or not (args.so_ocr or args.so_digitacao or args.so_filtro) else _carregar_ou_criar_slug()

    staging_root = Path(tempfile.gettempdir()) / f"producao_equatorial_go_{slug}"
    pasta_staging_bt = staging_root / "BT"
    pasta_staging_mt = staging_root / "MT"
    xlsx_bt = OCR_SAIDA_DIR / f"ocr_equatorial_go_BT_{slug}.xlsx"
    xlsx_mt = OCR_SAIDA_DIR / f"ocr_equatorial_go_MT_{slug}.xlsx"
    saida_bt = PIPELINE_SAIDA_DIR / slug / "BT"
    saida_mt = PIPELINE_SAIDA_DIR / slug / "MT"

    log.info("=" * 60)
    log.info("  PIPELINE PRODUCAO EQUATORIAL GO")
    log.info("=" * 60)
    log.info("  Slug        : %s", slug)
    log.info("  Pasta PDFs  : %s", pasta_pdfs)
    log.info("  Staging     : %s", staging_root)
    log.info("  XLSX BT     : %s", xlsx_bt)
    log.info("  XLSX MT     : %s", xlsx_mt)
    log.info("  Saida BT    : %s", saida_bt)
    log.info("  Saida MT    : %s", saida_mt)
    log.info("  Digitadas   : %s", DIGITADAS_DIR)

    mapa_carimbo: dict[Path, str] = {}
    mapa_reverso: dict[str, Path] = {}
    cache_info: dict[Path, dict] = {}
    falhas: list[str] = []

    if not args.so_ocr and not args.so_digitacao and not args.so_filtro:
        if not pasta_pdfs.exists():
            log.error("Pasta nao encontrada: %s", pasta_pdfs)
            return 1

        pdfs = sorted(p for p in pasta_pdfs.glob("*.pdf") if p.is_file())
        if not pdfs:
            log.warning("Nenhum PDF encontrado na raiz de: %s", pasta_pdfs)
            return 0

        log.info("  PDFs encontrados na raiz: %d", len(pdfs))
        for pdf in pdfs:
            cache_info[pdf] = _identificar_pdf_equatorial(pdf)

        eq_encontrados = sum(1 for info in cache_info.values() if info.get("sistema") == "EQUATORIAL")
        log.info("  PDFs Equatorial GO identificados: %d", eq_encontrados)
        if not eq_encontrados:
            log.warning("Nenhum PDF Equatorial GO identificado na pasta raiz.")
            return 0

        master = MasterIndice()
        mapa_carimbo = etapa_carimbos(pdfs, master, cache_info)
        if not mapa_carimbo:
            log.warning("Nenhum PDF Equatorial GO mapeado para processamento.")
            return 0

        pasta_staging_bt, pasta_staging_mt, mapa_reverso, bt_carimbos, mt_carimbos = etapa_staging(
            mapa_carimbo, staging_root, cache_info
        )
        if args.so_carimbo:
            log.info("  --so-carimbo: carimbo concluido e staging preparado; encerrando antes do OCR.")
            return 0
    else:
        bt_carimbos = _carimbos_da_pasta(pasta_staging_bt)
        mt_carimbos = _carimbos_da_pasta(pasta_staging_mt)
        if args.so_ocr and not bt_carimbos and not mt_carimbos:
            log.info("  Staging vazio para o slug %s; tentando reconstruir a partir da pasta raiz.", slug)
            pasta_staging_bt, pasta_staging_mt, mapa_reverso, bt_carimbos, mt_carimbos = _reconstruir_staging_da_raiz(
                pasta_pdfs, staging_root
            )

    if not args.so_digitacao and not args.so_filtro:
        cod = etapa_ocr_tipo(pasta_staging_bt, "bt", xlsx_bt, bt_carimbos)
        if cod != 0:
            falhas.append("OCR_BT")
        cod = etapa_ocr_tipo(pasta_staging_mt, "mt", xlsx_mt, mt_carimbos)
        if cod != 0:
            falhas.append("OCR_MT")

    if not args.so_ocr and not args.so_filtro:
        cod = etapa_digitacao_bt(xlsx_bt, saida_bt)
        if cod != 0:
            falhas.append("DIGITACAO_BT")

    if not args.so_ocr and not args.so_digitacao:
        if args.so_filtro and (not pasta_staging_bt.exists() or not any(pasta_staging_bt.glob("*.pdf"))):
            mapa_reverso = _reconstruir_staging_para_filtro(saida_bt / "auditoria_resultados.csv", pasta_staging_bt, MasterIndice())
        cod = etapa_filtro_bt(saida_bt, pasta_staging_bt, mapa_reverso)
        if cod != 0:
            falhas.append("FILTRO_BT")

    # -- MT: digitação e filtro ------------------------------------------------
    if not args.so_ocr and not args.so_filtro:
        cod = etapa_digitacao_mt(xlsx_mt, saida_mt)
        if cod != 0:
            falhas.append("DIGITACAO_MT")

    if not args.so_ocr and not args.so_digitacao:
        mapa_reverso_mt = mapa_reverso
        if args.so_filtro and (not pasta_staging_mt.exists() or not any(pasta_staging_mt.glob("*.pdf"))):
            mapa_reverso_mt = _reconstruir_staging_para_filtro(
                saida_mt / "auditoria_resultados.csv", pasta_staging_mt, MasterIndice()
            )
        cod = etapa_filtro_mt(saida_mt, pasta_staging_mt, mapa_reverso_mt)
        if cod != 0:
            falhas.append("FILTRO_MT")

    if not args.manter_staging and staging_root.exists():
        try:
            shutil.rmtree(staging_root, ignore_errors=True)
        except Exception as exc:
            log.warning("  Falha ao remover staging: %s", exc)

    log.info("=" * 60)
    if falhas:
        log.info("  PIPELINE PRODUCAO EQUATORIAL GO FINALIZADO COM FALHAS: %s", ", ".join(falhas))
        log.info("=" * 60)
        return 1

    log.info("  PIPELINE PRODUCAO EQUATORIAL GO FINALIZADO COM SUCESSO")
    log.info("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
