#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import logging
import os
import subprocess
import sys
import threading
from pathlib import Path

from openpyxl import Workbook, load_workbook


LOCAL_DIR = Path(__file__).parent.parent
DIGITACAO_SCRIPT = LOCAL_DIR / "digitacao_consen" / "digitacao_consen_enel.py"
FILTRO_SCRIPT = LOCAL_DIR / "digitacao_consen" / "enel_filtro.py"
SERVIDOR = Path("//10.10.250.21/Energia")
OCR_SAIDA_DIR = SERVIDOR / "ARQUIVOS ENZO" / "OCR ENEL"
PIPELINE_SAIDA = SERVIDOR / "ARQUIVOS ENZO" / "ENEL_pipeline_saida"
PDFS_DIR = SERVIDOR / "ARQUIVOS ENZO" / "DOWNLOAD ENEL" / "Faltantes"
DIGITADAS_DIR = SERVIDOR / "CONTASDEENERGIAELETRICA" / "BB" / "ENZO" / "Digitadas"
PYTHON_EXE = str(Path(sys.executable).parent / "python.exe")

CORRECOES_PADRAO = {
    "2003132": "BTE0016958",
    "2003133": "BTE0017836",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def _coluna_por_prefixo(headers: list[str], prefixo: str) -> int:
    prefixo = prefixo.lower()
    for idx, header in enumerate(headers, 1):
        if str(header or "").lower().startswith(prefixo):
            return idx
    raise ValueError(f"Coluna com prefixo '{prefixo}' nao encontrada.")


def gerar_planilha_corrigida(origem: Path, destino: Path, correcoes: dict[str, str]) -> int:
    wb = load_workbook(origem, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    idx_inst = _coluna_por_prefixo(headers, "instala")
    idx_carimbo = headers.index("fatCarimbo") + 1

    novo_wb = Workbook()
    novo_ws = novo_wb.active
    novo_ws.title = ws.title
    novo_ws.append(headers)

    total = 0
    for row_idx in range(2, ws.max_row + 1):
        valores = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        carimbo = str(valores[idx_carimbo - 1] or "").strip()
        if carimbo not in correcoes:
            continue
        valores[idx_inst - 1] = correcoes[carimbo]
        novo_ws.append(valores)
        total += 1

    destino.parent.mkdir(parents=True, exist_ok=True)
    novo_wb.save(destino)
    return total


def etapa_digitacao(xlsx: Path) -> int:
    env = os.environ.copy()
    env["ENEL_EXCEL_PATH"] = str(xlsx)
    env["CONSEN_PIPELINE_SAIDA"] = str(PIPELINE_SAIDA)
    env["CONSEN_INTERATIVO_FECHAR"] = "0"
    env["DIGITACAO_FATOR_VELOCIDADE"] = "0.25"
    env["CONSEN_PERMITIR_LOTE_COMPLETO"] = "1"
    env["CONSEN_SENHA"]              = "Acao2026"
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUNBUFFERED"] = "1"

    log.info("=" * 60)
    log.info(f"  Digitacao ENEL BTE corrigidas - {xlsx.name}")
    log.info("=" * 60)

    proc = subprocess.Popen(
        [PYTHON_EXE, str(DIGITACAO_SCRIPT)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
        env=env,
        creationflags=subprocess.CREATE_NEW_PROCESS_GROUP,
    )

    def _drenar(stream, prefixo):
        for linha in iter(stream.readline, ""):
            linha = linha.rstrip()
            if linha:
                log.info(f"  [{prefixo}] {linha}")

    t_out = threading.Thread(target=_drenar, args=(proc.stdout, "DIG"), daemon=True)
    t_err = threading.Thread(target=_drenar, args=(proc.stderr, "DIG-ERR"), daemon=True)
    t_out.start()
    t_err.start()
    t_out.join()
    t_err.join()

    proc.wait()
    codigo = proc.returncode
    log.info(f"Digitacao exit {codigo}")
    return codigo


def _atualizar_master_pos_filtro(auditoria_csv: Path) -> None:
    try:
        import sys as _sys
        _sys.path.insert(0, str(LOCAL_DIR))
        from indice_master import MasterIndice, marcar_digitados_do_auditoria
        contadores = marcar_digitados_do_auditoria(auditoria_csv, MasterIndice())
        log.info(f"  [MASTER] Digitacao atualizada: {contadores}")
    except Exception as exc:
        log.warning(f"  [MASTER] Nao foi possivel atualizar o indice master: {exc}")


def etapa_filtro(mes: str, ano: str) -> int:
    auditoria_csv = PIPELINE_SAIDA / "auditoria_resultados.csv"
    env = os.environ.copy()
    env["ENEL_FILTRO_CSV"] = str(auditoria_csv)
    env["ENEL_FILTRO_PDFS"] = str(PDFS_DIR / f"{mes}-{ano}" / "BT")
    env["ENEL_FILTRO_DESTINO"] = str(DIGITADAS_DIR)
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUNBUFFERED"] = "1"

    log.info("=" * 60)
    log.info("  Filtro ENEL BTE corrigidas")
    log.info("=" * 60)

    proc = subprocess.Popen(
        [PYTHON_EXE, str(FILTRO_SCRIPT)],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
        env=env,
    )

    for linha in iter(proc.stdout.readline, ""):
        linha = linha.rstrip()
        if linha:
            log.info(f"  [FILTRO] {linha}")

    proc.wait()
    codigo = proc.returncode
    log.info(f"Filtro exit {codigo}")
    if codigo == 0:
        _atualizar_master_pos_filtro(auditoria_csv)
    return codigo


def parse_args():
    p = argparse.ArgumentParser(description="Corrige 2 linhas BTE da ENEL e roda digitacao+filtro.")
    p.add_argument("--mes", default="03")
    p.add_argument("--ano", default="2026")
    return p.parse_args()


def main():
    args = parse_args()
    origem = OCR_SAIDA_DIR / f"ocr_enel_BT_{args.mes}{args.ano}.xlsx"
    destino = OCR_SAIDA_DIR / f"ocr_enel_BT_{args.mes}{args.ano}_bte_corrigidas.xlsx"

    total = gerar_planilha_corrigida(origem, destino, CORRECOES_PADRAO)
    log.info(f"Planilha corrigida gerada: {destino} | linhas={total}")
    if total == 0:
        log.error("Nenhuma linha BTE foi encontrada para correcao.")
        sys.exit(1)

    dig = etapa_digitacao(destino)
    if dig != 0:
        sys.exit(dig)

    filtro = etapa_filtro(args.mes, args.ano)
    sys.exit(filtro)


if __name__ == "__main__":
    main()
