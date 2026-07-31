#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline_enel.py
----------------
Pipeline sequencial ENEL BT:
  1. OCR     extrai faturas da pasta DOWNLOAD ENEL ? OCR ENEL (xlsx)
  2. Digitação  digita xlsx no Consen via Selenium
  3. Filtro  move os PDFs digitados para a pasta Digitadas

Uso:
    python pipeline_enel.py                    # mês/ano atual (padrão)
    python pipeline_enel.py --mes 03 --ano 2026
    python pipeline_enel.py --tipo bt          # só BT (padrão)
    python pipeline_enel.py --tipo mt
    python pipeline_enel.py --so-ocr           # só OCR
    python pipeline_enel.py --so-digitacao     # só digitação
    python pipeline_enel.py --so-filtro        # só filtro

Config principal em config.py:
    PIPELINE_ENEL_ATIVAR   = True
    PIPELINE_ENEL_MES      = "03"      # deixar vazio para mês atual
    PIPELINE_ENEL_ANO      = "2026"    # deixar vazio para ano atual
    PIPELINE_ENEL_TIPO     = "bt"      # "bt" | "mt" | "ambos"
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import shutil
import subprocess
import sys
import threading
from pathlib import Path

from openpyxl import load_workbook

try:
    from pipelines._visual_safe import _p, _info, _ok, _fail, _warn, _sep, _banner, _rodar as _rodar_visual
except ModuleNotFoundError:
    from _visual_safe import _p, _info, _ok, _fail, _warn, _sep, _banner, _rodar as _rodar_visual

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

# Raiz local do projeto
LOCAL_DIR         = Path(__file__).parent.parent  # C:\Users\Revit\Desktop\ENERGIA
ROOT_DIR          = LOCAL_DIR.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

OCR_SCRIPT        = LOCAL_DIR / "ocr"             / "ocr_enel.py"
DIGITACAO_SCRIPT  = LOCAL_DIR / "digitacao_consen" / "digitacao_consen_enel.py"
FILTRO_SCRIPT     = LOCAL_DIR / "digitacao_consen" / "enel_filtro.py"

# Dados ficam no servidor
SERVIDOR          = Path("//10.10.250.21/Energia")
DOWNLOAD_DIR      = SERVIDOR / "ARQUIVOS ENZO" / "DOWNLOAD ENEL"
OCR_SAIDA_DIR     = SERVIDOR / "ARQUIVOS ENZO" / "OCR ENEL"
PIPELINE_SAIDA    = SERVIDOR / "ARQUIVOS ENZO" / "ENEL_pipeline_saida"
DIGITADAS_DIR     = SERVIDOR / "CONTASDEENERGIAELETRICA" / "BB" / "ENZO" / "Digitadas"
FALLBACK_ROOT     = LOCAL_DIR / "_runtime_fallback" / "pipeline_enel"

PYTHON_EXE = str(Path(sys.executable).parent / "python.exe")

# =============================================================================
# HELPERS
# =============================================================================

def _rodar(descricao: str, cmd: list[str]) -> int:
    return _rodar_visual(descricao, cmd)


def _safe_exists(path: Path) -> bool:
    try:
        return path.exists()
    except OSError:
        return False


def _resolver_dir(preferido: Path, fallback: Path, rotulo: str) -> Path:
    try:
        preferido.mkdir(parents=True, exist_ok=True)
        return preferido
    except OSError as e:
        fallback.mkdir(parents=True, exist_ok=True)
        _warn(f"{rotulo} indisponível em {preferido}: {e}. Usando fallback local {fallback}")
        return fallback


def _xlsx_bt(mes: str, ano: str) -> Path:
    """Caminho esperado do xlsx BT gerado pelo OCR."""
    return OCR_SAIDA_DIR / f"ocr_enel_BT_{mes}{ano}.xlsx"


def _xlsx_mt(mes: str, ano: str) -> Path:
    return OCR_SAIDA_DIR / f"ocr_enel_MT_{mes}{ano}.xlsx"


def _pasta_download_bt(mes: str, ano: str) -> Path:
    """Pasta de download BT para o filtro."""
    return DOWNLOAD_DIR / f"{mes}-{ano}" / "BT"


def _pasta_download_mt(mes: str, ano: str) -> Path:
    return DOWNLOAD_DIR / f"{mes}-{ano}" / "MT"


# =============================================================================
# ETAPAS
# =============================================================================

def etapa_ocr(mes: str, ano: str, tipo: str, recriar: bool = False) -> int:
    """Roda ocr_enel.py para o mês/tipo especificado."""
    cmd = [PYTHON_EXE, str(OCR_SCRIPT), "--mes", mes, "--ano", ano, "--tipo", tipo]
    if recriar:
        cmd.append("--recriar")
    env_extra = {
        "OCR_ENEL_DOWNLOAD_DIR": str(DOWNLOAD_DIR),
        "OCR_ENEL_SAIDA_DIR": str(OCR_SAIDA_DIR),
    }
    return _rodar_visual(f"OCR ENEL {tipo.upper()} {mes}/{ano}", cmd, env_extra=env_extra)


def etapa_digitacao(xlsx: Path, pasta_pdfs: Path | None = None) -> int:
    if not xlsx.exists():
        _fail(f"Planilha não encontrada: {xlsx}")
        return 1
    env_extra: dict[str, str] = {
        "ENEL_EXCEL_PATH":               str(xlsx),
        "CONSEN_PIPELINE_SAIDA":         str(PIPELINE_SAIDA),
        "CONSEN_INTERATIVO_FECHAR":      "0",
        "CONSEN_INVESTIGAR_ZEROS":       "0",
        "DIGITACAO_FATOR_VELOCIDADE":    "0.25",
        "CONSEN_SENHA":                  "Acao2026",
        "ENEL_AJUSTAR_LEITURA_ULTIMO_DIA": "1",
    }
    if pasta_pdfs is not None:
        env_extra["ENEL_DIGITACAO_PASTA_PDFS"] = str(pasta_pdfs)
    else:
        env_extra["CONSEN_PERMITIR_LOTE_COMPLETO"] = "1"
    return _rodar_visual(f"DIGITAÇÃO ENEL  ({xlsx.name})", [PYTHON_EXE, str(DIGITACAO_SCRIPT)], env_extra=env_extra)

def _valor_cel(row, colunas: dict[str, int], nome: str) -> str:
    idx = colunas.get(nome)
    if not idx:
        return ""
    return str(row[idx - 1].value or "").strip()


def _normalizar_carimbo_xlsx(v: object) -> str:
    txt = str(v or "").strip().upper().replace("BB_", "")
    if txt.endswith(".0"):
        txt = txt[:-2]
    return f"BB_{txt}" if txt else ""


def preparar_digitacao_bb(xlsx: Path, pasta_pdfs: Path) -> tuple[Path, Path]:
    """Cria cópias BB_<carimbo>.pdf e ajusta ARQUIVO no XLSX para o digitador."""
    if not xlsx.exists():
        return xlsx, pasta_pdfs
    wb = load_workbook(xlsx)
    ws = wb.active
    header = [str(c.value or "").strip() for c in ws[1]]
    colunas = {nome: i + 1 for i, nome in enumerate(header) if nome}
    col_carimbo = colunas.get("fatCarimbo")
    col_arquivo = colunas.get("ARQUIVO")
    if not col_carimbo or not col_arquivo:
        return xlsx, pasta_pdfs

    destino = PIPELINE_SAIDA / "_bb_digitacao" / xlsx.stem
    destino.mkdir(parents=True, exist_ok=True)

    for row in ws.iter_rows(min_row=2):
        carimbo = _normalizar_carimbo_xlsx(row[col_carimbo - 1].value)
        arquivo = _valor_cel(row, colunas, "ARQUIVO")
        if not carimbo or not arquivo:
            continue
        src: Path | None = pasta_pdfs / arquivo
        if not src.exists():
            src = next((p for p in pasta_pdfs.rglob(arquivo) if p.is_file()), None)
        if not src or not src.exists():
            continue
        nome_bb = f"{carimbo}.pdf"
        dst = destino / nome_bb
        if not dst.exists():
            shutil.copy2(src, dst)
        row[col_arquivo - 1].value = nome_bb

    xlsx_bb = destino / xlsx.name
    wb.save(xlsx_bb)
    return xlsx_bb, destino


def _atualizar_master_pos_filtro(auditoria_csv: Path) -> None:
    try:
        import sys as _sys
        _sys.path.insert(0, str(LOCAL_DIR))
        from indice_master import MasterIndice, marcar_digitados_do_auditoria
        contadores = marcar_digitados_do_auditoria(auditoria_csv, MasterIndice())
        _info(f"[MASTER] Digitacao atualizada: {contadores}")
    except Exception as _e:
        _warn(f"[MASTER] Nao foi possivel atualizar o indice master: {_e}")


def _resetar_auditoria(pasta_saida: Path) -> None:
    arq = pasta_saida / "auditoria_resultados.csv"
    if _safe_exists(arq):
        arq.unlink()
        _info("[reset] auditoria_resultados.csv removido antes da digitacao.")


def etapa_filtro(pasta_pdfs: Path, pasta_destino: Path) -> int:
    auditoria_csv = PIPELINE_SAIDA / "auditoria_resultados.csv"
    if not auditoria_csv.exists():
        _warn(f"auditoria_resultados.csv não encontrado em {PIPELINE_SAIDA}  filtro pulado")
        return 1
    env_extra = {
        "ENEL_FILTRO_CSV":     str(auditoria_csv),
        "ENEL_FILTRO_PDFS":    str(pasta_pdfs),
        "ENEL_FILTRO_DESTINO": str(pasta_destino),
    }
    codigo = _rodar_visual(f"FILTRO ENEL  {pasta_pdfs.name}", [PYTHON_EXE, str(FILTRO_SCRIPT)], env_extra=env_extra)
    if codigo == 0:
        _atualizar_master_pos_filtro(auditoria_csv)
    else:
        _warn("[MASTER] Atualização do índice pulada porque o filtro falhou.")
    return codigo


# =============================================================================
# ATUALIZAR enel_filtro.py PARA LER VARIÁVEIS DE AMBIENTE
# (injeção de compatibilidade  o filtro original usa constantes hardcoded)
# =============================================================================
# Nota: o enel_filtro.py foi atualizado para respeitar as env vars
# ENEL_FILTRO_CSV / ENEL_FILTRO_PDFS / ENEL_FILTRO_DESTINO quando presentes.
# Ver seção CONFIGURAÇÕES no topo do filtro.


# =============================================================================
# MAIN
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(description="Pipeline ENEL BT/MT: OCR ? Digitação ? Filtro")
    p.add_argument("--mes",           type=str, help="Mês (ex: 03)")
    p.add_argument("--ano",           type=str, help="Ano (ex: 2026)")
    p.add_argument("--tipo",          choices=["bt", "mt", "ambos"], default="bt")
    p.add_argument("--so-ocr",        action="store_true")
    p.add_argument("--so-digitacao",  action="store_true")
    p.add_argument("--so-filtro",     action="store_true")
    p.add_argument("--recriar",       action="store_true",
                   help="Apaga o xlsx de OCR antes de processar (recria do zero, ignora dedup)")
    p.add_argument("--pasta",         type=str, default="",
                   help="Pasta alternativa de PDFs (modo lote avulso, substitui DOWNLOAD_DIR)")
    p.add_argument("--preservar-auditoria", action="store_true",

                   help="Nao apaga auditoria_resultados.csv antes da digitacao; util para retomar lote interrompido")

    return p.parse_args()


def main():
    args = parse_args()

    hoje = dt.date.today()
    mes  = args.mes  or f"{hoje.month:02d}"
    ano  = args.ano  or str(hoje.year)
    tipo = args.tipo

    if args.pasta:
        global DOWNLOAD_DIR
        DOWNLOAD_DIR = Path(args.pasta.strip())

    # Se nenhum flag --so-*, roda tudo
    tudo          = not (args.so_ocr or args.so_digitacao or args.so_filtro)
    fazer_ocr     = tudo or args.so_ocr
    fazer_dig     = tudo or args.so_digitacao
    # --so-digitacao inclui filtro (move PDFs digitados para Digitadas após digitação)
    fazer_filtro  = tudo or args.so_filtro or args.so_digitacao

    preservar_auditoria = args.preservar_auditoria or os.environ.get(
        "PIPELINE_PRESERVAR_AUDITORIA", ""
    ).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}

    global OCR_SAIDA_DIR, PIPELINE_SAIDA, DIGITADAS_DIR
    OCR_SAIDA_DIR = _resolver_dir(OCR_SAIDA_DIR, FALLBACK_ROOT / "ocr_saida", "Saída OCR ENEL")
    PIPELINE_SAIDA = _resolver_dir(PIPELINE_SAIDA, FALLBACK_ROOT / "saida", "Saída do pipeline ENEL")
    DIGITADAS_DIR = _resolver_dir(DIGITADAS_DIR, FALLBACK_ROOT / "digitadas", "Pasta Digitadas ENEL")

    falhou = False
    _banner(f"PIPELINE ENEL  {tipo.upper()}  {mes}/{ano}", [
        f"Download : {DOWNLOAD_DIR}",
        f"OCR      : {OCR_SAIDA_DIR}",
        f"Saída    : {PIPELINE_SAIDA}",
    ])

    tipos = []
    if tipo in ("bt", "ambos"):
        tipos.append("bt")
    if tipo in ("mt", "ambos"):
        tipos.append("mt")

    for t in tipos:
        xlsx   = _xlsx_bt(mes, ano) if t == "bt" else _xlsx_mt(mes, ano)
        pdfs   = _pasta_download_bt(mes, ano) if t == "bt" else _pasta_download_mt(mes, ano)
        xlsx_digitacao = xlsx
        pdfs_digitacao = pdfs
        digitacao_falhou = False

        # -- 1. OCR ----------------------------------------------------------
        if fazer_ocr:
            cod = etapa_ocr(mes, ano, t, recriar=getattr(args, "recriar", False))
            if cod != 0:
                _fail(f"OCR {t.upper()} falhou  abortando pipeline {t.upper()}", cod)
                falhou = True
                continue

        # -- 2. Digitação ----------------------------------------------------
        if fazer_dig or fazer_filtro:
            xlsx_digitacao, pdfs_digitacao = preparar_digitacao_bb(xlsx, pdfs)

        if fazer_dig:
            if preservar_auditoria:

                _info("[resume] Preservando auditoria_resultados.csv para retomar o lote.")

            else:

                _resetar_auditoria(PIPELINE_SAIDA)
            cod = etapa_digitacao(xlsx_digitacao, pdfs_digitacao)
            if cod != 0:
                digitacao_falhou = True
                _warn(f"Digitação {t.upper()} terminou com exit {cod}  continuando para filtro")
                # Não aborta: digitação parcial ainda gera auditoria_resultados.csv

        # -- 3. Filtro -------------------------------------------------------
        # Passa a raiz DOWNLOAD_DIR para que o filtro busque recursivamente
        # em todos os subdiretórios de mês/tipo (03-2026/BT, 04-2026/MT, etc.)
        if fazer_filtro:
            cod = etapa_filtro(pdfs_digitacao, DIGITADAS_DIR)
            if cod != 0:
                _fail(f"Filtro {t.upper()} falhou", cod)
                falhou = True
            elif digitacao_falhou:
                _warn(f"Digitação {t.upper()} falhou parcialmente, mas o filtro conseguiu aproveitar a auditoria gerada.")
        elif digitacao_falhou:
            _fail(f"Digitação {t.upper()} falhou e o filtro não foi executado.")
            falhou = True

    _p()
    _sep("-")
    if falhou:
        _fail("PIPELINE ENEL FINALIZADO COM FALHAS")
    else:
        _ok("PIPELINE ENEL FINALIZADO COM SUCESSO")
    _sep("-")
    sys.exit(1 if falhou else 0)


if __name__ == "__main__":
    main()
