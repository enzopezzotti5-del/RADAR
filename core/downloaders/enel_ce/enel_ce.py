#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         ENEL CEARÁ — Downloader Selenium (portal legado ASP.NET)            ║
║         Ação Engenharia e Instalações Ltda.                                 ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Fonte de dados : senhas_enel_ce.xlsx (mesma pasta do script)               ║
║  Portal         : https://www.eneldistribuicao.com.br/ce/Corporativo.aspx   ║
║  Fluxo          :                                                            ║
║    1) Lê a planilha e agrupa UCs por (Login, Senha)                         ║
║    2) Faz login no portal legado para cada grupo de credenciais              ║
║    3) Para cada UC: seleciona na tabela → abre 2ª Via                       ║
║    4) Se todas Paga → pula; senão baixa cada fatura pendente                ║
║    5) Classifica BT / MT / NAO_IDENTIFICADA via pdfplumber                  ║
║    6) Grava no indice_master.py (carimbo BB_XXXXXX) + CSV local             ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
import _venv_check  # noqa

import csv
import os
import re
import shutil
import sys
import tempfile
import time
import importlib.util as _ilu
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pdfplumber

from core.metrics.radar_metrics import (
    emit_downloaded,
    emit_item_error,
    emit_progress,
    emit_skipped_existing,
)

import subprocess
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, ElementClickInterceptedException,
    StaleElementReferenceException,
)


# ══════════════════════════════════════════════════════════════════════════════
# ÍNDICE MASTER
# ══════════════════════════════════════════════════════════════════════════════

_MASTER_LOCAL  = Path(__file__).resolve().parent.parent.parent / "indice_master.py"

_master_mod_path = next((p for p in [_MASTER_LOCAL] if p.exists()), None)

if _master_mod_path:
    print(f"[master] Carregando: {_master_mod_path}")
    _spec = _ilu.spec_from_file_location("indice_master", str(_master_mod_path))
    _mod  = _ilu.module_from_spec(_spec)
    _spec.loader.exec_module(_mod)
    MasterIndice    = _mod.MasterIndice
    MASTER_FILE     = _mod.MASTER_FILE
    _normalizar_ref = _mod.normalizar_mes_ref
    _chave_dedup    = _mod.chave_dedup
    _FILELOCK_OK    = getattr(_mod, "_FILELOCK_OK", False)
    _usar_master    = True
    if not _FILELOCK_OK:
        print("[master] AVISO: filelock não instalado — pip install filelock")
else:
    MasterIndice    = None
    MASTER_FILE     = None
    _normalizar_ref = lambda r: r
    _chave_dedup    = lambda uc, ref: f"{str(uc).strip().lstrip('0') or '0'}|{ref}"
    _usar_master    = False
    print("[master] indice_master.py não encontrado — modo local")


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

URL_PUBLICA     = "https://www.enel.com.br/pt-ceara.html"
URL_CORPORATIVA = "https://www.eneldistribuicao.com.br/ce/Corporativo.aspx"

ROOT_DIR_DEFAULT = "//10.10.250.21/Energia/ARQUIVOS ENZO/DOWNLOAD ENEL CE"
PLANILHA_DEFAULT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "senhas_enel_ce.xlsx")

TIMEOUT_PADRAO   = 25    # WebDriverWait padrão (s)
TIMEOUT_LOGIN    = 35    # espera tela de seleção após login (s)
TIMEOUT_DOWNLOAD = 60    # espera PDF na pasta temp (s)
PAUSA_POSTBACK   = 2.5   # após click ASP.NET (__doPostBack)
PAUSA_NAVEGACAO  = 2.0   # após driver.get()
PAUSA_ENTRE_DL   = 1.5   # entre downloads na mesma UC

INDEX_HEADERS = [
    "INDICE", "UC", "CLASSIFICACAO", "MES_REF",
    "FATURA_ID", "DATA_DOWNLOAD", "ARQUIVO",
    "LOGIN", "CNPJ", "INSTALACAO",
]


# ══════════════════════════════════════════════════════════════════════════════
# DATACLASS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class UnidadeAcesso:
    uc: str
    instalacao: str
    cnpj: str
    tensao: str
    login: str
    senha: str
    prefixo:         str = ""
    medidor:         str = ""
    urgente:         str = ""
    data_vencimento: str = ""
    data_emissao:    str = ""


# ══════════════════════════════════════════════════════════════════════════════
# LOG RICO COM CORES ANSI
# ══════════════════════════════════════════════════════════════════════════════

_TERM = sys.stdout.isatty()

def _c(code: str) -> str:
    return code if _TERM else ""

_COR = {
    "RST": _c("\033[0m"),
    "CZ":  _c("\033[90m"),   # cinza
    "VD":  _c("\033[92m"),   # verde
    "AM":  _c("\033[93m"),   # amarelo
    "VM":  _c("\033[91m"),   # vermelho
    "CI":  _c("\033[96m"),   # ciano
    "MG":  _c("\033[95m"),   # magenta
    "BR":  _c("\033[97m"),   # branco brilhante
    "AZ":  _c("\033[94m"),   # azul
}

_NIVEL: Dict[str, Tuple[str, str]] = {
    #           símbolo   cor
    "INFO":    ("→",  _COR["BR"]),
    "SUCCESS": ("✓",  _COR["VD"]),
    "ERROR":   ("✗",  _COR["VM"]),
    "WARNING": ("⚠",  _COR["AM"]),
    "SKIP":    ("⏭", _COR["CZ"]),
    "PROG":    ("📊", _COR["CI"]),
    "DUP":     ("🔁", _COR["MG"]),
    "ALERT":   ("🚨", _COR["VM"]),
    "STEP":    ("▶",  _COR["CI"]),
    "WAIT":    ("⏳", _COR["CZ"]),
}


def _log(msg: str, level: str = "INFO", arquivo_log: Optional[str] = None):
    sim, cor = _NIVEL.get(level, ("•", ""))
    ts       = datetime.now().strftime("%H:%M:%S")
    linha    = f"{_COR['CZ']}[{ts}]{_COR['RST']} {cor}{sim} {msg}{_COR['RST']}"
    print(linha, flush=True)
    if arquivo_log:
        novo = not os.path.exists(arquivo_log)
        try:
            with open(arquivo_log, "a", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                if novo:
                    w.writerow(["DATA", "NIVEL", "MSG"])
                w.writerow([datetime.now().strftime("%d/%m/%Y %H:%M:%S"), level, msg])
        except Exception:
            pass


def _sep(char: str = "─", n: int = 78):
    print(_COR["CZ"] + char * n + _COR["RST"], flush=True)


def _banner(titulo: str):
    _sep("═")
    print(f"{_COR['CI']}  {titulo}{_COR['RST']}", flush=True)
    _sep("═")


# ══════════════════════════════════════════════════════════════════════════════
# DOWNLOADER
# ══════════════════════════════════════════════════════════════════════════════

class EnelCELegacyPlanilhaDownloader:

    def __init__(
        self,
        planilha_path: str,
        root_dir: str  = ROOT_DIR_DEFAULT,
        headless: bool = False,
        timeout: int   = TIMEOUT_PADRAO,
    ):
        self.planilha_path = planilha_path
        self.timeout       = timeout
        self.script_dir    = os.path.dirname(os.path.abspath(__file__))

        # ── Pasta temporária de download ──────────────────────────────────────
        self.temp_dir = os.path.join(self.script_dir, "downloads_temp_enel_ce_legacy")
        os.makedirs(self.temp_dir, exist_ok=True)

        # ── Pasta de saída ────────────────────────────────────────────────────
        if str(root_dir).endswith("DOWNLOAD ENEL CE"):
            self.output_base_dir = root_dir
        else:
            self.output_base_dir = os.path.join(root_dir, "DOWNLOAD ENEL CE")
        try:
            os.makedirs(self.output_base_dir, exist_ok=True)
        except Exception as e:
            print(f"[init] ⚠ Não foi possível criar '{self.output_base_dir}': {e}")
            self.output_base_dir = os.path.join(self.script_dir, "DOWNLOAD ENEL CE")
            os.makedirs(self.output_base_dir, exist_ok=True)
            print(f"[init] → Usando fallback local: {self.output_base_dir}")

        # ── Arquivos de log / índice ──────────────────────────────────────────
        self.index_file               = os.path.join(self.output_base_dir, "indice_faturas.csv")
        self.log_pulos_file           = os.path.join(self.output_base_dir, "log_ucs_puladas_todas_pagas.csv")
        self.log_execucao_file        = os.path.join(self.output_base_dir, "log_execucao_legacy.csv")
        self.log_nao_encontradas_file = os.path.join(self.output_base_dir, "log_ucs_nao_encontradas_no_portal.csv")
        self.log_sem_link_file        = os.path.join(self.output_base_dir, "log_sem_link_download.csv")

        # ── Master index ──────────────────────────────────────────────────────
        try:
            self._master = MasterIndice(MASTER_FILE) if (_usar_master and MASTER_FILE is not None) else None
        except Exception as e:
            print(f"[master] Falha ao instanciar MasterIndice: {e} — continuando sem master.")
            self._master = None

        # ── Contadores e memória ──────────────────────────────────────────────
        self.indice_fatura             = 2000000
        self.memoria_download:   set   = set()
        self.faturas_baixadas:   set   = set()
        self.faturas_do_indice:  set   = set()
        self.meses_por_uc              = defaultdict(set)

        self.qtd_baixadas_hoje       = 0
        self.qtd_puladas_todas_pagas = 0
        self.qtd_sem_tabela          = 0
        self.qtd_sem_link_download   = 0
        self.qtd_ucs_nao_encontradas = 0
        self.qtd_erros               = 0

        self._carregar_indice()

        # ── Driver Selenium ───────────────────────────────────────────────────
        self.driver         = self._criar_driver(headless=headless)
        self.wait           = WebDriverWait(self.driver, self.timeout)
        self.selection_url: Optional[str] = None

    # ──────────────────────────────────────────────────────────────────────────
    # LOG
    # ──────────────────────────────────────────────────────────────────────────

    def log(self, msg: str, level: str = "INFO"):
        _log(msg, level, self.log_execucao_file)

    # ──────────────────────────────────────────────────────────────────────────
    # DRIVER
    # ──────────────────────────────────────────────────────────────────────────

    def _criar_driver(self, headless: bool = False) -> webdriver.Chrome:
        opts = Options()
        profile_root = Path(self.script_dir) / "chrome_profiles"
        profile_root.mkdir(parents=True, exist_ok=True)
        profile_dir = Path(tempfile.mkdtemp(prefix="enel_ce_", dir=str(profile_root)))
        if headless:
            opts.add_argument("--headless=new")

        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--disable-infobars")
        opts.add_argument(f"--user-data-dir={profile_dir.resolve()}")
        opts.add_argument("--remote-debugging-port=0")
        opts.add_argument("--no-first-run")
        opts.add_argument("--no-default-browser-check")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_experimental_option("prefs", {
            "download.default_directory":         os.path.abspath(self.temp_dir),
            "download.prompt_for_download":       False,
            "download.directory_upgrade":         True,
            "plugins.always_open_pdf_externally": True,
            "safebrowsing.enabled":               True,
        })

        service = Service()
        service.creationflags = subprocess.CREATE_NEW_PROCESS_GROUP
        driver = webdriver.Chrome(service=service, options=opts)
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
        )
        driver.set_page_load_timeout(90)
        driver._profile_dir = profile_dir
        _log("Chrome iniciado (modo visível, janela maximizada).", "STEP")
        return driver

    # ──────────────────────────────────────────────────────────────────────────
    # PLANILHA
    # ──────────────────────────────────────────────────────────────────────────

    def _fmt_data(self, val) -> str:
        if val is None:
            return ""
        return val.strftime("%d/%m/%Y") if hasattr(val, "strftime") else str(val).strip()

    def _so_digitos(self, s) -> str:
        return re.sub(r"\D+", "", str(s or ""))

    def carregar_unidades_da_planilha(self) -> List[UnidadeAcesso]:
        self.log(f"Lendo planilha: {self.planilha_path}", "STEP")
        wb = load_workbook(self.planilha_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        idx    = {nome: i for i, nome in enumerate(header)}

        obrig = ["Concessionária", "Conta Contrato", "Instalacao", "Login", "Senha"]
        faltantes = [c for c in obrig if c not in idx]
        if faltantes:
            raise ValueError(f"Colunas obrigatórias ausentes na planilha: {faltantes}")

        unidades: List[UnidadeAcesso] = []
        vistos: set = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            conc = str(row[idx["Concessionária"]] or "").strip().upper()
            if "ENEL CEARA" not in conc and "COELCE" not in conc:
                continue

            login  = str(row[idx["Login"]]  or "").strip()
            senha  = str(row[idx["Senha"]]  or "").strip()
            cc     = self._so_digitos(row[idx["Conta Contrato"]])
            inst   = self._so_digitos(row[idx["Instalacao"]])
            cnpj   = self._so_digitos(row[idx["CNPJ"]])       if "CNPJ"   in idx else ""
            tensao = str(row[idx["Tensão"]] or "").strip()    if "Tensão" in idx else ""

            uc = cc or inst
            if not (uc and login and senha):
                continue

            chave = (login.lower(), senha, uc)
            if chave in vistos:
                continue
            vistos.add(chave)

            unidades.append(UnidadeAcesso(
                uc=uc, instalacao=inst, cnpj=cnpj, tensao=tensao,
                login=login, senha=senha,
                prefixo         = str(row[idx["Prefixo"]]       or "").strip() if "Prefixo"        in idx else "",
                medidor         = str(row[idx["Medidor"]]       or "").strip() if "Medidor"        in idx else "",
                urgente         = str(row[idx["Urgente"]]       or "").strip() if "Urgente"        in idx else "",
                data_vencimento = self._fmt_data(row[idx["DataVencimento"]]) if "DataVencimento" in idx else "",
                data_emissao    = self._fmt_data(row[idx["DataEmissão"]])    if "DataEmissão"    in idx else "",
            ))

        self.log(f"Planilha carregada: {len(unidades)} UCs ENEL CE válidas.", "SUCCESS")
        return unidades

    def agrupar_por_credenciais(
        self, unidades: List[UnidadeAcesso]
    ) -> Dict[Tuple[str, str], List[UnidadeAcesso]]:
        grupos: Dict[Tuple[str, str], List[UnidadeAcesso]] = defaultdict(list)
        for u in unidades:
            grupos[(u.login, u.senha)].append(u)
        return grupos

    # ──────────────────────────────────────────────────────────────────────────
    # CLASSIFICAÇÃO PDF
    # ──────────────────────────────────────────────────────────────────────────

    def _classificar_pdf(self, pdf_path: str, tensao_planilha: str = "") -> str:
        texto = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages[:2]:
                    texto += (page.extract_text() or "")
        except Exception:
            pass

        total = f"{texto}\n{(tensao_planilha or '').upper()}".upper()
        if not total.strip():
            return "NAO_IDENTIFICADA"

        padroes_mt = [
            r"\bA4\b", r"\bA3\b", r"\bA3A\b", r"\bAS\b",
            r"M[ÉE]DIA\s*TENS[ÃA]O", r"GRUPO\s*A",
            r"SUBGRUPO\s*A", r"TARIFA\s*A", r"M[ÉE]DIATENS[ÃA]O",
        ]
        padroes_bt = [
            r"\bB1\b", r"\bB2\b", r"\bB3\b", r"\bB4\b",
            r"BAIXA\s*TENS[ÃA]O", r"GRUPO\s*B",
            r"SUBGRUPO\s*B", r"TARIFA\s*B",
        ]
        for p in padroes_mt:
            if re.search(p, total, re.IGNORECASE):
                return "MT"
        for p in padroes_bt:
            if re.search(p, total, re.IGNORECASE):
                return "BT"
        return "NAO_IDENTIFICADA"

    # ──────────────────────────────────────────────────────────────────────────
    # ÍNDICE LOCAL
    # ──────────────────────────────────────────────────────────────────────────

    def _carregar_indice(self):
        if self._master is not None:
            for chave in self._master._ja_baixados:
                self.memoria_download.add(chave)
            self.indice_fatura = self._master._proximo_num
            _log(
                f"Master carregado: {len(self._master._ja_baixados)} registros "
                f"| próximo: {self._master.proximo_carimbo}",
                "SUCCESS",
            )

        if os.path.exists(self.index_file):
            with open(self.index_file, encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    uc  = (row.get("UC")        or "").strip()
                    ref = (row.get("MES_REF")   or "").strip()
                    fid = (row.get("FATURA_ID") or "").strip()
                    bb  = (row.get("INDICE")    or "").strip()
                    if uc and ref:
                        self.memoria_download.add(_chave_dedup(uc, ref))
                        self.meses_por_uc[uc].add(ref)
                    if fid:
                        self.faturas_do_indice.add(fid)
                    if bb.startswith("BB_"):
                        try:
                            n = int(bb.replace("BB_", ""))
                            if n >= self.indice_fatura:
                                self.indice_fatura = n + 1
                        except ValueError:
                            pass
            _log(f"Índice CSV local: {len(self.memoria_download)} entradas.", "SUCCESS")
        else:
            with open(self.index_file, "w", encoding="utf-8-sig", newline="") as f:
                csv.writer(f).writerow(INDEX_HEADERS)
            _log("Índice CSV criado do zero.", "INFO")

        # Demais arquivos de log de apoio
        for path, cols in [
            (self.log_pulos_file,           ["UC", "DATA", "MOTIVO"]),
            (self.log_nao_encontradas_file,  ["UC", "LOGIN", "DATA", "MOTIVO"]),
            (self.log_sem_link_file,         ["UC", "REF", "DATA", "SITUACAO"]),
        ]:
            if not os.path.exists(path):
                with open(path, "w", encoding="utf-8-sig", newline="") as f:
                    csv.writer(f).writerow(cols)

    def _registrar_no_indice(
        self,
        indice_bb: str,
        unidade: UnidadeAcesso,
        classificacao: str,
        ref: str,
        fatura_id: str,
        arquivo: str = "",
    ):
        with open(self.index_file, "a", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerow([
                indice_bb, unidade.uc, classificacao, ref, fatura_id,
                datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                arquivo, unidade.login, unidade.cnpj, unidade.instalacao,
            ])

        if self._master is not None:
            try:
                self._master.registrar(
                    indice_bb=indice_bb, sistema="ENEL_CE",
                    uc=unidade.uc, mes_ref=_normalizar_ref(ref),
                    fatura_id=fatura_id, cnpj=unidade.cnpj,
                    estado="CEARÁ", instalacao=unidade.instalacao,
                    arquivo=arquivo,
                )
            except Exception as e2:
                print(f"[master] Falha ao registrar {indice_bb}: {e2}")
            except Exception as e:
                print(f"[master] Falha ao registrar {indice_bb}: {e}")

        self.memoria_download.add(_chave_dedup(unidade.uc, ref))
        self.meses_por_uc[unidade.uc].add(ref)
        self.faturas_baixadas.add(fatura_id)

    def _registrar_uc_pulada(self, unidade: UnidadeAcesso, motivo: str):
        with open(self.log_pulos_file, "a", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerow([
                unidade.uc, datetime.now().strftime("%d/%m/%Y %H:%M:%S"), motivo,
            ])
        self.qtd_puladas_todas_pagas += 1

    def _registrar_uc_nao_encontrada(self, unidade: UnidadeAcesso):
        with open(self.log_nao_encontradas_file, "a", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerow([
                unidade.uc, unidade.login,
                datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "UC não apareceu na tabela de agrupamento do login",
            ])
        self.qtd_ucs_nao_encontradas += 1

    def _registrar_sem_link(self, unidade: UnidadeAcesso, ref: str, situacao: str):
        with open(self.log_sem_link_file, "a", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerow([
                unidade.uc, ref,
                datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                situacao,
            ])
        self.qtd_sem_link_download += 1

    # ──────────────────────────────────────────────────────────────────────────
    # HELPERS SELENIUM
    # ──────────────────────────────────────────────────────────────────────────

    def _safe_click(self, by: By, selector: str, timeout: Optional[int] = None):
        w  = WebDriverWait(self.driver, timeout or self.timeout)
        el = w.until(EC.element_to_be_clickable((by, selector)))
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.3)
        try:
            el.click()
        except (ElementClickInterceptedException, StaleElementReferenceException):
            self.driver.execute_script("arguments[0].click();", el)
        return el

    def _safe_send(self, by: By, selector: str, value: str, clear: bool = True):
        el = self.wait.until(EC.presence_of_element_located((by, selector)))
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        if clear:
            el.clear()
            time.sleep(0.15)
            el.send_keys(Keys.CONTROL, "a")
            el.send_keys(Keys.DELETE)
        el.send_keys(value)
        return el

    def _esperar_texto(self, trecho: str, timeout: int = 15) -> bool:
        try:
            WebDriverWait(self.driver, timeout).until(
                lambda d: trecho.lower() in (d.page_source or "").lower()
            )
            return True
        except TimeoutException:
            return False

    def _limpar_temp(self):
        for p in Path(self.temp_dir).glob("*"):
            try:
                p.unlink() if p.is_file() else shutil.rmtree(p, ignore_errors=True)
            except Exception:
                pass

    def _esperar_download_pdf(self, timeout: int = TIMEOUT_DOWNLOAD) -> Optional[str]:
        """
        Aguarda um novo PDF aparecer em self.temp_dir.
        Registra snapshot inicial para detectar somente arquivos novos
        ou modificados após o click de download.
        """
        vistos = {str(p): p.stat().st_mtime for p in Path(self.temp_dir).glob("*.pdf")}
        inicio = time.time()
        ultimo_log = -1

        while time.time() - inicio < timeout:
            em_andamento = list(Path(self.temp_dir).glob("*.crdownload"))
            pdfs         = list(Path(self.temp_dir).glob("*.pdf"))

            for p in pdfs:
                sp = str(p)
                if sp not in vistos or p.stat().st_mtime > vistos[sp]:
                    if not em_andamento:
                        time.sleep(0.8)   # garante que o SO fechou o arquivo
                        return sp

            if not em_andamento and pdfs:
                pdfs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                return str(pdfs[0])

            decorrido = int(time.time() - inicio)
            if decorrido % 10 == 0 and decorrido != ultimo_log and decorrido > 0:
                self.log(f"    ⏳ Aguardando download... {decorrido}s/{timeout}s", "WAIT")
                ultimo_log = decorrido

            time.sleep(1)

        return None

    # ──────────────────────────────────────────────────────────────────────────
    # COOKIES
    # ──────────────────────────────────────────────────────────────────────────

    def _aceitar_cookies(self):
        seletores = [
            (By.ID,    "truste-consent-button"),
            (By.XPATH, "//button[@id='truste-consent-button']"),
            (By.XPATH, "//button[contains(normalize-space(.),'Aceitar tudo')]"),
            (By.XPATH, "//button[contains(normalize-space(.),'Aceitar')]"),
        ]
        for by, sel in seletores:
            try:
                btn = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((by, sel)))
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(0.3)
                try:
                    btn.click()
                except Exception:
                    self.driver.execute_script("arguments[0].click();", btn)
                self.log("Banner de cookies aceito.", "INFO")
                time.sleep(1.0)
                return
            except Exception:
                continue

    # ──────────────────────────────────────────────────────────────────────────
    # LOGIN
    # ──────────────────────────────────────────────────────────────────────────

    def _abrir_portal(self):
        self.log(f"Abrindo: {URL_CORPORATIVA}", "STEP")
        self.driver.get(URL_CORPORATIVA)
        time.sleep(PAUSA_NAVEGACAO)
        self._aceitar_cookies()

        try:
            self.wait.until(EC.presence_of_element_located(
                (By.ID, "WEBDOOR_headercorporativo_UserName")
            ))
            self.log("Formulário de login pronto.", "INFO")
            return
        except TimeoutException:
            pass

        # Fallback: fluxo pelo site público
        self.log("Formulário não apareceu direto — tentando via site público...", "WARNING")
        self.driver.get(URL_PUBLICA)
        time.sleep(PAUSA_NAVEGACAO)
        self._aceitar_cookies()
        try:
            self._safe_click(
                By.XPATH,
                "//a[contains(., 'NEGÓCIOS E GOVERNO') or contains(., 'NEGOCIOS E GOVERNO')]",
                timeout=8,
            )
            time.sleep(2)
            self._safe_click(
                By.XPATH,
                "//a[contains(@href,'Corporativo.aspx') and "
                "   (.//span[contains(.,'Segunda Via')] or contains(.,'Segunda Via'))]",
                timeout=8,
            )
            time.sleep(2)
            self._aceitar_cookies()
        except Exception:
            self.log("Fluxo público falhou — forçando URL corporativa.", "WARNING")
            self.driver.get(URL_CORPORATIVA)
            time.sleep(PAUSA_NAVEGACAO)
            self._aceitar_cookies()

    def fazer_login(self, email: str, senha: str):
        _sep()
        self.log(f"LOGIN  →  {email}", "STEP")
        _sep()
        self._abrir_portal()

        self.log("Preenchendo e-mail...", "INFO")
        self._safe_send(By.ID, "WEBDOOR_headercorporativo_UserName", email)
        time.sleep(0.3)

        self.log("Preenchendo senha...", "INFO")
        self._safe_send(By.ID, "WEBDOOR_headercorporativo_Password", senha)
        time.sleep(0.3)

        btn_candidatos = [
            (By.ID,    "WEBDOOR_headercorporativo_Ok"),
            (By.XPATH, "//input[@id='WEBDOOR_headercorporativo_Ok']"),
            (By.XPATH, "//input[@type='submit' and contains(@value,'Entrar')]"),
            (By.XPATH, "//button[contains(.,'Entrar') or contains(.,'Acessar')]"),
            (By.XPATH, "//a[contains(.,'Entrar') or contains(.,'Acessar')]"),
        ]
        clicou = False
        for by, sel in btn_candidatos:
            try:
                self._safe_click(by, sel, timeout=4)
                clicou = True
                self.log("Botão de login clicado.", "INFO")
                break
            except Exception:
                continue

        if not clicou:
            self.log("Botão não encontrado — enviando ENTER no campo senha.", "WARNING")
            self.driver.find_element(
                By.ID, "WEBDOOR_headercorporativo_Password"
            ).send_keys(Keys.ENTER)

        self.log("Aguardando tela de seleção de cliente...", "WAIT")
        ok = self._esperar_texto("selecione o número de cliente", timeout=TIMEOUT_LOGIN)
        if not ok:
            self.wait.until(EC.presence_of_element_located(
                (By.ID, "CONTENT_gdEscolherClienteDoAgrupamento")
            ))

        self.selection_url = self.driver.current_url
        self.log(f"Login OK — seleção carregada.  URL: {self.selection_url}", "SUCCESS")

    # ──────────────────────────────────────────────────────────────────────────
    # NAVEGAÇÃO ENTRE UCs
    # ──────────────────────────────────────────────────────────────────────────

    def voltar_para_tela_de_selecao(self):
        if self._esperar_texto("selecione o número de cliente", timeout=2):
            return

        self.log("Voltando para a tela de seleção...", "WAIT")

        for tentativa in range(1, 4):
            try:
                self.driver.back()
                time.sleep(1.5)
                if self._esperar_texto("selecione o número de cliente", timeout=5):
                    self.log(f"Retorno via back() OK (tentativa {tentativa}).", "INFO")
                    return
            except Exception:
                pass

        if self.selection_url:
            self.log("Reabrindo URL de seleção salva...", "WARNING")
            self.driver.get(self.selection_url)
            if self._esperar_texto("selecione o número de cliente", timeout=15):
                return

        self.log("Reabrindo URL corporativa base...", "WARNING")
        self.driver.get(URL_CORPORATIVA)
        if not self._esperar_texto("selecione o número de cliente", timeout=20):
            raise RuntimeError("Não foi possível voltar para a tela de seleção de cliente.")

    # ──────────────────────────────────────────────────────────────────────────
    # TABELA DE AGRUPAMENTO
    # ──────────────────────────────────────────────────────────────────────────

    def _mapear_ucs_da_tabela(self) -> Dict[str, str]:
        """
        Retorna {numero_uc: id_do_checkbox}.
        Também preenche self._ucs_ja_selecionadas com os números de UC
        cujo checkbox vem checked+disabled (UC Principal única — portal já selecionou,
        não precisa e não aceita novo clique).
        """
        self.wait.until(EC.presence_of_element_located(
            (By.ID, "CONTENT_gdEscolherClienteDoAgrupamento")
        ))
        soup   = BeautifulSoup(self.driver.page_source, "html.parser")
        tabela = soup.find("table", id="CONTENT_gdEscolherClienteDoAgrupamento")
        mapa: Dict[str, str] = {}
        self._ucs_ja_selecionadas: set = set()
        if not tabela:
            return mapa

        for tr in tabela.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 4:
                continue
            uc_num = re.sub(r"\D+", "", tds[0].get_text(" ", strip=True))
            chk    = tr.find("input", {"type": "checkbox"})
            if uc_num and chk and chk.get("id"):
                mapa[uc_num] = chk["id"]
                # checked + disabled → UC Principal já selecionada pelo portal
                if chk.get("checked") and chk.get("disabled"):
                    self._ucs_ja_selecionadas.add(uc_num)

        return mapa

    def selecionar_uc(self, unidade: UnidadeAcesso) -> bool:
        self.voltar_para_tela_de_selecao()
        mapa = self._mapear_ucs_da_tabela()   # também popula _ucs_ja_selecionadas

        self.log(
            f"Tabela de agrupamento: {len(mapa)} UCs visíveis "
            f"| Buscando UC={unidade.uc} / Instalação={unidade.instalacao}",
            "INFO",
        )

        chk_id = mapa.get(unidade.uc) or mapa.get(unidade.instalacao)
        uc_key = unidade.uc if unidade.uc in mapa else unidade.instalacao

        if not chk_id:
            self.log(
                f"UC {unidade.uc} NÃO encontrada na tabela do login {unidade.login}. "
                f"UCs disponíveis: {sorted(mapa.keys())[:15]}",
                "WARNING",
            )
            self._registrar_uc_nao_encontrada(unidade)
            return False

        # ── UC Principal única: checkbox já vem checked+disabled pelo portal ──
        ja_sel = getattr(self, "_ucs_ja_selecionadas", set())
        if uc_key in ja_sel:
            self.log(
                f"UC {unidade.uc} é UC Principal (única) — já selecionada pelo portal. "
                "Indo direto para 2ª Via.",
                "INFO",
            )
            return True

        # ── Caso normal: clica no checkbox e aguarda postback ─────────────────
        self.log(f"Checkbox #{chk_id} — clicando para selecionar a UC...", "STEP")
        html_antes = self.driver.page_source
        self._safe_click(By.ID, chk_id)
        time.sleep(PAUSA_POSTBACK)

        try:
            WebDriverWait(self.driver, 15).until(lambda d: d.page_source != html_antes)
            self.log("Postback ASP.NET concluído — UC selecionada.", "INFO")
        except TimeoutException:
            self.log("Postback não detectado (página pode não ter mudado).", "WARNING")

        return True

    # ──────────────────────────────────────────────────────────────────────────
    # 2ª VIA
    # ──────────────────────────────────────────────────────────────────────────

    def abrir_2via(self):
        self.log("Clicando em '2ª Via'...", "STEP")
        candidatos = [
            "//a[contains(.,'2ª Via') or contains(.,'2a Via') or contains(@href,'SegundaViaGa.aspx')]",
            "//a[contains(@href,'SegundaVia') or contains(@href,'segunda-via')]",
        ]
        for xp in candidatos:
            try:
                self._safe_click(By.XPATH, xp, timeout=10)
                time.sleep(PAUSA_NAVEGACAO)
                self.log("Tela de 2ª Via aberta.", "INFO")
                return
            except Exception:
                continue
        raise RuntimeError("Link '2ª Via' não encontrado — verifique os seletores em abrir_2via().")

    # ──────────────────────────────────────────────────────────────────────────
    # PARSE DA TABELA DE FATURAS
    # ──────────────────────────────────────────────────────────────────────────

    def _obter_tabela_faturas(self):
        soup = BeautifulSoup(self.driver.page_source, "html.parser")
        for tabela in soup.find_all("table"):
            hdrs = [th.get_text(" ", strip=True).lower() for th in tabela.find_all("th")]
            if any("situação" in h or "situacao" in h for h in hdrs):
                return tabela
        return None

    def _parse_faturas(self) -> List[Dict[str, str]]:
        tabela = self._obter_tabela_faturas()
        if not tabela:
            return []

        headers      = [th.get_text(" ", strip=True) for th in tabela.find_all("th")]
        headers_norm = [h.lower() for h in headers]
        rows: List[Dict[str, str]] = []

        for tr in tabela.find_all("tr")[1:]:
            tds = tr.find_all("td")
            if not tds:
                continue
            vals = [td.get_text(" ", strip=True) for td in tds]
            row  = {headers[i]: vals[i] if i < len(vals) else "" for i in range(len(headers))}
            row["_row_text"]      = " | ".join(vals)
            row["_row_index_dom"] = str(len(rows) + 1)

            # Situação
            situacao = ""
            for i, h in enumerate(headers_norm):
                if "situação" in h or "situacao" in h:
                    situacao = vals[i] if i < len(vals) else ""
                    break
            row["_situacao"] = situacao.strip()

            # Referência (mês/ano)
            ref = "SEM_REF"
            for i, h in enumerate(headers_norm):
                if any(k in h for k in ["refer", "mês", "mes", "compet", "venc"]):
                    txt = vals[i] if i < len(vals) else ""
                    m   = re.search(r"(\d{2}/\d{4}|\d{2}-\d{4}|\d{6})", txt)
                    if m:
                        ref = m.group(1).replace("-", "/")
                    elif txt:
                        ref = txt
                    break
            row["_ref"]       = ref
            row["_fatura_id"] = re.sub(r"\s+", "_", row["_row_text"])[:180]
            rows.append(row)

        return rows

    def _todas_pagas(self, rows: List[Dict]) -> bool:
        sits = [r["_situacao"].strip().lower() for r in rows if r["_situacao"].strip()]
        return bool(sits) and all(s == "paga" for s in sits)

    def _nao_pagas(self, rows: List[Dict]) -> List[Dict]:
        return [r for r in rows if r["_situacao"].strip().lower() != "paga"]

    # ──────────────────────────────────────────────────────────────────────────
    # FLUXO DE DOWNLOAD: checkbox por linha + botão "Imprimir"
    # ──────────────────────────────────────────────────────────────────────────
    #
    # Estrutura do portal:
    #   <table id="gdHistoricoDeFaturamento"> (ou tabela com th "Situação")
    #     <tr>  <td>02/2026</td> ... <td>A Pagar</td>
    #           <td><input id="...chkSegundaVia_0" type="checkbox"></td>
    #     </tr>
    #     ...
    #   </table>
    #   <input id="CONTENT_btnEmitirSegundaVia" value="Imprimir" type="submit">
    #
    # Para cada fatura não paga:
    #   1) Garante que nenhum checkbox esteja marcado (estado limpo)
    #   2) Marca o checkbox da linha correspondente
    #   3) Clica em "Imprimir" — o portal gera e baixa o PDF
    #   4) Aguarda o PDF aparecer na pasta temp
    #   5) Classifica, move e registra
    # ──────────────────────────────────────────────────────────────────────────

    # IDs reais do portal (ajuste aqui se necessário)
    _ID_TABELA_HIST  = "CONTENT_gdHistoricoDeFaturamento"
    _ID_BTN_IMPRIMIR = "CONTENT_btnEmitirSegundaVia"
    _PREFIX_CHK      = "CONTENT_gdHistoricoDeFaturamento_chkSegundaVia_"

    def _desmarcar_todos_checkboxes_2via(self):
        """Garante estado limpo — desmarca qualquer checkbox já marcado na tabela."""
        try:
            checkboxes = self.driver.find_elements(
                By.XPATH,
                f"//table[@id='{self._ID_TABELA_HIST}']//input[@type='checkbox']",
            )
            for chk in checkboxes:
                if chk.is_selected():
                    self.driver.execute_script("arguments[0].click();", chk)
                    time.sleep(0.3)
        except Exception:
            pass

    def _checkbox_id_da_linha(self, dom_row_index: int) -> Optional[str]:
        """
        Retorna o id do checkbox da linha dom_row_index (0-based).
        Tenta primeiro o padrão conhecido; se não achar, varre a tabela.
        """
        # Padrão direto (mais rápido)
        candidato = f"{self._PREFIX_CHK}{dom_row_index}"
        try:
            el = self.driver.find_element(By.ID, candidato)
            return candidato if el else None
        except NoSuchElementException:
            pass

        # Fallback: pega todos os checkboxes da tabela e escolhe pelo índice
        try:
            checkboxes = self.driver.find_elements(
                By.XPATH,
                f"//table[@id='{self._ID_TABELA_HIST}']//input[@type='checkbox']",
            )
            if dom_row_index < len(checkboxes):
                chk_id = checkboxes[dom_row_index].get_attribute("id")
                return chk_id or None
        except Exception:
            pass

        return None

    def _salvar_pdf_baixado(
        self,
        unidade: UnidadeAcesso,
        ref: str,
        fatura_id: str,
        pdf_path: str,
    ) -> Optional[str]:
        """
        Atribui carimbo, classifica, move para a pasta final e registra.
        Retorna o caminho de destino ou None em caso de erro.
        """
        if self._master is not None:
            carimbo = self._master.consumir_carimbo()
            self.indice_fatura = self._master._proximo_num
        else:
            carimbo = f"BB_{self.indice_fatura}"
            self.indice_fatura += 1

        self.log("    🔍 Classificando PDF...", "INFO")
        classe    = self._classificar_pdf(pdf_path, unidade.tensao)
        ref_pasta = ref.replace("/", "-")
        mes_dir   = os.path.join(self.output_base_dir, ref_pasta, classe)
        Path(mes_dir).mkdir(parents=True, exist_ok=True)
        destino   = os.path.join(mes_dir, f"{carimbo}.pdf")

        try:
            if os.path.abspath(pdf_path) != os.path.abspath(destino):
                shutil.move(pdf_path, destino)
        except Exception as e:
            self.log(f"    ✗  Erro ao mover PDF: {e}", "ERROR")
            return None

        self._registrar_no_indice(
            indice_bb=carimbo, unidade=unidade,
            classificacao=classe, ref=ref_pasta,
            fatura_id=fatura_id, arquivo=destino,
        )
        emit_downloaded(
            utility="ENEL CE", account_id=unidade.uc, competence=ref_pasta, invoice_id=fatura_id,
        )
        return destino

    # ──────────────────────────────────────────────────────────────────────────
    # DOWNLOAD DAS FATURAS DE UMA UC
    # ──────────────────────────────────────────────────────────────────────────

    def baixar_faturas_nao_pagas_da_tabela(self, unidade: UnidadeAcesso) -> int:
        rows = self._parse_faturas()

        if not rows:
            self.log(f"UC {unidade.uc}: tabela de faturas NÃO encontrada na página.", "WARNING")
            self.qtd_sem_tabela += 1
            return 0

        sits_resumo = [r["_situacao"] for r in rows]
        self.log(
            f"UC {unidade.uc}: {len(rows)} fatura(s) | situações: {sits_resumo}",
            "INFO",
        )

        if self._todas_pagas(rows):
            self.log(f"UC {unidade.uc}: TODAS PAGAS — pulando.", "SKIP")
            self._registrar_uc_pulada(unidade, "Todas as faturas estão Paga")
            return 0

        candidatas = self._nao_pagas(rows)
        self.log(
            f"UC {unidade.uc}: {len(candidatas)} fatura(s) pendente(s) para baixar.",
            "INFO",
        )

        # ── Filtro MT — ignora BT e NAO_IDENTIFICADA por enquanto ────────────
        # A tensão é lida da planilha (unidade.tensao). Se vier vazio,
        # tenta inferir pelo texto da própria linha da tabela.
        def _e_mt(row: Dict, tensao_planilha: str) -> bool:
            tensao = tensao_planilha.upper()
            # Planilha diz explicitamente
            if tensao:
                return any(k in tensao for k in ("MEDIA", "MÉDIA", "MT", "A4", "A3", "AS"))
            # Fallback: varre o texto da linha em busca de indicadores MT
            texto = row.get("_row_text", "").upper()
            mt_patt = [r"\bA4\b", r"\bA3\b", r"\bA3A\b", r"\bAS\b",
                       r"M[ÉE]DIA", r"GRUPO\s*A", r"SUBGRUPO\s*A"]
            bt_patt = [r"\bB1\b", r"\bB2\b", r"\bB3\b", r"\bB4\b",
                       r"BAIXA", r"GRUPO\s*B", r"SUBGRUPO\s*B"]
            for p in mt_patt:
                if re.search(p, texto, re.IGNORECASE):
                    return True
            for p in bt_patt:
                if re.search(p, texto, re.IGNORECASE):
                    return False
            return True   # sem indicador → assume MT (conservador para não perder faturas)

        mt_candidatas = [r for r in candidatas if _e_mt(r, unidade.tensao)]
        ignoradas_bt  = len(candidatas) - len(mt_candidatas)

        if ignoradas_bt:
            self.log(
                f"UC {unidade.uc}: {ignoradas_bt} fatura(s) ignorada(s) — BT (fluxo MT apenas).",
                "SKIP",
            )

        if not mt_candidatas:
            self.log(f"UC {unidade.uc}: nenhuma fatura MT pendente. Pulando.", "SKIP")
            return 0

        candidatas = mt_candidatas
        self.log(
            f"UC {unidade.uc}: {len(candidatas)} fatura(s) MT para baixar.",
            "INFO",
        )

        # Verifica se o botão "Imprimir" existe na página
        try:
            WebDriverWait(self.driver, 8).until(
                EC.presence_of_element_located((By.ID, self._ID_BTN_IMPRIMIR))
            )
        except TimeoutException:
            self.log(
                f"UC {unidade.uc}: botão '{self._ID_BTN_IMPRIMIR}' não encontrado na página. "
                "Verifique se a tela de 2ª Via carregou corretamente.",
                "ERROR",
            )
            self.qtd_erros += 1
            return 0

        baixadas = 0

        for seq, row in enumerate(candidatas, start=1):
            situacao      = row["_situacao"] or "SEM_SITUACAO"
            ref           = (row["_ref"] or "SEM_REF").replace("-", "/")
            fatura_id     = row["_fatura_id"] or f"{unidade.uc}_{ref}_{seq}"
            dom_row_index = int(row["_row_index_dom"]) - 1   # converte para 0-based

            self.log(
                f"  [{seq}/{len(candidatas)}]  ref={ref}  situação='{situacao}'",
                "INFO",
            )

            # ── Deduplicação ──────────────────────────────────────────────────
            if _chave_dedup(unidade.uc, ref) in self.memoria_download:
                self.log(f"    ⏭  ref={ref} já no índice. Pulando.", "SKIP")
                emit_skipped_existing(
                    utility="ENEL CE", account_id=unidade.uc, competence=ref, invoice_id=fatura_id,
                )
                continue
            if fatura_id in self.faturas_baixadas:
                self.log("    🔁 Fatura já baixada nesta execução. Pulando.", "DUP")
                continue

            # ── Aguarda a tabela estabilizar (ASP.NET UpdatePanel pode ter recarregado) ──
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, self._ID_TABELA_HIST))
                )
                time.sleep(0.5)
            except TimeoutException:
                self.log(f"    ✗  Tabela de faturas não encontrada antes da fatura {seq}.", "ERROR")
                self.qtd_erros += 1
                emit_item_error(utility="ENEL CE", account_id=unidade.uc, competence=ref, invoice_id=fatura_id)
                break

            # ── Estado limpo: garante que nenhum checkbox está marcado ─────────
            self._desmarcar_todos_checkboxes_2via()
            time.sleep(0.4)

            # ── Localiza o checkbox desta linha (após eventual reload da tabela) ─
            chk_id = self._checkbox_id_da_linha(dom_row_index)
            if not chk_id:
                self.log(
                    f"    ✗  Checkbox da linha {dom_row_index} não encontrado. "
                    f"Prefixo esperado: '{self._PREFIX_CHK}{dom_row_index}'",
                    "WARNING",
                )
                self._registrar_sem_link(unidade, ref, situacao)
                emit_item_error(utility="ENEL CE", account_id=unidade.uc, competence=ref, invoice_id=fatura_id)
                continue

            self.log(f"    ▶  Checkbox #{chk_id} — marcando fatura...", "STEP")

            # ── Marca SOMENTE o checkbox da fatura atual ───────────────────────
            try:
                chk_el = WebDriverWait(self.driver, 8).until(
                    EC.presence_of_element_located((By.ID, chk_id))
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", chk_el)
                time.sleep(0.3)
                if not chk_el.is_selected():
                    self.driver.execute_script("arguments[0].click();", chk_el)
                    time.sleep(0.5)
                # Verificação extra: se ainda não está marcado, tenta native click
                try:
                    if not chk_el.is_selected():
                        chk_el.click()
                        time.sleep(0.5)
                except StaleElementReferenceException:
                    # Elemento ficou stale após postback — re-localiza
                    chk_el = self.driver.find_element(By.ID, chk_id)
                    self.driver.execute_script("arguments[0].click();", chk_el)
                    time.sleep(0.5)
                self.log("    ✓  Checkbox marcado.", "INFO")
            except Exception as e:
                self.log(f"    ✗  Erro ao marcar checkbox #{chk_id}: {e}", "ERROR")
                self.qtd_erros += 1
                emit_item_error(utility="ENEL CE", account_id=unidade.uc, competence=ref, invoice_id=fatura_id)
                self._desmarcar_todos_checkboxes_2via()
                continue

            # ── Confirma que apenas 1 checkbox está marcado (evita o aviso do portal) ─
            try:
                marcados = self.driver.find_elements(
                    By.XPATH,
                    f"//table[@id='{self._ID_TABELA_HIST}']//input[@type='checkbox'][@checked]"
                )
                marcados_sel = [c for c in self.driver.find_elements(
                    By.XPATH,
                    f"//table[@id='{self._ID_TABELA_HIST}']//input[@type='checkbox']"
                ) if c.is_selected()]
                if len(marcados_sel) > 1:
                    self.log(
                        f"    ⚠  {len(marcados_sel)} checkboxes marcados — desmarcando extras.",
                        "WARNING",
                    )
                    for c in marcados_sel:
                        try:
                            if c.get_attribute("id") != chk_id and c.is_selected():
                                self.driver.execute_script("arguments[0].click();", c)
                                time.sleep(0.2)
                        except Exception:
                            pass
            except Exception:
                pass

            # ── Clica no botão "Imprimir" ─────────────────────────────────────
            self.log(f"    ▶  Clicando em 'Imprimir' ({self._ID_BTN_IMPRIMIR})...", "STEP")
            try:
                btn = WebDriverWait(self.driver, 8).until(
                    EC.element_to_be_clickable((By.ID, self._ID_BTN_IMPRIMIR))
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(0.3)
                self.driver.execute_script("arguments[0].click();", btn)
            except Exception as e:
                self.log(f"    ✗  Erro ao clicar em Imprimir: {e}", "ERROR")
                self.qtd_erros += 1
                emit_item_error(utility="ENEL CE", account_id=unidade.uc, competence=ref, invoice_id=fatura_id)
                self._desmarcar_todos_checkboxes_2via()
                continue

            # ── Aguarda o PDF aparecer na pasta temp ──────────────────────────
            self.log(f"    ⏳ Aguardando PDF ({TIMEOUT_DOWNLOAD}s máx.)...", "WAIT")
            pdf_path = self._esperar_download_pdf(timeout=TIMEOUT_DOWNLOAD)

            if not pdf_path or not os.path.exists(pdf_path):
                self.log(f"    ✗  PDF não chegou para ref={ref}.", "ERROR")
                self.qtd_erros += 1
                emit_item_error(utility="ENEL CE", account_id=unidade.uc, competence=ref, invoice_id=fatura_id)
                self._desmarcar_todos_checkboxes_2via()
                time.sleep(PAUSA_POSTBACK)
                continue

            self.log(f"    ✓  PDF recebido: {os.path.basename(pdf_path)}", "INFO")

            # ── Organiza nas pastas e registra ────────────────────────────────
            destino = self._salvar_pdf_baixado(unidade, ref, fatura_id, pdf_path)
            if not destino:
                self.qtd_erros += 1
                emit_item_error(utility="ENEL CE", account_id=unidade.uc, competence=ref, invoice_id=fatura_id)
                continue

            self.qtd_baixadas_hoje += 1
            baixadas += 1
            ref_pasta = ref.replace("/", "-")
            self.log(
                f"    ✓  SALVO  →  {ref_pasta}/{Path(destino).parent.name}/{Path(destino).name}",
                "SUCCESS",
            )

            # ── Aguarda o portal resetar antes da próxima fatura ─────────────
            time.sleep(PAUSA_ENTRE_DL)

        return baixadas

    # ──────────────────────────────────────────────────────────────────────────
    # LOOP PRINCIPAL
    # ──────────────────────────────────────────────────────────────────────────

    def processar(self):
        _banner("ENEL CEARÁ — Início do processamento")
        self._limpar_temp()

        unidades = self.carregar_unidades_da_planilha()
        if not unidades:
            self.log("Nenhuma UC ENEL CE encontrada na planilha. Encerrando.", "WARNING")
            return

        grupos     = self.agrupar_por_credenciais(unidades)
        total      = sum(len(v) for v in grupos.values())
        start_time = time.time()
        self.log(f"Total de UCs: {total}  |  Grupos de login: {len(grupos)}", "INFO")

        proc       = 0
        prox_marca = 10

        for (login, senha), ucs_do_login in grupos.items():
            _banner(f"Grupo: {login}  ({len(ucs_do_login)} UCs)")

            try:
                self.fazer_login(login, senha)
            except Exception as e:
                self.log(f"Falha no login '{login}': {e}", "ERROR")
                self.qtd_erros += 1
                continue

            for unidade in ucs_do_login:
                proc += 1
                emit_progress(uc_current=proc, uc_total=total)
                pct   = proc / total * 100 if total else 100

                if pct >= prox_marca:
                    decorrido = time.time() - start_time
                    restante  = (decorrido / proc) * (total - proc) if proc else 0
                    self.log(
                        f"PROGRESSO: {int(pct)}% ({proc}/{total}) "
                        f"| ~{int(restante // 60)}min restantes",
                        "PROG",
                    )
                    prox_marca += 10

                _sep()
                self.log(
                    f"[{proc}/{total}]  UC={unidade.uc} | Inst={unidade.instalacao} "
                    f"| Tensão={unidade.tensao} | Login={unidade.login}",
                    "STEP",
                )

                # ── Filtro de tensão: somente MT ──────────────────────────────
                tensao_upper = (unidade.tensao or "").upper()
                e_bt = any(k in tensao_upper for k in (
                    "BAIXA", "BT", "B1", "B2", "B3", "B4",
                ))
                if e_bt:
                    self.log(
                        f"⏭  Tensão '{unidade.tensao}' — BT ignorada (fluxo MT apenas).",
                        "SKIP",
                    )
                    continue
                # ─────────────────────────────────────────────────────────────

                try:
                    if not self.selecionar_uc(unidade):
                        continue
                    self.abrir_2via()
                    self.baixar_faturas_nao_pagas_da_tabela(unidade)

                except Exception as e:
                    self.log(f"ERRO na UC {unidade.uc}: {e}", "ERROR")
                    self.qtd_erros += 1

                finally:
                    try:
                        self.voltar_para_tela_de_selecao()
                    except Exception as e_volta:
                        self.log(
                            f"Falha ao voltar para seleção após UC {unidade.uc}: {e_volta} — forçando URL...",
                            "ERROR",
                        )
                        try:
                            self.driver.get(URL_CORPORATIVA)
                            time.sleep(PAUSA_NAVEGACAO)
                        except Exception:
                            pass

            # Limpa sessão antes do próximo grupo de credenciais
            try:
                self.driver.delete_all_cookies()
                self.log("Sessão encerrada — cookies limpos.", "INFO")
            except Exception:
                pass

        # ── Resumo final ──────────────────────────────────────────────────────
        duracao = int(time.time() - start_time)
        _banner("EXECUÇÃO CONCLUÍDA")
        self.log(f"Tempo total              : {duracao // 60}min {duracao % 60}s", "INFO")
        self.log(f"UCs processadas          : {proc}", "INFO")
        self.log(f"Faturas baixadas ✓       : {self.qtd_baixadas_hoje}", "SUCCESS")
        self.log(f"UCs todas pagas ⏭        : {self.qtd_puladas_todas_pagas}", "SKIP")
        self.log(f"UCs não encontradas ⚠    : {self.qtd_ucs_nao_encontradas}", "WARNING")
        self.log(f"Sem tabela de faturas ⚠  : {self.qtd_sem_tabela}", "WARNING")
        self.log(f"Sem link de download ⚠   : {self.qtd_sem_link_download}", "WARNING")
        self.log(f"Erros gerais ✗           : {self.qtd_erros}",
                 "ERROR" if self.qtd_erros else "INFO")
        ultimo = (
            f"BB_{self._master._proximo_num - 1}"
            if self._master is not None else f"BB_{self.indice_fatura - 1}"
        )
        self.log(f"Último carimbo           : {ultimo}", "INFO")
        self.log(f"Saída                    : {self.output_base_dir}", "INFO")
        _sep("═")

    def fechar(self):
        try:
            self.driver.quit()
            shutil.rmtree(getattr(self.driver, "_profile_dir", None) or "", ignore_errors=True)
            _log("Navegador fechado.", "INFO")
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ROOT_DIR   = ROOT_DIR_DEFAULT
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Localiza a planilha de senhas (primeiro na pasta do script)
    PLANILHA = next(
        (p for p in [
            os.path.join(script_dir, "senhas_enel_ce.xlsx"),
            PLANILHA_DEFAULT,
        ] if os.path.exists(p)),
        None,
    )

    if not PLANILHA:
        print(
            "❌  Planilha 'senhas_enel_ce.xlsx' não encontrada.\n"
            f"    Esperada em : {script_dir}\n"
            f"    Ou em       : {PLANILHA_DEFAULT}"
        )
        sys.exit(1)

    _banner("ENEL CEARÁ — Download Selenium (portal legado)")
    print(f"  Planilha : {PLANILHA}")
    print(f"  Saída    : {ROOT_DIR}")
    print(f"  Python   : {sys.version.split()[0]}")
    print(f"  Início   : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    _sep("═")

    bot = EnelCELegacyPlanilhaDownloader(
        planilha_path=PLANILHA,
        root_dir=ROOT_DIR,
        headless=False,          # janela visível
        timeout=TIMEOUT_PADRAO,
    )

    try:
        bot.processar()
    finally:
        bot.fechar()


if __name__ == "__main__":
    main()
