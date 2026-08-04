#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Teste Light RJ: login automático e continuação automática do fluxo.

Fluxo:
  1. Abre browser no portal
  2. Preenche login e clica Entrar
  3. Script supera o acesso rápido Banco do Brasil automaticamente
  4. Detecta o campo de UC e assume
  5. Para cada UC: busca -> accordion -> dots -> Baixar
"""

import argparse
import os
import shutil
import sys
import time
import importlib.util as _ilu
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
import _venv_check  # noqa: F401

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)

URL_PORTAL = "https://agenciavirtual.light.com.br/portal/"
LOGIN = "bbenergia@acaoenge.com.br"
SENHA = "2024*Aca"
LIGHT_MODAL_BB_ID = "AGV_Acesso_VW_wt2_block_wtListRecords2_ctl00_wt36"
LIGHT_CAMPO_UC_SELECTOR = "input[id*='wtInputBuscarUnidadeConsumidora']"

TEMP_DOWNLOAD = Path(os.environ.get("LOCALAPPDATA", r"C:\Users\Public")) / "light_test_temp"
OUT_DIR = Path(__file__).resolve().parent / "_debug_out"
FINAL_DOWNLOAD_ROOT = Path(r"\\10.10.250.21\Energia\ARQUIVOS ENZO\DOWNLOAD LIGHT")
PLANILHA_UCS = FINAL_DOWNLOAD_ROOT / "instalações light.xlsx"
ROOT_LOCAL = Path(__file__).resolve().parents[3]
MASTER_CANDIDATOS = [
    ROOT_LOCAL / "scripts" / "indice_master.py",
    ROOT_LOCAL / "indice_master.py",
    Path(__file__).resolve().parent / "indice_master.py",
]

COLUNAS_UC = ["Instalacao", "Instalação Antiga", "Instalacao Antiga"]
INICIO_LOTE = max(1, int(os.environ.get("LIGHT_INICIO", "1")))
MOTIVO_SEGUNDA_VIA = "Perda ou esquecimento"
MAPA_MESES = {
    "JAN": "01",
    "FEV": "02",
    "MAR": "03",
    "ABR": "04",
    "MAI": "05",
    "JUN": "06",
    "JUL": "07",
    "AGO": "08",
    "SET": "09",
    "OUT": "10",
    "NOV": "11",
    "DEZ": "12",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Downloader LIGHT RJ: login automatico e execucao do fluxo do portal."
    )
    parser.add_argument("--url", default=URL_PORTAL, help="URL inicial do portal LIGHT.")
    parser.add_argument(
        "--timeout-modal",
        type=int,
        default=180,
        help="Tempo maximo em segundos para superar o modal e aguardar a area da UC.",
    )
    parser.add_argument(
        "--inicio-lote",
        type=int,
        default=INICIO_LOTE,
        help="Indice inicial 1-based da UC na planilha para retomar o lote.",
    )
    parser.add_argument(
        "--preflight", action="store_true",
        help="Valida browser, login, modal e campo inicial sem carregar UCs ou baixar arquivos.",
    )
    return parser.parse_args()


def _encontrar_coluna(headers: list[str], candidatos: list[str]) -> str | None:
    mapa = {str(header or "").strip().casefold(): str(header or "").strip() for header in headers}
    for candidato in candidatos:
        achado = mapa.get(candidato.casefold())
        if achado:
            return achado
    return None


def _somente_digitos(valor) -> str:
    return "".join(ch for ch in str(valor or "") if ch.isdigit())


def carregar_ucs_planilha(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Planilha da LIGHT nao encontrada: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    coluna_uc = _encontrar_coluna(headers, COLUNAS_UC)
    if not coluna_uc:
        wb.close()
        raise ValueError(f"Nao achei nenhuma coluna de UC/instalacao em: {headers}")

    idx_uc = headers.index(coluna_uc)
    vistos: set[str] = set()
    ucs: list[str] = []
    repetidas = 0

    for row in rows:
        if idx_uc >= len(row):
            continue
        uc = _somente_digitos(row[idx_uc])
        if not uc:
            continue
        if uc in vistos:
            repetidas += 1
            continue
        vistos.add(uc)
        ucs.append(uc)

    wb.close()
    print(
        f"[planilha] coluna usada: {coluna_uc} | UCs unicas: {len(ucs)} | repetidas ignoradas: {repetidas}"
    )
    if ucs:
        print(f"[planilha] primeira UC: {ucs[0]} | ultima UC: {ucs[-1]}")
    return ucs


def carregar_master():
    for caminho in MASTER_CANDIDATOS:
        try:
            if caminho.exists():
                spec = _ilu.spec_from_file_location("indice_master", str(caminho))
                mod = _ilu.module_from_spec(spec)
                sys.modules["indice_master"] = mod
                assert spec.loader is not None
                spec.loader.exec_module(mod)
                print(f"[master] carregado: {caminho}")
                return mod.MasterIndice(), mod
        except Exception as exc:
            print(f"[master] falha ao carregar {caminho}: {exc}")
    raise FileNotFoundError("indice_master.py nao encontrado nos candidatos configurados.")


def criar_driver() -> webdriver.Chrome:
    TEMP_DOWNLOAD.mkdir(parents=True, exist_ok=True)
    prefs = {
        "download.default_directory": str(TEMP_DOWNLOAD),
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "download.open_pdf_in_system_reader": False,
    }
    opts = Options()
    opts.add_experimental_option("prefs", prefs)
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--lang=pt-BR")
    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )
    try:
        driver.execute_cdp_cmd(
            "Browser.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": str(TEMP_DOWNLOAD),
                "eventsEnabled": False,
            },
        )
    except Exception:
        pass
    return driver


def salvar_html(driver: webdriver.Chrome, nome: str) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / f"{nome}.html").write_text(driver.page_source, encoding="utf-8")
    print(f"  [HTML salvo] {nome}.html")


def listar_arquivos_recentes(base: Path, limite: int = 10) -> list[str]:
    if not base.exists():
        return []
    arquivos = [p for p in base.rglob("*") if p.is_file()]
    arquivos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return [str(p) for p in arquivos[:limite]]


def _normalizar_ref_light(ref_bruta: str) -> str:
    valor = (ref_bruta or "").strip().upper()
    match = None
    if "/" in valor:
        match = valor.split("/")
        if len(match) == 2 and match[0] in MAPA_MESES:
            return f"{MAPA_MESES[match[0]]}.{match[1]}"
    raise ValueError(f"Referencia da LIGHT em formato inesperado: {ref_bruta!r}")


def limpar_downloads_temporarios() -> None:
    removidos = 0
    for pattern in ("*.pdf", "*.crdownload"):
        for arquivo in TEMP_DOWNLOAD.glob(pattern):
            try:
                arquivo.unlink()
                removidos += 1
            except OSError:
                pass
    if removidos:
        print(f"  pasta temporaria limpa: {removidos} arquivo(s) removido(s)")


def fechar_modais_residuais(driver: webdriver.Chrome, tentativas: int = 3) -> None:
    for _ in range(tentativas):
        fechou_algo = False
        seletores = [
            (By.XPATH, "//input[contains(@value, 'VOLTAR')]"),
            (By.XPATH, "//button[contains(normalize-space(.), 'VOLTAR')]"),
            (By.XPATH, "//span[contains(@class, 'close') or normalize-space(.)='×']"),
            (By.XPATH, "//button[contains(@class, 'close')]"),
        ]
        for by, selector in seletores:
            try:
                elementos = driver.find_elements(by, selector)
            except WebDriverException:
                continue
            for el in elementos:
                try:
                    if not el.is_displayed():
                        continue
                except WebDriverException:
                    continue
                if _tentar_click(driver, el):
                    fechou_algo = True
                    time.sleep(1)
                    break
            if fechou_algo:
                break
        if not fechou_algo:
            break


def garantir_aba_principal(driver: webdriver.Chrome, main_handle: str) -> None:
    handles = list(driver.window_handles)
    for handle in handles:
        if handle == main_handle:
            continue
        try:
            driver.switch_to.window(handle)
            if driver.current_url == "about:blank":
                driver.close()
        except WebDriverException:
            continue
    driver.switch_to.window(main_handle)


def _clicar_js(driver: webdriver.Chrome, elemento) -> bool:
    try:
        driver.execute_script(
            """
            arguments[0].scrollIntoView({block: 'center', inline: 'center'});
            arguments[0].click();
            """,
            elemento,
        )
        return True
    except WebDriverException:
        return False


def _tentar_click(driver: webdriver.Chrome, elemento) -> bool:
    try:
        elemento.click()
        return True
    except WebDriverException:
        return _clicar_js(driver, elemento)


def _campo_uc_pronto(driver: webdriver.Chrome):
    try:
        campo = driver.find_element(By.CSS_SELECTOR, LIGHT_CAMPO_UC_SELECTOR)
    except WebDriverException:
        return None
    try:
        if campo.is_displayed() and campo.is_enabled():
            return campo
    except WebDriverException:
        return None
    return None


def fazer_login(driver: webdriver.Chrome, wait: WebDriverWait) -> None:
    print("Preenchendo login automaticamente...")
    campo_email = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id*='wtUserNameInput']"))
    )
    campo_email.click()
    campo_email.clear()
    campo_email.send_keys(LOGIN)

    campo_senha = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id*='wtPasswordInput']"))
    )
    campo_senha.click()
    campo_senha.clear()
    campo_senha.send_keys(SENHA)

    btn_entrar = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input.btn-entrar, input[id*='wtEntrar']"))
    )
    print(f"  clicando Entrar: {btn_entrar.get_attribute('value')!r}")
    _tentar_click(driver, btn_entrar)


def _clicar_modal_banco_brasil(driver: webdriver.Chrome) -> bool:
    seletores = [
        (By.ID, LIGHT_MODAL_BB_ID),
        (By.XPATH, f"//a[@id='{LIGHT_MODAL_BB_ID}']"),
        (
            By.XPATH,
            "//a[contains(@class, 'card-nagv__action') and contains(@href, '__doPostBack')]",
        ),
    ]

    for by, seletor in seletores:
        try:
            elementos = driver.find_elements(by, seletor)
        except WebDriverException:
            continue
        for elemento in elementos:
            try:
                if not elemento.is_displayed():
                    continue
            except WebDriverException:
                continue

            href = (elemento.get_attribute("href") or "").strip()
            if _tentar_click(driver, elemento):
                print("  modal LIGHT: clique automatico executado")
                return True

            if "__doPostBack" in href:
                try:
                    driver.execute_script(href.removeprefix("javascript:").rstrip(";"))
                    print("  modal LIGHT: postback automatico executado")
                    return True
                except WebDriverException:
                    continue
    return False


def aguardar_area_uc(driver: webdriver.Chrome, timeout: int = 180):
    print(f"\n{'='*55}")
    print("AGUARDANDO AREA DA UC")
    print("  Tentando superar automaticamente o modal Banco do Brasil.")
    print(f"  (timeout: {timeout}s)")
    print(f"{'='*55}\n")

    ultima_tentativa = -10
    for tick in range(timeout):
        campo = _campo_uc_pronto(driver)
        if campo is not None:
            print(f"Campo da UC detectado apos {tick+1}s.")
            return campo

        if tick - ultima_tentativa >= 3:
            if _clicar_modal_banco_brasil(driver):
                ultima_tentativa = tick

        if (tick + 1) % 10 == 0:
            print(f"  [{tick+1}s] ainda aguardando a area da UC ficar pronta...")
        time.sleep(1)

    raise TimeoutException("Campo de busca da UC nao apareceu apos tentar superar o modal da LIGHT.")


def _clicar_primeiro_visivel(driver: webdriver.Chrome, seletores) -> bool:
    for by, selector in seletores:
        try:
            elementos = driver.find_elements(by, selector)
        except WebDriverException:
            continue
        for el in elementos:
            try:
                if not el.is_displayed():
                    continue
            except WebDriverException:
                continue
            if _tentar_click(driver, el):
                return True
    return False


def aguardar_spinners_busca(driver: webdriver.Chrome, timeout: int = 45) -> None:
    print("  aguardando spins da busca terminarem...")
    fim = time.time() + timeout
    while time.time() < fim:
        try:
            spinner_ids = [
                "span[id*='wtEstaBuscando']",
                "span[class*='Loading']",
                "div[class*='loading']",
                "div[class*='Loading']",
            ]
            ativos = False
            for selector in spinner_ids:
                for el in driver.find_elements(By.CSS_SELECTOR, selector):
                    try:
                        if el.is_displayed():
                            ativos = True
                            break
                    except WebDriverException:
                        continue
                if ativos:
                    break

            pronto = driver.execute_script(
                """
                const spans = Array.from(document.querySelectorAll("span[id*='wtEstaBuscando']"));
                const loading = Array.from(document.querySelectorAll("[class*='loading'], [class*='Loading']"));
                const visible = (el) => {
                  const s = window.getComputedStyle(el);
                  return s && s.display !== 'none' && s.visibility !== 'hidden' && el.offsetParent !== null;
                };
                return !spans.some(visible) && !loading.some(visible);
                """
            )
            if pronto and not ativos:
                print("  spins finalizados")
                return
        except WebDriverException:
            pass
        time.sleep(1)

    print("  timeout aguardando spins; seguindo assim mesmo")


def aguardar_uc_normalizada(driver: webdriver.Chrome, uc_digits: str, timeout: int = 20) -> str:
    print("  aguardando a UC normalizar no campo...")
    fim = time.time() + timeout
    while time.time() < fim:
        try:
            campo = driver.find_element(By.CSS_SELECTOR, "input[id*='wtInputBuscarUnidadeConsumidora']")
            valor = (campo.get_attribute("value") or "").strip()
            apenas_digitos = "".join(ch for ch in valor if ch.isdigit())
            if apenas_digitos.endswith(uc_digits) and len(apenas_digitos) >= len(uc_digits):
                print(f"  UC normalizada para: {valor}")
                return valor
        except WebDriverException:
            pass
        time.sleep(1)
    raise TimeoutException("UC nao normalizou no campo apos a primeira busca.")


def clicar_buscar(driver: webdriver.Chrome) -> None:
    btn_buscar = driver.find_element(By.CSS_SELECTOR, "input[id*='wtBuscarCodigoInstalacaoBtn']")
    _tentar_click(driver, btn_buscar)
    print("  Buscar clicado")


def selecionar_motivo_e_confirmar(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    timeout_modal: int = 12,
) -> bool:
    print("  aguardando modal de motivo...")
    select_el = WebDriverWait(driver, timeout_modal).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "select[id*='wtMotivoSegundaViaSelect']"))
    )
    Select(select_el).select_by_visible_text(MOTIVO_SEGUNDA_VIA)
    print(f"  motivo selecionado: {MOTIVO_SEGUNDA_VIA}")
    time.sleep(1.5)

    fim = time.time() + 20
    ultimo_erro: Exception | None = None
    while time.time() < fim:
        try:
            btn_confirmar = driver.find_element(By.CSS_SELECTOR, "input[id*='wtbtnConfirmarDownload']")
            disabled = (btn_confirmar.get_attribute("disabled") or "").strip().lower()
            classes = (btn_confirmar.get_attribute("class") or "").lower()
            if disabled or "disabled" in classes:
                time.sleep(0.5)
                continue
            if _tentar_click(driver, btn_confirmar):
                print("  Confirmar download clicado")
                return True
        except (NoSuchElementException, StaleElementReferenceException, WebDriverException) as exc:
            ultimo_erro = exc
            time.sleep(0.5)
            continue

    if ultimo_erro:
        raise ultimo_erro
    raise TimeoutException("Botao de confirmar download nao ficou clicavel apos selecionar o motivo.")


def capturar_referencia_fatura(driver: webdriver.Chrome) -> tuple[str, str]:
    candidatos = [
        "//*[contains(normalize-space(.), 'Ref: Mês/Ano') or contains(normalize-space(.), 'Ref: Mes/Ano')]",
        "//*[contains(normalize-space(.), 'Escolha qual mês de referência') or contains(normalize-space(.), 'Escolha qual mês de referencia')]",
    ]
    for xpath in candidatos:
        try:
            elementos = driver.find_elements(By.XPATH, xpath)
        except WebDriverException:
            continue
        for el in elementos:
            try:
                texto = (el.text or "").strip()
            except WebDriverException:
                continue
            if not texto:
                continue
            match = None
            import re
            match = re.search(r"\b(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)/(\d{4})\b", texto.upper())
            if match:
                ref_bruta = f"{match.group(1)}/{match.group(2)}"
                ref_normalizada = _normalizar_ref_light(ref_bruta)
                print(f"  referencia detectada na tela: {ref_bruta} -> {ref_normalizada}")
                return ref_bruta, ref_normalizada
    raise TimeoutException("Nao consegui identificar a referencia da fatura na tela antes do download.")


def montar_destino_final(ref_normalizada: str) -> Path:
    destino = FINAL_DOWNLOAD_ROOT / ref_normalizada / "BT"
    destino.mkdir(parents=True, exist_ok=True)
    print(f"  destino final LIGHT: {destino}")
    return destino


def aguardar_pdf_temporario(arquivos_antes: set[str], timeout: int = 40) -> Path | None:
    fim = time.time() + timeout
    while time.time() < fim:
        pdfs = [p for p in TEMP_DOWNLOAD.glob("*.pdf") if p.is_file()]
        crdownloads = [p for p in TEMP_DOWNLOAD.glob("*.crdownload") if p.is_file()]
        pdfs_novos = [p for p in pdfs if p.name not in arquivos_antes]
        if pdfs_novos and not crdownloads:
            pdfs_novos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            return pdfs_novos[0]
        time.sleep(1)
    return None


def mover_pdf_para_destino(origem: Path, destino: Path, carimbo: str) -> Path:
    alvo = destino / f"{carimbo}.pdf"
    contador = 1
    while alvo.exists():
        alvo = destino / f"{carimbo}_{contador}.pdf"
        contador += 1
    shutil.move(str(origem), str(alvo))
    print(f"  PDF movido: {origem.name} -> {alvo}")
    return alvo


def coletar_evidencias_download(
    driver: webdriver.Chrome,
    arquivos_antes: set[str],
    main_handle: str,
) -> Path | None:
    time.sleep(10)
    handle_inicial = driver.current_window_handle
    print(f"  handles abertos: {len(driver.window_handles)}")
    for idx, handle in enumerate(driver.window_handles):
        try:
            driver.switch_to.window(handle)
            print(f"    aba[{idx}] url={driver.current_url}")
        except WebDriverException:
            continue

    pdfs = list(TEMP_DOWNLOAD.glob("*.pdf"))
    crdownloads = list(TEMP_DOWNLOAD.glob("*.crdownload"))
    print(f"  PDFs em temp: {[p.name for p in pdfs]}")
    print(f"  CRDOWNLOAD em temp: {[p.name for p in crdownloads]}")
    recentes = listar_arquivos_recentes(TEMP_DOWNLOAD, limite=10)
    if recentes:
        print("  arquivos recentes temp:")
        for caminho in recentes:
            print(f"    - {caminho}")

    try:
        mensagens = driver.execute_script(
            """
            return Array.from(document.querySelectorAll("body *"))
              .map((el) => (el.innerText || "").trim())
              .filter((txt) => txt && txt.length <= 200)
              .filter((txt) =>
                /download|baixar|erro|sucesso|motivo|fatura|segunda via|referência|referencia/i.test(txt)
              )
              .slice(0, 40);
            """
        )
        if mensagens:
            print("  textos relevantes apos confirmar:")
            for msg in mensagens:
                print(f"    - {msg}")
    except WebDriverException:
        pass
    try:
        if main_handle in driver.window_handles:
            driver.switch_to.window(main_handle)
        elif handle_inicial in driver.window_handles:
            driver.switch_to.window(handle_inicial)
    except WebDriverException:
        pass
    return aguardar_pdf_temporario(arquivos_antes, timeout=40)


def processar_uc(
    driver: webdriver.Chrome,
    uc_digits: str,
    wait: WebDriverWait,
    master,
    master_mod,
    main_handle: str,
) -> str:
    print(f"\n--- UC={uc_digits} ---")
    garantir_aba_principal(driver, main_handle)
    fechar_modais_residuais(driver)
    limpar_downloads_temporarios()

    try:
        campo_uc = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id*='wtInputBuscarUnidadeConsumidora']"))
        )
        _tentar_click(driver, campo_uc)
        campo_uc.clear()
        campo_uc.send_keys(uc_digits)
        print(f"  UC digitada: {uc_digits}")
    except TimeoutException:
        return "ERRO: campo UC nao encontrado"

    try:
        clicar_buscar(driver)
    except NoSuchElementException:
        return "ERRO: botao Buscar nao encontrado"

    aguardar_spinners_busca(driver, timeout=45)
    try:
        aguardar_uc_normalizada(driver, uc_digits, timeout=20)
    except TimeoutException:
        return "ERRO: UC nao normalizou depois da primeira busca"

    try:
        clicar_buscar(driver)
    except NoSuchElementException:
        return "ERRO: botao Buscar nao encontrado na segunda busca"

    aguardar_spinners_busca(driver, timeout=45)
    time.sleep(2)
    salvar_html(driver, f"resultado_busca_{uc_digits}")

    if not _clicar_primeiro_visivel(
        driver,
        [
            (By.CSS_SELECTOR, "div.accordion-fi__btn"),
            (By.XPATH, "//div[contains(@class,'accordion-fi__btn')]"),
        ],
    ):
        return "SEM_RESULTADO: accordion nao apareceu (UC sem fatura?)"
    print("  Accordion expandido")
    time.sleep(2)

    if not _clicar_primeiro_visivel(
        driver,
        [
            (By.CSS_SELECTOR, "div.dots-btn"),
            (By.XPATH, "//div[contains(@class,'dots-btn')]"),
        ],
    ):
        return "ERRO: dots-btn nao encontrado"
    print("  Dots clicado")
    time.sleep(1)

    try:
        _, ref_normalizada = capturar_referencia_fatura(driver)
        destino_final = montar_destino_final(ref_normalizada)
    except TimeoutException:
        return "ERRO: nao consegui capturar a referencia da fatura"

    if not _clicar_primeiro_visivel(
        driver,
        [
            (By.CSS_SELECTOR, "a[id*='wtlnkDownloadDesktop']"),
            (By.XPATH, "//a[normalize-space(.)='Baixar']"),
            (By.XPATH, "//a[contains(normalize-space(.), 'Baixar')]"),
        ],
    ):
        return "ERRO: link Baixar nao encontrado"

    print("  Baixar clicado")
    arquivos_antes = {p.name for p in TEMP_DOWNLOAD.glob("*") if p.is_file()}
    try:
        selecionar_motivo_e_confirmar(driver, wait, timeout_modal=12)
    except TimeoutException:
        print("  modal de motivo nao apareceu; seguindo para aguardar download direto")
    salvar_html(driver, f"apos_confirmar_{uc_digits}")
    pdf_temporario = coletar_evidencias_download(driver, arquivos_antes, main_handle)
    if pdf_temporario is None:
        return "ERRO: download nao apareceu na pasta temporaria"

    carimbo = master.consumir_carimbo()
    pdf_movido = mover_pdf_para_destino(pdf_temporario, destino_final, carimbo)
    mes_ref_master = master_mod.normalizar_mes_ref(ref_normalizada.replace(".", "-"))
    master.registrar(
        indice_bb=carimbo,
        sistema="LIGHT_RJ",
        uc=uc_digits,
        mes_ref=mes_ref_master,
        arquivo=str(pdf_movido),
        estado="RIO DE JANEIRO",
        concessionaria="LIGHT",
    )
    print(f"  master registrado: {carimbo} | UC={uc_digits} | REF={mes_ref_master}")
    return f"OK: {carimbo}"


def main() -> int:
    args = parse_args()
    driver = criar_driver()
    wait = WebDriverWait(driver, 20)

    try:
        print(f"Abrindo: {args.url}")
        driver.get(args.url)
        time.sleep(3)

        fazer_login(driver, wait)
        salvar_html(driver, "apos_login_submit")

        aguardar_area_uc(driver, timeout=max(10, int(args.timeout_modal)))
        print(f"Area da UC pronta. URL={driver.current_url}")
        salvar_html(driver, "area_logada")
        if args.preflight:
            print("PREFLIGHT_PASS: login, modal e campo inicial validados; nenhum download iniciado.")
            return 0

        master, master_mod = carregar_master()
        ucs_lote = carregar_ucs_planilha(PLANILHA_UCS)
        main_handle = driver.current_window_handle

        print(
            f"\nIniciando downloads do lote LIGHT ({len(ucs_lote)} UCs) | "
            f"inicio configurado: {args.inicio_lote}"
        )
        resultados = {}

        for idx, uc in enumerate(ucs_lote, start=1):
            if idx < args.inicio_lote:
                continue
            print(f"\n[{idx}/{len(ucs_lote)}] processando UC {uc}")
            try:
                status = processar_uc(driver, uc, wait, master, master_mod, main_handle)
            except Exception as exc:
                status = f"ERRO_EXCECAO: {exc.__class__.__name__}: {exc}"
                try:
                    salvar_html(driver, f"erro_uc_{idx}_{uc}")
                except Exception:
                    pass
                try:
                    garantir_aba_principal(driver, main_handle)
                    fechar_modais_residuais(driver)
                except Exception:
                    pass
            resultados[uc] = status
            print(f"  -> {uc}: {status}")
            time.sleep(2)

        print("\n=== RESUMO ===")
        for uc, st in resultados.items():
            print(f"  {uc}: {st}")
        print("\nFluxo finalizado. Fechando navegador em 20s...")
        time.sleep(20)
        return 0

    except Exception as exc:
        print(f"\nERRO: {exc}")
        import traceback

        traceback.print_exc()
        try:
            salvar_html(driver, "erro_final")
        except Exception:
            pass
        print("\nFechando navegador em 30s...")
        time.sleep(30)
        return 1
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    raise SystemExit(main())
