#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cemig_190.py  —  Downloader dedicado para as 190 instalacoes CEMIG especificas
================================================================================
Le:  \\\\10.10.250.21\\Energia\\ARQUIVOS ENZO\\DOWNLOAD CEMIG\\
         instalacoes_cemig_190_status.csv

Salva em:
    \\\\10.10.250.21\\Energia\\ARQUIVOS ENZO\\DOWNLOAD CEMIG\\
        Faltantes\\MM.AAAA\\BT|MT|NAO_IDENTIFICADA\\BB_xxxxxxx.pdf

Diferenca do fluxo normal (cemig.py):
  - Lista vem do CSV, nao da planilha Senhas_CEMIG.xlsx
  - Baixa a fatura mais recente de 2026 INDEPENDENTE do status de pagamento
    (o fluxo normal ignora faturas ja pagas — por isso essas nao foram baixadas)
  - UCs com status_senha_cemig=FALTANTE sao logadas e puladas
  - Sem filtro len(uc) < 10 — aceita numeros curtos
  - Pasta de destino: <mes>/<BT|MT|NAO_IDENTIFICADA>
"""

from __future__ import annotations

import argparse
import sys
import ctypes as _ctypes
from pathlib import Path

if sys.platform == "win32":
    try:
        _ctypes.windll.kernel32.SetConsoleCtrlHandler(None, True)
    except Exception:
        pass

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# Ajusta path para importar cemig.py e o _venv_check dentro de core/.
_THIS_DIR = Path(__file__).resolve().parent
_ROOT_DIR = _THIS_DIR.parent.parent
sys.path.insert(0, str(_THIS_DIR))   # permite: import cemig
sys.path.insert(0, str(_ROOT_DIR))   # permite: import _venv_check

import _venv_check  # noqa

import csv
import io
import re
import shutil
import traceback
import urllib.parse
from dataclasses import dataclass
from typing import List, Optional

from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Importa toda a infraestrutura de cemig.py (Selenium, login, indices, etc.)
import cemig as C


# =============================================================================
# CONFIGURACAO
# =============================================================================

# Pasta raiz exclusiva para este script
DESTINO_DIR = C.ROOT_DIR / "DOWNLOAD CEMIG"

# CSV de entrada — tenta servidor, cai no local
CSV_190_SERVIDOR = C.ROOT_DIR / "DOWNLOAD CEMIG" / "instalacoes_cemig_190_status.csv"
CSV_190_LOCAL    = _ROOT_DIR / "logs" / "consultas_cemig" / "instalacoes_cemig_190_status.csv"
SENHAS_CEMIG_XLSX = C.ROOT_DIR / "DOWNLOAD CEMIG" / "Senhas CEMIG.xlsx"
SENHAS_CEMIG_LOCAL = _ROOT_DIR / "logs" / "consultas_cemig" / "Senhas CEMIG.xlsx"

# Mes alvo: apenas faturas deste mes serao baixadas.
# Formato "MM-AAAA". None = baixa a mais recente disponivel.
MES_ALVO: str | None = "03-2026"
RELATORIO_190 = DESTINO_DIR / f"relatorio_190_{(MES_ALVO or 'geral').replace('-','_')}.txt"

# Se True: processa apenas linhas com buscar=SIM no CSV (modo retry).
# Se False: processa todas as PRESENTE independente da coluna buscar.
APENAS_BUSCAR_SIM: bool = False
CSV_190_OVERRIDE: Path | None = None
RELATORIO_190_OVERRIDE: Path | None = None


# =============================================================================
# DATACLASS
# =============================================================================

@dataclass
class Item190:
    instalacao:   str   # "7.016.325.018-23"
    uc:           str   # so digitos: "701632501823"
    cnpj_texto:   str   # "00.000.000/5435-60"
    cnpj_digitos: str   # "00000000543560"
    faltante:     bool  # True se status_senha_cemig=FALTANTE


# =============================================================================
# CSV / RELATORIO HELPERS
# =============================================================================

def _resolver_csv_190() -> Path | None:
    for caminho in [CSV_190_SERVIDOR, CSV_190_LOCAL]:
        try:
            if caminho.exists():
                return caminho
        except OSError:
            continue
    return None


def _ler_csv_rows(caminho: Path) -> list[dict[str, str]]:
    texto = caminho.read_text(encoding="utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(texto)))
    normalizadas: list[dict[str, str]] = []
    for row in rows:
        normalizadas.append({(k or "").replace("\ufeff", ""): v for k, v in row.items()})
    return normalizadas


def _mapa_senhas_cemig(path_xlsx: Path | None = None) -> dict[str, str]:
    """
    Retorna mapa {instalacao_normalizada: cnpj}.
    """
    ultimo_erro: Exception | None = None
    for candidato in [path_xlsx, SENHAS_CEMIG_XLSX, SENHAS_CEMIG_LOCAL]:
        if not candidato:
            continue
        try:
            wb = load_workbook(candidato, read_only=True, data_only=True)
            break
        except Exception as exc:
            ultimo_erro = exc
            continue
    else:
        raise RuntimeError(f"Nao foi possivel abrir Senhas CEMIG.xlsx: {ultimo_erro}")

    ws = wb.active

    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    idx_inst = next((i for i, h in enumerate(headers) if h.strip().lower() == "instalacao"), None)
    idx_cnpj = next((i for i, h in enumerate(headers) if h.strip().lower() == "cnpj"), None)
    if idx_inst is None or idx_cnpj is None:
        raise RuntimeError(f"Cabecalhos nao encontrados em Senhas CEMIG.xlsx: {headers}")

    mapa: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        inst = str(row[idx_inst]).strip() if row[idx_inst] is not None else ""
        cnpj = str(row[idx_cnpj]).strip() if row[idx_cnpj] is not None else ""
        inst_norm = re.sub(r"\D", "", inst)
        if inst_norm and cnpj:
            mapa[inst_norm] = cnpj
    wb.close()
    return mapa


def _extrair_instalacao_relatorio(linha: str, prefixo: str) -> str | None:
    linha = linha.strip()
    if not linha.startswith(prefixo):
        return None
    m = re.match(rf"^{re.escape(prefixo)}\s+(.+?)\s+UC\s+\d+", linha)
    return m.group(1).strip() if m else None


def _marcacoes_retry_do_relatorio(path_relatorio: Path) -> tuple[set[str], set[str]]:
    """
    Retorna dois conjuntos de instalacoes:
      - baixados: marcar buscar=NAO
      - retry:    linhas de nao baixados/erro no relatorio

    Regra operacional atual:
      - tudo que ja baixou nao precisa repetir;
      - o restante segue como candidato a retry.
    """
    texto = path_relatorio.read_text(encoding="utf-8")
    baixados: set[str] = set()
    retry: set[str] = set()

    for linha in texto.splitlines():
        inst_ok = _extrair_instalacao_relatorio(linha, "OK")
        if inst_ok:
            baixados.add(inst_ok)
            continue

        inst_retry = _extrair_instalacao_relatorio(linha, "--")
        if inst_retry:
            retry.add(inst_retry)

    return baixados, retry


def preparar_csv_retry(path_csv: Path | None = None, path_relatorio: Path | None = None) -> Path:
    """
    Atualiza a coluna buscar do CSV com base no relatorio:
      - BAIXADOS => buscar=NAO
      - NAO BAIXADOS / ERRO / FALTANTE => buscar=SIM
      - linhas fora do relatorio atual mantêm o valor existente
    """
    path_csv = path_csv or _resolver_csv_190()
    if path_csv is None:
        raise FileNotFoundError("CSV 190 nao encontrado.")

    path_relatorio = path_relatorio or RELATORIO_190
    if not path_relatorio.exists():
        raise FileNotFoundError(f"Relatorio nao encontrado: {path_relatorio}")

    rows = _ler_csv_rows(path_csv)
    baixados, retry = _marcacoes_retry_do_relatorio(path_relatorio)
    mapa_senhas = _mapa_senhas_cemig()

    atualizados = 0
    promovidas = 0
    for row in rows:
        instalacao = (row.get("instalacao") or "").strip()
        inst_norm = re.sub(r"\D", "", (row.get("instalacao_normalizada") or instalacao))
        status = (row.get("status_senha_cemig") or "").strip().upper()
        novo_buscar = row.get("buscar", "NAO").strip().upper() or "NAO"
        promovido_agora = False

        cnpj_senhas = mapa_senhas.get(inst_norm, "")
        if cnpj_senhas:
            if status != "PRESENTE":
                promovidas += 1
                promovido_agora = True
            row["status_senha_cemig"] = "PRESENTE"
            row["cnpj"] = cnpj_senhas
            status = "PRESENTE"

        if instalacao in baixados:
            novo_buscar = "NAO"
        elif instalacao in retry:
            novo_buscar = "SIM"

        if row.get("buscar", "").strip().upper() != novo_buscar:
            atualizados += 1
        row["buscar"] = novo_buscar

    fieldnames = list(rows[0].keys()) if rows else [
        "instalacao", "instalacao_normalizada", "status_senha_cemig", "cnpj", "buscar"
    ]
    with path_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    sim = sum(1 for row in rows if (row.get("buscar") or "").strip().upper() == "SIM")
    nao = sum(1 for row in rows if (row.get("buscar") or "").strip().upper() == "NAO")
    C.log(
        f"CSV retry atualizado: {path_csv} | alteradas={atualizados} | promovidas={promovidas} | buscar=SIM {sim} | buscar=NAO {nao}",
        "OK",
    )
    return path_csv


# =============================================================================
# LEITURA DO CSV
# =============================================================================

def _ler_csv_190(path_csv: Path | None = None) -> List[Item190]:
    """
    Le instalacoes_cemig_190_status.csv.
    Colunas: instalacao, instalacao_normalizada, status_senha_cemig, cnpj, buscar
    """
    caminhos = [path_csv] if path_csv else [CSV_190_OVERRIDE, CSV_190_SERVIDOR, CSV_190_LOCAL]
    caminhos = [c for c in caminhos if c]

    for caminho in caminhos:
        conteudo = C._ler_arquivo_unc(caminho, timeout=10)
        if conteudo:
            C.log(f"CSV 190 carregado: {caminho}", "OK")
            break
    else:
        C.log(f"CSV nao encontrado:\n  {CSV_190_SERVIDOR}\n  {CSV_190_LOCAL}", "ERR")
        return []

    itens: List[Item190] = []
    faltantes: List[str] = []

    for row in csv.DictReader(io.StringIO(conteudo)):
        instalacao  = row.get("instalacao", "").strip()
        normalizada = row.get("instalacao_normalizada", "").strip()
        status      = row.get("status_senha_cemig", "").strip().upper()
        cnpj_raw    = row.get("cnpj", "").strip()
        buscar      = row.get("buscar", "").strip().upper()

        if not instalacao:
            continue

        # Modo retry: pula linhas sem buscar=SIM
        if APENAS_BUSCAR_SIM and buscar != "SIM":
            continue

        uc = re.sub(r"\D", "", normalizada) if normalizada else re.sub(r"\D", "", instalacao)
        if not uc:
            C.log(f"UC vazia, ignorando: {instalacao!r}", "WARN")
            continue

        if status == "FALTANTE":
            faltantes.append(instalacao)
            itens.append(Item190(instalacao=instalacao, uc=uc,
                                 cnpj_texto="", cnpj_digitos="", faltante=True))
            continue

        cnpj_d = re.sub(r"\D", "", cnpj_raw)
        if not cnpj_d:
            C.log(f"PRESENTE sem CNPJ: {instalacao} — ignorando", "WARN")
            continue

        itens.append(Item190(instalacao=instalacao, uc=uc,
                             cnpj_texto=cnpj_raw, cnpj_digitos=cnpj_d, faltante=False))

    presentes = sum(1 for x in itens if not x.faltante)
    C.log(f"CSV: {len(itens)} itens | {presentes} PRESENTES | {len(faltantes)} FALTANTES", "OK")
    if faltantes:
        C.log("FALTANTES (sem senha, serao pulados): " + ", ".join(faltantes), "WARN")
    return itens


# =============================================================================
# DOWNLOAD COM DESTINO FALTANTES
# =============================================================================

def _baixar_faltante(driver, fat: C.Fatura) -> Optional[Path]:
    """
    Replica o clique + espera de cemig.baixar_pdf(), mas salva em:
        DESTINO_DIR / MM.AAAA / BT|MT|NAO_IDENTIFICADA / _tmp_190.pdf

    Retorna o Path temporario (ainda sem carimbo). O chamador renomeia.
    Tambem define fat.classificacao com BT|MT|NAO_IDENTIFICADA.
    """
    temp_dir = C.CEMIG_DIR / "_temp"
    C._mkdir_seguro(temp_dir)
    antes = {p.name for p in temp_dir.glob("*.pdf")}

    # --- Clica no botao Baixar PDF da linha correta ---
    doc_q  = urllib.parse.quote(fat.documento_impressao)
    clicou = False
    for xp in [
        f"//a[contains(@onclick,'{doc_q}')][.//span[normalize-space()='Baixar PDF']]",
        (f"//td[normalize-space()='{fat.mes_ano}']/following-sibling::td"
         f"//a[contains(@onclick,'BaixarPDF')]"),
        f"//a[contains(@onclick,'BaixarPDF')][contains(@onclick,'{fat.documento_impressao[:15]}')]",
    ]:
        try:
            el = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xp)))
            C.clicar(driver, el, "btn_baixar_pdf")
            clicou = True
            break
        except Exception:
            continue

    # Fallback: URL direta
    if not clicou and fat.url_pdf:
        driver.get(fat.url_pdf)
        clicou = True

    if not clicou:
        C.log(f"  Botao de download nao encontrado ({fat.mes_ano})", "WARN")
        return None

    # --- Aguarda PDF aparecer em temp ---
    pdf_temp = C.aguardar_pdf(temp_dir, antes)
    if not pdf_temp:
        C.log(f"  Timeout — PDF nao chegou ({fat.mes_ano})", "WARN")
        return None

    # --- Classifica BT / MT ---
    classificacao    = C.classificar_pdf(pdf_temp)
    fat.classificacao = classificacao

    # --- Move para DESTINO_DIR / MM.AAAA / classificacao / _tmp_190.pdf ---
    pasta_dest  = DESTINO_DIR / fat.pasta / classificacao
    C._mkdir_seguro(pasta_dest)
    destino_tmp = pasta_dest / "_tmp_190.pdf"

    try:
        pdf_temp.rename(destino_tmp)
    except Exception:
        shutil.copy2(pdf_temp, destino_tmp)
        pdf_temp.unlink(missing_ok=True)

    C.log(f"  Classificado: {classificacao} | temp em {fat.pasta}/{classificacao}/",
          "DBG")
    return destino_tmp


# =============================================================================
# RELATORIO FINAL
# =============================================================================

def _imprimir_relatorio(baixados: list, sem_marco: list,
                        n_faltantes: int, erros: int, master) -> None:
    """Imprime e salva em TXT um relatorio mini com baixados e indisponiveis."""
    from datetime import datetime

    linhas = []
    W = linhas.append

    W("")
    W("=" * 72)
    W(f"  RELATORIO CEMIG 190  —  MES ALVO: {MES_ALVO or 'MAIS RECENTE'}")
    W(f"  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    W("=" * 72)

    # ── Baixados com sucesso ──────────────────────────────────────────────
    W(f"\n  BAIXADOS ({len(baixados)}):")
    W("  " + "-" * 68)
    if baixados:
        for instalacao, uc, carimbo, classif in baixados:
            W(f"  OK  {instalacao:<25}  UC {uc:<15}  {carimbo}  [{classif}]")
    else:
        W("  (nenhum)")

    # ── Sem fatura de marco / indisponiveis ───────────────────────────────
    W(f"\n  SEM FATURA DE {MES_ALVO or 'ALVO'} / NAO BAIXADOS ({len(sem_marco)}):")
    W("  " + "-" * 68)
    if sem_marco:
        for instalacao, uc, motivo in sem_marco:
            W(f"  --  {instalacao:<25}  UC {uc:<15}  {motivo}")
    else:
        W("  (nenhum)")

    # ── Faltantes (sem senha) ─────────────────────────────────────────────
    if n_faltantes:
        W(f"\n  FALTANTES sem senha no portal (nao processados): {n_faltantes}")

    W("")
    W(f"  Total baixados   : {len(baixados)}")
    W(f"  Total s/ fatura  : {len(sem_marco)}")
    W(f"  Erros tecnicos   : {erros}")
    if master:
        W(f"  Proximo BB_      : {master.proximo_carimbo}")
    W("=" * 72)
    W("")

    # Imprime no terminal
    print("\n".join(linhas))

    # Salva arquivo TXT na pasta Faltantes
    try:
        C._mkdir_seguro(DESTINO_DIR)
        nome_rel = f"relatorio_190_{(MES_ALVO or 'geral').replace('-','_')}.txt"
        path_rel = DESTINO_DIR / nome_rel
        path_rel.write_text("\n".join(linhas), encoding="utf-8")
        C.log(f"Relatorio salvo: {path_rel}", "OK")
    except Exception as e:
        C.log(f"Nao foi possivel salvar relatorio TXT: {e}", "WARN")


# =============================================================================
# EXECUCAO PRINCIPAL
# =============================================================================

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Downloader CEMIG 190 com suporte a preparo de retry pelo relatorio."
    )
    p.add_argument(
        "--preparar-retry",
        action="store_true",
        help="Atualiza a coluna buscar do CSV usando o relatorio_190 atual e encerra.",
    )
    p.add_argument("--csv", type=Path, default=None, help="Caminho alternativo do CSV 190.")
    p.add_argument(
        "--relatorio", type=Path, default=None, help="Caminho alternativo do relatorio 190."
    )
    p.add_argument(
        "--apenas-buscar-sim",
        action="store_true",
        help="Processa somente linhas com buscar=SIM no CSV.",
    )
    p.add_argument(
        "--mes-ref",
        default=None,
        help="Mes alvo no formato MM-AAAA. Ex.: 03-2026. Se vazio, usa o padrao do script.",
    )
    p.add_argument(
        "--mais-recente",
        action="store_true",
        help="Ignora MES_ALVO e baixa apenas a fatura mais recente disponivel de cada UC.",
    )
    return p.parse_args()


def executar_190() -> None:
    print("\n" + "=" * 72)
    print("  CEMIG 190 — Instalacoes especificas → <mes>/<BT|MT>")
    print("=" * 72)

    # Cria estrutura de diretorios
    C._mkdir_seguro(C.CEMIG_DIR)
    C._mkdir_seguro(C.DEBUG_DIR)
    C._mkdir_seguro(DESTINO_DIR)
    C.log(f"Pasta destino: {DESTINO_DIR}", "INFO")

    master = C._carregar_master()
    indice = C.IndiceLocal()

    # Carrega lista
    itens = _ler_csv_190(CSV_190_OVERRIDE)
    if not itens:
        C.log("Nenhum item. Abortando.", "ERR")
        return

    presentes = [x for x in itens if not x.faltante]
    n_faltantes = len(itens) - len(presentes)
    if not presentes:
        C.log("Todos os itens sao FALTANTES. Nada a fazer.", "WARN")
        return

    C.log(f"Processando {len(presentes)} PRESENTES (pulando {n_faltantes} FALTANTES)...", "INFO")

    # Driver + login
    driver = C.build_driver()
    try:
        if not C.fazer_login(driver, C.LOGIN_CEMIG, C.SENHA_CEMIG):
            C.log("Login falhou. Abortando.", "ERR")
            return

        # Le o select ddCliente uma vez apos o login
        mapa_select = C._ler_select(driver)
        inv_mapa    = {v: k for k, v in mapa_select.items() if len(k) > 8}

        erros      = 0
        cnpj_atual = ""

        # Listas para o relatorio final
        rel_baixados:     list = []   # (instalacao, uc, carimbo, classificacao)
        rel_sem_marco:    list = []   # (instalacao, uc, motivo)

        for i, item in enumerate(presentes, start=1):
            print(f"\n{'─' * 72}")
            C.log(f"[{i}/{len(presentes)}] {item.instalacao} | UC={item.uc} | CNPJ={item.cnpj_texto}")

            # Resolve PN do CNPJ no select do portal
            pn_val = C._buscar_pn(item.cnpj_digitos, mapa_select)
            if not pn_val:
                C.log(f"  CNPJ {item.cnpj_texto} nao encontrado no select", "WARN")
                erros += 1
                rel_sem_marco.append((item.instalacao, item.uc, "CNPJ nao encontrado no portal"))
                continue

            pn_txt = inv_mapa.get(pn_val, item.cnpj_texto)

            try:
                # Trocar UC (revela ddCliente + limparInst)
                C.clicar_trocar_uc(driver)

                # Selecionar CNPJ apenas quando muda
                if pn_val != cnpj_atual:
                    if not C.selecionar_cnpj(driver, pn_val, pn_txt):
                        C.log("  Falha CNPJ — pulando", "WARN")
                        erros += 1
                        rel_sem_marco.append((item.instalacao, item.uc, "Falha ao selecionar CNPJ"))
                        C._voltar_home(driver)
                        continue
                    cnpj_atual = pn_val

                # Digitar UC e pesquisar
                if not C.digitar_uc_e_pesquisar(driver, item.uc):
                    C.log("  Falha ao digitar UC — pulando", "WARN")
                    erros += 1
                    rel_sem_marco.append((item.instalacao, item.uc, "Falha ao pesquisar UC"))
                    C._voltar_home(driver)
                    continue

                # Historico de Contas
                if not C.clicar_historico(driver):
                    C.log("  Historico nao abriu — pulando", "WARN")
                    erros += 1
                    rel_sem_marco.append((item.instalacao, item.uc, "Historico nao carregou"))
                    C._voltar_home(driver)
                    continue

                # Le faturas (>= ANO_MINIMO)
                faturas = C.ler_faturas(driver, item.uc)
                C.logar_faturas(faturas, item.uc)

                if not faturas:
                    C.log(f"  Sem faturas >= {C.ANO_MINIMO}", "WARN")
                    rel_sem_marco.append((item.instalacao, item.uc, f"Sem faturas >= {C.ANO_MINIMO}"))
                    C._voltar_home(driver)
                    continue

                # Filtra pelo mes alvo
                if MES_ALVO:
                    candidatas = [f for f in faturas if f.mes_ref == MES_ALVO]
                    if not candidatas:
                        meses_disp = ", ".join(sorted({f.mes_ano for f in faturas}))
                        C.log(f"  Sem fatura de {MES_ALVO} (disponiveis: {meses_disp})", "WARN")
                        rel_sem_marco.append((
                            item.instalacao, item.uc,
                            f"Mes {MES_ALVO} nao disponivel (tem: {meses_disp})",
                        ))
                        C._voltar_home(driver)
                        continue
                else:
                    candidatas = [
                        max(
                            faturas,
                            key=lambda f: tuple(map(int, f.mes_ref.split("-")[::-1])),
                        )
                    ]

                # Baixa cada candidata — SEM checar se ja existe no master/indice
                for fat in candidatas:
                    destino_tmp = _baixar_faltante(driver, fat)
                    if not destino_tmp:
                        erros += 1
                        rel_sem_marco.append((item.instalacao, item.uc,
                                              f"Falha no download ({fat.mes_ano})"))
                        continue

                    # Usa master apenas para gerar o carimbo
                    carimbo = (master.consumir_carimbo() if master
                               else f"BB_{indice.proximo:07d}")

                    destino_final = destino_tmp.parent / f"{carimbo}.pdf"
                    try:
                        destino_tmp.rename(destino_final)
                    except Exception:
                        shutil.copy2(destino_tmp, destino_final)
                        destino_tmp.unlink(missing_ok=True)

                    if master:
                        master.registrar(
                            indice_bb  = carimbo,
                            sistema    = "CEMIG",
                            uc         = item.uc,
                            mes_ref    = fat.mes_ref,
                            fatura_id  = fat.documento_impressao,
                            cnpj       = item.cnpj_digitos,
                            estado     = "MINAS GERAIS",
                            arquivo    = str(destino_final),
                        )
                    indice.gravar(
                        indice_bb  = carimbo,
                        uc         = item.uc,
                        mes_ref    = fat.mes_ref,
                        fatura_id  = fat.documento_impressao,
                        cnpj       = item.cnpj_digitos,
                        arquivo    = str(destino_final),
                    )

                    rel_baixados.append((item.instalacao, item.uc, carimbo, fat.classificacao))
                    C.log(
                        f"  OK {fat.mes_ano} ({fat.status}) "
                        f"[{fat.classificacao}] -> {carimbo}.pdf",
                        "OK",
                    )

                C._voltar_home(driver)

            except Exception as e:
                C.log(f"Erro UC {item.uc}: {e}", "ERR")
                traceback.print_exc()
                C.salvar_debug(driver, f"190_erro_{item.uc}")
                erros += 1
                rel_sem_marco.append((item.instalacao, item.uc, f"Erro: {e}"))
                C._voltar_home(driver)

        # ── Relatorio final ────────────────────────────────────────────────
        _imprimir_relatorio(rel_baixados, rel_sem_marco, n_faltantes, erros, master)

    except KeyboardInterrupt:
        C.log("Interrompido.", "WARN")
    except Exception as e:
        C.log(f"Erro fatal: {e}", "ERR")
        traceback.print_exc()
        C.salvar_debug(driver, "190_erro_fatal")
    finally:
        try:
            driver.quit()
            print("\n[CEMIG 190] Navegador fechado.")
        except Exception:
            pass
        print("[CEMIG 190] Processo finalizado.")


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    args = parse_args()
    if args.mais_recente:
        MES_ALVO = None
        RELATORIO_190 = DESTINO_DIR / "relatorio_190_geral.txt"
    if args.mes_ref:
        MES_ALVO = str(args.mes_ref).strip()
        RELATORIO_190 = DESTINO_DIR / f"relatorio_190_{(MES_ALVO or 'geral').replace('-','_')}.txt"
    if args.apenas_buscar_sim:
        APENAS_BUSCAR_SIM = True
    if args.csv:
        CSV_190_OVERRIDE = args.csv
    if args.relatorio:
        RELATORIO_190_OVERRIDE = args.relatorio
    if args.preparar_retry:
        preparar_csv_retry(CSV_190_OVERRIDE, RELATORIO_190_OVERRIDE)
    else:
        executar_190()
