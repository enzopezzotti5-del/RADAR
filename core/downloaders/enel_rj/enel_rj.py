#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import csv
import argparse
import importlib.util as _ilu
import logging
import os
import re
import shutil
import sys
import tempfile
import time
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import pdfplumber
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

try:
    from core.metrics.radar_metrics import emit_outcome as _emit_rj_outcome
    def _emit(outcome: str, *, uc: str, ref: str, carimbo: str = "") -> None:
        _emit_rj_outcome(outcome, utility="ENEL RJ", account_id=uc,
                         competence=ref, invoice_id=carimbo or ref)
except Exception:
    def _emit(outcome: str, **_: str) -> None:  # type: ignore[misc]
        pass


# =============================================================================
# CONFIG
# =============================================================================

BASE_DIR = Path("//10.10.250.21/Energia//ARQUIVOS ENZO/DOWNLOAD ENEL RJ")
PLANILHA_PATH = BASE_DIR / "senhas_enel_rj.xlsx"
INDICE_LOCAL_PATH = BASE_DIR / "indice_faturas_enel_rj.csv"
LOG_DIR = BASE_DIR / "LOGS"
TEMP_DOWNLOAD_DIR = Path(os.environ.get("LOCALAPPDATA", r"C:\Users\Public")) / "enel_rj_temp"
FALLBACK_BASE_DIR = Path(__file__).resolve().parent / "_runtime_fallback"
PROFILE_ROOT = Path(__file__).resolve().parent / "chrome_profiles"

URL_PORTAL = "https://www.eneldistribuicao.com.br/rj/Corporativo.aspx"

TIMEOUT_PADRAO = 20
TIMEOUT_DOWNLOAD = 90
MIN_PDF_SIZE = 5_000
PAUSA_ENTRE_UCS = 2.0

INDICE_FIELDS = [
    "INDICE",
    "UC",
    "SUBTIPO",
    "MES_REF",
    "FATURA_ID",
    "DATA_DOWNLOAD",
    "CAMINHO",
]

_ROOT_LOCAL = Path(__file__).resolve().parents[2]
MASTER_CANDIDATOS = [
    # Usar SEMPRE o modulo local; nao carregar .py da rede.
    _ROOT_LOCAL / "indice_master.py",
    Path(__file__).parent / "indice_master.py",
]


# =============================================================================
# LOG
# =============================================================================

def _configurar_logging() -> logging.Logger:
    log_dir = LOG_DIR
    try:
        log_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        log_dir = FALLBACK_BASE_DIR / "LOGS"
        log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    fmt = logging.Formatter("%(asctime)s    %(message)s", datefmt="%H:%M:%S")

    h_console = logging.StreamHandler(sys.stdout)
    h_console.setFormatter(fmt)
    h_console.setLevel(logging.INFO)

    h_file = logging.FileHandler(str(log_dir / f"enel_rj_{ts}.log"), encoding="utf-8")
    h_file.setFormatter(fmt)
    h_file.setLevel(logging.DEBUG)

    logger = logging.getLogger("enel_rj")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    logger.addHandler(h_console)
    logger.addHandler(h_file)
    return logger


log = _configurar_logging()


# =============================================================================
# HELPERS GERAIS
# =============================================================================

def _limpar_digitos(valor: str) -> str:
    return re.sub(r"\D", "", str(valor or "")).strip()


def _normalizar_uc(valor: str) -> str:
    """
    Normaliza UC para comparação:
    - remove não dígitos
    - remove zeros à esquerda
    - mantém '0' se a UC virar vazia
    """
    s = _limpar_digitos(valor)
    s = s.lstrip("0")
    return s or "0"


def _normalizar_texto(s) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s.strip()).encode("ASCII", "ignore").decode("ASCII")
    return " ".join(s.split()).lower()


def _xpath_literal(s: str) -> str:
    if "'" not in s:
        return f"'{s}'"
    if '"' not in s:
        return f'"{s}"'
    partes = s.split("'")
    return "concat(" + ", \"'\", ".join(f"'{p}'" for p in partes) + ")"


def _normalizar_situacao(txt: str) -> str:
    s = unicodedata.normalize("NFKD", str(txt or "")).encode("ASCII", "ignore").decode("ASCII")
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s


def _situacao_esta_paga(txt: str) -> bool:
    return _normalizar_situacao(txt) in {"PAGA", "PAGO"}


# =============================================================================
# ÍNDICE MASTER
# =============================================================================

def _carregar_master():
    for caminho in MASTER_CANDIDATOS:
        if caminho.exists():
            spec = _ilu.spec_from_file_location("indice_master", str(caminho))
            mod = _ilu.module_from_spec(spec)
            sys.modules["indice_master"] = mod
            assert spec.loader is not None
            spec.loader.exec_module(mod)
            log.info("Master carregado: %s", caminho)
            return mod.MasterIndice()
    raise FileNotFoundError(
        "indice_master.py não encontrado em:\n" + "\n".join(str(p) for p in MASTER_CANDIDATOS)
    )


# =============================================================================
# ÍNDICE LOCAL
# =============================================================================

def _carregar_indice_local() -> set[tuple[str, str]]:
    baixados: set[tuple[str, str]] = set()

    if not INDICE_LOCAL_PATH.exists():
        return baixados

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(INDICE_LOCAL_PATH, newline="", encoding=enc) as f:
                for row in csv.DictReader(f):
                    uc = _normalizar_uc(row.get("UC") or "")
                    mes = (row.get("MES_REF") or "").strip()
                    if uc and mes:
                        baixados.add((uc, mes))
            return baixados
        except UnicodeDecodeError:
            continue

    return baixados


def _registrar_indice_local(
    indice_bb: str,
    uc: str,
    subtipo: str,
    mes_ref: str,
    fatura_id: str,
    caminho: str,
) -> None:
    novo = not INDICE_LOCAL_PATH.exists()
    with open(INDICE_LOCAL_PATH, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=INDICE_FIELDS)
        if novo:
            w.writeheader()
        w.writerow({
            "INDICE": indice_bb,
            "UC": _normalizar_uc(uc),
            "SUBTIPO": subtipo.strip(),
            "MES_REF": mes_ref.strip(),
            "FATURA_ID": fatura_id.strip(),
            "DATA_DOWNLOAD": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "CAMINHO": caminho,
        })


# =============================================================================
# PLANILHA
# =============================================================================

@dataclass(frozen=True)
class LinhaAcesso:
    login: str
    senha: str
    uc_original: str
    uc_norm: str
    cnpj: str
    nome: str = ""
    tensao: str = ""


def _encontrar_coluna(headers: list[str], opcoes: list[str]) -> Optional[str]:
    mapa = {_normalizar_texto(c): c for c in headers}
    for opc in opcoes:
        k = _normalizar_texto(opc)
        if k in mapa:
            return mapa[k]
    return None


def carregar_planilha(path: Path) -> Dict[Tuple[str, str], List[LinhaAcesso]]:
    if not path.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Planilha vazia")

    headers = [str(h or "").strip() for h in rows[0]]

    col_login = _encontrar_coluna(headers, ["Login", "E-mail", "Email", "Usuario"])
    col_senha = _encontrar_coluna(headers, ["Senha", "Password"])
    col_uc = _encontrar_coluna(headers, ["UC", "Instalacao", "Instalação", "Unidade"])
    col_cnpj = _encontrar_coluna(headers, ["CNPJ", "CPF/CNPJ", "Documento"])
    col_nome = _encontrar_coluna(headers, ["Nome", "Cliente"])
    col_tensao = _encontrar_coluna(headers, ["Tensão", "Tensao", "Tipo"])

    faltando = [n for n, v in [
        ("Login", col_login),
        ("Senha", col_senha),
        ("UC", col_uc),
        ("CNPJ", col_cnpj),
    ] if v is None]

    if faltando:
        raise ValueError(f"Colunas obrigatórias não encontradas: {faltando}")

    idx = {h: i for i, h in enumerate(headers)}
    grupos: Dict[Tuple[str, str], List[LinhaAcesso]] = defaultdict(list)

    for raw in rows[1:]:
        def cel(nome_col: Optional[str]) -> str:
            if not nome_col:
                return ""
            i = idx[nome_col]
            return str(raw[i] or "").strip()

        login = cel(col_login)
        senha = cel(col_senha)
        uc_original = _limpar_digitos(cel(col_uc))
        uc_norm = _normalizar_uc(uc_original)
        cnpj = _limpar_digitos(cel(col_cnpj))
        nome = cel(col_nome)
        tensao = cel(col_tensao)

        if not login or not senha or not uc_original:
            continue

        grupos[(login, senha)].append(
            LinhaAcesso(
                login=login,
                senha=senha,
                uc_original=uc_original,
                uc_norm=uc_norm,
                cnpj=cnpj,
                nome=nome,
                tensao=tensao,
            )
        )

    total = sum(len(v) for v in grupos.values())
    log.info("Planilha carregada: %d grupos | %d UCs", len(grupos), total)
    return grupos


# =============================================================================
# PDF / REFERÊNCIA
# =============================================================================

def classificar_pdf(pdf_path: Path) -> str:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages[:2]:
                texto = (page.extract_text() or "").upper()

                if any(t in texto for t in [
                    "MÉDIA TENSÃO", "MEDIA TENSAO", "A4", "A3", "DEMANDA",
                    "HORO-SAZONAL", "HOROSAZONAL", "VERDE", "AZUL",
                ]):
                    return "MT"

                if any(t in texto for t in [
                    "BAIXA TENSÃO", "BAIXA TENSAO", "B1", "B3",
                    "RESIDENCIAL", "CONVENCIONAL",
                ]):
                    return "BT"
    except Exception as exc:
        log.debug("[PDF] Erro ao classificar %s: %s", pdf_path.name, exc)

    return "NAO_IDENTIFICADA"


_MESES_ABREV = {
    "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
    "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
    "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12",
}


def extrair_mes_ref(texto: str) -> Optional[str]:
    if not texto:
        return None

    up = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII").upper()

    m = re.search(r"\b(" + "|".join(_MESES_ABREV) + r")[/\-](20\d{2})\b", up)
    if m:
        return f"{_MESES_ABREV[m.group(1)]}-{m.group(2)}"

    m = re.search(r"\b(0[1-9]|1[0-2])[/-](20\d{2})\b", texto)
    if m:
        return f"{m.group(1)}-{m.group(2)}"

    return None


def extrair_mes_ref_do_pdf(pdf_path: Path) -> Optional[str]:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages[:2]:
                texto = page.extract_text() or ""
                for linha in texto.splitlines():
                    linha = linha.strip()
                    if linha and len(linha) <= 40:
                        ref = extrair_mes_ref(linha)
                        if ref:
                            return ref
    except Exception as exc:
        log.warning("[PDF] Erro ao extrair referência: %s", exc)

    return None


# =============================================================================
# DRIVER
# =============================================================================

def criar_driver() -> webdriver.Chrome:
    TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    prefs = {
        "download.default_directory": str(TEMP_DOWNLOAD_DIR),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "download.open_pdf_in_system_reader": False,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_settings.popups": 0,
    }

    opts = Options()
    PROFILE_ROOT.mkdir(parents=True, exist_ok=True)
    profile_dir = Path(tempfile.mkdtemp(prefix="enel_rj_", dir=str(PROFILE_ROOT)))
    opts.add_experimental_option("prefs", prefs)
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--lang=pt-BR")
    opts.add_argument(f"--user-data-dir={profile_dir.resolve()}")
    opts.add_argument("--remote-debugging-port=0")
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")

    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )

    try:
        driver.execute_cdp_cmd("Browser.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": str(TEMP_DOWNLOAD_DIR),
            "eventsEnabled": False,
        })
    except Exception:
        pass

    return driver


# =============================================================================
# HELPERS SELENIUM
# =============================================================================

def _wait(driver, timeout: int = TIMEOUT_PADRAO) -> WebDriverWait:
    return WebDriverWait(driver, timeout)


def _clicar(driver: webdriver.Chrome, by: By, locator: str, timeout: int = TIMEOUT_PADRAO) -> None:
    elem = _wait(driver, timeout).until(EC.element_to_be_clickable((by, locator)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    time.sleep(0.3)
    try:
        elem.click()
    except (ElementClickInterceptedException, StaleElementReferenceException):
        driver.execute_script("arguments[0].click();", elem)


def _digitar_lento(elem, texto: str) -> None:
    elem.clear()
    for ch in texto:
        elem.send_keys(ch)
        time.sleep(0.05)


def _aguardar_novo_pdf(antes: set[str], timeout: int = TIMEOUT_DOWNLOAD) -> Optional[Path]:
    fim = time.time() + timeout
    while time.time() < fim:
        atuais = {p.name for p in TEMP_DOWNLOAD_DIR.glob("*")}
        novos = {
            TEMP_DOWNLOAD_DIR / n
            for n in (atuais - antes)
            if n.lower().endswith(".pdf") and not n.lower().endswith(".crdownload")
        }

        if novos:
            pdf = max(novos, key=lambda p: p.stat().st_mtime)
            for _ in range(20):
                try:
                    s1 = pdf.stat().st_size
                    time.sleep(0.5)
                    s2 = pdf.stat().st_size
                    if s1 == s2 and s1 >= MIN_PDF_SIZE:
                        return pdf
                except FileNotFoundError:
                    break
        time.sleep(0.8)

    return None


# =============================================================================
# LOGIN / NAV
# =============================================================================

def _fazer_login(driver: webdriver.Chrome, login: str, senha: str) -> bool:
    log.info("[login] Abrindo portal RJ...")
    driver.get(URL_PORTAL)

    try:
        campo_email = _wait(driver).until(
            EC.visibility_of_element_located((By.ID, "WEBDOOR_headercorporativo_UserName"))
        )
        campo_senha = driver.find_element(By.ID, "WEBDOOR_headercorporativo_Password")
    except TimeoutException:
        log.error("[login] Campos de login não encontrados")
        return False

    log.info("[login] Preenchendo credenciais para %s...", login)
    _digitar_lento(campo_email, login)
    _digitar_lento(campo_senha, senha)

    try:
        _clicar(
            driver,
            By.XPATH,
            "//input[@type='submit' and contains(@value,'Entrar')]"
            " | //button[contains(normalize-space(.),'Entrar')]",
            timeout=10,
        )
    except Exception:
        campo_senha.send_keys(Keys.RETURN)

    time.sleep(3)

    # ── Verifica se o portal abriu um alert de erro (ex: "Cliente não encontrado") ──
    try:
        alert = driver.switch_to.alert
        msg_alert = alert.text
        log.warning("[login] Alert do portal: %s — credencial inválida para %s", msg_alert, login)
        alert.accept()
        return False
    except Exception:
        pass  # Nenhum alert presente, fluxo normal

    # ── Verifica URL para confirmar que saiu da tela de login ──────────────────
    try:
        url = driver.current_url.lower()
    except Exception as exc:
        log.warning("[login] Não foi possível ler URL após login (%s)", exc)
        return False

    if "corporativo.aspx" in url or "login" in url:
        log.warning("[login] Ainda na tela de login")
        return False

    log.info("[login] Login OK → %s", driver.current_url)
    return True


def _fechar_modal_pos_login(driver: webdriver.Chrome) -> None:
    time.sleep(1.5)

    seletores = [
        (By.CSS_SELECTOR, "button.close[data-dismiss='modal']"),
        (By.XPATH, "//button[@data-dismiss='modal']"),
        (By.XPATH, "//button[contains(@class,'ModalButton') and @data-dismiss='modal']"),
        (By.XPATH, "//button[normalize-space(.)='×' or normalize-space(.)='X' or normalize-space(.)='OK']"),
    ]

    for by, loc in seletores:
        try:
            elems = driver.find_elements(by, loc)
            alvo = next((e for e in elems if e.is_displayed()), None)
            if alvo:
                driver.execute_script("arguments[0].click();", alvo)
                time.sleep(0.8)
                return
        except Exception:
            pass


def _aguardar_tabela_ucs(driver: webdriver.Chrome, timeout: int = 15) -> bool:
    try:
        _wait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "CONTENT_gdEscolherClienteDoAgrupamento"))
        )
        return True
    except Exception:
        return False


def _aguardar_tabela_faturas(driver: webdriver.Chrome, timeout: int = 15) -> bool:
    try:
        _wait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "CONTENT_gdHistoricoDeFaturamento"))
        )
        return True
    except Exception:
        return False


# =============================================================================
# UCs
# =============================================================================

def _obter_ucs_da_tabela(driver: webdriver.Chrome) -> List[dict]:
    time.sleep(1.5)
    ucs: List[dict] = []

    try:
        _aguardar_tabela_ucs(driver, timeout=15)

        linhas = driver.find_elements(
            By.XPATH,
            "//table[@id='CONTENT_gdEscolherClienteDoAgrupamento']//tr[td]"
        )

        for i, tr in enumerate(linhas):
            try:
                tds = tr.find_elements(By.TAG_NAME, "td")
                if len(tds) < 4:
                    continue

                uc_raw = _limpar_digitos(tds[0].text.strip())
                uc_norm = _normalizar_uc(uc_raw)

                municipio = tds[1].text.strip() if len(tds) > 1 else ""
                endereco = tds[2].text.strip() if len(tds) > 2 else ""

                cb = None
                try:
                    cb = tds[3].find_element(By.XPATH, ".//input[@type='checkbox']")
                except Exception:
                    cb = None

                cb_id = cb.get_attribute("id") if cb else ""
                checked = cb.is_selected() if cb else False
                disabled_attr = cb.get_attribute("disabled") if cb else None
                disabled = disabled_attr is not None and str(disabled_attr).lower() != "false"

                if uc_raw:
                    ucs.append({
                        "uc_raw": uc_raw,
                        "uc_norm": uc_norm,
                        "nome": endereco or municipio or "",
                        "municipio": municipio,
                        "endereco": endereco,
                        "cb_id": cb_id,
                        "checked": checked,
                        "disabled": disabled,
                        "row_index": i,
                    })

            except StaleElementReferenceException:
                continue

    except Exception as exc:
        log.debug("[tabela] Erro ao ler UCs: %s", exc)

    return ucs


def _selecionar_uc(driver: webdriver.Chrome, uc_info: dict) -> bool:
    uc_raw = _limpar_digitos(uc_info.get("uc_raw", "") or uc_info.get("uc", ""))
    uc_norm = _normalizar_uc(uc_raw)

    if not uc_norm:
        log.warning("[UC] UC vazia para seleção")
        return False

    try:
        _aguardar_tabela_ucs(driver, timeout=12)

        linhas = driver.find_elements(
            By.XPATH,
            "//table[@id='CONTENT_gdEscolherClienteDoAgrupamento']//tr[td]"
        )

        for tr in linhas:
            try:
                tds = tr.find_elements(By.TAG_NAME, "td")
                if len(tds) < 4:
                    continue

                linha_uc_raw = _limpar_digitos(tds[0].text.strip())
                linha_uc_norm = _normalizar_uc(linha_uc_raw)

                if linha_uc_norm != uc_norm:
                    continue

                checkbox = tds[3].find_element(By.XPATH, ".//input[@type='checkbox']")
                checked = checkbox.is_selected()
                disabled_attr = checkbox.get_attribute("disabled")
                disabled = disabled_attr is not None and str(disabled_attr).lower() != "false"

                if checked:
                    log.info("[UC] UC=%s já estava selecionada", uc_raw or uc_norm)
                    return True

                if disabled and not checked:
                    log.warning("[UC] Checkbox da UC=%s está desabilitado sem marcar", uc_raw or uc_norm)
                    return False

                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", checkbox)
                time.sleep(0.3)

                try:
                    checkbox.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", checkbox)

                time.sleep(1.0)

                try:
                    return checkbox.is_selected() or True
                except Exception:
                    return True

            except StaleElementReferenceException:
                continue

    except Exception as exc:
        log.warning("[UC] Não foi possível selecionar UC=%s | %s", uc_raw or uc_norm, exc)
        return False

    log.warning("[UC] UC=%s não encontrada na tabela após normalização", uc_raw or uc_norm)
    return False


def _abrir_segunda_via(driver: webdriver.Chrome) -> bool:
    seletores = [
        (By.XPATH, "//a[contains(normalize-space(.),'2ª Via')]"),
        (By.XPATH, "//a[contains(normalize-space(.),'2 Via')]"),
        (By.XPATH, "//a[contains(normalize-space(.),'Segunda Via')]"),
        (By.XPATH, "//a[contains(@href,'SegundaVia') or contains(@href,'segunda-via') or contains(@href,'2via')]"),
        (By.LINK_TEXT, "2ª Via de Conta"),
        (By.PARTIAL_LINK_TEXT, "2ª Via"),
        (By.PARTIAL_LINK_TEXT, "Segunda Via"),
    ]

    for by, loc in seletores:
        try:
            _clicar(driver, by, loc, timeout=10)
            time.sleep(2)
            if _aguardar_tabela_faturas(driver, timeout=10):
                return True
        except Exception:
            pass

    log.warning("[2via] Link de Segunda Via não encontrado")
    return False


def _voltar_para_tela_de_ucs(driver: webdriver.Chrome) -> bool:
    try:
        if _aguardar_tabela_ucs(driver, timeout=3):
            return True
    except Exception:
        pass

    try:
        driver.back()
        if _aguardar_tabela_ucs(driver, timeout=8):
            return True
    except Exception:
        pass

    try:
        driver.get("https://www.eneldistribuicao.com.br/agencia/DefaultGa.aspx")
        if _aguardar_tabela_ucs(driver, timeout=10):
            return True
    except Exception:
        pass

    log.warning("[nav] Não foi possível retornar à tabela de UCs")
    return False


# =============================================================================
# FATURAS
# =============================================================================

def _ler_tabela_faturas(driver: webdriver.Chrome) -> List[dict]:
    time.sleep(1.5)
    faturas: List[dict] = []
    refs_vistas: set[str] = set()

    try:
        _aguardar_tabela_faturas(driver, timeout=12)

        linhas = driver.find_elements(
            By.XPATH,
            "//table[@id='CONTENT_gdHistoricoDeFaturamento']//tr[td]"
        )

        for i, tr in enumerate(linhas):
            try:
                tds = tr.find_elements(By.TAG_NAME, "td")
                if len(tds) < 5:
                    continue

                referencia_raw = tds[0].text.strip()
                vencimento = tds[1].text.strip()
                valor = tds[2].text.strip()
                situacao = tds[3].text.strip()

                checkbox = None
                try:
                    checkbox = tds[4].find_element(By.XPATH, ".//input[@type='checkbox']")
                except Exception:
                    checkbox = None

                chk_id = checkbox.get_attribute("id") if checkbox else ""
                ref = extrair_mes_ref(referencia_raw)

                if ref and ref in refs_vistas:
                    log.info("  [faturas] Linha duplicada ignorada para ref=%s", ref)
                    continue

                if ref:
                    refs_vistas.add(ref)

                faturas.append({
                    "row_index": i,
                    "referencia": ref,
                    "referencia_raw": referencia_raw,
                    "vencimento": vencimento,
                    "valor": valor,
                    "situacao": situacao,
                    "checkbox_id": chk_id,
                })

            except StaleElementReferenceException:
                continue

    except Exception as exc:
        log.debug("[faturas] Erro ao ler tabela: %s", exc)

    return faturas


def _todas_pagas(faturas: List[dict]) -> bool:
    if not faturas:
        return True
    return all(_situacao_esta_paga(f.get("situacao", "")) for f in faturas)


def _baixar_fatura(driver: webdriver.Chrome, fatura: dict) -> Optional[Path]:
    antes = {p.name for p in TEMP_DOWNLOAD_DIR.glob("*")}
    chk_id = fatura.get("checkbox_id", "")

    if not chk_id:
        log.warning("[download] Checkbox não encontrado para ref=%s", fatura.get("referencia"))
        return None

    try:
        chk = _wait(driver, 10).until(
            EC.presence_of_element_located((By.ID, chk_id))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", chk)
        time.sleep(0.3)

        if not chk.is_selected():
            try:
                chk.click()
            except Exception:
                driver.execute_script("arguments[0].click();", chk)

        time.sleep(0.5)
        _clicar(driver, By.ID, "CONTENT_btnEmitirSegundaVia", timeout=10)

    except Exception as exc:
        log.warning("[download] Falha ao marcar/imprimir ref=%s | %s", fatura.get("referencia"), exc)
        return None

    return _aguardar_novo_pdf(antes)


# =============================================================================
# PROCESSAMENTO
# =============================================================================

def processar_uc(
    driver: webdriver.Chrome,
    linha: LinhaAcesso,
    uc_info: dict,
    baixados: set[tuple[str, str]],
    master,
) -> List[dict]:
    uc_norm = linha.uc_norm
    uc_exibir = linha.uc_original
    resultados: List[dict] = []
    refs_processadas_no_lote: set[str] = set()

    log.info("UC=%s | %s", uc_exibir, uc_info.get("nome", "") or "-")

    if not _selecionar_uc(driver, uc_info):
        resultados.append({"uc": uc_exibir, "status": "ERRO", "mensagem": "Falha ao selecionar UC"})
        return resultados

    if not _abrir_segunda_via(driver):
        resultados.append({"uc": uc_exibir, "status": "ERRO", "mensagem": "2ª Via não encontrada"})
        return resultados

    faturas = _ler_tabela_faturas(driver)

    if not faturas:
        resultados.append({"uc": uc_exibir, "status": "SEM_FATURA", "mensagem": "Tabela de faturas vazia"})
        return resultados

    if _todas_pagas(faturas):
        log.info("  → Todas pagas — pulando UC")
        resultados.append({
            "uc": uc_exibir,
            "status": "TODAS_PAGAS",
            "mensagem": f"{len(faturas)} fatura(s) — todas pagas",
        })
        return resultados

    for fatura in faturas:
        ref = fatura.get("referencia")
        situacao = fatura.get("situacao", "")

        if _situacao_esta_paga(situacao):
            continue

        if not ref:
            resultados.append({
                "uc": uc_exibir,
                "status": "ERRO",
                "mensagem": "Fatura sem referência válida",
            })
            continue

        if ref in refs_processadas_no_lote:
            log.info("  → %s já apareceu nesta leitura — pulando duplicata", ref)
            resultados.append({
                "uc": uc_exibir,
                "mes_ref": ref,
                "status": "DUPLICADA_PORTAL",
                "mensagem": f"Duplicata ignorada na grade: {ref}",
            })
            continue

        if (uc_norm, ref) in baixados:
            log.info("  → %s já existe no índice — pulando", ref)
            _emit("skipped_existing", uc=uc_norm, ref=ref)
            resultados.append({
                "uc": uc_exibir,
                "mes_ref": ref,
                "status": "JA_EXISTE",
                "mensagem": f"Já baixado: {ref}",
            })
            continue

        log.info("  → Baixando %s (situação: %s)...", ref, situacao or "—")

        # ── Aguarda tabela estabilizar após eventual postback anterior ────────
        try:
            _aguardar_tabela_faturas(driver, timeout=10)
            time.sleep(0.5)
        except Exception:
            pass

        # ── Garante estado limpo — desmarca todos os checkboxes ───────────────
        try:
            todos_chks = driver.find_elements(
                By.XPATH,
                "//table[@id='CONTENT_gdHistoricoDeFaturamento']//input[@type='checkbox']"
            )
            for c in todos_chks:
                try:
                    if c.is_selected():
                        driver.execute_script("arguments[0].click();", c)
                        time.sleep(0.2)
                except Exception:
                    pass
        except Exception:
            pass

        pdf = _baixar_fatura(driver, fatura)
        if not pdf or not pdf.exists():
            resultados.append({
                "uc": uc_exibir,
                "mes_ref": ref,
                "status": "ERRO",
                "mensagem": "PDF não gerado",
            })
            # Pausa para portal resetar antes da próxima tentativa
            time.sleep(2.0)
            continue

        ref_pdf = extrair_mes_ref_do_pdf(pdf)
        if ref_pdf:
            ref = ref_pdf
            if ref in refs_processadas_no_lote or (uc_norm, ref) in baixados:
                log.warning("  → PDF retornou referência já processada (%s); ignorando duplicata", ref)
                try:
                    pdf.unlink(missing_ok=True)
                except Exception:
                    pass
                resultados.append({
                    "uc": uc_exibir,
                    "mes_ref": ref,
                    "status": "DUPLICADA_PDF",
                    "mensagem": f"PDF duplicado ignorado: {ref}",
                })
                time.sleep(1.0)
                continue

        subtipo = classificar_pdf(pdf)
        carimbo = master.consumir_carimbo()

        pasta_dest = BASE_DIR / ref / subtipo
        pasta_dest.mkdir(parents=True, exist_ok=True)
        destino = pasta_dest / f"{carimbo}.pdf"
        shutil.move(str(pdf), str(destino))

        master.registrar(
            indice_bb=carimbo,
            sistema="ENEL_RJ",
            uc=uc_norm,
            mes_ref=ref,
            fatura_id=fatura.get("checkbox_id", "")[:60],
            cnpj=linha.cnpj,
            estado="RIO DE JANEIRO",
            instalacao=uc_norm,
            arquivo=str(destino),
        )

        _registrar_indice_local(
            indice_bb=carimbo,
            uc=uc_norm,
            subtipo=subtipo,
            mes_ref=ref,
            fatura_id=fatura.get("checkbox_id", "")[:60],
            caminho=str(destino),
        )
        baixados.add((uc_norm, ref))
        refs_processadas_no_lote.add(ref)
        _emit("downloaded", uc=uc_norm, ref=ref, carimbo=carimbo)

        log.info("  ✓ %s → %s/%s/%s.pdf", ref, ref, subtipo, carimbo)
        resultados.append({
            "uc": uc_exibir,
            "mes_ref": ref,
            "subtipo": subtipo,
            "status": "OK",
            "indice": carimbo,
            "arquivo": str(destino),
        })

        # ── Pausa para portal resetar os checkboxes antes da próxima fatura ──
        time.sleep(1.5)

    return resultados


# =============================================================================
# MAIN
# =============================================================================

def main(*, preflight: bool = False) -> None:
    global BASE_DIR, PLANILHA_PATH, INDICE_LOCAL_PATH, LOG_DIR
    try:
        BASE_DIR.mkdir(parents=True, exist_ok=True)
        LOG_DIR.mkdir(parents=True, exist_ok=True)
    except OSError:
        BASE_DIR = FALLBACK_BASE_DIR
        LOG_DIR = BASE_DIR / "LOGS"
        PLANILHA_PATH = Path(__file__).resolve().parent / "senhas_enel_rj.xlsx"
        INDICE_LOCAL_PATH = BASE_DIR / "indice_faturas_enel_rj.csv"
        BASE_DIR.mkdir(parents=True, exist_ok=True)
        LOG_DIR.mkdir(parents=True, exist_ok=True)
    TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    log.info("=" * 62)
    log.info("ENEL RJ — DOWNLOAD DE FATURAS  %s", datetime.now().strftime("%d/%m/%Y %H:%M"))
    log.info("Base dir   : %s", BASE_DIR)
    log.info("Planilha   : %s", PLANILHA_PATH)
    log.info("Índice     : %s", INDICE_LOCAL_PATH)
    log.info("Temp local : %s", TEMP_DOWNLOAD_DIR)
    log.info("=" * 62)

    grupos = carregar_planilha(PLANILHA_PATH)
    master = None if preflight else _carregar_master()
    baixados = set() if preflight else _carregar_indice_local()

    resultados_todos: List[dict] = []
    total_grupos = len(grupos)

    for idx_grupo, ((login, senha), ucs_grupo) in enumerate(grupos.items(), 1):
        log.info("")
        log.info("─" * 62)
        log.info("GRUPO %d/%d | login=%s | %d UC(s)", idx_grupo, total_grupos, login, len(ucs_grupo))
        log.info("─" * 62)

        driver = criar_driver()
        try:
            if not _fazer_login(driver, login, senha):
                log.error("GRUPO %s — falha no login", login)
                for u in ucs_grupo:
                    resultados_todos.append({
                        "uc": u.uc_original,
                        "login": login,
                        "status": "ERRO_LOGIN",
                        "mensagem": "Falha no login",
                    })
                continue

            _fechar_modal_pos_login(driver)

            ucs_portal = _obter_ucs_da_tabela(driver)
            log.info("Portal retornou %d UC(s) na tabela", len(ucs_portal))
            log.debug("UCs lidas do portal: %s", [u["uc_raw"] for u in ucs_portal])
            if preflight:
                if not ucs_portal:
                    raise RuntimeError("preflight: tabela inicial de UCs vazia")
                log.info("PREFLIGHT_PASS: login e tabela de UCs validados; nenhum download iniciado.")
                return

            for idx_uc, linha in enumerate(ucs_grupo, 1):
                if idx_uc > 1:
                    _voltar_para_tela_de_ucs(driver)
                    _fechar_modal_pos_login(driver)
                    ucs_portal = _obter_ucs_da_tabela(driver)

                uc_info = next(
                    (u for u in ucs_portal if u.get("uc_norm") == linha.uc_norm),
                    None,
                )

                if uc_info is None:
                    log.warning(
                        "UC=%s não encontrada na tabela do portal — tentando direto pela UC normalizada=%s",
                        linha.uc_original,
                        linha.uc_norm,
                    )
                    uc_info = {
                        "uc_raw": linha.uc_original,
                        "uc_norm": linha.uc_norm,
                        "nome": linha.nome,
                        "checked": False,
                        "disabled": False,
                    }

                try:
                    res = processar_uc(driver, linha, uc_info, baixados, master)
                    resultados_todos.extend(res)

                    for r in res:
                        st = r["status"]
                        if st == "OK":
                            log.info(
                                "✓ UC=%s | %s → %s | %s",
                                r["uc"],
                                r.get("mes_ref", "?"),
                                r.get("subtipo", "?"),
                                r.get("indice", "?"),
                            )
                        elif st in ("JA_EXISTE", "TODAS_PAGAS"):
                            log.info(
                                "= UC=%s | %s → %s",
                                r["uc"],
                                r.get("mes_ref", ""),
                                r.get("status"),
                            )
                        else:
                            log.warning("✗ UC=%s → %s", r["uc"], r.get("mensagem", ""))

                except Exception as exc:
                    log.warning("EXCEÇÃO UC=%s: %s", linha.uc_original, exc)
                    log.debug("TRACEBACK:", exc_info=True)
                    resultados_todos.append({
                        "uc": linha.uc_original,
                        "status": "ERRO",
                        "mensagem": f"Exceção: {exc}",
                    })

                time.sleep(PAUSA_ENTRE_UCS)

        finally:
            try:
                driver.quit()
            except Exception:
                pass

    ok = sum(1 for r in resultados_todos if r["status"] == "OK")
    ja = sum(1 for r in resultados_todos if r["status"] == "JA_EXISTE")
    pg = sum(1 for r in resultados_todos if r["status"] == "TODAS_PAGAS")
    sf = sum(1 for r in resultados_todos if r["status"] == "SEM_FATURA")
    er = sum(1 for r in resultados_todos if r["status"] not in ("OK", "JA_EXISTE", "TODAS_PAGAS", "SEM_FATURA"))
    tot = len(resultados_todos)

    log.info("")
    log.info("=" * 62)
    log.info("FINALIZADO  %s", datetime.now().strftime("%d/%m/%Y %H:%M"))
    log.info("✓ OK          : %d", ok)
    log.info("= JÁ EXISTE   : %d", ja)
    log.info("- TODAS PAGAS : %d", pg)
    log.info("~ SEM FATURA  : %d", sf)
    log.info("✗ ERRO        : %d", er)
    log.info("TOTAL         : %d", tot)
    log.info("=" * 62)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Downloader ENEL RJ")
    parser.add_argument("--preflight", action="store_true")
    cli_args = parser.parse_args()
    main(preflight=cli_args.preflight)
