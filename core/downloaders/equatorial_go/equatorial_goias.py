#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Download Selenium — EQUATORIAL GOIÁS
=====================================

Fluxo por UC:
  1. Login: instalação + CNPJ → botão Entrar
  2. Fecha modal pós-login (X ou OK)
  3. Menu Contas → Segunda Via de Fatura
  4. Tipo=Completa, Motivo=Outros → Emitir
  5. Download → aguarda PDF local → OK modal pós-download
  6. Move PDF para rede: DOWNLOAD EQUATORIAL\MM-AAAA\BT|MT\BB_xxxxxxx.pdf
  7. Registra no índice local CSV e no master

Notas técnicas:
  - Driver Chrome único por execução — limpeza agressiva entre UCs via CDP
  - Reinício automático do driver a cada REINICIAR_DRIVER_A_CADA UCs (padrão 15)
    ou após MAX_ERROS_CONSECUTIVOS falhas seguidas (padrão 3)
  - Duplicados verificados exclusivamente pelo índice local (INSTALACAO + MES_REF)
  - Chrome não salva PDFs em UNC: usa %LOCALAPPDATA%\equatorial_temp como intermediário
  - Alerta #002 (servidor instável) → refresh sem contar tentativa
  - Log terminal: separadores visuais por UC, fases nomeadas, status compacto ao final
  - Usa undetected-chromedriver (UC) para evitar detecção de bot pelo portal Equatorial

Dependências: pip install selenium pandas openpyxl undetected-chromedriver
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
import _venv_check  # noqa

try:
    from core.metrics.radar_metrics import emit_outcome as _emit_eq_outcome
    def _emit(outcome: str, *, instalacao: str, mes_ref: str, carimbo: str = "") -> None:
        _emit_eq_outcome(outcome, utility="EQUATORIAL GO", account_id=instalacao,
                         competence=mes_ref, invoice_id=carimbo or mes_ref)
except Exception:
    def _emit(outcome: str, **_: str) -> None:  # type: ignore[misc]
        pass

import csv
import importlib.util
import logging
import os
import random
import re
import shutil
import sys
import time
import traceback
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

# undetected-chromedriver evita detecção de bot pelo portal Equatorial.
# Fallback para Selenium padrão se UC não estiver instalado.
try:
    import undetected_chromedriver as uc
    _UC_DISPONIVEL = True
except ImportError:
    _UC_DISPONIVEL = False

from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoAlertPresentException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver import ChromeOptions
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait


# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

URL_LOGIN = "https://goias.equatorialenergia.com.br/LoginGO.aspx"

BASE_DIR              = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/DOWNLOAD EQUATORIAL")
PLANILHA_PATH         = BASE_DIR / "Acessos Equatorial.xlsx"
PLANILHA_FATURAS_PATH = BASE_DIR / "Faturas Equatoriais BB - 2026.xlsx"
MASTER_PY_PATH    = Path(__file__).resolve().parent.parent.parent.parent / "scripts" / "indice_master.py"
INDICE_LOCAL_PATH = BASE_DIR / "indice_downloads_equatorial.csv"
LOG_DIR           = BASE_DIR / "LOGS"
RESUMO_XLSX       = LOG_DIR / "resumo_execucao_equatorial.xlsx"

# Chrome não salva em UNC — pasta local como intermediário
TEMP_DOWNLOAD_DIR = Path(os.environ.get("LOCALAPPDATA", r"C:\Users\Public")) / "equatorial_temp"

HEADLESS                = False
TIMEOUT_PADRAO          = 25
TIMEOUT_DOWNLOAD        = 45
TIMEOUT_MODAL_OK        = 45
LOGIN_MAX_TENTATIVAS    = 3
MAX_ALERTAS_SERVIDOR    = 4
MAX_ERROS_CONSECUTIVOS  = 3   # reinicia driver após N erros seguidos
REINICIAR_DRIVER_A_CADA = 15  # reinicia driver a cada N UCs (0 = nunca)
PAUSA_PREVENTIVA_A_CADA = 0   # 0 = desabilitado
PAUSA_PREVENTIVA_SEG    = 90  # duração da pausa preventiva em segundos
CONCESSIONARIA_FILTRO   = "EQUATORIAL"
MIN_DIGITOS_INSTALACAO  = 5
MIN_DIGITOS_CNPJ        = 11
PAUSA_ENTRE_UCS_MIN     = 5.0
PAUSA_ENTRE_UCS_MAX     = 10.0
CHROMEDRIVER_PATH: Optional[str] = None
FILTRO_INSTALACOES_CSV_PADRAO = Path(__file__).resolve().parent / "equatorial_go_acessos_especificos.csv"
TIMEOUT_PAGE_LOAD       = 70
TIMEOUT_LOGIN_RESULTADO = 45

LOGIN_UC_SELECTORS = [
    (By.ID, "WEBDOOR_headercorporativogo_txtUC"),
    (By.CSS_SELECTOR, "input[id*='txtUC']"),
    (By.CSS_SELECTOR, "input[name*='txtUC']"),
]

LOGIN_DOC_SELECTORS = [
    (By.ID, "WEBDOOR_headercorporativogo_txtDocumento"),
    (By.CSS_SELECTOR, "input[id*='txtDocumento']"),
    (By.CSS_SELECTOR, "input[name*='txtDocumento']"),
]

LOGIN_BTN_SELECTORS = [
    (By.XPATH, "//button[@type='button' and contains(@onclick,'ValidarCamposAreaLogada')]"),
    (By.XPATH, "//input[@type='submit' and contains(@value,'Entrar')]"),
    (By.XPATH, "//button[contains(normalize-space(.),'Entrar')]"),
]

LOGIN_OK_SELECTORS = [
    (By.ID, "LinkSegundaVia"),
    (By.CSS_SELECTOR, "label[for='A']"),
    (By.XPATH, "//h1[contains(normalize-space(.),'Segunda Via de Conta')]"),
    (By.XPATH, "//strong[contains(normalize-space(.),'BANCO DO BRASIL SA')]"),
]

LOGIN_ERROS_SERVIDOR = (
    "#002",
    "nao foi possivel",
    "tente novamente mais tarde",
    "instabilidade",
    "temporariamente indisponivel",
    "erro interno",
)

LOGIN_ERROS_CREDENCIAL = (
    "cliente nao encontrado",
    "instalacao nao encontrada",
    "documento nao encontrado",
    "documento invalido",
    "dados invalidos",
    "dados incorretos",
)

# ---------------------------------------------------------------------------
# FILTRO DE INSTALAÇÕES ESPECÍFICAS
# Quando preenchido, processa APENAS estas instalações (ignora o resto da
# planilha). Chave = instalação (só dígitos), valor = CNPJ (só dígitos).
# Valor vazio ("") → usa o CNPJ da planilha; se não tiver → 00000000000191.
# Para processar TODAS as UCs da planilha, deixe o dicionário vazio: {}
# ---------------------------------------------------------------------------
FILTRO_INSTALACOES: dict[str, str] = {
    "1160024021":   "00000000549339",
    "580018258":    "00000000000191",
    "1390023014":   "00000000065994",
    "630019496":    "00000000000191",
    "610012757":    "00000000037788",
    "20239361":     "00000000253480",
    "910037139":    "00000000000191",
    "60000806":     "00000000000191",
    "930009009":    "00000000000191",
    "16641942":     "00000000385808",
    "1990013354":   "00000000000191",
    "13153237":     "00000000538809",
    "220023360":    "00000000133159",
    "16624269":     "00000000405760",
    "10008724987":  "00000000551902",
    "170023000":    "00000000000191",
    "1480004343":   "00000000000191",
    "21290933":     "00000000000191",
    "400063165":    "00000000000191",
    "10008080265":  "00000000000191",
    "730032528":    "00000000000191",
    "80020331":     "00000000000191",
    "150002907":    "00000000000191",
    "190131469":    "00000000572900",
    "1290040737":   "00000000000191",
    "320018167":    "00000000000191",
    "10012035775":  "00000000331546",
    "10007883097":  "00000000000191",
    "12509905":     "00000000000191",
    "2300004209":   "00000000000191",
    "90061330":     "00000000000191",
    "290004731":    "00000000000191",
    "690175103":    "00000000037605",
    "460085141":    "00000000253480",
    "140027555":    "00000000000191",
    "30059604":     "00000000000191",
    "840025786":    "00000000000191",
    "590016684":    "00000000000191",
    "640015645":    "00000000000191",
    "70002952":     "00000000064246",
    "550083959":    "00000000000191",
    "980013185":    "00000000000191",
    "10235954":     "00000000075280",
    "830034183":    "00000000000191",
    "780006811":    "00000000000191",
    "1140024647":   "00000000000191",
    "1930000896":   "00000000000191",
    "660065721":    "00000000098817",
    "1060036608":   "00000000000191",
    "2260019359":   "00000000000191",
    "10005869070":  "00000000220558",
    "760239680":    "00000000159204",
    "15311739":     "00000000000191",
    "13601751":     "00000000000191",
    "10003084548":  "00000000000191",
    "1420016780":   "00000000265578",
    "2960085298":   "00000000000191",
    "10003283257":  "00000000000191",
    "1880018448":   "00000000253480",
    "10002677421":  "00000000000191",
    "2860301155":   "00000000410683",
    "10006579270":  "00000000000191",
    "10011847970":  "00000000000191",
    "13161787":     "00000000400610",
    "1470004139":   "00000000000191",
    "10072524":     "00000000000191",
    "16280994":     "00000000000191",
    "17222400":     "00000000549681",
    "10014542380":  "00000000000191",
    "10008843250":  "00000000261580",
    "10008311925":  "00000000000191",
    "10012318424":  "00000000728691",
    "120105056":    "00000000000191",
    "10015559007":  "00000000747483",
    "480031605":    "00000000307165",
    "1210017651":   "00000000000191",
    "10010440702":  "00000000000191",
    "10009314456":  "00000000000191",
    "10023108973":  "00000000000191",
    "2040015456":   "00000000000191",
    "9880010509":   "00000000000191",
    "2250005145":   "00000000000191",
    "10013577490":  "00000000000191",
    "2310006161":   "00000000000191",
    "160066116":    "00000000000191",
    "570015078":    "00000000000191",
    "13297053":     "00000000000191",
    "10007299611":  "00000000451959",
    "2810014250":   "00000000000191",
    "10018828181":  "05339508000132",
    "13477638":     "00000000000191",
    "13548761":     "00000000000191",
    "91386101201":  "00000000000191",  # 913.861.012-01
    "1180002456":   "00000000000191",
    "690365676":    "00000000000191",
    "13550860":     "00000000000191",
    "10010462277":  "00000000503509",
    "311926701232": "00000000000191",  # 3.119.267.012-32
    "2010002370":   "00000000000191",
    "420003459":    "00000000000191",
    "8570039398":   "00000000549410",
    "10026269935":  "00000000542407",
    "2840007235":   "00000000000191",
    "3870021261":   "00000000542911",
    "1440009007":   "00000000000191",
    "10028591168":  "00000000000191",
    "13876107":     "00000000585807",
    "10015434255":  "00000000000191",
    "10016395865":  "00000000742686",
    "10507220":     "00000000000191",
    "10001928722":  "00000000000191",
    "1090012187":   "00000000000191",
    "410002197":    "00000000078034",
    "2490006259":   "00000000000191",
    "1080007500":   "00000000000191",
    "290157900":    "00000000000191",
    "2760017662":   "00000000000191",
    "230013004":    "00000000000191",
    "2170006625":   "00000000448628",
}


# =============================================================================
# LOGGING
# Console: limpo e por fases | Arquivo: DEBUG completo
# =============================================================================

class _ConsoleFmt(logging.Formatter):
    """Formata mensagens de console com prefixo visual por nível."""
    _PREF = {
        logging.DEBUG:    "      ",
        logging.INFO:     "  ",
        logging.WARNING:  "  ⚠  ",
        logging.ERROR:    "  ✗  ",
        logging.CRITICAL: "  ✗✗ ",
    }

    def format(self, record: logging.LogRecord) -> str:
        pref = self._PREF.get(record.levelno, "  ")
        ts   = datetime.fromtimestamp(record.created).strftime("%H:%M:%S")
        return f"{ts} {pref}{record.getMessage()}"


def _configurar_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    import io
    stdout_utf8 = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)
    h_con = logging.StreamHandler(stdout_utf8)
    h_con.setFormatter(_ConsoleFmt())
    h_con.setLevel(logging.INFO)

    h_arq = logging.FileHandler(
        str(LOG_DIR / f"equatorial_{ts}.log"), encoding="utf-8"
    )
    h_arq.setFormatter(
        logging.Formatter("%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
    )
    h_arq.setLevel(logging.DEBUG)

    log = logging.getLogger("equatorial")
    log.setLevel(logging.DEBUG)
    log.handlers.clear()
    log.addHandler(h_con)
    log.addHandler(h_arq)
    return log


logger = _configurar_logging()

# Separadores e atalhos de log ────────────────────────────────────────────────
_SEP  = "─" * 64
_SEP2 = "═" * 64


def _fase(msg: str) -> None:
    """Etapa do fluxo — aparece no terminal como linha de progresso."""
    logger.info("  ▸ %s", msg)


def _resultado_ok(inst: str, ref: str, indice: str, nome_arq: str) -> None:
    logger.info("  ✓  %-14s  %s  →  %s  (%s)", inst, ref, indice, nome_arq)


def _resultado_ja(inst: str, ref: str) -> None:
    logger.info("  =  %-14s  %s  (já baixado)", inst, ref)


def _resultado_sf(inst: str, motivo: str) -> None:
    logger.info("  -  %-14s  sem fatura — %s", inst, motivo)


def _resultado_erro(inst: str, motivo: str) -> None:
    logger.warning("  ✗  %-14s  ERRO: %s", inst, motivo)


class SessaoChromeInstavel(RuntimeError):
    """Sinaliza que o Chrome precisa ser recriado antes de repetir a UC."""


def _erro_sessao_driver(exc: Exception) -> bool:
    msg = normalizar_texto(exc)
    sinais = (
        "invalid session id",
        "no such window",
        "target window already closed",
        "web view not found",
        "session deleted because of page crash",
        "disconnected: not connected to devtools",
    )
    return any(s in msg for s in sinais)


# =============================================================================
# UTILITÁRIOS
# =============================================================================

def normalizar_texto(s) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s.strip()).encode("ASCII", "ignore").decode("ASCII")
    return " ".join(s.split()).lower()


def limpar_numero(s) -> str:
    return re.sub(r"\D+", "", "" if s is None else str(s))


def encontrar_coluna(df: pd.DataFrame, opcoes: list[str]) -> Optional[str]:
    mapa = {normalizar_texto(c): c for c in df.columns}
    for o in opcoes:
        if normalizar_texto(o) in mapa:
            return mapa[normalizar_texto(o)]
    return None


def classificar_tensao(valor: str) -> str:
    """'Média Tensão' → MT | qualquer outra coisa → BT"""
    t = normalizar_texto(valor)
    return "MT" if "media" in t or t == "mt" else "BT"


def pausa_humana(mn: float = 0.3, mx: float = 0.8) -> None:
    mu    = (mn + mx) / 2
    sigma = (mx - mn) / 6 if mx > mn else 0.01
    time.sleep(max(mn, min(mx, random.gauss(mu, sigma))))


def _valida_ref(ref: Optional[str]) -> Optional[str]:
    if not ref:
        return None
    return ref if re.fullmatch(r"(0[1-9]|1[0-2])-20\d{2}", ref) else None


def extrair_referencia(texto: str) -> Optional[str]:
    """
    Extrai MM-YYYY de uma string curta.
    Ordem: abreviado (MAR/2026) → extenso (MARÇO/2026) → numérico (03/2026).
    Sempre valida antes de retornar.
    """
    if not texto:
        return None
    bruto  = str(texto).strip()
    txt_up = unicodedata.normalize("NFKD", bruto).encode("ASCII", "ignore").decode("ASCII").upper()

    MESES_ABREV = {
        "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
        "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
        "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12",
    }
    m = re.search(r"\b(" + "|".join(MESES_ABREV) + r")[/\-. ](20\d{2})\b", txt_up)
    if m:
        return _valida_ref(f"{MESES_ABREV[m.group(1)]}-{m.group(2)}")

    MESES_EXT = {
        "JANEIRO": "01", "FEVEREIRO": "02", "MARCO": "03", "ABRIL": "04",
        "MAIO": "05", "JUNHO": "06", "JULHO": "07", "AGOSTO": "08",
        "SETEMBRO": "09", "OUTUBRO": "10", "NOVEMBRO": "11", "DEZEMBRO": "12",
    }
    m = re.search(r"\b(" + "|".join(MESES_EXT) + r")[ /\-]+(20\d{2})\b", txt_up)
    if m:
        return _valida_ref(f"{MESES_EXT[m.group(1)]}-{m.group(2)}")

    if len(bruto) <= 30:
        m = re.search(r"\b(0[1-9]|1[0-2])[/-](20\d{2})\b", bruto)
        if m:
            return _valida_ref(f"{m.group(1)}-{m.group(2)}")

    return None


# =============================================================================
# ÍNDICE LOCAL
# =============================================================================

def _carregar_indice_local() -> set[tuple[str, str]]:
    baixados: set[tuple[str, str]] = set()
    if not INDICE_LOCAL_PATH.exists():
        return baixados
    with open(INDICE_LOCAL_PATH, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            inst = (row.get("INSTALACAO") or "").strip()
            ref  = (row.get("MES_REF")    or "").strip()
            if inst and ref:
                baixados.add((inst, ref))
    return baixados


def ja_foi_baixado_local(baixados: set[tuple[str, str]],
                         instalacao: str, mes_ref: str) -> bool:
    return (instalacao.strip(), mes_ref.strip()) in baixados


def registrar_indice_local(indice_bb: str, instalacao: str, mes_ref: str,
                           arquivo: str, tensao: str) -> None:
    novo = not INDICE_LOCAL_PATH.exists()
    with open(INDICE_LOCAL_PATH, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[
            "INDICE", "INSTALACAO", "MES_REF", "ARQUIVO", "TENSAO", "DATA_DOWNLOAD",
        ])
        if novo:
            w.writeheader()
        w.writerow({
            "INDICE":        indice_bb,
            "INSTALACAO":    instalacao.strip(),
            "MES_REF":       mes_ref.strip(),
            "ARQUIVO":       arquivo,
            "TENSAO":        tensao.strip(),
            "DATA_DOWNLOAD": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        })


# =============================================================================
# MASTER INDEX
# =============================================================================

def carregar_master():
    if not MASTER_PY_PATH.exists():
        raise FileNotFoundError(f"indice_master.py não encontrado: {MASTER_PY_PATH}")
    spec = importlib.util.spec_from_file_location("indice_master", str(MASTER_PY_PATH))
    mod  = importlib.util.module_from_spec(spec)
    sys.modules["indice_master"] = mod
    spec.loader.exec_module(mod)
    logger.info("Master carregado de %s", MASTER_PY_PATH)
    return mod.MasterIndice()


# =============================================================================
# PLANILHA
# =============================================================================

@dataclass(frozen=True)
class LinhaAcesso:
    instalacao:     str
    cnpj:           str
    tensao:         str
    concessionaria: str


def _carregar_filtro_instalacoes_csv(caminho: Path) -> dict[str, str]:
    if not caminho.exists():
        raise FileNotFoundError(f"Lista específica não encontrada: {caminho}")

    texto = caminho.read_text(encoding="utf-8-sig").strip()
    if not texto:
        return {}

    try:
        dialect = csv.Sniffer().sniff(texto[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(texto.splitlines(), dialect=dialect)
    if reader.fieldnames is None:
        raise ValueError(f"CSV sem cabeçalho: {caminho}")

    mapa_campos = {normalizar_texto(nome): nome for nome in reader.fieldnames}
    col_inst = (
        mapa_campos.get("instalacao")
        or mapa_campos.get("instalacao/uc")
        or mapa_campos.get("uc")
    )
    col_cnpj = (
        mapa_campos.get("cnpj")
        or mapa_campos.get("cpf/cnpj")
        or mapa_campos.get("documento")
    )
    if not col_inst or not col_cnpj:
        raise ValueError(
            f"CSV precisa ter colunas Instalacao e CNPJ: {caminho}"
        )

    filtro: dict[str, str] = {}
    for row in reader:
        inst = limpar_numero(row.get(col_inst, "") or "")
        if not inst:
            continue
        cnpj = limpar_numero(row.get(col_cnpj, "") or "") or "00000000000191"
        filtro[inst] = cnpj
    return filtro


def _resolver_filtro_instalacoes(args: argparse.Namespace) -> tuple[dict[str, str], str]:
    if args.sem_filtro:
        return {}, "sem filtro especifico"
    if args.lista_especifica:
        caminho = Path(args.lista_especifica).expanduser()
        return _carregar_filtro_instalacoes_csv(caminho), f"lista especifica: {caminho}"
    if FILTRO_INSTALACOES_CSV_PADRAO.exists():
        return (
            _carregar_filtro_instalacoes_csv(FILTRO_INSTALACOES_CSV_PADRAO),
            f"lista especifica padrao: {FILTRO_INSTALACOES_CSV_PADRAO}",
        )
    return FILTRO_INSTALACOES.copy(), "filtro hardcoded no downloader"


def _resolver_colunas(df: pd.DataFrame) -> dict[str, str]:
    col_inst   = encontrar_coluna(df, ["Instalacao", "Instalação", "UC", "Unidade Consumidora"])
    col_cnpj   = encontrar_coluna(df, ["CNPJ", "CPF/CNPJ", "Documento"])
    col_tensao = encontrar_coluna(df, ["Tensão", "Tensao"])
    col_conc   = encontrar_coluna(df, ["Concessionária", "Concessionaria"])

    faltando = [n for n, v in [
        ("Instalacao", col_inst), ("CNPJ", col_cnpj),
        ("Tensão", col_tensao), ("Concessionária", col_conc),
    ] if v is None]
    if faltando:
        raise ValueError(
            f"Colunas não encontradas: {faltando} | Disponíveis: {list(df.columns)}"
        )
    return {
        "instalacao": col_inst, "cnpj": col_cnpj,
        "tensao": col_tensao,   "concessionaria": col_conc,
    }


def carregar_unidades(planilha: Path, filtro_instalacoes: Optional[dict[str, str]] = None) -> list[LinhaAcesso]:
    if not planilha.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {planilha}")

    wb  = load_workbook(planilha, read_only=True)
    aba = None
    for nome in wb.sheetnames:
        ws   = wb[nome]
        hdrs = [normalizar_texto(c.value) for c in next(ws.iter_rows(max_row=1))]
        if (any(h in hdrs for h in ["instalacao", "uc"]) and
                any(h in hdrs for h in ["cnpj", "cpf/cnpj", "documento"])):
            aba = nome
            break
    wb.close()
    if aba is None:
        aba = load_workbook(planilha, read_only=True).sheetnames[0]

    df   = pd.read_excel(planilha, sheet_name=aba, dtype=str)
    cols = _resolver_colunas(df)

    df["_conc"] = df[cols["concessionaria"]].map(normalizar_texto)
    df = df[df["_conc"].str.contains(normalizar_texto(CONCESSIONARIA_FILTRO), na=False)].copy()

    # Mapa auxiliar: instalação → (cnpj, tensao, concessionaria) da planilha
    planilha_map: dict[str, tuple[str, str, str]] = {}
    for _, row in df.iterrows():
        inst = limpar_numero(row.get(cols["instalacao"], "") or "")
        if inst:
            planilha_map[inst] = (
                limpar_numero(row.get(cols["cnpj"], "") or ""),
                str(row.get(cols["tensao"]) or "").strip(),
                str(row.get(cols["concessionaria"]) or "").strip(),
            )

    filtro_instalacoes = filtro_instalacoes or {}

    if filtro_instalacoes:
        # Modo filtro: processa somente as instalações da lista
        linhas: list[LinhaAcesso] = []
        for inst, cnpj_filtro in filtro_instalacoes.items():
            inst_limpa = limpar_numero(inst)
            if not inst_limpa:
                continue
            # CNPJ: usa o do filtro; se vazio tenta planilha; senão 191
            if cnpj_filtro:
                cnpj = cnpj_filtro
            else:
                cnpj = planilha_map.get(inst_limpa, ("", "", ""))[0] or "00000000000191"
            # Tensão: usa da planilha quando disponível; senão BT
            tensao       = planilha_map.get(inst_limpa, ("", "BT", ""))[1] or "BT"
            concessionaria = planilha_map.get(inst_limpa, ("", "", "EQUATORIAL"))[2] or "EQUATORIAL"
            linhas.append(LinhaAcesso(
                instalacao=inst_limpa,
                cnpj=cnpj,
                tensao=tensao,
                concessionaria=concessionaria,
            ))
        logger.info("Planilha: %d UCs no filtro específico", len(linhas))
        return linhas

    # Modo normal: todas as UCs da planilha
    linhas = []
    for inst, (cnpj, tensao, conc) in planilha_map.items():
        if not cnpj:
            continue
        linhas.append(LinhaAcesso(
            instalacao=inst,
            cnpj=cnpj,
            tensao=tensao,
            concessionaria=conc,
        ))

    logger.info("Planilha: %d UCs EQUATORIAL encontradas", len(linhas))
    return linhas


# =============================================================================
# DRIVER
# =============================================================================

def _aplicar_stealth(driver: webdriver.Chrome) -> None:
    try:
        driver.execute_cdp_cmd("Network.enable", {})
    except Exception:
        pass

    try:
        driver.execute_cdp_cmd(
            "Network.setExtraHTTPHeaders",
            {"headers": {"Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7"}},
        )
    except Exception:
        pass

    try:
        driver.execute_cdp_cmd("Emulation.setTimezoneOverride", {"timezoneId": "America/Sao_Paulo"})
    except Exception:
        pass

    try:
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                    Object.defineProperty(navigator, 'platform', {get: () => 'Win32'});
                    Object.defineProperty(navigator, 'languages', {get: () => ['pt-BR', 'pt', 'en-US', 'en']});
                    Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
                    window.chrome = window.chrome || { runtime: {} };

                    const originalQuery = window.navigator.permissions && window.navigator.permissions.query;
                    if (originalQuery) {
                        window.navigator.permissions.query = (parameters) => (
                            parameters && parameters.name === 'notifications'
                                ? Promise.resolve({ state: Notification.permission })
                                : originalQuery(parameters)
                        );
                    }
                """
            },
        )
    except Exception:
        pass


def criar_driver() -> webdriver.Chrome:
    """
    Cria driver Chrome para o portal Equatorial.

    Usa undetected-chromedriver (UC) quando disponível — ele patcha o binário
    do ChromeDriver para remover assinaturas de automação que causam o erro #002.
    Fallback para Selenium padrão se UC não estiver instalado.

    Service instanciado a cada chamada para evitar reuso de handle morto.
    """
    TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    prefs = {
        "download.default_directory":                                 str(TEMP_DOWNLOAD_DIR),
        "download.prompt_for_download":                               False,
        "download.directory_upgrade":                                 True,
        "download.open_pdf_in_system_reader":                         False,
        "plugins.always_open_pdf_externally":                         True,
        "safebrowsing.enabled":                                       True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_setting_values.notifications":       2,
        "profile.default_content_settings.popups":                    0,
    }

    if _UC_DISPONIVEL:
        logger.info("  [driver] usando undetected-chromedriver (anti-bot)")
        opts = uc.ChromeOptions()
        opts.add_experimental_option("prefs", prefs)
        # NÃO usar --incognito com UC: incógnito ignora prefs de download,
        # causando o "Salvar como". UC já cria perfil temporário isolado.
        # excludeSwitches/useAutomationExtension são gerenciados internamente pelo UC.
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--disable-infobars")
        opts.add_argument("--disable-notifications")
        opts.add_argument("--disable-popup-blocking")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--no-first-run")
        opts.add_argument("--no-default-browser-check")
        opts.add_argument("--disable-default-apps")
        opts.add_argument("--lang=pt-BR")
        opts.add_argument("--log-level=3")
        if HEADLESS:
            opts.add_argument("--headless=new")
            opts.add_argument("--window-size=1920,1080")

        uc_kwargs: dict = {"options": opts, "use_subprocess": True, "version_main": 149}
        if CHROMEDRIVER_PATH:
            uc_kwargs["driver_executable_path"] = CHROMEDRIVER_PATH
        driver = uc.Chrome(**uc_kwargs)
    else:
        logger.warning("  [driver] undetected-chromedriver não instalado — "
                       "usando Selenium padrão (pode causar erro #002)")
        opts = ChromeOptions()
        opts.add_experimental_option("prefs", prefs)
        opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--disable-infobars")
        opts.add_argument("--disable-notifications")
        opts.add_argument("--disable-popup-blocking")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--no-first-run")
        opts.add_argument("--no-default-browser-check")
        opts.add_argument("--disable-default-apps")
        opts.add_argument("--lang=pt-BR")
        opts.add_argument("--log-level=3")
        opts.add_argument("--silent")
        if HEADLESS:
            opts.add_argument("--headless=new")
            opts.add_argument("--window-size=1920,1080")

        svc = Service(CHROMEDRIVER_PATH) if CHROMEDRIVER_PATH else Service()
        driver = webdriver.Chrome(service=svc, options=opts)

    driver.set_page_load_timeout(TIMEOUT_PAGE_LOAD)
    driver.implicitly_wait(1)
    _aplicar_stealth(driver)

    # Configura destino de download via CDP — tenta ambos os comandos
    # (Browser é mais moderno; Page é fallback para versões antigas)
    for cmd, params in [
        ("Browser.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": str(TEMP_DOWNLOAD_DIR),
            "eventsEnabled": False,
        }),
        ("Page.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": str(TEMP_DOWNLOAD_DIR),
        }),
    ]:
        try:
            driver.execute_cdp_cmd(cmd, params)
        except Exception:
            pass

    return driver


def _encerrar_driver(driver: Optional[webdriver.Chrome]) -> None:
    """Fecha o Chrome com tolerância a falhas."""
    if driver is None:
        return

    try:
        driver.quit()
    except Exception:
        pass

    time.sleep(1.0)


def limpar_sessao_entre_ucs(driver: webdriver.Chrome) -> None:
    """
    Limpeza agressiva entre UCs sem recriar o driver.
    Apaga cookies, storage e cache de rede via CDP para isolar cada sessão.
    O reinício completo do Chrome ocorre a cada REINICIAR_DRIVER_A_CADA UCs
    pelo GerenciadorDriver.
    """
    # Fecha abas extras primeiro
    try:
        handles = driver.window_handles
        if len(handles) > 1:
            principal = handles[0]
            for h in handles[1:]:
                try:
                    driver.switch_to.window(h)
                    driver.close()
                except Exception:
                    pass
            driver.switch_to.window(principal)
    except Exception:
        pass

    # Navega para página neutra antes de limpar
    try:
        driver.get("about:blank")
        time.sleep(0.4)
    except Exception:
        pass

    # Cookies via Selenium
    try:
        driver.delete_all_cookies()
    except Exception:
        pass

    # localStorage + sessionStorage via JS
    try:
        driver.execute_script(
            "try { window.localStorage.clear(); } catch(e) {}"
            "try { window.sessionStorage.clear(); } catch(e) {}"
        )
    except Exception:
        pass

    # Cache e cookies via CDP
    for cmd in ("Network.clearBrowserCache", "Network.clearBrowserCookies"):
        try:
            driver.execute_cdp_cmd(cmd, {})
        except Exception:
            pass


# =============================================================================
# GERENCIADOR DE DRIVER com reinício automático
# =============================================================================

class GerenciadorDriver:
    """
    Mantém um driver Chrome ativo e o reinicia automaticamente quando:
      - REINICIAR_DRIVER_A_CADA UCs foram processadas (ciclo preventivo)
      - MAX_ERROS_CONSECUTIVOS falhas seguidas (driver possivelmente travado)
    """

    def __init__(self) -> None:
        self._driver: Optional[webdriver.Chrome] = None
        self._ucs_no_ciclo:   int = 0
        self._erros_seguidos: int = 0

    def _driver_esta_vivo(self) -> bool:
        if self._driver is None:
            return False
        try:
            _ = self._driver.current_window_handle
            handles = self._driver.window_handles
            return bool(handles)
        except Exception:
            return False

    def obter(self) -> webdriver.Chrome:
        """Retorna o driver ativo, recriando-o se necessário."""
        precisa = (
            self._driver is None
            or not self._driver_esta_vivo()
            or (REINICIAR_DRIVER_A_CADA > 0
                and self._ucs_no_ciclo >= REINICIAR_DRIVER_A_CADA)
            or self._erros_seguidos >= MAX_ERROS_CONSECUTIVOS
        )
        if precisa:
            self._recriar()
        return self._driver

    def _recriar(self) -> None:
        if self._driver is not None:
            if self._ucs_no_ciclo >= REINICIAR_DRIVER_A_CADA > 0:
                motivo = f"ciclo de {REINICIAR_DRIVER_A_CADA} UCs"
            else:
                motivo = f"{self._erros_seguidos} erros consecutivos"
            logger.info("↺  Reiniciando Chrome (%s)...", motivo)
            _encerrar_driver(self._driver)
            self._driver = None
            time.sleep(2.0)   # aguarda o processo Chrome encerrar no Windows

        # Tenta criar o driver até 3 vezes antes de desistir
        ultimo_erro: Optional[Exception] = None
        for tentativa in range(1, 4):
            try:
                logger.debug("  [driver] abrindo Chrome (tentativa %d/3)...", tentativa)
                self._driver         = criar_driver()
                self._ucs_no_ciclo   = 0
                self._erros_seguidos = 0
                logger.debug("  [driver] Chrome pronto")
                return
            except Exception as exc:
                ultimo_erro = exc
                logger.warning("  [driver] falha ao abrir Chrome (t%d/3): %s",
                               tentativa, exc)
                time.sleep(3.0 * tentativa)

        # Se chegou aqui, as 3 tentativas falharam — propaga o erro
        raise RuntimeError(
            f"Não foi possível iniciar o Chrome após 3 tentativas: {ultimo_erro}"
        )

    def registrar_sucesso(self) -> None:
        self._ucs_no_ciclo   += 1
        self._erros_seguidos  = 0

    def registrar_erro(self) -> None:
        self._ucs_no_ciclo   += 1
        self._erros_seguidos += 1

    def reiniciar_agora(self, motivo: str = "reinicio solicitado") -> None:
        if self._driver:
            logger.info("↺  Reiniciando Chrome (%s)...", motivo)
            _encerrar_driver(self._driver)
            self._driver = None
            time.sleep(2.0)
        self._ucs_no_ciclo = 0
        self._erros_seguidos = 0

    def fechar(self) -> None:
        if self._driver:
            _encerrar_driver(self._driver)
            self._driver = None


# =============================================================================
# HELPERS SELENIUM
# =============================================================================

def lento_digitar(element, texto: str) -> None:
    element.parent.execute_script("arguments[0].focus(); arguments[0].value = '';", element)
    element.click()
    for ch in texto:
        element.send_keys(ch)
        time.sleep(random.uniform(0.05, 0.13))


def safe_click(driver: webdriver.Chrome, by: By, locator: str,
               timeout: int = TIMEOUT_PADRAO):
    elem = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, locator)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    pausa_humana(0.15, 0.4)
    try:
        elem.click()
    except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
        driver.execute_script("arguments[0].click();", elem)
    return elem


def safe_visible(driver: webdriver.Chrome, by: By, locator: str,
                 timeout: int = TIMEOUT_PADRAO):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, locator))
    )


def clicar_reforcado(
    driver: webdriver.Chrome,
    descricao: str,
    seletores: list[tuple],
    timeout: int = TIMEOUT_PADRAO,
    tentativas: int = 3,
    pausa: float = 1.2,
    ignorar_erro: bool = False,
) -> None:
    ultimo_erro = None
    for t in range(1, tentativas + 1):
        for by, loc in seletores:
            try:
                safe_click(driver, by, loc, timeout=timeout)
                logger.debug("  [click ok] %s (t%d)", descricao, t)
                return
            except Exception as exc:
                ultimo_erro = exc
        if t < tentativas:
            time.sleep(pausa)
    if ignorar_erro:
        logger.debug("  [click skip] %s — não encontrado", descricao)
        return
    raise RuntimeError(f"Falha ao clicar '{descricao}': {ultimo_erro}")


def fechar_abas_extras(driver: webdriver.Chrome, aba_principal: str) -> None:
    for handle in list(driver.window_handles):
        if handle != aba_principal:
            try:
                driver.switch_to.window(handle)
                driver.close()
            except Exception:
                pass
    try:
        driver.switch_to.window(aba_principal)
    except Exception:
        pass


def aguardar_novo_pdf(dir_path: Path, antes: set[str],
                      timeout: int = TIMEOUT_DOWNLOAD) -> Optional[Path]:
    fim = time.time() + timeout
    while time.time() < fim:
        novos = {p.name for p in dir_path.glob("*")} - antes
        candidatos = [
            dir_path / n for n in novos
            if n.lower().endswith(".pdf") and not n.endswith(".crdownload")
        ]
        if candidatos:
            candidatos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            pdf = candidatos[0]
            for _ in range(20):
                try:
                    t1 = pdf.stat().st_size
                    time.sleep(0.5)
                    if pdf.stat().st_size == t1 and t1 > 512:
                        return pdf
                except FileNotFoundError:
                    break
        time.sleep(0.6)
    return None


def _esperar_primeiro_visivel(
    driver: webdriver.Chrome,
    seletores: list[tuple],
    timeout: int = TIMEOUT_PADRAO,
    exigir_habilitado: bool = True,
):
    fim = time.time() + timeout
    ultimo_erro = None
    while time.time() < fim:
        for by, loc in seletores:
            try:
                for elem in driver.find_elements(by, loc):
                    if not elem.is_displayed():
                        continue
                    if exigir_habilitado and not elem.is_enabled():
                        continue
                    return elem
            except Exception as exc:
                ultimo_erro = exc
        time.sleep(0.25)
    raise TimeoutException(f"Elemento não localizado pelos seletores: {seletores} ({ultimo_erro})")


def _valor_campo(elem) -> str:
    try:
        return elem.get_attribute("value") or ""
    except Exception:
        return ""


def _preencher_campo_validando(
    driver: webdriver.Chrome,
    seletores: list[tuple],
    valor: str,
    descricao: str,
    tentativas: int = 3,
) -> None:
    esperado = str(valor).strip()
    esperado_digits = limpar_numero(esperado)

    for tentativa in range(1, tentativas + 1):
        campo = _esperar_primeiro_visivel(driver, seletores, timeout=TIMEOUT_PADRAO)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
        try:
            campo.click()
        except Exception:
            driver.execute_script("arguments[0].click();", campo)
        pausa_humana(0.05, 0.12)

        try:
            campo.send_keys(Keys.CONTROL, "a")
            campo.send_keys(Keys.DELETE)
        except Exception:
            pass
        try:
            campo.clear()
        except Exception:
            pass
        try:
            driver.execute_script("arguments[0].value = '';", campo)
        except Exception:
            pass

        lento_digitar(campo, esperado)

        try:
            driver.execute_script(
                """
                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('change', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('blur', {bubbles: true}));
                """,
                campo,
            )
        except Exception:
            pass

        pausa_humana(0.15, 0.35)
        atual = _valor_campo(campo)
        atual_digits = limpar_numero(atual)

        if esperado_digits:
            if atual_digits == esperado_digits:
                return
        elif atual.strip() == esperado:
            return

        logger.warning(
            "  [login] %s digitado incompleto (t%d/%d): '%s' esperado '%s'",
            descricao,
            tentativa,
            tentativas,
            atual,
            esperado,
        )
        pausa_humana(0.35, 0.8)

    raise RuntimeError(f"Não foi possível preencher corretamente o campo {descricao}")


def _texto_body_normalizado(driver: webdriver.Chrome) -> str:
    try:
        body = driver.find_element(By.TAG_NAME, "body")
        return normalizar_texto(body.text)
    except Exception:
        return ""


# =============================================================================
# FLUXO DO PORTAL
# =============================================================================

def abrir_login(driver: webdriver.Chrome) -> None:
    ultimo_erro = None
    for tentativa in range(1, 3):
        try:
            driver.get(URL_LOGIN)
            WebDriverWait(driver, TIMEOUT_PADRAO).until(
                lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
            )
            _esperar_primeiro_visivel(driver, LOGIN_UC_SELECTORS, timeout=TIMEOUT_PADRAO)
            _esperar_primeiro_visivel(driver, LOGIN_DOC_SELECTORS, timeout=8)
            return
        except Exception as exc:
            ultimo_erro = exc
            logger.warning("  [login] falha ao abrir tela inicial (t%d/2): %s", tentativa, exc)
            try:
                driver.get("about:blank")
            except Exception:
                pass
            time.sleep(1.0)

    raise RuntimeError(f"Portal de login não carregou corretamente: {ultimo_erro}")


def _detectar_alerta(driver: webdriver.Chrome) -> Optional[str]:
    try:
        alert = driver.switch_to.alert
        txt   = alert.text
        alert.accept()
        time.sleep(0.8)
        return txt
    except (NoAlertPresentException, Exception):
        return None


def _esta_logado(driver: webdriver.Chrome) -> bool:
    try:
        url = driver.current_url.lower()
        if "logingo.aspx" not in url and not url.rstrip("/").endswith("/login"):
            return True

        for by, loc in LOGIN_OK_SELECTORS:
            try:
                elems = driver.find_elements(by, loc)
                if any(e.is_displayed() for e in elems):
                    return True
            except Exception:
                pass

        # Se os campos de login ainda estão visíveis, seguimos tratando como não logado.
        for seletores in (LOGIN_UC_SELECTORS, LOGIN_DOC_SELECTORS):
            for by, loc in seletores:
                try:
                    elems = driver.find_elements(by, loc)
                    if any(e.is_displayed() for e in elems):
                        return False
                except Exception:
                    pass

        return False
    except Exception:
        return False


def _aguardar_select_pronto(driver: webdriver.Chrome, select_id: str,
                            timeout: int = 20):
    """Espera o select existir, ficar visível e carregar opções úteis."""
    def _ok(_driver):
        try:
            elem = _driver.find_element(By.ID, select_id)
            if not elem.is_displayed() or not elem.is_enabled():
                return False
            opcoes = [
                o for o in Select(elem).options
                if (o.text or "").strip() and "selecione" not in normalizar_texto(o.text)
            ]
            return elem if opcoes else False
        except Exception:
            return False

    return WebDriverWait(driver, timeout).until(_ok)


def _classificar_resultado_login(driver: webdriver.Chrome) -> tuple[str, str]:
    txt_alerta = _detectar_alerta(driver)
    if txt_alerta:
        norm = normalizar_texto(txt_alerta)
        if any(sig in norm or sig in txt_alerta for sig in LOGIN_ERROS_SERVIDOR):
            return "server", txt_alerta
        return "credencial", txt_alerta

    if _esta_logado(driver):
        return "ok", driver.current_url

    body = _texto_body_normalizado(driver)
    if any(sig in body for sig in LOGIN_ERROS_SERVIDOR):
        return "server", body
    if any(sig in body for sig in LOGIN_ERROS_CREDENCIAL):
        return "credencial", body
    return "pendente", body


def _aguardar_resultado_login(driver: webdriver.Chrome, timeout: int = TIMEOUT_LOGIN_RESULTADO) -> tuple[str, str]:
    fim = time.time() + timeout
    ultimo_status = ("pendente", "")
    while time.time() < fim:
        ultimo_status = _classificar_resultado_login(driver)
        if ultimo_status[0] != "pendente":
            return ultimo_status
        time.sleep(0.35)
    return _classificar_resultado_login(driver)


def efetuar_login(driver: webdriver.Chrome, instalacao: str, cnpj: str) -> bool:
    tentativa      = 0
    alertas_server = 0

    while tentativa < LOGIN_MAX_TENTATIVAS:
        try:
            url_atual = driver.current_url.lower()

            # Se já saímos da página de login, o portal pode ter aceitado as credenciais.
            # Não navega de volta — verifica diretamente.
            if "logingo.aspx" not in url_atual:
                if _esta_logado(driver):
                    return True
                # Página desconhecida — volta ao login
                abrir_login(driver)

            if tentativa > 0:
                logger.debug("  [login] recarregando página de login (t%d/%d)",
                             tentativa + 1, LOGIN_MAX_TENTATIVAS)
                driver.refresh()
                _esperar_primeiro_visivel(driver, LOGIN_UC_SELECTORS, timeout=TIMEOUT_PADRAO)
                pausa_humana(0.5, 0.9)

            _preencher_campo_validando(driver, LOGIN_UC_SELECTORS, instalacao, "instalação")
            _preencher_campo_validando(driver, LOGIN_DOC_SELECTORS, cnpj, "documento")
            pausa_humana(0.2, 0.5)

            clicar_reforcado(driver, "Entrar", LOGIN_BTN_SELECTORS, timeout=12, tentativas=3)
            pausa_humana(0.9, 1.6)

            status, detalhe = _aguardar_resultado_login(driver, timeout=TIMEOUT_LOGIN_RESULTADO)
            if status == "ok":
                return True

            if status == "server":
                alertas_server += 1
                logger.warning("  [login] alerta/instabilidade do portal (%d/%d): %s",
                               alertas_server, MAX_ALERTAS_SERVIDOR, detalhe)
                if alertas_server >= MAX_ALERTAS_SERVIDOR:
                    return False
                if alertas_server >= 2:
                    raise SessaoChromeInstavel(
                        "Portal Equatorial GO instável no login; recriando Chrome"
                    )
                try:
                    driver.refresh()
                except Exception:
                    abrir_login(driver)
                pausa_humana(1.0, 1.6)
                continue

            if status == "credencial":
                logger.debug("  [login] portal rejeitou acesso: %s", detalhe)
                tentativa += 1
                continue

            # "pendente" após timeout — verifica pelo DOM antes de tentar de novo
            if _esta_logado(driver):
                logger.info("  [login] confirmado pelo DOM (t%d)", tentativa + 1)
                return True

            logger.warning("  [login] sem confirmação de sucesso após submit (t%d/%d)",
                           tentativa + 1, LOGIN_MAX_TENTATIVAS)
            tentativa += 1

        except Exception as exc:
            logger.debug("  [login] exceção t%d: %s", tentativa + 1,
                         traceback.format_exc().strip().splitlines()[-1])
            if _erro_sessao_driver(exc):
                raise
            tentativa += 1
            time.sleep(2)

    return False


def _inspecionar_modais_bs4(driver: webdriver.Chrome) -> None:
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        logger.debug("  [BS4] beautifulsoup4 não instalado")
        return

    try:
        soup = BeautifulSoup(driver.page_source, "html.parser")
    except Exception as exc:
        logger.debug("  [BS4] falha ao parsear: %s", exc)
        return

    PALAVRAS = {"modal", "overlay", "popup", "dialog", "lightbox", "backdrop"}
    candidatos = [
        tag for tag in soup.find_all(True)
        if any(
            p in " ".join(tag.get("class") or []).lower()
            or p in (tag.get("id") or "").lower()
            for p in PALAVRAS
        )
    ]

    if not candidatos:
        logger.debug("  [BS4] nenhum modal no DOM")
        return

    logger.debug("  [BS4] %d modal(is) encontrado(s)", len(candidatos))
    for el in candidatos:
        botoes = [
            f"<{b.name} id='{b.get('id','')}' "
            f"txt='{b.get_text(strip=True) or b.get('value','')}'>"
            for b in el.find_all(["button", "input", "a"])
        ]
        logger.debug("  [BS4] <%s> id='%s' botões=%s",
                     el.name, el.get("id", ""), botoes)


def _modal_pos_login_visivel(driver: webdriver.Chrome) -> bool:
    seletores = [
        (By.CSS_SELECTOR, ".modal.show"),
        (By.CSS_SELECTOR, ".modal.in"),
        (By.CSS_SELECTOR, ".modal-backdrop"),
        (By.CSS_SELECTOR, ".popup"),
        (By.CSS_SELECTOR, ".overlay"),
        (By.XPATH, "//div[contains(@class,'modal') and .//*[contains(.,'PIX')]]"),
    ]
    for by, loc in seletores:
        try:
            elems = driver.find_elements(by, loc)
            if any(e.is_displayed() for e in elems):
                return True
        except Exception:
            pass
    return False


def fechar_modal_pos_login(driver: webdriver.Chrome) -> None:
    try:
        WebDriverWait(driver, 2).until(lambda d: _modal_pos_login_visivel(d))
    except Exception:
        return

    _inspecionar_modais_bs4(driver)

    seletores = [
        (By.CSS_SELECTOR, "button.close[data-dismiss='modal']"),
        (By.CSS_SELECTOR, "a.close[data-dismiss='modal']"),
        (By.CSS_SELECTOR, "span.close"),
        (By.CSS_SELECTOR, "[aria-label='Close']"),
        (By.CSS_SELECTOR, "[aria-label='Fechar']"),
        (By.XPATH, "//button[@type='button' and contains(@class,'close') and @data-dismiss='modal']"),
        (By.XPATH, "//a[contains(@class,'close') or @data-dismiss='modal']"),
        (By.CSS_SELECTOR, "button.btn.btn-info.ModalButton[data-dismiss='modal']"),
        (By.XPATH, "//button[contains(@class,'ModalButton') and @data-dismiss='modal']"),
        (By.XPATH, "//button[@data-dismiss='modal']"),
        (By.XPATH, "//*[self::button or self::a or self::span][normalize-space(.)='?' or normalize-space(.)='X']"),
    ]

    clicou = False
    for by, loc in seletores:
        try:
            elems = driver.find_elements(by, loc)
            alvo = next((e for e in elems if e.is_displayed()), None)
            if alvo:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
                pausa_humana(0.1, 0.25)
                try:
                    alvo.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", alvo)
                logger.debug("  [MODAL] fechado via %s", loc)
                clicou = True
                break
        except Exception:
            pass

    if clicou:
        try:
            WebDriverWait(driver, 2).until(lambda d: not _modal_pos_login_visivel(d))
        except Exception:
            pass

    if not clicou or _modal_pos_login_visivel(driver):
        try:
            driver.execute_script(
                """
                document.querySelectorAll('.modal.in,.modal.show,.popup,.overlay').forEach(function(m){
                  var b = m.querySelector('[data-dismiss="modal"], .close, [aria-label="Close"], [aria-label="Fechar"]');
                  if (b) {
                    try { b.click(); } catch(e) {}
                  }
                  try { m.classList.remove('show', 'in'); } catch(e) {}
                  try { m.style.display = 'none'; } catch(e) {}
                  try { m.setAttribute('aria-hidden', 'true'); } catch(e) {}
                  try { if (window.jQuery && jQuery(m).modal) { jQuery(m).modal('hide'); } } catch(e) {}
                });
                document.querySelectorAll('.modal-backdrop,.backdrop,.overlay').forEach(function(bd){
                  try { bd.remove(); } catch(e) { try { bd.style.display = 'none'; } catch(e2) {} }
                });
                try { document.body.classList.remove('modal-open'); } catch(e) {}
                try { document.body.style.overflow = 'auto'; } catch(e) {}
                """
            )
            logger.debug("  [MODAL] fechado via JS agressivo")
        except Exception:
            pass

    try:
        WebDriverWait(driver, 3).until(lambda d: not _modal_pos_login_visivel(d))
    except Exception:
        logger.debug("  [MODAL] overlay permaneceu vis?vel ap?s tentativas")

    time.sleep(0.2)


def abrir_segunda_via(driver: webdriver.Chrome) -> None:
    # Se o submenu já estiver visível, não clicamos em "Contas" de novo para
    # evitar recolher o acordeão e ficar preso no Index.
    submenu_visivel = False
    for by, loc in [
        (By.ID, "LinkSegundaVia"),
        (By.XPATH, "//a[contains(normalize-space(.),'Segunda Via de Fatura')]"),
    ]:
        try:
            elems = driver.find_elements(by, loc)
            if any(e.is_displayed() for e in elems):
                submenu_visivel = True
                break
        except Exception:
            pass

    if not submenu_visivel:
        clicar_reforcado(driver, "Menu Contas", [
            (By.XPATH,        "//label[@for='A' and contains(normalize-space(.),'Contas')]"),
            (By.CSS_SELECTOR, "label[for='A']"),
            (By.XPATH,
             "//label[contains(normalize-space(.),'Contas') "
             "and .//i[contains(@class,'menu-icon')]]"),
        ], timeout=15, tentativas=4, pausa=1.0)
        time.sleep(0.7)

    clicar_reforcado(driver, "Segunda Via de Fatura", [
        (By.ID,    "LinkSegundaVia"),
        (By.XPATH, "//a[contains(@href,'SegundaVia.aspx')]"),
        (By.XPATH, "//a[contains(normalize-space(.),'Segunda Via de Fatura')]"),
    ], timeout=15, tentativas=4, pausa=1.0)
    WebDriverWait(driver, 20).until(
        lambda d: "segundavia" in (d.current_url or "").lower()
        or any(
            e.is_displayed() for e in d.find_elements(By.ID, "CONTENT_cbTipoEmissao")
        )
    )
    time.sleep(0.8)


def _select_com_fallback(sel: Select, valor: str, texto_parcial: str) -> None:
    try:
        sel.select_by_value(valor)
        return
    except Exception:
        pass
    for opt in sel.options:
        if texto_parcial.lower() in opt.text.lower():
            opt.click()
            return
    raise NoSuchElementException(f"Op??o n?o encontrada: '{valor}' / '{texto_parcial}'")


def _selecionar_select_robusto(driver: webdriver.Chrome, select_id: str,
                               valor: str, texto_parcial: str,
                               timeout: int = 20) -> None:
    elem = _aguardar_select_pronto(driver, select_id, timeout=timeout)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    pausa_humana(0.1, 0.25)

    try:
        _select_com_fallback(Select(elem), valor, texto_parcial)
    except Exception:
        elem = driver.find_element(By.ID, select_id)
        driver.execute_script(
            """
            const sel = arguments[0];
            const wantedValue = arguments[1].toLowerCase();
            const wantedText = arguments[2].toLowerCase();
            const opts = Array.from(sel.options || []);
            const opt = opts.find(o =>
                String(o.value || '').toLowerCase() === wantedValue ||
                String(o.text || '').toLowerCase().includes(wantedText)
            );
            if (!opt) return false;
            sel.value = opt.value;
            opt.selected = true;
            sel.dispatchEvent(new Event('input', { bubbles: true }));
            sel.dispatchEvent(new Event('change', { bubbles: true }));
            return true;
            """,
            elem, valor, texto_parcial,
        )

    escolhido = (driver.execute_script(
        "return (arguments[0].options[arguments[0].selectedIndex] || {}).text || '';",
        driver.find_element(By.ID, select_id),
    ) or "").strip()
    if not escolhido or "selecione" in normalizar_texto(escolhido):
        raise NoSuchElementException(
            f"Select {select_id} permaneceu sem escolha v?lida: '{escolhido}'"
        )


def preencher_opcoes_fatura(driver: webdriver.Chrome) -> None:
    _selecionar_select_robusto(driver, "CONTENT_cbTipoEmissao", "completa", "completa")
    pausa_humana(0.2, 0.4)

    _selecionar_select_robusto(driver, "CONTENT_cbMotivo", "ESV05", "outro")
    pausa_humana(0.2, 0.4)

    clicar_reforcado(driver, "Emitir", [
        (By.ID,    "CONTENT_btEnviar"),
        (By.XPATH, "//input[@type='submit' and @value='Emitir']"),
        (By.XPATH, "//input[contains(@name,'btEnviar')]"),
    ], timeout=15, tentativas=4)
    time.sleep(2.2)


# =============================================================================
# FLUXO POR CNPJ — mapeamento antigo→novo e dropdown de UC
# =============================================================================

URL_SEGUNDA_VIA_GO = "https://goias.equatorialenergia.com.br/SegundaVia.aspx"

# Mapeamento instalação_antiga → instalacao_nova (carregado de Faturas Equatoriais BB)
# Ex: "580018258" → "1302962012-07" (sem pontos, com hífen)
_MAPA_INSTALACAO: dict[str, str] = {}
_MAPA_INSTALACAO_REVERSO: dict[str, str] = {}


def _normalizar_instalacao(s: str) -> str:
    """Remove pontos mas mantém hífen: '1.302.962.012-07' → '1302962012-07'."""
    return re.sub(r"\.", "", str(s).strip())


def carregar_mapa_instalacoes() -> None:
    """Lê Faturas Equatoriais BB e popula _MAPA_INSTALACAO (antigo→novo)."""
    global _MAPA_INSTALACAO, _MAPA_INSTALACAO_REVERSO
    if not PLANILHA_FATURAS_PATH.exists():
        logger.warning("  [mapa] planilha de faturas não encontrada: %s", PLANILHA_FATURAS_PATH)
        return
    try:
        df = pd.read_excel(str(PLANILHA_FATURAS_PATH), dtype=str)
        col_ant = next((c for c in df.columns if "antiga" in normalizar_texto(c)), None)
        col_nov = next((c for c in df.columns if normalizar_texto(c) == "instalacao"
                        or ("instalac" in normalizar_texto(c) and "antiga" not in normalizar_texto(c))), None)
        if not col_ant or not col_nov:
            logger.warning("  [mapa] colunas não encontradas: %s", df.columns.tolist())
            return
        for _, row in df.iterrows():
            ant = limpar_numero(str(row.get(col_ant) or ""))
            nov = _normalizar_instalacao(str(row.get(col_nov) or ""))
            if ant and nov:
                _MAPA_INSTALACAO[ant] = nov
                _MAPA_INSTALACAO_REVERSO[nov] = ant
                _MAPA_INSTALACAO_REVERSO[limpar_numero(nov)] = ant
        logger.info("  [mapa] %d instalações mapeadas (antigo→novo)", len(_MAPA_INSTALACAO))
    except Exception as exc:
        logger.warning("  [mapa] erro ao carregar: %s", exc)


def instalacao_nova(inst_antiga: str) -> str:
    """Retorna número novo de instalação, ou o próprio antigo se não mapeado."""
    return _MAPA_INSTALACAO.get(limpar_numero(inst_antiga), inst_antiga)


def instalacao_para_dropdown(inst_antiga: str) -> str:
    """
    Converte instalação antiga no valor exato do dropdown (15 dígitos com zeros).
    Ex: '580018258' → '1.302.962.012-07' → '130296201207' → '000130296201207'
    """
    nova = instalacao_nova(inst_antiga)          # ex: "1.302.962.012-07"
    apenas_digitos = limpar_numero(nova)         # ex: "130296201207"
    return apenas_digitos.zfill(15)              # ex: "000130296201207"


def instalacao_antiga(inst_nova: str) -> str:
    """Retorna número antigo a partir do novo (com ou sem pontos)."""
    chave = _normalizar_instalacao(inst_nova)
    return _MAPA_INSTALACAO_REVERSO.get(chave, _MAPA_INSTALACAO_REVERSO.get(limpar_numero(chave), inst_nova))


def _encontrar_select_uc_js(driver: webdriver.Chrome):
    """
    Usa JavaScript para achar o <select> cujas opções parecem números de instalação
    (contêm dígitos e hífen no formato X.XXX.XXX.XXX-XX ou similar).
    Retorna o elemento ou None.
    """
    try:
        elem = driver.execute_script("""
            var sels = document.querySelectorAll('select');
            for (var i = 0; i < sels.length; i++) {
                var opts = Array.from(sels[i].options || []);
                var parece_inst = opts.some(function(o) {
                    var v = (o.value || o.text || '').trim();
                    return v.length > 3 && /[0-9]/.test(v) && v.indexOf('selecione') < 0
                           && v.toLowerCase().indexOf('selecione') < 0;
                });
                if (parece_inst && sels[i].offsetParent !== null) {
                    return sels[i];
                }
            }
            return null;
        """)
        return elem
    except Exception:
        return None


def _aguardar_select_uc(driver: webdriver.Chrome, timeout: int = 15):
    """Aguarda e retorna o select de Unidade Consumidora na página de Segunda Via."""
    # IDs conhecidos a tentar primeiro
    ids_candidatos = [
        "CONTENT_comboBoxUC",
        "CONTENT_cbInstalacao", "CONTENT_cbUC", "CONTENT_ddlInstalacao",
        "CONTENT_ddlUC", "CONTENT_cbUnidadeConsumidora",
    ]
    fim = time.time() + timeout
    while time.time() < fim:
        # 1. Tenta IDs específicos
        for sel_id in ids_candidatos:
            try:
                elems = driver.find_elements(By.ID, sel_id)
                for elem in elems:
                    if not elem.is_displayed() or not elem.is_enabled():
                        continue
                    opcoes = [o for o in Select(elem).options
                              if (o.get_attribute("value") or "").strip()
                              and "selecione" not in normalizar_texto(o.text)]
                    if opcoes:
                        logger.debug("  [uc_dropdown] encontrado por ID: %s", sel_id)
                        return elem
            except Exception:
                pass

        # 2. XPath por label com texto "instalação"
        try:
            for xpath in [
                "//*[contains(translate(normalize-space(.),'INSTALAÇÃO','instalacao'),'instalac')]/following::select[1]",
                "//select[not(contains(@id,'Tipo')) and not(contains(@id,'Motivo')) and not(contains(@id,'tipo')) and not(contains(@id,'motivo'))]",
            ]:
                elems = driver.find_elements(By.XPATH, xpath)
                for elem in elems:
                    if not elem.is_displayed() or not elem.is_enabled():
                        continue
                    opcoes = [o for o in Select(elem).options
                              if (o.get_attribute("value") or "").strip()
                              and "selecione" not in normalizar_texto(o.text)]
                    if opcoes:
                        logger.debug("  [uc_dropdown] encontrado por XPath")
                        return elem
        except Exception:
            pass

        # 3. Busca via JS (qualquer select com opções numéricas visível)
        elem = _encontrar_select_uc_js(driver)
        if elem:
            logger.debug("  [uc_dropdown] encontrado via JS")
            return elem

        time.sleep(0.4)

    raise TimeoutException(f"Select de Unidade Consumidora não encontrado em {timeout}s")


def listar_ucs_segunda_via(driver: webdriver.Chrome) -> list[str]:
    """
    Retorna lista de valores do dropdown de UC (no formato que o portal usa,
    ex: '1302962012-07'). Retorna [] se não encontrado (UC única).
    """
    # Loga todos os selects para diagnóstico
    try:
        todos = driver.execute_script("""
            return Array.from(document.querySelectorAll('select')).map(function(s) {
                return {id: s.id, name: s.name,
                        opts: Array.from(s.options).map(function(o){return o.value+'|'+o.text;})};
            });
        """)
        logger.debug("  [uc_dropdown] selects na página: %s", todos)
    except Exception:
        pass

    try:
        elem = _aguardar_select_uc(driver, timeout=10)
        opcoes = []
        for o in Select(elem).options:
            val  = (o.get_attribute("value") or "").strip()
            txt  = (o.text or "").strip()
            chave = val or txt
            if chave and "selecione" not in normalizar_texto(chave):
                opcoes.append({"value": val, "text": txt})
        logger.debug("  [uc_dropdown] primeiras opções (value|text): %s",
                     [(x["value"], x["text"]) for x in opcoes[:5]])
        return opcoes
    except Exception as exc:
        logger.debug("  [uc_dropdown] não encontrado: %s", exc)
        return []


def _normalizar_para_match(s: str) -> str:
    """Remove tudo que não é dígito ou hífen, para comparação."""
    return re.sub(r"[^\d\-]", "", s.strip())


def selecionar_uc_segunda_via(driver: webdriver.Chrome, instalacao_ant: str) -> bool:
    """
    Seleciona a UC no dropdown usando o valor de 15 dígitos com zeros à esquerda.
    Ex: '580018258' → '1.302.962.012-07' → '000130296201207'
    """
    alvo = instalacao_para_dropdown(instalacao_ant)
    logger.debug("  [uc_dropdown] buscando value='%s' para inst=%s", alvo, instalacao_ant)

    try:
        elem = _aguardar_select_uc(driver, timeout=10)
        sel  = Select(elem)
        sel.select_by_value(alvo)
        try:
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                elem,
            )
        except Exception:
            pass
        pausa_humana(0.4, 0.8)
        logger.info("  [uc_dropdown] selecionada: %s → %s", instalacao_ant, alvo)
        return True
    except Exception as exc:
        logger.warning("  [uc_dropdown] falha ao selecionar %s (value='%s'): %s",
                       instalacao_ant, alvo, exc)
        return False


_PAGINA_INSTAVEL_SINAIS = (
    "indisponivel no momento",
    "sistema encontra-se indisponivel",
    "tente novamente mais tarde",
    "instabilidade",
    "temporariamente indisponivel",
    "#002",
)


def _pagina_instavel(driver: webdriver.Chrome) -> bool:
    try:
        body = normalizar_texto(driver.find_element(By.TAG_NAME, "body").text)
        return any(s in body for s in _PAGINA_INSTAVEL_SINAIS)
    except Exception:
        return False


def _esta_na_segunda_via(driver: webdriver.Chrome) -> bool:
    """Verifica se o formulário de Segunda Via está visível (pelo DOM, não pela URL)."""
    try:
        # Elementos presentes no formulário de emissão
        for sel_id in ("CONTENT_cbTipoEmissao", "CONTENT_cbMotivo", "CONTENT_btEnviar"):
            elems = driver.find_elements(By.ID, sel_id)
            if any(e.is_displayed() for e in elems):
                return True
        # UC dropdown também indica que estamos no formulário
        elem = _encontrar_select_uc_js(driver)
        if elem and elem.is_displayed():
            return True
    except Exception:
        pass
    return False


def _aguardar_portal_estavel(driver: webdriver.Chrome, max_espera: int = 120) -> bool:
    """
    Aguarda o portal sair do estado 'indisponível' fazendo refreshes periódicos.
    Retorna True se estabilizou, False se esgotou o tempo.
    """
    fim = time.time() + max_espera
    intervalo = 15
    while time.time() < fim:
        try:
            driver.refresh()
        except Exception:
            pass
        time.sleep(intervalo)
        if not _pagina_instavel(driver):
            logger.info("  [nav] portal estabilizou após refresh")
            return True
        restante = int(fim - time.time())
        logger.info("  [nav] portal ainda instável — aguardando %ds (restam %ds)...",
                    intervalo, restante)
        intervalo = min(intervalo + 10, 30)
    return False


def voltar_para_segunda_via(driver: webdriver.Chrome) -> None:
    """
    Retorna ao formulário de Segunda Via após um download.
    Se o portal estiver instável faz refreshes até estabilizar (até 2min),
    depois navega via menu. Levanta RuntimeError se não conseguir.
    """
    # O modal OK pode já ter retornado ao formulário
    pausa_humana(1.0, 2.0)
    if _esta_na_segunda_via(driver):
        return

    # Portal instável: aguarda estabilizar com refreshes
    if _pagina_instavel(driver):
        logger.info("  [nav] portal instável — aguardando estabilizar (até 90s)...")
        if not _aguardar_portal_estavel(driver, max_espera=90):
            raise RuntimeError("Portal permaneceu indisponível por mais de 90s")

    # Tenta via menu (mais natural, menos chance de bloqueio)
    try:
        abrir_segunda_via(driver)
        if _esta_na_segunda_via(driver):
            return
    except Exception as exc:
        logger.debug("  [nav] falha via menu: %s", exc)

    # Fallback: URL direta com até 3 tentativas + refresh entre elas
    for tentativa in range(1, 4):
        try:
            driver.get(URL_SEGUNDA_VIA_GO)
            WebDriverWait(driver, TIMEOUT_PADRAO).until(
                lambda d: _esta_na_segunda_via(d) or _pagina_instavel(d)
            )
            if _esta_na_segunda_via(driver):
                pausa_humana(0.5, 1.0)
                return
            if _pagina_instavel(driver):
                logger.info("  [nav] instável após navegação (t%d/3) — refresh em 20s...", tentativa)
                time.sleep(20)
                try:
                    driver.refresh()
                    time.sleep(5)
                except Exception:
                    pass
                if _esta_na_segunda_via(driver):
                    return
        except Exception as exc:
            logger.debug("  [nav] URL direta t%d falha: %s", tentativa, exc)
            time.sleep(5.0)

    raise RuntimeError("Não foi possível retornar à página de Segunda Via após tentativas")


def _processar_download_uc(
    driver:   webdriver.Chrome,
    linha:    "LinhaAcesso",
    baixados: set[tuple[str, str]],
    master,
) -> dict:
    """
    Executa download de uma UC assumindo que a página está no estado
    pós-Emitir (tabela de faturas visível). Reutilizável pelo fluxo por CNPJ.
    """
    res          = _resultado_base(linha)
    tensao_class = classificar_tensao(linha.tensao)
    refs_pendentes: list[str] = []

    mes_ref_pre = _inspecionar_pagina_pre_download(driver)
    if mes_ref_pre:
        if ja_foi_baixado_local(baixados, linha.instalacao, mes_ref_pre):
            _emit("skipped_existing", instalacao=linha.instalacao, mes_ref=mes_ref_pre)
            res.update(status="JA_EXISTE", mensagem=f"Já existe: {mes_ref_pre}", mes_ref=mes_ref_pre)
            return res
        refs_pendentes = [mes_ref_pre]

    if not _botao_download_existe(driver):
        motivo = "sem botão Download" + (f" (ref: {mes_ref_pre})" if mes_ref_pre else "")
        res.update(status="SEM_FATURA", mensagem=motivo)
        return res

    _fase("Baixando PDF...")
    antes = {p.name for p in TEMP_DOWNLOAD_DIR.glob("*")}
    clicar_download(driver)
    time.sleep(2.5)

    try:
        clicar_ok_modal_pos_download(driver)
    except Exception as exc:
        logger.debug("  modal OK não imediato: %s", exc)

    pdf = aguardar_novo_pdf(TEMP_DOWNLOAD_DIR, antes, timeout=TIMEOUT_DOWNLOAD)

    if _tem_modal_ok_visivel(driver):
        try:
            clicar_ok_modal_pos_download(driver)
        except Exception:
            pass

    mes_ref = refs_pendentes[0] if refs_pendentes else None
    if not mes_ref and pdf and pdf.exists():
        _fase("Extraindo referência do PDF...")
        mes_ref = extrair_referencia_do_pdf(pdf)

    if not mes_ref:
        if pdf and pdf.exists():
            try: pdf.unlink()
            except Exception: pass
        res.update(status="ERRO", mensagem="Referência da fatura não identificada")
        return res

    res["mes_ref"] = mes_ref

    if ja_foi_baixado_local(baixados, linha.instalacao, mes_ref):
        if pdf and pdf.exists():
            try: pdf.unlink()
            except Exception: pass
        _emit("skipped_existing", instalacao=linha.instalacao, mes_ref=mes_ref)
        res.update(status="JA_EXISTE", mensagem=f"Já existe: {mes_ref}")
        return res

    if not pdf or not pdf.exists():
        res.update(status="ERRO", mensagem="PDF não encontrado após download")
        return res

    if not re.fullmatch(r"(0[1-9]|1[0-2])-20\d{2}", mes_ref):
        if pdf and pdf.exists():
            try: pdf.unlink()
            except Exception: pass
        res.update(status="ERRO", mensagem=f"mes_ref inválido: '{mes_ref}'")
        return res

    _fase(f"Salvando {mes_ref} ({tensao_class})...")
    indice_bb     = master.consumir_carimbo()
    pasta_destino = BASE_DIR / mes_ref / tensao_class
    pasta_destino.mkdir(parents=True, exist_ok=True)
    caminho_final = pasta_destino / f"{indice_bb}.pdf"
    if caminho_final.exists():
        caminho_final.unlink()
    shutil.move(str(pdf), str(caminho_final))

    master.registrar(
        indice_bb=indice_bb, sistema="EQUATORIAL", uc=linha.instalacao,
        mes_ref=mes_ref, fatura_id="", cnpj=linha.cnpj, estado="GO",
        instalacao=linha.instalacao, arquivo=str(caminho_final),
    )
    registrar_indice_local(
        indice_bb=indice_bb, instalacao=linha.instalacao,
        mes_ref=mes_ref, arquivo=str(caminho_final), tensao=tensao_class,
    )
    baixados.add((linha.instalacao.strip(), mes_ref.strip()))

    _emit("downloaded", instalacao=linha.instalacao, mes_ref=mes_ref, carimbo=indice_bb)
    res.update(status="OK", mensagem="Download concluído", indice=indice_bb, arquivo=str(caminho_final))
    return res


def _imprimir_resultado_uc(res: dict, linha: "LinhaAcesso") -> None:
    if res["status"] == "OK":
        _resultado_ok(linha.instalacao, res["mes_ref"], res.get("indice", "?"), Path(res["arquivo"]).name)
    elif res["status"] == "JA_EXISTE":
        _resultado_ja(linha.instalacao, res["mes_ref"])
    elif res["status"] == "SEM_FATURA":
        _resultado_sf(linha.instalacao, res["mensagem"])
    else:
        _resultado_erro(linha.instalacao, res["mensagem"])


def processar_grupo_cnpj(
    driver:       webdriver.Chrome,
    cnpj:         str,
    linhas:       list,
    baixados:     set[tuple[str, str]],
    master,
    total_geral:  int,
    offset:       int,
) -> list[dict]:
    """
    Processa todas as UCs de um mesmo CNPJ com UM ÚNICO LOGIN.
    Itera pelo dropdown de UC na página de Segunda Via.
    """
    resultados: list[dict] = []

    # Pré-índice: instalações já baixadas (sem mes_ref — qualquer mês)
    instalacoes_baixadas = {inst for inst, _ in baixados}

    # Se todas já foram baixadas, pula sem abrir browser
    if all(l.instalacao in instalacoes_baixadas for l in linhas):
        for linha in linhas:
            r = _resultado_base(linha)
            r.update(status="JA_EXISTE", mensagem="Já no índice local (pré-check)")
            resultados.append(r)
            logger.info("  inst=%-14s  → já baixada (pulando)", linha.instalacao)
        return resultados

    # Login sempre com a primeira UC do grupo (padrão estável)
    uc_login = linhas[0]
    _fase(f"Login CNPJ ...{cnpj[-6:]}  (UC ref: {uc_login.instalacao})")
    abrir_login(driver)
    login_ok = efetuar_login(driver, uc_login.instalacao, cnpj)
    if not login_ok:
        # Tenta confirmar pelo DOM mesmo sem resposta positiva do loop de login
        login_ok = _esta_logado(driver)
        if login_ok:
            logger.info("  [login] confirmado pelo DOM após timeout — continuando")
        else:
            for linha in linhas:
                r = _resultado_base(linha)
                r.update(status="ERRO", mensagem="Falha no login")
                resultados.append(r)
                _imprimir_resultado_uc(r, linha)
            return resultados

    fechar_modal_pos_login(driver)
    abrir_segunda_via(driver)

    ucs_disponiveis = listar_ucs_segunda_via(driver)   # list[dict] com value e text
    tem_dropdown = bool(ucs_disponiveis)
    if tem_dropdown:
        logger.info("  [cnpj ...%s] %d UC(s) no dropdown", cnpj[-6:], len(ucs_disponiveis))
    else:
        logger.info("  [cnpj ...%s] dropdown de UC não encontrado — modo UC única", cnpj[-6:])

    browser_usado    = False  # True após a primeira UC processada via browser
    downloads_bloco  = 0      # contador para pausa preventiva
    for i, linha in enumerate(linhas):
        n            = offset + i + 1
        tensao_class = classificar_tensao(linha.tensao)

        # Pré-check: já existe no índice local → pula sem tocar no browser
        if linha.instalacao in instalacoes_baixadas:
            res = _resultado_base(linha)
            res.update(status="JA_EXISTE", mensagem="Já no índice local (pré-check)")
            resultados.append(res)
            logger.info("  [%d/%d]  inst=%-14s  → já baixada (pulando)",
                        n, total_geral, linha.instalacao)
            continue

        logger.info("")
        logger.info(_SEP)
        logger.info("  [%d/%d]  inst=%-14s  cnpj=...%-6s  %s",
                    n, total_geral, linha.instalacao, cnpj[-6:], tensao_class)
        logger.info(_SEP)

        res = _resultado_base(linha)

        # Volta à Segunda Via somente após a primeira UC processada no browser
        if browser_usado:
            try:
                voltar_para_segunda_via(driver)
            except Exception as exc:
                logger.warning("  falha ao voltar para Segunda Via (%s) — tentando re-login...", exc)
                # Re-login: trata como novo grupo a partir desta UC
                try:
                    abrir_login(driver)
                    relogin_ok = efetuar_login(driver, linha.instalacao, cnpj)
                    if not relogin_ok:
                        relogin_ok = _esta_logado(driver)
                        if relogin_ok:
                            logger.info("  [re-login] confirmado pelo DOM após timeout")
                    if not relogin_ok:
                        raise RuntimeError("Falha no re-login")
                    fechar_modal_pos_login(driver)
                    abrir_segunda_via(driver)
                    # Atualiza dropdown após re-login
                    ucs_disponiveis = listar_ucs_segunda_via(driver)
                    tem_dropdown = bool(ucs_disponiveis)
                    logger.info("  [re-login] sessão restaurada na UC %s", linha.instalacao)
                except Exception as exc2:
                    logger.warning("  re-login falhou: %s", exc2)
                    res.update(status="ERRO", mensagem=f"Navegação+re-login: {exc2}")
                    resultados.append(res)
                    _imprimir_resultado_uc(res, linha)
                    continue

        # Seleciona UC no dropdown (quando existe dropdown com múltiplas opções)
        if tem_dropdown and len(ucs_disponiveis) > 1:
            alvo_dropdown = instalacao_para_dropdown(linha.instalacao)
            valores_disponiveis = [o["value"] for o in ucs_disponiveis]
            if alvo_dropdown not in valores_disponiveis:
                logger.warning("  [%s] value='%s' ausente no dropdown",
                               linha.instalacao, alvo_dropdown)
                res.update(status="ERRO", mensagem=f"UC ausente no dropdown (value={alvo_dropdown})")
                resultados.append(res)
                _imprimir_resultado_uc(res, linha)
                continue
            if not selecionar_uc_segunda_via(driver, linha.instalacao):
                res.update(status="ERRO", mensagem="Falha ao selecionar UC no dropdown")
                resultados.append(res)
                _imprimir_resultado_uc(res, linha)
                continue

        try:
            _fase("Emitindo fatura...")
            preencher_opcoes_fatura(driver)

            _fase("Verificando fatura na página...")
            res = _processar_download_uc(driver, linha, baixados, master)
        except Exception as exc:
            logger.debug("  TRACEBACK:\n%s", traceback.format_exc())
            res.update(status="ERRO", mensagem=str(exc))

        resultados.append(res)
        _imprimir_resultado_uc(res, linha)
        browser_usado = True

        if res.get("status") == "OK":
            downloads_bloco += 1
            if PAUSA_PREVENTIVA_A_CADA > 0 and downloads_bloco >= PAUSA_PREVENTIVA_A_CADA:
                logger.info(
                    "  [pausa] %d downloads — aguardando %ds para não sobrecarregar o portal...",
                    downloads_bloco, PAUSA_PREVENTIVA_SEG,
                )
                time.sleep(PAUSA_PREVENTIVA_SEG)
                downloads_bloco = 0

        pausa_humana(PAUSA_ENTRE_UCS_MIN, PAUSA_ENTRE_UCS_MAX)

    return resultados


def extrair_referencia_do_pdf(pdf_path: Path) -> Optional[str]:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("  [PDF] pdfplumber não instalado — pip install pdfplumber")
        return None

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                # Estratégias 1 e 2: células de tabela
                try:
                    for tabela in (page.extract_tables() or []):
                        celulas = [
                            str(cel).strip()
                            for linha in tabela
                            for cel in (linha or [])
                            if cel
                        ]
                        # Célula seguinte a "Conta mês"
                        for i, cel in enumerate(celulas):
                            norm = (unicodedata.normalize("NFKD", cel)
                                    .encode("ASCII", "ignore").decode("ASCII").upper())
                            if "CONTA MES" in norm:
                                for prox in celulas[i + 1: i + 4]:
                                    ref = extrair_referencia(prox)
                                    if ref:
                                        return ref
                        # Qualquer célula curta com data
                        for cel in celulas:
                            if len(cel) <= 30:
                                ref = extrair_referencia(cel)
                                if ref:
                                    return ref
                except Exception:
                    pass

                # Estratégia 3: texto livre linha a linha
                try:
                    texto  = page.extract_text() or ""
                    linhas = [l.strip() for l in texto.splitlines() if l.strip()]
                    for i, linha in enumerate(linhas):
                        norm = (unicodedata.normalize("NFKD", linha)
                                .encode("ASCII", "ignore").decode("ASCII").upper())
                        if "CONTA MES" in norm:
                            for prox in linhas[i + 1: i + 4]:
                                ref = extrair_referencia(prox)
                                if ref:
                                    return ref
                    for linha in linhas:
                        if len(linha) <= 30:
                            ref = extrair_referencia(linha)
                            if ref:
                                return ref
                except Exception:
                    pass

    except Exception as exc:
        logger.debug("  [PDF] erro ao abrir: %s", exc)

    return None


def _ref_para_yyyymm(ref: str) -> int:
    """Converte 'MM-YYYY' para inteiro YYYYMM para comparação cronológica."""
    try:
        mm, yyyy = ref.split("-")
        return int(yyyy) * 100 + int(mm)
    except Exception:
        return 0


@dataclass
class _LinhaFatura:
    """Representa uma linha da tabela de faturas na página."""
    ref:      str       # MM-YYYY
    elem_dl:  object    # WebElement do link Download


def _ler_tabela_faturas(driver: webdriver.Chrome) -> list[_LinhaFatura]:
    """
    Lê a tabela de faturas disponíveis na página de Segunda Via.
    Retorna lista de _LinhaFatura ordenada do mais recente para o mais antigo.

    A tabela tem colunas: Mês/Ano de referência | Valor | Download | PIX
    Cada <tr> contém um <td> com o mês e outro com o link Download.
    """
    linhas: list[_LinhaFatura] = []

    try:
        # Aguarda a tabela estar presente
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "table"))
        )
    except Exception:
        pass

    try:
        rows = driver.find_elements(By.XPATH, "//table//tr")
        for row in rows:
            tds = row.find_elements(By.TAG_NAME, "td")
            if not tds:
                continue

            # Procura referência de mês em qualquer td da linha
            ref = None
            for td in tds:
                txt = (td.text or "").strip()
                ref = extrair_referencia(txt)
                if ref:
                    break
            if not ref:
                continue

            # Procura o link Download na mesma linha
            elem_dl = None
            SELS_DL = [
                (By.XPATH,
                 ".//a[contains(@onclick,'mostrarFaturaCompleta.aspx')"
                 " and contains(normalize-space(.),'Download')]"),
                (By.XPATH,            ".//a[normalize-space(.)='Download']"),
                (By.PARTIAL_LINK_TEXT, "Download"),
            ]
            for by, loc in SELS_DL:
                try:
                    elems = row.find_elements(by, loc)
                    vis   = [e for e in elems if e.is_displayed()]
                    if vis:
                        elem_dl = vis[0]
                        break
                except Exception:
                    pass

            if elem_dl:
                linhas.append(_LinhaFatura(ref=ref, elem_dl=elem_dl))
                logger.debug("  [tabela] linha: ref=%s  dl_visivel=True", ref)

    except Exception as exc:
        logger.debug("  [tabela] erro ao ler linhas: %s", exc)

    # Ordena do mais recente para o mais antigo
    linhas.sort(key=lambda l: _ref_para_yyyymm(l.ref), reverse=True)
    return linhas


def _inspecionar_pagina_pre_download(driver: webdriver.Chrome) -> Optional[str]:
    """
    Lê a tabela de faturas e retorna a referência MM-YYYY mais recente disponível.
    (Mantida para compatibilidade com o fluxo principal.)
    """
    linhas = _ler_tabela_faturas(driver)
    if linhas:
        logger.debug("  [tabela] %d fatura(s): %s",
                     len(linhas), [l.ref for l in linhas])
        return linhas[0].ref   # mais recente
    return None


def _botao_download_existe(driver: webdriver.Chrome) -> bool:
    """Retorna True se há ao menos uma linha de fatura com botão Download visível."""
    linhas = _ler_tabela_faturas(driver)
    if linhas:
        return True
    # Aguarda um pouco e tenta novamente — página pode ainda estar carregando
    time.sleep(3)
    linhas = _ler_tabela_faturas(driver)
    return bool(linhas)


def clicar_download(driver: webdriver.Chrome) -> None:
    """
    Clica no botão Download da fatura mais recente da tabela.
    Lê a tabela novamente para obter o elemento atualizado (evita StaleElement).
    """
    aba_principal = driver.current_window_handle
    handles_antes = set(driver.window_handles)

    linhas = _ler_tabela_faturas(driver)
    if not linhas:
        raise RuntimeError("Nenhuma linha de fatura encontrada para clicar Download")

    alvo = linhas[0]   # mais recente
    logger.debug("  [dl] clicando Download da fatura %s", alvo.ref)

    ultimo_erro: Optional[Exception] = None
    for t in range(1, 5):
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo.elem_dl)
            pausa_humana(0.2, 0.5)
            try:
                alvo.elem_dl.click()
            except Exception:
                driver.execute_script("arguments[0].click();", alvo.elem_dl)
            break
        except Exception as exc:
            ultimo_erro = exc
            # Elemento pode ter ficado stale — relê a tabela
            try:
                linhas = _ler_tabela_faturas(driver)
                if linhas:
                    alvo = linhas[0]
            except Exception:
                pass
            time.sleep(1.8)
    else:
        raise RuntimeError(f"Falha ao clicar Download ({alvo.ref}): {ultimo_erro}")

    time.sleep(2.5)
    novos = set(driver.window_handles) - handles_antes
    if novos:
        fechar_abas_extras(driver, aba_principal)


def _tem_modal_ok_visivel(driver: webdriver.Chrome) -> bool:
    try:
        for sel in [
            "div.modal-footer input#CONTENT_btnModal",
            "div.modal.show input#CONTENT_btnModal",
            "div.modal.in input#CONTENT_btnModal",
            "input#CONTENT_btnModal",
        ]:
            if any(e.is_displayed() for e in driver.find_elements(By.CSS_SELECTOR, sel)):
                return True
    except Exception:
        pass
    return False


def _clicar_forte(driver: webdriver.Chrome, elem) -> None:
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    pausa_humana(0.1, 0.3)
    for fn in [
        lambda: elem.click(),
        lambda: ActionChains(driver).move_to_element(elem).pause(0.1).click(elem).perform(),
        lambda: driver.execute_script(
            "arguments[0].removeAttribute('disabled');arguments[0].click();", elem
        ),
        lambda: driver.execute_script(
            "var e=arguments[0];"
            "['mousedown','mouseup','click'].forEach(function(t){"
            "  e.dispatchEvent(new MouseEvent(t,"
            "    {view:window,bubbles:true,cancelable:true}));"
            "});",
            elem,
        ),
    ]:
        try:
            fn()
            return
        except Exception:
            pass


def clicar_ok_modal_pos_download(driver: webdriver.Chrome) -> None:
    SELETORES = [
        (By.CSS_SELECTOR, "div.modal-footer input#CONTENT_btnModal"),
        (By.XPATH,
         "//div[contains(@class,'modal-footer')]//input[@id='CONTENT_btnModal']"),
        (By.XPATH,
         "//div[contains(@class,'modal-footer')]"
         "//input[@name='ctl00$CONTENT$btnModal']"),
        (By.CSS_SELECTOR, "div.modal.show input#CONTENT_btnModal"),
        (By.CSS_SELECTOR, "div.modal.in input#CONTENT_btnModal"),
        (By.ID,           "CONTENT_btnModal"),
        (By.XPATH,        "//input[@name='ctl00$CONTENT$btnModal']"),
    ]

    fim        = time.time() + TIMEOUT_MODAL_OK
    ultimo_err = None

    while time.time() < fim:
        _detectar_alerta(driver)
        for by, loc in SELETORES:
            try:
                elems = driver.find_elements(by, loc)
                if not elems:
                    continue
                alvo = next((e for e in elems if e.is_displayed()), elems[0])
                try:
                    WebDriverWait(driver, 2).until(lambda d: alvo.is_displayed())
                except Exception:
                    pass
                _clicar_forte(driver, alvo)
                time.sleep(0.8)
                if not _tem_modal_ok_visivel(driver):
                    time.sleep(1.2)
                    return
                try:
                    driver.execute_script(
                        "var f=arguments[0].closest('form');if(f)f.submit();", alvo
                    )
                    time.sleep(0.8)
                except Exception:
                    pass
                if not _tem_modal_ok_visivel(driver):
                    time.sleep(1.2)
                    return
            except Exception as exc:
                ultimo_err = exc
        time.sleep(1.0)

    raise RuntimeError(
        f"Modal OK pós-download não fechou em {TIMEOUT_MODAL_OK}s: {ultimo_err}"
    )


# =============================================================================
# PROCESSAMENTO DE UMA UNIDADE
# =============================================================================

def _resultado_base(linha: LinhaAcesso) -> dict:
    return {
        "instalacao": linha.instalacao,
        "cnpj":       linha.cnpj,
        "tensao":     linha.tensao,
        "status":     "",
        "mensagem":   "",
        "mes_ref":    "",
        "arquivo":    "",
    }


def processar_uma_unidade(
    driver: webdriver.Chrome,
    linha: LinhaAcesso,
    baixados: set[tuple[str, str]],
    master,
) -> dict:
    res = _resultado_base(linha)

    if len(linha.instalacao) < MIN_DIGITOS_INSTALACAO:
        res.update(status="ERRO", mensagem=f"Instalação inválida: '{linha.instalacao}'")
        return res
    if len(linha.cnpj) < MIN_DIGITOS_CNPJ:
        res.update(status="ERRO", mensagem=f"CNPJ inválido: '{linha.cnpj}'")
        return res

    tensao_class = classificar_tensao(linha.tensao)

    # ── Login ──────────────────────────────────────────────────────────────
    _fase("Login...")
    abrir_login(driver)
    if not efetuar_login(driver, linha.instalacao, linha.cnpj):
        res.update(status="ERRO", mensagem="Falha no login")
        return res

    # ── Modal pós-login ────────────────────────────────────────────────────
    _fase("Fechando modal pós-login...")
    fechar_modal_pos_login(driver)

    # ── Segunda Via ────────────────────────────────────────────────────────
    _fase("Abrindo Segunda Via...")
    abrir_segunda_via(driver)

    # ── Opções + Emitir ────────────────────────────────────────────────────
    _fase("Emitindo fatura...")
    preencher_opcoes_fatura(driver)

    # ── Referência pré-download + verificação de duplicata ─────────────────
    _fase("Verificando fatura na página...")
    mes_ref_pre    = _inspecionar_pagina_pre_download(driver)
    refs_pendentes: list[str] = []

    if mes_ref_pre:
        if ja_foi_baixado_local(baixados, linha.instalacao, mes_ref_pre):
            res.update(
                status="JA_EXISTE",
                mensagem=f"Já existe: {mes_ref_pre}",
                mes_ref=mes_ref_pre,
            )
            return res
        refs_pendentes = [mes_ref_pre]

    if not _botao_download_existe(driver):
        motivo = "sem botão Download" + (
            f" (ref: {mes_ref_pre})" if mes_ref_pre else " (sem ref na página)"
        )
        res.update(status="SEM_FATURA", mensagem=motivo)
        return res

    # ── Download ───────────────────────────────────────────────────────────
    _fase("Baixando PDF...")
    antes = {p.name for p in TEMP_DOWNLOAD_DIR.glob("*")}
    clicar_download(driver)
    time.sleep(2.5)

    try:
        clicar_ok_modal_pos_download(driver)
    except Exception as exc:
        logger.debug("  modal OK não imediato: %s", exc)

    pdf = aguardar_novo_pdf(TEMP_DOWNLOAD_DIR, antes, timeout=TIMEOUT_DOWNLOAD)

    if _tem_modal_ok_visivel(driver):
        try:
            clicar_ok_modal_pos_download(driver)
        except Exception:
            pass

    if pdf:
        logger.debug("  PDF: %s (%d KB)", pdf.name, pdf.stat().st_size // 1024)
    else:
        logger.debug("  PDF não encontrado dentro do timeout")

    # ── Resolve referência ─────────────────────────────────────────────────
    mes_ref = refs_pendentes[0] if refs_pendentes else None

    if not mes_ref and pdf and pdf.exists():
        _fase("Extraindo referência do PDF...")
        mes_ref = extrair_referencia_do_pdf(pdf)

    if not mes_ref:
        if pdf and pdf.exists():
            try:
                pdf.unlink()
            except Exception:
                pass
        res.update(status="ERRO", mensagem="Referência da fatura não identificada")
        return res

    res["mes_ref"] = mes_ref

    # Duplicata final (descoberta via PDF)
    if ja_foi_baixado_local(baixados, linha.instalacao, mes_ref):
        if pdf and pdf.exists():
            try:
                pdf.unlink()
            except Exception:
                pass
        res.update(status="JA_EXISTE", mensagem=f"Já existe: {mes_ref}")
        return res

    if not pdf or not pdf.exists():
        res.update(status="ERRO", mensagem="PDF não encontrado após download")
        return res

    if not re.fullmatch(r"(0[1-9]|1[0-2])-20\d{2}", mes_ref):
        if pdf and pdf.exists():
            try:
                pdf.unlink()
            except Exception:
                pass
        res.update(status="ERRO", mensagem=f"mes_ref inválido: '{mes_ref}'")
        return res

    # ── Move PDF para a rede ───────────────────────────────────────────────
    _fase(f"Salvando {mes_ref} ({tensao_class})...")
    indice_bb     = master.consumir_carimbo()
    pasta_destino = BASE_DIR / mes_ref / tensao_class
    pasta_destino.mkdir(parents=True, exist_ok=True)

    caminho_final = pasta_destino / f"{indice_bb}.pdf"
    if caminho_final.exists():
        caminho_final.unlink()
    shutil.move(str(pdf), str(caminho_final))
    logger.debug("  %s → %s", pdf.name, caminho_final)

    # ── Registros ──────────────────────────────────────────────────────────
    master.registrar(
        indice_bb=indice_bb,
        sistema="EQUATORIAL",
        uc=linha.instalacao,
        mes_ref=mes_ref,
        fatura_id="",
        cnpj=linha.cnpj,
        estado="GO",
        instalacao=linha.instalacao,
        arquivo=str(caminho_final),
    )
    registrar_indice_local(
        indice_bb=indice_bb,
        instalacao=linha.instalacao,
        mes_ref=mes_ref,
        arquivo=str(caminho_final),
        tensao=tensao_class,
    )
    baixados.add((linha.instalacao.strip(), mes_ref.strip()))

    res.update(
        status="OK",
        mensagem="Download concluído",
        indice=indice_bb,
        arquivo=str(caminho_final),
    )
    return res


# =============================================================================
# RESUMO
# =============================================================================

def salvar_resumo(resultados: list[dict]) -> None:
    if not resultados:
        return
    try:
        RESUMO_XLSX.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(resultados).to_excel(RESUMO_XLSX, index=False)
    except Exception:
        logger.error("Falha ao salvar resumo: %s",
                     traceback.format_exc().strip().splitlines()[-1])


# =============================================================================
# MAIN
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Downloader Equatorial Goias")
    parser.add_argument(
        "--lista-especifica",
        help="CSV com colunas Instalacao e CNPJ para rodar somente esse lote.",
    )
    parser.add_argument(
        "--sem-filtro",
        action="store_true",
        help="Ignora a lista especifica e processa todas as UCs da planilha.",
    )
    return parser.parse_args()


def main() -> None:
    from collections import defaultdict

    args = parse_args()

    BASE_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    logger.info(_SEP2)
    logger.info("  EQUATORIAL GOIÁS  —  %s", datetime.now().strftime("%d/%m/%Y %H:%M"))
    logger.info("  Destino     : %s", BASE_DIR)
    logger.info("  Temp local  : %s", TEMP_DOWNLOAD_DIR)
    logger.info(_SEP2)

    filtro_instalacoes, origem_filtro = _resolver_filtro_instalacoes(args)
    logger.info("  Filtro      : %s", origem_filtro)
    if filtro_instalacoes:
        logger.info("  Instalacoes : %d acessos especificos", len(filtro_instalacoes))

    unidades = carregar_unidades(PLANILHA_PATH, filtro_instalacoes)
    total    = len(unidades)

    if not unidades:
        logger.info("Nenhuma UC encontrada na planilha.")
        return

    master   = carregar_master()
    baixados = _carregar_indice_local()
    logger.info("  Índice local : %d registros já processados", len(baixados))
    carregar_mapa_instalacoes()

    # Agrupa por CNPJ — 1 login por grupo
    grupos: dict[str, list[LinhaAcesso]] = defaultdict(list)
    for linha in unidades:
        grupos[linha.cnpj].append(linha)
    logger.info("  CNPJs distintos: %d  |  UCs total: %d", len(grupos), total)
    logger.info(_SEP2)

    resultados: list[dict] = []
    ger    = GerenciadorDriver()
    offset = 0

    try:
        for cnpj, linhas_grupo in grupos.items():
            tentativa_grupo = 0
            while True:
                driver = ger.obter()
                try:
                    res_grupo = processar_grupo_cnpj(
                        driver, cnpj, linhas_grupo, baixados, master,
                        total_geral=total, offset=offset,
                    )
                    ger.registrar_sucesso()
                    resultados.extend(res_grupo)
                    break
                except SessaoChromeInstavel as exc:
                    tentativa_grupo += 1
                    logger.warning("  [driver] %s", exc)
                    ger.reiniciar_agora("alerta #002 recorrente no login")
                    if tentativa_grupo >= 2:
                        for linha in linhas_grupo:
                            r = _resultado_base(linha)
                            r.update(status="ERRO", mensagem=str(exc))
                            resultados.append(r)
                        ger.registrar_erro()
                        break
                    logger.info("  [driver] repetindo grupo com Chrome novo (%d/2)", tentativa_grupo)
                    continue
                except Exception as exc:
                    logger.debug("  TRACEBACK:\n%s", traceback.format_exc())
                    if _erro_sessao_driver(exc) and tentativa_grupo < 1:
                        tentativa_grupo += 1
                        logger.warning("  [driver] sessão perdida — reiniciando para grupo", )
                        ger.reiniciar_agora("sessao do Chrome perdida")
                        continue
                    for linha in linhas_grupo:
                        r = _resultado_base(linha)
                        r.update(status="ERRO", mensagem=str(exc))
                        resultados.append(r)
                    ger.registrar_erro()
                    break

            offset += len(linhas_grupo)

    finally:
        ger.fechar()

    salvar_resumo(resultados)

    ok = sum(1 for r in resultados if r["status"] == "OK")
    ja = sum(1 for r in resultados if r["status"] == "JA_EXISTE")
    sf = sum(1 for r in resultados if r["status"] == "SEM_FATURA")
    er = sum(1 for r in resultados if r["status"] == "ERRO")

    logger.info("")
    logger.info(_SEP2)
    logger.info(
        "  FINALIZADO  —  ✓ OK=%-3d  = Existia=%-3d"
        "  - Sem fatura=%-3d  ✗ Erro=%-3d  / Total=%d",
        ok, ja, sf, er, total,
    )
    logger.info("  Resumo: %s", RESUMO_XLSX)
    logger.info(_SEP2)


if __name__ == "__main__":
    main()
