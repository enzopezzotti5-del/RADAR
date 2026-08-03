# -*- coding: utf-8 -*-
# pip install selenium openpyxl beautifulsoup4
# Consen_CEMIG.py — digitação de faturas CEMIG no sistema Consen

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import _venv_check  # noqa

import argparse
import json
import os
import traceback
import time
import csv
import re
import unicodedata
from datetime import datetime, date

try:
    from dotenv import load_dotenv
    for _env_path in [
        Path(__file__).resolve().parent.parent / ".env",
        Path(__file__).resolve().parent.parent.parent / ".env",
    ]:
        if _env_path.exists():
            load_dotenv(_env_path)
except ImportError:
    pass

import openpyxl
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

try:
    from digitacao_consen.auditoria_schema import (
        AUDITORIA_HEADERS,
        append_resultado_auditoria,
        extrair_status_auditoria,
    )
    from digitacao_consen.consen_credentials import resolver_credenciais_consen
except ModuleNotFoundError:
    from auditoria_schema import (  # type: ignore
        AUDITORIA_HEADERS,
        append_resultado_auditoria,
        extrair_status_auditoria,
    )
    from consen_credentials import resolver_credenciais_consen  # type: ignore


# =========================================================
# CONFIGURAÇÕES
# =========================================================

LOGIN_URL = "https://consen.acaoengenharia.com.br/login.php"
TARGET_HASH = "#bpg/gestao/fatura/cadastroTabFatura.php"
TARGET_URL = f"{LOGIN_URL.rsplit('/', 1)[0]}/index.php{TARGET_HASH}"
LINK_HREF = "bpg/gestao/fatura/cadastroTabFatura.php"
LINK_TEXTO = "Instalacao"

USUARIO, SENHA = resolver_credenciais_consen()
os.environ["CONSEN_USUARIO"] = USUARIO
os.environ["CONSEN_SENHA"] = SENHA

# ── Configuração padrão (pode ser sobrescrita via CLI --xlsx / --linha-inicio) ──
# Quando chamado pelo pipeline_cemig.py, esses valores vêm por argumento.
# Quando chamado manualmente, edite _EXCEL_PATH_PADRAO abaixo.
_EXCEL_PATH_PADRAO = r"\\fs01\Energia\ARQUIVOS ENZO\OCR CEMIG\ocr_cemig_BT_032026.xlsx"
_LINHA_INICIO_PADRAO = 2

EXCEL_PATH: str = _EXCEL_PATH_PADRAO
LINHA_INICIO: int = _LINHA_INICIO_PADRAO
PASTA_ORIGEM_PDFS: str = ""

PASTA_SAIDA = Path(
    os.environ.get(
        "CONSEN_PIPELINE_SAIDA",
        str(Path(__file__).resolve().parent.parent / "pipelines" / "saida_importacao"),
    )
)
try:
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
except OSError:
    pass
AUDITORIA_CSV = PASTA_SAIDA / "auditoria_resultados.csv"

MAPEAMENTO_CSV = Path(r"C:\SEU\CAMINHO\mapeamento_campos_planilha.csv")
SCORE_MINIMO_MAPEAMENTO = 35
HEADLESS = False


def _parse_args():
    """Processa argumentos CLI e atualiza EXCEL_PATH / LINHA_INICIO globais."""
    global EXCEL_PATH, LINHA_INICIO, PASTA_ORIGEM_PDFS
    p = argparse.ArgumentParser(
        description="Digitação CEMIG no Consen",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p.add_argument(
        "--xlsx",
        type=str,
        default=None,
        help="Caminho completo do xlsx gerado pelo OCR.",
    )
    p.add_argument(
        "--linha-inicio",
        type=int,
        default=None,
        dest="linha_inicio",
        help="Linha Excel a partir da qual processar (padrão: 2).",
    )
    p.add_argument(
        "--pasta-origem",
        type=str,
        default=None,
        help="Pasta BT de origem para limitar a execucao aos PDFs presentes nela.",
    )
    args, _ = p.parse_known_args()
    if args.xlsx:
        EXCEL_PATH = args.xlsx
    if args.linha_inicio is not None:
        LINHA_INICIO = args.linha_inicio
    if args.pasta_origem:
        PASTA_ORIGEM_PDFS = args.pasta_origem

# =========================================================
# MAPA DE CONVERSÃO PARA SELECTS
# =========================================================

MAP_SELECTS: dict[str, dict[str, str]] = {
    "cb-dados-contratuais-fatura-tarifa": {
        "Convencional": "Convencional",
        "THS Verde": "HS - Verde",
        "TUSD Livre Verde": "HS - Verde",
        "HS - Verde": "HS - Verde",
    },
    "cb-dados-contratuais-fatura-subgrupo": {
        "B3": "B3",
        "B3 [<2,3kV]": "B3 [<2,3kV]",
        "B1": "B1",
        "A4": "A4",
        "A4 [2,3kV a 25kV]": "A4 [2,3kV a 25kV]",
        "A4 [2,3 a 25 kV]": "A4 [2,3kV a 25kV]",
        "A4 [<13,8kV]": "A4 [2,3kV a 25kV]",
        "A3": "A3",
        "A3 [69 kV]": "A3 [69 kV]",
        "A3 [<44kV]": "A3 [69 kV]",
        "A3A": "A3a [30kV a 44kV]",
        "A3a": "A3a [30kV a 44kV]",
        "A3a [30kV a 44kV]": "A3a [30kV a 44kV]",
        "A2": "A2",
        "A2  [88 kV a 138 kV]": "A2  [88 kV a 138 kV]",
        "A1": "A1",
        "AS": "AS [<2,3kV]",
        "AS [<2,3kV]": "AS [<2,3kV]",
    },
    "cb-dados-financeiros-obs": {
        "1": "1", "6": "6", "7": "7", "8": "8", "10": "10", "11": "11",
        "12": "12", "13": "13", "14": "14", "16": "16", "17": "17", "18": "18",
        "23": "23", "34": "34", "35": "35", "36": "36", "44": "44", "47": "47",
        "48": "48", "49": "49", "51": "51", "54": "54", "55": "55", "56": "56",
        "57": "57", "58": "58", "59": "59", "60": "60", "63": "63", "64": "64",
        "65": "65", "67": "67", "68": "68", "69": "69", "70": "70", "71": "71",
        "72": "72", "73": "73", "74": "74", "75": "75", "76": "76", "77": "77",
        "78": "78", "79": "79", "80": "80", "81": "81", "82": "82", "83": "83",
        "84": "84", "86": "86", "87": "87", "88": "88", "90": "90", "91": "91",
        "92": "92", "93": "93", "94": "94", "95": "95", "96": "96", "97": "97",
        "98": "98", "99": "99", "100": "100", "101": "101", "102": "102",
        "103": "103", "108": "108", "109": "109", "110": "110", "111": "111",
        "113": "113", "114": "114", "115": "115", "116": "116", "119": "119",
        "123": "123", "124": "124", "125": "125", "126": "126", "129": "129",
        "130": "130", "131": "131", "132": "132", "133": "133", "134": "134",
        "135": "135", "136": "136", "137": "137", "139": "139", "140": "140",
        "141": "141", "142": "142", "143": "143", "144": "144", "145": "145",
        "146": "146", "147": "147", "148": "148", "149": "149", "150": "150",
        "151": "151", "152": "152", "154": "154", "155": "155", "156": "156",
        "157": "157", "158": "158", "159": "159", "160": "160", "161": "161",
        "162": "162", "163": "163", "164": "164", "165": "165", "166": "166",
        "167": "167", "168": "168", "169": "169", "170": "170", "171": "171",
        "172": "172", "173": "173", "174": "174", "175": "175", "176": "176",
        "177": "177", "178": "178", "180": "180", "181": "181", "182": "182",
        "183": "183", "184": "184", "185": "185", "186": "186", "187": "187",
        "188": "188", "189": "189", "190": "190", "191": "191", "192": "192",
        "193": "193", "194": "194", "195": "195", "196": "196", "197": "197",
        "198": "198", "200": "200", "201": "201", "204": "204", "206": "206",
        "208": "208", "213": "213", "219": "219", "220": "220", "222": "222",
        "223": "223", "224": "224", "225": "225", "229": "229", "230": "230",
        "231": "231", "234": "234", "235": "235", "236": "236", "237": "237",
        "242": "242", "244": "244", "245": "245", "246": "246", "247": "247",
        "248": "248", "249": "249", "250": "250", "251": "251", "252": "252",
        "253": "253", "254": "254", "255": "255", "256": "256", "257": "257",
        "258": "258", "259": "259", "260": "260", "261": "261", "262": "262",
        "263": "263", "264": "264", "265": "265", "266": "266", "267": "267",
        "268": "268", "269": "269", "270": "270", "271": "271", "272": "272",
        "273": "273", "274": "274", "275": "275", "276": "276", "277": "277",
        "278": "278", "279": "279", "280": "280", "281": "281", "282": "282",
        "283": "283", "284": "284", "285": "285", "286": "286", "287": "287",
        "288": "288", "289": "289", "290": "290", "291": "291", "292": "292",
        "293": "293", "294": "294", "295": "295", "296": "296", "297": "297",
        "298": "298", "299": "299", "300": "300", "301": "301", "302": "302",
    },
}


# =========================================================
# ALIASES FORÇADOS
# =========================================================

ALIASES_FORCADOS = {
    # ── Datas
    "dataVencimento": "fatDataVcto",
    "dataEmissao": "fatDataEmissao",

    # ── Selects
    "cb-dados-contratuais-fatura-tarifa": "cadTarifaCod",
    "cb-dados-contratuais-fatura-subgrupo": "cadSubGrupoCod",

    # ── CONSUMO
    "txt-consumo-registrada-pta": "fatConPontaRegistrado",
    "txt-consumo-faturada-pta": "fatConPontaFaturado",
    "txt-consumo-pta-valor-reais": "fatConPontaValorReais",
    "txt-consumo-registrada-fpind": "fatConFPontaIndRegistrado",
    "txt-consumo-faturada-fpind": "fatConFPontaIndFaturado",
    "txt-consumo-fpind-valor-reais": "fatConFPontaIndValorReais",
    "txt-consumo-registrada-inter": "fatConIntermediarioRegistrado",
    "txt-consumo-faturada-inter": "fatConIntermediarioFaturado",
    "txt-consumo-inter-valor-reais": "fatConIntermediarioValorReais",
    "txt-consumo-excedente-registrada-pta": "fatConPontaExcRegistrado",
    "txt-consumo-excedente-faturada-pta": "fatConPontaExcFaturado",
    "txt-consumo-excedente-pta-valor-reais": "fatConPontaExcValorReais",
    "txt-consumo-excedente-registrada-fpind": "fatConFPontaIndExcRegistrado",
    "txt-consumo-excedente-faturada-fpind": "fatConFPontaIndExcFaturado",
    "txt-consumo-excedente-fpind-valor-reais": "fatConFPontaIndExcValorReais",
    "txt-consumo-excedente-registrada-fpcap": "fatConFPontaCapExcRegistrado",
    "txt-consumo-excedente-faturada-fpcap": "fatConFPontaCapExcFaturado",

    # ── DEMANDAS
    "txt-demandas-registrada-pta": "fatDemPontaRegistrada",
    "txt-demandas-faturada-pta": "fatDemPontaFaturada",
    "txt-demandas-pta-valor-reais": "fatDemPontaValorReais",
    "txt-demandas-registrada-fpind": "fatDemFPontaIndRegistrada",
    "txt-demandas-faturada-fpind": "fatDemFPontaIndFaturada",
    "txt-demandas-fpind-valor-reais": "fatDemFPontaIndValorReais",

    # ── DEMANDAS CONTRATADAS
    "txt-dados-contratuais-fatura-dem-cont-p": "fatDemContratadaPonta",
    "txt-dados-contratuais-fatura-dem-cont-fp": "fatDemContratadaFPonta",
    "txt-dados-contratuais-fatura-dem-cont-geracao-p": "fatDemContratadaGeracaoPonta",
    "txt-dados-contratuais-fatura-dem-cont-geracao-fp": "fatDemContratadaGeracaoFPonta",

    # ── INJETADO / GD
    "txt-consumo-injetado-registrado-fpta": "fatConFPontaInjetadoRegistrado",
    "txt-consumo-injetado-faturado-fpta": "fatConFPontaInjetadoFaturado",
    "txt-consumo-injetado-fpta-valor-reais": "fatConFPontaInjetadoValorReais",
    "txt-consumo-injetado-usina-fpta": "fatConFPontaInjetadoUsina",
    "txt-consumo-injetado-usina-pta": "fatConPontaInjetadoUsina",

    # ── SALDOS USINA
    "txt-consumo-injetado-usina-fpta-saldo": "fatConFPontaInjetadoUsinaSaldoAcumulado",
    "txt-consumo-injetado-usina-pta-saldo": "fatConPontaInjetadoUsinaSaldoAcumulado",

    # ── CÓDIGO DE BARRAS
    "fatCodigoBarras": "fatCodigoBarras",
    "txt-dados-financeiros-codigo-barra": "fatCodigoBarras",

    # ── FINANCEIROS
    "camposFinanIlumimnacaoPublica": "fatIlumPublica",
    "camposFinanICMS": "fatICMS",
    "txt-dados-financeiros-pis-pasep": "fatPIS",
    "txt-dados-financeiros-cofins": "fatCOFINS",
    "txt-dados-financeiros-valor-nota-fiscal": "fatValorNotaFiscal",
    "txt-dados-financeiros-valor-fatura-a-pagar": "fatValorFatura",
}

_OBS_PARES_MAX = 5

ALIASES_FORCADOS_HEADERS_ALTERNATIVOS = {
    # OCR MT CEMIG pode expor os campos com nome explícito de reativo.
    # No Consen eles entram nos mesmos inputs de consumo excedente.
    "txt-consumo-excedente-registrada-pta": [
        "fatConPontaReativoExcedente",
    ],
    "txt-consumo-excedente-faturada-pta": [
        "fatConPontaReativoFaturado",
        "fatConPontaReativoExcedente",
    ],
    "txt-consumo-excedente-registrada-fpind": [
        "fatConFPontaIndReativoExcedente",
    ],
    "txt-consumo-excedente-faturada-fpind": [
        "fatConFPontaIndReativoFaturado",
        "fatConFPontaIndReativoExcedente",
    ],
}

CAMPOS_BLOQUEADOS_CSV = {
    "txt-consumo-excedente-registrada-fpind",
    "txt-consumo-excedente-faturada-fpind",
    "txt-consumo-excedente-registrada-fpcap",
    "txt-consumo-excedente-faturada-fpcap",
    "txt-consumo-registrada-fpcap",
    "txt-consumo-faturada-fpcap",
    "txt-demandas-geracao-registrada-fpta",
    "txt-demandas-geracao-faturada-fpta",
    "txt-demandas-excedente-registrada-fpta",
    "txt-demandas-excedente-faturada-fpta",
    "txt-demandas-registrada-fpcap",
    "txt-demandas-ultrapassagem-faturada-fpind",
    "txt-demandas-ultrapassagem-faturada-pta",
    "instalacaoMedidor",
    "enderecoUnidade",
    "enderecoMedidor",
    "cidadeUnidade",
    "cidadeMedidor",
}

# Headers da planilha que NUNCA devem ser digitados no Consen CEMIG,
# independentemente do mapeamento CSV.
# fatTributoFederal* não é campo recorrente no Consen — quando enviado
# causa preenchimento indevido. Bloqueado aqui na origem.
HEADERS_BLOQUEADOS_PLANILHA = {
    "fatTributoFederalPerc",
    "fatTributoFederalVal",
    "fatDIC",
    "fatFIC",
}


# =========================================================
# LOG
# =========================================================

def log(msg: str):
    print(f"[INFO] {msg}")


def warn(msg: str):
    print(f"[WARN] {msg}")


def erro(msg: str):
    print(f"[ERRO] {msg}")


# =========================================================
# UTILITÁRIOS
# =========================================================

def preencher_saldo_usina_tabela(driver, wait, campo_id, valor):
    if valor is None or str(valor).strip() == "":
        warn(f"[USINA TABELA] Valor vazio para {campo_id}")
        return False

    valor_fmt = formatar_numero_br(valor)

    try:
        campo = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.ID, campo_id))
        )
    except Exception:
        warn(f"[USINA TABELA] Campo não encontrado: {campo_id}")
        return False

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
        time.sleep(0.3)

        driver.execute_script("arguments[0].focus();", campo)
        time.sleep(0.1)

        try:
            campo.click()
            time.sleep(0.1)
        except Exception:
            pass

        try:
            campo.send_keys(Keys.CONTROL, "a")
            campo.send_keys(Keys.DELETE)
        except Exception:
            pass

        driver.execute_script("""
            arguments[0].value = '';
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
        """, campo)

        time.sleep(0.15)

        campo.send_keys(valor_fmt)
        time.sleep(0.15)

        driver.execute_script("""
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new KeyboardEvent('keyup', {bubbles:true, key:'0'}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));
        """, campo)

        try:
            campo.send_keys(Keys.TAB)
        except Exception:
            pass

        time.sleep(0.3)

        valor_final = (campo.get_attribute("value") or "").strip()
        log(f"[USINA TABELA] {campo_id} | enviado={valor_fmt!r} | final={valor_final!r}")

        vf = valor_final.replace(".", "").replace(",", "")
        ve = valor_fmt.replace(".", "").replace(",", "")
        return vf == ve

    except Exception as e:
        warn(f"[USINA TABELA] Erro ao preencher {campo_id}: {type(e).__name__} - {e}")
        return False

def normalizar_texto(txt):
    if txt is None:
        return ""
    txt = str(txt).strip()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return txt.lower().strip()


def normalizar_slug(txt):
    txt = normalizar_texto(txt)
    txt = re.sub(r"[^a-z0-9]+", "-", txt)
    txt = re.sub(r"-+", "-", txt).strip("-")
    return txt


def tokenizar(txt):
    slug = normalizar_slug(txt)
    return set(slug.split("-")) if slug else set()


def iniciar_driver(headless=False):
    options = Options()
    if headless:
        options.add_argument("--headless=new")

    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-save-password-bubble")
    options.add_argument("--disable-notifications")

    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
        "autofill.profile_enabled": False,
        "autofill.credit_card_enabled": False,
    }

    options.add_experimental_option("prefs", prefs)
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)

    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            """
        },
    )

    return driver


def formatar_ddmmyyyy(dt):
    return dt.strftime("%d/%m/%Y")


def formatar_yyyy_mm_dd(dt):
    return dt.strftime("%Y-%m-%d")


def parse_data_ddmmyyyy(txt):
    return datetime.strptime(txt.strip(), "%d/%m/%Y").date()


def primeiro_dia_mes(dt):
    return date(dt.year, dt.month, 1)


def valor_excel_para_date(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        if float(valor) == 0:
            return None
    if isinstance(valor, str):
        valor = valor.strip()
        if valor in ("", "0", "0,0", "0,00", "None", "nan"):
            return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(valor, fmt).date()
            except ValueError:
                pass
    raise Exception(f"Não consegui converter para data: {valor!r}")


def valor_para_numero(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    s = str(valor).strip()
    if not s:
        return None

    s = s.replace("R$", "").replace(" ", "")

    # Formato brasileiro: tem vírgula E ponto → ponto é milhar, vírgula é decimal
    # Ex: "5.843,21" → 5843.21  |  "1.234.567,89" → 1234567.89
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    # Só vírgula → vírgula é decimal (padrão BR sem separador de milhar)
    # Ex: "220,45" → 220.45
    elif "," in s:
        s = s.replace(",", ".")
    # Só ponto: ambíguo — se há exatamente 3 dígitos após o ponto, pode ser milhar
    # Ex: "112.56" → decimal americano (2 casas) → 112.56
    # Ex: "9.820" → milhar BR (3 casas) → 9820
    elif "." in s:
        partes = s.split(".")
        if len(partes) == 2 and len(partes[1]) == 3 and partes[1].isdigit():
            # Provávelmente separador de milhar BR: "9.820" → 9820
            s = s.replace(".", "")
        # Senão mantém como decimal americano: "112.56" → 112.56

    try:
        return float(s)
    except ValueError:
        return None


def formatar_numero_br(valor):
    num = valor_para_numero(valor)
    if num is None:
        return str(valor).strip()
    # SEM separador de milhar — campo do Consen tem máscara que interpreta
    # ponto como decimal se vier "5.956,26", gerando "595.626,00"
    if abs(num - round(num)) < 1e-9:
        return f"{int(round(num))},00"
    return f"{num:.2f}".replace(".", ",")


def obter_valor_planilha_por_headers(dados_planilha, *headers_possiveis):
    for header in headers_possiveis:
        if header in dados_planilha:
            valor = dados_planilha[header]
            if valor is not None and str(valor).strip() != "":
                return valor, header
    return None, None


# =========================================================
# LOGIN
# =========================================================

def localizar_campos_login(driver, wait):
    seletores_usuario = [
        (By.NAME, "usuario"),
        (By.NAME, "username"),
        (By.NAME, "login"),
        (By.ID, "usuario"),
        (By.ID, "username"),
        (By.ID, "login"),
        (By.CSS_SELECTOR, "input[type='text']"),
        (By.CSS_SELECTOR, "input[type='email']"),
    ]
    seletores_senha = [
        (By.NAME, "senha"),
        (By.NAME, "password"),
        (By.ID, "senha"),
        (By.ID, "password"),
        (By.CSS_SELECTOR, "input[type='password']"),
    ]

    campo_usuario = None
    campo_senha = None

    for by, sel in seletores_usuario:
        try:
            campo_usuario = wait.until(EC.presence_of_element_located((by, sel)))
            log(f"Campo usuário encontrado: {by} = {sel}")
            break
        except TimeoutException:
            pass

    for by, sel in seletores_senha:
        try:
            campo_senha = wait.until(EC.presence_of_element_located((by, sel)))
            log(f"Campo senha encontrado: {by} = {sel}")
            break
        except TimeoutException:
            pass

    if not campo_usuario or not campo_senha:
        raise Exception("Não foi possível localizar os campos de login.")

    return campo_usuario, campo_senha


def enviar_login(driver, wait, usuario, senha):
    campo_usuario, campo_senha = localizar_campos_login(driver, wait)

    campo_usuario.clear()
    campo_usuario.send_keys(usuario)

    campo_senha.clear()
    campo_senha.send_keys(senha)

    seletores_submit = [
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.CSS_SELECTOR, "input[type='submit']"),
        (By.XPATH, "//button[contains(., 'Entrar')]"),
        (By.XPATH, "//button[contains(., 'Login')]"),
        (By.XPATH, "//button[contains(., 'Acessar')]"),
        (By.XPATH, "//input[contains(@value, 'Entrar')]"),
    ]

    for by, sel in seletores_submit:
        try:
            btn = driver.find_element(by, sel)
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.3)
            btn.click()
            log(f"Botão de login clicado: {by} = {sel}")
            return
        except Exception:
            pass

    warn("Botão de login não encontrado. Tentando ENTER.")
    campo_senha.send_keys(Keys.ENTER)


# =========================================================
# PLANILHA
# =========================================================

def encontrar_coluna_por_header(ws, header_desejado):
    desejado = normalizar_texto(header_desejado)

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if normalizar_texto(valor) == desejado:
            return col

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if desejado in normalizar_texto(valor):
            return col

    raise Exception(f"Não encontrei a coluna '{header_desejado}'.")


def obter_headers_planilha(caminho_arquivo):
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    return [str(h).strip() for h in headers if h is not None and str(h).strip()]


def ler_primeira_linha_completa_planilha(caminho_arquivo):
    if not os.path.exists(caminho_arquivo):
        raise Exception(f"Planilha não encontrada: {caminho_arquivo}")

    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        headers.append(str(h).strip() if h is not None else "")

    linha = 2
    dados = {}

    for col, header in enumerate(headers, start=1):
        if not header:
            continue
        dados[header] = ws.cell(row=linha, column=col).value

    return dados


def _normalizar_nome_coluna(nome) -> str:
    txt = str(nome or "").strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", txt)


def _obter_valor_por_alias(dados: dict, aliases: list[str]):
    if not dados:
        return None
    mapa = {_normalizar_nome_coluna(k): v for k, v in dados.items()}
    for alias in aliases:
        chave = _normalizar_nome_coluna(alias)
        if chave in mapa:
            return mapa[chave]
    return None


def normalizar_datas_leitura_por_referencia(data_referencia, leitura_anterior, leitura_atual):
    if data_referencia is None or leitura_anterior is None or leitura_atual is None:
        return leitura_anterior, leitura_atual

    ano_atual_esperado = data_referencia.year
    ano_anterior_esperado = (
        data_referencia.year - 1
        if leitura_anterior.month > leitura_atual.month
        else data_referencia.year
    )

    leitura_atual_norm = leitura_atual.replace(year=ano_atual_esperado)
    leitura_anterior_norm = leitura_anterior.replace(year=ano_anterior_esperado)
    return leitura_anterior_norm, leitura_atual_norm


def ler_todas_as_linhas_planilha(caminho_arquivo, linha_inicio=2):
    if not os.path.exists(caminho_arquivo):
        raise Exception(f"Planilha não encontrada: {caminho_arquivo}")

    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[wb.sheetnames[0]]
    log(f"Aba lida: {ws.title}")

    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        headers.append(str(h).strip() if h is not None else "")

    registros = []
    for linha in range(linha_inicio, ws.max_row + 1):
        dados = {}
        for col, header in enumerate(headers, start=1):
            if not header:
                continue
            dados[header] = ws.cell(row=linha, column=col).value

        instalacao = _obter_valor_por_alias(dados, ["Instalacao", "Instalação", "instalacao"])
        if instalacao is None or str(instalacao).strip() == "":
            continue

        leitura_anterior_raw = _obter_valor_por_alias(
            dados,
            ["fatDataLeituraAnterior", "Data Leitura Anterior", "dataLeituraAnterior"],
        )
        leitura_atual_raw = _obter_valor_por_alias(
            dados,
            ["fatDataLeituraAtual", "Data Leitura Atual", "dataLeituraAtual"],
        )
        data_vcto_raw = _obter_valor_por_alias(dados, ["fatDataVcto", "Data Vencimento", "dataVencimento"])
        carimbo = _obter_valor_por_alias(dados, ["fatCarimbo", "carimbo", "Carimbo"])
        data_referencia_raw = _obter_valor_por_alias(
            dados,
            ["fatDataReferencia", "Data Referencia", "dataReferencia"],
        )

        leitura_anterior = valor_excel_para_date(leitura_anterior_raw)
        leitura_atual = valor_excel_para_date(leitura_atual_raw)
        data_vcto = valor_excel_para_date(data_vcto_raw) if data_vcto_raw else None
        data_referencia = valor_excel_para_date(data_referencia_raw) if data_referencia_raw else None

        if leitura_anterior is None or leitura_atual is None:
            continue

        leitura_anterior, leitura_atual = normalizar_datas_leitura_por_referencia(
            data_referencia,
            leitura_anterior,
            leitura_atual,
        )

        # Quando o OCR trouxer a referencia explicitamente, ela eh mais confiavel
        # do que recalcular o mes a partir das datas de leitura.
        data_referencia_esperada = (
            primeiro_dia_mes(data_referencia)
            if data_referencia is not None
            else primeiro_dia_mes(leitura_atual)
        )

        registros.append({
            "linha_excel": linha,
            "instalacao": str(instalacao).strip(),
            "fatDataLeituraAnterior": leitura_anterior,
            "fatDataLeituraAtual": leitura_atual,
            "fatCarimbo": "" if carimbo is None else str(carimbo).strip(),
            "fatDataVcto": data_vcto,
            "dataReferenciaEsperada": data_referencia_esperada,
            "dados_completos": dados,
        })

    if not registros:
        raise Exception(f"Nenhuma linha válida encontrada a partir da linha {linha_inicio}.")

    log(f"Total de linhas válidas (a partir de {linha_inicio}): {len(registros)}")
    return registros


# =========================================================
# NAVEGAÇÃO
# =========================================================

def abrir_tela_instalacao(driver, wait):
    log("Tentando abrir a tela Instalacao via link JS...")

    js_click = f"""
    const links = Array.from(document.querySelectorAll('a'));
    const alvo = links.find(a =>
        (a.getAttribute('href') === '{LINK_HREF}') ||
        ((a.getAttribute('href') || '').includes('{LINK_HREF}')) ||
        ((a.textContent || '').trim() === '{LINK_TEXTO}')
    );
    if (alvo) {{
        alvo.scrollIntoView({{block:'center'}});
        alvo.click();
        return 'clicked';
    }}
    return 'not_found';
    """

    try:
        resultado = driver.execute_script(js_click)
        log(f"Resultado do clique JS em Instalacao: {resultado}")
        time.sleep(0.5)
    except Exception as e:
        warn(f"Falha no clique JS da aba Instalacao: {type(e).__name__}")

    try:
        if "cadastroTabFatura.php" in driver.page_source or 'id="instalacao"' in driver.page_source:
            log("Tela Instalacao aparentemente já aberta após clique JS.")
            return True
    except Exception:
        pass

    log("Tentando abrir a tela Instalacao via alteração do hash...")
    try:
        driver.execute_script("window.location.hash = arguments[0];", "bpg/gestao/fatura/cadastroTabFatura.php")
        time.sleep(0.5)

        if "cadastroTabFatura.php" in driver.page_source or 'id="instalacao"' in driver.page_source:
            log("Tela Instalacao aberta via hash.")
            return True
    except Exception as e:
        warn(f"Falha ao alterar hash: {type(e).__name__}")

    log("Tentando abrir a tela Instalacao via URL direta...")
    driver.get(TARGET_URL)
    try:
        WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.ID, "instalacao")))
    except Exception:
        time.sleep(1)

    if 'id="instalacao"' in driver.page_source or "btnInstalacao" in driver.page_source:
        log("Tela Instalacao aberta via URL direta.")
        return True

    raise Exception("Não foi possível abrir a tela 'Instalacao'.")


# =========================================================
# CAMPOS DA TELA
# =========================================================

def localizar_input_generico(driver, wait, campo_id_ou_name):
    seletores = [
        (By.ID, campo_id_ou_name),
        (By.NAME, campo_id_ou_name),
        (By.CSS_SELECTOR, f"input#{campo_id_ou_name}"),
        (By.CSS_SELECTOR, f"input[name='{campo_id_ou_name}']"),
        (By.CSS_SELECTOR, f"textarea#{campo_id_ou_name}"),
        (By.CSS_SELECTOR, f"textarea[name='{campo_id_ou_name}']"),
        (By.CSS_SELECTOR, f"select#{campo_id_ou_name}"),
        (By.CSS_SELECTOR, f"select[name='{campo_id_ou_name}']"),
    ]

    ultimo = None
    for by, sel in seletores:
        try:
            wait.until(EC.presence_of_element_located((by, sel)))
            elementos = driver.find_elements(by, sel)
            candidatos = []
            for el in elementos:
                try:
                    if el.is_displayed() and el.is_enabled():
                        candidatos.append(el)
                except Exception:
                    pass
            if candidatos:
                ultimo = candidatos[-1]
                log(f"Campo encontrado: {by} = {sel} (último visível)")
                return ultimo
            elif elementos:
                ultimo = elementos[-1]
        except TimeoutException:
            pass

    if ultimo is not None:
        return ultimo

    raise Exception(f"Não encontrei o campo '{campo_id_ou_name}'.")


def localizar_elemento_por_id_ou_name(driver, campo_id=None, campo_name=None):
    candidatos = []

    tentativas = []
    if campo_id:
        tentativas.extend([
            (By.ID, campo_id),
            (By.CSS_SELECTOR, f"#{campo_id}"),
            (By.CSS_SELECTOR, f"input#{campo_id}"),
            (By.CSS_SELECTOR, f"textarea#{campo_id}"),
            (By.CSS_SELECTOR, f"select#{campo_id}"),
        ])

    if campo_name:
        tentativas.extend([
            (By.NAME, campo_name),
            (By.CSS_SELECTOR, f"[name='{campo_name}']"),
            (By.CSS_SELECTOR, f"input[name='{campo_name}']"),
            (By.CSS_SELECTOR, f"textarea[name='{campo_name}']"),
            (By.CSS_SELECTOR, f"select[name='{campo_name}']"),
        ])

    for by, sel in tentativas:
        try:
            encontrados = driver.find_elements(by, sel)
            for el in encontrados:
                if el not in candidatos:
                    candidatos.append(el)
        except Exception:
            pass

    if not candidatos:
        return None

    visiveis_habilitados = []
    for el in candidatos:
        try:
            if el.is_displayed() and el.is_enabled():
                visiveis_habilitados.append(el)
        except Exception:
            pass

    if visiveis_habilitados:
        return visiveis_habilitados[-1]

    habilitados = []
    for el in candidatos:
        try:
            if el.is_enabled():
                habilitados.append(el)
        except Exception:
            pass

    if habilitados:
        return habilitados[-1]

    return candidatos[-1]


def preencher_input_texto(driver, wait, campo_id_ou_name, valor, pausa_antes=0.0):
    campo = localizar_input_generico(driver, wait, campo_id_ou_name)

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
    time.sleep(0.15)

    if pausa_antes > 0:
        time.sleep(pausa_antes)

    try:
        campo.click()
        time.sleep(0.1)
    except Exception:
        pass

    try:
        campo.send_keys(Keys.CONTROL, "a")
        campo.send_keys(Keys.DELETE)
    except Exception:
        pass

    try:
        driver.execute_script("arguments[0].value = '';", campo)
    except Exception:
        pass

    time.sleep(0.2)

    try:
        campo.send_keys(str(valor))
    except Exception:
        driver.execute_script("arguments[0].value = arguments[1];", campo, str(valor))

    driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", campo)
    driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", campo)
    driver.execute_script("arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));", campo)

    try:
        valor_final = campo.get_attribute("value")
    except Exception:
        valor_final = "<não lido>"

    log(f"Campo {campo_id_ou_name} preenchido com {valor!r} | valor final na tela = {valor_final!r}")


def preencher_input_date(driver, wait, campo_id_ou_name, valor_date):
    campo = localizar_input_generico(driver, wait, campo_id_ou_name)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
    time.sleep(0.1)

    valor_iso = formatar_yyyy_mm_dd(valor_date)
    driver.execute_script("arguments[0].value = arguments[1];", campo, valor_iso)
    driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", campo)
    driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", campo)

    log(f"Campo {campo_id_ou_name} preenchido com {valor_iso}")


def preencher_elemento_html(driver, elemento, valor):
    tag = (elemento.tag_name or "").lower()
    tipo = (elemento.get_attribute("type") or "").lower()

    if tag == "select":
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elemento)
        time.sleep(0.05)
        valor_str = str(valor).strip()

        try:
            select = Select(elemento)

            try:
                select.select_by_visible_text(valor_str)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", elemento)
                log(f"[SELECT] '{valor_str}' selecionado por texto visível")
                return True
            except Exception:
                pass

            try:
                select.select_by_value(valor_str)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", elemento)
                log(f"[SELECT] '{valor_str}' selecionado por value")
                return True
            except Exception:
                pass

            for opt in select.options:
                opt_text = (opt.text or "").strip()
                if valor_str in opt_text or opt_text in valor_str:
                    select.select_by_visible_text(opt_text)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", elemento)
                    log(f"[SELECT] '{valor_str}' selecionado por match parcial em '{opt_text}'")
                    return True
        except Exception:
            pass

        ok = driver.execute_script("""
            const el = arguments[0];
            const v = arguments[1];
            let achou = false;
            for (const opt of el.options) {
                if ((opt.value || '') === v || (opt.textContent || '').trim() === v) {
                    el.value = opt.value;
                    el.dispatchEvent(new Event('input', {bubbles:true}));
                    el.dispatchEvent(new Event('change', {bubbles:true}));
                    achou = true;
                    break;
                }
            }
            return achou;
        """, elemento, valor_str)

        if not ok:
            try:
                opcoes = driver.execute_script("""
                    return Array.from(arguments[0].options).map(o => ({
                        value: o.value, text: o.textContent.trim()
                    }));
                """, elemento)
                warn(f"[SELECT] value='{valor_str}' não encontrado. Opções disponíveis: {opcoes}")
            except Exception:
                pass

        return bool(ok)

    # Inputs/textareas — sem scroll nem sleeps fixos, mantém send_keys para respeitar máscara
    valor_str = "" if valor is None else str(valor).strip()

    try:
        elemento.click()
    except Exception:
        pass

    try:
        elemento.send_keys(Keys.CONTROL, "a")
        elemento.send_keys(Keys.DELETE)
    except Exception:
        pass

    try:
        driver.execute_script("arguments[0].value = '';", elemento)
    except Exception:
        pass

    try:
        elemento.send_keys(valor_str)
    except Exception:
        pass

    try:
        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));
        """, elemento, valor_str)
    except Exception:
        return False

    try:
        valor_final = (elemento.get_attribute("value") or "").strip()
    except Exception:
        valor_final = ""

    log(
        f"[INPUT] id={elemento.get_attribute('id')} name={elemento.get_attribute('name')} | "
        f"valor enviado={valor_str!r} | valor final={valor_final!r}"
    )

    vf = valor_final.replace(".", "").replace(",", "")
    vs = valor_str.replace(".", "").replace(",", "")
    return valor_final == valor_str or vf == vs


def preencher_input_mascarado_por_id(driver, campo_id, valor):
    elementos = driver.find_elements(By.ID, campo_id)
    if not elementos:
        raise Exception(f"Campo não encontrado: {campo_id}")

    alvo = None
    for el in elementos:
        try:
            if el.is_displayed() and el.is_enabled():
                alvo = el
        except Exception:
            pass

    if alvo is None:
        alvo = elementos[-1]

    valor_str = "" if valor is None else str(valor).strip()

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    time.sleep(0.2)

    try:
        alvo.click()
        time.sleep(0.1)
    except Exception:
        pass

    driver.execute_script("""
        const el = arguments[0];
        const val = arguments[1];

        const nativeSetter =
            Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;

        nativeSetter.call(el, '');
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));

        nativeSetter.call(el, val);
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        el.dispatchEvent(new Event('blur', { bubbles: true }));
    """, alvo, valor_str)

    time.sleep(0.2)
    valor_final = (alvo.get_attribute("value") or "").strip()

    log(f"[CAMPO MASCARADO] id={campo_id} | valor enviado={valor_str!r} | valor final={valor_final!r}")
    return valor_final


def clicar_botao(driver, wait, seletores, descricao):
    botao = None

    for by, sel in seletores:
        try:
            encontrados = driver.find_elements(by, sel)
            candidatos = []
            for el in encontrados:
                try:
                    if el.is_displayed() and el.is_enabled():
                        candidatos.append(el)
                except Exception:
                    pass
            if candidatos:
                botao = candidatos[-1]
                log(f"Botão {descricao} encontrado: {by} = {sel}")
                break
            if encontrados:
                botao = encontrados[-1]
        except Exception:
            pass

    if botao is None:
        raise Exception(f"Não encontrei o botão '{descricao}'.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", botao)
    time.sleep(0.2)

    try:
        botao.click()
        log(f"Botão {descricao} clicado com clique normal.")
    except Exception:
        driver.execute_script("arguments[0].click();", botao)
        log(f"Botão {descricao} clicado via JavaScript.")


def clicar_botao_carregar_instalacao(driver, wait):
    seletores = [
        (By.ID, "btnInstalacao"),
        (By.NAME, "btnInstalacao"),
        (By.CSS_SELECTOR, "button#btnInstalacao"),
        (By.CSS_SELECTOR, "button[name='btnInstalacao']"),
        (By.XPATH, "//button[@id='btnInstalacao']"),
        (By.XPATH, "//button[contains(normalize-space(), 'Carregar')]"),
    ]
    clicar_botao(driver, wait, seletores, "Carregar Instalação")


def clicar_botao_carregar_leitura(driver, wait):
    seletores = [
        (By.ID, "botaoLeitura"),
        (By.NAME, "botaoLeitura"),
        (By.CSS_SELECTOR, "button#botaoLeitura"),
        (By.CSS_SELECTOR, "button[name='botaoLeitura']"),
        (By.XPATH, "//button[@id='botaoLeitura']"),
    ]
    clicar_botao(driver, wait, seletores, "Carregar Leitura")


def aguardar_carregamento_tabela(driver, wait):
    try:
        wait.until(EC.presence_of_element_located((By.ID, "datatable_tabletools")))
        wait.until(lambda d: len(d.find_elements(By.CSS_SELECTOR, "#datatable_tabletools tbody tr")) >= 1)
        log("Tabela carregada.")
    except Exception:
        warn("A tabela não confirmou carregamento completo, seguindo com o HTML atual.")
    time.sleep(1.2)


def scrollar_para_baixo(driver):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(0.3)
    log("Scroll executado até o final da página.")


def aguardar_campos_finais(driver):
    deadline = time.time() + 4.0
    total = 0
    while time.time() < deadline:
        total = len(driver.find_elements(By.CSS_SELECTOR, "input, select, textarea"))
        if total >= 5:
            break
        time.sleep(0.2)
    log(f"Campos visíveis no DOM: {total}")


def _aguardar_sem_spinner(driver, timeout: float = 6.0, min_wait: float = 0.3):
    time.sleep(min_wait)
    deadline = time.time() + timeout

    spinner_sels = [
        ".loading", ".spinner", ".overlay", "#loading",
        "[class*='loading']", "[class*='spinner']", "[id*='loading']",
        ".blockUI", ".blockOverlay", ".wait",
    ]
    spinner_css = ", ".join(spinner_sels)

    while time.time() < deadline:
        try:
            spinners = driver.find_elements(By.CSS_SELECTOR, spinner_css)
            visivel = any(
                s.is_displayed() for s in spinners
                if s.size.get("height", 0) > 0
            )
            if visivel:
                time.sleep(0.15)
                continue
        except Exception:
            pass

        try:
            state = driver.execute_script("return document.readyState;")
            if state != "complete":
                time.sleep(0.15)
                continue
        except Exception:
            pass

        try:
            jq_active = driver.execute_script(
                "return (typeof jQuery !== 'undefined') ? jQuery.active : 0;")
            if jq_active and int(jq_active) > 0:
                time.sleep(0.15)
                continue
        except Exception:
            pass

        return


def aguardar_tela_instalacao_pronta(driver, wait, timeout=8, pausa_extra=1.2):
    log("Aguardando tela de instalação ficar pronta...")
    campo = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.ID, "instalacao"))
    )

    time.sleep(pausa_extra)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
    time.sleep(0.2)

    try:
        if campo.get_attribute("disabled") or campo.get_attribute("readonly"):
            warn("Campo instalação ainda estava disabled/readonly. Aguardando mais um pouco...")
            time.sleep(1.0)
    except Exception:
        pass

    log("Tela de instalação pronta para digitação.")
    return campo


def campo_carimbo_existe(driver):
    try:
        elementos = driver.find_elements(By.ID, "carimbo")
        for el in elementos:
            try:
                if el.is_displayed():
                    return True
            except Exception:
                return True
    except Exception:
        pass

    try:
        elementos = driver.find_elements(By.NAME, "carimbo")
        for el in elementos:
            try:
                if el.is_displayed():
                    return True
            except Exception:
                return True
    except Exception:
        pass

    return False


def aguardar_campo_carimbo(driver, timeout=6.0, intervalo=0.2):
    fim = time.time() + timeout
    while time.time() < fim:
        if campo_carimbo_existe(driver):
            return True
        time.sleep(intervalo)
    return False


# =========================================================
# TABELA
# =========================================================

def localizar_tabela_faturas(driver, wait):
    seletores = [
        (By.ID, "datatable_tabletools"),
        (By.CSS_SELECTOR, "table#datatable_tabletools"),
    ]

    for by, sel in seletores:
        try:
            tabela = wait.until(EC.presence_of_element_located((by, sel)))
            log(f"Tabela encontrada: {by} = {sel}")
            return tabela
        except TimeoutException:
            pass

    raise Exception("Não encontrei a tabela de faturas.")


def obter_datas_referencia_tabela(driver, wait):
    tabela = localizar_tabela_faturas(driver, wait)
    linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody tr")
    log(f"Linhas encontradas na tabela: {len(linhas)}")

    datas = []

    for idx, linha in enumerate(linhas, start=1):
        colunas = linha.find_elements(By.TAG_NAME, "td")
        if not colunas:
            continue

        texto_data_ref = colunas[0].text.strip()
        if not texto_data_ref:
            continue

        try:
            data_ref = parse_data_ddmmyyyy(texto_data_ref)
            datas.append(data_ref)
            log(f"Linha {idx} | Data Referência = {texto_data_ref}")
        except Exception:
            warn(f"Não consegui interpretar a Data Referência da linha {idx}: {texto_data_ref!r}")

    return datas


def obter_ultima_data_referencia_tabela(driver, wait):
    datas = obter_datas_referencia_tabela(driver, wait)

    if not datas:
        warn("Tabela sem datas de referência. Considerando sem registro anterior.")
        return None

    ultima = max(datas)
    log(f"Última Data Referência da tabela: {formatar_ddmmyyyy(ultima)}")
    return ultima


def preencher_datas_e_carimbo_se_necessario(driver, wait, registro_planilha):
    ultima_data_tabela = obter_ultima_data_referencia_tabela(driver, wait)
    data_esperada = registro_planilha["dataReferenciaEsperada"]

    log(f"Data Referência esperada da planilha: {formatar_ddmmyyyy(data_esperada)}")
    if ultima_data_tabela:
        log(f"Última Data Referência encontrada na tabela: {formatar_ddmmyyyy(ultima_data_tabela)}")
    else:
        log("Nenhuma Data Referência encontrada na tabela.")

    # Sempre tenta datas; a decisão de pular será pela ausência do carimbo
    log("Preenchendo datas de leitura...")
    preencher_input_date(driver, wait, "dataLeituraAtual", registro_planilha["fatDataLeituraAtual"])
    preencher_input_date(driver, wait, "dataLeituraAnterior", registro_planilha["fatDataLeituraAnterior"])

    log("Clicando em carregar leitura após preencher datas...")
    clicar_botao_carregar_leitura(driver, wait)
    _aguardar_sem_spinner(driver, timeout=6, min_wait=0.8)
    time.sleep(0.8)

    apareceu_carimbo = aguardar_campo_carimbo(driver, timeout=6.0, intervalo=0.2)

    if not apareceu_carimbo:
        log("Campo 'carimbo' NÃO apareceu após carregar as datas. Referência já existente. Pulando instalação.")
        return False

    log("Campo 'carimbo' apareceu. Data aceita. Seguindo para preencher carimbo...")

    valor_carimbo = registro_planilha.get("fatCarimbo", "")
    if valor_carimbo is None:
        valor_carimbo = ""
    valor_carimbo = str(valor_carimbo).strip()

    if valor_carimbo:
        preencher_input_texto(driver, wait, "carimbo", valor_carimbo, pausa_antes=0.4)

        log("Clicando em carregar leitura após preencher carimbo...")
        clicar_botao_carregar_leitura(driver, wait)
        _aguardar_sem_spinner(driver, timeout=6, min_wait=0.8)
        time.sleep(0.8)
    else:
        warn("fatCarimbo veio vazio na planilha. Seguindo sem preencher carimbo.")

    scrollar_para_baixo(driver)
    aguardar_campos_finais(driver)
    return True


# =========================================================
# BS4 - EXPORTAR CAMPOS DISPONÍVEIS
# =========================================================

def extrair_campos_editaveis_bs4(driver):
    soup = BeautifulSoup(driver.page_source, "html.parser")
    campos = []

    for tag in soup.find_all("input"):
        tipo = (tag.get("type") or "text").strip().lower()

        if tipo in {"hidden", "submit", "button", "reset", "image", "file", "checkbox", "radio"}:
            continue

        if tag.has_attr("disabled") or tag.has_attr("readonly"):
            continue

        campo = {
            "tag": "input",
            "type": tipo,
            "name": tag.get("name"),
            "id": tag.get("id"),
            "value": tag.get("value"),
            "placeholder": tag.get("placeholder"),
            "class": " ".join(tag.get("class", [])) if tag.get("class") else "",
            "maxlength": tag.get("maxlength"),
            "required": tag.has_attr("required"),
        }
        campos.append(campo)

    for tag in soup.find_all("select"):
        if tag.has_attr("disabled") or tag.has_attr("readonly"):
            continue

        campo = {
            "tag": "select",
            "type": "select",
            "name": tag.get("name"),
            "id": tag.get("id"),
            "value": None,
            "placeholder": None,
            "class": " ".join(tag.get("class", [])) if tag.get("class") else "",
            "maxlength": None,
            "required": tag.has_attr("required"),
            "options": [
                {
                    "value": opt.get("value"),
                    "text": opt.get_text(" ", strip=True)
                }
                for opt in tag.find_all("option")
            ],
        }
        campos.append(campo)

    for tag in soup.find_all("textarea"):
        if tag.has_attr("disabled") or tag.has_attr("readonly"):
            continue

        campo = {
            "tag": "textarea",
            "type": "textarea",
            "name": tag.get("name"),
            "id": tag.get("id"),
            "value": tag.get_text("", strip=True),
            "placeholder": tag.get("placeholder"),
            "class": " ".join(tag.get("class", [])) if tag.get("class") else "",
            "maxlength": tag.get("maxlength"),
            "required": tag.has_attr("required"),
        }
        campos.append(campo)

    return campos


# =========================================================
# CORRELAÇÃO CAMPOS HTML x HEADERS PLANILHA
# =========================================================

def pontuar_correlacao(nome_campo, header_planilha):
    nome = normalizar_slug(nome_campo)
    header = normalizar_slug(header_planilha)

    if not nome or not header:
        return 0

    score = 0

    if nome == header:
        score += 100

    if nome in header or header in nome:
        score += 40

    tokens_nome = tokenizar(nome)
    tokens_header = tokenizar(header)
    inter = tokens_nome.intersection(tokens_header)
    score += len(inter) * 8

    aliases = [
        (["instalacao"], "Instalação"),
        (["carimbo"], "fatCarimbo"),
        (["tarifa"], "cadTarifaCod"),
        (["subgrupo"], "cadSubGrupoCod"),
        (["data", "leitura", "atual"], "fatDataLeituraAtual"),
        (["data", "leitura", "anterior"], "fatDataLeituraAnterior"),
        (["data", "vencimento"], "fatDataVcto"),
        (["data", "vcto"], "fatDataVcto"),
        (["data", "emissao"], "fatDataEmissao"),
        (["valor", "fatura"], "fatValorFatura"),
        (["valor", "nota", "fiscal"], "fatValorNotaFiscal"),
        (["ilum"], "fatIlumPublica"),
        (["icms"], "fatICMS"),
        (["pis"], "fatPIS"),
        (["cofins"], "fatCOFINS"),
        (["dem", "contratada", "ponta"], "fatDemContratadaPonta"),
        (["dem", "contratada", "fp"], "fatDemContratadaFPonta"),
        (["dem", "contratada", "geracao", "ponta"], "fatDemContratadaGeracaoPonta"),
        (["dem", "cont", "geracao", "ponta"], "fatDemContratadaGeracaoPonta"),
        (["dem", "contratada", "geracao", "fp"], "fatDemContratadaGeracaoFPonta"),
        (["dem", "cont", "geracao", "fp"], "fatDemContratadaGeracaoFPonta"),
        (["dem", "ponta", "registrada"], "fatDemPontaRegistrada"),
        (["dem", "fpind", "registrada"], "fatDemFPontaIndRegistrada"),
        (["dem", "fp", "registrada"], "fatDemFPontaIndRegistrada"),
        (["dem", "ponta", "faturada"], "fatDemPontaFaturada"),
        (["dem", "fpind", "faturada"], "fatDemFPontaIndFaturada"),
        (["dem", "fp", "faturada"], "fatDemFPontaIndFaturada"),
        (["consumo", "registrada", "ponta"], "fatConPontaRegistrado"),
        (["consumo", "registrada", "fpind"], "fatConFPontaIndRegistrado"),
        (["consumo", "registrada", "fp"], "fatConFPontaIndRegistrado"),
        (["consumo", "registrada", "intermediario"], "fatConIntermediarioRegistrado"),
        (["consumo", "faturada", "ponta"], "fatConPontaFaturado"),
        (["consumo", "faturada", "fpind"], "fatConFPontaIndFaturado"),
        (["consumo", "faturada", "fp"], "fatConFPontaIndFaturado"),
        (["consumo", "faturada", "intermediario"], "fatConIntermediarioFaturado"),
        (["rs", "kwh"], "fatRsKWh"),
    ]

    for palavras, header_alvo in aliases:
        if all(p in nome for p in palavras):
            if normalizar_slug(header_planilha) == normalizar_slug(header_alvo):
                score += 70

    return score


def correlacionar_campos_com_headers(campos, headers_planilha):
    correlacoes = []

    for campo in campos:
        nome_base = campo.get("name") or campo.get("id") or ""
        melhor_header = None
        melhor_score = 0

        for header in headers_planilha:
            score = pontuar_correlacao(nome_base, header)
            if score > melhor_score:
                melhor_score = score
                melhor_header = header

        correlacoes.append({
            "campo_html": nome_base,
            "id": campo.get("id"),
            "name": campo.get("name"),
            "tag": campo.get("tag"),
            "type": campo.get("type"),
            "valor_atual_tela": campo.get("value"),
            "header_planilha_sugerido": melhor_header if melhor_score > 0 else "",
            "score_correlacao": melhor_score,
        })

    return correlacoes


# =========================================================
# SALVAR EXPORTS
# =========================================================

def salvar_campos_disponiveis(campos):
    caminho_json = PASTA_SAIDA / "campos_disponiveis.json"
    caminho_csv = PASTA_SAIDA / "campos_disponiveis.csv"

    caminho_json.write_text(json.dumps(campos, ensure_ascii=False, indent=2), encoding="utf-8")

    with caminho_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "tag", "type", "name", "id", "value", "placeholder",
            "class", "maxlength", "required"
        ])

        for c in campos:
            writer.writerow([
                c.get("tag"), c.get("type"), c.get("name"), c.get("id"), c.get("value"),
                c.get("placeholder"), c.get("class"), c.get("maxlength"), c.get("required"),
            ])

    log(f"Campos disponíveis salvos em: {caminho_json}")
    log(f"Campos disponíveis salvos em: {caminho_csv}")


def salvar_correlacoes(correlacoes):
    caminho = PASTA_SAIDA / "mapeamento_campos_planilha.csv"

    with caminho.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "campo_html", "id", "name", "tag", "type",
            "valor_atual_tela", "header_planilha_sugerido", "score_correlacao"
        ])

        for item in correlacoes:
            writer.writerow([
                item.get("campo_html"),
                item.get("id"),
                item.get("name"),
                item.get("tag"),
                item.get("type"),
                item.get("valor_atual_tela"),
                item.get("header_planilha_sugerido"),
                item.get("score_correlacao"),
            ])

    log(f"Mapeamento sugerido salvo em: {caminho}")


def exportar_campos_e_correlacoes(driver, excel_path):
    headers = obter_headers_planilha(excel_path)
    campos = extrair_campos_editaveis_bs4(driver)
    correlacoes = correlacionar_campos_com_headers(campos, headers)

    salvar_campos_disponiveis(campos)
    salvar_correlacoes(correlacoes)

    return campos, correlacoes


# =========================================================
# CONSUMIR CSV DE MAPEAMENTO E PREENCHER
# =========================================================

def carregar_mapeamento_csv(caminho_csv):
    caminho_csv = Path(caminho_csv)
    if not caminho_csv.exists():
        raise Exception(f"CSV de mapeamento não encontrado: {caminho_csv}")

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with caminho_csv.open("r", newline="", encoding=enc) as f:
                return list(csv.DictReader(f, delimiter=";"))
        except UnicodeDecodeError:
            continue

    raise Exception(f"Não consegui ler o CSV de mapeamento: {caminho_csv}")


def escolher_csv_mapeamento():
    if MAPEAMENTO_CSV.exists():
        log(f"Usando CSV de mapeamento externo: {MAPEAMENTO_CSV}")
        return MAPEAMENTO_CSV

    interno = PASTA_SAIDA / "mapeamento_campos_planilha.csv"
    if interno.exists():
        log(f"Usando CSV de mapeamento gerado pelo script: {interno}")
        return interno

    raise Exception("Nenhum CSV de mapeamento encontrado.")


def formatar_valor_para_campo(header, valor, tipo_html):
    if valor is None:
        return ""

    if isinstance(valor, datetime):
        valor = valor.date()

    if isinstance(valor, date):
        if str(tipo_html).lower() == "date":
            return valor.strftime("%Y-%m-%d")
        return valor.strftime("%d/%m/%Y")

    header_norm = normalizar_texto(header)
    tipo_norm = str(tipo_html).lower()

    if "cod" in header_norm:
        return str(valor).strip()

    if "data" in header_norm and tipo_norm == "date":
        try:
            return formatar_yyyy_mm_dd(valor_excel_para_date(valor))
        except Exception:
            return str(valor).strip()

    if any(chave in header_norm for chave in [
        "instalacao", "carimbo", "cnpj", "faturaid", "referencia",
        "codigobarras", "codigobarra"
    ]):
        return str(valor).strip()

    num = valor_para_numero(valor)
    if num is not None:
        # Proteção extra para fatValorNotaFiscal: o Consen tem máscara que
        # pode interpretar errado se o valor vier com separador de milhar.
        # Garantimos que o número tem no máximo 2 casas decimais e sem milhar.
        if "valornotafiscal" in header_norm or "valorfatura" in header_norm:
            # Faturas CEMIG MT (média tensão) chegam a R$ 65.000+ legítimos.
            # Só rejeita valores acima de R$ 500.000 que seriam erro de parsing
            # (ex: "112,56" lido pelo Excel como 11256000,00).
            if num > 500000:
                warn(f"[NOTA FISCAL] Valor suspeito ({num:.2f}) para '{header}' — campo ignorado para evitar erro de máscara")
                return ""
        if tipo_norm in ("number", "range"):
            return str(num)
        return formatar_numero_br(num)

    return str(valor).strip()


def _resolver_valor_select(campo_html: str, valor_bruto) -> str:
    chave = str(valor_bruto).strip()

    if campo_html in MAP_SELECTS:
        mapa = MAP_SELECTS[campo_html]
        if chave in mapa:
            return mapa[chave]
        warn(f"[MAP_SELECTS] Texto '{chave}' não está no mapa de '{campo_html}'. "
             f"Opções conhecidas: {list(mapa.keys())}. Tentando usar o texto diretamente no select.")
    return chave


def _normalizar_select_mt_cemig_consen(campo_html: str, valor_formatado: str) -> str:
    """
    Os extratores MT da CEMIG permanecem separados no OCR, mas no Consen
    ambos devem convergir para a mesma tarifa final, preservando o subgrupo
    real vindo do OCR.
    """
    if campo_html == "cb-dados-contratuais-fatura-tarifa" and valor_formatado in (
        "THS Verde", "TUSD Livre Verde", "HS - Verde"
    ):
        return "HS - Verde"

    return valor_formatado


def preencher_aliases_forcados_linha(driver, dados_planilha):
    preenchidos = []
    pulados = []

    for campo_html, header_planilha in ALIASES_FORCADOS.items():
        headers_alternativos = [header_planilha]
        headers_alternativos.extend(ALIASES_FORCADOS_HEADERS_ALTERNATIVOS.get(campo_html, []))

        # Fallbacks
        if header_planilha == "fatConFPontaInjetadoUsinaSaldoAcumulado":
            headers_alternativos.append("fatConFPontaInjetadoUsinaSaldo")
        if header_planilha == "fatConPontaInjetadoUsinaSaldoAcumulado":
            headers_alternativos.append("fatConPontaInjetadoUsinaSaldo")

        valor_bruto, header_usado = obter_valor_planilha_por_headers(dados_planilha, *headers_alternativos)

        if header_usado is None:
            pulados.append((campo_html, f"header/valor ausente: {headers_alternativos}"))
            continue

        elemento = localizar_elemento_por_id_ou_name(driver, campo_id=campo_html, campo_name=campo_html)
        if elemento is None:
            pulados.append((campo_html, "elemento não encontrado na tela"))
            continue

        tag = (elemento.tag_name or "").lower()
        tipo = (elemento.get_attribute("type") or tag or "").lower()

        if tag == "select":
            valor_formatado = _resolver_valor_select(campo_html, valor_bruto)
            valor_formatado = _normalizar_select_mt_cemig_consen(campo_html, valor_formatado)
        else:
            valor_formatado = formatar_valor_para_campo(header_usado, valor_bruto, tipo)

        log(f"[ALIAS] campo={campo_html} | header={header_usado} | valor_bruto={valor_bruto!r} | valor_formatado={valor_formatado!r}")

        try:
            ok = preencher_elemento_html(driver, elemento, valor_formatado)
            if ok:
                preenchidos.append((campo_html, header_usado, valor_formatado))
                # Após mudar a tarifa (MT = THS Verde / TUSD Livre Verde), o Consen
                # dispara um AJAX que recarrega os campos de demanda/consumo.
                # É obrigatório aguardar o AJAX antes de preencher os campos seguintes,
                # caso contrário os valores são apagados pelo reload da tela.
                if campo_html == "cb-dados-contratuais-fatura-tarifa" and valor_formatado == "HS - Verde":
                    log("[ALIAS] Tarifa MT detectada — aguardando estabilização AJAX (timeout=15s)...")
                    _aguardar_sem_spinner(driver, timeout=15, min_wait=1.5)
                    log("[ALIAS] AJAX estabilizado. Prosseguindo com aliases de demanda/consumo.")
            else:
                pulados.append((campo_html, f"falha ao preencher (value='{valor_formatado}')"))
        except Exception as e:
            pulados.append((campo_html, f"erro: {type(e).__name__}"))

    print("=" * 100)
    print("ALIASES FORÇADOS")
    print("=" * 100)
    for campo, header, valor in preenchidos:
        print(f"{campo} <- {header} | valor={valor}")

    return preenchidos, pulados


def preencher_campos_via_mapeamento_csv_linha(driver, dados_planilha, caminho_csv, score_minimo=35):
    mapeamentos = carregar_mapeamento_csv(caminho_csv)
    preenchidos = []
    pulados = []

    campos_pulados_fluxo = {
        "instalacao", "dataLeituraAtual", "dataLeituraAnterior",
        "carimbo", "dataVencimento", "dataEmissao",
        "fatCarimbo",
    }
    campos_pulados_alias = set(ALIASES_FORCADOS.keys()) | CAMPOS_BLOQUEADOS_CSV

    for item in mapeamentos:
        campo_id = (item.get("id") or "").strip()
        campo_name = (item.get("name") or "").strip()
        tag = (item.get("tag") or "").strip().lower()
        tipo = (item.get("type") or "").strip().lower()
        header = (item.get("header_planilha_sugerido") or "").strip()

        try:
            score = int(float(item.get("score_correlacao") or 0))
        except Exception:
            score = 0

        identificador = campo_id or campo_name or item.get("campo_html") or "<sem_id>"

        if not header:
            pulados.append((identificador, "sem header sugerido"))
            continue
        if score < score_minimo:
            pulados.append((identificador, f"score baixo ({score})"))
            continue
        if campo_id in campos_pulados_fluxo or campo_name in campos_pulados_fluxo:
            pulados.append((identificador, "campo já tratado pelo fluxo principal"))
            continue
        if campo_id in campos_pulados_alias or campo_name in campos_pulados_alias:
            pulados.append((identificador, "campo já tratado por alias forçado"))
            continue
        if header in HEADERS_BLOQUEADOS_PLANILHA:
            pulados.append((identificador, f"header bloqueado explicitamente: {header}"))
            continue
        if header not in dados_planilha:
            pulados.append((identificador, f"header não existe na planilha: {header}"))
            continue

        valor_bruto = dados_planilha[header]
        if valor_bruto is None or str(valor_bruto).strip() == "":
            pulados.append((identificador, f"valor vazio para {header}"))
            continue

        elemento = localizar_elemento_por_id_ou_name(driver, campo_id=campo_id, campo_name=campo_name)
        if elemento is None:
            pulados.append((identificador, "elemento não encontrado na tela"))
            continue

        valor_formatado = formatar_valor_para_campo(header, valor_bruto, tipo)
        try:
            ok = preencher_elemento_html(driver, elemento, valor_formatado)
            if ok:
                preenchidos.append((identificador, header, valor_formatado, score, tag, tipo))
            else:
                pulados.append((identificador, "falha ao preencher"))
        except Exception as e:
            pulados.append((identificador, f"erro: {type(e).__name__}"))

    caminho_log = PASTA_SAIDA / "resultado_preenchimento_csv.csv"
    with caminho_log.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["status", "campo", "header", "valor", "score_ou_motivo"])
        for campo, header, valor, score, _, _ in preenchidos:
            writer.writerow(["preenchido", campo, header, valor, score])
        for campo, motivo in pulados:
            writer.writerow(["pulado", campo, "", "", motivo])

    log(f"Resultado do preenchimento salvo em: {caminho_log}")
    return preenchidos, pulados


def preencher_saldos_injetado_usina(driver, wait, dados_planilha):
    mapa = {
        "txt-consumo-injetado-usina-pta-saldo": "fatConPontaInjetadoUsinaSaldoAcumulado",
        "txt-consumo-injetado-usina-fpta-saldo": "fatConFPontaInjetadoUsinaSaldoAcumulado",
    }

    for campo_consen, header_planilha in mapa.items():
        valor_bruto, header_usado = obter_valor_planilha_por_headers(
            dados_planilha,
            header_planilha,
            header_planilha.replace("Acumulado", "")
        )

        if header_usado is None:
            warn(f"[SALDO USINA] Valor ausente na planilha para {campo_consen}")
            continue

        ok = preencher_saldo_usina_tabela(driver, wait, campo_consen, valor_bruto)

        if ok:
            log(f"[SALDO USINA] OK | {campo_consen} <- {header_usado} | valor={valor_bruto!r}")
        else:
            warn(f"[SALDO USINA] FALHOU | {campo_consen} <- {header_usado} | valor={valor_bruto!r}")


# =========================================================
# OBSERVAÇÕES / BOTÕES FINAIS
# =========================================================

def _normalizar_valor_obs(cod: str, val) -> float:
    num = valor_para_numero(val)
    if num is None:
        num = 0.0

    codigos_negativos = {"8", "11", "23", "58", "59", "109", "149"}
    if cod in codigos_negativos:
        return -abs(float(num))
    return round(float(num), 2)


def _coletar_pares_obs(dados_planilha: dict) -> list[tuple[str, float]]:
    agregados: dict[str, float] = {}
    ordem: list[str] = []

    for i in range(1, _OBS_PARES_MAX + 1):
        cod = str(dados_planilha.get(f"obsCod_{i}") or "").strip()
        if not cod or cod == "0":
            continue

        valor = _normalizar_valor_obs(cod, dados_planilha.get(f"obsValor_{i}"))
        if cod not in agregados:
            agregados[cod] = 0.0
            ordem.append(cod)
        agregados[cod] = round(agregados[cod] + valor, 2)

    # Fallback defensivo: alguns fluxos carregam a restituição em campo auxiliar
    # mesmo quando os pares obsCod_/obsValor_ não vieram materializados.
    if "109" not in agregados:
        restit = dados_planilha.get("_restituicaoPagamento")
        valor_restit = _normalizar_valor_obs("109", restit)
        if abs(valor_restit) > 0.004:
            agregados["109"] = valor_restit
            ordem.append("109")

    if not any(abs(agregados.get(cod, 0.0)) > 0.004 for cod in ("58", "149", "11")):
        fat_dic = _normalizar_valor_obs("58", dados_planilha.get("fatDIC"))
        fat_fic = _normalizar_valor_obs("11", dados_planilha.get("fatFIC"))
        if abs(fat_dic) > 0.004:
            agregados["58"] = fat_dic
            ordem.append("58")
        elif abs(fat_fic) > 0.004:
            agregados["11"] = fat_fic
            ordem.append("11")

    # A família DIC/FIC da CEMIG deve entrar em uma única observação.
    dic_codigos = ("58", "149", "11")
    presentes_dic = [cod for cod in dic_codigos if abs(agregados.get(cod, 0.0)) > 0.004]
    if len(presentes_dic) > 1:
        cod_escolhido = presentes_dic[0]
        valor_escolhido = agregados[cod_escolhido]
        for cod in dic_codigos:
            agregados.pop(cod, None)
            if cod in ordem:
                ordem.remove(cod)
        agregados[cod_escolhido] = valor_escolhido
        ordem.append(cod_escolhido)

    pares = []
    for cod in ordem:
        val = round(agregados[cod], 2)
        if abs(val) > 0.004:
            pares.append((cod, val))
    return pares


def preencher_obs_multiplas(driver, wait, dados_planilha: dict):
    pares = _coletar_pares_obs(dados_planilha)

    if not pares:
        log("[OBS] Nenhuma observação para preencher.")
        return

    for idx, (cod, val) in enumerate(pares, start=1):
        try:
            el_sel = wait.until(EC.presence_of_element_located((By.ID, "cb-dados-financeiros-obs")))
            ok = preencher_elemento_html(driver, el_sel, cod)
            if ok:
                log(f"[OBS] Linha {idx}: select cod={cod}")
            else:
                warn(f"[OBS] Linha {idx}: falha ao selecionar cod={cod}")
                continue
        except Exception as e:
            warn(f"[OBS] Linha {idx}: erro no select — {type(e).__name__}")
            continue

        try:
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));",
                el_sel)
        except Exception:
            pass

        try:
            driver.execute_script("document.body.click();")
        except Exception:
            pass
        time.sleep(0.4)

        val_fmt = formatar_numero_br(val) if val is not None else "0,00"
        try:
            el_val = wait.until(EC.element_to_be_clickable((By.ID, "txt-dados-financeiros-outros")))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el_val)
            time.sleep(0.2)
            driver.execute_script("arguments[0].value = '';", el_val)
            driver.execute_script("arguments[0].dispatchEvent(new Event('focus',{bubbles:true}));", el_val)
            el_val.click()
            el_val.send_keys(Keys.CONTROL, "a")
            el_val.send_keys(Keys.DELETE)
            driver.execute_script("arguments[0].value = '';", el_val)
            time.sleep(0.1)
            el_val.send_keys(val_fmt)
            driver.execute_script("""
                arguments[0].dispatchEvent(new Event('input',{bubbles:true}));
                arguments[0].dispatchEvent(new Event('change',{bubbles:true}));
            """, el_val)
            log(f"[OBS] Linha {idx}: valor={val_fmt}")
        except Exception as e:
            warn(f"[OBS] Linha {idx}: erro no campo valor — {type(e).__name__}")
            try:
                driver.execute_script("""
                    var el = document.getElementById('txt-dados-financeiros-outros');
                    if(el){
                        el.value = arguments[0];
                        el.dispatchEvent(new Event('input',{bubbles:true}));
                        el.dispatchEvent(new Event('change',{bubbles:true}));
                    }
                """, val_fmt)
                log(f"[OBS] Linha {idx}: valor via JS={val_fmt}")
            except Exception:
                pass

        time.sleep(0.3)

        try:
            btn = wait.until(EC.element_to_be_clickable((By.ID, "btnIncluiLinha")))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.2)
            btn.click()
            log(f"[OBS] Linha {idx}: btnIncluiLinha clicado")
        except Exception as e:
            warn(f"[OBS] Linha {idx}: erro ao clicar btnIncluiLinha — {type(e).__name__}")
            try:
                driver.execute_script("document.getElementById('btnIncluiLinha').click();")
                log(f"[OBS] Linha {idx}: btnIncluiLinha via JS")
            except Exception:
                pass

        time.sleep(0.6)

    log(f"[OBS] {len(pares)} observação(ões) incluída(s).")


def clicar_botao_salvar(driver, wait):
    seletores = [
        (By.ID, "btnSalvar"),
        (By.NAME, "btnSalvar"),
        (By.CSS_SELECTOR, "button#btnSalvar"),
        (By.CSS_SELECTOR, "button[name='btnSalvar']"),
        (By.XPATH, "//button[@id='btnSalvar']"),
        (By.XPATH, "//button[contains(normalize-space(), 'Salvar')]")
    ]
    clicar_botao(driver, wait, seletores, "Salvar")


def clicar_link_auditoria(driver, wait):
    seletores = [
        (By.ID, "linkAuditoria"),
        (By.CSS_SELECTOR, "a#linkAuditoria"),
        (By.XPATH, "//a[@id='linkAuditoria']"),
        (By.XPATH, "//a[contains(normalize-space(), 'Auditoria')]")
    ]

    link = None
    for by, sel in seletores:
        try:
            encontrados = driver.find_elements(by, sel)
            candidatos = []
            for el in encontrados:
                try:
                    if el.is_displayed() and el.is_enabled():
                        candidatos.append(el)
                except Exception:
                    pass
            if candidatos:
                link = candidatos[-1]
                log(f"Link Auditoria encontrado: {by} = {sel}")
                break
            if encontrados:
                link = encontrados[-1]
        except Exception:
            pass

    if link is None:
        raise Exception("Não encontrei o link 'Auditoria'.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
    time.sleep(0.5)

    try:
        link.click()
        log("Link Auditoria clicado com clique normal.")
    except Exception:
        driver.execute_script("arguments[0].click();", link)
        log("Link Auditoria clicado via JavaScript.")


def clicar_botao_proxima_fatura(driver, wait):
    seletores = [
        (By.ID, "btnProxima"),
        (By.NAME, "btnProxima"),
        (By.CSS_SELECTOR, "button#btnProxima"),
        (By.CSS_SELECTOR, "button[name='btnProxima']"),
        (By.XPATH, "//button[@id='btnProxima']"),
        (By.XPATH, "//button[contains(normalize-space(), 'Proxima Fatura')]")
    ]
    clicar_botao(driver, wait, seletores, "Proxima Fatura")


def registrar_resultado_auditoria(
    linha_excel, instalacao, data_referencia_esperada, carimbo,
    valor_auditoria, status,
    pct_diferenca="", itens_divergentes="", memoria_calculo="",
):
    caminho = AUDITORIA_CSV
    append_resultado_auditoria(caminho, {
        "linha_excel": linha_excel,
        "instalacao": instalacao,
        "data_referencia_esperada": data_referencia_esperada,
        "carimbo": carimbo,
        "valor_auditoria": valor_auditoria,
        "pct_diferenca": pct_diferenca,
        "itens_divergentes": itens_divergentes,
        "memoria_calculo": memoria_calculo,
        "status": status,
    })
    log(f"Resultado da auditoria registrado em: {caminho}")


def _extrair_status_auditoria(row: dict) -> str:
    """Lê o status mesmo quando o CSV ficou com colunas extras desalinhadas."""
    return extrair_status_auditoria(row)


def _capturar_detalhe_auditoria(driver) -> tuple:
    """
    Le a pagina de auditoria do Consen e retorna:
      (total_auditoria, pct_diferenca, itens_divergentes, memoria_calculo)
    - total_auditoria : 1o span.auditoria.sucesso  (ex: '-0,01')
    - pct_diferenca   : 2o span.auditoria.sucesso  (ex: '0,00%')
    - itens_divergentes: linhas da tabela onde Diferenca != 0,
                         formato: 'Consumo Fora Ponta Ind=R$-0,03|Total Fatura=R$-0,01'
    """
    total = ""
    pct   = ""
    itens = ""
    memoria = ""

    try:
        spans = driver.find_elements(By.CSS_SELECTOR, "span.auditoria.sucesso")
        if spans:
            total = spans[0].text.strip()
        if len(spans) >= 2:
            pct = spans[1].text.strip()
        log(f"Auditoria capturada: total={total!r}  pct={pct!r}")
    except Exception as _e:
        warn(f"Nao foi possivel capturar spans de auditoria: {_e}")

    try:
        tabela = driver.find_element(By.CSS_SELECTOR, "table.table-bordered")
        linhas = tabela.find_elements(By.TAG_NAME, "tr")
        divs = []
        memoria_rows = []
        for linha in linhas[1:]:
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if len(colunas) >= 7:
                valores = [c.text.strip() for c in colunas[:7]]
                descricao = valores[0]
                diferenca = valores[6]
                memoria_rows.append({
                    "descricao": descricao,
                    "volumes_fatura": valores[1],
                    "tarifa_calculo": valores[2],
                    "tarifa_calculo_com_imposto": valores[3],
                    "valores_fatura": valores[4],
                    "valores_calculados": valores[5],
                    "diferenca_auditoria": diferenca,
                })
                if diferenca and "0,00" not in diferenca:
                    divs.append(f"{descricao}={diferenca}")
        itens = "|".join(divs)
        memoria = json.dumps(memoria_rows, ensure_ascii=False)
        if itens:
            log(f"Itens divergentes: {itens}")
    except Exception as _e:
        warn(f"Nao foi possivel ler tabela de auditoria: {_e}")

    return total, pct, itens, memoria


def salvar_abrir_auditoria_capturar_fechar(driver, wait, registro_planilha):
    aba_principal = driver.current_window_handle
    abas_antes = driver.window_handles[:]

    log("Clicando em Salvar...")
    clicar_botao_salvar(driver, wait)
    _aguardar_sem_spinner(driver, timeout=10, min_wait=1.0)

    try:
        WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.ID, "linkAuditoria")))
    except Exception:
        time.sleep(1.5)

    log("Clicando em Auditoria...")
    clicar_link_auditoria(driver, wait)

    nova_aba = None
    for _ in range(20):
        abas_agora = driver.window_handles
        if len(abas_agora) > len(abas_antes):
            novas = [h for h in abas_agora if h not in abas_antes]
            if novas:
                nova_aba = novas[0]
                break
        time.sleep(0.2)

    if nova_aba:
        driver.switch_to.window(nova_aba)
        log("Mudou para a aba da auditoria.")
    else:
        log("Auditoria não abriu nova aba; seguindo na aba atual.")

    try:
        WebDriverWait(driver, 6).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.auditoria.sucesso")))
    except Exception:
        warn("Timeout aguardando span.auditoria.sucesso")

    valor_auditoria, pct_diferenca, itens_divergentes, memoria_calculo = _capturar_detalhe_auditoria(driver)

    registrar_resultado_auditoria(
        linha_excel=registro_planilha["linha_excel"],
        instalacao=registro_planilha["instalacao"],
        data_referencia_esperada=formatar_ddmmyyyy(registro_planilha["dataReferenciaEsperada"]),
        carimbo=registro_planilha.get("fatCarimbo", ""),
        valor_auditoria=valor_auditoria,
        pct_diferenca=pct_diferenca,
        itens_divergentes=itens_divergentes,
        memoria_calculo=memoria_calculo,
        status="sucesso_auditoria" if valor_auditoria else "auditoria_sem_valor",
    )

    if nova_aba:
        driver.close()
        driver.switch_to.window(aba_principal)
        log("Retornou à aba principal.")

    log("Clicando em Proxima Fatura...")
    clicar_botao_proxima_fatura(driver, wait)
    _aguardar_sem_spinner(driver, timeout=5, min_wait=0.4)

    return valor_auditoria


# =========================================================
# FLUXO
# =========================================================

def voltar_para_tela_inicial_instalacao(driver, wait):
    log("Voltando para a tela inicial de instalação...")
    abrir_tela_instalacao(driver, wait)
    _aguardar_sem_spinner(driver, timeout=5, min_wait=0.5)

    try:
        aguardar_tela_instalacao_pronta(driver, wait, timeout=8, pausa_extra=1.2)
    except Exception:
        time.sleep(1.5)
        aguardar_tela_instalacao_pronta(driver, wait, timeout=5, pausa_extra=0.8)


def iniciar_auditoria_execucao() -> None:
    """Zera o CSV de auditoria a cada nova execucao."""
    with AUDITORIA_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(AUDITORIA_HEADERS)
    log(f"Auditoria reiniciada para esta execucao: {AUDITORIA_CSV}")


def _classificar_status_erro(exc: Exception) -> str:
    msg = str(exc).lower()
    if "carimbo" in msg and "exist" in msg:
        return "pulado_carimbo_existente"
    return f"erro_no_fluxo:{type(exc).__name__}"


def filtrar_registros_pela_origem(registros: list[dict]) -> list[dict]:
    """
    Mantem apenas registros cujo carimbo aparece em algum PDF da pasta de origem.
    Se --pasta-origem nao for informado, retorna a lista original.
    """
    if not PASTA_ORIGEM_PDFS:
        warn("Sem --pasta-origem: processando todas as linhas da planilha.")
        return registros

    pasta = Path(PASTA_ORIGEM_PDFS)
    if not pasta.exists():
        raise Exception(f"Pasta de origem informada nao existe: {pasta}")

    arquivos = list(pasta.glob("*.pdf"))
    nomes = [p.name.lower() for p in arquivos]
    log(f"Origem BT: {pasta} | PDFs encontrados: {len(arquivos)}")

    filtrados = []
    for r in registros:
        carimbo = str(r.get("fatCarimbo", "") or "").strip().lower()
        if not carimbo:
            continue
        if any(carimbo in nome for nome in nomes):
            filtrados.append(r)

    log(f"Filtro por origem: {len(registros)} -> {len(filtrados)} registro(s) para digitar.")
    return filtrados


def validar_execucao_auditoria(total_esperado: int) -> bool:
    """Valida se todos os itens desta execucao foram resolvidos."""
    if not AUDITORIA_CSV.exists():
        warn("auditoria_resultados.csv nao encontrado apos execucao.")
        return False

    status_resolvidos = {
        "sucesso_auditoria",
        "auditoria_sem_valor",
        "pulado_referencia_existente",
        "pulado_carimbo_existente",
    }

    total_linhas = 0
    total_resolvidos = 0
    total_erros = 0

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with AUDITORIA_CSV.open("r", newline="", encoding=enc) as f:
                for row in csv.DictReader(f, delimiter=";"):
                    total_linhas += 1
                    status = _extrair_status_auditoria(row)
                    if status in status_resolvidos:
                        total_resolvidos += 1
                    else:
                        total_erros += 1
            break
        except UnicodeDecodeError:
            continue

    log(f"Validacao auditoria: esperado={total_esperado} | registrado={total_linhas} | resolvidos={total_resolvidos} | erros={total_erros}")
    if total_linhas != total_esperado:
        warn("Quantidade registrada na auditoria difere do esperado para esta execucao.")
        return False
    if total_erros > 0:
        warn("Existem linhas com erro/pendencia na auditoria desta execucao.")
        return False
    return True


def main():
    _parse_args()  # resolve --xlsx e --linha-inicio antes de qualquer acesso
    driver = None

    try:
        if not str(SENHA or "").strip():
            raise RuntimeError(
                "CONSEN_SENHA nao configurada para o usuario informado. Defina CONSEN_USUARIO/CONSEN_SENHA no ambiente ou use o login padrao do Robo Digitador."
            )

        registros = ler_todas_as_linhas_planilha(EXCEL_PATH, linha_inicio=LINHA_INICIO)
        iniciar_auditoria_execucao()
        registros = filtrar_registros_pela_origem(registros)
        total_esperado = len(registros)

        driver = iniciar_driver(headless=HEADLESS)
        wait = WebDriverWait(driver, 20)

        log("Abrindo login...")
        driver.get(LOGIN_URL)

        log("Fazendo login...")
        enviar_login(driver, wait, USUARIO, SENHA)
        time.sleep(1.5)

        total_auditorias = 0

        for idx, registro in enumerate(registros, start=1):
            try:
                log("=" * 80)
                log(f"PROCESSANDO LINHA {registro['linha_excel']} | instalação={registro['instalacao']} | item {idx}/{len(registros)}")

                voltar_para_tela_inicial_instalacao(driver, wait)

                log("Preparando campo de instalação...")
                aguardar_tela_instalacao_pronta(driver, wait, timeout=8, pausa_extra=1.2)

                log("Digitando instalação...")
                preencher_input_texto(driver, wait, "instalacao", registro["instalacao"], pausa_antes=0.8)

                log("Clicando em Carregar da instalação...")
                clicar_botao_carregar_instalacao(driver, wait)
                _aguardar_sem_spinner(driver, timeout=8, min_wait=0.5)
                aguardar_carregamento_tabela(driver, wait)

                prosseguir_fluxo = preencher_datas_e_carimbo_se_necessario(driver, wait, registro)
                if not prosseguir_fluxo:
                    registrar_resultado_auditoria(
                        linha_excel=registro["linha_excel"],
                        instalacao=registro["instalacao"],
                        data_referencia_esperada=formatar_ddmmyyyy(registro["dataReferenciaEsperada"]),
                        carimbo=registro.get("fatCarimbo", ""),
                        valor_auditoria="",
                        status="pulado_referencia_existente",
                    )
                    continue

                log("Preenchendo campos forçados essenciais...")
                preencher_aliases_forcados_linha(driver, registro["dados_completos"])

                log("Preenchendo saldos acumulados da usina (tratamento especial)...")
                preencher_saldos_injetado_usina(driver, wait, registro["dados_completos"])

                log("Preenchendo observações múltiplas...")
                preencher_obs_multiplas(driver, wait, registro["dados_completos"])

                caminho_csv_existente = PASTA_SAIDA / "mapeamento_campos_planilha.csv"
                if not caminho_csv_existente.exists() or idx == 1:
                    log("Exportando campos disponíveis e correlação...")
                    exportar_campos_e_correlacoes(driver, EXCEL_PATH)

                caminho_csv = escolher_csv_mapeamento()
                log("Preenchendo campos com base no CSV de mapeamento...")
                preencher_campos_via_mapeamento_csv_linha(
                    driver=driver,
                    dados_planilha=registro["dados_completos"],
                    caminho_csv=caminho_csv,
                    score_minimo=SCORE_MINIMO_MAPEAMENTO,
                )

                salvar_abrir_auditoria_capturar_fechar(driver, wait, registro)
                total_auditorias += 1

            except Exception as e:
                warn(f"Falha na linha {registro['linha_excel']} ({registro['instalacao']}): {type(e).__name__} - {e}")
                warn(traceback.format_exc())
                registrar_resultado_auditoria(
                    linha_excel=registro["linha_excel"],
                    instalacao=registro["instalacao"],
                    data_referencia_esperada=formatar_ddmmyyyy(registro["dataReferenciaEsperada"]),
                    carimbo=registro.get("fatCarimbo", ""),
                    valor_auditoria="",
                    status=_classificar_status_erro(e),
                )
                continue

        print(f"Processo concluído. Linhas com auditoria registrada: {total_auditorias}")
        gerar_csv_pendentes()
        ok_execucao = validar_execucao_auditoria(total_esperado=total_esperado)
        if not ok_execucao:
            sys.exit(1)
        sys.exit(0)

    except Exception as e:
        erro(str(e))
        sys.exit(1)

    finally:
        if driver:
            driver.quit()


def gerar_csv_pendentes():
    """
    Lê auditoria_resultados.csv e gera pendentes_HHMMSS.csv com instalação
    e carimbo de todas as contas que não foram auditadas com sucesso.
    """
    caminho_auditoria = AUDITORIA_CSV
    if not caminho_auditoria.exists():
        warn("auditoria_resultados.csv não encontrado — nenhum pendente gerado.")
        return

    STATUS_OK = {
        "sucesso_auditoria",
        "auditoria_sem_valor",
        "pulado_referencia_existente",
        "pulado_carimbo_existente",
    }

    pendentes = []
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with caminho_auditoria.open("r", newline="", encoding=enc) as f:
                for row in csv.DictReader(f, delimiter=";"):
                    status = _extrair_status_auditoria(row)
                    if status not in STATUS_OK:
                        pendentes.append({
                            "instalacao":  row.get("instalacao", ""),
                            "carimbo":     row.get("carimbo", ""),
                            "status":      status,
                            "linha_excel": row.get("linha_excel", ""),
                        })
            break
        except UnicodeDecodeError:
            continue

    if not pendentes:
        log("Nenhum pendente — todas as contas foram digitadas com sucesso.")
        return

    ts = datetime.now().strftime("%H%M%S")
    caminho_saida = PASTA_SAIDA / f"pendentes_{ts}.csv"
    with caminho_saida.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f, fieldnames=["instalacao", "carimbo", "status", "linha_excel"], delimiter=";"
        )
        writer.writeheader()
        writer.writerows(pendentes)

    log(f"Pendentes: {len(pendentes)} conta(s) → {caminho_saida}")


if __name__ == "__main__":
    main()
