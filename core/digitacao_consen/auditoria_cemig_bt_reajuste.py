#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Auditoria CEMIG BT — Reajuste tarifário de 28/05/2026.

Para cada carimbo CEMIG BT da planilha de auditoria:
  1. Abre no CONSEN via ConsenSession
  2. Lê fatDataLeituraAnterior e fatDataLeituraAtual
  3. Se o período cruza 28/05/2026 → registra análise "Tarifa proporcional ao reajuste"
  4. Caso contrário → sinaliza para investigação manual

Uso:
    python auditoria_cemig_bt_reajuste.py [--salvar] [--limite N] [--carimbo BB_XXXXXXX]

    --salvar     Efetiva o registro de análise no CONSEN (sem isso, só lê e classifica)
    --limite N   Processa só os N primeiros carimbos (teste)
    --carimbo    Processa carimbo específico (repetível)
"""

from __future__ import annotations

import argparse
import csv
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from digitacao_consen.consen_api import ConsenSession, carregar_lista_carimbos_args
    from digitacao_consen.correcao_fluxo_base import log, warn
except ModuleNotFoundError:
    from consen_api import ConsenSession, carregar_lista_carimbos_args  # type: ignore
    from correcao_fluxo_base import log, warn  # type: ignore

# ── Configuração ──────────────────────────────────────────────────────────────

AUDITORIA_XLSX = Path(
    "//10.10.250.21/Energia/ARQUIVOS ENZO"
    "/ALTERACOES APOS AUDITORIA/Auditoria"
    "/Auditoria a Menor de Julho.xlsx"
)
SAIDA_DIR = Path(
    "//10.10.250.21/Energia/ARQUIVOS ENZO"
    "/ALTERACOES APOS AUDITORIA/Auditoria"
)
EXECUCAO_CSV = SAIDA_DIR / "auditoria_cemig_bt_reajuste_execucao.csv"

DATA_REAJUSTE = date(2026, 5, 28)
TEXTO_ANALISE = "Tarifa proporcional ao reajuste"

CAMPOS_DATAS = ("fatDataLeituraAnterior", "fatDataLeituraAtual")
_ESPERA_PAGINA = 1.5  # segundos após navegação antes de ler campos


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(valor: str) -> Optional[date]:
    if not valor or str(valor).strip() in ("", "None", "nan"):
        return None
    s = str(valor).strip().split(" ")[0].split("T")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _cruza_reajuste(d_ant: Optional[date], d_atu: Optional[date]) -> bool:
    if not d_ant or not d_atu:
        return False
    return d_ant < DATA_REAJUSTE <= d_atu


def _carregar_carimbos_auditoria_cemig() -> list[str]:
    """Lê carimbos CEMIG da planilha de auditoria."""
    wb = openpyxl.load_workbook(str(AUDITORIA_XLSX))
    ws = wb["Export"]
    headers = [str(c.value or "") for c in ws[1]]
    carimbos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        car = str(rec.get("Carimbo") or "").strip()
        conc = str(rec.get("Concessionaria") or "")
        analise = rec.get("Análise") or rec.get("Análise") or ""
        if car.startswith("2") and "CEMIG" in conc and not analise:
            carimbos.append(car)
    return carimbos


def _registrar(carimbo: str, classificacao: str, detalhe: str, d_ant: str, d_atu: str) -> None:
    existe = EXECUCAO_CSV.exists()
    with open(EXECUCAO_CSV, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if not existe:
            w.writerow(["carimbo", "classificacao", "detalhe", "leit_anterior", "leit_atual", "data_hora"])
        w.writerow([
            carimbo, classificacao, detalhe, d_ant, d_atu,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        ])


# ── Lógica principal ──────────────────────────────────────────────────────────

def rodar(carimbos: list[str], salvar: bool) -> None:
    SAIDA_DIR.mkdir(parents=True, exist_ok=True)

    if not salvar:
        log("MODO LEITURA: use --salvar para registrar análise no CONSEN.")

    contadores = {"proporcional": 0, "investigar": 0, "sem_datas": 0, "erro": 0}

    with ConsenSession.abrir() as s:
        for carimbo in carimbos:
            log(f"=== BB_{carimbo} ===")
            try:
                s.buscar_carimbo(carimbo, aguardar_campos=CAMPOS_DATAS)
                val_ant = s.obter_campo("fatDataLeituraAnterior")
                val_atu = s.obter_campo("fatDataLeituraAtual")
                d_ant = _parse_date(val_ant)
                d_atu = _parse_date(val_atu)

                log(f"  Leitura anterior: {val_ant}  |  Leitura atual: {val_atu}")

                if not d_ant or not d_atu:
                    warn(f"  BB_{carimbo}: datas de leitura nao encontradas no CONSEN.")
                    contadores["sem_datas"] += 1
                    _registrar(carimbo, "sem_datas", "campos de data nao localizados", val_ant, val_atu)
                    continue

                if _cruza_reajuste(d_ant, d_atu):
                    log("  Cruza 28/05 -> Tarifa proporcional ao reajuste")
                    contadores["proporcional"] += 1
                    if salvar:
                        s.salvar_e_auditar()
                        log("  Analise registrada no CONSEN.")
                    _registrar(carimbo, "proporcional", TEXTO_ANALISE, val_ant, val_atu)
                else:
                    warn(f"  Nao cruza 28/05 ({d_ant} -> {d_atu}) -> INVESTIGAR")
                    contadores["investigar"] += 1
                    _registrar(carimbo, "investigar",
                               f"periodo {d_ant} a {d_atu} nao cruza reajuste", val_ant, val_atu)

            except Exception as exc:
                warn(f"  BB_{carimbo}: {type(exc).__name__}: {exc}")
                contadores["erro"] += 1
                _registrar(carimbo, "erro", str(exc)[:200], "", "")

    log("\n=== RESUMO ===")
    log(f"  Tarifa proporcional: {contadores['proporcional']}")
    log(f"  Investigar:          {contadores['investigar']}")
    log(f"  Sem datas:           {contadores['sem_datas']}")
    log(f"  Erros:               {contadores['erro']}")
    log(f"\nResultados em: {EXECUCAO_CSV}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Auditoria CEMIG BT — Reajuste 28/05/2026"
    )
    p.add_argument("--salvar", action="store_true",
                   help="Registra análise no CONSEN (sem esta flag só lê)")
    p.add_argument("--limite", type=int, default=0,
                   help="Processa só os N primeiros carimbos (0 = todos)")
    p.add_argument("--carimbo", action="append", default=[],
                   help="Carimbo específico (BB_XXXXXXX ou só número). Repetível.")
    p.add_argument("--carimbos-arquivo", type=str, default="",
                   help="Arquivo TXT com um carimbo por linha")
    return p.parse_args()


def main() -> int:
    args = parse_args()

    # Carimbos explícitos via CLI
    carimbos_cli = carregar_lista_carimbos_args(args) if (args.carimbo or args.carimbos_arquivo) else []

    if carimbos_cli:
        carimbos = carimbos_cli
    else:
        log(f"Carregando carimbos CEMIG sem análise de: {AUDITORIA_XLSX.name}")
        carimbos = _carregar_carimbos_auditoria_cemig()

    if not carimbos:
        print("Nenhum carimbo CEMIG pendente encontrado.")
        return 0

    if args.limite:
        carimbos = carimbos[: args.limite]

    log(f"{len(carimbos)} carimbos a processar. Data reajuste: {DATA_REAJUSTE}")

    rodar(carimbos, salvar=args.salvar)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
