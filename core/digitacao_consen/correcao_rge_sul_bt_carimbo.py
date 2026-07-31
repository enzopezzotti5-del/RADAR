#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Correcao RGE Sul BT por carimbo.

Fluxo:
1. Carrega os carimbos a partir das planilhas de auditoria RGE B3.
2. Localiza os PDFs BB_<carimbo>.pdf na arvore de CARIMBOS DIGITADOS.
3. Reprocessa com o OCR dedicado da RGE.
4. Sobrepoe no Consen os campos de consumo, energia injetada, bandeiras
   e retencoes tributarias.

Por seguranca, nao salva por padrao. Use --salvar para efetivar.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import sys
from pathlib import Path
from typing import Any

import openpyxl
from selenium.common.exceptions import InvalidSessionIdException, WebDriverException

DEFAULT_SAIDA_DIR = "//10.10.250.21/Energia/ARQUIVOS ENZO/RGE_pipeline_saida/correcoes_por_carimbo"
DEFAULT_AUDIT_DIR = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/APOIO/Audit - xxx - B3 por concessionaria")
DEFAULT_AUDIT_FILES = (
    DEFAULT_AUDIT_DIR / "Audit - xxx - B3 - RGE.xlsx",
    DEFAULT_AUDIT_DIR / "Audit - xxx - B3 - RGE - ANTIGA AES SUL.xlsx",
)
NETWORK_PDFS_ROOT = Path("//10.10.250.21/Energia/CONTROLE BB/DIGITADOS/CARIMBOS DIGITADOS")
os.environ.setdefault("CONSEN_PIPELINE_SAIDA", DEFAULT_SAIDA_DIR)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import _venv_check  # noqa: E402,F401

try:
    from digitacao_consen import correcao_fluxo_base as fluxo_base  # noqa: E402
    from digitacao_consen.correcao_fluxo_base import (  # noqa: E402
        CorrecaoFluxoConfig,
        formatar_valor_para_campo,
        log,
        warn,
    )
    from ocr.ocr_rge_sul_bt import processar_pdf  # noqa: E402
except ModuleNotFoundError:
    import correcao_fluxo_base as fluxo_base  # type: ignore  # noqa: E402
    from correcao_fluxo_base import (  # type: ignore  # noqa: E402
        CorrecaoFluxoConfig,
        formatar_valor_para_campo,
        log,
        warn,
    )
    from ocr_rge_sul_bt import processar_pdf  # type: ignore  # noqa: E402

LOGIN_URL = fluxo_base.LOGIN_URL
BASE_URL = LOGIN_URL.rsplit("login.php", 1)[0]
EDIT_HASH = "#bpg/gestao/fatura/editaFaturaCarimbo.php"
EDIT_URL = os.environ.get("CONSEN_EDITA_FATURA_CARIMBO_URL", f"{BASE_URL}index.php{EDIT_HASH}")

SAIDA_DIR = Path(os.environ.get("CONSEN_CORRECAO_SAIDA", DEFAULT_SAIDA_DIR))
PDFS_ROOT = Path(os.environ.get("CONSEN_CORRECAO_RGE_BT_ROOT", str(NETWORK_PDFS_ROOT)))
FECHAR_AO_FINAL = os.environ.get("CONSEN_CORRECAO_FECHAR", "1").strip().lower() not in {"0", "false", "nao", "não"}
EXECUCAO_CSV = SAIDA_DIR / "correcao_execucao.csv"

CAMPOS_CRITICOS_TELA: tuple[str, ...] = (
    "fatConFPontaIndRegistrado",
    "fatConFPontaInjetadoRegistrado",
    "fatValorNFiscal",
)

CAMPO_OCR_PARA_TELA: tuple[tuple[str, str], ...] = (
    ("cb-subgrupo", "cadSubGrupoCod"),
    ("fatConFPontaIndRegistrado", "fatConFPontaIndRegistrado"),
    ("fatConFPontaIndutivo", "fatConFPontaIndFaturado"),
    ("fatConFPontaIndValorReais", "fatConFPontaIndValorReais"),
    ("fatConFPontaInjetadoRegistrado", "fatConFPontaInjetadoRegistrado"),
    ("fatConFPontaInjetado", "fatConFPontaInjetadoFaturado"),
    ("fatConFPontaInjetadoValorReais", "fatConFPontaInjetadoValorReais"),
    ("fatConFPontaInjetadoUsina", "fatConFPontaInjetadoUsina"),
    ("fatICMS", "fatICMS"),
    ("fatValBandeira", "fatValBandeira"),
    ("fatValBandeira2", "fatValBandeira2"),
    ("fatValorNFiscal", "fatValorNotaFiscal"),
    ("fatDescPisPercRetImposto", "fatDescPisPercRetImposto"),
    ("fatDescPisValRetImposto", "fatDescPisValRetImposto"),
    ("fatDescCofinsPercRetImposto", "fatDescCofinsPercRetImposto"),
    ("fatDescCofinsValRetImposto", "fatDescCofinsValRetImposto"),
    ("fatDescCsllPercRetImposto", "fatDescCsllPercRetImposto"),
    ("fatDescCsllValRetImposto", "fatDescCsllValRetImposto"),
    ("fatDescIrpjPercRetImposto", "fatDescIrpjPercRetImposto"),
    ("fatDescIrpjValRetImposto", "fatDescIrpjValRetImposto"),
    ("fatDescIrrfPercRetImposto", "fatDescIrrfPercRetImposto"),
    ("fatDescIrrfValRetImposto", "fatDescIrrfValRetImposto"),
    ("fatDescConsumoPercRetImposto", "fatDescConsumoPercRetImposto"),
    ("fatDescConsumoValRetImposto", "fatDescConsumoValRetImposto"),
)

ORDEM_CAMPOS_CORRECAO: tuple[str, ...] = (
    "cb-subgrupo",
    "fatConFPontaIndRegistrado",
    "fatConFPontaIndutivo",
    "fatConFPontaIndValorReais",
    "fatConFPontaInjetadoRegistrado",
    "fatConFPontaInjetado",
    "fatConFPontaInjetadoValorReais",
    "fatConFPontaInjetadoUsina",
    "fatICMS",
    "fatValBandeira",
    "fatValBandeira2",
    "fatValorNFiscal",
    "fatDescPisPercRetImposto",
    "fatDescPisValRetImposto",
    "fatDescCofinsPercRetImposto",
    "fatDescCofinsValRetImposto",
    "fatDescCsllPercRetImposto",
    "fatDescCsllValRetImposto",
    "fatDescIrpjPercRetImposto",
    "fatDescIrpjValRetImposto",
    "fatDescIrrfPercRetImposto",
    "fatDescIrrfValRetImposto",
    "fatDescConsumoPercRetImposto",
    "fatDescConsumoValRetImposto",
)

CONFIG = CorrecaoFluxoConfig(
    saida_dir=SAIDA_DIR,
    execucao_csv=EXECUCAO_CSV,
    edit_url=EDIT_URL,
    ordem_campos=ORDEM_CAMPOS_CORRECAO,
    fechar_ao_final=FECHAR_AO_FINAL,
)


def normalizar_carimbo(carimbo: str) -> str:
    return fluxo_base.normalizar_carimbo(carimbo)


def valor_vazio(valor: Any) -> bool:
    return fluxo_base.valor_vazio(valor)


def formatar_valor_correcao(header: str, valor: Any) -> str:
    return formatar_valor_para_campo(header, valor, "text")


def _normalizar_data_pasta(valor: Any) -> str:
    if isinstance(valor, dt.datetime):
        data = valor.date()
    elif isinstance(valor, dt.date):
        data = valor
    else:
        txt = str(valor or "").strip()
        if not txt:
            return ""
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
            try:
                data = dt.datetime.strptime(txt, fmt).date()
                break
            except ValueError:
                continue
        else:
            return ""
    return data.strftime("%d%m%Y")


def carregar_carimbos_auditoria(paths: list[Path] | tuple[Path, ...]) -> list[dict[str, str]]:
    registros: list[dict[str, str]] = []
    vistos: set[str] = set()

    for path in paths:
        if not path.exists():
            warn(f"Planilha de auditoria nao encontrada: {path}")
            continue

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            headers = [str(h or "").strip() for h in next(rows, ())]
            idx = {h: i for i, h in enumerate(headers)}
            if "Carimbo" not in idx:
                warn(f"Planilha sem coluna Carimbo: {path}")
                continue

            for row in rows:
                carimbo_bruto = row[idx["Carimbo"]] if idx["Carimbo"] < len(row) else ""
                if not str(carimbo_bruto or "").strip():
                    continue
                carimbo = normalizar_carimbo(str(carimbo_bruto))
                if carimbo in vistos:
                    continue
                cadastro = row[idx["Cadastro"]] if "Cadastro" in idx and idx["Cadastro"] < len(row) else ""
                instalacao = row[idx["Instalacao"]] if "Instalacao" in idx and idx["Instalacao"] < len(row) else ""
                registros.append(
                    {
                        "carimbo": carimbo,
                        "cadastro_pasta": _normalizar_data_pasta(cadastro),
                        "instalacao": str(instalacao or "").strip(),
                        "origem_auditoria": str(path),
                    }
                )
                vistos.add(carimbo)
        finally:
            wb.close()

    return registros


def _mapa_por_carimbo(registros: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {r["carimbo"]: r for r in registros if r.get("carimbo")}


def localizar_pdfs_por_carimbo(
    raiz: Path,
    registros_auditoria: list[dict[str, str]],
    carimbos_filtro: set[str],
) -> dict[str, Path]:
    encontrados: dict[str, Path] = {}
    pendentes = set(carimbos_filtro)
    if not pendentes:
        return encontrados

    meta = _mapa_por_carimbo(registros_auditoria)

    for carimbo in sorted(pendentes):
        pasta_data = meta.get(carimbo, {}).get("cadastro_pasta", "")
        if not pasta_data:
            continue
        candidato = raiz / pasta_data / f"BB_{carimbo}.pdf"
        if candidato.exists():
            encontrados[carimbo] = candidato

    pendentes -= set(encontrados)
    if not pendentes:
        return encontrados

    for atual, _, arquivos in os.walk(raiz):
        if not pendentes:
            break
        pasta = Path(atual)
        for nome in arquivos:
            if not nome.lower().endswith(".pdf"):
                continue
            stem = Path(nome).stem
            try:
                carimbo = normalizar_carimbo(stem)
            except Exception:
                continue
            if carimbo in pendentes and carimbo not in encontrados:
                encontrados[carimbo] = pasta / nome
                pendentes.remove(carimbo)
                log(f"[LOCALIZADO] BB_{carimbo} -> {encontrados[carimbo]}")
                if not pendentes:
                    break
    return encontrados


def correcoes_da_linha_ocr(row: dict[str, Any]) -> dict[str, str]:
    correcoes: dict[str, str] = {
        "cb-subgrupo": "B3 [<2,3kV]",
    }

    for campo_tela, header in CAMPO_OCR_PARA_TELA:
        valor = row.get(header)
        if valor_vazio(valor):
            continue
        correcoes[campo_tela] = formatar_valor_correcao(header, valor)

    return correcoes


def carregar_correcoes_de_raiz_pdfs(
    raiz: Path,
    registros_auditoria: list[dict[str, str]],
    carimbos_filtro: set[str],
) -> dict[str, dict[str, str]]:
    encontrados = localizar_pdfs_por_carimbo(raiz, registros_auditoria, carimbos_filtro)
    meta = _mapa_por_carimbo(registros_auditoria)
    correcoes: dict[str, dict[str, str]] = {}
    linhas_log: list[dict[str, Any]] = []

    for carimbo in sorted(carimbos_filtro):
        info = meta.get(carimbo, {})
        pdf = encontrados.get(carimbo)
        if not pdf:
            linhas_log.append(
                {
                    "carimbo": carimbo,
                    "instalacao": info.get("instalacao", ""),
                    "cadastro_pasta": info.get("cadastro_pasta", ""),
                    "arquivo": "",
                    "status": "nao_encontrado",
                    "campos": 0,
                }
            )
            warn(f"[LOCALIZADO] BB_{carimbo}: PDF nao encontrado em {raiz}")
            continue

        try:
            row = processar_pdf(pdf)
        except Exception as exc:
            linhas_log.append(
                {
                    "carimbo": carimbo,
                    "instalacao": info.get("instalacao", ""),
                    "cadastro_pasta": info.get("cadastro_pasta", ""),
                    "arquivo": str(pdf),
                    "status": f"erro_ocr:{type(exc).__name__}",
                    "campos": 0,
                }
            )
            warn(f"[OCR] {pdf.name}: {type(exc).__name__}: {exc}")
            continue

        mapa = correcoes_da_linha_ocr(row)
        if mapa:
            correcoes[carimbo] = mapa

        linhas_log.append(
            {
                "carimbo": carimbo,
                "instalacao": info.get("instalacao", ""),
                "cadastro_pasta": info.get("cadastro_pasta", ""),
                "arquivo": str(pdf),
                "status": "ok" if mapa else "sem_campos",
                "campos": len(mapa),
                "fponta_reg": row.get("fatConFPontaIndRegistrado", ""),
                "fponta_fat": row.get("fatConFPontaIndFaturado", ""),
                "fponta_valor": row.get("fatConFPontaIndValorReais", ""),
                "inj_reg": row.get("fatConFPontaInjetadoRegistrado", ""),
                "inj_fat": row.get("fatConFPontaInjetadoFaturado", ""),
                "inj_valor": row.get("fatConFPontaInjetadoValorReais", ""),
                "inj_usina": row.get("fatConFPontaInjetadoUsina", ""),
                "bandeira": row.get("fatValBandeira", ""),
                "bandeira2": row.get("fatValBandeira2", ""),
                "valor_nf": row.get("fatValorNotaFiscal", ""),
                "pis_ret": row.get("fatDescPisValRetImposto", ""),
                "cofins_ret": row.get("fatDescCofinsValRetImposto", ""),
                "csll_ret": row.get("fatDescCsllValRetImposto", ""),
                "irpj_ret": row.get("fatDescIrpjValRetImposto", ""),
            }
        )

    salvar_log_preparacao(linhas_log)
    return correcoes


def salvar_log_preparacao(linhas: list[dict[str, Any]]) -> None:
    SAIDA_DIR.mkdir(parents=True, exist_ok=True)
    path = SAIDA_DIR / "preparacao_correcao_rge_bt.csv"
    campos = [
        "carimbo",
        "instalacao",
        "cadastro_pasta",
        "arquivo",
        "status",
        "campos",
        "fponta_reg",
        "fponta_fat",
        "fponta_valor",
        "inj_reg",
        "inj_fat",
        "inj_valor",
        "inj_usina",
        "bandeira",
        "bandeira2",
        "valor_nf",
        "pis_ret",
        "cofins_ret",
        "csll_ret",
        "irpj_ret",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore", delimiter=";")
        writer.writeheader()
        writer.writerows(linhas)
    log(f"Log de preparacao salvo: {path}")

    carimbos_ok = sorted({str(l.get("carimbo") or "") for l in linhas if l.get("status") == "ok"})
    txt = SAIDA_DIR / "carimbos_preparados_rge_bt.txt"
    txt.write_text("\n".join(f"BB_{c}" for c in carimbos_ok if c), encoding="utf-8")
    log(f"Lista de carimbos preparados: {txt}")


def registrar_execucao(carimbo: str, status: str, detalhe: str = "") -> None:
    fluxo_base.registrar_execucao(CONFIG.execucao_csv, carimbo, status, detalhe)


def carregar_status_execucao() -> dict[str, str]:
    return fluxo_base.carregar_status_execucao(CONFIG.execucao_csv)


def abrir_driver_logado():
    return fluxo_base.abrir_driver_logado()


def abrir_tela_edicao_carimbo(driver, wait) -> None:
    fluxo_base.abrir_tela_edicao_carimbo(driver, wait, CONFIG.edit_url)


def carregar_fatura_por_carimbo(driver, wait, carimbo: str) -> None:
    fluxo_base.carregar_fatura_por_carimbo(driver, wait, carimbo, CAMPOS_CRITICOS_TELA)


def aplicar_correcoes(driver, wait, carimbo: str, correcoes: dict[str, Any]) -> tuple[int, int, int]:
    return fluxo_base.aplicar_correcoes(driver, wait, carimbo, correcoes, CONFIG.ordem_campos)


def salvar_auditar_e_avancar(driver, wait, carimbo: str) -> None:
    fluxo_base.salvar_auditar_e_avancar(driver, wait, carimbo)


def salvar_snapshot(driver, carimbo: str) -> None:
    fluxo_base.salvar_snapshot(driver, CONFIG.saida_dir, carimbo)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Correcao RGE Sul BT por carimbo no Consen")
    p.add_argument("--carimbo", action="append", default=[], help="Ex: --carimbo BB_2011186")
    p.add_argument("--carimbos-arquivo", type=str, default="", help="TXT com um carimbo por linha")
    p.add_argument(
        "--planilha-auditoria",
        action="append",
        default=[],
        help="Pode ser informado mais de uma vez. Se omitido, usa as planilhas RGE padrao.",
    )
    p.add_argument("--pasta-pdfs", type=str, default=str(PDFS_ROOT), help="Raiz para busca recursiva dos PDFs")
    p.add_argument("--preparar-apenas", action="store_true", help="Apenas OCR + monta correcoes, sem abrir Consen")
    p.add_argument("--retomar-apos", type=str, default="", help="Retoma o lote apos este carimbo")
    p.add_argument("--reprocessar-ok", action="store_true", help="Nao pula carimbos ja concluidos com status ok")
    p.add_argument("--salvar", action="store_true", help="Salva a fatura apos aplicar correcoes")
    p.add_argument("--sem-snapshot", action="store_true", help="Nao salva HTML/JSON da tela")
    p.add_argument("--limite", type=int, default=0)
    return p.parse_args()


def _carregar_filtro_carimbos(args: argparse.Namespace) -> set[str]:
    carimbos_filtro: set[str] = set()
    for item in list(args.carimbo or []):
        carimbos_filtro.add(normalizar_carimbo(item))
    if args.carimbos_arquivo:
        path = Path(args.carimbos_arquivo)
        for line in path.read_text(encoding="utf-8-sig").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                carimbos_filtro.add(normalizar_carimbo(line))
    return carimbos_filtro


def _resolver_planilhas(args: argparse.Namespace) -> list[Path]:
    if args.planilha_auditoria:
        return [Path(p) for p in args.planilha_auditoria]
    return list(DEFAULT_AUDIT_FILES)


def main() -> int:
    args = parse_args()
    planilhas = _resolver_planilhas(args)
    registros_auditoria = carregar_carimbos_auditoria(planilhas)
    if not registros_auditoria:
        print("Nenhum carimbo encontrado nas planilhas de auditoria informadas.")
        return 2

    carimbos_filtro = _carregar_filtro_carimbos(args)
    if not carimbos_filtro:
        carimbos_filtro = {r["carimbo"] for r in registros_auditoria}

    raiz_pdfs = Path(args.pasta_pdfs)
    log("=== Correcao RGE Sul BT ===")
    log(f"Planilhas  : {len(planilhas)}")
    log(f"Pasta PDFs : {raiz_pdfs}")
    log(f"Carimbos   : {len(carimbos_filtro)}")

    correcoes = carregar_correcoes_de_raiz_pdfs(raiz_pdfs, registros_auditoria, carimbos_filtro)

    if args.preparar_apenas:
        log(f"[preparar-apenas] {len(correcoes)} carimbos preparados. Encerrando sem abrir Consen.")
        return 0

    if not correcoes:
        log("Nenhuma correcao encontrada. Encerrando.")
        return 0

    carimbos_ordenados = sorted(correcoes)
    if args.retomar_apos:
        ancora = normalizar_carimbo(args.retomar_apos)
        if ancora in carimbos_ordenados:
            idx = carimbos_ordenados.index(ancora) + 1
            carimbos_ordenados = carimbos_ordenados[idx:]
            log(f"Retomando apos BB_{ancora}: {len(carimbos_ordenados)} restantes.")

    status_execucao = carregar_status_execucao()
    if not args.reprocessar_ok:
        antes = len(carimbos_ordenados)
        carimbos_ordenados = [c for c in carimbos_ordenados if status_execucao.get(c) != "ok"]
        log(f"Pulados (ok anteriores): {antes - len(carimbos_ordenados)}")

    if args.limite > 0:
        carimbos_ordenados = carimbos_ordenados[: args.limite]

    log(f"Processando {len(carimbos_ordenados)} carimbos.")
    if not carimbos_ordenados:
        log("Nenhum carimbo para processar.")
        return 0

    driver = None
    try:
        driver, wait = abrir_driver_logado()
        abrir_tela_edicao_carimbo(driver, wait)

        for i, carimbo in enumerate(carimbos_ordenados, start=1):
            log(f"--- [{i}/{len(carimbos_ordenados)}] BB_{carimbo} ---")
            try:
                carregar_fatura_por_carimbo(driver, wait, carimbo)
                if not args.sem_snapshot:
                    salvar_snapshot(driver, carimbo)
                apl, conf, total = aplicar_correcoes(driver, wait, carimbo, correcoes[carimbo])
                if args.salvar and apl > 0:
                    salvar_auditar_e_avancar(driver, wait, carimbo)
                    registrar_execucao(carimbo, "ok", f"apl={apl} conf={conf}/{total}")
                elif not args.salvar:
                    log(f"BB_{carimbo}: --salvar nao informado. Correcoes NAO salvas.")
                    registrar_execucao(carimbo, "simulado", f"apl={apl} conf={conf}/{total}")
                else:
                    log(f"BB_{carimbo}: nenhum campo alterado.")
                    registrar_execucao(carimbo, "sem_alteracao")
            except (InvalidSessionIdException, WebDriverException) as exc:
                warn(f"BB_{carimbo}: sessao perdida ({type(exc).__name__}). Tentando relogin...")
                registrar_execucao(carimbo, "erro_sessao", str(exc))
                try:
                    driver.quit()
                except Exception:
                    pass
                driver, wait = abrir_driver_logado()
                abrir_tela_edicao_carimbo(driver, wait)
            except Exception as exc:
                warn(f"BB_{carimbo}: {type(exc).__name__}: {exc}")
                registrar_execucao(carimbo, "erro", str(exc))

    finally:
        if FECHAR_AO_FINAL and driver is not None:
            try:
                driver.quit()
            except Exception:
                pass

    log("=== Correcao RGE Sul BT concluida ===")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
