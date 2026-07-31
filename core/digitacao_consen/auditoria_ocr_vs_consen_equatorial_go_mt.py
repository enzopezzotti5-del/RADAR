#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Auditoria OCR vs Consen — Equatorial GO MT.

Para cada carimbo MT da massa informada:
  1. Carrega a fatura no Consen (tela de edição por carimbo).
  2. Lê os campos do formulário via coletar_campos_visiveis.
  3. Reprocessa o PDF com o OCR atual.
  4. Compara valor a valor e gera CSV de divergências.

Uso:
    python auditoria_ocr_vs_consen_equatorial_go_mt.py \\
        --carimbos-arquivo lista.txt \\
        --raiz-pdfs "//servidor/CARIMBOS DIGITADOS"

    python auditoria_ocr_vs_consen_equatorial_go_mt.py \\
        --carimbo BB_964503 --carimbo BB_964504 \\
        --raiz-pdfs "//servidor/CARIMBOS DIGITADOS"

    # Apenas lê OCR sem abrir navegador (útil para testar):
    python auditoria_ocr_vs_consen_equatorial_go_mt.py \\
        --xlsx "ocr_equatorial_go_MT_07112025_SGO.xlsx" \\
        --raiz-pdfs "//servidor/CONTROLE BB/DIGITADOS/BB 2025/07112025_SGO" \\
        --so-ocr
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    import _venv_check  # noqa: F401
except ImportError:
    pass

try:
    from digitacao_consen import correcao_fluxo_base as fluxo_base
    from digitacao_consen.correcao_fluxo_base import (
        CorrecaoFluxoConfig,
        log,
        warn,
        normalizar_carimbo,
        abrir_driver_logado,
        abrir_tela_edicao_carimbo,
        carregar_fatura_por_carimbo,
        coletar_campos_visiveis,
    )
    from ocr import ocr_equatorial_go as ocr_eq
except ModuleNotFoundError:
    import correcao_fluxo_base as fluxo_base  # type: ignore
    from correcao_fluxo_base import (  # type: ignore
        CorrecaoFluxoConfig,
        log,
        warn,
        normalizar_carimbo,
        abrir_driver_logado,
        abrir_tela_edicao_carimbo,
        carregar_fatura_por_carimbo,
        coletar_campos_visiveis,
    )
    import ocr_equatorial_go as ocr_eq  # type: ignore

# ---------------------------------------------------------------------------
# Mapeamento campo lógico → IDs HTML possíveis na tela Equatorial GO MT
# ---------------------------------------------------------------------------
CAMPO_IDS: dict[str, tuple[str, ...]] = {
    "fatConFPontaIndRegistrado":            ("txt-consumo-registrada-fpind",    "fatConFPontaIndRegistrado"),
    "fatConFPontaIndFaturado":              ("txt-consumo-faturada-fpind",       "fatConFPontaIndFaturado"),
    "fatConFPontaIndValorReais":            ("txt-consumo-fpind-valor-reais",    "fatConFPontaIndValorReais"),
    "fatDemFPontaIndRegistrada":            ("txt-demandas-registrada-fpind",    "fatDemFPontaIndRegistrada"),
    "fatDemFPontaIndFaturada":              ("txt-demandas-faturada-fpind",      "fatDemFPontaIndFaturada"),
    "fatDemFPontaIndValorReais":            ("txt-demandas-fpind-valor-reais",   "fatDemFPontaIndValorReais"),
    "fatDemPontaIndRegistrada":             ("txt-demandas-registrada-pind",     "fatDemPontaIndRegistrada"),
    "fatDemPontaIndFaturada":               ("txt-demandas-faturada-pind",       "fatDemPontaIndFaturada"),
    "fatConPontaRegistrado":                ("txt-consumo-registrada-pind",      "fatConPontaRegistrado"),
    "fatConPontaFaturado":                  ("txt-consumo-faturada-pind",        "fatConPontaFaturado"),
    "fatConPontaValorReais":                ("txt-consumo-pind-valor-reais",     "fatConPontaValorReais"),
    "fatDescontoFio":                       ("fatDescontoFio",),
    "fatDescontoFioKWh":                    ("fatDescontoFioKWh",),
    "fatEscassezHidrica":                   ("fatEscassezHidrica",),
    "fatEscassezHidricaValorReais":         ("fatEscassezHidricaValorReais",),
    "fatMultas":                            ("fatMultas",),
    "fatValorNotaFiscal":                   ("txt-dados-financeiros-valor-nota-fiscal", "fatValorNFiscal", "fatValorNotaFiscal"),
    "fatBeneficioTarifarioBrutoValorReais": ("fatBeneficioTarifarioBrutoValorReais",),
    "fatBeneficioLiquidoValorReais":        ("fatBeneficioLiquidoValorReais",),
    "fatICMS":                              ("fatICMS",),
    "fatPIS":                               ("fatPIS",),
    "fatCOFINS":                            ("fatCOFINS",),
    "fatValorFatura":                       ("fatValorFatura",),
    "fatIlumPublica":                       ("fatIlumPublica",),
}

CAMPOS_ORDEM = list(CAMPO_IDS.keys())

DEFAULT_SAIDA_DIR = Path(
    os.environ.get(
        "CONSEN_AUDITORIA_SAIDA",
        "//10.10.250.21/Energia/ARQUIVOS ENZO/EQUATORIAL_GO_producao_saida/auditoria_ocr_vs_consen",
    )
)
DEFAULT_PDFS_ROOT = Path(
    os.environ.get("CONSEN_CORRECAO_EQUATORIAL_GO_MT_ROOT",
                   "//10.10.250.21/Energia/CONTROLE BB/DIGITADOS/CARIMBOS DIGITADOS")
)
EDIT_URL = os.environ.get(
    "CONSEN_EDITA_FATURA_CARIMBO_URL",
    "https://consen.acaoengenharia.com.br/index.php#bpg/gestao/fatura/editaFaturaCarimbo.php",
)
CAMPOS_CRITICOS_TELA: tuple[str, ...] = ("btnSalvar", "instalacao")


# ---------------------------------------------------------------------------
# Leitura de campos do Consen
# ---------------------------------------------------------------------------

def _br_to_float(txt: str) -> float | None:
    """Converte '1.234,56' ou '1234.56' para float."""
    txt = re.sub(r"[^\d,.-]", "", (txt or "").strip())
    if not txt:
        return None
    txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return None


def ler_campos_consen(campos_form: list[dict[str, str]]) -> dict[str, str]:
    """
    Dado o resultado de coletar_campos_visiveis(), produz
    {campo_logico: valor_consen} usando CAMPO_IDS como índice.
    """
    by_id: dict[str, str] = {}
    by_name: dict[str, str] = {}
    for f in campos_form:
        fid = (f.get("id") or "").strip()
        fname = (f.get("name") or "").strip()
        val = (f.get("value") or f.get("text") or "").strip()
        if fid:
            by_id[fid] = val
        if fname:
            by_name[fname] = val

    resultado: dict[str, str] = {}
    for campo_logico, ids in CAMPO_IDS.items():
        for fid in ids:
            v = by_id.get(fid) or by_name.get(fid)
            if v is not None:
                resultado[campo_logico] = v
                break
        else:
            resultado[campo_logico] = ""
    return resultado


# ---------------------------------------------------------------------------
# OCR
# ---------------------------------------------------------------------------

def _score_data_pasta(pdf: Path) -> tuple[int, int, int]:
    for parte in (pdf.parent.name, pdf.parent.parent.name):
        m = re.fullmatch(r"(\d{2})(\d{2})(\d{4})", str(parte).strip())
        if m:
            return int(m.group(3)), int(m.group(2)), int(m.group(1))
    return (0, 0, 0)


def localizar_pdf(raiz: Path, carimbo_norm: str) -> Path | None:
    for atual, _, arquivos in os.walk(raiz):
        pasta = Path(atual)
        for nome in arquivos:
            if not nome.lower().endswith(".pdf"):
                continue
            try:
                c = normalizar_carimbo(Path(nome).stem)
            except Exception:
                continue
            if c == carimbo_norm:
                return pasta / nome
    return None


def ocr_para_carimbo(raiz: Path, carimbo_norm: str) -> dict[str, Any]:
    pdf = localizar_pdf(raiz, carimbo_norm)
    if not pdf:
        return {"_erro": f"PDF nao encontrado para BB_{carimbo_norm}"}
    try:
        return ocr_eq.processar_pdf(str(pdf), "mt")
    except Exception as exc:
        return {"_erro": f"{type(exc).__name__}: {exc}"}


def ocr_campo_float(dados: dict[str, Any], campo: str) -> float | None:
    v = dados.get(campo)
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", "."))
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Comparação
# ---------------------------------------------------------------------------

TOLERANCIA_RELATIVA = 0.005  # 0,5 % de tolerância para arredondamentos


def _comparar(ocr_v: float | None, consen_raw: str) -> tuple[str, str, bool, float | None]:
    """
    Retorna (ocr_str, consen_str, match, delta_abs).
    match=True se ambos são 0/vazio, ou se diferença ≤ TOLERANCIA_RELATIVA.
    """
    ocr_str = f"{ocr_v:.2f}" if ocr_v is not None else ""
    consen_f = _br_to_float(consen_raw)
    consen_str = f"{consen_f:.2f}" if consen_f is not None else consen_raw

    if ocr_v is None and not consen_raw:
        return ocr_str, consen_str, True, None
    if ocr_v is None or consen_f is None:
        return ocr_str, consen_str, False, None

    delta = abs(ocr_v - consen_f)
    ref = max(abs(ocr_v), abs(consen_f), 1e-9)
    match = (delta / ref) <= TOLERANCIA_RELATIVA
    return ocr_str, consen_str, match, round(delta, 4)


def comparar_carimbo(
    carimbo_norm: str,
    dados_ocr: dict[str, Any],
    dados_consen: dict[str, str],
) -> list[dict[str, Any]]:
    linhas = []
    for campo in CAMPOS_ORDEM:
        ocr_v = ocr_campo_float(dados_ocr, campo)
        consen_raw = dados_consen.get(campo, "")
        ocr_str, consen_str, match, delta = _comparar(ocr_v, consen_raw)

        # Só inclui campos onde ao menos um lado tem valor
        if not ocr_str and not consen_str:
            continue

        linhas.append({
            "carimbo":      f"BB_{carimbo_norm}",
            "campo":        campo,
            "ocr":          ocr_str,
            "consen":       consen_str,
            "match":        "SIM" if match else "NAO",
            "delta":        str(delta) if delta is not None else "",
        })
    return linhas


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _carregar_lista(args) -> list[str]:
    carimbos: list[str] = []
    for c in (args.carimbo or []):
        try:
            carimbos.append(normalizar_carimbo(c))
        except Exception:
            pass
    if args.carimbos_arquivo:
        for line in Path(args.carimbos_arquivo).read_text(encoding="utf-8-sig").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                try:
                    carimbos.append(normalizar_carimbo(line))
                except Exception:
                    pass
    if args.xlsx:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = next((i for i, h in enumerate(headers) if "carimbo" in h.lower()), None)
            if idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    v = str(row[idx] or "").strip()
                    if v:
                        try:
                            carimbos.append(normalizar_carimbo(v))
                        except Exception:
                            pass
            wb.close()
        except Exception as exc:
            warn(f"Falha ao ler xlsx: {exc}")
    return list(dict.fromkeys(carimbos))  # deduplica mantendo ordem


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Auditoria OCR vs Consen — Equatorial GO MT")
    p.add_argument("--carimbo", action="append", default=[])
    p.add_argument("--carimbos-arquivo", default="")
    p.add_argument("--xlsx", default="", help="xlsx MT para extrair lista de carimbos")
    p.add_argument("--raiz-pdfs", default=str(DEFAULT_PDFS_ROOT))
    p.add_argument("--saida-dir", default=str(DEFAULT_SAIDA_DIR))
    p.add_argument("--so-ocr", action="store_true", help="Apenas roda OCR, sem abrir Consen")
    p.add_argument("--limite", type=int, default=0)
    return p.parse_args()


def main() -> int:
    args = parse_args()
    carimbos = _carregar_lista(args)
    if not carimbos:
        print("Informe carimbos via --carimbo, --carimbos-arquivo ou --xlsx.")
        return 2

    if args.limite > 0:
        carimbos = carimbos[: args.limite]

    saida_dir = Path(args.saida_dir)
    saida_dir.mkdir(parents=True, exist_ok=True)
    raiz = Path(args.raiz_pdfs)

    csv_path = saida_dir / "auditoria_ocr_vs_consen_mt.csv"
    resumo_path = saida_dir / "auditoria_resumo_mt.csv"

    todas_linhas: list[dict[str, Any]] = []
    resumo: list[dict[str, Any]] = []

    # ── Fase 1: OCR ─────────────────────────────────────────────────────────
    log(f"Fase 1 — OCR de {len(carimbos)} carimbo(s)...")
    ocr_cache: dict[str, dict[str, Any]] = {}
    for c in carimbos:
        log(f"  OCR BB_{c}")
        ocr_cache[c] = ocr_para_carimbo(raiz, c)

    if args.so_ocr:
        for c, dados in ocr_cache.items():
            erro = dados.get("_erro", "")
            log(f"  BB_{c}: {'ERRO: ' + erro if erro else 'OK'}")
        log("Modo --so-ocr: sem conexao com Consen.")
        return 0

    # ── Fase 2: Consen ───────────────────────────────────────────────────────
    log(f"Fase 2 — leitura de {len(carimbos)} fatura(s) no Consen...")
    driver = None
    try:
        driver, wait = abrir_driver_logado()
        time.sleep(2.0)

        for carimbo_norm in carimbos:
            log(f"  Consen BB_{carimbo_norm}")
            dados_ocr = ocr_cache.get(carimbo_norm, {})
            erro_ocr = dados_ocr.get("_erro", "")

            dados_consen: dict[str, str] = {}
            erro_consen = ""

            try:
                abrir_tela_edicao_carimbo(driver, wait, EDIT_URL)
                time.sleep(0.8)
                carregar_fatura_por_carimbo(driver, wait, carimbo_norm, CAMPOS_CRITICOS_TELA)
                time.sleep(0.8)
                campos_form = coletar_campos_visiveis(driver)
                dados_consen = ler_campos_consen(campos_form)
            except Exception as exc:
                erro_consen = f"{type(exc).__name__}: {str(exc)[:120]}"
                warn(f"  Erro Consen BB_{carimbo_norm}: {erro_consen}")

            linhas = comparar_carimbo(carimbo_norm, dados_ocr, dados_consen)
            todas_linhas.extend(linhas)

            match_count = sum(1 for l in linhas if l["match"] == "SIM")
            nao_match = sum(1 for l in linhas if l["match"] == "NAO")
            resumo.append({
                "carimbo":      f"BB_{carimbo_norm}",
                "campos_total": len(linhas),
                "match":        match_count,
                "divergencia":  nao_match,
                "erro_ocr":     erro_ocr,
                "erro_consen":  erro_consen,
            })

            log(f"    -> {match_count} OK / {nao_match} divergencia(s)")
            time.sleep(0.5)

    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    # ── Saída ────────────────────────────────────────────────────────────────
    _cols = ["carimbo", "campo", "ocr", "consen", "match", "delta"]
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=_cols, delimiter=";")
        w.writeheader()
        w.writerows(todas_linhas)

    _res_cols = ["carimbo", "campos_total", "match", "divergencia", "erro_ocr", "erro_consen"]
    with resumo_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=_res_cols, delimiter=";")
        w.writeheader()
        w.writerows(resumo)

    total_div = sum(r["divergencia"] for r in resumo)
    total_ok = sum(r["match"] for r in resumo)
    log(f"\nAuditoria concluida: {total_ok} campos OK / {total_div} divergencias")
    log(f"Detalhe : {csv_path}")
    log(f"Resumo  : {resumo_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
