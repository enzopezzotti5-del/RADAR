# -*- coding: utf-8 -*-
# pip install selenium openpyxl beautifulsoup4
# Consen_ENEL.py — digitação de faturas ENEL no sistema Consen

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import _venv_check  # noqa

import os
import subprocess
import shutil
import tempfile
import time
import json
import csv
import re
import unicodedata
from datetime import datetime, date

try:
    from dotenv import load_dotenv
    for _dotenv_path in (
        Path(__file__).resolve().parent.parent / ".env",
        Path(__file__).resolve().parent.parent.parent / ".env",
    ):
        if _dotenv_path.exists():
            load_dotenv(_dotenv_path)
except ImportError:
    pass

import openpyxl
from bs4 import BeautifulSoup
from urllib3.exceptions import MaxRetryError

from selenium import webdriver
from selenium.common.exceptions import (
    InvalidSessionIdException,
    SessionNotCreatedException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

try:
    from digitacao_consen.auditoria_schema import (
        append_resultado_auditoria,
        extrair_status_auditoria,
        upsert_resultado_auditoria,
    )
    from digitacao_consen.consen_credentials import resolver_credenciais_consen
except ModuleNotFoundError:
    from auditoria_schema import (  # type: ignore
        append_resultado_auditoria,
        extrair_status_auditoria,
        upsert_resultado_auditoria,
    )
    from consen_credentials import resolver_credenciais_consen  # type: ignore

try:
    from digitacao_consen.checkpoint import gravar_salvar_confirmado, salvar_confirmado
except ModuleNotFoundError:
    try:
        from checkpoint import gravar_salvar_confirmado, salvar_confirmado  # type: ignore
    except ModuleNotFoundError:
        def gravar_salvar_confirmado(*a, **k): pass  # type: ignore[misc]
        def salvar_confirmado(*a, **k): return False  # type: ignore[misc]


# =========================================================
# CONFIGURAÇÕES
# =========================================================

LOGIN_URL = os.environ.get("CONSEN_LOGIN_URL", "https://consen.acaoengenharia.com.br/login.php")
TARGET_HASH = os.environ.get("CONSEN_TARGET_HASH", "#bpg/gestao/fatura/cadastroTabFatura.php")
_BASE_URL = LOGIN_URL.rsplit("login.php", 1)[0]
TARGET_URL = os.environ.get("CONSEN_TARGET_URL", f"{_BASE_URL}index.php{TARGET_HASH}")
LINK_HREF = os.environ.get("CONSEN_LINK_HREF", "bpg/gestao/fatura/cadastroTabFatura.php")
LINK_TEXTO = os.environ.get("CONSEN_LINK_TEXTO", "Instalacao")

USUARIO, SENHA = resolver_credenciais_consen()
os.environ["CONSEN_USUARIO"] = USUARIO
os.environ["CONSEN_SENHA"] = SENHA

# Planilha gerada pelo ocr_enel.py
import os as _os_dig


class _ErroNaoRecuperavel(RuntimeError):
    """Erro de dados/validação que não deve ser repetido pelo retry do driver.

    Exemplos: campo obrigatório ausente na planilha, salvamento rejeitado pelo
    CONSEN com .is-invalid, arquivo PDF fora do padrão BB_.
    Quando esta exceção é capturada no loop principal, o registro é marcado
    como erro imediatamente sem aguardar as 3 tentativas.
    """
EXCEL_PATH = _os_dig.environ.get(
    "ENEL_EXCEL_PATH",
    "//10.10.250.21/Energia/ARQUIVOS ENZO/OCR ENEL/ocr_enel_BT_032026.xlsx"
)

PASTA_SAIDA = Path(
    _os_dig.environ.get(
        "CONSEN_PIPELINE_SAIDA",
        _os_dig.environ.get(
            "ENEL_PIPELINE_SAIDA",
            "//10.10.250.21/Energia/ARQUIVOS ENZO/ENEL_pipeline_saida",
        ),
    )
)
try:
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
except OSError:
    pass
PASTA_PDFS_ATUAIS = _os_dig.environ.get("ENEL_DIGITACAO_PASTA_PDFS", "").strip()
CONSEN_TRACE_CAMPOS = _os_dig.environ.get("CONSEN_TRACE_CAMPOS", "0") == "1"
CONSEN_TRACE_ARQUIVO = Path(
    _os_dig.environ.get("CONSEN_TRACE_ARQUIVO", str(PASTA_SAIDA / "trace_campos_consen.jsonl"))
)
INVESTIGAR_DIR = Path(
    _os_dig.environ.get(
        "ENEL_DIGITACAO_INVESTIGAR_DIR",
        "//10.10.250.21/Energia/CONTASDEENERGIAELETRICA/BB/ENZO/Investigar",
    )
)


def _trace_campos_evento(evento: str, **dados) -> None:
    """Registra diagnóstico opt-in de preenchimento/persistência de campos.

    Ativado apenas com CONSEN_TRACE_CAMPOS=1. Não interfere no fluxo normal.
    """
    if not CONSEN_TRACE_CAMPOS:
        return
    try:
        CONSEN_TRACE_ARQUIVO.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "evento": evento,
            **dados,
        }
        with CONSEN_TRACE_ARQUIVO.open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")
    except Exception as exc:
        warn(f"[TRACE_CAMPOS] Falha ao registrar evento {evento}: {type(exc).__name__}: {exc}")


def _valor_elemento_seguro(elemento) -> str:
    try:
        tag = (elemento.tag_name or "").lower()
        if tag == "select":
            return _texto_select_atual(elemento)
        return (elemento.get_attribute("value") or elemento.text or "").strip()
    except Exception:
        return ""
# Arquivo .txt com um nome de PDF por linha — restringe digitação a esses arquivos apenas.
# Gerado pelo resgate_direcionado.py para garantir que só os PDFs baixados nesta sessão sejam digitados.
ENEL_SOMENTE_NOMES_FILE = _os_dig.environ.get("ENEL_DIGITACAO_SOMENTE_NOMES", "").strip()
# Sobrescreve data_ref_esperada para todos os registros — usa "MM/YYYY" (ex: "03/2026").
# Útil para MT onde o OCR extrai a competência errada do PDF.
ENEL_DATA_REF_OVERRIDE = _os_dig.environ.get("ENEL_DATA_REF_OVERRIDE", "").strip()
CONSEN_PERMITIR_LOTE_COMPLETO = _os_dig.environ.get("CONSEN_PERMITIR_LOTE_COMPLETO", "0").strip().lower() in {
    "1",
    "true",
    "sim",
    "s",
    "yes",
    "y",
}
CONSEN_EDITAR_REFERENCIA_EXISTENTE = _os_dig.environ.get(
    "CONSEN_EDITAR_REFERENCIA_EXISTENTE", "0"
).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}
# Se True, ajusta datas de leitura que caem no último dia do mês para o 1º do mês seguinte.
# Ativar apenas para ENEL (define-se no pipeline_enel.py). Outras concs NÃO devem ajustar.
ENEL_AJUSTAR_LEITURA_ULTIMO_DIA = _os_dig.environ.get(
    "ENEL_AJUSTAR_LEITURA_ULTIMO_DIA", "0"
).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}

CONSEN_INTERATIVO_FECHAR = _os_dig.environ.get("CONSEN_INTERATIVO_FECHAR", "1").strip().lower() not in {
    "0",
    "false",
    "nao",
    "n",
}
CONSEN_REINICIAR_NAVEGADOR_CADA_LINHA = _os_dig.environ.get(
    "CONSEN_REINICIAR_NAVEGADOR_CADA_LINHA", "0"
).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}
CONSEN_USAR_PERFIL_TEMP = _os_dig.environ.get(
    "CONSEN_USAR_PERFIL_TEMP", "1"
).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}
CONSEN_RESETAR_SELENIUM_ANTES = _os_dig.environ.get(
    "CONSEN_RESETAR_SELENIUM_ANTES", "0"
).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}
CONSEN_USAR_CHROMEDRIVER_CACHE = _os_dig.environ.get(
    "CONSEN_USAR_CHROMEDRIVER_CACHE", "0"
).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}

# Timeout (segundos) para aguardar span.auditoria.sucesso após salvar.
# Reduzir para 3 agiliza contas que nunca passam na auditoria automática.
CONSEN_TIMEOUT_AUDITORIA = max(1, int(_os_dig.environ.get("CONSEN_TIMEOUT_AUDITORIA", "3")))

MAPEAMENTO_CSV = Path(r"C:\SEU\CAMINHO\mapeamento_campos_planilha.csv")
SCORE_MINIMO_MAPEAMENTO = 35
HEADLESS = False

# Linha Excel a partir da qual processar (1 = cabeçalho, 2 = primeiro dado)
LINHA_INICIO = 2

# Fator de velocidade: 1.0 = comportamento original, 0.5 = 2× mais rápido.
# Defina via env: DIGITACAO_FATOR_VELOCIDADE=0.5
# Default mais rapido para o fluxo atual, mantendo margem de estabilidade.
DIGITACAO_FATOR_VELOCIDADE = max(0.1, min(float(os.environ.get("DIGITACAO_FATOR_VELOCIDADE", "0.25")), 2.0))
CHROME_PROFILE_ROOT = Path(__file__).resolve().parents[2] / ".runtime" / "chrome_profiles"
CHROME_PROFILE_PREFIX = "consen_enel_"
CHROMEDRIVER_CACHE_ROOT = Path.home() / ".cache" / "selenium" / "chromedriver" / "win64"
CHROMEDRIVER_LOG_ROOT = Path(__file__).resolve().parents[2] / ".runtime" / "chromedriver_logs"
RE_ARQUIVO_BB_ESTRITO = re.compile(r"^BB_(\d{7})\.pdf$", re.IGNORECASE)


def _s(sec: float) -> None:
    """Sleep escalado por DIGITACAO_FATOR_VELOCIDADE. Mínimo absoluto de 20 ms."""
    time.sleep(max(0.02, sec * DIGITACAO_FATOR_VELOCIDADE))


def _smin(sec: float, minval: float) -> None:
    """Sleep escalado mas com piso mínimo (para esperas de rede/servidor)."""
    time.sleep(max(minval, sec * DIGITACAO_FATOR_VELOCIDADE))


# =========================================================
# MAPA DE CONVERSÃO PARA SELECTS
# =========================================================

MAP_SELECTS: dict[str, dict[str, str]] = {
    "cb-dados-contratuais-fatura-tarifa": {
        "1":            "HS - Verde",
        "2":            "HS - Azul",
        "Convencional": "Convencional",
        "Branca":       "Branca",
        "HS - Verde":   "HS - Verde",
        "HS - Azul":    "HS - Azul",
    },
    "cb-dados-contratuais-fatura-subgrupo": {
        "5":            "A4 [2,3kV a 25kV]",
        "B3 [<2,3kV]":  "B3 [<2,3kV]",
        "A4":           "A4 [2,3kV a 25kV]",
        "A4 [2,3 a 25 kV]": "A4 [2,3kV a 25kV]",
        "A4 [<13,8kV]": "A4 [2,3kV a 25kV]",
        "A3A":          "A3a [30kV a 44kV]",
        "A3a":          "A3a [30kV a 44kV]",
        "A3 [<44kV]":   "A3a [30kV a 44kV]",
        "A3":           "A3 [69 kV]",
        "A2":           "A2 [88 kV a 138 kV]",
        "A1":           "A1",
    },
    "cb-dados-financeiros-obs": {
        "1": "1", "6": "6", "7": "7", "8": "8", "10": "10", "11": "11",
        "12": "12", "13": "13", "14": "14", "16": "16", "17": "17", "18": "18",
        "23": "23", "34": "34", "35": "35", "36": "36", "44": "44", "47": "47",
        "48": "48", "49": "49", "51": "51", "54": "54", "55": "55", "56": "56",
        "57": "57", "58": "58", "59": "59", "60": "60", "63": "63", "64": "64",
        "65": "65", "67": "67", "68": "68", "69": "69", "70": "70", "71": "71",
        "72": "72", "73": "73", "74": "74", "75": "75", "76": "76", "77": "77",
        "78": "78", "79": "79", "80": "80", "81": "81", "82": "82", "83": "83",
        "84": "84", "86": "86", "87": "87", "88": "88", "90": "90", "91": "91",
        "92": "92", "93": "93", "94": "94", "95": "95", "96": "96", "97": "97",
        "98": "98", "99": "99", "100": "100", "101": "101", "102": "102",
        "103": "103", "108": "108", "109": "109", "110": "110", "111": "111",
        "113": "113", "114": "114", "115": "115", "116": "116", "119": "119",
        "123": "123", "124": "124", "125": "125", "126": "126", "129": "129",
        "130": "130", "131": "131", "132": "132", "133": "133", "134": "134",
        "135": "135", "136": "136", "137": "137", "139": "139", "140": "140",
        "141": "141", "142": "142", "143": "143", "144": "144", "145": "145",
        "146": "146", "147": "147", "148": "148", "149": "149", "150": "150",
        "151": "151", "152": "152", "154": "154", "155": "155", "156": "156",
        "157": "157", "158": "158", "159": "159", "160": "160", "161": "161",
        "162": "162", "163": "163", "164": "164", "165": "165", "166": "166",
        "167": "167", "168": "168", "169": "169", "170": "170", "171": "171",
        "172": "172", "173": "173", "174": "174", "175": "175", "176": "176",
        "177": "177", "178": "178", "180": "180", "181": "181", "182": "182",
        "183": "183", "184": "184", "185": "185", "186": "186", "187": "187",
        "188": "188", "189": "189", "190": "190", "191": "191", "192": "192",
        "193": "193", "194": "194", "195": "195", "196": "196", "197": "197",
        "198": "198", "200": "200", "201": "201", "204": "204", "206": "206",
        "208": "208", "213": "213", "219": "219", "220": "220", "222": "222",
        "223": "223", "224": "224", "225": "225", "229": "229", "230": "230",
        "231": "231", "234": "234", "235": "235", "236": "236", "237": "237",
        "242": "242", "244": "244", "245": "245", "246": "246", "247": "247",
        "248": "248", "249": "249", "250": "250", "251": "251", "252": "252",
        "253": "253", "254": "254", "255": "255", "256": "256", "257": "257",
        "258": "258", "259": "259", "260": "260", "261": "261", "262": "262",
        "263": "263", "264": "264", "265": "265", "266": "266", "267": "267",
        "268": "268", "269": "269", "270": "270", "271": "271", "272": "272",
        "273": "273", "274": "274", "275": "275", "276": "276", "277": "277",
        "278": "278", "279": "279", "280": "280", "281": "281", "282": "282",
        "283": "283", "284": "284", "285": "285", "286": "286", "287": "287",
        "288": "288", "289": "289", "290": "290", "291": "291", "292": "292",
        "293": "293", "294": "294", "295": "295", "296": "296", "297": "297",
        "298": "298", "299": "299", "300": "300", "301": "301", "302": "302",
    },
}


# =========================================================
# ALIASES FORÇADOS
# =========================================================

ALIASES_FORCADOS = {
    # ── Datas
    "dataVencimento": "fatDataVcto",
    "dataEmissao":    "fatDataEmissao",

    # ── Selects
    "cb-dados-contratuais-fatura-tarifa":   "cadTarifaCod",
    "cb-dados-contratuais-fatura-subgrupo": "cadSubGrupoCod",

    # ── CONSUMO
    "txt-consumo-registrada-pta":    "fatConPontaRegistrado",
    "txt-consumo-faturada-pta":      "fatConPontaFaturado",
    "txt-consumo-pta-valor-reais":   "fatConPontaValorReais",
    "txt-consumo-registrada-fpind":  "fatConFPontaIndRegistrado",
    "txt-consumo-faturada-fpind":    "fatConFPontaIndFaturado",
    "txt-consumo-fpind-valor-reais": "fatConFPontaIndValorReais",
    "txt-consumo-registrada-inter":  "fatConIntermediarioRegistrado",
    "txt-consumo-faturada-inter":    "fatConIntermediarioFaturado",
    "txt-consumo-inter-valor-reais": "fatConIntermediarioValorReais",

    # ── CONSUMO EXCEDENTE REATIVO (COPEL MT)
    "txt-consumo-excedente-registrada-pta":    "fatConPontaExcRegistrado",
    "txt-consumo-excedente-faturada-pta":      "fatConPontaExcFaturado",
    "txt-consumo-excedente-pta-valor-reais":   "fatConPontaExcValorReais",
    "txt-consumo-excedente-registrada-fpind":  "fatConFPontaIndExcRegistrado",
    "txt-consumo-excedente-faturada-fpind":    "fatConFPontaIndExcFaturado",
    "txt-consumo-excedente-fpind-valor-reais": "fatConFPontaIndExcValorReais",
    # Aliases diretos pelo id HTML (CELESC MT energia reativa / UFER)
    "fatConFPontaIndExcRegistrado":  "fatConFPontaIndExcRegistrado",
    "fatConFPontaIndExc":            "fatConFPontaIndExcFaturado",
    "fatConFPontaIndExcValorReais":  "fatConFPontaIndExcValorReais",

    # ── DEMANDAS
    "txt-demandas-registrada-pta":    "fatDemPontaRegistrada",
    "txt-demandas-faturada-pta":      "fatDemPontaFaturada",
    "txt-demandas-pta-valor-reais":   "fatDemPontaValorReais",
    "txt-demandas-registrada-fpind":  "fatDemFPontaIndRegistrada",
    "txt-demandas-faturada-fpind":    "fatDemFPontaIndFaturada",
    "txt-demandas-fpind-valor-reais": "fatDemFPontaIndValorReais",
    "txt-demandas-ultrapassagem-faturada-fpind": "fatDemFPontaIndUltra",
    "txt-demandas-ultrapassagem-fpind-valor-reais": "fatDemFPontaIndUltraValorReais",
    "txt-demandas-ultrapassagem-faturada-pta": "fatDemPontaUltra",
    "txt-demandas-ultrapassagem-pta-valor-reais": "fatDemPontaUltraValorReais",

    # ── DEMANDAS CONTRATADAS
    "txt-dados-contratuais-fatura-dem-cont-p":          "fatDemContratadaPonta",
    "txt-dados-contratuais-fatura-dem-cont-fp":         "fatDemContratadaFPonta",
    "txt-dados-contratuais-fatura-dem-cont-geracao-p":  "fatDemContratadaGeracaoPonta",
    "txt-dados-contratuais-fatura-dem-cont-geracao-fp": "fatDemContratadaGeracaoFPonta",

    # ── INJETADO / GD
    "txt-consumo-injetado-registrado-fpta":  "fatConFPontaInjetadoRegistrado",
    "txt-consumo-injetado-faturado-fpta":    "fatConFPontaInjetadoFaturado",
    "txt-consumo-injetado-fpta-valor-reais": "fatConFPontaInjetadoValorReais",
    "txt-consumo-injetado-usina-fpta":       "fatConFPontaInjetadoUsina",
    "txt-consumo-injetado-usina-pta":        "fatConPontaInjetadoUsina",

    # ── SALDOS USINA
    "txt-consumo-injetado-usina-fpta-saldo": "fatConFPontaInjetadoUsinaSaldoAcumulado",
    "txt-consumo-injetado-usina-pta-saldo":  "fatConPontaInjetadoUsinaSaldoAcumulado",

    # ── CÓDIGO DE BARRAS
    "fatCodigoBarras":                    "fatCodigoBarras",
    "txt-dados-financeiros-codigo-barra": "fatCodigoBarras",

    # ── FINANCEIROS — alíquotas (%) e valores monetários separados
    "camposFinanIlumimnacaoPublica":                   "fatIlumPublica",
    "txt-dados-financeiros-iluminacao-publica":        "fatIlumPublica",
    # ICMS
    "camposFinanICMS":                                 "fatICMS",           # valor ICMS R$
    "fatDesIcmsAliquota":                              "fatDesIcmsAliquota", # alíquota %
    "fatICMS":                                         "fatICMS",            # valor R$
    # PIS
    "fatDescPisAliquota":                              "fatDescPisAliquota",
    "txt-dados-financeiros-pis-pasep":                 "fatPIS",
    # COFINS
    "fatDesCofinsAliquota":                            "fatDesCofinsAliquota",
    "txt-dados-financeiros-cofins":                    "fatCOFINS",
    # Retenção PIS/COFINS — bloqueados por padrão para ENEL BT (sem retenção).
    # Para COPEL B3 a lógica em preencher_aliases_forcados_linha() usa os valores reais do xlsx.
    "fatDescPisPercRetImposto":                        "_zero",
    "fatDescPisValRetImposto":                         "_zero",
    "fatDescCofinsPercRetImposto":                     "_zero",
    "fatDescCofinsValRetImposto":                      "_zero",
    # Retenção CSLL/IRPJ — preenchidos do xlsx; cooperativas (PIS=COF=0) zeramos via lógica abaixo
    "fatDescCsllPercRetImposto":                       "fatDescCsllPercRetImposto",
    "fatDescCsllValRetImposto":                        "fatDescCsllValRetImposto",
    "fatDescIrpjPercRetImposto":                       "fatDescIrpjPercRetImposto",
    "fatDescIrpjValRetImposto":                        "fatDescIrpjValRetImposto",
    # Retenção consolidada em consumo — AMBAR_AM (Amazonas): CONSEN usa campo único em vez do desdobramento
    "fatDescConsumoPercRetImposto":                    "fatDescConsumoPercRetImposto",
    "fatDescConsumoValRetImposto":                     "fatDescConsumoValRetImposto",
    # Nota fiscal e fatura
    "txt-dados-financeiros-valor-nota-fiscal":        "fatValorNotaFiscal",
    "txt-dados-financeiros-valor-fatura-a-pagar":     "fatValorFatura",

    # ── FINANCEIROS EXTRAS (Equatorial GO MT e outros)
    "fatEscassezHidrica":           "fatEscassezHidrica",
    "fatEscassezHidricaValorReais": "fatEscassezHidricaValorReais",
    "fatDescontoFio":               "fatDescontoFio",
    "fatDescontoFioKWh":            "fatDescontoFioKWh",
    "fatMultas":                    "fatMultas",
    # ── UFER / Reativo Excedente F. Ponta (ENEL SP MT) — IndExc já mapeado acima via CELESC
    "fatConFPontaIndExcRegistrado": "fatConFPontaIndExcRegistrado",
    "fatConFPontaIndExc":           "fatConFPontaIndExcFaturado",
    "fatConFPontaIndExcValorReais": "fatConFPontaIndExcValorReais",
}

# Campos de retenção CSLL/IRPJ que devem ser zerados em cooperativas (PIS=COFINS=0)
_CAMPOS_RET_CSLL_IRPJ = frozenset({
    "fatDescCsllPercRetImposto", "fatDescCsllValRetImposto",
    "fatDescIrpjPercRetImposto", "fatDescIrpjValRetImposto",
})

_OBS_PARES_MAX = 5

CAMPOS_BLOQUEADOS_CSV = {
    "txt-consumo-excedente-registrada-fpcap",
    "txt-consumo-excedente-faturada-fpcap",
    "txt-consumo-registrada-fpcap",
    "txt-consumo-faturada-fpcap",
    "txt-demandas-geracao-registrada-fpta",
    "txt-demandas-geracao-faturada-fpta",
    "txt-demandas-excedente-registrada-fpta",
    "txt-demandas-excedente-faturada-fpta",
    "txt-demandas-registrada-fpcap",
    "txt-demandas-ultrapassagem-faturada-fpind",
    "txt-demandas-ultrapassagem-faturada-pta",
    "instalacaoMedidor",
    "enderecoUnidade",
    "enderecoMedidor",
    "cidadeUnidade",
    "cidadeMedidor",
}

TOKENS_CAMPOS_BLOQUEADOS_GENERICOS = (
    "subvenc",
    "subven",
    "beneficio",
)

# Headers da planilha que NUNCA devem ser digitados no Consen ENEL.
HEADERS_BLOQUEADOS_PLANILHA = {}
HEADERS_BLOQUEADOS_PLANILHA_POR_CONCESSIONARIA = {
    "COPEL": {"fatTributoFederalPerc", "fatTributoFederalVal"},
    "EQUATORIAL": {"fatTributoFederalPerc", "fatTributoFederalVal"},
}


_CAMPOS_TRIB_FEDERAL = frozenset({"fatTributoFederalPerc", "fatTributoFederalVal"})
_CAMPOS_RETENCAO_INDIVIDUAL = frozenset({
    "fatDescPisPercRetImposto",
    "fatDescPisValRetImposto",
    "fatDescCofinsPercRetImposto",
    "fatDescCofinsValRetImposto",
    "fatDescCsllPercRetImposto",
    "fatDescCsllValRetImposto",
    "fatDescIrpjPercRetImposto",
    "fatDescIrpjValRetImposto",
})
_CONCS_NEO_TRIBUTO_FEDERAL = frozenset({"COELBA", "CELPE", "COSERN", "ELEKTRO"})


def _usa_somente_tributo_federal(dados_planilha: dict) -> bool:
    """Neoenergia MT e ENEL MT usam o total de Tributo Federal em vez do desdobramento."""
    conc = str(dados_planilha.get("concCod") or "").strip().upper()
    subgrupo = str(dados_planilha.get("cadSubGrupoCod") or "").strip().upper()
    # ENEL MT: concCod="3", subgrupo comeca com "A" (A4, A3a, etc.)
    if conc == "3" and subgrupo.startswith("A"):
        return True
    if conc not in _CONCS_NEO_TRIBUTO_FEDERAL:
        return False
    try:
        trib_perc = abs(float(dados_planilha.get("fatTributoFederalPerc") or 0))
        trib_val = abs(float(dados_planilha.get("fatTributoFederalVal") or 0))
    except (ValueError, TypeError):
        trib_perc = 0.0
        trib_val = 0.0
    return trib_perc > 0 or trib_val > 0


def _tem_retencoes_individuais(dados_planilha: dict) -> bool:
    """True se o xlsx tem retenções individuais com valor não-nulo/zero."""
    for campo in ("fatDescCsllValRetImposto", "fatDescIrpjValRetImposto",
                  "fatDescPisValRetImposto", "fatDescCofinsValRetImposto"):
        val = dados_planilha.get(campo)
        if val is None:
            continue
        try:
            if abs(float(str(val).replace(",", ".").replace(" ", ""))) > 0:
                return True
        except (ValueError, TypeError):
            pass
    return False


def _tem_ret_pis_cof_individuais(dados_planilha: dict) -> bool:
    """True se o xlsx tem valor não-zero de retenção individual de PIS ou COFINS."""
    for campo in ("fatDescPisValRetImposto", "fatDescCofinsValRetImposto"):
        val = dados_planilha.get(campo)
        if val is None:
            continue
        try:
            if abs(float(str(val).replace(",", ".").replace(" ", ""))) > 0:
                return True
        except (ValueError, TypeError):
            pass
    return False


def header_bloqueado_planilha(header: str, dados_planilha: dict | None = None) -> bool:
    if header in HEADERS_BLOQUEADOS_PLANILHA:
        return True
    if not dados_planilha:
        return False

    conc = str(dados_planilha.get("concCod") or "").strip().upper()
    arquivo = str(dados_planilha.get("ARQUIVO") or "").strip().upper()
    if not conc and "COPEL" in arquivo:
        conc = "COPEL"

    bloqueados = HEADERS_BLOQUEADOS_PLANILHA_POR_CONCESSIONARIA.get(conc, set())
    if header in bloqueados:
        return True

    # Regra geral: se há retenções individuais no xlsx, bloquear TributoFederal total
    if (
        header in _CAMPOS_TRIB_FEDERAL
        and _tem_retencoes_individuais(dados_planilha)
        and not _usa_somente_tributo_federal(dados_planilha)
    ):
        return True

    return False


def campo_bloqueado_generico(*nomes: str) -> bool:
    for nome in nomes:
        normalizado = normalizar_slug(nome or "")
        if not normalizado:
            continue
        if any(token in normalizado for token in TOKENS_CAMPOS_BLOQUEADOS_GENERICOS):
            return True
    return False


def permitir_campo_beneficio_copel(header: str, dados_planilha: dict | None = None) -> bool:
    if header not in {
        "fatBeneficioTarifarioBrutoValorReais",
        "fatBeneficioLiquidoValorReais",
    }:
        return False
    if not dados_planilha:
        return False

    conc = str(dados_planilha.get("concCod") or "").strip().upper()
    arquivo = str(dados_planilha.get("ARQUIVO") or "").strip().upper()
    if not conc and "COPEL" in arquivo:
        conc = "COPEL"
    if not conc and "EQUATORIAL" in arquivo:
        conc = "EQUATORIAL"
    if conc in {"COPEL", "EQUATORIAL"}:
        return True
    # ENEL MT (A4) — subgrupo contém "A4"
    subgrupo = str(dados_planilha.get("cadSubGrupoCod") or "").strip().upper()
    return "A4" in subgrupo


# =========================================================
# LOG
# =========================================================

def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [INFO] {msg}")


def warn(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [WARN] {msg}")


def erro(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [ERRO] {msg}")


# =========================================================
# UTILITÁRIOS
# =========================================================

def preencher_saldo_usina_tabela(driver, wait, campo_id, valor):
    if valor is None or str(valor).strip() == "":
        warn(f"[USINA TABELA] Valor vazio para {campo_id}")
        return False

    valor_fmt = formatar_numero_br(valor)

    try:
        campo = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.ID, campo_id))
        )
    except Exception:
        warn(f"[USINA TABELA] Campo não encontrado: {campo_id}")
        return False

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
        _s(0.3)

        driver.execute_script("arguments[0].focus();", campo)
        _s(0.1)

        try:
            campo.click()
            _s(0.1)
        except Exception:
            pass

        try:
            campo.send_keys(Keys.CONTROL, "a")
            campo.send_keys(Keys.DELETE)
        except Exception:
            pass

        driver.execute_script("""
            arguments[0].value = '';
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
        """, campo)

        _s(0.15)

        campo.send_keys(valor_fmt)
        _s(0.15)

        driver.execute_script("""
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new KeyboardEvent('keyup', {bubbles:true, key:'0'}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));
        """, campo)

        try:
            campo.send_keys(Keys.TAB)
        except Exception:
            pass

        _s(0.3)

        valor_final = (campo.get_attribute("value") or "").strip()
        log(f"[USINA TABELA] {campo_id} | enviado={valor_fmt!r} | final={valor_final!r}")

        vf = valor_final.replace(".", "").replace(",", "")
        ve = valor_fmt.replace(".", "").replace(",", "")
        return vf == ve

    except Exception as e:
        warn(f"[USINA TABELA] Erro ao preencher {campo_id}: {type(e).__name__} - {e}")
        return False

def normalizar_texto(txt):
    if txt is None:
        return ""
    txt = str(txt).strip()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return txt.lower().strip()


def normalizar_slug(txt):
    txt = normalizar_texto(txt)
    txt = re.sub(r"[^a-z0-9]+", "-", txt)
    txt = re.sub(r"-+", "-", txt).strip("-")
    return txt


def tokenizar(txt):
    slug = normalizar_slug(txt)
    return set(slug.split("-")) if slug else set()


def _cleanup_profile_dir(profile_dir: Path) -> None:
    try:
        shutil.rmtree(profile_dir, ignore_errors=True)
    except Exception:
        pass


def _cleanup_stale_temp_profiles(max_age_hours: int = 12) -> None:
    CHROME_PROFILE_ROOT.mkdir(parents=True, exist_ok=True)
    limite = time.time() - (max_age_hours * 3600)
    removidos = 0

    for pasta in CHROME_PROFILE_ROOT.glob(f"{CHROME_PROFILE_PREFIX}*"):
        try:
            if not pasta.is_dir():
                continue
            if pasta.stat().st_mtime > limite:
                continue
            shutil.rmtree(pasta, ignore_errors=True)
            removidos += 1
        except Exception:
            continue

    if removidos:
        log(f"[driver] {removidos} perfil(is) temporario(s) antigo(s) removido(s).")


def _novo_log_chromedriver() -> Path:
    CHROMEDRIVER_LOG_ROOT.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return CHROMEDRIVER_LOG_ROOT / f"chromedriver_{stamp}.log"


def _encerrar_selenium_orfao() -> None:
    """Encerra apenas processos órfãos do Selenium/ChromeDriver.

    Mantém o Chrome normal do usuário intacto e mira só processos de automação
    identificados por command-line de WebDriver. Também limpa diretórios
    scoped_dir* no %TEMP% que ChromeDriver deixa quando Chrome crasha.
    """
    script = r"""
$ErrorActionPreference = 'SilentlyContinue'
$targets = Get-CimInstance Win32_Process | Where-Object {
  ($_.Name -eq 'chromedriver.exe') -or
  ($_.Name -eq 'chrome.exe' -and $_.CommandLine -match '--test-type=webdriver|--remote-debugging-port=0|data:,')
}
$killed = @()
foreach($proc in $targets){
  try {
    Stop-Process -Id $proc.ProcessId -Force -ErrorAction SilentlyContinue
    $killed += ($proc.Name + ':' + $proc.ProcessId)
  } catch {}
}
if($killed.Count -gt 0){
  $killed -join ','
}
"""
    try:
        resultado = subprocess.run(
            ["powershell", "-NoProfile", "-Command", script],
            capture_output=True,
            text=True,
            timeout=15,
            creationflags=0x08000000,
        )
        saida = (resultado.stdout or "").strip()
        if saida:
            log(f"[driver] processos Selenium encerrados antes da inicializacao: {saida}")
    except Exception as exc:
        warn(f"[driver] nao foi possivel limpar processos Selenium orfaos: {type(exc).__name__}")

    # Limpa scoped_dir* residuais no %TEMP% — ChromeDriver não os remove quando Chrome crasha,
    # e esses diretórios orphaned impedem o próximo Chrome de inicializar corretamente.
    try:
        import tempfile as _tf
        import shutil as _sh
        _tmp = _os_dig.environ.get("TEMP") or _tf.gettempdir()
        _removidos = 0
        for _d in _os_dig.scandir(_tmp):
            if _d.name.startswith("scoped_dir") and _d.is_dir(follow_symlinks=False):
                try:
                    _sh.rmtree(_d.path, ignore_errors=True)
                    _removidos += 1
                except Exception:
                    pass
        if _removidos:
            log(f"[driver] {_removidos} scoped_dir(s) residual(is) removido(s) de {_tmp}")
    except Exception as exc:
        warn(f"[driver] nao foi possivel limpar scoped_dirs: {type(exc).__name__}")

    # Remove SingletonLock do perfil padrão do Chrome (evita trava após crash)
    try:
        _local_app = _os_dig.environ.get("LOCALAPPDATA", "")
        if _local_app:
            _lock = _os_dig.path.join(_local_app, "Google", "Chrome", "User Data", "SingletonLock")
            if _os_dig.path.exists(_lock):
                _os_dig.remove(_lock)
                log("[driver] SingletonLock do perfil padrao removido.")
    except Exception:
        pass


def _versao_tuple(texto: str) -> tuple[int, ...]:
    nums = re.findall(r"\d+", str(texto or ""))
    return tuple(int(n) for n in nums) if nums else (0,)


def _chrome_version() -> str:
    candidatos = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
    ]
    for chrome_exe in candidatos:
        if not chrome_exe.exists():
            continue
        try:
            cmd = [
                "powershell",
                "-NoProfile",
                "-Command",
                f"(Get-Item '{chrome_exe}').VersionInfo.ProductVersion",
            ]
            resultado = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=5,
                creationflags=0x08000000,
            )
            versao = (resultado.stdout or "").strip()
            if versao:
                return versao
        except Exception:
            continue
    return ""


def _find_cached_chromedriver() -> str | None:
    """Retorna o chromedriver em cache para a versão do Chrome instalado.

    Prioridade:
    1. Match exato de versão (ex: 150.0.7871.101 == 150.0.7871.101)
    2. Mesma major version (ex: driver 150.0.7871.115 serve Chrome 150.0.7871.101)
       — patch mismatch dentro da mesma major é seguro e evita falha de download
       no contexto do watcher/Task Scheduler onde o Selenium Manager pode não
       conseguir baixar um novo driver.

    Retorna None só se não houver nenhum driver da mesma major version em cache,
    deixando o Selenium Manager tentar o download.
    """
    if not CHROMEDRIVER_CACHE_ROOT.exists():
        return None

    chrome_ver = _chrome_version()
    if not chrome_ver:
        return None

    chrome_major = chrome_ver.split(".")[0]

    try:
        candidatos = sorted(
            [p for p in CHROMEDRIVER_CACHE_ROOT.iterdir() if p.is_dir()],
            key=lambda p: _versao_tuple(p.name),
            reverse=True,
        )
    except Exception:
        return None

    fallback_mesmo_major: str | None = None
    for pasta in candidatos:
        exe = pasta / "chromedriver.exe"
        if not exe.exists():
            continue
        if pasta.name == chrome_ver:
            return str(exe)
        if pasta.name.split(".")[0] == chrome_major and fallback_mesmo_major is None:
            fallback_mesmo_major = str(exe)

    if fallback_mesmo_major:
        log(f"[driver] sem match exato para Chrome {chrome_ver} — usando driver de mesma major: {fallback_mesmo_major}")
    return fallback_mesmo_major


def _corrigir_se_metadata() -> None:
    """Remove se-metadata.json se o driver cacheado não corresponde à versão atual do Chrome."""
    import json as _json
    meta = Path.home() / ".cache" / "selenium" / "se-metadata.json"
    if not meta.exists():
        return
    chrome_ver = _chrome_version()
    if not chrome_ver:
        return
    try:
        data = _json.loads(meta.read_text(encoding="utf-8"))
        for d in data.get("drivers", []):
            if d.get("driver_name") == "chromedriver" and d.get("driver_version", "") != chrome_ver:
                log(f"[driver] se-metadata.json desatualizado ({d.get('driver_version')} ≠ {chrome_ver}) — removendo")
                meta.unlink(missing_ok=True)
                return
    except Exception:
        try:
            meta.unlink(missing_ok=True)
        except Exception:
            pass


def _erro_driver_recuperavel(exc: Exception) -> bool:
    if isinstance(exc, (InvalidSessionIdException, SessionNotCreatedException)):
        return True
    if isinstance(exc, MaxRetryError):
        return True
    if not isinstance(exc, WebDriverException):
        return False
    msg = str(exc).lower()
    sinais = (
        "invalid session id",
        "session deleted",
        "disconnected",
        "not connected to devtools",
        "chrome not reachable",
        "target window already closed",
        "devtoolsactiveport",
        "failed to create chrome process",
        "timed out receiving message from renderer",
        "unable to connect to renderer",
        "connection refused",
    )
    return any(sinal in msg for sinal in sinais)


def iniciar_driver(headless=False):
    if CONSEN_RESETAR_SELENIUM_ANTES:
        _encerrar_selenium_orfao()

    ultimo_erro = None
    ultimo_log_driver: Path | None = None
    cached_driver = _find_cached_chromedriver()
    if not cached_driver:
        _corrigir_se_metadata()

    for tentativa in range(1, 5):
        options = Options()
        options.add_argument("--start-maximized")
        options.add_argument("--window-size=1600,1000")

        driver_log_path = _novo_log_chromedriver()
        ultimo_log_driver = driver_log_path

        service_kwargs = {
            "log_output": str(driver_log_path),
            "service_args": ["--verbose"],
        }

        try:
            log(f"[driver] tentativa {tentativa}/4 iniciando Chrome")
            if cached_driver:
                log(f"[driver] usando ChromeDriver em cache: {cached_driver}")
                driver = webdriver.Chrome(
                    service=Service(executable_path=cached_driver, **service_kwargs),
                    options=options,
                )
            else:
                driver = webdriver.Chrome(service=Service(**service_kwargs), options=options)
            driver._codex_driver_log = str(driver_log_path)
            driver.set_page_load_timeout(60)
            return driver
        except (SessionNotCreatedException, WebDriverException) as exc:
            ultimo_erro = exc
            warn(
                f"[driver] falha ao iniciar Chrome (tentativa {tentativa}/4): "
                f"{type(exc).__name__}: {exc}"
            )
            warn(f"[driver] log verboso salvo em: {driver_log_path}")
            time.sleep(min(1.0 * tentativa, 3.0))

    if ultimo_log_driver is not None:
        warn(f"[driver] ultima tentativa registrada em: {ultimo_log_driver}")
    raise ultimo_erro or RuntimeError("Falha ao iniciar o Chrome.")


def formatar_ddmmyyyy(dt):
    return dt.strftime("%d/%m/%Y")


def formatar_yyyy_mm_dd(dt):
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d")


def parse_data_ddmmyyyy(txt):
    return datetime.strptime(txt.strip(), "%d/%m/%Y").date()


def primeiro_dia_mes(dt):
    return date(dt.year, dt.month, 1)


def _ajustar_leitura_ultimo_dia(dt_val: "date | None") -> "date | None":
    """Se a data de leitura cai no último dia do mês, avança para o dia 1 do mês seguinte.
    ENEL às vezes registra a leitura no último dia do mês, mas o CONSEN exige o 1º do próximo."""
    if dt_val is None:
        return None
    import calendar as _cal
    ultimo_dia = _cal.monthrange(dt_val.year, dt_val.month)[1]
    if dt_val.day == ultimo_dia:
        if dt_val.month == 12:
            return date(dt_val.year + 1, 1, 1)
        return date(dt_val.year, dt_val.month + 1, 1)
    return dt_val


def _ref_mmaaaa_para_date(valor) -> "date | None":
    """Converte 'MM/YYYY' (campo fatDataReferencia do OCR) para o 1º dia daquele mês.
    Também aceita date/datetime objects diretamente."""
    if not valor:
        return None
    if isinstance(valor, datetime):
        return date(valor.year, valor.month, 1)
    if isinstance(valor, date):
        return date(valor.year, valor.month, 1)
    txt = str(valor).strip()
    try:
        return datetime.strptime(txt, "%m/%Y").date()
    except ValueError:
        pass
    # Tenta parsear como data completa e extrair mês/ano
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(txt[:10], fmt)
            return date(dt.year, dt.month, 1)
        except ValueError:
            continue
    return None


def valor_excel_para_date(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        if float(valor) == 0:
            return None
    if isinstance(valor, str):
        valor = valor.strip()
        if valor in ("", "0", "0,0", "0,00", "None", "nan"):
            return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(valor, fmt).date()
            except ValueError:
                pass
    raise Exception(f"Não consegui converter para data: {valor!r}")


def valor_para_numero(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    s = str(valor).strip()
    if not s:
        return None

    s = s.replace("R$", "").replace(" ", "")
    # Formato brasileiro: tem vírgula E ponto → ponto é milhar, vírgula é decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    # Só vírgula → vírgula é decimal
    elif "," in s:
        s = s.replace(",", ".")
    # Só ponto: se exatamente 3 dígitos após o ponto → milhar BR (ex: "9.820" → 9820)
    # caso contrário decimal americano (ex: "112.56" → 112.56)
    elif "." in s:
        partes = s.split(".")
        if len(partes) == 2 and len(partes[1]) == 3 and partes[1].isdigit():
            s = s.replace(".", "")

    try:
        return float(s)
    except ValueError:
        return None


def formatar_numero_br(valor):
    num = valor_para_numero(valor)
    if num is None:
        return str(valor).strip()
    # SEM separador de milhar — o campo do Consen tem máscara que interpreta
    # o ponto como decimal se vier "5.956,26", gerando "595.626,00".
    # Formato correto: "5956,26"
    if abs(num - round(num)) < 1e-9:
        return f"{int(round(num))},00"
    return f"{num:.2f}".replace(".", ",")


def normalizar_carimbo(valor):
    txt = "" if valor is None else str(valor).strip().upper()
    txt = txt.replace("BB_", "").replace(".PDF", "")
    if txt.endswith(".0"):
        txt = txt[:-2]
    txt = re.sub(r"\D", "", txt)
    if len(txt) > 7:
        txt = txt[:7]
    return txt


def extrair_carimbo_estrito_do_arquivo(nome_arquivo: str) -> str:
    nome = Path(str(nome_arquivo or "").strip()).name
    match = RE_ARQUIVO_BB_ESTRITO.fullmatch(nome)
    if not match:
        return ""
    return match.group(1)


def _resolver_pdf_registro(registro: dict) -> Path | None:
    dados = registro.get("dados_completos", {}) or {}
    arquivo_raw = str(dados.get("ARQUIVO") or "").strip()
    if not arquivo_raw:
        return None

    arquivo_path = Path(arquivo_raw)
    if arquivo_path.exists():
        return arquivo_path

    nome = arquivo_path.name
    if not nome or not PASTA_PDFS_ATUAIS:
        return None

    pasta = Path(PASTA_PDFS_ATUAIS)
    if not pasta.exists():
        return None

    candidato_direto = pasta / nome
    if candidato_direto.exists():
        return candidato_direto

    try:
        return next((p for p in pasta.rglob(nome) if p.is_file()), None)
    except Exception:
        return None


def mover_pdf_para_investigar(pdf_path: Path, motivo: str) -> str:
    motivo = str(motivo or "").strip() or "arquivo invalido"
    destino_dir = INVESTIGAR_DIR / "arquivo_sem_bb"
    try:
        destino_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return f"{motivo} | falha ao preparar Investigar: {exc}"

    destino = destino_dir / pdf_path.name
    if destino.exists():
        destino = destino_dir / f"{pdf_path.stem}_{int(time.time())}{pdf_path.suffix}"

    try:
        shutil.move(str(pdf_path), str(destino))
        sidecar = destino.with_name(f"{destino.name}.motivo.txt")
        sidecar.write_text(
            f"motivo={motivo}\norigem={pdf_path}\ndestino={destino}\n",
            encoding="utf-8",
        )
        return f"{motivo} | movido_para={destino}"
    except Exception as exc:
        return f"{motivo} | falha_ao_mover={exc}"


def validar_registro_bb_arquivo(registro: dict) -> tuple[bool, str]:
    dados = registro.get("dados_completos", {}) or {}
    arquivo_raw = str(dados.get("ARQUIVO") or "").strip()
    nome = Path(arquivo_raw).name if arquivo_raw else ""
    if not nome:
        return False, "ARQUIVO vazio na planilha"

    carimbo_arquivo = extrair_carimbo_estrito_do_arquivo(nome)
    if not carimbo_arquivo:
        return False, f"arquivo fora do padrao BB_<7 digitos>.pdf: {nome}"

    carimbo_planilha = normalizar_carimbo(registro.get("fatCarimbo", ""))
    if carimbo_planilha and carimbo_planilha != carimbo_arquivo:
        return (
            False,
            "fatCarimbo divergente do nome do arquivo: "
            f"planilha=BB_{carimbo_planilha} arquivo=BB_{carimbo_arquivo}.pdf",
        )

    registro["fatCarimbo"] = carimbo_arquivo
    dados["fatCarimbo"] = f"BB_{carimbo_arquivo}"
    registro["dados_completos"] = dados
    return True, ""


def obter_valor_planilha_por_headers(dados_planilha, *headers_possiveis):
    for header in headers_possiveis:
        if header in dados_planilha:
            valor = dados_planilha[header]
            if valor is not None and str(valor).strip() != "":
                return valor, header
    return None, None


def _perfil_pula_zero_forcado(dados_planilha: dict | None) -> bool:
    conc = str((dados_planilha or {}).get("concCod") or "").strip().upper()
    return conc in {"AMBAR_AM"}


def _valor_zero_numerico(valor) -> bool:
    try:
        num = valor_para_numero(valor)
    except Exception:
        return False
    return num is not None and abs(num) < 1e-9


def _pular_preenchimento_zero(header: str | None, valor, dados_planilha: dict | None = None) -> bool:
    if not _perfil_pula_zero_forcado(dados_planilha):
        return False
    if not header:
        return False
    if header.startswith("fatData") or header in {"cadTarifaCod", "cadSubGrupoCod", "fatCarimbo", "fatCodigoBarras"}:
        return False
    return _valor_zero_numerico(valor)


# =========================================================
# LOGIN
# =========================================================

def _esta_na_login_page(driver) -> bool:
    """Retorna True se o browser está na página de login do Consen."""
    try:
        url = driver.current_url
        return "login.php" in url or "login" in url.split("?")[0].rstrip("/").split("/")[-1].lower()
    except Exception:
        return False


def localizar_campos_login(driver, wait):
    seletores_usuario = [
        (By.NAME, "usuario"),
        (By.NAME, "username"),
        (By.NAME, "login"),
        (By.ID, "usuario"),
        (By.ID, "username"),
        (By.ID, "login"),
        (By.CSS_SELECTOR, "input[type='text']"),
        (By.CSS_SELECTOR, "input[type='email']"),
    ]
    seletores_senha = [
        (By.NAME, "senha"),
        (By.NAME, "password"),
        (By.ID, "senha"),
        (By.ID, "password"),
        (By.CSS_SELECTOR, "input[type='password']"),
    ]

    # Primeiro espera a página carregar com o wait cheio (20s),
    # depois varre cada seletor com timeout CURTO (1s) para não gastar
    # dezenas de segundos tentando nomes de campo que não existem.
    # Ao encontrar, confirma clickability com o wait original.
    wait_scan = WebDriverWait(driver, 1)

    def _achar(seletores: list, nome_campo: str):
        # Tenta rápido primeiro
        for by, sel in seletores:
            try:
                el = wait_scan.until(EC.element_to_be_clickable((by, sel)))
                log(f"Campo {nome_campo} encontrado: {by} = {sel}")
                return el
            except TimeoutException:
                pass
        # Nenhum achou em 1s — espera page load completo e tenta de novo com wait cheio
        try:
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
        except Exception:
            pass
        for by, sel in seletores:
            try:
                el = wait.until(EC.element_to_be_clickable((by, sel)))
                log(f"Campo {nome_campo} encontrado (2ª tentativa): {by} = {sel}")
                return el
            except TimeoutException:
                pass
        return None

    campo_usuario = _achar(seletores_usuario, "usuário")
    campo_senha   = _achar(seletores_senha,   "senha")

    if not campo_usuario or not campo_senha:
        raise Exception("Não foi possível localizar os campos de login.")

    return campo_usuario, campo_senha


def enviar_login(driver, wait, usuario, senha):
    campo_usuario, campo_senha = localizar_campos_login(driver, wait)

    # Clica explicitamente antes de digitar para garantir foco.
    # Sleeps aqui são FIXOS (não escalados por velocidade) — login deve ser confiável.
    campo_usuario.click()
    time.sleep(0.3)
    campo_usuario.clear()
    time.sleep(0.2)
    campo_usuario.send_keys(usuario)
    time.sleep(0.3)

    campo_senha.click()
    time.sleep(0.3)
    campo_senha.clear()
    time.sleep(0.2)
    campo_senha.send_keys(senha)
    time.sleep(0.3)

    # Verifica que os campos foram preenchidos; re-tenta uma vez se não
    try:
        val_u = campo_usuario.get_attribute("value") or ""
        val_s = campo_senha.get_attribute("value") or ""
        if not val_u.strip() or not val_s.strip():
            warn("Campos de login vazios após send_keys — re-tentando preenchimento.")
            campo_usuario.click(); time.sleep(0.2)
            campo_usuario.send_keys(Keys.CONTROL, "a"); campo_usuario.send_keys(usuario)
            time.sleep(0.3)
            campo_senha.click(); time.sleep(0.2)
            campo_senha.send_keys(Keys.CONTROL, "a"); campo_senha.send_keys(senha)
            time.sleep(0.3)
    except Exception:
        pass

    seletores_submit = [
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.CSS_SELECTOR, "input[type='submit']"),
        (By.XPATH, "//button[contains(., 'Entrar')]"),
        (By.XPATH, "//button[contains(., 'Login')]"),
        (By.XPATH, "//button[contains(., 'Acessar')]"),
        (By.XPATH, "//input[contains(@value, 'Entrar')]"),
    ]

    for by, sel in seletores_submit:
        try:
            btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((by, sel)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.3)
            btn.click()
            log(f"Botão de login clicado: {by} = {sel}")
            return
        except Exception:
            pass

    warn("Botão de login não encontrado. Tentando ENTER.")
    campo_senha.send_keys(Keys.ENTER)


_MAX_LOGIN_TENTATIVAS = 3


def _fazer_login_com_retry(driver, wait, usuario, senha) -> None:
    """Envia credenciais e confirma que saiu da login page. Retenta até 3x."""
    for tentativa in range(1, _MAX_LOGIN_TENTATIVAS + 1):
        try:
            enviar_login(driver, wait, usuario, senha)
        except Exception as e:
            warn(f"Login tentativa {tentativa}: enviar_login falhou: {e}")
            if tentativa == _MAX_LOGIN_TENTATIVAS:
                raise
            time.sleep(1.5)
            try:
                driver.get(LOGIN_URL)
                time.sleep(1.0)
            except Exception:
                pass
            continue

        try:
            WebDriverWait(driver, 30).until(EC.url_changes(LOGIN_URL))
        except Exception:
            warn(f"Login tentativa {tentativa}: timeout aguardando redirect (30s).")

        if not _esta_na_login_page(driver):
            log(f"Login concluido (tentativa {tentativa}).")
            return

        warn(f"Login tentativa {tentativa}: ainda na login page após redirect.")
        if tentativa < _MAX_LOGIN_TENTATIVAS:
            time.sleep(1.5)
            try:
                driver.get(LOGIN_URL)
                time.sleep(1.0)
            except Exception:
                pass

    if _esta_na_login_page(driver):
        raise Exception(
            f"Login no Consen falhou após {_MAX_LOGIN_TENTATIVAS} tentativas — "
            "verifique credenciais e disponibilidade do Consen."
        )


# =========================================================
# PLANILHA
# =========================================================

def encontrar_coluna_por_header(ws, header_desejado):
    desejado = normalizar_texto(header_desejado)

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if normalizar_texto(valor) == desejado:
            return col

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if desejado in normalizar_texto(valor):
            return col

    raise Exception(f"Não encontrei a coluna '{header_desejado}'.")


def obter_headers_planilha(caminho_arquivo):
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    return [str(h).strip() for h in headers if h is not None and str(h).strip()]


def ler_primeira_linha_completa_planilha(caminho_arquivo):
    if not os.path.exists(caminho_arquivo):
        raise Exception(f"Planilha não encontrada: {caminho_arquivo}")

    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        headers.append(str(h).strip() if h is not None else "")

    linha = 2
    dados = {}

    for col, header in enumerate(headers, start=1):
        if not header:
            continue
        dados[header] = ws.cell(row=linha, column=col).value

    return dados


def ler_todas_as_linhas_planilha(caminho_arquivo, linha_inicio=2):
    if not os.path.exists(caminho_arquivo):
        raise Exception(f"Planilha não encontrada: {caminho_arquivo}")

    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[wb.sheetnames[0]]
    log(f"Aba lida: {ws.title}")

    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        headers.append(str(h).strip() if h is not None else "")

    registros = []
    for linha in range(linha_inicio, ws.max_row + 1):
        dados = {}
        for col, header in enumerate(headers, start=1):
            if not header:
                continue
            dados[header] = ws.cell(row=linha, column=col).value

        instalacao = dados.get("Instalação") or dados.get("Instalacao")
        if instalacao is None or str(instalacao).strip() == "":
            continue

        leitura_anterior = valor_excel_para_date(dados.get("fatDataLeituraAnterior"))
        leitura_atual = valor_excel_para_date(dados.get("fatDataLeituraAtual"))
        data_emissao = valor_excel_para_date(dados.get("fatDataEmissao")) if dados.get("fatDataEmissao") else None
        data_vcto = valor_excel_para_date(dados.get("fatDataVcto")) if dados.get("fatDataVcto") else None
        carimbo = dados.get("fatCarimbo")

        # Usa fatDataReferencia ("MM/YYYY") quando disponível — para concessionárias como
        # COPEL MT onde o mês de referência é posterior ao mês da leitura atual.
        data_ref_esperada = (
            _ref_mmaaaa_para_date(dados.get("fatDataReferencia"))
            or (primeiro_dia_mes(leitura_atual) if leitura_atual else None)
        )

        # Permite linhas sem datas de leitura quando fatDataReferencia está presente
        # (ex: NF3e CPFL — o fallback busca as datas na tabela do Consen)
        if data_ref_esperada is None and (leitura_anterior is None or leitura_atual is None):
            continue

        instalacao_fmt = _formatar_instalacao_para_digitacao(instalacao, dados)

        # Permite sobrescrever data_ref_esperada via env var (ex: MT onde OCR extrai mês errado)
        if ENEL_DATA_REF_OVERRIDE:
            data_ref_esperada = _ref_mmaaaa_para_date(ENEL_DATA_REF_OVERRIDE) or data_ref_esperada

        registros.append({
            "linha_excel": linha,
            "instalacao": instalacao_fmt,
            "fatDataLeituraAnterior": leitura_anterior,
            "fatDataLeituraAtual": leitura_atual,
            "fatCarimbo": "" if carimbo is None else str(carimbo).strip(),
            "fatDataEmissao": data_emissao,
            "fatDataVcto": data_vcto,
            "dataReferenciaEsperada": data_ref_esperada,
            "erro_planilha": "" if dados.get("ERRO") is None else str(dados.get("ERRO")).strip(),
            "dados_completos": dados,
        })

    if not registros:
        raise Exception(f"Nenhuma linha válida encontrada a partir da linha {linha_inicio}.")

    log(f"Total de linhas válidas (a partir de {linha_inicio}): {len(registros)}")
    return registros


def _formatar_instalacao_para_digitacao(instalacao: str, dados_planilha: dict) -> str:
    """
    Equatorial GO pode exigir UC nova com mascara (x.xxx.xxx.xxx-xx)
    para localizar a instalacao no Consen.
    """
    valor = str(instalacao or "").strip()
    conc = str(dados_planilha.get("concCod") or "").strip().upper()
    arquivo = str(dados_planilha.get("ARQUIVO") or "").strip().upper()
    dig = re.sub(r"\D", "", valor)

    if conc == "COPEL" or "COPEL" in arquivo:
        if 11 <= len(dig) <= 15:
            return dig.zfill(15)
        return valor

    if conc == "EQUATORIAL" or "EQUATORIAL" in arquivo:
        if len(dig) == 12:
            return f"{dig[0]}.{dig[1:4]}.{dig[4:7]}.{dig[7:10]}-{dig[10:12]}"
        if len(dig) == 11:
            return f"{dig[0:3]}.{dig[3:6]}.{dig[6:9]}-{dig[9:11]}"
        if len(dig) == 10:
            return f"{dig[0:2]}.{dig[2:5]}.{dig[5:8]}-{dig[8:10]}"

    return valor


# =========================================================
# NAVEGAÇÃO
# =========================================================

def abrir_tela_instalacao(driver, wait):
    log("Tentando abrir a tela Instalacao via link JS...")

    js_click = f"""
    const links = Array.from(document.querySelectorAll('a'));
    const alvo = links.find(a =>
        (a.getAttribute('href') === '{LINK_HREF}') ||
        ((a.getAttribute('href') || '').includes('{LINK_HREF}')) ||
        ((a.textContent || '').trim() === '{LINK_TEXTO}')
    );
    if (alvo) {{
        alvo.scrollIntoView({{block:'center'}});
        alvo.click();
        return 'clicked';
    }}
    return 'not_found';
    """

    try:
        resultado = driver.execute_script(js_click)
        log(f"Resultado do clique JS em Instalacao: {resultado}")
        _s(0.5)
    except Exception as e:
        warn(f"Falha no clique JS da aba Instalacao: {type(e).__name__}")

    try:
        # Apenas considera "já aberta" se o campo instalacao está de fato no DOM
        if 'id="instalacao"' in driver.page_source:
            log("Tela Instalacao aparentemente já aberta após clique JS.")
            return True
    except Exception:
        pass

    log("Tentando abrir a tela Instalacao via alteração do hash...")
    try:
        try:
            driver.switch_to.alert.dismiss()
        except Exception:
            pass
        driver.execute_script("window.location.hash = arguments[0];", "bpg/gestao/fatura/cadastroTabFatura.php")
        _s(0.5)
        try:
            driver.switch_to.alert.dismiss()
        except Exception:
            pass

        if 'id="instalacao"' in driver.page_source:
            log("Tela Instalacao aberta via hash.")
            return True
    except Exception as e:
        warn(f"Falha ao alterar hash: {type(e).__name__}")

    log("Tentando abrir a tela Instalacao via URL direta...")
    try:
        driver.switch_to.alert.dismiss()
    except Exception:
        pass
    driver.get(TARGET_URL)
    try:
        WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.ID, "instalacao")))
    except Exception:
        _smin(1.0, 0.3)

    if 'id="instalacao"' in driver.page_source or "btnInstalacao" in driver.page_source:
        log("Tela Instalacao aberta via URL direta.")
        return True

    raise Exception("Não foi possível abrir a tela 'Instalacao'.")


# =========================================================
# CAMPOS DA TELA
# =========================================================

def localizar_input_generico(driver, wait, campo_id_ou_name):
    seletores = [
        (By.ID, campo_id_ou_name),
        (By.NAME, campo_id_ou_name),
        (By.CSS_SELECTOR, f"input#{campo_id_ou_name}"),
        (By.CSS_SELECTOR, f"input[name='{campo_id_ou_name}']"),
        (By.CSS_SELECTOR, f"textarea#{campo_id_ou_name}"),
        (By.CSS_SELECTOR, f"textarea[name='{campo_id_ou_name}']"),
        (By.CSS_SELECTOR, f"select#{campo_id_ou_name}"),
        (By.CSS_SELECTOR, f"select[name='{campo_id_ou_name}']"),
    ]

    ultimo = None
    for by, sel in seletores:
        try:
            wait.until(EC.presence_of_element_located((by, sel)))
            elementos = driver.find_elements(by, sel)
            candidatos = []
            for el in elementos:
                try:
                    if el.is_displayed() and el.is_enabled():
                        candidatos.append(el)
                except Exception:
                    pass
            if candidatos:
                ultimo = candidatos[-1]
                log(f"Campo encontrado: {by} = {sel} (último visível)")
                return ultimo
            elif elementos:
                ultimo = elementos[-1]
        except TimeoutException:
            pass

    if ultimo is not None:
        return ultimo

    raise Exception(f"Não encontrei o campo '{campo_id_ou_name}'.")


def localizar_elemento_por_id_ou_name(driver, campo_id=None, campo_name=None):
    candidatos = []

    tentativas = []
    if campo_id:
        tentativas.extend([
            (By.ID, campo_id),
            (By.CSS_SELECTOR, f"#{campo_id}"),
            (By.CSS_SELECTOR, f"input#{campo_id}"),
            (By.CSS_SELECTOR, f"textarea#{campo_id}"),
            (By.CSS_SELECTOR, f"select#{campo_id}"),
        ])

    if campo_name:
        tentativas.extend([
            (By.NAME, campo_name),
            (By.CSS_SELECTOR, f"[name='{campo_name}']"),
            (By.CSS_SELECTOR, f"input[name='{campo_name}']"),
            (By.CSS_SELECTOR, f"textarea[name='{campo_name}']"),
            (By.CSS_SELECTOR, f"select[name='{campo_name}']"),
        ])

    for by, sel in tentativas:
        try:
            encontrados = driver.find_elements(by, sel)
            for el in encontrados:
                if el not in candidatos:
                    candidatos.append(el)
        except Exception:
            pass

    if not candidatos:
        return None

    visiveis_habilitados = []
    for el in candidatos:
        try:
            if el.is_displayed() and el.is_enabled():
                visiveis_habilitados.append(el)
        except Exception:
            pass

    if visiveis_habilitados:
        return visiveis_habilitados[-1]

    habilitados = []
    for el in candidatos:
        try:
            if el.is_enabled():
                habilitados.append(el)
        except Exception:
            pass

    if habilitados:
        return habilitados[-1]

    return candidatos[-1]


def preencher_input_texto(driver, wait, campo_id_ou_name, valor, pausa_antes=0.0):
    campo = localizar_input_generico(driver, wait, campo_id_ou_name)

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
    _s(0.15)

    if pausa_antes > 0:
        _s(pausa_antes)

    try:
        campo.click()
        _s(0.1)
    except Exception:
        pass

    try:
        campo.send_keys(Keys.CONTROL, "a")
        campo.send_keys(Keys.DELETE)
    except Exception:
        pass

    try:
        driver.execute_script("arguments[0].value = '';", campo)
    except Exception:
        pass

    _s(0.2)

    try:
        campo.send_keys(str(valor))
    except Exception:
        driver.execute_script("arguments[0].value = arguments[1];", campo, str(valor))

    driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", campo)
    driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", campo)
    driver.execute_script("arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));", campo)

    try:
        valor_final = campo.get_attribute("value")
    except Exception:
        valor_final = "<não lido>"

    log(f"Campo {campo_id_ou_name} preenchido com {valor!r} | valor final na tela = {valor_final!r}")


def preencher_input_date(driver, wait, campo_id_ou_name, valor_date):
    campo = localizar_input_generico(driver, wait, campo_id_ou_name)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
    _s(0.1)

    valor_iso = formatar_yyyy_mm_dd(valor_date)
    for tentativa in range(1, 4):
        try:
            campo.click()
        except Exception:
            pass

        try:
            driver.execute_script("""
                arguments[0].value = '';
                arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
                arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
            """, campo)
        except Exception:
            pass

        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));
        """, campo, valor_iso)

        try:
            campo.send_keys(Keys.TAB)
        except Exception:
            pass

        _s(0.15)

        try:
            valor_final = (campo.get_attribute("value") or "").strip()
        except Exception:
            valor_final = ""

        log(
            f"Campo {campo_id_ou_name} preenchido com {valor_iso} | "
            f"tentativa={tentativa} | valor final na tela = {valor_final!r}"
        )

        if valor_final == valor_iso:
            return

        _s(0.15)

    warn(f"Campo {campo_id_ou_name} nao confirmou a data {valor_iso} apos 3 tentativas.")


def preencher_elemento_html(driver, elemento, valor):
    tag = (elemento.tag_name or "").lower()
    tipo = (elemento.get_attribute("type") or "").lower()

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elemento)
    _s(0.05)

    if tag == "select":
        valor_str = str(valor).strip()

        try:
            select = Select(elemento)

            try:
                select.select_by_visible_text(valor_str)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", elemento)
                log(f"[SELECT] '{valor_str}' selecionado por texto visível")
                return True
            except Exception:
                pass

            try:
                select.select_by_value(valor_str)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", elemento)
                log(f"[SELECT] '{valor_str}' selecionado por value")
                return True
            except Exception:
                pass

            for opt in select.options:
                opt_text = (opt.text or "").strip()
                if valor_str in opt_text or opt_text in valor_str:
                    select.select_by_visible_text(opt_text)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", elemento)
                    log(f"[SELECT] '{valor_str}' selecionado por match parcial em '{opt_text}'")
                    return True
        except Exception:
            pass

        ok = driver.execute_script("""
            const el = arguments[0];
            const v = arguments[1];
            let achou = false;
            for (const opt of el.options) {
                if ((opt.value || '') === v || (opt.textContent || '').trim() === v) {
                    el.value = opt.value;
                    el.dispatchEvent(new Event('input', {bubbles:true}));
                    el.dispatchEvent(new Event('change', {bubbles:true}));
                    achou = true;
                    break;
                }
            }
            return achou;
        """, elemento, valor_str)

        if not ok:
            try:
                opcoes = driver.execute_script("""
                    return Array.from(arguments[0].options).map(o => ({
                        value: o.value, text: o.textContent.trim()
                    }));
                """, elemento)
                warn(f"[SELECT] value='{valor_str}' não encontrado. Opções disponíveis: {opcoes}")
            except Exception:
                pass

        return bool(ok)

    valor_str = "" if valor is None else str(valor).strip()
    if tipo == "date" and valor_str:
        try:
            valor_str = formatar_yyyy_mm_dd(valor_excel_para_date(valor))
        except Exception:
            pass

    # Valor zero: JS direto sem send_keys (mais rápido; evita scrollIntoView+click por campo)
    _valor_limpo = valor_str.replace(".", "").replace(",", "").lstrip("-")
    _eh_zero = _valor_limpo in ("", "0", "00", "000")

    if not _eh_zero:
        try:
            elemento.click()
        except Exception:
            pass

        try:
            elemento.send_keys(Keys.CONTROL, "a")
            elemento.send_keys(Keys.DELETE)
        except Exception:
            pass

        try:
            driver.execute_script("arguments[0].value = '';", elemento)
        except Exception:
            pass

        try:
            elemento.send_keys(valor_str)
        except Exception:
            pass

    try:
        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));
        """, elemento, valor_str)
    except Exception:
        return False

    try:
        valor_final = (elemento.get_attribute("value") or "").strip()
    except Exception:
        valor_final = ""

    log(
        f"[INPUT] id={elemento.get_attribute('id')} name={elemento.get_attribute('name')} | "
        f"valor enviado={valor_str!r} | valor final={valor_final!r}"
    )

    vf = valor_final.replace(".", "").replace(",", "")
    vs = valor_str.replace(".", "").replace(",", "")
    return valor_final == valor_str or vf == vs


def preencher_input_mascarado_por_id(driver, campo_id, valor):
    elementos = driver.find_elements(By.ID, campo_id)
    if not elementos:
        raise Exception(f"Campo não encontrado: {campo_id}")

    alvo = None
    for el in elementos:
        try:
            if el.is_displayed() and el.is_enabled():
                alvo = el
        except Exception:
            pass

    if alvo is None:
        alvo = elementos[-1]

    valor_str = "" if valor is None else str(valor).strip()

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    _s(0.2)

    try:
        alvo.click()
        _s(0.1)
    except Exception:
        pass

    driver.execute_script("""
        const el = arguments[0];
        const val = arguments[1];

        const nativeSetter =
            Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;

        nativeSetter.call(el, '');
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));

        nativeSetter.call(el, val);
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        el.dispatchEvent(new Event('blur', { bubbles: true }));
    """, alvo, valor_str)

    _s(0.2)
    valor_final = (alvo.get_attribute("value") or "").strip()

    log(f"[CAMPO MASCARADO] id={campo_id} | valor enviado={valor_str!r} | valor final={valor_final!r}")
    return valor_final


def clicar_botao(driver, wait, seletores, descricao):
    botao = None

    for by, sel in seletores:
        try:
            encontrados = driver.find_elements(by, sel)
            candidatos = []
            for el in encontrados:
                try:
                    if el.is_displayed() and el.is_enabled():
                        candidatos.append(el)
                except Exception:
                    pass
            if candidatos:
                botao = candidatos[-1]
                log(f"Botão {descricao} encontrado: {by} = {sel}")
                break
            if encontrados:
                botao = encontrados[-1]
        except Exception:
            pass

    if botao is None:
        raise Exception(f"Não encontrei o botão '{descricao}'.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", botao)
    _s(0.2)

    try:
        botao.click()
        log(f"Botão {descricao} clicado com clique normal.")
    except Exception:
        driver.execute_script("arguments[0].click();", botao)
        log(f"Botão {descricao} clicado via JavaScript.")


def clicar_botao_carregar_instalacao(driver, wait):
    seletores = [
        (By.ID, "btnInstalacao"),
        (By.NAME, "btnInstalacao"),
        (By.CSS_SELECTOR, "button#btnInstalacao"),
        (By.CSS_SELECTOR, "button[name='btnInstalacao']"),
        (By.XPATH, "//button[@id='btnInstalacao']"),
        (By.XPATH, "//button[contains(normalize-space(), 'Carregar')]"),
    ]
    clicar_botao(driver, wait, seletores, "Carregar Instalação")


def clicar_botao_carregar_leitura(driver, wait):
    seletores = [
        (By.ID, "botaoLeitura"),
        (By.NAME, "botaoLeitura"),
        (By.CSS_SELECTOR, "button#botaoLeitura"),
        (By.CSS_SELECTOR, "button[name='botaoLeitura']"),
        (By.XPATH, "//button[@id='botaoLeitura']"),
    ]
    clicar_botao(driver, wait, seletores, "Carregar Leitura")


def aguardar_carregamento_tabela(driver, wait):
    try:
        wait.until(EC.presence_of_element_located((By.ID, "datatable_tabletools")))
        wait.until(lambda d: len(d.find_elements(By.CSS_SELECTOR, "#datatable_tabletools tbody tr")) >= 1)
        log("Tabela carregada.")
    except Exception:
        warn("A tabela não confirmou carregamento completo, seguindo com o HTML atual.")
    _s(0.5)


def scrollar_para_baixo(driver):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    _s(0.3)
    log("Scroll executado até o final da página.")


def aguardar_campos_finais(driver):
    deadline = time.time() + 4.0
    total = 0
    while time.time() < deadline:
        total = len(driver.find_elements(By.CSS_SELECTOR, "input, select, textarea"))
        if total >= 5:
            break
        _s(0.2)
    log(f"Campos visíveis no DOM: {total}")
    return total


def garantir_formulario_carregado(driver, wait, tentativas: int = 2):
    total = aguardar_campos_finais(driver)
    if total >= 5:
        return

    for tentativa in range(1, tentativas + 1):
        warn(
            f"Formulario veio incompleto apos carregar leitura (campos={total}). "
            f"Tentando recarregar ({tentativa}/{tentativas})..."
        )
        clicar_botao_carregar_leitura(driver, wait)
        _aguardar_sem_spinner(driver, timeout=6, min_wait=0.3)
        _s(0.4)
        scrollar_para_baixo(driver)
        total = aguardar_campos_finais(driver)
        if total >= 5:
            log("Formulario carregado apos recarregar leitura.")
            return

    raise Exception(f"Formulario nao carregou apos o carimbo (campos={total}).")


def _aguardar_sem_spinner(driver, timeout: float = 6.0, min_wait: float = 0.2):
    _smin(min_wait, 0.1)
    deadline = time.time() + timeout

    spinner_sels = [
        ".loading", ".spinner", ".overlay", "#loading",
        "[class*='loading']", "[class*='spinner']", "[id*='loading']",
        ".blockUI", ".blockOverlay", ".wait",
    ]
    spinner_css = ", ".join(spinner_sels)

    while time.time() < deadline:
        try:
            spinners = driver.find_elements(By.CSS_SELECTOR, spinner_css)
            visivel = any(
                s.is_displayed() for s in spinners
                if s.size.get("height", 0) > 0
            )
            if visivel:
                _s(0.15)
                continue
        except Exception:
            pass

        try:
            state = driver.execute_script("return document.readyState;")
            if state != "complete":
                _s(0.15)
                continue
        except Exception:
            pass

        try:
            jq_active = driver.execute_script(
                "return (typeof jQuery !== 'undefined') ? jQuery.active : 0;")
            if jq_active and int(jq_active) > 0:
                _s(0.15)
                continue
        except Exception:
            pass

        return


def aguardar_tela_instalacao_pronta(driver, wait, timeout=8, pausa_extra=0.5):
    log("Aguardando tela de instalação ficar pronta...")
    campo = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.ID, "instalacao"))
    )

    _smin(pausa_extra, 0.2)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
    _s(0.2)

    try:
        if campo.get_attribute("disabled") or campo.get_attribute("readonly"):
            warn("Campo instalação ainda estava disabled/readonly. Aguardando mais um pouco...")
            _smin(1.0, 0.3)
    except Exception:
        pass

    log("Tela de instalação pronta para digitação.")
    return campo


def campo_carimbo_existe(driver):
    try:
        elementos = driver.find_elements(By.ID, "carimbo")
        for el in elementos:
            try:
                if el.is_displayed():
                    return True
            except Exception:
                return True
    except Exception:
        pass

    try:
        elementos = driver.find_elements(By.NAME, "carimbo")
        for el in elementos:
            try:
                if el.is_displayed():
                    return True
            except Exception:
                return True
    except Exception:
        pass

    return False


def aguardar_campo_carimbo(driver, timeout=6.0, intervalo=0.2):
    fim = time.time() + timeout
    while time.time() < fim:
        if campo_carimbo_existe(driver):
            return True
        time.sleep(intervalo)
    return False


# =========================================================
# TABELA
# =========================================================

def localizar_tabela_faturas(driver, wait):
    seletores = [
        (By.ID, "datatable_tabletools"),
        (By.CSS_SELECTOR, "table#datatable_tabletools"),
    ]

    for by, sel in seletores:
        try:
            tabela = wait.until(EC.presence_of_element_located((by, sel)))
            log(f"Tabela encontrada: {by} = {sel}")
            return tabela
        except TimeoutException:
            pass

    raise Exception("Não encontrei a tabela de faturas.")


def obter_datas_referencia_tabela(driver, wait):
    tabela = localizar_tabela_faturas(driver, wait)
    linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody tr")
    log(f"Linhas encontradas na tabela: {len(linhas)}")

    datas = []

    for idx, linha in enumerate(linhas, start=1):
        colunas = linha.find_elements(By.TAG_NAME, "td")
        if not colunas:
            continue

        # Loga todas as colunas da primeira linha para diagnóstico
        if idx == 1:
            textos = [c.text.strip() for c in colunas]
            log(f"Linha 1 da tabela — colunas: {textos}")

        # Tenta a coluna 0 primeiro, depois demais colunas se falhar
        for col_idx, coluna in enumerate(colunas):
            texto = coluna.text.strip()
            if not texto:
                continue
            try:
                data_ref = parse_data_ddmmyyyy(texto)
                datas.append(data_ref)
                log(f"Linha {idx} col {col_idx} | Data Referência = {texto}")
                break  # achou data nesta linha, passa para a próxima
            except Exception:
                continue

        if not datas or len(datas) < idx:
            # Nenhuma coluna desta linha tinha data válida
            textos_linha = [c.text.strip() for c in colunas]
            warn(f"Linha {idx} da tabela sem data reconhecível: {textos_linha}")

    return datas


def obter_datas_leitura_da_tabela(driver, wait, data_ref_alvo=None) -> "tuple[date|None, date|None]":
    """Lê dataLeituraAtual (col 1) e dataLeituraAnterior (col 2) da linha com data_ref_alvo.
    Se data_ref_alvo for None, usa a primeira linha."""
    try:
        tabela = localizar_tabela_faturas(driver, wait)
        linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody tr")
        if not linhas:
            return None, None
        linha_alvo = None if data_ref_alvo else linhas[0]
        if data_ref_alvo:
            alvo_fmt = formatar_ddmmyyyy(data_ref_alvo)
            for ln in linhas:
                cols = ln.find_elements(By.TAG_NAME, "td")
                if cols and cols[0].text.strip() == alvo_fmt:
                    linha_alvo = ln
                    break
        if linha_alvo is None:
            return None, None
        colunas = linha_alvo.find_elements(By.TAG_NAME, "td")
        atual = anterior = None
        for col_idx in (1, 2):
            if col_idx < len(colunas):
                txt = colunas[col_idx].text.strip()
                try:
                    d = parse_data_ddmmyyyy(txt)
                    if col_idx == 1:
                        atual = d
                    else:
                        anterior = d
                except Exception:
                    pass
        return atual, anterior
    except Exception:
        return None, None


def obter_ultima_data_referencia_tabela(driver, wait):
    datas = obter_datas_referencia_tabela(driver, wait)

    if not datas:
        warn("Tabela sem datas de referência reconhecíveis.")
        return None

    ultima = max(datas)
    log(f"Última Data Referência da tabela: {formatar_ddmmyyyy(ultima)}")
    return ultima


def aguardar_formulario_edicao_pronto(driver, wait, timeout=12):
    wait.until(EC.presence_of_element_located((By.ID, "btnSalvar")))
    deadline = time.time() + timeout
    while time.time() < deadline:
        total = len(driver.find_elements(By.CSS_SELECTOR, "input, select, textarea"))
        if total >= 5:
            log(f"Formulario de edicao pronto (campos={total}).")
            return
        _s(0.2)
    raise TimeoutException("Formulario de edicao nao carregou apos abrir referencia existente.")


def abrir_referencia_existente_para_edicao(driver, wait, data_esperada):
    tabela = localizar_tabela_faturas(driver, wait)
    linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody tr")
    alvo = None

    for idx, linha in enumerate(linhas, start=1):
        colunas = linha.find_elements(By.TAG_NAME, "td")
        if not colunas:
            continue
        try:
            data_ref = parse_data_ddmmyyyy(colunas[0].text.strip())
        except Exception:
            continue
        if data_ref != data_esperada:
            continue

        links = linha.find_elements(By.CSS_SELECTOR, "a[href*='editaTabFatura.php']")
        if links:
            alvo = (idx, links[-1])
            break

    if alvo is None:
        raise Exception(
            "Referencia existente detectada, mas nao encontrei link de edicao "
            f"para {formatar_ddmmyyyy(data_esperada)}."
        )

    idx, link = alvo
    href = (link.get_attribute("href") or "").strip()
    log(
        "Abrindo referencia existente para edicao: "
        f"linha={idx} ref={formatar_ddmmyyyy(data_esperada)} href={href or '(sem href)'}"
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
    _s(0.2)
    try:
        link.click()
    except Exception:
        driver.execute_script("arguments[0].click();", link)

    _aguardar_sem_spinner(driver, timeout=8, min_wait=0.4)
    aguardar_formulario_edicao_pronto(driver, wait)
    scrollar_para_baixo(driver)


def preencher_datas_e_carimbo_se_necessario(driver, wait, registro_planilha):
    ultima_data_tabela = obter_ultima_data_referencia_tabela(driver, wait)
    data_esperada = registro_planilha["dataReferenciaEsperada"]

    log(f"Data Referência esperada da planilha: {formatar_ddmmyyyy(data_esperada)}")
    if ultima_data_tabela:
        log(f"Última Data Referência encontrada na tabela: {formatar_ddmmyyyy(ultima_data_tabela)}")
    else:
        log("Nenhuma Data Referência encontrada na tabela.")

    # Sempre tenta preencher datas; a decisão de pular vem da ausência do campo carimbo
    # (sinal real do Consen de que a referência já existe — evita falso-positivo por
    # leitura incorreta da coluna de vencimento na tabela de faturas)
    log("Preenchendo datas de leitura...")
    fat_lc = registro_planilha["fatDataLeituraAtual"]
    fat_la = registro_planilha["fatDataLeituraAnterior"]
    # Fallback para datas da tabela quando OCR não conseguiu extrair (None ou 0).
    # Com ENEL_DATA_REF_OVERRIDE, consulta a tabela mas só substitui datas inválidas;
    # datas válidas do OCR são mantidas mesmo com o override ativo.
    _lc_invalida = not fat_lc
    _la_invalida = not fat_la
    if _lc_invalida or _la_invalida or ENEL_DATA_REF_OVERRIDE:
        lc_tabela, la_tabela = obter_datas_leitura_da_tabela(driver, wait, data_ref_alvo=data_esperada)
        # Heurística de saneamento: se a leitura anterior do OCR não for menor do que
        # a leitura atual disponível (do OCR ou da tabela), trata como inválida e
        # deixa a tabela corrigir. Isso evita pares invertidos como 01/03 atual e
        # 05/05 anterior.
        referencia_atual = fat_lc or lc_tabela
        if fat_la and referencia_atual and fat_la >= referencia_atual:
            log(
                "[fallback] dataLeituraAnterior do OCR invalida "
                f"({fat_la} >= {referencia_atual}). "
                "Usando leitura anterior da tabela se disponível."
            )
            _la_invalida = True
        if lc_tabela and _lc_invalida:
            log(f"[fallback] dataLeituraAtual da tabela: {lc_tabela}")
            fat_lc = lc_tabela
        if la_tabela and _la_invalida:
            log(f"[fallback] dataLeituraAnterior da tabela: {la_tabela}")
            fat_la = la_tabela
    if not fat_lc or not fat_la:
        raise RuntimeError(
            "Datas de leitura ausentes apos OCR/fallback "
            f"(linha={registro_planilha.get('linha_excel')}, "
            f"instalacao={registro_planilha.get('instalacao')}, "
            f"ref={formatar_ddmmyyyy(data_esperada)})."
        )
    preencher_input_date(driver, wait, "dataLeituraAtual", fat_lc)
    preencher_input_date(driver, wait, "dataLeituraAnterior", fat_la)

    log("Clicando em carregar leitura após preencher datas...")
    clicar_botao_carregar_leitura(driver, wait)
    _aguardar_sem_spinner(driver, timeout=6, min_wait=0.3)
    _s(0.3)

    apareceu_carimbo = aguardar_campo_carimbo(driver, timeout=6.0, intervalo=0.2)

    if not apareceu_carimbo and ENEL_AJUSTAR_LEITURA_ULTIMO_DIA:
        # Referência já existente com as datas originais.
        # Tenta novamente com datas ajustadas (último dia do mês → 1º do mês seguinte).
        fat_lc_adj = _ajustar_leitura_ultimo_dia(fat_lc)
        fat_la_adj = _ajustar_leitura_ultimo_dia(fat_la)
        if fat_lc_adj != fat_lc or fat_la_adj != fat_la:
            log(
                f"[leitura] Referência existente com datas originais ({fat_lc}/{fat_la}). "
                f"Tentando ajustadas: {fat_lc_adj}/{fat_la_adj}"
            )
            preencher_input_date(driver, wait, "dataLeituraAtual", fat_lc_adj)
            preencher_input_date(driver, wait, "dataLeituraAnterior", fat_la_adj)
            clicar_botao_carregar_leitura(driver, wait)
            _aguardar_sem_spinner(driver, timeout=6, min_wait=0.3)
            _s(0.3)
            apareceu_carimbo = aguardar_campo_carimbo(driver, timeout=6.0, intervalo=0.2)
            if apareceu_carimbo:
                log("[leitura] Datas ajustadas aceitas pelo CONSEN.")
            else:
                log("[leitura] Datas ajustadas também não abriram referência nova.")

    if not apareceu_carimbo:
        if CONSEN_EDITAR_REFERENCIA_EXISTENTE:
            log(
                "Campo 'carimbo' NÃO apareceu após carregar as datas. "
                "Referência já existente. Abrindo linha existente para edição."
            )
            abrir_referencia_existente_para_edicao(driver, wait, data_esperada)
            return True
        log("Campo 'carimbo' NÃO apareceu após carregar as datas. Referência já existente. Pulando instalação.")
        return False

    log("Campo 'carimbo' apareceu. Data aceita. Seguindo para preencher carimbo...")

    valor_carimbo = normalizar_carimbo(registro_planilha.get("fatCarimbo", ""))

    if valor_carimbo:
        preencher_input_texto(driver, wait, "carimbo", valor_carimbo, pausa_antes=0.4)

        log("Clicando em carregar leitura após preencher carimbo...")
        clicar_botao_carregar_leitura(driver, wait)
        _aguardar_sem_spinner(driver, timeout=6, min_wait=0.3)
        _s(0.3)
    else:
        warn("fatCarimbo veio vazio na planilha. Seguindo sem preencher carimbo.")

    data_emissao = registro_planilha.get("fatDataEmissao")
    data_vcto = registro_planilha.get("fatDataVcto")
    if data_emissao:
        preencher_input_date(driver, wait, "dataEmissao", data_emissao)
    else:
        warn("fatDataEmissao veio vazio na planilha. Campo dataEmissao nao sera preenchido.")
    if data_vcto:
        preencher_input_date(driver, wait, "dataVencimento", data_vcto)
    else:
        raise _ErroNaoRecuperavel(
            "fatDataVcto ausente na planilha — dataVencimento é campo obrigatório no CONSEN. "
            "Corrija o OCR antes de redigitar."
        )

    scrollar_para_baixo(driver)
    garantir_formulario_carregado(driver, wait)
    return True


# =========================================================
# BS4 - EXPORTAR CAMPOS DISPONÍVEIS
# =========================================================

def extrair_campos_editaveis_bs4(driver):
    soup = BeautifulSoup(driver.page_source, "html.parser")
    campos = []

    for tag in soup.find_all("input"):
        tipo = (tag.get("type") or "text").strip().lower()

        if tipo in {"hidden", "submit", "button", "reset", "image", "file", "checkbox", "radio"}:
            continue

        if tag.has_attr("disabled") or tag.has_attr("readonly"):
            continue

        campo = {
            "tag": "input",
            "type": tipo,
            "name": tag.get("name"),
            "id": tag.get("id"),
            "value": tag.get("value"),
            "placeholder": tag.get("placeholder"),
            "class": " ".join(tag.get("class", [])) if tag.get("class") else "",
            "maxlength": tag.get("maxlength"),
            "required": tag.has_attr("required"),
        }
        campos.append(campo)

    for tag in soup.find_all("select"):
        if tag.has_attr("disabled") or tag.has_attr("readonly"):
            continue

        campo = {
            "tag": "select",
            "type": "select",
            "name": tag.get("name"),
            "id": tag.get("id"),
            "value": None,
            "placeholder": None,
            "class": " ".join(tag.get("class", [])) if tag.get("class") else "",
            "maxlength": None,
            "required": tag.has_attr("required"),
            "options": [
                {
                    "value": opt.get("value"),
                    "text": opt.get_text(" ", strip=True)
                }
                for opt in tag.find_all("option")
            ],
        }
        campos.append(campo)

    for tag in soup.find_all("textarea"):
        if tag.has_attr("disabled") or tag.has_attr("readonly"):
            continue

        campo = {
            "tag": "textarea",
            "type": "textarea",
            "name": tag.get("name"),
            "id": tag.get("id"),
            "value": tag.get_text("", strip=True),
            "placeholder": tag.get("placeholder"),
            "class": " ".join(tag.get("class", [])) if tag.get("class") else "",
            "maxlength": tag.get("maxlength"),
            "required": tag.has_attr("required"),
        }
        campos.append(campo)

    return campos


# =========================================================
# CORRELAÇÃO CAMPOS HTML x HEADERS PLANILHA
# =========================================================

def pontuar_correlacao(nome_campo, header_planilha):
    nome = normalizar_slug(nome_campo)
    header = normalizar_slug(header_planilha)

    if not nome or not header:
        return 0

    score = 0

    if nome == header:
        score += 100

    if nome in header or header in nome:
        score += 40

    tokens_nome = tokenizar(nome)
    tokens_header = tokenizar(header)
    inter = tokens_nome.intersection(tokens_header)
    score += len(inter) * 8

    aliases = [
        (["instalacao"], "Instalação"),
        (["carimbo"], "fatCarimbo"),
        (["tarifa"], "cadTarifaCod"),
        (["subgrupo"], "cadSubGrupoCod"),
        (["data", "leitura", "atual"], "fatDataLeituraAtual"),
        (["data", "leitura", "anterior"], "fatDataLeituraAnterior"),
        (["data", "vencimento"], "fatDataVcto"),
        (["data", "vcto"], "fatDataVcto"),
        (["data", "emissao"], "fatDataEmissao"),
        (["valor", "fatura"], "fatValorFatura"),
        (["valor", "nota", "fiscal"], "fatValorNotaFiscal"),
        (["ilum"], "fatIlumPublica"),
        (["icms"], "fatICMS"),
        (["pis"], "fatPIS"),
        (["cofins"], "fatCOFINS"),
        (["dem", "contratada", "ponta"], "fatDemContratadaPonta"),
        (["dem", "contratada", "fp"], "fatDemContratadaFPonta"),
        (["dem", "contratada", "geracao", "ponta"], "fatDemContratadaGeracaoPonta"),
        (["dem", "cont", "geracao", "ponta"], "fatDemContratadaGeracaoPonta"),
        (["dem", "contratada", "geracao", "fp"], "fatDemContratadaGeracaoFPonta"),
        (["dem", "cont", "geracao", "fp"], "fatDemContratadaGeracaoFPonta"),
        (["dem", "ponta", "registrada"], "fatDemPontaRegistrada"),
        (["dem", "fpind", "registrada"], "fatDemFPontaIndRegistrada"),
        (["dem", "fp", "registrada"], "fatDemFPontaIndRegistrada"),
        (["dem", "ponta", "faturada"], "fatDemPontaFaturada"),
        (["dem", "fpind", "faturada"], "fatDemFPontaIndFaturada"),
        (["dem", "fp", "faturada"], "fatDemFPontaIndFaturada"),
        (["consumo", "registrada", "ponta"], "fatConPontaRegistrado"),
        (["consumo", "registrada", "fpind"], "fatConFPontaIndRegistrado"),
        (["consumo", "registrada", "fp"], "fatConFPontaIndRegistrado"),
        (["consumo", "registrada", "intermediario"], "fatConIntermediarioRegistrado"),
        (["consumo", "faturada", "ponta"], "fatConPontaFaturado"),
        (["consumo", "faturada", "fpind"], "fatConFPontaIndFaturado"),
        (["consumo", "faturada", "fp"], "fatConFPontaIndFaturado"),
        (["consumo", "faturada", "intermediario"], "fatConIntermediarioFaturado"),
        (["rs", "kwh"], "fatRsKWh"),
    ]

    for palavras, header_alvo in aliases:
        if all(p in nome for p in palavras):
            # Não dar boost a campos "excedente" para headers de consumo normal (e vice-versa):
            # ex: txt-consumo-excedente-registrada-fpind não deve virar fatConFPontaIndRegistrado
            _nome_tem_exc = "excedente" in nome
            _alvo_tem_exc = "exc" in normalizar_slug(header_alvo)
            if _nome_tem_exc != _alvo_tem_exc:
                continue
            if normalizar_slug(header_planilha) == normalizar_slug(header_alvo):
                score += 70

    return score


def correlacionar_campos_com_headers(campos, headers_planilha):
    correlacoes = []

    for campo in campos:
        nome_base = campo.get("name") or campo.get("id") or ""
        melhor_header = None
        melhor_score = 0

        for header in headers_planilha:
            score = pontuar_correlacao(nome_base, header)
            if score > melhor_score:
                melhor_score = score
                melhor_header = header

        correlacoes.append({
            "campo_html": nome_base,
            "id": campo.get("id"),
            "name": campo.get("name"),
            "tag": campo.get("tag"),
            "type": campo.get("type"),
            "valor_atual_tela": campo.get("value"),
            "header_planilha_sugerido": melhor_header if melhor_score > 0 else "",
            "score_correlacao": melhor_score,
        })

    return correlacoes


# =========================================================
# SALVAR EXPORTS
# =========================================================

def salvar_campos_disponiveis(campos):
    caminho_json = PASTA_SAIDA / "campos_disponiveis.json"
    caminho_csv = PASTA_SAIDA / "campos_disponiveis.csv"

    caminho_json.write_text(json.dumps(campos, ensure_ascii=False, indent=2), encoding="utf-8")

    with caminho_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "tag", "type", "name", "id", "value", "placeholder",
            "class", "maxlength", "required"
        ])

        for c in campos:
            writer.writerow([
                c.get("tag"), c.get("type"), c.get("name"), c.get("id"), c.get("value"),
                c.get("placeholder"), c.get("class"), c.get("maxlength"), c.get("required"),
            ])

    log(f"Campos disponíveis salvos em: {caminho_json}")
    log(f"Campos disponíveis salvos em: {caminho_csv}")


def salvar_correlacoes(correlacoes):
    caminho = PASTA_SAIDA / "mapeamento_campos_planilha.csv"

    with caminho.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "campo_html", "id", "name", "tag", "type",
            "valor_atual_tela", "header_planilha_sugerido", "score_correlacao"
        ])

        for item in correlacoes:
            writer.writerow([
                item.get("campo_html"),
                item.get("id"),
                item.get("name"),
                item.get("tag"),
                item.get("type"),
                item.get("valor_atual_tela"),
                item.get("header_planilha_sugerido"),
                item.get("score_correlacao"),
            ])

    log(f"Mapeamento sugerido salvo em: {caminho}")


def exportar_campos_e_correlacoes(driver, excel_path):
    headers = obter_headers_planilha(excel_path)
    campos = extrair_campos_editaveis_bs4(driver)
    correlacoes = correlacionar_campos_com_headers(campos, headers)

    salvar_campos_disponiveis(campos)
    salvar_correlacoes(correlacoes)

    return campos, correlacoes


# =========================================================
# CONSUMIR CSV DE MAPEAMENTO E PREENCHER
# =========================================================

def carregar_mapeamento_csv(caminho_csv):
    caminho_csv = Path(caminho_csv)
    if not caminho_csv.exists():
        raise Exception(f"CSV de mapeamento não encontrado: {caminho_csv}")

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with caminho_csv.open("r", newline="", encoding=enc) as f:
                return list(csv.DictReader(f, delimiter=";"))
        except UnicodeDecodeError:
            continue

    raise Exception(f"Não consegui ler o CSV de mapeamento: {caminho_csv}")


def escolher_csv_mapeamento():
    if MAPEAMENTO_CSV.exists():
        log(f"Usando CSV de mapeamento externo: {MAPEAMENTO_CSV}")
        return MAPEAMENTO_CSV

    interno = PASTA_SAIDA / "mapeamento_campos_planilha.csv"
    if interno.exists():
        log(f"Usando CSV de mapeamento gerado pelo script: {interno}")
        return interno

    raise Exception("Nenhum CSV de mapeamento encontrado.")


def formatar_valor_para_campo(header, valor, tipo_html):
    if valor is None:
        return ""

    if isinstance(valor, datetime):
        valor = valor.date()

    if isinstance(valor, date):
        if str(tipo_html).lower() == "date":
            return valor.strftime("%Y-%m-%d")
        return valor.strftime("%d/%m/%Y")

    header_norm = normalizar_texto(header)
    tipo_norm = str(tipo_html).lower()

    if "cod" in header_norm:
        return str(valor).strip()

    if "data" in header_norm and tipo_norm == "date":
        try:
            return formatar_yyyy_mm_dd(valor_excel_para_date(valor))
        except Exception:
            return str(valor).strip()

    # "nota" removido — fatValorNotaFiscal é numérico e deve ser formatado como R$
    if any(chave in header_norm for chave in [
        "instalacao", "carimbo", "cnpj", "faturaid", "referencia",
        "codigobarras", "codigobarra"
    ]):
        return str(valor).strip()

    num = valor_para_numero(valor)
    if num is not None:
        # Proteção para fatValorNotaFiscal / fatValorFatura:
        # valores > R$ 999.999 são quase certamente erro de parsing do Excel
        # (ex: "112,56" serializado como 112560000,00)
        if "valornotafiscal" in header_norm or "valorfatura" in header_norm:
            if num > 999999:
                warn(f"[NOTA FISCAL] Valor suspeito ({num:.2f}) para '{header}' — campo ignorado para evitar erro de máscara")
                return ""
        if tipo_norm in ("number", "range"):
            return str(num)
        return formatar_numero_br(num)

    return str(valor).strip()


def _resolver_valor_select(campo_html: str, valor_bruto) -> str:
    chave = str(valor_bruto).strip()

    if campo_html in MAP_SELECTS:
        mapa = MAP_SELECTS[campo_html]
        if chave in mapa:
            return mapa[chave]
        warn(f"[MAP_SELECTS] Texto '{chave}' não está no mapa de '{campo_html}'. "
             f"Opções conhecidas: {list(mapa.keys())}. Tentando usar o texto diretamente no select.")
    return chave


def _texto_select_atual(elemento) -> str:
    try:
        select = Select(elemento)
        return (select.first_selected_option.text or "").strip()
    except Exception:
        try:
            return (elemento.get_attribute("value") or "").strip()
        except Exception:
            return ""


def preencher_classificacao_contratual_inicio(driver, dados_planilha):
    """
    Alguns fluxos de media tensao so estabilizam os demais campos depois que
    tarifa e subgrupo ficam definidos na tela.
    """
    pares = [
        ("cb-dados-contratuais-fatura-tarifa", "cadTarifaCod", "tarifa"),
        ("cb-dados-contratuais-fatura-subgrupo", "cadSubGrupoCod", "subgrupo"),
    ]

    for campo_html, header_planilha, rotulo in pares:
        valor_bruto, header_usado = obter_valor_planilha_por_headers(dados_planilha, header_planilha)
        if header_usado is None:
            continue

        elemento = localizar_elemento_por_id_ou_name(driver, campo_id=campo_html, campo_name=campo_html)
        if elemento is None:
            warn(f"[CLASSIFICACAO] Campo de {rotulo} nao encontrado na tela.")
            continue

        valor_formatado = _resolver_valor_select(campo_html, valor_bruto)
        log(f"[CLASSIFICACAO] Definindo {rotulo}: {valor_formatado!r}")

        confirmado = False
        for tentativa in range(1, 4):
            ok = preencher_elemento_html(driver, elemento, valor_formatado)
            texto_atual = _texto_select_atual(elemento)
            if ok and texto_atual and "Selecione" not in texto_atual:
                log(f"[CLASSIFICACAO] {rotulo.title()} confirmado na tela: {texto_atual}")
                confirmado = True
                break

            warn(
                f"[CLASSIFICACAO] Tentativa {tentativa}/3 de {rotulo} ainda nao confirmou "
                f"(atual={texto_atual!r}). Reaplicando..."
            )
            _s(0.15)
            elemento = localizar_elemento_por_id_ou_name(driver, campo_id=campo_html, campo_name=campo_html)
            if elemento is None:
                break

        if not confirmado:
            warn(f"[CLASSIFICACAO] Falha ao definir {rotulo} no inicio do fluxo.")


def preencher_aliases_forcados_linha(driver, dados_planilha):
    preenchidos = []
    pulados = []
    campos_alias_ok = set()
    campos_alias_falhos = set()

    # Detecta cooperativa: PIS e COFINS alíquota = 0 → retentões CSLL/IRPJ devem ser zeradas
    try:
        _pis = float(dados_planilha.get("fatDescPisAliquota") or 0)
        _cof = float(dados_planilha.get("fatDesCofinsAliquota") or 0)
        _eh_cooperativa = (_pis == 0.0 and _cof == 0.0)
    except Exception:
        _eh_cooperativa = False

    # Detecta concessionárias que cadastram retenção PIS/COFINS no Consen
    # Regra: explícito por nome (COPEL, CPFL, EDP) OU pelo dado do xlsx (qualquer
    # concessionária cujo OCR extraiu valores não-zero de PIS/COFINS — ex: CEEE)
    _conc = str(dados_planilha.get("concCod") or "").strip().upper()
    _arq = str(dados_planilha.get("ARQUIVO") or "").strip().upper()
    _eh_copel = _conc == "COPEL" or "COPEL" in _arq
    _usa_ret_pis_cof = (
        _eh_copel
        or _conc in {"CPFL", "EDP SP", "EDP ES"}
        or "CPFL" in _arq
        or _tem_ret_pis_cof_individuais(dados_planilha)
    )
    _usa_somente_trib_federal = _usa_somente_tributo_federal(dados_planilha)

    # Mapeamento efetivo para retenção PIS/COFINS: usa xlsx se COPEL/CPFL, _zero para os demais
    _ALIASES_RET_PIS_COF = {
        "fatDescPisPercRetImposto":    "fatDescPisPercRetImposto"    if _usa_ret_pis_cof else "_zero",
        "fatDescPisValRetImposto":     "fatDescPisValRetImposto"     if _usa_ret_pis_cof else "_zero",
        "fatDescCofinsPercRetImposto": "fatDescCofinsPercRetImposto" if _usa_ret_pis_cof else "_zero",
        "fatDescCofinsValRetImposto":  "fatDescCofinsValRetImposto"  if _usa_ret_pis_cof else "_zero",
    }

    for campo_html, header_planilha in ALIASES_FORCADOS.items():
        # Para concessionárias com retenção PIS/COFINS (COPEL, CEEE, CPFL, EDP, etc.)
        if campo_html in _ALIASES_RET_PIS_COF:
            header_planilha = _ALIASES_RET_PIS_COF[campo_html]
        if _usa_somente_trib_federal and campo_html in _CAMPOS_RETENCAO_INDIVIDUAL:
            header_planilha = "_zero"
        headers_alternativos = [header_planilha]

        # Fallbacks
        if header_planilha == "fatConFPontaInjetadoUsinaSaldoAcumulado":
            headers_alternativos.append("fatConFPontaInjetadoUsinaSaldo")
        if header_planilha == "fatConPontaInjetadoUsinaSaldoAcumulado":
            headers_alternativos.append("fatConPontaInjetadoUsinaSaldo")

        valor_bruto, header_usado = obter_valor_planilha_por_headers(dados_planilha, *headers_alternativos)

        if header_usado is None:
            pulados.append((campo_html, f"header/valor ausente: {headers_alternativos}"))
            campos_alias_falhos.add(campo_html)
            continue

        # Cooperativa (PIS=COF=0): zera CSLL/IRPJ — Consen rejeita retenção sem PIS/COFINS
        # Só zera se o xlsx também for 0: evita falso-positivo quando OCR não extraiu alíquotas (EDP layout antigo)
        if campo_html in _CAMPOS_RET_CSLL_IRPJ and _eh_cooperativa:
            xlsx_val = abs(float(dados_planilha.get(campo_html, 0) or 0))
            if xlsx_val == 0:
                valor_bruto = 0
                header_usado = "_zero_cooperativa"

        # Fallback nota fiscal: se 0, reconstrói como fatura + retenções que serão preenchidas no Consen
        if campo_html == "txt-dados-financeiros-valor-nota-fiscal":
            try:
                _nf = float(str(valor_bruto).replace(",", ".").replace(" ", "") or 0)
            except Exception:
                _nf = 0.0
            if _nf == 0.0:
                try:
                    _fatura = float(dados_planilha.get("fatValorFatura") or 0)
                    if _eh_cooperativa:
                        _rets = 0.0  # cooperativa: CSLL/IRPJ zerados → nota fiscal = fatura
                    elif _usa_somente_trib_federal:
                        _rets = abs(float(dados_planilha.get("fatTributoFederalVal") or 0))
                    else:
                        _ret_campos = ["fatDescCsllValRetImposto", "fatDescIrpjValRetImposto"]
                        if _usa_ret_pis_cof:
                            _ret_campos += ["fatDescPisValRetImposto", "fatDescCofinsValRetImposto"]
                        _rets = sum(
                            abs(float(dados_planilha.get(c) or 0))
                            for c in _ret_campos
                        )
                    if _fatura > 0:
                        valor_bruto = round(_fatura + _rets, 2)
                        header_usado = "_calc_nota_via_fatura"
                        log(f"[ALIAS] fatValorNotaFiscal=0 -> calculado via fatura+rets: {valor_bruto}")
                except Exception:
                    pass

        if _pular_preenchimento_zero(header_usado, valor_bruto, dados_planilha):
            pulados.append((campo_html, f"zero ignorado para {header_usado}"))
            continue

        elemento = localizar_elemento_por_id_ou_name(driver, campo_id=campo_html, campo_name=campo_html)
        if elemento is None:
            pulados.append((campo_html, "elemento não encontrado na tela"))
            campos_alias_falhos.add(campo_html)
            continue

        tag = (elemento.tag_name or "").lower()
        tipo = (elemento.get_attribute("type") or tag or "").lower()
        valor_antes = _valor_elemento_seguro(elemento)

        if tag == "select":
            valor_formatado = _resolver_valor_select(campo_html, valor_bruto)
        else:
            valor_formatado = formatar_valor_para_campo(header_usado, valor_bruto, tipo)

        log(f"[ALIAS] campo={campo_html} | header={header_usado} | valor_bruto={valor_bruto!r} | valor_formatado={valor_formatado!r}")

        try:
            ok = preencher_elemento_html(driver, elemento, valor_formatado)
            valor_depois = _valor_elemento_seguro(elemento)
            _trace_campos_evento(
                "alias_forcado",
                campo_consen=campo_html,
                header_planilha=header_usado,
                valor_bruto=valor_bruto,
                valor_enviado=valor_formatado,
                valor_antes=valor_antes,
                valor_depois=valor_depois,
                ok=ok,
                url=getattr(driver, "current_url", ""),
            )
            if ok:
                preenchidos.append((campo_html, header_usado, valor_formatado))
                campos_alias_ok.add(campo_html)
            else:
                pulados.append((campo_html, f"falha ao preencher (value='{valor_formatado}')"))
                campos_alias_falhos.add(campo_html)
        except Exception as e:
            _trace_campos_evento(
                "alias_forcado_erro",
                campo_consen=campo_html,
                header_planilha=header_usado,
                valor_bruto=valor_bruto,
                valor_enviado=valor_formatado,
                valor_antes=valor_antes,
                erro=f"{type(e).__name__}: {e}",
                url=getattr(driver, "current_url", ""),
            )
            pulados.append((campo_html, f"erro: {type(e).__name__}"))
            campos_alias_falhos.add(campo_html)

    print("=" * 100)
    print("ALIASES FORÇADOS")
    print("=" * 100)
    for campo, header, valor in preenchidos:
        print(f"{campo} <- {header} | valor={valor}")

    return preenchidos, pulados, campos_alias_ok, campos_alias_falhos


def preencher_campos_via_mapeamento_csv_linha(
    driver,
    dados_planilha,
    caminho_csv,
    score_minimo=35,
    campos_alias_ok: set[str] | None = None,
    campos_alias_falhos: set[str] | None = None,
):
    mapeamentos = carregar_mapeamento_csv(caminho_csv)
    preenchidos = []
    pulados = []
    _zero_batch: list = []  # campos zero acumulados para batch JS
    campos_alias_ok = campos_alias_ok or set()
    campos_alias_falhos = campos_alias_falhos or set()

    campos_pulados_fluxo = {
        "instalacao", "dataLeituraAtual", "dataLeituraAnterior",
        "carimbo", "dataVencimento", "dataEmissao",
        "fatCarimbo",
    }
    campos_pulados_alias = set(ALIASES_FORCADOS.keys()) | CAMPOS_BLOQUEADOS_CSV

    # Todos os campos input numéricos acumulados para batch JS (uma roundtrip)
    _input_batch: list = []

    for item in mapeamentos:
        campo_id = (item.get("id") or "").strip()
        campo_name = (item.get("name") or "").strip()
        tag = (item.get("tag") or "").strip().lower()
        tipo = (item.get("type") or "").strip().lower()
        header = (item.get("header_planilha_sugerido") or "").strip()

        try:
            score = int(float(item.get("score_correlacao") or 0))
        except Exception:
            score = 0

        identificador = campo_id or campo_name or item.get("campo_html") or "<sem_id>"

        if not header:
            pulados.append((identificador, "sem header sugerido"))
            continue
        if campo_bloqueado_generico(campo_id, campo_name, identificador) and not permitir_campo_beneficio_copel(header, dados_planilha):
            pulados.append((identificador, "campo bloqueado por heuristica (subvencao/beneficio)"))
            continue
        if score < score_minimo:
            pulados.append((identificador, f"score baixo ({score})"))
            continue
        if campo_id in campos_pulados_fluxo or campo_name in campos_pulados_fluxo:
            pulados.append((identificador, "campo já tratado pelo fluxo principal"))
            continue
        if (
            (campo_id in campos_pulados_alias or campo_name in campos_pulados_alias)
            and campo_id not in campos_alias_falhos
            and campo_name not in campos_alias_falhos
        ):
            pulados.append((identificador, "campo já tratado por alias forçado"))
            continue
        if header_bloqueado_planilha(header, dados_planilha):
            pulados.append((identificador, f"header bloqueado explicitamente: {header}"))
            continue
        if header not in dados_planilha:
            pulados.append((identificador, f"header não existe na planilha: {header}"))
            continue

        valor_bruto = dados_planilha[header]
        if valor_bruto is None or str(valor_bruto).strip() == "":
            pulados.append((identificador, f"valor vazio para {header}"))
            continue
        if _pular_preenchimento_zero(header, valor_bruto, dados_planilha):
            pulados.append((identificador, f"zero ignorado para {header}"))
            continue

        valor_formatado = formatar_valor_para_campo(header, valor_bruto, tipo)

        # Selects e datas precisam de tratamento individual (não acumulam no batch)
        _pode_batch = tag != "select" and tipo not in ("date", "datetime-local", "datetime")
        if _pode_batch:
            _input_batch.append((campo_id or campo_name, valor_formatado, identificador, header, score, tag, tipo))
            continue

        # Select/date: tratamento individual (mantém lógica original)
        elemento = localizar_elemento_por_id_ou_name(driver, campo_id=campo_id, campo_name=campo_name)
        if elemento is None:
            pulados.append((identificador, "elemento não encontrado na tela"))
            continue

        try:
            ok = preencher_elemento_html(driver, elemento, valor_formatado)
            if ok:
                preenchidos.append((identificador, header, valor_formatado, score, tag, tipo))
            else:
                pulados.append((identificador, "falha ao preencher"))
        except Exception as e:
            pulados.append((identificador, f"erro: {type(e).__name__}"))

    # Batch JS para TODOS os campos input (zeros e não-zeros): uma única roundtrip ao browser
    if _input_batch:
        pares = [[item[0], item[1]] for item in _input_batch]
        try:
            driver.execute_script("""
                var pares = arguments[0];
                pares.forEach(function(p) {
                    var id = p[0], val = p[1];
                    var el = document.getElementById(id) || document.getElementsByName(id)[0];
                    if (el) {
                        el.value = val;
                        el.dispatchEvent(new Event('input',  {bubbles:true}));
                        el.dispatchEvent(new Event('change', {bubbles:true}));
                        el.dispatchEvent(new Event('blur',   {bubbles:true}));
                    }
                });
            """, pares)
            for campo_id_b, val_b, idf_b, hdr_b, sc_b, tg_b, tp_b in _input_batch:
                preenchidos.append((idf_b, hdr_b, val_b, sc_b, tg_b, tp_b))
                log(f"[INPUT batch] id={campo_id_b} | valor={val_b!r}")
        except Exception as e:
            # Fallback: preenche campo a campo se o batch falhar
            for campo_id_b, val_b, idf_b, hdr_b, sc_b, tg_b, tp_b in _input_batch:
                elem = localizar_elemento_por_id_ou_name(driver, campo_id=campo_id_b, campo_name="")
                if elem is None:
                    pulados.append((idf_b, f"batch falhou e elemento nao encontrado: {type(e).__name__}"))
                    continue
                try:
                    ok = preencher_elemento_html(driver, elem, val_b)
                    if ok:
                        preenchidos.append((idf_b, hdr_b, val_b, sc_b, tg_b, tp_b))
                    else:
                        pulados.append((idf_b, "fallback falhou ao preencher"))
                except Exception as e2:
                    pulados.append((idf_b, f"fallback erro: {type(e2).__name__}"))


    caminho_log = PASTA_SAIDA / "resultado_preenchimento_csv.csv"
    with caminho_log.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["status", "campo", "header", "valor", "score_ou_motivo"])
        for campo, header, valor, score, _, _ in preenchidos:
            writer.writerow(["preenchido", campo, header, valor, score])
        for campo, motivo in pulados:
            writer.writerow(["pulado", campo, "", "", motivo])

    log(f"Resultado do preenchimento salvo em: {caminho_log}")
    return preenchidos, pulados


def preencher_saldos_injetado_usina(driver, wait, dados_planilha):
    mapa = {
        "txt-consumo-injetado-usina-pta-saldo": "fatConPontaInjetadoUsinaSaldoAcumulado",
        "txt-consumo-injetado-usina-fpta-saldo": "fatConFPontaInjetadoUsinaSaldoAcumulado",
    }

    for campo_consen, header_planilha in mapa.items():
        valor_bruto, header_usado = obter_valor_planilha_por_headers(
            dados_planilha,
            header_planilha,
            header_planilha.replace("Acumulado", "")
        )

        if header_usado is None:
            warn(f"[SALDO USINA] Valor ausente na planilha para {campo_consen}")
            continue

        ok = preencher_saldo_usina_tabela(driver, wait, campo_consen, valor_bruto)

        if ok:
            log(f"[SALDO USINA] OK | {campo_consen} <- {header_usado} | valor={valor_bruto!r}")
        else:
            warn(f"[SALDO USINA] FALHOU | {campo_consen} <- {header_usado} | valor={valor_bruto!r}")


# =========================================================
# OBSERVAÇÕES / BOTÕES FINAIS
# =========================================================

def preencher_obs_multiplas(driver, wait, dados_planilha: dict):
    pares = []
    for i in range(1, _OBS_PARES_MAX + 1):
        cod = str(dados_planilha.get(f"obsCod_{i}") or "").strip()
        val = dados_planilha.get(f"obsValor_{i}")
        if cod and cod not in ("", "0"):
            pares.append((cod, val))

    if not pares:
        log("[OBS] Nenhuma observação para preencher.")
        return

    for idx, (cod, val) in enumerate(pares, start=1):
        try:
            el_sel = wait.until(EC.presence_of_element_located((By.ID, "cb-dados-financeiros-obs")))
            ok = preencher_elemento_html(driver, el_sel, cod)
            if ok:
                log(f"[OBS] Linha {idx}: select cod={cod}")
            else:
                warn(f"[OBS] Linha {idx}: falha ao selecionar cod={cod}")
                continue
        except Exception as e:
            warn(f"[OBS] Linha {idx}: erro no select — {type(e).__name__}")
            continue

        try:
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));",
                el_sel)
        except Exception:
            pass

        try:
            driver.execute_script("document.body.click();")
        except Exception:
            pass
        _s(0.2)

        val_fmt = formatar_numero_br(val) if val is not None else "0,00"
        try:
            el_val = wait.until(EC.element_to_be_clickable((By.ID, "txt-dados-financeiros-outros")))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el_val)
            _s(0.2)
            driver.execute_script("arguments[0].value = '';", el_val)
            driver.execute_script("arguments[0].dispatchEvent(new Event('focus',{bubbles:true}));", el_val)
            el_val.click()
            el_val.send_keys(Keys.CONTROL, "a")
            el_val.send_keys(Keys.DELETE)
            driver.execute_script("arguments[0].value = '';", el_val)
            _s(0.1)
            el_val.send_keys(val_fmt)
            driver.execute_script("""
                arguments[0].dispatchEvent(new Event('input',{bubbles:true}));
                arguments[0].dispatchEvent(new Event('change',{bubbles:true}));
            """, el_val)
            log(f"[OBS] Linha {idx}: valor={val_fmt}")
        except Exception as e:
            warn(f"[OBS] Linha {idx}: erro no campo valor — {type(e).__name__}")
            try:
                driver.execute_script("""
                    var el = document.getElementById('txt-dados-financeiros-outros');
                    if(el){
                        el.value = arguments[0];
                        el.dispatchEvent(new Event('input',{bubbles:true}));
                        el.dispatchEvent(new Event('change',{bubbles:true}));
                    }
                """, val_fmt)
                log(f"[OBS] Linha {idx}: valor via JS={val_fmt}")
            except Exception:
                pass

        _s(0.3)

        try:
            btn = wait.until(EC.element_to_be_clickable((By.ID, "btnIncluiLinha")))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            _s(0.2)
            btn.click()
            log(f"[OBS] Linha {idx}: btnIncluiLinha clicado")
        except Exception as e:
            warn(f"[OBS] Linha {idx}: erro ao clicar btnIncluiLinha — {type(e).__name__}")
            try:
                driver.execute_script("document.getElementById('btnIncluiLinha').click();")
                log(f"[OBS] Linha {idx}: btnIncluiLinha via JS")
            except Exception:
                pass

        _s(0.3)

    log(f"[OBS] {len(pares)} observação(ões) incluída(s).")


def _verificar_erros_validacao_pos_salvar(driver) -> None:
    """Detecta campos obrigatórios marcados como inválidos após clicar Salvar.

    Se o CONSEN manteve o formulário aberto com .is-invalid, o save NÃO foi
    confirmado. Levanta RuntimeError para que o registro seja marcado como erro.
    """
    campos_invalidos = driver.find_elements(
        By.CSS_SELECTOR, ".is-invalid, .invalid-feedback:not(:empty)"
    )
    if campos_invalidos:
        msgs = [e.text.strip() for e in campos_invalidos if e.text.strip()]
        raise _ErroNaoRecuperavel(
            f"CONSEN rejeitou o formulário (campo obrigatório vazio após Salvar): {msgs}. "
            "Salvamento NÃO confirmado."
        )


def clicar_botao_salvar(driver, wait):
    seletores = [
        (By.ID, "btnSalvar"),
        (By.NAME, "btnSalvar"),
        (By.CSS_SELECTOR, "button#btnSalvar"),
        (By.CSS_SELECTOR, "button[name='btnSalvar']"),
        (By.XPATH, "//button[@id='btnSalvar']"),
        (By.XPATH, "//button[contains(normalize-space(), 'Salvar')]")
    ]
    clicar_botao(driver, wait, seletores, "Salvar")


def clicar_link_auditoria(driver, wait):
    seletores = [
        (By.ID, "linkAuditoria"),
        (By.CSS_SELECTOR, "a#linkAuditoria"),
        (By.XPATH, "//a[@id='linkAuditoria']"),
        (By.XPATH, "//a[contains(normalize-space(), 'Auditoria')]")
    ]

    link = None
    for by, sel in seletores:
        try:
            encontrados = driver.find_elements(by, sel)
            candidatos = []
            for el in encontrados:
                try:
                    if el.is_displayed() and el.is_enabled():
                        candidatos.append(el)
                except Exception:
                    pass
            if candidatos:
                link = candidatos[-1]
                log(f"Link Auditoria encontrado: {by} = {sel}")
                break
            if encontrados:
                link = encontrados[-1]
        except Exception:
            pass

    if link is None:
        raise Exception("Não encontrei o link 'Auditoria'.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
    _s(0.5)

    try:
        link.click()
        log("Link Auditoria clicado com clique normal.")
    except Exception:
        driver.execute_script("arguments[0].click();", link)
        log("Link Auditoria clicado via JavaScript.")


def clicar_botao_proxima_fatura(driver, wait):
    seletores = [
        (By.ID, "btnProxima"),
        (By.NAME, "btnProxima"),
        (By.CSS_SELECTOR, "button#btnProxima"),
        (By.CSS_SELECTOR, "button[name='btnProxima']"),
        (By.XPATH, "//button[@id='btnProxima']"),
        (By.XPATH, "//button[contains(normalize-space(), 'Proxima Fatura')]")
    ]
    clicar_botao(driver, wait, seletores, "Proxima Fatura")


STATUS_AUDITORIA_OK = {
    "sucesso_auditoria",
    "auditoria_sem_valor",
    "pulado_referencia_existente",
    "pulado_carimbo_existente",
}


def _status_auditoria_row(row: dict) -> str:
    return extrair_status_auditoria(row)


def _migrar_auditoria_legacy(caminho: Path) -> None:
    if not caminho.exists():
        return

    try:
        with caminho.open("r", newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f, delimiter=";"))
    except Exception:
        return

    if not rows:
        return

    header = [str(col).strip() for col in rows[0]]
    header_legacy = [
        "linha_excel",
        "instalacao",
        "data_referencia_esperada",
        "carimbo",
        "valor_auditoria",
        "status",
    ]
    header_atual = [
        "linha_excel",
        "instalacao",
        "data_referencia_esperada",
        "carimbo",
        "valor_auditoria",
        "pct_diferenca",
        "itens_divergentes",
        "memoria_calculo",
        "status",
    ]

    if header == header_atual or header != header_legacy:
        if header == header_atual:
            return
        header_sem_memoria = [
            "linha_excel",
            "instalacao",
            "data_referencia_esperada",
            "carimbo",
            "valor_auditoria",
            "pct_diferenca",
            "itens_divergentes",
            "status",
        ]
        if header == header_sem_memoria:
            convertidas = [header_atual]
            for row in rows[1:]:
                vals = list(row) + [""] * max(0, 8 - len(row))
                convertidas.append(vals[:7] + [""] + [vals[7]])
            with caminho.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerows(convertidas)
        return

    convertidas = [header_atual]
    for row in rows[1:]:
        if not row:
            continue
        vals = list(row)
        if len(vals) >= 8:
            nova = vals[:7] + [""] + [vals[7]]
        elif len(vals) == 7:
            nova = [vals[0], vals[1], vals[2], vals[3], vals[4], "", vals[5], "", vals[6]]
        else:
            vals += [""] * max(0, 6 - len(vals))
            nova = [vals[0], vals[1], vals[2], vals[3], vals[4], "", "", "", vals[5]]
        convertidas.append(nova)

    with caminho.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerows(convertidas)


def registrar_resultado_auditoria(
    linha_excel, instalacao, data_referencia_esperada, carimbo,
    valor_auditoria, status,
    pct_diferenca="", itens_divergentes="", memoria_calculo="",
):
    caminho = PASTA_SAIDA / "auditoria_resultados.csv"
    upsert_resultado_auditoria(caminho, {
        "linha_excel": linha_excel,
        "instalacao": instalacao,
        "data_referencia_esperada": data_referencia_esperada,
        "carimbo": carimbo,
        "valor_auditoria": valor_auditoria,
        "pct_diferenca": pct_diferenca,
        "itens_divergentes": itens_divergentes,
        "memoria_calculo": memoria_calculo,
        "status": status,
    }, key_fields=("carimbo", "data_referencia_esperada"))
    log(f"Resultado da auditoria registrado em: {caminho}")


def _parse_decimal_brl(valor: str) -> float:
    texto = str(valor or "").strip()
    if not texto:
        return 0.0
    texto = texto.replace("R$", "").replace("%", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(texto)
    except Exception:
        return 0.0


def _capturar_resumo_calculo(driver) -> dict[str, str]:
    ids = (
        "calcValorNF",
        "calcValorFatura",
        "calcFatorImpostoCaculado",
        "fatCalcDiferenca",
        "fatCalcDiferencaPerc",
    )
    resumo: dict[str, str] = {}
    for field_id in ids:
        try:
            elem = driver.find_element(By.ID, field_id)
        except Exception:
            continue
        valor = (elem.get_attribute("value") or elem.text or "").strip()
        if valor:
            resumo[field_id] = valor
    return resumo


def _capturar_detalhe_auditoria(driver) -> tuple:
    """
    Le a pagina de auditoria do Consen e retorna:
      (total_auditoria, pct_diferenca, itens_divergentes, memoria_calculo)
    - total_auditoria : valor do 1o span.auditoria.sucesso  (ex: '-0,01')
    - pct_diferenca   : valor do 2o span.auditoria.sucesso  (ex: '0,00%')
    - itens_divergentes: linhas da tabela onde Diferenca != 'R$ 0,00',
                         formato: 'Descricao=R$-0,03|Total Fatura=R$-0,01'
    """
    total = ""
    pct   = ""
    itens = ""
    memoria = ""

    try:
        spans = driver.find_elements(By.CSS_SELECTOR, "span.auditoria.sucesso")
        if spans:
            total = spans[0].text.strip()
        if len(spans) >= 2:
            pct = spans[1].text.strip()
        log(f"Auditoria capturada: total={total!r}  pct={pct!r}")
    except Exception as _e:
        warn(f"Nao foi possivel capturar spans de auditoria: {_e}")

    try:
        tabela = driver.find_element(By.CSS_SELECTOR, "table.table-bordered")
        linhas = tabela.find_elements(By.TAG_NAME, "tr")
        divs = []
        memoria_rows = []
        for linha in linhas[1:]:  # pula cabecalho
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if len(colunas) >= 7:
                valores = [c.text.strip() for c in colunas[:7]]
                descricao = valores[0]
                diferenca = valores[6]
                memoria_rows.append({
                    "descricao": descricao,
                    "volumes_fatura": valores[1],
                    "tarifa_calculo": valores[2],
                    "tarifa_calculo_com_imposto": valores[3],
                    "valores_fatura": valores[4],
                    "valores_calculados": valores[5],
                    "diferenca_auditoria": diferenca,
                })
                # guarda apenas linhas com diferenca real (ignora 0,00 e vazio)
                if diferenca and "0,00" not in diferenca:
                    divs.append(f"{descricao}={diferenca}")
        itens = "|".join(divs)
        memoria = json.dumps(memoria_rows, ensure_ascii=False)
        if itens:
            log(f"Itens divergentes: {itens}")
    except Exception as _e:
        warn(f"Nao foi possivel ler tabela de auditoria: {_e}")

    return total, pct, itens, memoria


_TRACE_CAMPOS_CRITICOS = (
    "dataVencimento",
    "dataEmissao",
    "txt-consumo-registrada-fpind",
    "txt-consumo-faturada-fpind",
    "txt-consumo-fpind-valor-reais",
    "txt-consumo-injetado-registrado-fpta",
    "txt-consumo-injetado-faturado-fpta",
    "txt-consumo-injetado-fpta-valor-reais",
    "txt-consumo-injetado-usina-fpta",
    "txt-consumo-injetado-usina-fpta-saldo",
    "fatValBandeira",
    "fatValBandeira2",
    "camposFinanIlumimnacaoPublica",
    "camposFinanICMS",
    "fatDesIcmsAliquota",
    "fatDescPisAliquota",
    "txt-dados-financeiros-pis-pasep",
    "fatDesCofinsAliquota",
    "txt-dados-financeiros-cofins",
    "fatDescPisPercRetImposto",
    "fatDescPisValRetImposto",
    "fatDescCofinsPercRetImposto",
    "fatDescCofinsValRetImposto",
    "fatDescCsllPercRetImposto",
    "fatDescCsllValRetImposto",
    "fatDescIrpjPercRetImposto",
    "fatDescIrpjValRetImposto",
    "txt-dados-financeiros-valor-nota-fiscal",
    "txt-dados-financeiros-valor-fatura-a-pagar",
)


def _capturar_campos_criticos_tela(driver) -> dict[str, str]:
    resumo: dict[str, str] = {}
    for campo in _TRACE_CAMPOS_CRITICOS:
        elemento = localizar_elemento_por_id_ou_name(driver, campo_id=campo, campo_name=campo)
        if elemento is None:
            resumo[campo] = "<nao_encontrado>"
        else:
            resumo[campo] = _valor_elemento_seguro(elemento)
    return resumo


def salvar_abrir_auditoria_capturar_fechar(driver, wait, registro_planilha):
    aba_principal = driver.current_window_handle
    abas_antes = driver.window_handles[:]

    # Blindagem: avisa se calcValorNF e calcValorFatura estiverem ambos zerados.
    # O preenchimento pode nao ter surtido efeito. Salva mesmo assim mas pausa
    # para inspeção manual (CONSEN_INVESTIGAR_ZEROS=1 força pausa interativa).
    resumo_pre = _capturar_resumo_calculo(driver)
    calc_nf_texto = resumo_pre.get("calcValorNF", "")
    calc_fat_texto = resumo_pre.get("calcValorFatura", "")
    calc_nf_disponivel = bool(str(calc_nf_texto).strip())
    calc_fat_disponivel = bool(str(calc_fat_texto).strip())
    nf_pre  = _parse_decimal_brl(calc_nf_texto)
    fat_pre = _parse_decimal_brl(calc_fat_texto)
    _zeros_detectados = (
        calc_nf_disponivel
        and calc_fat_disponivel
        and nf_pre == 0.0
        and fat_pre == 0.0
    )
    dados_planilha = registro_planilha.get("dados_completos", {}) or {}
    _trace_campos_evento(
        "pre_salvar",
        carimbo=registro_planilha.get("fatCarimbo", ""),
        instalacao=registro_planilha.get("instalacao", ""),
        referencia=formatar_ddmmyyyy(registro_planilha["dataReferenciaEsperada"]),
        url=getattr(driver, "current_url", ""),
        resumo_calculo=resumo_pre,
        campos=_capturar_campos_criticos_tela(driver),
    )
    valor_nf_esperado = _parse_decimal_brl(dados_planilha.get("fatValorNotaFiscal", ""))
    valor_fatura_esperado = _parse_decimal_brl(dados_planilha.get("fatValorFatura", ""))
    if not calc_nf_disponivel and not calc_fat_disponivel:
        warn(
            "Resumo de calculo do Consen nao esta disponivel antes de salvar "
            "(calcValorNF/calcValorFatura vazios ou ausentes). Seguindo para salvar."
        )
    elif _zeros_detectados:
        if abs(valor_nf_esperado) > 0 or abs(valor_fatura_esperado) > 0:
            raise RuntimeError(
                "Calculo do Consen permaneceu zerado antes de salvar "
                f"(calcValorNF={resumo_pre.get('calcValorNF', '')}, "
                f"calcValorFatura={resumo_pre.get('calcValorFatura', '')}, "
                f"valor_nf_esperado={dados_planilha.get('fatValorNotaFiscal', '')}, "
                f"valor_fatura_esperado={dados_planilha.get('fatValorFatura', '')})."
            )
        warn(
            "ATENCAO: calcValorNF e calcValorFatura estao zerados antes de salvar — "
            "formulario pode nao ter sido preenchido corretamente. Salvando mesmo assim."
        )
        if os.environ.get("CONSEN_INVESTIGAR_ZEROS", "0") == "1":
            input(
                ">>> [INVESTIGACAO] Valores zerados detectados. Inspecione o Consen e "
                "pressione ENTER para salvar e continuar..."
            )

    log("Clicando em Salvar...")
    clicar_botao_salvar(driver, wait)
    _aguardar_sem_spinner(driver, timeout=10, min_wait=0.5)
    _verificar_erros_validacao_pos_salvar(driver)
    resumo_pos_salvar = _capturar_resumo_calculo(driver)
    _trace_campos_evento(
        "pos_salvar",
        carimbo=registro_planilha.get("fatCarimbo", ""),
        instalacao=registro_planilha.get("instalacao", ""),
        referencia=formatar_ddmmyyyy(registro_planilha["dataReferenciaEsperada"]),
        url=getattr(driver, "current_url", ""),
        resumo_calculo=resumo_pos_salvar,
        campos=_capturar_campos_criticos_tela(driver),
    )

    try:
        gravar_salvar_confirmado(
            PASTA_SAIDA,
            registro_planilha.get("fatCarimbo", ""),
            registro_planilha["instalacao"],
            formatar_ddmmyyyy(registro_planilha["dataReferenciaEsperada"]),
        )
    except Exception:
        pass

    try:
        WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.ID, "linkAuditoria")))
    except Exception:
        _s(0.5)

    if os.environ.get("CONSEN_PARAR_ANTES_AUDITORIA", "0") == "1":
        input(">>> Salvo. Verifique o Consen e pressione ENTER para continuar para Auditoria...")
    log("Clicando em Auditoria...")
    clicar_link_auditoria(driver, wait)

    nova_aba = None
    for _ in range(20):
        abas_agora = driver.window_handles
        if len(abas_agora) > len(abas_antes):
            novas = [h for h in abas_agora if h not in abas_antes]
            if novas:
                nova_aba = novas[0]
                break
        _s(0.2)

    if nova_aba:
        driver.switch_to.window(nova_aba)
        log("Mudou para a aba da auditoria.")
    else:
        log("Auditoria não abriu nova aba; seguindo na aba atual.")

    try:
        WebDriverWait(driver, CONSEN_TIMEOUT_AUDITORIA).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.auditoria.sucesso")))
    except Exception:
        warn("Timeout aguardando span.auditoria.sucesso")

    valor_auditoria, pct_diferenca, itens_divergentes, memoria_calculo = _capturar_detalhe_auditoria(driver)
    resumo_calculo = _capturar_resumo_calculo(driver)
    _trace_campos_evento(
        "auditoria",
        carimbo=registro_planilha.get("fatCarimbo", ""),
        instalacao=registro_planilha.get("instalacao", ""),
        referencia=formatar_ddmmyyyy(registro_planilha["dataReferenciaEsperada"]),
        url=getattr(driver, "current_url", ""),
        valor_auditoria=valor_auditoria,
        pct_diferenca=pct_diferenca,
        itens_divergentes=itens_divergentes,
        memoria_calculo=memoria_calculo,
        resumo_calculo=resumo_calculo,
    )
    status_auditoria = "sucesso_auditoria" if valor_auditoria else "auditoria_sem_valor"

    # Quando o Consen "parece" salvar mas a auditoria real não abre,
    # esse resumo fica preenchido com diferencas positivas e o caso
    # nao deve entrar como salvo.
    if not memoria_calculo and resumo_calculo:
        diferenca_rs = _parse_decimal_brl(resumo_calculo.get("fatCalcDiferenca", ""))
        diferenca_pct = _parse_decimal_brl(resumo_calculo.get("fatCalcDiferencaPerc", ""))
        if abs(diferenca_rs) > 0 or abs(diferenca_pct) > 0:
            status_auditoria = "erro_no_fluxo:salvamento_invalido"
            itens_resumo = [f"{k}={v}" for k, v in resumo_calculo.items()]
            itens_divergentes = "|".join(filter(None, [itens_divergentes, *itens_resumo]))
            warn(
                "Resumo de calculo indica salvamento invalido "
                f"(dif_R$={resumo_calculo.get('fatCalcDiferenca', '')}, "
                f"dif_%={resumo_calculo.get('fatCalcDiferencaPerc', '')})."
            )

    registrar_resultado_auditoria(
        linha_excel=registro_planilha["linha_excel"],
        instalacao=registro_planilha["instalacao"],
        data_referencia_esperada=formatar_ddmmyyyy(registro_planilha["dataReferenciaEsperada"]),
        carimbo=registro_planilha.get("fatCarimbo", ""),
        valor_auditoria=valor_auditoria,
        pct_diferenca=pct_diferenca,
        itens_divergentes=itens_divergentes,
        memoria_calculo=memoria_calculo,
        status=status_auditoria,
    )

    if nova_aba:
        driver.close()
        driver.switch_to.window(aba_principal)
        log("Retornou à aba principal.")

    try:
        log("Clicando em Proxima Fatura...")
        clicar_botao_proxima_fatura(driver, wait)
        _aguardar_sem_spinner(driver, timeout=5, min_wait=0.2)
    except Exception as _e_prox:
        # btnProxima pode não existir ou não funcionar neste contexto do Consen.
        # A auditoria já foi registrada; a próxima iteração navega via
        # voltar_para_tela_inicial_instalacao, então a falha aqui não é crítica.
        warn(f"Proxima Fatura nao clicada ({type(_e_prox).__name__}) — navegacao sera feita no proximo ciclo.")

    return valor_auditoria


# =========================================================
# FLUXO
# =========================================================

def voltar_para_tela_inicial_instalacao(driver, wait):
    log("Voltando para a tela inicial de instalação...")
    abrir_tela_instalacao(driver, wait)
    _aguardar_sem_spinner(driver, timeout=5, min_wait=0.2)

    try:
        aguardar_tela_instalacao_pronta(driver, wait, timeout=8, pausa_extra=0.5)
    except Exception:
        # Campo instalacao não apareceu — força navegação completa para TARGET_URL
        warn("Campo instalacao não encontrado após clique JS. Forçando navegação completa para TARGET_URL...")
        try:
            driver.get(TARGET_URL)
            _smin(1.5, 0.5)
            _aguardar_sem_spinner(driver, timeout=8, min_wait=0.3)
        except Exception as nav_e:
            warn(f"Falha na navegação forçada: {type(nav_e).__name__}")
        aguardar_tela_instalacao_pronta(driver, wait, timeout=12, pausa_extra=0.5)


def fechar_driver_seguro(driver):
    if not driver:
        return
    profile_dir = getattr(driver, "_codex_profile_dir", "")
    try:
        driver.quit()
    except Exception:
        pass
    if profile_dir:
        try:
            shutil.rmtree(profile_dir, ignore_errors=True)
        except Exception:
            pass


def abrir_driver_logado():
    driver = iniciar_driver(headless=HEADLESS)
    wait = WebDriverWait(driver, 20)
    log("Abrindo login...")
    driver.get(LOGIN_URL)
    log("Fazendo login...")
    _fazer_login_com_retry(driver, wait, USUARIO, SENHA)
    _smin(1.5, 0.5)
    try:
        _aguardar_sem_spinner(driver, timeout=8, min_wait=0.3)
    except Exception:
        pass
    try:
        abrir_tela_instalacao(driver, wait)
        _aguardar_sem_spinner(driver, timeout=8, min_wait=0.3)
    except Exception as e:
        warn(f"Falha ao preparar a tela Instalacao logo apos o login: {type(e).__name__}")
    return driver, wait


def carregar_linhas_ja_digitadas() -> tuple[set[int], set[str]]:
    """
    Lê saida_importacao/auditoria_resultados.csv e retorna o set de
    linha_excel que já foram digitadas com sucesso ou puladas por já existirem.
    Linhas com status de erro NÃO entram no set — serão tentadas novamente.
    """
    caminho = PASTA_SAIDA / "auditoria_resultados.csv"
    if not caminho.exists():
        return set(), set()

    ja_feitas_linhas = set()
    ja_feitos_carimbos = set()
    status_ok = STATUS_AUDITORIA_OK

    try:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with caminho.open("r", newline="", encoding=enc) as f:
                    reader = csv.DictReader(f, delimiter=";")
                    for row in reader:
                        status = _status_auditoria_row(row).strip().lower()
                        if status in status_ok:
                            carimbo = normalizar_carimbo(row.get("carimbo", ""))
                            if carimbo:
                                ja_feitos_carimbos.add(carimbo)
                            try:
                                ja_feitas_linhas.add(int(row["linha_excel"]))
                            except (KeyError, ValueError):
                                pass
                break
            except UnicodeDecodeError:
                continue
    except Exception as e:
        warn(f"Não consegui ler auditoria_resultados.csv: {e}")

    if ja_feitas_linhas or ja_feitos_carimbos:
        log(
            f"Auditoria: {len(ja_feitas_linhas)} linha(s) e "
            f"{len(ja_feitos_carimbos)} carimbo(s) já digitados — serão pulados automaticamente."
        )
    return ja_feitas_linhas, ja_feitos_carimbos


def filtrar_registros_por_pdfs_existentes(registros: list[dict]) -> list[dict]:
    """
    Quando a digitação roda pelo pipeline, mantém somente linhas cujo PDF
    original ainda existe na pasta atual do lote.
    Se ENEL_DIGITACAO_SOMENTE_NOMES estiver definido, restringe adicionalmente
    ao conjunto de nomes listados no arquivo manifesto (um nome por linha).
    """
    if not PASTA_PDFS_ATUAIS:
        return registros

    pasta = Path(PASTA_PDFS_ATUAIS)
    if not pasta.exists():
        warn(f"Pasta de PDFs da digitação não encontrada: {pasta}")
        return registros

    arquivos_existentes = {
        pdf.name.strip().lower()
        for pdf in pasta.rglob("*.pdf")
        if pdf.is_file()
    }
    if not arquivos_existentes:
        warn(f"Nenhum PDF encontrado na pasta informada para a digitação: {pasta}")
        return []

    # Filtro adicional por manifesto de nomes (apenas PDFs baixados nesta sessão)
    if ENEL_SOMENTE_NOMES_FILE:
        manifesto = Path(ENEL_SOMENTE_NOMES_FILE)
        if manifesto.exists():
            nomes_sessao = {
                linha.strip().lower()
                for linha in manifesto.read_text(encoding="utf-8").splitlines()
                if linha.strip()
            }
            antes = len(arquivos_existentes)
            arquivos_existentes &= nomes_sessao
            log(
                f"[manifesto] Restrição a PDFs da sessão: {antes} na pasta → "
                f"{len(arquivos_existentes)} permitidos pelo manifesto ({ENEL_SOMENTE_NOMES_FILE})."
            )
        else:
            warn(f"[manifesto] Arquivo ENEL_DIGITACAO_SOMENTE_NOMES não encontrado: {manifesto} — ignorado.")

    total_antes = len(registros)
    filtrados = []
    pulados_sem_pdf = 0

    for registro in registros:
        arquivo_raw = str(registro.get("dados_completos", {}).get("ARQUIVO") or "").strip()
        arquivo = _os_dig.path.basename(arquivo_raw).lower() if arquivo_raw else ""
        if arquivo and arquivo in arquivos_existentes:
            filtrados.append(registro)
            continue
        pulados_sem_pdf += 1

    log(
        f"Filtro por PDFs atuais: {total_antes} linhas na planilha -> "
        f"{len(filtrados)} com arquivo presente em {pasta} ({pulados_sem_pdf} puladas)."
    )
    return filtrados


def validar_escopo_digitacao() -> None:
    if PASTA_PDFS_ATUAIS:
        log(f"Escopo restrito por pasta de PDFs: {PASTA_PDFS_ATUAIS}")
        if ENEL_SOMENTE_NOMES_FILE:
            log(f"Escopo adicional por manifesto de sessão: {ENEL_SOMENTE_NOMES_FILE}")
        return
    if CONSEN_PERMITIR_LOTE_COMPLETO:
        log("Escopo liberado para lote completo pela pipeline (CONSEN_PERMITIR_LOTE_COMPLETO=1).")
        return
    raise RuntimeError(
        "Digitacao ENEL bloqueada sem filtro de PDFs. "
        "Informe ENEL_DIGITACAO_PASTA_PDFS para execucoes isoladas "
        "ou defina CONSEN_PERMITIR_LOTE_COMPLETO=1 apenas nas pipelines de lote inteiro."
    )


def _erro_planilha_bloqueia_digitacao(erro_planilha: str) -> bool:
    erro = str(erro_planilha or "").strip().upper()
    if not erro:
        return False
    # Faturas zeradas sao validas para o CONSEN; nao devem ser barradas.
    if erro == "FATURA_ZERADA":
        return False
    return True


_CAMPOS_OBRIGATORIOS: tuple[str, ...] = ("fatDataVcto",)


def _campos_obrigatorios_ausentes(registro: dict) -> list[str]:
    """Retorna lista de campos obrigatórios ausentes em um registro.

    Verifica antes de abrir o driver/Chrome. fatDataVcto ausente causaria
    que o CONSEN mostrasse 'This field is required' após Salvar, mas o
    pipeline antigo retornava exit 0 silenciosamente.
    """
    ausentes: list[str] = []
    dados = registro.get("dados_completos") or registro
    for campo in _CAMPOS_OBRIGATORIOS:
        val = dados.get(campo) or registro.get(campo)
        if val is None or str(val).strip() in ("", "0", "0.0", "None"):
            ausentes.append(campo)
    return ausentes


def main():
    driver = None

    try:
        validar_escopo_digitacao()
        registros = ler_todas_as_linhas_planilha(EXCEL_PATH, linha_inicio=LINHA_INICIO)
        registros = filtrar_registros_por_pdfs_existentes(registros)

        ja_digitadas_linhas, ja_digitadas_carimbos = carregar_linhas_ja_digitadas()
        if ja_digitadas_linhas or ja_digitadas_carimbos:
            total_antes = len(registros)
            filtrados = []
            puladas = 0
            for r in registros:
                carimbo = normalizar_carimbo(r.get("fatCarimbo", ""))
                if carimbo and carimbo in ja_digitadas_carimbos:
                    puladas += 1
                    continue
                if not carimbo and r["linha_excel"] in ja_digitadas_linhas:
                    puladas += 1
                    continue
                filtrados.append(r)
            registros = filtrados
            log(
                f"Filtro de já digitadas: {total_antes} linhas na planilha -> "
                f"{len(registros)} novas para processar ({puladas} puladas)."
            )
        else:
            log(f"Nenhuma digitação anterior registrada. Processando todas as {len(registros)} linhas.")

        if not registros:
            log("Nenhuma linha nova para digitar. Encerrando.")
            gerar_csv_pendentes()
            return

        registros_validos = []
        registros_bloqueados = []
        for registro in registros:
            if _erro_planilha_bloqueia_digitacao(registro.get("erro_planilha", "")):
                registros_bloqueados.append(registro)
            else:
                registros_validos.append(registro)

        for registro in registros_bloqueados:
            motivo = registro.get("erro_planilha", "")
            warn(
                f"Linha {registro['linha_excel']} bloqueada antes da digitacao "
                f"({registro['instalacao']}): {motivo}"
            )
            registrar_resultado_auditoria(
                linha_excel=registro["linha_excel"],
                instalacao=registro["instalacao"],
                data_referencia_esperada=formatar_ddmmyyyy(registro["dataReferenciaEsperada"]),
                carimbo=registro.get("fatCarimbo", ""),
                valor_auditoria="",
                status="erro_extracao",
                itens_divergentes=motivo,
            )

        registros = registros_validos
        if not registros:
            log("Nenhuma linha apta para digitacao apos filtrar erros de extracao.")
            gerar_csv_pendentes()
            return

        registros_validos = []
        for registro in registros:
            ok_bb, motivo_bb = validar_registro_bb_arquivo(registro)
            if ok_bb:
                registros_validos.append(registro)
                continue

            detalhe = motivo_bb
            pdf_invalido = _resolver_pdf_registro(registro)
            if pdf_invalido is not None:
                detalhe = mover_pdf_para_investigar(pdf_invalido, motivo_bb)

            warn(
                f"Linha {registro['linha_excel']} bloqueada antes da digitacao "
                f"({registro['instalacao']}): {detalhe}"
            )
            registrar_resultado_auditoria(
                linha_excel=registro["linha_excel"],
                instalacao=registro["instalacao"],
                data_referencia_esperada=formatar_ddmmyyyy(registro["dataReferenciaEsperada"]),
                carimbo=registro.get("fatCarimbo", ""),
                valor_auditoria="",
                status="erro_arquivo_invalido",
                itens_divergentes=detalhe,
            )

        registros = registros_validos
        if not registros:
            log("Nenhuma linha apta para digitacao apos validar o nome BB_ dos PDFs.")
            gerar_csv_pendentes()
            return

        # Validação pré-driver: rejeita linhas com campos obrigatórios ausentes
        # ANTES de abrir o Chrome. fatDataVcto ausente causaria salvamento silencioso.
        registros_validos = []
        for registro in registros:
            campos_ausentes = _campos_obrigatorios_ausentes(registro)
            if campos_ausentes:
                motivo = f"campo(s) obrigatorio(s) ausente(s): {campos_ausentes}"
                warn(
                    f"Linha {registro['linha_excel']} bloqueada pre-CONSEN "
                    f"({registro['instalacao']}): {motivo}"
                )
                registrar_resultado_auditoria(
                    linha_excel=registro["linha_excel"],
                    instalacao=registro["instalacao"],
                    data_referencia_esperada=formatar_ddmmyyyy(registro["dataReferenciaEsperada"]),
                    carimbo=registro.get("fatCarimbo", ""),
                    valor_auditoria="",
                    status="erro_campo_obrigatorio_ausente",
                    itens_divergentes=motivo,
                )
            else:
                registros_validos.append(registro)

        registros = registros_validos
        if not registros:
            log("Nenhuma linha apta apos validacao de campos obrigatorios pre-CONSEN.")
            gerar_csv_pendentes()
            return

        driver, wait = abrir_driver_logado()

        total_auditorias = 0

        for idx, registro in enumerate(registros, start=1):
            tentativa_linha = 0
            sucesso_linha = False

            while tentativa_linha < 3 and not sucesso_linha:
                tentativa_linha += 1
                try:
                    if CONSEN_REINICIAR_NAVEGADOR_CADA_LINHA and idx > 1 and tentativa_linha == 1:
                        log("Reiniciando navegador antes da proxima linha (modo de teste)...")
                        fechar_driver_seguro(driver)
                        driver, wait = abrir_driver_logado()

                    log("=" * 80)
                    log(
                        f"PROCESSANDO LINHA {registro['linha_excel']} | instalação={registro['instalacao']} "
                        f"| item {idx}/{len(registros)} | tentativa {tentativa_linha}/3"
                    )

                    if _esta_na_login_page(driver):
                        warn("Sessão Consen expirada (redirecionado para login). Re-autenticando...")
                        _fazer_login_com_retry(driver, wait, USUARIO, SENHA)
                        time.sleep(1.0)
                        try:
                            _aguardar_sem_spinner(driver, timeout=8, min_wait=0.3)
                        except Exception:
                            pass

                    carimbo_atual = registro.get("fatCarimbo", "")
                    if carimbo_atual and salvar_confirmado(PASTA_SAIDA, carimbo_atual):
                        warn(
                            f"SALVAR_CONFIRMADO detectado para {carimbo_atual} — "
                            "consultando CONSEN antes de qualquer redigitacao."
                        )

                    voltar_para_tela_inicial_instalacao(driver, wait)

                    log("Preparando campo de instalação...")
                    aguardar_tela_instalacao_pronta(driver, wait, timeout=8, pausa_extra=0.5)

                    log("Digitando instalação...")
                    preencher_input_texto(driver, wait, "instalacao", registro["instalacao"], pausa_antes=0.3)

                    log("Clicando em Carregar da instalação...")
                    clicar_botao_carregar_instalacao(driver, wait)
                    _aguardar_sem_spinner(driver, timeout=8, min_wait=0.3)
                    aguardar_carregamento_tabela(driver, wait)

                    prosseguir_fluxo = preencher_datas_e_carimbo_se_necessario(driver, wait, registro)
                    if not prosseguir_fluxo:
                        registrar_resultado_auditoria(
                            linha_excel=registro["linha_excel"],
                            instalacao=registro["instalacao"],
                            data_referencia_esperada=formatar_ddmmyyyy(registro["dataReferenciaEsperada"]),
                            carimbo=registro.get("fatCarimbo", ""),
                            valor_auditoria="",
                            status="erro_referencia_nao_abriu",
                        )
                        sucesso_linha = True
                        continue

                    log("Preenchendo campos forçados essenciais...")
                    log("Definindo classificacao contratual no inicio...")
                    preencher_classificacao_contratual_inicio(driver, registro["dados_completos"])

                    _, _, campos_alias_ok, campos_alias_falhos = preencher_aliases_forcados_linha(
                        driver, registro["dados_completos"]
                    )

                    log("Preenchendo saldos acumulados da usina (tratamento especial)...")
                    preencher_saldos_injetado_usina(driver, wait, registro["dados_completos"])

                    log("Preenchendo observações múltiplas...")
                    preencher_obs_multiplas(driver, wait, registro["dados_completos"])

                    caminho_csv_existente = PASTA_SAIDA / "mapeamento_campos_planilha.csv"
                    if not caminho_csv_existente.exists() or idx == 1:
                        log("Exportando campos disponíveis e correlação...")
                        exportar_campos_e_correlacoes(driver, EXCEL_PATH)

                    caminho_csv = escolher_csv_mapeamento()
                    log("Preenchendo campos com base no CSV de mapeamento...")
                    preencher_campos_via_mapeamento_csv_linha(
                        driver=driver,
                        dados_planilha=registro["dados_completos"],
                        caminho_csv=caminho_csv,
                        score_minimo=SCORE_MINIMO_MAPEAMENTO,
                        campos_alias_ok=campos_alias_ok,
                        campos_alias_falhos=campos_alias_falhos,
                    )

                    salvar_abrir_auditoria_capturar_fechar(driver, wait, registro)
                    total_auditorias += 1
                    sucesso_linha = True

                except Exception as e:
                    warn(
                        f"Falha na linha {registro['linha_excel']} ({registro['instalacao']}) "
                        f"[tentativa {tentativa_linha}/3]: {type(e).__name__} - {e}"
                    )

                    if _erro_driver_recuperavel(e):
                        if tentativa_linha < 3:
                            warn("Sessao/driver do navegador perdeu estabilidade. Reabrindo e repetindo a mesma linha...")
                            fechar_driver_seguro(driver)
                            driver, wait = abrir_driver_logado()
                            continue
                    elif _esta_na_login_page(driver):
                        warn("Sessão Consen expirada (detectada no handler de erro). Re-autenticando...")
                        try:
                            _fazer_login_com_retry(driver, wait, USUARIO, SENHA)
                            time.sleep(1.0)
                            if tentativa_linha < 3:
                                continue
                        except Exception as re_e:
                            warn(f"Re-login falhou: {re_e}. Reabrindo driver...")
                            fechar_driver_seguro(driver)
                            driver, wait = abrir_driver_logado()
                            if tentativa_linha < 3:
                                continue

                    if tentativa_linha >= 3:
                        registrar_resultado_auditoria(
                            linha_excel=registro["linha_excel"],
                            instalacao=registro["instalacao"],
                            data_referencia_esperada=formatar_ddmmyyyy(registro["dataReferenciaEsperada"]),
                            carimbo=registro.get("fatCarimbo", ""),
                            valor_auditoria="",
                            status=f"erro_no_fluxo:{type(e).__name__}",
                        )

        print(f"Processo concluído. Linhas com auditoria registrada: {total_auditorias}")
        gerar_csv_pendentes()

    except Exception as e:
        erro(str(e))
        sys.exit(1)

    finally:
        if driver:
            if CONSEN_INTERATIVO_FECHAR:
                input("Pressione ENTER para fechar o navegador...")
            fechar_driver_seguro(driver)


def gerar_csv_pendentes():
    """
    Lê auditoria_resultados.csv e gera pendentes_HHMMSS.csv com as contas
    que não foram auditadas com sucesso.
    Colunas: instalacao, carimbo, status, linha_excel.
    """
    caminho_auditoria = PASTA_SAIDA / "auditoria_resultados.csv"
    if not caminho_auditoria.exists():
        warn("auditoria_resultados.csv não encontrado — nenhum pendente gerado.")
        return

    STATUS_OK = {"sucesso_auditoria", "pulado_referencia_existente"}

    pendentes = []
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with caminho_auditoria.open("r", newline="", encoding=enc) as f:
                for row in csv.DictReader(f, delimiter=";"):
                    status = _status_auditoria_row(row).strip()
                    if status not in STATUS_OK:
                        pendentes.append({
                            "instalacao":  row.get("instalacao", ""),
                            "carimbo":     row.get("carimbo", ""),
                            "status":      status,
                            "linha_excel": row.get("linha_excel", ""),
                        })
            break
        except UnicodeDecodeError:
            continue

    if not pendentes:
        log("Nenhum pendente — todas as contas foram digitadas com sucesso.")
        return

    ts = datetime.now().strftime("%H%M%S")
    caminho_saida = PASTA_SAIDA / f"pendentes_{ts}.csv"
    with caminho_saida.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f, fieldnames=["instalacao", "carimbo", "status", "linha_excel"], delimiter=";"
        )
        writer.writeheader()
        writer.writerows(pendentes)

    log(f"Pendentes: {len(pendentes)} conta(s) -> {caminho_saida}")


if __name__ == "__main__":
    main()
