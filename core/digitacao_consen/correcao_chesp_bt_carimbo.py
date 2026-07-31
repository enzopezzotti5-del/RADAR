#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Correcao CHESP BT por carimbo.

Le o XLSX de OCR ja gerado e corrige no CONSEN os campos que a digitacao
errou (subgrupo, tarifa, consumo, bandeira, aliquotas, retencoes).

Uso:
    python correcao_chesp_bt_carimbo.py --carimbo BB_2017132 [--salvar]
    python correcao_chesp_bt_carimbo.py --carimbos-arquivo lista.txt --salvar
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from digitacao_consen.consen_api import (
        ConsenSession,
        carregar_lista_carimbos_args,
        formatar_correcao,
    )
    from digitacao_consen.correcao_fluxo_base import normalizar_carimbo, log, warn
except ModuleNotFoundError:
    from consen_api import ConsenSession, carregar_lista_carimbos_args, formatar_correcao  # type: ignore
    from correcao_fluxo_base import normalizar_carimbo, log, warn  # type: ignore

# ── Configuração (único por distribuidora) ────────────────────────────────────

SAIDA_DIR = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/CHESP_pipeline_saida/correcoes_por_carimbo")
EXECUCAO_CSV = SAIDA_DIR / "correcao_execucao.csv"
XLSX_PADRAO = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/OCR CHESP/ocr_chesp_BT_062026.xlsx")

CAMPOS_CRITICOS: tuple[str, ...] = (
    "fatConFPontaIndRegistrado",
    "fatDescIrpjPercRetImposto",
)

ORDEM_CAMPOS: tuple[str, ...] = (
    "cb-tarifa",
    "cb-subgrupo",
    "fatConPontaRegistrado",
    "fatConPonta",
    "fatConPontaValorReais",
    "fatConIntermediarioRegistrado",
    "fatConIntermediario",
    "fatConIntermediarioValorReais",
    "fatConFPontaIndRegistrado",
    "fatConFPontaIndutivo",
    "fatConFPontaIndValorReais",
    "fatICMS",
    "fatDescPisAliquota",
    "fatDesCofinsAliquota",
    "fatValBandeira",
    "fatDescIrpjPercRetImposto",
    "fatDescConsumoPercRetImposto",
    "fatDescConsumoValRetImposto",
)


# ── Lógica única: leitura do XLSX + mapeamento OCR → CONSEN ──────────────────

def _ler_xlsx(xlsx: Path, carimbo_norm: str) -> dict[str, Any] | None:
    if not xlsx.exists():
        warn(f"XLSX nao encontrado: {xlsx}")
        return None
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        car = str(rec.get("fatCarimbo") or "").strip()
        if normalizar_carimbo(car) == carimbo_norm:
            return rec
    return None


def _build_correcoes(rec: dict[str, Any]) -> dict[str, str]:
    """Mapeia campos do OCR para IDs do formulário CONSEN. (Único por distribuidora)"""
    def fmt(header: str, valor: Any) -> str:
        return formatar_correcao(header, valor)

    zero = fmt("fatConPontaRegistrado", 0)
    return {
        "cb-tarifa":                        str(rec.get("cadTarifaCod") or "Convencional").strip(),
        "cb-subgrupo":                      str(rec.get("cadSubGrupoCod") or "B3 [<2,3kV]").strip(),
        "fatConPontaRegistrado":            fmt("fatConPontaRegistrado",            rec.get("fatConPontaRegistrado", 0)),
        "fatConPonta":                      fmt("fatConPontaFaturado",              rec.get("fatConPontaFaturado", 0)),
        "fatConPontaValorReais":            fmt("fatConPontaValorReais",            rec.get("fatConPontaValorReais", 0)),
        "fatConIntermediarioRegistrado":    fmt("fatConIntermediarioRegistrado",    rec.get("fatConIntermediarioRegistrado", 0)),
        "fatConIntermediario":              fmt("fatConIntermediarioFaturado",      rec.get("fatConIntermediarioFaturado", 0)),
        "fatConIntermediarioValorReais":    fmt("fatConIntermediarioValorReais",    rec.get("fatConIntermediarioValorReais", 0)),
        "fatConFPontaIndRegistrado":        fmt("fatConFPontaIndRegistrado",        rec.get("fatConFPontaIndRegistrado", 0)),
        "fatConFPontaIndutivo":             fmt("fatConFPontaIndFaturado",          rec.get("fatConFPontaIndFaturado", 0)),
        "fatConFPontaIndValorReais":        fmt("fatConFPontaIndValorReais",        rec.get("fatConFPontaIndValorReais", 0)),
        "fatDescPisAliquota":               fmt("fatDescPisAliquota",               rec.get("fatDescPisAliquota", 0)),
        "fatDesCofinsAliquota":             fmt("fatDesCofinsAliquota",             rec.get("fatDesCofinsAliquota", 0)),
        "fatValBandeira":                   fmt("fatValBandeira",                   rec.get("fatValBandeira", 0)),
        "fatDescIrpjPercRetImposto":        fmt("fatDescIrpjPercRetImposto",        rec.get("fatDescIrpjPercRetImposto", 0)),
        # Zera campos MT preenchidos erroneamente na primeira digitação
        "fatDescConsumoPercRetImposto":     zero,
        "fatDescConsumoValRetImposto":      zero,
    }


# ── Fluxo padrão (idêntico em todos os scripts que usam consen_api) ──────────

def rodar(carimbos: list[str], xlsx: Path, salvar: bool) -> dict[str, str]:
    SAIDA_DIR.mkdir(parents=True, exist_ok=True)

    if not salvar:
        log("MODO SIMULACAO: use --salvar para efetivar as correcoes.")

    resultados: dict[str, str] = {}

    with ConsenSession.abrir() as s:
        for carimbo in carimbos:
            log(f"=== BB_{carimbo} ===")
            rec = _ler_xlsx(xlsx, carimbo)
            if not rec:
                warn(f"BB_{carimbo}: nao encontrado no XLSX.")
                s.registrar(EXECUCAO_CSV, carimbo, "nao_encontrado_xlsx")
                resultados[carimbo] = "nao_encontrado_xlsx"
                continue

            correcoes = _build_correcoes(rec)
            log(f"  {len(correcoes)} campo(s) a corrigir")

            try:
                s.buscar_carimbo(carimbo, CAMPOS_CRITICOS)
                s.snapshot(SAIDA_DIR, carimbo)
                resultado = s.editar_campos(correcoes, ORDEM_CAMPOS)

                if not salvar:
                    log(f"  {resultado.resumo()} [simulado]")
                    s.registrar(EXECUCAO_CSV, carimbo, "simulado",
                                f"{resultado.n_alterados}/{len(correcoes)} campos")
                    resultados[carimbo] = "simulado"
                    continue

                s.salvar_e_auditar()
                log(f"  {resultado.resumo()} [SALVO]")
                s.registrar(EXECUCAO_CSV, carimbo, "corrigido",
                             f"{resultado.n_alterados}/{len(correcoes)} campos")
                resultados[carimbo] = "corrigido"

            except Exception as exc:
                warn(f"  BB_{carimbo}: {type(exc).__name__}: {exc}")
                s.registrar(EXECUCAO_CSV, carimbo, "erro", str(exc)[:200])
                resultados[carimbo] = "erro"

    return resultados


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Correcao CHESP BT por carimbo no CONSEN")
    p.add_argument("--carimbo", action="append", default=[],
                   help="BB_XXXXXXX ou so o numero. Repetivel.")
    p.add_argument("--carimbos-arquivo", type=str, default="",
                   help="TXT com um carimbo por linha")
    p.add_argument("--xlsx", type=str, default=str(XLSX_PADRAO),
                   help="XLSX de OCR BT gerado pelo pipeline")
    p.add_argument("--salvar", action="store_true",
                   help="Efetivar a correcao (sem esta flag apenas simula)")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    carimbos = carregar_lista_carimbos_args(args)

    if not carimbos:
        print("Informe ao menos um carimbo com --carimbo BB_XXXXXXX ou --carimbos-arquivo arquivo.txt")
        return 1

    resultados = rodar(carimbos, Path(str(args.xlsx).strip()), salvar=args.salvar)

    log("=== RESUMO ===")
    for c, s in resultados.items():
        log(f"  BB_{c}: {s}")

    return 1 if any(s == "erro" for s in resultados.values()) else 0


if __name__ == "__main__":
    raise SystemExit(main())
