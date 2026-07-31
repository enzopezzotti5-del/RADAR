#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Correcao ENEL SP MT por carimbo.

Modos:
  padrao (sem --direto): le CSV de auditoria e corrige so campos divergentes.
  --direto:              le xlsx do OCR diretamente e sobrescreve todos os campos
                         mapeados (sem etapa de comparacao).

Campos excluidos da correcao automatica:
  - fatDescontoFio / fatDescontoFioKWh  -- corretos no Consen
  - fatConFPontaCapRegistrado / Faturado -- bloqueados no formulario
  - fatBeneficioTarifarioBrutoValorReais / Liquido -- bloqueados

Uso:
    python correcao_enel_mt_carimbo.py [--salvar] [--carimbo 2008133 ...]
    python correcao_enel_mt_carimbo.py --direto --salvar [--carimbo 2008139 ...]
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from pathlib import Path
from typing import Any

import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import _venv_check  # noqa

try:
    from digitacao_consen import correcao_fluxo_base as fluxo_base
    from digitacao_consen.correcao_fluxo_base import log, warn
    from digitacao_consen.digitacao_consen_enel import (
        _aguardar_sem_spinner,
        formatar_numero_br,
        preencher_elemento_html,
    )
except ModuleNotFoundError:
    import correcao_fluxo_base as fluxo_base  # type: ignore
    from correcao_fluxo_base import log, warn  # type: ignore
    from digitacao_consen_enel import (  # type: ignore
        _aguardar_sem_spinner,
        formatar_numero_br,
        preencher_elemento_html,
    )

LOGIN_URL = fluxo_base.LOGIN_URL
BASE_URL  = LOGIN_URL.rsplit("login.php", 1)[0]

CSV_AUDITORIA = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/OCR ENEL/auditoria_mt_enel_052026.csv")
XLSX_OCR      = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/OCR ENEL/ocr_enel_MT_052026.xlsx")
CSV_EXECUCAO  = Path("//10.10.250.21/Energia/ARQUIVOS ENZO/OCR ENEL/correcao_enel_mt_execucao.csv")

# Campos que NAO devem ser atualizados (independente do modo)
CAMPOS_EXCLUIDOS = {
    "_obs_pairs",
    "_ERRO_CARGA",
}

# Mapeamento campo_xlsx -> id_html real no DOM do Consen
CAMPO_PARA_ID: dict[str, str] = {
    "fatDemFPontaIndRegistrada":       "fatDemFPontaIndRegistrada",
    "fatDemFPontaIndFaturada":         "fatDemFPontaIndutivo",
    "fatDemFPontaIndValorReais":       "fatDemFPontaIndValorReais",
    "fatDemFPontaIndUltra":            "fatDemFPontaIndUltra",
    "fatDemFPontaIndUltraValorReais":  "fatDemFPontaIndUltraValorReais",
    "fatConFPontaIndValorReais":       "fatConFPontaIndValorReais",
    "fatConFPontaIndExcRegistrado":      "fatConFPontaIndExcRegistrado",
    "fatConFPontaIndExcFaturado":        "fatConFPontaIndExc",
    "fatConFPontaIndExcValorReais":      "fatConFPontaIndExcValorReais",
    "fatEscassezHidrica":              "fatEscassezHidrica",
    "fatEscassezHidricaValorReais":    "fatEscassezHidricaValorReais",
    "fatICMS":                         "fatICMS",
    "fatIlumPublica":                  "fatIluminacaoPublica",
    "fatMultas":                       "fatMultas",
    "fatDescontoFio":                              "fatDescontoFio",
    "fatDescontoFioKWh":                           "fatDescontoFioKWh",
    "fatBeneficioTarifarioBrutoValorReais":        "fatBeneficioTarifarioBrutoValorReais",
    "fatBeneficioLiquidoValorReais":               "fatBeneficioLiquidoValorReais",
}
# compat modo auditoria
CAMPO_PARA_ABA_ID = {c: (None, v) for c, v in CAMPO_PARA_ID.items()}

OBS_SEL_ID  = "cb-dados-financeiros-obs"
OBS_VAL_ID  = "fatValorObs"
OBS_BTN_ID  = "btnIncluiLinha"


# ─── helpers ──────────────────────────────────────────────────────────────────

def _br2f(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _registrar(carimbo: str, status: str, detalhe: str = "") -> None:
    existe = CSV_EXECUCAO.exists()
    with CSV_EXECUCAO.open("a", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        if not existe:
            w.writerow(["timestamp", "carimbo", "status", "detalhe"])
        w.writerow([time.strftime("%Y-%m-%d %H:%M:%S"), carimbo, status, detalhe])


# ─── carregamento: CSV de auditoria ou xlsx direto ────────────────────────────

def _br2f_xlsx(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def carregar_xlsx_direto(xlsx_path: Path, filtro_carimbos: set[str] | None = None) -> dict[str, dict]:
    """Le o xlsx do OCR e retorna {carimbo: {"campos": {campo: valor_str}, "obs": []}}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]

    resultado: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {headers[i]: row[i] for i in range(len(headers))}
        carimbo_raw = str(rec.get("fatCarimbo") or "").strip()
        if not carimbo_raw:
            continue
        num = re.sub(r"\D", "", carimbo_raw)
        carimbo = f"BB_{num}"
        if filtro_carimbos and carimbo not in filtro_carimbos:
            continue

        campos: dict[str, str] = {}
        for campo, _html_id in CAMPO_PARA_ID.items():
            if campo in CAMPOS_EXCLUIDOS:
                continue
            v = rec.get(campo)
            f = _br2f_xlsx(v)
            campos[campo] = str(f)

        obs: list[tuple[str, str]] = []
        for i in range(1, 6):
            cod = str(rec.get(f"obsCod_{i}") or "").strip()
            val = rec.get(f"obsValor_{i}")
            if cod and cod not in ("", "0", "None"):
                obs.append((cod, str(_br2f_xlsx(val))))

        resultado[carimbo] = {"campos": campos, "obs": obs}
    return resultado


def carregar_auditoria(csv_path: Path, filtro_carimbos: set[str] | None = None) -> dict[str, dict]:
    """
    Retorna:
      {carimbo: {"campos": {campo: ocr_novo}, "obs": [(cod, val), ...]}}
    Apenas linhas com diferente=SIM.
    """
    resultado: dict[str, dict] = {}
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            carimbo = row["carimbo"].strip()
            if filtro_carimbos and carimbo not in filtro_carimbos:
                continue
            if row["diferente"].strip().upper() != "SIM":
                continue
            campo = row["campo"].strip()
            if carimbo not in resultado:
                resultado[carimbo] = {"campos": {}, "obs": []}
            if campo == "_obs_pairs":
                # parse "97:6.32 | 190:3709.77 | 192:-3709.77"
                ocr_obs = row["ocr_novo"].strip()
                if ocr_obs and ocr_obs != "(vazio)":
                    for parte in ocr_obs.split("|"):
                        parte = parte.strip()
                        if ":" in parte:
                            cod, val = parte.split(":", 1)
                            resultado[carimbo]["obs"].append((cod.strip(), val.strip()))
            elif campo not in CAMPOS_EXCLUIDOS:
                resultado[carimbo]["campos"][campo] = row["ocr_novo"].strip()
    return resultado


# ─── Consen: escrita de campos ─────────────────────────────────────────────────

_aba_ativa: str | None = None


def _ativar_aba(driver, texto_aba: str | None) -> None:
    global _aba_ativa
    if not texto_aba or texto_aba == _aba_ativa:
        return
    try:
        driver.execute_script("""
            var links = document.querySelectorAll('a, li, button, [role=tab]');
            for (var i = 0; i < links.length; i++) {
                var t = (links[i].innerText || links[i].textContent || '').trim();
                if (t.toLowerCase().indexOf(arguments[0].toLowerCase()) >= 0) {
                    links[i].click();
                    break;
                }
            }
        """, texto_aba)
        time.sleep(0.6)
        _aba_ativa = texto_aba
        log(f"  Aba '{texto_aba}' ativada")
    except Exception as e:
        warn(f"  Falha ao ativar aba '{texto_aba}': {e}")


def _preencher_campo(driver, campo_id: str, valor_str: str) -> bool:
    """Escreve valor em campo input via JS + dispara eventos."""
    try:
        ok = driver.execute_script("""
            var el = document.getElementById(arguments[0]);
            if (!el) return false;
            el.scrollIntoView({block:'center'});
            el.focus();
            el.value = arguments[1];
            el.dispatchEvent(new Event('input',  {bubbles:true}));
            el.dispatchEvent(new Event('change', {bubbles:true}));
            el.dispatchEvent(new Event('blur',   {bubbles:true}));
            return true;
        """, campo_id, valor_str)
        return bool(ok)
    except Exception as e:
        warn(f"    [{campo_id}] erro JS: {e}")
        return False


def _excluir_obs_existentes(driver) -> int:
    """Clica nos botões de excluir de todas as linhas da tabela de obs."""
    removidas = 0
    for _ in range(10):
        try:
            clicou = driver.execute_script("""
                var btns = document.querySelectorAll('button[onclick*="Excluir"], button[onclick*="excluir"], img[onclick*="excluir"], img[onclick*="Excluir"]');
                if (!btns.length) {
                    // fallback: botão "X" ou "Excluir" na tabela de obs
                    var trs = document.querySelectorAll('#tabelaObs tr, table tr');
                    for (var i = 0; i < trs.length; i++) {
                        var btn = trs[i].querySelector('button, input[type=button], img[title*="xcluir"], img[alt*="xcluir"]');
                        if (btn) { btn.click(); return true; }
                    }
                    return false;
                }
                btns[0].click();
                return true;
            """)
            if clicou:
                removidas += 1
                time.sleep(0.3)
            else:
                break
        except Exception:
            break
    return removidas


def _adicionar_obs(driver, wait, pares: list[tuple[str, str]]) -> None:
    for idx, (cod, val_str) in enumerate(pares, 1):
        try:
            el_sel = wait.until(EC.presence_of_element_located((By.ID, OBS_SEL_ID)))
            preencher_elemento_html(driver, el_sel, cod)
            log(f"  [OBS] {idx}: cod={cod}")
        except Exception as e:
            warn(f"  [OBS] {idx}: erro select cod={cod} — {e}")
            continue

        time.sleep(0.15)
        val_fmt = formatar_numero_br(_br2f(val_str))
        try:
            el_val = wait.until(EC.element_to_be_clickable((By.ID, OBS_VAL_ID)))
            driver.execute_script("arguments[0].value='';", el_val)
            el_val.click()
            el_val.send_keys(val_fmt)
            driver.execute_script("""
                arguments[0].dispatchEvent(new Event('input',{bubbles:true}));
                arguments[0].dispatchEvent(new Event('change',{bubbles:true}));
            """, el_val)
            log(f"  [OBS] {idx}: val={val_fmt}")
        except Exception as e:
            warn(f"  [OBS] {idx}: erro valor — {e}")
            continue

        time.sleep(0.2)
        try:
            driver.execute_script("document.getElementById(arguments[0]).click();", OBS_BTN_ID)
            time.sleep(0.3)
        except Exception as e:
            warn(f"  [OBS] {idx}: erro btnIncluiLinha — {e}")


def _salvar(driver, wait, salvar: bool) -> bool:
    if not salvar:
        log("  [DRY-RUN] Não salvando.")
        return True
    try:
        wait.until(EC.element_to_be_clickable((By.ID, "btnSalvar"))).click()
    except Exception:
        try:
            driver.execute_script("document.getElementById('btnSalvar').click();")
        except Exception as e:
            warn(f"  Erro ao clicar Salvar: {e}")
            return False
    _aguardar_sem_spinner(driver, timeout=10, min_wait=0.3)
    return True


# ─── fluxo por carimbo ────────────────────────────────────────────────────────

EDIT_URL = f"{BASE_URL}index.php#bpg/gestao/fatura/editaFaturaCarimbo.php"


def _abrir_fatura(driver, wait, carimbo: str) -> None:
    """Usa fluxo editaFaturaCarimbo → Carregar, igual ao digitação.
    Ignora timeout de btnSalvar — ENEL MT demora mais que 12s."""
    fluxo_base.abrir_tela_edicao_carimbo(driver, wait, EDIT_URL)
    try:
        fluxo_base.carregar_fatura_por_carimbo(driver, wait, carimbo, ())
    except Exception:
        # btnSalvar timeout esperado; verifica se URL está correta
        num = re.sub(r"\D", "", carimbo)
        url = driver.current_url or ""
        if "editaTabFatura" not in url or f"carimbo={num}" not in url:
            raise
    time.sleep(1.0)


def corrigir_carimbo(driver, wait, carimbo: str, dados: dict, salvar: bool, escrever_obs: bool = True) -> str:
    global _aba_ativa
    _aba_ativa = None  # reset a cada carimbo

    _abrir_fatura(driver, wait, carimbo)

    campos = dados.get("campos", {})
    obs    = dados.get("obs", [])

    erros = []

    # Campos numéricos — agrupados por aba para minimizar cliques
    for campo, ocr_val_str in campos.items():
        entry = CAMPO_PARA_ABA_ID.get(campo)
        if not entry:
            warn(f"  [{campo}] sem mapeamento de ID — pulando")
            continue
        aba_texto, campo_id = entry
        _ativar_aba(driver, aba_texto)
        val_f = _br2f(ocr_val_str)
        val_fmt = formatar_numero_br(val_f)
        ok = _preencher_campo(driver, campo_id, val_fmt)
        if ok:
            log(f"  {campo} -> {val_fmt}")
        else:
            warn(f"  {campo} ({campo_id}) não encontrado")
            erros.append(campo)

    # Obs
    if escrever_obs and obs:
        n_rem = _excluir_obs_existentes(driver)
        if n_rem:
            log(f"  {n_rem} obs existentes removidas")
        _adicionar_obs(driver, wait, obs)

    # Salvar
    ok_salvar = _salvar(driver, wait, salvar)
    if not ok_salvar:
        erros.append("SALVAR")

    return "ok" if not erros else f"erros:{','.join(erros)}"


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--salvar", action="store_true", help="Efetiva as correcoes (sem flag = dry-run)")
    parser.add_argument("--direto", action="store_true", help="Le xlsx do OCR diretamente (sem auditoria)")
    parser.add_argument("--xlsx", default=None, help="Caminho alternativo para o xlsx do OCR (com --direto)")
    parser.add_argument("--carimbo", action="append", dest="carimbos", help="Filtrar por carimbo(s)")
    parser.add_argument("--sem-obs", action="store_true", help="Nao escreve observacoes (apenas campos numericos)")
    parser.add_argument("--apenas-cap", action="store_true",
                        help="Modo restrito: escreve so capacitivo/UFER + fatDescontoFioKWh; "
                             "obs apenas cod 213 (DIC) se houver; nao toca em mais nada")
    args = parser.parse_args()

    filtro = {f"BB_{re.sub(chr(92)+'D','',c)}" for c in (args.carimbos or [])} or None

    if args.direto:
        xlsx_path = Path(args.xlsx) if args.xlsx else XLSX_OCR
        log(f"[DIRETO] Carregando xlsx: {xlsx_path.name}")
        dados = carregar_xlsx_direto(xlsx_path, filtro)
        # Modo --apenas-cap: restringir campos e obs
        if args.apenas_cap:
            CAMPOS_CAP = {"fatConFPontaIndExcRegistrado", "fatConFPontaIndExcFaturado",
                          "fatConFPontaIndExcValorReais", "fatDescontoFioKWh"}
            for carimbo, rec in dados.items():
                rec["campos"] = {k: v for k, v in rec["campos"].items() if k in CAMPOS_CAP}
                rec["obs"]    = [(c, v) for c, v in rec["obs"] if str(c) == "213"]
        log(f"  {len(dados)} carimbos carregados do OCR")
    else:
        log(f"Carregando auditoria: {CSV_AUDITORIA.name}")
        dados = carregar_auditoria(CSV_AUDITORIA, filtro)
        log(f"  {len(dados)} carimbos com diferencas a corrigir")

    if not args.salvar:
        log("  [DRY-RUN] Use --salvar para efetivar")

    driver, wait = fluxo_base.abrir_driver_logado()
    try:
        total = len(dados)
        for idx, (carimbo, rec) in enumerate(sorted(dados.items()), 1):
            n_campos = len(rec["campos"])
            n_obs    = len(rec["obs"])
            log(f"[{idx}/{total}] {carimbo}  campos={n_campos}  obs={n_obs}")
            for tentativa in range(2):
                try:
                    status = corrigir_carimbo(driver, wait, carimbo, rec, args.salvar, escrever_obs=not args.sem_obs)
                    _registrar(carimbo, status)
                    break
                except Exception as e:
                    msg = str(e)
                    warn(f"  Excecao (tentativa {tentativa+1}): {msg[:200]}")
                    if tentativa == 0 and any(k in msg for k in ("invalid session", "session id", "not reachable", "nao carregou", "carregou")):
                        warn("  Sessao morta — reabrindo driver...")
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        driver, wait = fluxo_base.abrir_driver_logado()
                    else:
                        _registrar(carimbo, "erro", msg[:120])
                        break
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    log("Concluido.")


if __name__ == "__main__":
    main()
