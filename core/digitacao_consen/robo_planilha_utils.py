#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from core.project_paths import resolve_indice_master_csv

ROOT_DIR = Path(__file__).resolve().parents[2]
ARQUIVOS_ENZO_DIR = Path(r"\\10.10.250.21\Energia\ARQUIVOS ENZO")
CONSOLIDADO_PADRAO = (
    ARQUIVOS_ENZO_DIR
    / "Denise"
    / "Consen - Correções"
    / "Consolidado_Robo_Digitador.csv"
)
INDICE_MASTER_PADRAO = resolve_indice_master_csv(prefer_network=False)

DOWNLOAD_ROOTS: dict[str, tuple[Path, ...]] = {
    "CEMIG": (ARQUIVOS_ENZO_DIR / "DOWNLOAD CEMIG",),
    "ENEL": (
        ARQUIVOS_ENZO_DIR / "DOWNLOAD ENEL",
        ARQUIVOS_ENZO_DIR / "DOWNLOAD ENEL RJ",
        ARQUIVOS_ENZO_DIR / "DOWNLOAD ENEL CE",
    ),
    "ENERGISA": (ARQUIVOS_ENZO_DIR / "DOWNLOAD ENERGISA",),
    "NEOENERGIA": (ARQUIVOS_ENZO_DIR / "DOWNLOAD NEOENERGIA",),
    "CPFL": (ARQUIVOS_ENZO_DIR / "DOWNLOAD CPFL",),
    "RGE": (ARQUIVOS_ENZO_DIR / "DOWNLOAD CPFL",),
    "EQUATORIAL": (ARQUIVOS_ENZO_DIR / "DOWNLOAD EQUATORIAL",),
    "CELESC": (ARQUIVOS_ENZO_DIR / "DOWNLOAD CELESC",),
    "CEEE": (ARQUIVOS_ENZO_DIR / "DOWNLOAD CEEE",),
    "LIGHT": (ARQUIVOS_ENZO_DIR / "DOWNLOAD LIGHT",),
    "COPEL": (ARQUIVOS_ENZO_DIR / "DOWNLOAD COPEL",),
}

OCR_ROOTS: dict[str, Path] = {
    "CEMIG": ARQUIVOS_ENZO_DIR / "OCR CEMIG",
    "ENEL": ARQUIVOS_ENZO_DIR / "OCR ENEL",
    "ENERGISA": ARQUIVOS_ENZO_DIR / "OCR ENERGISA",
    "NEOENERGIA": ARQUIVOS_ENZO_DIR / "OCR NEOENERGIA",
    "CPFL": ARQUIVOS_ENZO_DIR / "OCR CPFL",
    "RGE": ARQUIVOS_ENZO_DIR / "OCR RGE SUL",
    "EQUATORIAL": ARQUIVOS_ENZO_DIR / "OCR EQUATORIAL PA",
    "CELESC": ARQUIVOS_ENZO_DIR / "OCR CELESC",
    "CEEE": ARQUIVOS_ENZO_DIR / "OCR CEEE",
    "LIGHT": ARQUIVOS_ENZO_DIR / "OCR LIGHT RJ",
    "COPEL": ARQUIVOS_ENZO_DIR / "OCR COPEL",
}

CONSUMO_CAMPOS = (
    "fatConFPontaIndRegistrado",
    "fatConFPontaIndFaturado",
    "fatConFPontaIndValorReais",
)
TARIFA_BRANCA_CAMPOS = (
    "fatConPontaRegistrado",
    "fatConPontaFaturado",
    "fatConPontaValorReais",
    "fatConIntermediarioRegistrado",
    "fatConIntermediarioFaturado",
    "fatConIntermediarioValorReais",
    "fatConFPontaIndRegistrado",
    "fatConFPontaIndValorReais",
)
MULTA_CAMPOS_CANDIDATOS = (
    "fatMultasDiversas",
    "obsValor",
    "obsValor_1",
    "obsValor_2",
    "obsValor_3",
    "obsValor_4",
    "obsValor_5",
)


def normalizar_carimbo(valor: str) -> str:
    texto = str(valor or "").strip().upper().replace("BB_", "")
    texto = re.sub(r"\D", "", texto)
    if not texto:
        raise ValueError("Carimbo vazio.")
    return texto


def normalizar_data_referencia(valor: str) -> str:
    texto = str(valor or "").strip()
    if not texto:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return texto


def valor_decimal(valor: Any) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return 0.0
    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def formatar_zero(valor: Any = 0) -> str:
    return "0" if valor_decimal(valor) == 0 else str(valor)


@dataclass
class CasoRobo:
    queixa: str
    carimbo: str
    instalacao: str = ""
    data_referencia: str = ""
    concessionaria: str = ""
    arquivo_origem: str = ""
    linha: dict[str, str] = field(default_factory=dict)


@dataclass
class MetaIndice:
    indice: str
    concessionaria: str = ""
    sistema: str = ""
    estado: str = ""
    uc: str = ""
    mes_ref: str = ""
    arquivo: str = ""

    @property
    def carimbo(self) -> str:
        return normalizar_carimbo(self.indice)

    @property
    def is_mt(self) -> bool:
        texto = " ".join((self.sistema, self.arquivo, self.concessionaria)).upper()
        return " MT" in texto or "\\MT\\" in texto or texto.endswith("MT")

    @property
    def mes_ref_compacto(self) -> str:
        texto = str(self.mes_ref or "").strip()
        if not texto:
            return ""
        m = re.match(r"(\d{2})-(\d{4})", texto)
        if not m:
            return ""
        return f"{m.group(1)}{m.group(2)}"

    @property
    def sistema_chave(self) -> str:
        texto = " ".join((self.concessionaria, self.sistema)).upper()
        if "CEMIG" in texto:
            return "CEMIG"
        if "NEOENERGIA" in texto or "ELEKTRO" in texto:
            return "NEOENERGIA"
        if "ENERGISA" in texto or "CERON" in texto:
            return "ENERGISA"
        if "CPFL" in texto:
            return "CPFL"
        if "RGE" in texto:
            return "RGE"
        if "EQUATORIAL" in texto:
            return "EQUATORIAL"
        if "CELESC" in texto:
            return "CELESC"
        if "CEEE" in texto:
            return "CEEE"
        if "LIGHT" in texto:
            return "LIGHT"
        if "COPEL" in texto:
            return "COPEL"
        if "ENEL" in texto:
            return "ENEL"
        return texto.split()[0] if texto else ""


def carregar_casos_consolidado(caminho: Path = CONSOLIDADO_PADRAO) -> list[CasoRobo]:
    with caminho.open(encoding="utf-8-sig", newline="") as arquivo:
        leitor = csv.DictReader(arquivo)
        return [
            CasoRobo(
                queixa=str(row.get("Queixa") or "").strip(),
                carimbo=normalizar_carimbo(str(row.get("Carimbo") or "")),
                instalacao=str(row.get("Instalacao") or "").strip(),
                data_referencia=normalizar_data_referencia(row.get("DataReferencia") or row.get("Data Referencia") or ""),
                concessionaria=str(row.get("Concessionaria") or "").strip(),
                arquivo_origem=str(row.get("ArquivoOrigem") or "").strip(),
                linha={k: str(v or "").strip() for k, v in row.items()},
            )
            for row in leitor
            if str(row.get("Carimbo") or "").strip()
        ]


def filtrar_casos(
    casos: Iterable[CasoRobo],
    *,
    queixa: str | None = None,
    carimbos: set[str] | None = None,
) -> list[CasoRobo]:
    carimbos_norm = {normalizar_carimbo(c) for c in carimbos or set()}
    saida: list[CasoRobo] = []
    for caso in casos:
        if queixa and caso.queixa != queixa:
            continue
        if carimbos_norm and caso.carimbo not in carimbos_norm:
            continue
        saida.append(caso)
    return saida


def carregar_carimbos_arquivo(caminho: str) -> set[str]:
    if not caminho:
        return set()
    linhas = Path(caminho).read_text(encoding="utf-8-sig").splitlines()
    return {
        normalizar_carimbo(linha)
        for linha in linhas
        if linha.strip() and not linha.strip().startswith("#")
    }


def carregar_indice_master(caminho: Path = INDICE_MASTER_PADRAO) -> dict[str, MetaIndice]:
    with caminho.open(encoding="utf-8-sig", newline="") as arquivo:
        leitor = csv.DictReader(arquivo)
        resultado: dict[str, MetaIndice] = {}
        for row in leitor:
            indice = str(row.get("INDICE") or "").strip()
            if not indice:
                continue
            resultado[normalizar_carimbo(indice)] = MetaIndice(
                indice=indice,
                concessionaria=str(row.get("CONCESSIONARIA") or "").strip(),
                sistema=str(row.get("SISTEMA") or "").strip(),
                estado=str(row.get("ESTADO") or "").strip(),
                uc=str(row.get("UC") or "").strip(),
                mes_ref=str(row.get("MES_REF") or "").strip(),
                arquivo=str(row.get("ARQUIVO") or "").strip(),
            )
        return resultado


def obter_meta_caso(caso: CasoRobo, indice_map: dict[str, MetaIndice]) -> MetaIndice | None:
    return indice_map.get(caso.carimbo)


def selecionar_campos_multa_negativa(campos: dict[str, str]) -> dict[str, str]:
    correcoes: dict[str, str] = {}
    for campo in MULTA_CAMPOS_CANDIDATOS:
        atual = campos.get(campo, "")
        if valor_decimal(atual) < 0:
            correcoes[campo] = "0"
    return correcoes


def montar_payload_consumo_zerado(dados_ocr: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    registrado = valor_decimal(dados_ocr.get("fatConFPontaIndRegistrado"))
    faturado = valor_decimal(dados_ocr.get("fatConFPontaIndFaturado"))
    valor_rs = valor_decimal(dados_ocr.get("fatConFPontaIndValorReais"))

    if registrado > 0:
        payload["fatConFPontaIndRegistrado"] = registrado
    if faturado > 0:
        payload["fatConFPontaIndFaturado"] = faturado
    elif registrado > 0:
        payload["fatConFPontaIndFaturado"] = registrado
    if valor_rs > 0:
        payload["fatConFPontaIndValorReais"] = valor_rs
    return payload


def montar_payload_tarifa_branca(dados_ocr: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for campo in TARIFA_BRANCA_CAMPOS:
        valor = dados_ocr.get(campo)
        if valor_decimal(valor) > 0:
            payload[campo] = valor
    return payload


@lru_cache(maxsize=64)
def _carregar_rows_xlsx(caminho: str) -> list[dict[str, Any]]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(v).strip() if v is not None else "" for v in next(rows)]
        resultado: list[dict[str, Any]] = []
        for row in rows:
            item = {
                headers[i]: row[i]
                for i in range(min(len(headers), len(row)))
                if headers[i]
            }
            resultado.append(item)
        return resultado
    finally:
        wb.close()


def buscar_registro_ocr_exportado(caso: CasoRobo, meta: MetaIndice | None) -> dict[str, Any] | None:
    if meta is None:
        return None
    raiz = OCR_ROOTS.get(meta.sistema_chave)
    if raiz is None or not raiz.exists():
        return None

    arquivos = sorted(raiz.glob("*.xlsx"))
    if meta.mes_ref_compacto:
        priorizados = [p for p in arquivos if meta.mes_ref_compacto in p.name]
        arquivos = priorizados or arquivos

    for arquivo in arquivos:
        try:
            for row in _carregar_rows_xlsx(str(arquivo)):
                carimbo = str(row.get("fatCarimbo") or row.get("Carimbo") or "").strip()
                if carimbo and normalizar_carimbo(carimbo) == caso.carimbo:
                    return row
        except Exception:
            continue
    return None


def localizar_pdf_meta(meta: MetaIndice | None) -> Path | None:
    if meta is None:
        return None

    bruto = str(meta.arquivo or "").strip()
    if bruto:
        direto = Path(bruto)
        if direto.exists():
            return direto
        nome = direto.name
    else:
        nome = f"BB_{meta.carimbo}.pdf"

    for raiz in DOWNLOAD_ROOTS.get(meta.sistema_chave, (ARQUIVOS_ENZO_DIR,)):
        if not raiz.exists():
            continue
        encontrados = list(raiz.rglob(nome))
        if encontrados:
            return encontrados[0]
    return None


def extrair_dados_ocr_pdf(
    pdf_path: Path,
    *,
    meta: MetaIndice | None,
    data_referencia: str = "",
) -> dict[str, Any]:
    chave = meta.sistema_chave if meta else ""
    mes = 0
    ano = 0
    data_base = data_referencia or (meta.mes_ref if meta else "")
    m = re.search(r"(?P<mes>\d{2})[-/](?P<ano>\d{4})", data_base)
    if m:
        mes = int(m.group("mes"))
        ano = int(m.group("ano"))
    elif data_referencia:
        try:
            dt = datetime.strptime(data_referencia, "%Y-%m-%d")
            mes = dt.month
            ano = dt.year
        except ValueError:
            pass

    if chave == "CEMIG":
        from core.ocr.OCR_Cemig import processar_pdf

        return processar_pdf(str(pdf_path), "mt" if meta and meta.is_mt else "bt")
    if chave == "ENEL":
        from core.ocr.ocr_enel import processar_pdf

        return processar_pdf(str(pdf_path), "mt" if meta and meta.is_mt else "bt")
    if chave == "NEOENERGIA":
        from core.ocr.ocr_neoenergia import processar_pdf_direto

        return processar_pdf_direto(str(pdf_path), f"{mes or 1:02d}", str(ano or datetime.now().year))
    if chave == "ENERGISA":
        from core.ocr.ocr_energisa_bt import processar_pdf_direto

        return processar_pdf_direto(Path(pdf_path), mes or 1, ano or datetime.now().year)
    if chave == "CPFL":
        from core.ocr.ocr_cpfl_bt import processar_pdf

        return processar_pdf(str(pdf_path))
    if chave == "RGE":
        from core.ocr.ocr_rge_sul_bt import processar_pdf

        return processar_pdf(str(pdf_path))
    if chave == "EQUATORIAL":
        sistema = (meta.sistema or "").upper() if meta else ""
        if "PA MT" in sistema:
            from core.ocr.ocr_equatorial_pa_mt import processar_pdf
        elif "PI MT" in sistema:
            from core.ocr.ocr_equatorial_pi_mt import processar_pdf
        elif "MA MT" in sistema:
            from core.ocr.ocr_equatorial_ma_mt import processar_pdf
        elif "PI BT" in sistema:
            from core.ocr.ocr_equatorial_pi_bt_parser import processar_pdf
        elif "AL" in sistema:
            from core.ocr.ocr_equatorial_al_bt import processar_pdf_al as processar_pdf
        else:
            from core.ocr.ocr_bt_generico import processar_pdf

        return processar_pdf(str(pdf_path))
    if chave == "CELESC":
        if meta and meta.is_mt:
            from core.ocr.ocr_celesc_mt import extrair_campos
        else:
            from core.ocr.ocr_celesc_bt import extrair_campos

        return extrair_campos(str(pdf_path))
    if chave == "CEEE":
        from core.ocr.ocr_ceee_bt import processar_pdf_direto

        return processar_pdf_direto(str(pdf_path), f"{mes or 1:02d}", str(ano or datetime.now().year))
    if chave == "LIGHT":
        if meta and meta.is_mt:
            from core.ocr.ocr_light_rj_mt import processar_pdf
        else:
            from core.ocr.ocr_light_rj_bt import _parser_light_bt as processar_pdf

        return processar_pdf(str(pdf_path))
    if chave == "COPEL":
        if meta and meta.is_mt:
            from core.ocr.ocr_copel_mt import processar_pdf
            return processar_pdf(str(pdf_path))
        from core.ocr.ocr_copel_bt import processar_pdf as processar_pdf_bt

        return processar_pdf_bt(str(pdf_path))

    from core.ocr.ocr_bt_generico import processar_pdf

    return processar_pdf(str(pdf_path), str(pdf_path))


def resolver_dados_ocr_caso(
    caso: CasoRobo,
    indice_map: dict[str, MetaIndice],
) -> tuple[dict[str, Any] | None, str]:
    meta = obter_meta_caso(caso, indice_map)
    row_ocr = buscar_registro_ocr_exportado(caso, meta)
    if row_ocr:
        return row_ocr, "ocr_xlsx"

    pdf_path = localizar_pdf_meta(meta)
    if pdf_path is None:
        return None, "fonte_nao_encontrada"

    try:
        return extrair_dados_ocr_pdf(pdf_path, meta=meta, data_referencia=caso.data_referencia), str(pdf_path)
    except Exception as exc:
        return None, f"erro_ocr:{type(exc).__name__}:{exc}"
